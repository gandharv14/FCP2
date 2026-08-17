#!/usr/bin/env python3
"""Extract auditable formula-series evidence for the custom-formula gate."""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import math
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.formula import Tokenizer
from openpyxl.utils import get_column_letter, range_boundaries

try:
    import tomllib
except ImportError:  # pragma: no cover
    import tomli as tomllib


ROLE_PATTERNS = {
    "depreciation_amortization": (
        r"\bdepreciat", r"\bamorti[sz]", r"\bd&a\b", r"\basset depreciation\b",
    ),
    "interest_expense_income": (
        r"\binterest", r"\bfinance cost", r"\bfinancing cost",
    ),
    "tax": (r"\btax(?:es|ation)?\b", r"\bcurrent tax\b", r"\bcash tax\b"),
    "revenue": (r"\brevenue", r"\bsales\b", r"\bturnover\b"),
    "operating_expense": (
        r"\bopex\b", r"\boperating expense", r"\bsg&a\b", r"\bsalar",
        r"\brent\b", r"\bcost of sales\b",
    ),
    "working_capital": (
        r"\bworking capital\b", r"\breceivable", r"\binventor",
        r"\bpayable", r"\bdso\b", r"\bdio\b", r"\bdpo\b",
    ),
    "capital_expenditure": (
        r"\bcapex\b", r"\bcapital expenditure", r"\bfixed.asset addition",
    ),
    "debt_draw_repayment": (
        r"\bdebt\b", r"\bloan\b", r"\bdrawdown\b", r"\brepayment\b",
        r"\breimbursement\b", r"\bcash sweep\b",
    ),
    "discounting_valuation": (
        r"\bdiscount", r"\bpresent value\b", r"\bterminal value\b", r"\bnpv\b",
    ),
    "returns": (r"\birr\b", r"\bxirr\b", r"\bmoic\b", r"\bmultiple of money\b"),
    "cash_flow_metric": (
        r"\bcash flow\b", r"\bfcf\b", r"\bfcff\b", r"\bfcfe\b",
    ),
    "profit_metric": (
        r"\bebitda\b", r"\bebit\b", r"\bebt\b", r"\bnet profit\b",
        r"\bnet income\b",
    ),
    "balance_rollforward": (
        r"\bbop\b", r"\beop\b", r"\bbeginning balance\b", r"\bending balance\b",
        r"\bopening balance\b", r"\bclosing balance\b", r"\bretained earnings\b",
    ),
}

BORING_LITERALS = {0.0, 1.0, 2.0, 12.0, 100.0, 365.0, 1000.0, 10000.0}
CELL_REF_RE = re.compile(r"(?:'[^']+'|[A-Za-z_][^!]*)?!?\$?[A-Z]{1,3}\$?\d+")
COMPARE_RE = re.compile(r"(?:<=|>=|<>|=|<|>)")


def json_value(value):
    if isinstance(value, (dt.date, dt.datetime, dt.time)):
        return value.isoformat()
    if isinstance(value, float) and not math.isfinite(value):
        return str(value)
    return value


def parse_args():
    parser = argparse.ArgumentParser(
        description="Extract pre-package golden formula context from verified segmentation"
    )
    parser.add_argument(
        "workbook",
        help="Workbook id or path to the golden .xlsx workbook",
    )
    parser.add_argument(
        "--source",
        type=Path,
        default=Path("4-10 100"),
        help="Source workbook directory when workbook is an id",
    )
    parser.add_argument("--repo-root", type=Path, default=Path.cwd())
    parser.add_argument("--seg-dir", type=Path, default=None)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--neighbor-radius", type=int, default=2)
    return parser.parse_args()


def resolve_workbook(repo_root: Path, source: Path, value: str):
    candidate = Path(value)
    explicit = candidate if candidate.is_absolute() else repo_root / candidate
    if explicit.is_file():
        workbook = explicit.resolve()
        workbook_id = workbook.stem
    else:
        workbook_id = candidate.stem
        source_dir = source if source.is_absolute() else repo_root / source
        workbook = (source_dir / f"{workbook_id}.xlsx").resolve()
    if not workbook.exists():
        raise FileNotFoundError(f"missing raw workbook: {workbook}")
    return workbook_id, workbook


def load_curation(seg_dir: Path):
    curation_file = seg_dir / "curation.toml"
    if not curation_file.exists():
        raise FileNotFoundError(f"missing curation: {curation_file}")
    with curation_file.open("rb") as handle:
        data = tomllib.load(handle)
    outputs = [
        output for output in data.get("output") or [] if output.get("include")
    ]
    if not outputs:
        raise RuntimeError(f"{curation_file} contains no included outputs")
    return outputs


def load_lineage(seg_dir: Path):
    lineage_file = seg_dir / "lineage.json"
    if not lineage_file.exists():
        raise FileNotFoundError(f"missing lineage: {lineage_file}")
    data = json.loads(lineage_file.read_text(encoding="utf-8"))
    outputs_by_band = defaultdict(set)
    step_metadata = {}
    consumers = defaultdict(set)
    drivers_by_band = defaultdict(dict)
    for trace in data.get("traces") or []:
        output = str(trace.get("output") or "")
        steps = trace.get("band_steps") or []
        by_node = {
            str(step.get("node") or ""): step
            for step in steps
            if step.get("node")
        }
        for step in steps:
            node = str(step.get("node") or "")
            if node:
                outputs_by_band[node].add(output)
                current = step_metadata.setdefault(
                    node,
                    {
                        "node": node,
                        "bucket": step.get("bucket") or "",
                        "label": step.get("label") or "",
                        "formula": step.get("formula") or "",
                        "depth": int(step.get("depth") or 0),
                        "values": step.get("values") or [],
                    },
                )
                current["depth"] = max(
                    int(current.get("depth") or 0), int(step.get("depth") or 0)
                )
                for driver_node in step.get("inputs") or []:
                    driver_node = str(driver_node)
                    consumers[driver_node].add(node)
                    driver = by_node.get(driver_node) or {}
                    drivers_by_band[node][driver_node] = {
                        "node": driver_node,
                        "bucket": driver.get("bucket") or "",
                        "label": driver.get("label") or "",
                        "formula": driver.get("formula") or "",
                        "depth": int(driver.get("depth") or 0),
                        "values": driver.get("values") or [],
                    }
    return data, outputs_by_band, step_metadata, consumers, drivers_by_band


def load_bands(seg_dir: Path):
    bands_file = seg_dir / "bands.csv"
    if not bands_file.exists():
        raise FileNotFoundError(f"missing bands: {bands_file}")
    with bands_file.open(encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))


def split_band(band: str):
    sheet, separator, ref = band.rpartition("!")
    if not separator:
        raise ValueError(f"invalid band reference: {band}")
    min_col, min_row, max_col, max_row = range_boundaries(ref)
    return sheet, min_row, min_col, max_row, max_col


def row_label(ws_formula, ws_values, row: int, left_of_col: int) -> str:
    candidates = []
    for col in range(1, left_of_col):
        value = ws_values.cell(row, col).value
        if value is None:
            value = ws_formula.cell(row, col).value
        if isinstance(value, str) and value.strip() and not value.lstrip().startswith("="):
            candidates.append(value.strip())
    if not candidates:
        return ""
    if candidates[0] == candidates[-1]:
        return candidates[0]
    return f"{candidates[0]} / {candidates[-1]}"


def tokens(formula: str):
    try:
        return Tokenizer(formula).items
    except Exception:
        return []


def formula_details(formula: str):
    references = []
    literals = []
    for token in tokens(formula):
        if token.type == "OPERAND" and token.subtype == "RANGE":
            if CELL_REF_RE.search(token.value):
                references.append(token.value)
        elif token.type == "OPERAND" and token.subtype == "NUMBER":
            try:
                literals.append(float(token.value))
            except ValueError:
                pass

    upper = formula.upper()
    body = formula.lstrip()
    if body.startswith("="):
        body = body[1:]
    has_comparison = bool(COMPARE_RE.search(body))
    notable = sorted({number for number in literals if number not in BORING_LITERALS})
    signals = []
    if has_comparison and any(name in upper for name in ("IF(", "IFS(", "AND(", "OR(")):
        signals.append("boolean_gate")
    if has_comparison and ("IF(" in upper or "IFS(" in upper):
        signals.append("asymmetric_if")
    if 0.5 in literals:
        signals.append("hardcoded_half_year")
    if ("MAX(" in upper or "MIN(" in upper):
        signals.append("max_min_clamp")
    if "SUM(" in upper and ":" in formula:
        signals.append("range_aggregation")
    if re.search(r"(?:\*\s*-1|^\s*=\s*-)", formula):
        signals.append("sign_flip")
    signals.extend(f"literal_embedded:{number:g}" for number in notable)
    return {
        "references": list(dict.fromkeys(references)),
        "numeric_literals": literals,
        "notable_literals": notable,
        "signals": list(dict.fromkeys(signals)),
    }


def role_hints(text: str):
    lowered = text.lower()
    return [
        role
        for role, patterns in ROLE_PATTERNS.items()
        if any(re.search(pattern, lowered) for pattern in patterns)
    ]


def cell_payload(ws_formula, ws_values, coordinate: str):
    raw = ws_formula[coordinate].value
    cached = ws_values[coordinate].value
    return {
        "cell": f"{ws_formula.title}!{coordinate}",
        "formula": raw if isinstance(raw, str) and raw.startswith("=") else None,
        "value": json_value(cached),
    }


def extract_series(
    row,
    wb_formula,
    wb_values,
    outputs,
    radius: int,
    depth: int,
    fanout: int,
    drivers,
):
    band = row["band"]
    sheet, min_row, min_col, max_row, max_col = split_band(band)
    if min_row != max_row:
        raise ValueError(f"expected horizontal formula band, got {band}")
    ws_formula = wb_formula[sheet]
    ws_values = wb_values[sheet]

    cells = [
        f"{get_column_letter(col)}{min_row}"
        for col in range(min_col, max_col + 1)
    ]
    payloads = [cell_payload(ws_formula, ws_values, cell) for cell in cells]
    formulas = list(
        dict.fromkeys(item["formula"] for item in payloads if item["formula"])
    )

    references = []
    literals = []
    signals = []
    for formula in formulas:
        details = formula_details(formula)
        references.extend(details["references"])
        literals.extend(details["numeric_literals"])
        signals.extend(details["signals"])

    neighbors = []
    first_col = get_column_letter(min_col)
    last_col = get_column_letter(max_col)
    for neighbor_row in range(
        max(1, min_row - radius),
        min(ws_formula.max_row, min_row + radius) + 1,
    ):
        first = f"{first_col}{neighbor_row}"
        last = f"{last_col}{neighbor_row}"
        sample_cells = [first] if first == last else [first, last]
        neighbors.append(
            {
                "row": neighbor_row,
                "label": row_label(ws_formula, ws_values, neighbor_row, min_col),
                "samples": [
                    cell_payload(ws_formula, ws_values, coordinate)
                    for coordinate in sample_cells
                ],
            }
        )

    context_text = " / ".join(
        [row.get("label") or ""]
        + [neighbor.get("label") or "" for neighbor in neighbors]
    )
    return {
        "band": band,
        "sheet": sheet,
        "row": min_row,
        "columns": [first_col, last_col],
        "width": int(row.get("width") or len(cells)),
        "bucket": row.get("bucket") or "",
        "label": row.get("label") or "",
        "role_hints": role_hints(context_text),
        "r1c1_pattern": row.get("pattern") or "",
        "formula_samples": formulas[:3],
        "periods": payloads,
        "direct_references": list(dict.fromkeys(references)),
        "numeric_literals": list(dict.fromkeys(literals)),
        "signals": list(dict.fromkeys(signals)),
        "neighbors": neighbors,
        "downstream_outputs": sorted(outputs),
        "importance": {
            "output_coverage": len(outputs),
            "max_depth": depth,
            "downstream_fanout": fanout,
        },
        "direct_drivers": sorted(
            drivers,
            key=lambda driver: (
                int(driver.get("depth") or 0),
                str(driver.get("node") or ""),
            ),
        ),
    }


def main():
    args = parse_args()
    repo_root = args.repo_root.resolve()
    workbook_id, workbook = resolve_workbook(
        repo_root, args.source, args.workbook
    )

    seg_dir = args.seg_dir or (repo_root / "seg_out" / workbook_id)
    if not seg_dir.is_absolute():
        seg_dir = repo_root / seg_dir
    seg_dir = seg_dir.resolve()

    curated_outputs = load_curation(seg_dir)
    lineage, outputs_by_band, step_metadata, consumers, drivers_by_band = (
        load_lineage(seg_dir)
    )
    bands = load_bands(seg_dir)
    relevant_rows = []
    relevant_outputs = {}
    relevant_nodes = {}
    for row in bands:
        if row.get("kind") != "formula":
            continue
        band = row.get("band") or ""
        outputs = set(outputs_by_band.get(band, ()))
        component = row.get("component") or ""
        if component:
            outputs.update(outputs_by_band.get(component, ()))
        if outputs:
            relevant_rows.append(row)
            relevant_outputs[band] = outputs
            relevant_nodes[band] = (
                band if band in step_metadata else component
            )
    if not relevant_rows:
        raise RuntimeError("lineage contains no formula bands found in bands.csv")

    wb_formula = openpyxl.load_workbook(workbook, data_only=False, read_only=False)
    wb_values = openpyxl.load_workbook(workbook, data_only=True, read_only=False)
    try:
        series = [
            extract_series(
                row,
                wb_formula,
                wb_values,
                relevant_outputs[row["band"]],
                max(0, args.neighbor_radius),
                int(
                    (step_metadata.get(relevant_nodes[row["band"]]) or {}).get(
                        "depth", 0
                    )
                ),
                len(consumers.get(relevant_nodes[row["band"]], ())),
                list(
                    drivers_by_band.get(relevant_nodes[row["band"]], {}).values()
                ),
            )
            for row in relevant_rows
        ]
    finally:
        wb_formula.close()
        wb_values.close()

    series.sort(
        key=lambda item: (
            -item["importance"]["output_coverage"],
            -item["importance"]["max_depth"],
            -item["importance"]["downstream_fanout"],
            item["sheet"],
            item["row"],
            item["columns"][0],
        )
    )
    for rank, item in enumerate(series, 1):
        item["key_rank"] = rank

    output_names = [
        output.get("name") or output.get("label") or output.get("band")
        for output in curated_outputs
    ]
    answer_cells = []
    for output in curated_outputs:
        sheet, min_row, min_col, max_row, max_col = split_band(output["band"])
        answer_cells.extend(
            f"{sheet}!{get_column_letter(col)}{row}"
            for row in range(min_row, max_row + 1)
            for col in range(min_col, max_col + 1)
        )
    output = {
        "schema_version": "2.0",
        "task": {
            "name": f"{workbook_id}-outputs",
            "workbook": workbook_id,
            "output_names": output_names,
            "answer_cells": answer_cells,
        },
        "sources": {
            "raw_workbook": str(workbook),
            "segmentation": str(seg_dir),
            "curated_outputs": [
                item.get("output") for item in lineage.get("outputs") or []
            ],
        },
        "extraction": {
            "key_variables": len(series),
            "neighbor_radius": max(0, args.neighbor_radius),
            "ranking": [
                "output_coverage_desc",
                "max_depth_desc",
                "downstream_fanout_desc",
                "stable_band_order",
            ],
            "note": (
                "Every formula series in curated-output reverse closure is retained. "
                "Ranking prioritizes analysis but does not filter variables. Role "
                "hints and signals are candidate evidence only; final classes require "
                "closed-catalog agreement and assumption recoverability."
            ),
        },
        "key_variables": series,
    }

    output_path = args.output
    if not output_path.is_absolute():
        output_path = repo_root / output_path
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(output, indent=2) + "\n", encoding="utf-8")
    print(
        f"{workbook_id}: {len(series)} key golden formula series -> {output_path}"
    )


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"custom-formula gate extraction failed: {exc}", file=sys.stderr)
        raise SystemExit(2)
