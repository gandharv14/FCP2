#!/usr/bin/env python3
"""Unified pre-run disclosure for GDPval workbook rebuild tasks.

The hard invariant is selection by the task the agent actually sees:

  golden computes the cell
  + a graded answer reads the cell
  + the delivered workbook leaves the cell blank

Everything else is classification and rendering layered on top of that set.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import os
import re
import shutil
import sys
import time
from collections import Counter, defaultdict, deque
from pathlib import Path

try:
    import openpyxl
except ImportError:  # pragma: no cover
    sys.exit("openpyxl required: pip install openpyxl")


REPO_ROOT = Path(__file__).resolve().parents[4]
DEFAULT_TASKS_ROOT = REPO_ROOT.parent / "08_12_34_samples_tasks_outputs_hinted"
DEFAULT_RUNS_ROOT = REPO_ROOT / "runs" / "disclosure"
OUTPUT_HEADING = "\n## Output\n"
SECTION_START = "## Workbook disclosure"
STALE_AGENT_SECTIONS = (
    "## Custom formula hints",
    "## Modelling conventions in this workbook",
    SECTION_START,
)
FORMULA_RE = re.compile(r"`=|\b(?:IF|SUM|AVERAGE|NPV|IRR|XIRR|XNPV)\s*\(", re.I)
NUMBER_RE = re.compile(
    r"(?<![A-Za-z0-9_])[-+]?(?:\d+(?:,\d{3})*(?:\.\d+)?|\.\d+)"
    r"(?:[eE][-+]?\d+)?"
)


# --------------------------------------------------------------------------- refs

REF_RE = re.compile(
    r"""(?:(?P<sheet>'(?:[^']|'')+'|[A-Za-z_\\][A-Za-z0-9_. &-]*)!)?
        (?P<a>\$?[A-Z]{1,3}\$?[0-9]{1,7})
        (?::(?P<b>\$?[A-Z]{1,3}\$?[0-9]{1,7}))?
        (?![A-Za-z0-9_(])""",
    re.X,
)
STRING_RE = re.compile(r'"(?:[^"]|"")*"')
COORD_RE = re.compile(r"^\$?([A-Z]{1,3})\$?([0-9]{1,7})$")
NOT_A_REF = {"LOG10", "T", "N", "TRUE", "FALSE"}


def col_to_num(col: str) -> int:
    n = 0
    for ch in col.upper():
        n = n * 26 + ord(ch) - 64
    return n


def num_to_col(n: int) -> str:
    out = ""
    while n:
        n, r = divmod(n - 1, 26)
        out = chr(65 + r) + out
    return out


def split_coord(coord: str):
    m = COORD_RE.match(coord.replace("$", ""))
    if not m:
        return None
    return m.group(1), int(m.group(2))


def unquote_sheet(name: str) -> str:
    if name.startswith("'") and name.endswith("'"):
        return name[1:-1].replace("''", "'")
    return name


def quote_sheet(name: str) -> str:
    return "'%s'" % name if re.search(r"[^A-Za-z0-9_.]", name) else name


def key(sheet: str, coord: str) -> str:
    return "%s!%s" % (unquote_sheet(sheet), coord.replace("$", ""))


def pretty(k: str) -> str:
    sheet, coord = k.split("!", 1)
    return "%s!%s" % (quote_sheet(sheet), coord)


def parse_ref(ref: str, default_sheet: str) -> str:
    ref = ref.strip()
    if "!" in ref:
        sheet, coord = ref.rsplit("!", 1)
        return key(sheet, coord)
    return key(default_sheet, ref)


def expand(sheet: str, a: str, b: str | None, cap: int = 50000,
           truncated: set | None = None) -> list[str]:
    pa = split_coord(a)
    if not pa:
        return []
    if not b:
        return [key(sheet, "%s%d" % pa)]
    pb = split_coord(b)
    if not pb:
        return [key(sheet, "%s%d" % pa)]
    c1, c2 = sorted([col_to_num(pa[0]), col_to_num(pb[0])])
    r1, r2 = sorted([pa[1], pb[1]])
    if (c2 - c1 + 1) * (r2 - r1 + 1) > cap:
        if truncated is not None:
            truncated.add(sheet)
        return []
    return [
        key(sheet, "%s%d" % (num_to_col(c), r))
        for c in range(c1, c2 + 1)
        for r in range(r1, r2 + 1)
    ]


def refs_in(formula: str, default_sheet: str, truncated: set | None = None) -> list[str]:
    if not isinstance(formula, str) or not formula.startswith("="):
        return []
    body = STRING_RE.sub('""', formula)
    out = []
    for m in REF_RE.finditer(body):
        a = m.group("a")
        if a.upper() in NOT_A_REF:
            continue
        sheet = unquote_sheet(m.group("sheet")) if m.group("sheet") else default_sheet
        out.extend(expand(sheet, a, m.group("b"), truncated=truncated))
    return list(dict.fromkeys(out))


def cells_from_band_ref(ref: str) -> set[str]:
    if "!" not in ref:
        return set()
    sheet, coord = ref.rsplit("!", 1)
    sheet = unquote_sheet(sheet)
    if ":" in coord:
        a, b = coord.split(":", 1)
        return set(expand(sheet, a, b))
    return {key(sheet, coord)}


# A formula that is exactly one cell reference. BARE_LINK is unsigned (a plain
# link or `=+X`); BARE_ALIAS also admits a leading minus, because a negated
# alias still carries the source's construction up to sign.
_BARE_REF_BODY = r"(?:(?:'(?:[^']|'')+'|[A-Za-z_][A-Za-z0-9_. &-]*)!)?\$?[A-Z]{1,3}\$?[0-9]{1,7}"
BARE_LINK_RE = re.compile(r"^=\s*\+?\s*%s\s*$" % _BARE_REF_BODY)
BARE_ALIAS_RE = re.compile(r"^=\s*[+\-]?\s*%s\s*$" % _BARE_REF_BODY)


# --------------------------------------------------------------------------- workbook


class Book:
    """Flatten an xlsx to formula/value maps keyed as Sheet!A1."""

    def __init__(self, path: Path):
        self.path = Path(path)
        self.formula: dict[str, str] = {}
        self.value: dict[str, object] = {}
        self.sheets: list[str] = []
        wbf = openpyxl.load_workbook(path, data_only=False)
        wbv = openpyxl.load_workbook(path, data_only=True)
        try:
            for ws in wbf.worksheets:
                self.sheets.append(ws.title)
                vs = wbv[ws.title]
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value is None:
                            continue
                        k = key(ws.title, cell.coordinate)
                        if isinstance(cell.value, str) and cell.value.startswith("="):
                            self.formula[k] = cell.value
                            self.value[k] = vs[cell.coordinate].value
                        else:
                            self.value[k] = cell.value
        finally:
            wbf.close()
            wbv.close()
        self._labels: dict[str, str] = {}
        self._label_cols: dict[str, set] = {}
        self._stamp_cols: dict[str, set] = {}
        self._marker_cache: dict[str, dict] = {}
        self._max_cols: dict[str, int] = {}

    def max_col(self, sheet: str) -> int:
        """Rightmost used column on a sheet, so row scans never truncate scope."""
        if sheet in self._max_cols:
            return self._max_cols[sheet]
        hi = 1
        for k in list(self.value) + list(self.formula):
            s, coord = k.split("!", 1)
            if s != sheet:
                continue
            p = split_coord(coord)
            if p:
                hi = max(hi, col_to_num(p[0]))
        self._max_cols[sheet] = hi
        return hi

    def label_columns(self, sheet: str) -> set:
        """Columns that name rows, found from how the sheet is actually laid out.

        A label column carries text on many rows and numbers on almost none. This
        is structural rather than a blacklist, which matters because a scenario
        marker such as "Base" sitting between the label and the data is a real
        word and no token list will catch it. It also keeps side-by-side blocks
        working: each block's own label column qualifies on its own merits.
        """
        if sheet in self._label_cols:
            return self._label_cols[sheet]
        text = Counter()
        numeric = Counter()
        for k, v in self.value.items():
            s, coord = k.split("!", 1)
            if s != sheet:
                continue
            p = split_coord(coord)
            if not p:
                continue
            col = col_to_num(p[0])
            if isinstance(v, str) and v.strip() and not v.startswith("="):
                if not is_unit_stamp(v):
                    text[col] += 1
            elif isinstance(v, (int, float)):
                numeric[col] += 1
        cols = {c for c, n in text.items() if n >= 5 and n > numeric.get(c, 0)}
        self._label_cols[sheet] = cols
        return cols

    def stamp_columns(self, sheet: str) -> set:
        """Columns that carry units or bases rather than row names.

        0661's Assumptions column F holds "#", "date", "k€", "% of EBITDA",
        "years" - and, on two rows, the bare word "EBITDA", which is the basis
        of the "Acquisition multiple" assumption beside it, not a row name.
        No token list can reject "EBITDA" outright (it is a real row label
        elsewhere), but a column that is overwhelmingly stamps does not name
        rows even where one cell holds a real word.
        """
        if sheet in self._stamp_cols:
            return self._stamp_cols[sheet]
        stamps = Counter()
        other = Counter()
        for k, v in self.value.items():
            s, coord = k.split("!", 1)
            if s != sheet or k in self.formula:
                continue
            if not isinstance(v, str) or not v.strip():
                continue
            p = split_coord(coord)
            if not p:
                continue
            col = col_to_num(p[0])
            if is_unit_stamp(v) or is_soft_label(v):
                stamps[col] += 1
            else:
                other[col] += 1
        cols = {
            c for c, n in stamps.items()
            if n >= 5 and other.get(c, 0) <= max(1, n // 4)
        }
        self._stamp_cols[sheet] = cols
        return cols

    def linked_literal_text(self, k: str, hops: int = 4) -> str | None:
        """Text a bare single-reference link chain ultimately displays.

        A linked label - 0350's `Summary!G14 = Assumptions!B51` showing
        "Exit Year", 0666's `Calc_M!B386` linking "Initial Equity investment"
        - is a real row name and is accepted. Computed text (concatenations,
        lookups, scenario CHOOSEs) is not, because its cached display can
        describe a state rather than the row (0528's scenario side tables).
        """
        cur = k
        for _ in range(hops):
            formula = self.formula.get(cur)
            if formula is None:
                v = self.value.get(cur)
                return v if isinstance(v, str) and v.strip() else None
            if not BARE_LINK_RE.match(formula.strip()):
                return None
            refs = refs_in(formula, cur.split("!", 1)[0])
            if len(refs) != 1:
                return None
            cur = refs[0]
        return None

    def is_row_name(self, sheet: str, col: int, text: str) -> bool:
        """A row name identifies one row; a scenario marker repeats down a column.

        0463 cycles Base / Base Plus / Downside / Lender / Pancake beneath every
        assumption block, in the same column that legitimately labels those
        blocks. Naming a row "Base" identifies nothing.
        """
        return not bool(re.fullmatch(
            r"(?:base(?: case| plus| stressed)?|downside|lender|pancake|optimistic|pessimistic)",
            text.strip(),
            re.I,
        ))

    def has(self, k: str) -> bool:
        return k in self.value or k in self.formula

    def row_label(self, k: str) -> str:
        """The label governing this cell: nearest real name to its left on the row.

        Side-by-side assumption blocks share a row (0042 Summary: B/C keys, D/F
        areas, G/H unit costs, I/K fees). A sheet-wide "label column" list misses
        a short block (I has five names and more numbers than text) and then the
        nearest *accepted* column is a different table (K9 → D9 Parking Spaces).
        So walk left from the cell itself. Skip blanks, numbers, formulas, unit
        stamps, and repeating scenario markers. Take the first real name. Do not
        jump to the leftmost column on the row.
        """
        if k in self._labels:
            return self._labels[k]
        sheet, coord = k.split("!", 1)
        p = split_coord(coord)
        label = ""
        fallback = ""
        if p:
            col, row = col_to_num(p[0]), p[1]
            # Own cell first: a label sitting in a short block (I9) is the name.
            # Then walk left. Never jump to the leftmost column on the row.
            for c in range(col, 0, -1):
                candidate = key(sheet, "%s%d" % (num_to_col(c), row))
                v = self.value.get(candidate)
                # A formula's cached numeric, error or blank display is a
                # result, not a row name: keep walking. But a formula whose
                # cached display is semantic TEXT is a linked label - 0350's
                # `Summary!G14 = Assumptions!B51` shows "Exit Year", 0666's
                # `Calc_M!B386` links "Initial Equity investment" - and
                # skipping it lands on a different block's label ("Zip Code").
                # Cached text still runs the same unit/error/marker filters,
                # so an IFERROR's "n.a." or a scenario toggle's "Base" is
                # rejected exactly like literal text would be.
                if candidate in self.formula and not isinstance(v, str):
                    continue
                # Cached display text is only trusted when it is a genuine
                # linked label: a bare reference chain ending on literal text.
                # Computed cached text is skipped like a formula result
                # (0528's scenario side tables, 0605's review standard).
                if candidate in self.formula and isinstance(v, str):
                    if self.linked_literal_text(candidate) is None:
                        continue
                if isinstance(v, (int, float)) or v is None:
                    continue
                if not isinstance(v, str) or not v.strip() or v.startswith("="):
                    continue
                text = v.strip()
                # Calendar words are usually the unit column beside the real
                # name (0256 Summary!I6 "Year" beside "Uses & sources of
                # funds"; 0233 Model!G211 "date" beside "First Availability
                # Date") and are skipped. But on a row that carries nothing
                # else - 0233 Model!292's year header row names itself "Year"
                # in D292 - they are the nearest semantic name, so remember
                # the first one as a fallback rather than returning blank.
                if is_soft_label(text):
                    if not fallback:
                        fallback = text
                    continue
                if is_unit_stamp(text):
                    continue
                if re.fullmatch(r"scenario\s+\d+", text, re.I):
                    continue
                # A column that is overwhelmingly units and bases does not
                # name rows even where one cell holds a real word (0661's
                # F column stamps "EBITDA" beside "Acquisition multiple").
                if c in self.stamp_columns(sheet):
                    continue
                if not self.is_row_name(sheet, c, text):
                    continue
                label = text
                break
        label = label or fallback
        self._labels[k] = label
        return label

    def row_cells(self, sheet: str, rownum: int, lo: int = 1, hi: int | None = None) -> list[str]:
        # hi defaults to the sheet's real width. The old fixed cap of 80
        # columns silently truncated row scans at column CB, which is how
        # 0353's monthly waterfall rows lost their CC:DW periods.
        if hi is None:
            hi = self.max_col(sheet)
        return [key(sheet, "%s%d" % (num_to_col(c), rownum)) for c in range(lo, hi + 1)]


# Cells holding one of these are unit stamps, not row names. Taken from the
# segmentation stage's own list so the two agree on what a label is.
UNIT_TOKENS = {
    "aed", "usd", "eur", "gbp", "x", "%", "na", "n/a", "-", "bc",
    "000$", "$mm", "$bn", "$000s", "mm", "bn", "k", "m", "$", "yrs", "y",
    "none", "nil", "null", "yes", "tbd", "n.m.", "nm", "n.a.", "year",
    "years", "yr", "year(s)",
}


def is_soft_label(text: str) -> bool:
    """Calendar/type stamps: skipped while walking, usable as a last resort."""
    raw = text.strip()
    return len(raw) >= 4 and bool(re.fullmatch(r"(?:years?|year\(s\)|dates?)", raw, re.I))


def is_unit_stamp(text: str) -> bool:
    """Reject units, scale, and marker tokens so the reader keeps walking left.

    Exact listed tokens plus length < 3 are the original filter. The class
    matches below are what that filter missed: `flag`, `days`, `EURm`, `USD k`,
    `p&L`.
    """
    raw = text.strip()
    lowered = raw.lower()
    if lowered in UNIT_TOKENS or len(raw) < 3:
        return True
    if re.fullmatch(r"flags?", lowered):
        return True
    if re.fullmatch(r"days?", lowered):
        return True
    # Type/period stamps sitting between the label and the data. 0233's
    # Model!G211 holds "date" beside "First Availability Date".
    if re.fullmatch(r"dates?", lowered):
        return True
    # Physical energy/power units: 0233's Control!E20 holds "TWh" beside
    # "ElecLink flows".
    if re.fullmatch(r"[kmgt]?w(?:h|hrs?)?", lowered):
        return True
    # Scale stamps written as words or quantities: "in '000" (0233 Model!E292),
    # "'000", "000s".
    if re.fullmatch(r"(?:in\s*)?['’]?0{3}s?", lowered):
        return True
    # A currency code followed by a scale word: 0518's "AED million".
    if re.fullmatch(
        r"(?:usd|eur|gbp|aed|sar|zar|chf|jpy|cny|inr|cad|aud)"
        r"\s+(?:thousand|million|billion|trillion|bn|mn|mm|k|m)s?",
        lowered,
    ):
        return True
    # A country-prefixed currency symbol: 0646's `=BC` named range displays
    # "US$" beside the "Dividends" label.
    if re.fullmatch(r"[a-z]{0,3}\s*[$€£¥]", lowered):
        return True
    if re.fullmatch(r"p\s*&\s*l", lowered):
        return True
    if re.fullmatch(r"#(?:div/0!|n/a|value!|ref!|name\?|num!|null!)", lowered):
        return True
    # Currency-code scale stamps: 0620's Calc/FS/DCF grids hold "ZARm" in the
    # unit column beside every row name.
    if re.fullmatch(r"(usd|eur|gbp|aed|sar|zar)[kmbn]+", lowered):
        return True
    if re.fullmatch(r"(usd|eur|gbp|aed|\$)\s*['’]?(k|m|mm|bn|000s?)?", lowered):
        return True
    if re.fullmatch(r"[$€£]\s+in\s+(?:k|m|mm|bn|000s?)", lowered):
        return True
    if re.fullmatch(
        r"(?:usd|eur|gbp|aed|sar|[$€£])\s*(?:['’]?\s*0{3}s?|k|m|mm|bn)",
        lowered,
    ):
        return True
    if re.fullmatch(
        r"(?:usd|eur|gbp|aed|sar|[$€£])"
        r"\s*(?:/\s*[a-z][a-z0-9() -]*)+",
        lowered,
    ):
        return True
    # Currency-per-unit stamps may carry multi-word denominators: 0600's
    # InputT!D5 holds "USD/Metric T".
    if re.fullmatch(r"(?:usd|eur|gbp|aed|sar|[$€£])\s*/\s*[a-z0-9 .%-]+", lowered):
        return True
    if re.fullmatch(r"(?:usd|eur|gbp|aed|sar|[$€£])\s*\d{4}[kmbn]*", lowered):
        return True
    # Currency-then-zeros scale stamps with an optional trailing apostrophe:
    # 0635's Inputs!D56 holds "$000'" in the unit column beside the
    # "Land lease" label.
    if re.fullmatch(r"(?:usd|eur|gbp|aed|sar|[$€£])\s*['’]?0{3}s?['’]?", lowered):
        return True
    # Scale-prefixed currency stamps: 0598's Calculation!D265 holds "000A$"
    # beside the "Equity invested" label.
    if re.fullmatch(r"['’]?0{3}s?\s*[a-z]{0,3}\s*[$€£¥]", lowered):
        return True
    # A currency symbol with a trailing scale letter: 0622/0660's "US$m".
    if re.fullmatch(r"[a-z]{0,3}\s*[$€£¥]\s*(?:k|m|mm|mn|bn|b)", lowered):
        return True
    # Period-type qualifiers between the name and the data: 0669's
    # Calculations!C61 holds "historical" beside "Total Cost of Sales".
    if re.fullmatch(
        r"(?:historic(?:al)?|actuals?|forecasts?|projections?|projected|"
        r"budget(?:ed)?|estimated?|estimates)",
        lowered,
    ):
        return True
    # Basis stamps: 0661's Assumptions!F42 holds "% of EBITDA" as the basis
    # of the assumption beside it. "% net revenue" (0233's real row name)
    # carries no "of" and is kept.
    if re.fullmatch(r"%\s*of\s+[a-z0-9 .&/'’-]+", lowered):
        return True
    # Per-period quantity stamps and bare hour units: 0599's staffing block
    # stamps "hours / week" in the unit column beside every role name.
    if re.fullmatch(
        r"(?:hours?|hrs?|days?|weeks?|months?|shifts?|fte)\s*/\s*[a-z0-9 .-]+",
        lowered,
    ):
        return True
    if re.fullmatch(r"hours?|hrs?", lowered):
        return True
    # Template placeholders are not semantic names: 0599's dormant "Empty"
    # input slots.
    if re.fullmatch(r"empty", lowered):
        return True
    return False


def jsonable(value):
    if isinstance(value, (int, float, str, bool)) or value is None:
        if isinstance(value, float) and not math.isfinite(value):
            return str(value)
        return value
    return str(value)


def load_key(task_dir: Path):
    data = json.loads((task_dir / "tests" / "answer_key.json").read_text(encoding="utf-8"))
    return data.get("targets", {}), data.get("tolerance", {})


def task_id(task_dir: Path) -> str:
    return task_dir.name.split("-")[0]


def find_environment(task_dir: Path) -> Path:
    env = task_dir / "environment"
    # Same resolution contract as xl_artifact_paths.resolve_workbook_artifact:
    # .xlsx preferred, .xlsm accepted (workbooks 0635/0654 ship .xlsm inputs).
    for pattern in ("*.xlsx", "*.xlsm"):
        for path in sorted(env.glob(pattern)):
            return path
    raise FileNotFoundError(f"no delivered workbook under {env}")


def find_golden(task_dir: Path, explicit: str | None = None) -> Path:
    if explicit:
        return Path(explicit).resolve()
    tid = task_id(task_dir)
    root = task_dir.resolve().parents[1]
    for base in (root / "FCP Workbooks", REPO_ROOT.parent / "FCP Workbooks"):
        if not base.exists():
            continue
        for sub in sorted(base.iterdir()):
            for suffix in (".xlsx", ".xlsm"):
                cand = sub / f"{tid}{suffix}"
                if cand.exists():
                    return cand.resolve()
    raise FileNotFoundError(f"golden workbook not found for {tid}; pass --golden")


def run_dir(task_dir: Path, runs_root: Path = DEFAULT_RUNS_ROOT) -> Path:
    return runs_root / task_dir.name


def write_json(path: Path, payload: dict):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, default=jsonable) + "\n", encoding="utf-8")


def read_stage(task_dir: Path, name: str, runs_root: Path = DEFAULT_RUNS_ROOT) -> dict:
    return json.loads((run_dir(task_dir, runs_root) / f"{name}.json").read_text(encoding="utf-8"))


# --------------------------------------------------------------------------- selection


def regex_closure(gold: Book, targets: list[str], limit: int = 300000) -> list[str]:
    seen, order, queue = set(), [], deque(targets)
    while queue and len(seen) < limit:
        k = queue.popleft()
        if k in seen:
            continue
        seen.add(k)
        order.append(k)
        formula = gold.formula.get(k)
        if not formula:
            continue
        sheet = k.split("!", 1)[0]
        for ref in refs_in(formula, sheet):
            if ref not in seen:
                queue.append(ref)
    return order


def ast_closure_if_available(task_dir: Path, targets: list[str], ast_dir: Path | None):
    if not ast_dir:
        ast_dir = REPO_ROOT / "ast_out"
    wb = task_id(task_dir)
    graph_dir = ast_dir / wb
    if not (graph_dir / "nodes.csv").exists():
        return None, "missing_ast"
    try:
        sys.path.insert(0, str(REPO_ROOT))
        from xl_seg import model, project  # type: ignore

        graph = model.load(graph_dir, wb)
        cg = project.build(graph)
        seen, order, queue = set(), [], deque(targets)
        while queue:
            k = queue.popleft()
            if k in seen:
                continue
            seen.add(k)
            order.append(k)
            for pred in cg.radj.get(k, ()):
                if pred not in seen:
                    queue.append(pred)
        return order, "ok"
    except Exception as exc:  # fail closed to regex closure, but report it
        return None, f"ast_error:{exc}"


def blanked_formula_cells(gold: Book, delivered: Book, cells: list[str]) -> list[str]:
    return [k for k in cells if k in gold.formula and not delivered.has(k)]


def r1c1ish(formula: str, row: int, col: int) -> str:
    """Small local normalizer used only for grouping dragged formulas."""
    def repl(m):
        def one(raw: str) -> str:
            parsed = re.match(r"(\$?)([A-Z]{1,3})(\$?)(\d+)", raw)
            if not parsed:
                return "REF"
            abs_col, letters, abs_row, row_text = parsed.groups()
            ref_col, ref_row = col_to_num(letters), int(row_text)
            rpart = f"R{ref_row}" if abs_row else f"R[{ref_row - row}]"
            cpart = f"C{ref_col}" if abs_col else f"C[{ref_col - col}]"
            return rpart + cpart

        sheet = "S!" if m.group("sheet") else ""
        start = one(m.group("a"))
        end = one(m.group("b")) if m.group("b") else ""
        return sheet + start + (":" + end if end else "")
    return REF_RE.sub(repl, formula or "")


# ------------------------------------------------------------- copied scope
#
# The faithfulness reviewers judge a disclosed band against the *maximal
# same-mechanics span in the golden workbook*, including periods the selection
# never reached because the delivered file kept cached text or errors there, or
# because the graded closure did not pull them in. These helpers decide whether
# two golden cells carry the same copied mechanics and extend a band's stated
# scope to that maximal span. Selection semantics (which cells are blank and
# feed a graded answer) are untouched: only the stated scope widens.


def _mech_tokens(formula: str):
    """Split a formula into a normalized non-ref skeleton and its ref tokens."""
    body = STRING_RE.sub('""', formula)
    parts, refs, last = [], [], 0
    for m in REF_RE.finditer(body):
        if m.group("a").upper() in NOT_A_REF:
            continue
        parts.append(body[last : m.start()])
        refs.append(m)
        last = m.end()
    parts.append(body[last:])
    return re.sub(r"\s+", "", "".join(parts)).upper(), refs


def _mech_coord_match(ra: str, rb: str, dcol: int, drow: int) -> bool:
    pa = re.match(r"(\$?)([A-Z]{1,3})(\$?)(\d+)", ra.upper())
    pb = re.match(r"(\$?)([A-Z]{1,3})(\$?)(\d+)", rb.upper())
    if not pa or not pb:
        return False
    ca, rowa = col_to_num(pa.group(2)), int(pa.group(4))
    cb, rowb = col_to_num(pb.group(2)), int(pb.group(4))
    # A pinned anchor: both formulas read the same absolute cell even though
    # the author never typed the $ (0533's `D19=D18*B19`, `E19=E18*B19`).
    if ca == cb and rowa == rowb:
        return True
    # A dragged copy: matching $ flags, relative parts shifted by the offset.
    if pa.group(1) != pb.group(1) or pa.group(3) != pb.group(3):
        return False
    col_ok = (ca == cb) if pa.group(1) else (ca + dcol == cb)
    row_ok = (rowa == rowb) if pa.group(3) else (rowa + drow == rowb)
    return col_ok and row_ok


def _mech_ref_match(ma, mb, sheet: str, dcol: int, drow: int) -> bool:
    sa = unquote_sheet(ma.group("sheet")) if ma.group("sheet") else sheet
    sb = unquote_sheet(mb.group("sheet")) if mb.group("sheet") else sheet
    if sa != sb:
        return False
    for part in ("a", "b"):
        ra, rb = ma.group(part), mb.group(part)
        if (ra is None) != (rb is None):
            return False
        if ra is not None and not _mech_coord_match(ra, rb, dcol, drow):
            return False
    return True


def copy_equivalent(gold: Book, a: str, b: str) -> bool:
    """True when b's golden formula is a's copied to b's position.

    Each reference must either shift with the copy (a dragged relative ref) or
    stay pinned to the same absolute cell (a fixed anchor, with or without $).
    """
    fa, fb = gold.formula.get(a), gold.formula.get(b)
    if not fa or not fb:
        return False
    sheet_a, ca = a.split("!", 1)
    sheet_b, cb = b.split("!", 1)
    if sheet_a != sheet_b:
        return False
    pa, pb = split_coord(ca), split_coord(cb)
    if not pa or not pb:
        return False
    dcol, drow = col_to_num(pb[0]) - col_to_num(pa[0]), pb[1] - pa[1]
    ska, refsa = _mech_tokens(fa)
    skb, refsb = _mech_tokens(fb)
    if ska != skb or len(refsa) != len(refsb):
        return False
    return all(
        _mech_ref_match(x, y, sheet_a, dcol, drow) for x, y in zip(refsa, refsb)
    )


def bare_row_copy(gold: Book, k: str) -> bool:
    """A bare same-row reference to an earlier column: one hold-flat step.

    The first cell of a hold-flat run may skip a spacer column (0233's
    `DCF!AG31 = AE31` before `AH31 = AG31`), so strict copy-equivalence would
    exclude the seed cell the reviewers count as part of the run.
    """
    formula = gold.formula.get(k)
    if not formula:
        return False
    m = re.fullmatch(r"=\s*\+?\s*(\$?[A-Z]{1,3}\$?\d{1,7})\s*", formula)
    if not m:
        return False
    p = split_coord(k.split("!", 1)[1])
    rp = split_coord(m.group(1))
    return bool(p and rp and rp[1] == p[1] and col_to_num(rp[0]) < col_to_num(p[0]))


def copy_compatible(gold: Book, a: str, b: str) -> bool:
    return copy_equivalent(gold, a, b) or (
        bare_row_copy(gold, a) and bare_row_copy(gold, b)
    )


def _copy_lock_split(gold: Book, cells: list[str]):
    """Classify the representative formula's references as pinned or advancing.

    Compares the band's first two golden formulas token by token. A pinned
    reference reads the same absolute cell in every copy; an advancing one
    shifts with its position. Duplicate rendered addresses are disambiguated
    by occurrence, because 0658's `LBO!N84` reads N83 twice - once pinned,
    once advancing - and the representative alone cannot tell them apart.
    """
    if len(cells) < 2:
        return [], []
    a, b = cells[0], cells[1]
    fa, fb = gold.formula.get(a), gold.formula.get(b)
    if not fa or not fb:
        return [], []
    ska, refsa = _mech_tokens(fa)
    skb, refsb = _mech_tokens(fb)
    if ska != skb or len(refsa) != len(refsb):
        return [], []
    sheet = a.split("!", 1)[0]

    def shown(m) -> str:
        s = unquote_sheet(m.group("sheet")) if m.group("sheet") else sheet
        coord = m.group("a").replace("$", "")
        if m.group("b"):
            coord += ":" + m.group("b").replace("$", "")
        return "%s!%s" % (quote_sheet(s), coord)

    counts = Counter(shown(m) for m in refsa)
    seen: Counter = Counter()
    pinned, moving = [], []
    for x, y in zip(refsa, refsb):
        addr = shown(x)
        seen[addr] += 1
        desc = addr if counts[addr] == 1 else "occurrence %d of %s" % (seen[addr], addr)
        if shown(y) == addr:
            pinned.append(desc)
            continue
        # A range can pin one endpoint while the other advances with the
        # copy (0351's SUM($F$101:F101)). Name the fixed endpoint, or the
        # note's "every other reference shifts" clause would be false.
        if x.group("b") and y.group("b"):
            ax, bx = x.group("a").replace("$", ""), x.group("b").replace("$", "")
            ay, by = y.group("a").replace("$", ""), y.group("b").replace("$", "")
            if ax == ay and bx != by:
                pinned.append("the start %s of %s" % (ax, desc))
                continue
            if bx == by and ax != ay:
                pinned.append("the end %s of %s" % (bx, desc))
                continue
        moving.append(desc)
    return pinned, moving


def copy_lock_note(gold: Book, cells: list[str]) -> str:
    """A clause naming the references the copy pattern pins, when it pins any.

    0658's three blocked records each kept some references fixed
    (`LBO!$C$34:$C$37`, the first `$N83`, `$C$9`) while others advanced with
    the column, and the sentences gave only the representative addresses.
    Emitted only for a mixed pattern: an all-advancing or all-pinned copy
    needs no note.
    """
    pinned, moving = _copy_lock_split(gold, cells)
    # Endpoint-pinned ranges ("the start I250 of $I250:I250") are partially
    # moving: their other endpoint advances with the copy (0451's
    # SUM($I250:I250) -> SUM($I250:P250)). They must never trigger the
    # identical-calculation wording, which is only true when every reference
    # is a fully pinned plain address.
    all_plain = all(not d.startswith("the ") for d in pinned)
    if pinned and not moving and all_plain:
        # All references pinned: every copy holds the identical formula
        # (0441's =+IF($J238<=$F241,$E241/$F241,0) across J241:N241). Say so,
        # or a reader applying default copy translation shifts them all.
        return (
            ". As this formula is copied across the band, every reference "
            "stays fixed on exactly the cited cells, so each copy holds the "
            "identical calculation"
        )
    if not pinned:
        return ""
    listed = pinned[0] if len(pinned) == 1 else ", ".join(pinned[:-1]) + " and " + pinned[-1]
    plural = len(pinned) > 1
    # Endpoint descriptions ("the start F101 of ...") carry their own noun;
    # only plain addresses take the "the reference(s)" prefix.
    if all(not d.startswith("the ") for d in pinned):
        subject = "the reference%s %s" % ("s" if plural else "", listed)
    else:
        subject = listed
    return (
        ". As this formula is copied across the band, %s stay%s fixed "
        "while every other reference shifts with its position"
        % (subject, "" if plural else "s")
    )


def full_copied_scope(
    gold: Book, cells: list[str], targets: set | None = None, cap: int = 600
) -> list[str]:
    """Extend a contiguous same-row run to the maximal same-mechanics span.

    Walks outward over golden cells regardless of the delivered file's state,
    because a period kept as cached text or an error still belongs to the
    copied scope the sentence must state. Stops at a graded target so a wider
    stated scope can never claim a graded cell's construction.
    """
    if not cells:
        return cells
    sheet, coord = cells[0].split("!", 1)
    p0 = split_coord(coord)
    if not p0:
        return list(cells)
    row = p0[1]
    cols = sorted(
        col_to_num(split_coord(c.split("!", 1)[1])[0])
        for c in cells
        if split_coord(c.split("!", 1)[1])
    )
    if not cols:
        return list(cells)
    lo, hi = cols[0], cols[-1]
    targets = targets or set()

    def walk(start: int, step: int, edge_col: int) -> int:
        edge = key(sheet, "%s%d" % (num_to_col(edge_col), row))
        c, out, steps = start, edge_col, 0
        limit = gold.max_col(sheet)
        while 1 <= c <= limit and steps < cap:
            cand = key(sheet, "%s%d" % (num_to_col(c), row))
            if cand in targets or cand not in gold.formula:
                break
            if not copy_compatible(gold, edge, cand):
                break
            out, edge = c, cand
            c += step
            steps += 1
        return out

    out_lo = walk(lo - 1, -1, lo)
    out_hi = walk(hi + 1, 1, hi)
    return [key(sheet, "%s%d" % (num_to_col(c), row)) for c in range(out_lo, out_hi + 1)]


def full_copied_scope_vertical(
    gold: Book, cells: list[str], targets: set | None = None, cap: int = 600
) -> list[str]:
    """Vertical analog of full_copied_scope for copied-DOWN families.

    0528's `Inputs!U144:U146` is one contiguous copied-down run (`=U137*5`,
    `=U138*5`, `=U139*5`), but row-oriented banding fragments it into three
    one-cell records that each falsely claim their own scope. The stated
    scope of a same-column run must be its maximal same-mechanics span.
    """
    if not cells:
        return cells
    sheet, coord = cells[0].split("!", 1)
    p0 = split_coord(coord)
    if not p0:
        return list(cells)
    colnum = col_to_num(p0[0])
    rows = sorted(
        split_coord(c.split("!", 1)[1])[1]
        for c in cells
        if split_coord(c.split("!", 1)[1])
    )
    if not rows:
        return list(cells)
    lo, hi = rows[0], rows[-1]
    targets = targets or set()

    def walk(start: int, step: int, edge_row: int) -> int:
        edge = key(sheet, "%s%d" % (num_to_col(colnum), edge_row))
        r, out, steps = start, edge_row, 0
        while r >= 1 and steps < cap:
            cand = key(sheet, "%s%d" % (num_to_col(colnum), r))
            if cand in targets or cand not in gold.formula:
                break
            if not copy_equivalent(gold, edge, cand):
                break
            out, edge = r, cand
            r += step
            steps += 1
        return out

    out_lo = walk(lo - 1, -1, lo)
    out_hi = walk(hi + 1, 1, hi)
    return [key(sheet, "%s%d" % (num_to_col(colnum), r)) for r in range(out_lo, out_hi + 1)]


# ---------------------------------------------------------- graded closure


def bare_alias_ref(gold: Book, k: str) -> str | None:
    """The single cell a bare (optionally signed) reference formula reads."""
    formula = gold.formula.get(k)
    if not formula or len(formula) > 80 or not BARE_ALIAS_RE.match(formula.strip()):
        return None
    refs = refs_in(formula, k.split("!", 1)[0])
    return refs[0] if len(refs) == 1 else None


def graded_closure(gold: Book, targets) -> set[str]:
    """Cells whose construction is equivalent to a graded answer's.

    Two expansions, iterated to a fixpoint:

    - equality aliases: a bare-reference formula `=X` (optionally signed)
      makes the two cells the same figure, in either direction. 0441:
      `Model!O1 = +O24`, so disclosing O24's IRR calculation disclosed the
      graded O1; the same holds for a graded cell that is a bare reference
      into a band.
    - copied-method equivalents on the same row: a golden formula that is a
      closure cell's formula copied to another column carries the same
      method. 0672: `DCF!G58` is graded `DCF!C58` copied; 0620: `DCF!C37`
      leaked the mechanics of requested `DCF!E37`.

    Any disclosed record whose band, references, or rendered mechanics reach
    this set describes a graded answer and must be suppressed (0353's
    Summary!D50 leak is the direct case).
    """
    closure = set(targets)
    links = []
    for k, formula in gold.formula.items():
        ref = bare_alias_ref(gold, k)
        if ref:
            links.append((k, ref))
    expanded: set[str] = set()
    for _ in range(16):
        changed = False
        for a, b in links:
            if (a in closure) != (b in closure):
                closure.update((a, b))
                changed = True
        for c in sorted(closure - expanded):
            expanded.add(c)
            if c not in gold.formula:
                continue
            sheet, coord = c.split("!", 1)
            p = split_coord(coord)
            if not p:
                continue
            for other in gold.row_cells(sheet, p[1]):
                if (
                    other != c
                    and other not in closure
                    and other in gold.formula
                    and copy_compatible(gold, c, other)
                ):
                    closure.add(other)
                    changed = True
        if not changed:
            break
    return closure


def group_bands(gold: Book, cells: list[str], targets: set | None = None) -> list[dict]:
    """Band selected cells by row, splitting on any mechanics change.

    Grouping tests copied-mechanics equivalence rather than exact normalized
    patterns, so a pinned-anchor copy (`D19=D18*B19`, `E19=E18*B19`) and a
    hold-flat seed stay in one band. Each band's *stated* scope is then
    extended to the maximal same-mechanics span in the golden workbook, and
    bands whose stated scopes meet are merged, so one row never emits two
    records claiming the same copied range.
    """
    targets = targets or set()
    rows: dict[tuple, list[tuple[int, str]]] = defaultdict(list)
    for k in cells:
        sheet, coord = k.split("!", 1)
        p = split_coord(coord)
        if not p:
            continue
        col, row = p
        rows[(sheet, row)].append((col_to_num(col), k))

    bands_out = []
    for (sheet, row), entries in sorted(rows.items()):
        entries.sort()
        runs: list[list[tuple[int, str]]] = []
        run: list[tuple[int, str]] = []
        for item in entries:
            if run and not (
                item[0] == run[-1][0] + 1
                and copy_compatible(gold, run[-1][1], item[1])
            ):
                runs.append(run)
                run = []
            run.append(item)
        if run:
            runs.append(run)
        extended = [
            (r, full_copied_scope(gold, [k for _, k in r], targets)) for r in runs
        ]
        merged: list[tuple[list, list]] = []
        for r, full in extended:
            if merged:
                prev_run, prev_full = merged[-1]
                prev_hi = col_to_num(split_coord(prev_full[-1].split("!", 1)[1])[0])
                cur_lo = col_to_num(split_coord(full[0].split("!", 1)[1])[0])
                # Merge only when the extended scopes actually overlap, which
                # proves the walk crossed the gap on compatible mechanics.
                # Runs that merely touch are different copy families sitting
                # side by side (0233's Model!J224:R224 against S224:BT224,
                # which pin different rate cells).
                if cur_lo <= prev_hi:
                    span = full_copied_scope(
                        gold, [k for _, k in prev_run + r], targets
                    )
                    merged[-1] = (prev_run + r, span)
                    continue
            merged.append((r, full))
        for r, full in merged:
            bands_out.append(make_band(gold, sheet, row, r, full))
    return merge_vertical_singletons(gold, bands_out, targets)


def make_vertical_band(gold: Book, sheet: str, colnum: int, run_keys: list[str],
                       full: list[str] | None = None) -> dict:
    stated = full or run_keys
    rows = [split_coord(k.split("!", 1)[1])[1] for k in stated]
    col = num_to_col(colnum)
    ref = (
        f"{quote_sheet(sheet)}!{col}{rows[0]}"
        if rows[0] == rows[-1]
        else f"{quote_sheet(sheet)}!{col}{rows[0]}:{col}{rows[-1]}"
    )
    formulas = list(dict.fromkeys(gold.formula.get(k, "") for k in run_keys))
    first_row = split_coord(run_keys[0].split("!", 1)[1])[1]
    return {
        "band": ref,
        "sheet": sheet,
        "row": first_row,
        "col_lo": colnum,
        "col_hi": colnum,
        "orientation": "vertical",
        "cells": [pretty(k) for k in stated],
        "cell_keys": run_keys,
        "stated_cell_keys": stated,
        "label": gold.row_label(stated[0]) if stated else "",
        "pattern": r1c1ish(gold.formula.get(run_keys[0], ""), first_row, colnum),
        "formula_samples": formulas[:3],
        "values": [jsonable(gold.value.get(k)) for k in run_keys[:5]],
    }


def merge_vertical_singletons(gold: Book, bands: list[dict],
                              targets: set | None = None) -> list[dict]:
    """Rejoin copied-DOWN families that row banding left as one-cell bands.

    0528's `Inputs!U144:U146` was selected cell by cell, and each cell sits
    on its own row, so row banding emitted three singleton bands whose
    records each falsely claimed a self-contained scope. Contiguous
    same-column singletons in one copy family become a single vertical band,
    and a lone singleton's stated scope extends to its vertical span the
    same way horizontal bands extend theirs.
    """
    targets = targets or set()
    out, singles = [], []
    for band in bands:
        stated = band.get("stated_cell_keys") or band.get("cell_keys") or []
        if len(band.get("cell_keys") or []) == 1 and len(stated) == 1:
            singles.append(band)
        else:
            out.append(band)
    by_col: dict[tuple, list] = defaultdict(list)
    for band in singles:
        k = band["cell_keys"][0]
        sheet, coord = k.split("!", 1)
        p = split_coord(coord)
        if not p:
            out.append(band)
            continue
        by_col[(sheet, col_to_num(p[0]))].append((p[1], k, band))
    for (sheet, colnum), entries in sorted(by_col.items()):
        entries.sort()
        runs: list[list] = []
        run: list = []
        for item in entries:
            if run and not (
                item[0] == run[-1][0] + 1
                and copy_equivalent(gold, run[-1][1], item[1])
            ):
                runs.append(run)
                run = []
            run.append(item)
        if run:
            runs.append(run)
        for r in runs:
            keys = [k for _, k, _ in r]
            full = full_copied_scope_vertical(gold, keys, targets)
            if len(r) == 1 and len(full) <= 1:
                out.append(r[0][2])
            else:
                out.append(make_vertical_band(gold, sheet, colnum, keys, full))
    return out


def make_band(gold: Book, sheet: str, row: int, run: list[tuple[int, str]],
              full: list[str] | None = None) -> dict:
    cells = [k for _, k in run]
    stated = full or cells
    first_col = col_to_num(split_coord(stated[0].split("!", 1)[1])[0])
    last_col = col_to_num(split_coord(stated[-1].split("!", 1)[1])[0])
    ref = (
        f"{quote_sheet(sheet)}!{num_to_col(first_col)}{row}"
        if first_col == last_col
        else f"{quote_sheet(sheet)}!{num_to_col(first_col)}{row}:{num_to_col(last_col)}{row}"
    )
    formulas = list(dict.fromkeys(gold.formula.get(k, "") for k in cells))
    pattern = r1c1ish(gold.formula.get(cells[0], ""), row, run[0][0]) if cells else ""
    return {
        "band": ref,
        "sheet": sheet,
        "row": row,
        "col_lo": first_col,
        "col_hi": last_col,
        # The stated scope: the full copied-mechanics span the sentences and
        # reviewer records must quote. Selection semantics live in cell_keys.
        "cells": [pretty(k) for k in stated],
        "cell_keys": cells,
        "stated_cell_keys": stated,
        "label": gold.row_label(stated[0]) if stated else "",
        "pattern": pattern,
        "formula_samples": formulas[:3],
        "values": [jsonable(gold.value.get(k)) for k in cells[:5]],
    }


def select_payload(args) -> dict:
    task_dir = Path(args.task_dir).resolve()
    segmentation_mode = getattr(args, "segmentation_mode", "strict")
    pipeline_context = None
    pinned = bool(
        getattr(args, "release_root", None)
        or getattr(args, "source_generation_id", None)
        or getattr(args, "segmentation_generation_id", None)
    )
    delivered_path = find_environment(task_dir)
    if pinned:
        workbook = delivered_path.stem
        if workbook.endswith("-inputs"):
            workbook = workbook[:-7]
        from xl_release_publication import resolve_build_context

        pipeline_context = resolve_build_context(
            workbook,
            release_root=(
                Path(args.release_root) / workbook
                if getattr(args, "release_root", None) else None
            ),
            release_id=getattr(args, "release_id", None),
            source_root=Path(args.source_generation_root) / workbook,
            source_generation_id=getattr(args, "source_generation_id", None),
            segmentation_root=Path(args.seg_root) / workbook,
            segmentation_generation_id=getattr(
                args, "segmentation_generation_id", None
            ),
        )
    if segmentation_mode == "strict" and pipeline_context is None:
        missing = [
            name
            for name, value in (
                ("--golden", getattr(args, "golden", None)),
                ("--ast-dir", getattr(args, "ast_dir", None)),
                ("--seg-root", getattr(args, "seg_root", None)),
            )
            if not value
        ]
        if missing:
            raise SystemExit(
                "strict disclosure requires explicit " + ", ".join(missing)
            )
    golden_path = (
        Path(pipeline_context["source_path"])
        if pipeline_context is not None
        else find_golden(task_dir, args.golden)
    )
    generation = None
    seg_root = getattr(args, "seg_root", None)
    if seg_root:
        from xl_seg.publication import (
            INPUTS_POINTER_SCHEMA_VERSION,
            resolve_for_consumer,
        )

        workbook = Path(golden_path).stem
        expected_generation_id = getattr(
            args, "expected_generation_id", None
        )
        inputs_generation = {}
        packaged_manifest_path = (
            task_dir / "tests" / "segmentation_generation_manifest.json"
        )
        if packaged_manifest_path.is_file():
            packaged_manifest = json.loads(
                packaged_manifest_path.read_text(encoding="utf-8")
            )
            packaged_generation_id = packaged_manifest.get("generation_id")
            if (
                expected_generation_id is not None
                and packaged_generation_id != expected_generation_id
            ):
                raise SystemExit(
                    "packaged and requested segmentation generation IDs disagree"
                )
            expected_generation_id = packaged_generation_id
            inputs_generation_path = (
                task_dir / "tests" / "inputs_generation.json"
            )
            if not inputs_generation_path.is_file():
                raise SystemExit(
                    "packaged task has no inputs generation binding"
                )
            inputs_generation = json.loads(
                inputs_generation_path.read_text(encoding="utf-8")
            )
            delivered_sha256 = hashlib.sha256(
                delivered_path.read_bytes()
            ).hexdigest()
            packaged_manifest_sha256 = hashlib.sha256(
                packaged_manifest_path.read_bytes()
            ).hexdigest()
            if (
                inputs_generation.get("schema_version")
                != INPUTS_POINTER_SCHEMA_VERSION
                or inputs_generation.get("generation_id")
                != expected_generation_id
                or inputs_generation.get("inputs_file") != delivered_path.name
                or inputs_generation.get("inputs_sha256") != delivered_sha256
                or inputs_generation.get("generation_manifest_sha256")
                != packaged_manifest_sha256
            ):
                raise SystemExit(
                    "packaged inputs generation binding is invalid"
                )
        if pipeline_context is not None:
            generation_dir = Path(pipeline_context["segmentation_dir"])
            manifest = pipeline_context["segmentation_manifest"]
            if (
                expected_generation_id is not None
                and manifest.get("generation_id") != expected_generation_id
            ):
                raise SystemExit(
                    "packaged and pinned segmentation generation IDs disagree"
                )
            packaged_bindings = inputs_generation.get("pipeline_bindings")
            if packaged_manifest_path.is_file() and (
                packaged_bindings != pipeline_context["bindings"]
            ):
                raise SystemExit(
                    "packaged inputs do not preserve pinned pipeline bindings"
                )
        else:
            generation_dir, manifest = resolve_for_consumer(
                Path(seg_root) / workbook,
                mode=getattr(args, "segmentation_mode", "strict"),
                source_path=golden_path,
                ast_dir=(
                    Path(args.ast_dir) / workbook
                    if getattr(args, "ast_dir", None) else None
                ),
                require_pass=True,
                expected_generation_id=expected_generation_id,
            )
        generation = {
            "generation_id": (
                manifest.get("generation_id") if manifest is not None else None
            ),
            "directory": str(generation_dir),
            "mode": getattr(args, "segmentation_mode", "strict"),
        }
    gold = Book(golden_path)
    delivered = Book(delivered_path)
    targets_map, tolerance = load_key(task_dir)
    default_sheet = gold.sheets[0]
    targets = [parse_ref(t, default_sheet) for t in targets_map]

    regex = regex_closure(gold, targets)
    ast_root = (
        Path(pipeline_context["ast_root"])
        if pipeline_context is not None
        else Path(args.ast_dir) if args.ast_dir else None
    )
    ast, ast_status = ast_closure_if_available(task_dir, targets, ast_root)
    if segmentation_mode == "strict" and ast is None:
        raise SystemExit(f"strict disclosure requires bound AST closure: {ast_status}")
    closure = ast if ast is not None else regex
    selected = blanked_formula_cells(gold, delivered, closure)

    # Stated scopes stop at the graded closure, not just the raw targets, so
    # a widened span can never claim a graded alias's construction (0620,
    # 0672).
    bands = group_bands(gold, selected, graded_closure(gold, set(targets)))
    payload = {
        "schema_version": "1.0",
        "task": task_dir.name,
        "task_dir": str(task_dir),
        "golden": str(golden_path),
        "golden_sha256": hashlib.sha256(golden_path.read_bytes()).hexdigest(),
        "ast_dir": str(ast_root.resolve()) if ast_root else None,
        "seg_root": str(Path(seg_root).resolve()) if seg_root else None,
        "delivered": str(delivered_path),
        "targets": list(targets_map),
        "target_keys": targets,
        "tolerance": tolerance,
        "selection": {
            "closure_source": "ast" if ast is not None else "regex",
            "ast_status": ast_status,
            "regex_closure_cells": len(regex),
            "closure_cells": len(closure),
            "selected_cells": len(selected),
            "bands": len(bands),
            "non_blank_selected_cells": 0,
        },
        "bands": bands,
    }
    if generation is not None:
        payload["segmentation_generation"] = generation
    if pipeline_context is not None:
        payload["pipeline_bindings"] = pipeline_context["bindings"]
    return payload


def cmd_select(args):
    payload = select_payload(args)
    out = Path(args.out) if args.out else run_dir(Path(args.task_dir), Path(args.runs_root)) / "bands.json"
    write_json(out, payload)
    print(
        f"{payload['task']}: {payload['selection']['selected_cells']} selected cells, "
        f"{payload['selection']['bands']} bands ({payload['selection']['closure_source']} closure)"
    )


# --------------------------------------------------------------------------- custom methods

# Restored from the old custom-formula extractor, but used deterministically:
# the row label selects one method entry, then formula signatures decide whether
# the construction is a catalogued standard or genuinely out of catalogue.
METHOD_ROLE_PATTERNS = {
    "method_depreciation": (
        r"\bdepreciat", r"\bamorti[sz]", r"\bd&a\b", r"\basset depreciation\b",
    ),
    "method_interest": (
        r"\binterest", r"\bfinance cost", r"\bfinancing cost",
    ),
    "method_tax": (r"\btax(?:es|ation)?\b", r"\bcurrent tax\b", r"\bcash tax\b"),
    "method_revenue": (r"\brevenue", r"\bsales\b", r"\bturnover\b"),
    "method_operating_expense": (
        r"\bopex\b", r"\boperating expense", r"\bsg&a\b", r"\bsalar",
        r"\brent\b", r"\bcost of sales\b", r"\bmarketing\b", r"\bcrm\b",
        r"\bwebsite\b", r"\boffice\b", r"\bmaintenance\b", r"\bmonitoring fee\b",
        r"\bmanagement fee\b",
    ),
    "method_working_capital": (
        r"\bworking capital\b", r"\breceivable", r"\binventor",
        r"\bpayable", r"\bdso\b", r"\bdio\b", r"\bdpo\b",
    ),
    "method_capex": (
        r"\bcapex\b", r"\bcapital expenditure", r"\bfixed.asset addition",
    ),
    "method_debt_movement": (
        r"\bdebt\b", r"\bloan\b", r"\bdrawdown\b", r"\brepayment\b",
        r"\breimbursement\b", r"\bcash sweep\b",
    ),
    "method_discounting": (
        r"\bdiscount", r"\bpresent value\b", r"\bterminal value\b", r"\bnpv\b",
        r"\bvaluation\b", r"\benterprise value\b", r"\bequity value\b",
        r"\bmultiple\b",
    ),
    "method_returns": (
        r"\birr\b", r"\bxirr\b", r"\bmoic\b", r"\bmultiple of money\b",
    ),
}

# A role describes a calculated metric, not its assumption row. Without these
# exclusions "Interest Rate" becomes an interest calculation and "Tax Rate"
# becomes a tax calculation, flooding the custom path with assumptions.
METHOD_ROLE_EXCLUDES = {
    "method_interest": (r"\binterest rate\b", r"\brate\s*$"),
    "method_tax": (r"\btax rate\b",),
    "method_revenue": (r"\bgrowth", r"\bmargin", r"%"),
    "method_operating_expense": (r"\bgrowth", r"\bmargin", r"%"),
    "method_working_capital": (r"^\s*(dso|dio|dpo)\s*$", r"\bdays?\s*$"),
    "method_capex": (r"\bgrowth", r"\bmargin", r"%"),
    "method_discounting": (r"\bdiscount rate\b", r"\bdiscount period\b"),
}

METHOD_ENTRY_IDS = set(METHOD_ROLE_PATTERNS)
CONVENTION_ENTRY_IDS = {
    "discount_period", "inert_line", "terminal_value", "row_populated",
    "npv_timing", "aggregate_scope", "projection_rule", "distribution_policy",
    "liquidation_preference", "stake_scaling", "source_selection",
}
DISTRIBUTION_LABELS = ("dividend", "distribution", "shareholder payment")
# Wider than the entry's name, because the observed rows are labelled by what
# they pay out rather than by the convention that splits them: 0668's waterfall
# rows both read "Exit Equity Value".
EXIT_SPLIT_LABELS = (
    "preference", "preferred", "liquidation", "waterfall", "senior equity",
    "exit equity value", "exit proceeds", "proceeds to holders",
)
BORING_METHOD_LITERALS = {
    0.0, 1.0, -1.0, 2.0, 3.0, 4.0, 12.0, 100.0, 360.0, 365.0,
    1000.0, 10000.0, 1000000.0,
}


def method_roles(label: str) -> list[str]:
    """Registry method entries whose aliases match this visible row label."""
    lowered = (label or "").lower()
    if not lowered:
        return []
    hits = []
    for entry_id, patterns in METHOD_ROLE_PATTERNS.items():
        if not any(re.search(pattern, lowered) for pattern in patterns):
            continue
        if any(re.search(pattern, lowered) for pattern in METHOD_ROLE_EXCLUDES.get(entry_id, ())):
            continue
        hits.append(entry_id)
    return hits


def method_roles_for_band(gold: Book, band: dict) -> tuple[list[str], list[str]]:
    """Role from the row itself, falling back to two neighboring rows.

    Rows such as "Acq. 1" carry no finance word, while the section immediately
    above says Revenue. Neighborhood context is used only when the row itself
    gives no role, so it cannot override a clear label.
    """
    own = band.get("label") or ""
    direct = method_roles(own)
    if direct:
        return direct, [own]
    sheet, row = band.get("sheet"), band.get("row")
    col = band.get("col_lo", 1)
    context = []
    for nearby in range(max(1, int(row or 1) - 2), int(row or 1) + 3):
        cell = key(sheet, "%s%d" % (num_to_col(int(col)), nearby))
        label = gold.row_label(cell)
        if label and label not in context:
            context.append(label)
    context_text = " / ".join(context)
    if re.search(r"revenue.*ebitda.*margin", context_text, re.I):
        return ["method_revenue"], context
    hits = sorted({
        hit for context_label in context for hit in method_roles(context_label)
    })
    return hits, context


def role_triggers(label: str) -> dict[str, list[str]]:
    """Patterns that actually assigned a role after excludes."""
    lowered = (label or "").lower()
    if not lowered:
        return {}
    hits = {}
    for entry_id, patterns in METHOD_ROLE_PATTERNS.items():
        matched = [pattern for pattern in patterns if re.search(pattern, lowered)]
        if not matched:
            continue
        if any(re.search(pattern, lowered) for pattern in METHOD_ROLE_EXCLUDES.get(entry_id, ())):
            continue
        hits[entry_id] = matched
    return hits


def role_case_id(label: str, roles: list[str], role_context: list[str]) -> str:
    key = json.dumps(
        {
            "label": label or "",
            "roles": sorted(roles),
            "role_context": list(role_context or []),
        },
        sort_keys=True,
        ensure_ascii=False,
    )
    return hashlib.sha1(key.encode("utf-8")).hexdigest()[:12]


def prepare_method_items(gold: Book, bands: list[dict]) -> list[dict]:
    """Assign roles with the existing neighborhood rules. No arbitration yet."""
    prepared = []
    for original in bands:
        for band in method_subbands(gold, original):
            label = band.get("label") or ""
            if not label:
                prepared.append({
                    "band": band,
                    "label": "",
                    "roles": [],
                    "role_context": [],
                    "blank_label": True,
                })
                continue
            roles, role_context = method_roles_for_band(gold, band)
            prepared.append({
                "band": band,
                "label": label,
                "roles": roles,
                "role_context": role_context,
                "blank_label": False,
            })
    for item in prepared:
        if item["blank_label"] or item["roles"]:
            continue
        band = item["band"]
        nearby_roles = {
            other["roles"][0]
            for other in prepared
            if not other.get("blank_label")
            and len(other["roles"]) == 1
            and other["band"].get("sheet") == band.get("sheet")
            and other["band"].get("pattern") == band.get("pattern")
            and abs(int(other["band"].get("row", 0)) - int(band.get("row", 0))) <= 6
        }
        if len(nearby_roles) == 1:
            item["roles"] = list(nearby_roles)
            item["role_context"] = item["role_context"] + [
                "role propagated from same-shape neighboring row"
            ]
    return prepared


def triggered_by_for_item(item: dict) -> dict:
    own = item.get("label") or ""
    own_hits = role_triggers(own)
    neighbor_hits = {}
    for ctx in item.get("role_context") or []:
        if ctx == own or ctx == "role propagated from same-shape neighboring row":
            continue
        hits = role_triggers(ctx)
        if hits:
            neighbor_hits[ctx] = hits
    return {"own_label": own_hits, "neighbors": neighbor_hits}


def collect_ambiguous_role_cases(prepared: list[dict]) -> list[dict]:
    """One case per unique (label, roles, role_context) with two or more roles."""
    groups: dict[str, dict] = {}
    for item in prepared:
        if item.get("blank_label"):
            continue
        roles = item.get("roles") or []
        if len(roles) <= 1:
            continue
        case_id = role_case_id(item["label"], roles, item.get("role_context") or [])
        if case_id not in groups:
            groups[case_id] = {
                "case_id": case_id,
                "label": item["label"],
                "roles": sorted(roles),
                "role_context": list(item.get("role_context") or []),
                "band_count": 0,
                "sample_bands": [],
                "triggered_by": triggered_by_for_item(item),
                "questions": {
                    entry_id: registry().get(entry_id, {}).get("question", "")
                    for entry_id in sorted(roles)
                },
            }
        groups[case_id]["band_count"] += 1
        band_id = item["band"].get("band")
        samples = groups[case_id]["sample_bands"]
        if band_id and band_id not in samples and len(samples) < 5:
            samples.append(band_id)
    return sorted(groups.values(), key=lambda c: (c["label"], c["case_id"]))


def default_role_resolutions_path(task_dir: Path, runs_root: Path) -> Path:
    return run_dir(task_dir, runs_root) / "role_resolutions.json"


def load_role_resolutions(path: Path) -> dict | None:
    if not path.exists():
        return None
    payload = json.loads(path.read_text(encoding="utf-8"))
    if payload.get("validated") is not True:
        raise SystemExit(
            f"{path} has not passed roles-validate; run\n"
            f"  python3 disclose.py roles-validate --task-dir <task>\n"
            "first. Malformed or unvalidated arbitration files blocked "
            "workbooks 0350/0468/0522/0527/0534; detect refuses them now."
        )
    by_id = {}
    for row in payload.get("resolutions") or []:
        case_id = row.get("case_id")
        if case_id:
            by_id[case_id] = row
    payload["by_id"] = by_id
    return payload


# --------------------------------------------------------------------- roles-validate


ROLE_RESOLUTION_ROW_KEYS = {"case_id", "label", "chosen", "reason", "confidence"}
# Older arbitration prompts said "chosen_role"; normalize rather than reject.
LEGACY_ROLE_KEY = "chosen_role"


def validate_role_resolutions(
    path: Path, cases: list[dict]
) -> tuple[dict | None, list[str], list[str]]:
    """(normalized payload or None, errors, warnings) for an arbitration file.

    Models are asked to write exactly one JSON object, but the observed
    failure modes are prose wrappers and concatenated objects ("Extra data at
    line 2 column 1"). Extraction is tolerant -- the first balanced object is
    salvaged and the discard reported as a warning -- while the schema itself
    is strict: every row must name a known case_id at most once and choose one
    of that case's candidate roles. Any schema fault is an error carrying the
    exact detail the retry prompt needs.
    """
    errors: list[str] = []
    warnings: list[str] = []
    try:
        text = path.read_text(encoding="utf-8")
    except OSError as exc:
        return None, [f"cannot read {path}: {exc}"], []

    payload = None
    try:
        payload = json.loads(text)
    except json.JSONDecodeError as first_error:
        stripped = text.lstrip()
        start = stripped.find("{")
        if start < 0:
            return None, [f"no JSON object found: {first_error}"], []
        try:
            payload, end = json.JSONDecoder().raw_decode(stripped[start:])
        except json.JSONDecodeError:
            return None, [f"invalid JSON: {first_error}"], []
        prefix = stripped[:start].strip()
        trailing = stripped[start + end:].strip()
        warnings.append(
            "the file was not exactly one JSON object "
            f"(parser: {first_error}); salvaged the first object and "
            f"discarded {len(prefix)} prefix and {len(trailing)} trailing "
            "character(s). The arbitration agent must output one JSON object "
            "and nothing else."
        )

    if not isinstance(payload, dict):
        return None, errors + [
            "top level must be a JSON object with a 'resolutions' list"], warnings
    rows = payload.get("resolutions")
    if not isinstance(rows, list):
        return None, errors + ["'resolutions' must be a list"], warnings

    known = {case["case_id"]: case for case in cases}
    seen: set[str] = set()
    clean_rows: list[dict] = []
    for index, row in enumerate(rows):
        where = f"resolutions[{index}]"
        if not isinstance(row, dict):
            errors.append(f"{where}: must be an object")
            continue
        if LEGACY_ROLE_KEY in row and "chosen" not in row:
            row = dict(row)
            row["chosen"] = row.pop(LEGACY_ROLE_KEY)
            warnings.append(
                f"{where}: legacy key '{LEGACY_ROLE_KEY}' normalized to 'chosen'")
        unknown = sorted(set(row) - ROLE_RESOLUTION_ROW_KEYS)
        if unknown:
            errors.append(
                f"{where}: unknown key(s) {unknown}; allowed keys: "
                f"{sorted(ROLE_RESOLUTION_ROW_KEYS)}")
            continue
        case_id = row.get("case_id")
        if not isinstance(case_id, str) or not case_id.strip():
            errors.append(f"{where}: 'case_id' must be a non-empty string")
            continue
        if case_id in seen:
            errors.append(f"{where}: duplicate case_id {case_id!r}")
            continue
        seen.add(case_id)
        if known and case_id not in known:
            errors.append(
                f"{where}: case_id {case_id!r} is not one of the "
                f"{len(known)} case(s) in ambiguous_roles.json")
            continue
        chosen = row.get("chosen")
        if chosen is not None:
            # null means the agent abstained; candidates fall through in
            # registry order, exactly the pre-existing detect behavior.
            if not isinstance(chosen, str) or not chosen.strip():
                errors.append(
                    f"{where}: 'chosen' must be a candidate id or null")
                continue
            candidates = (known.get(case_id) or {}).get("roles") or []
            if known and chosen not in candidates:
                errors.append(
                    f"{where}: chosen {chosen!r} is not a candidate role for "
                    f"{case_id!r}; candidates: {candidates}")
                continue
        clean_rows.append(
            {key: row[key] for key in ROLE_RESOLUTION_ROW_KEYS if key in row})

    if errors:
        return None, errors, warnings
    normalized = {
        "schema_version": "1.0",
        "validated": True,
        "resolutions": clean_rows,
    }
    # Keep the arbitration provenance stamp: build_one.sh's gate-12 checker
    # requires it, and normalizing it away used to deadlock the gate.
    if isinstance(payload.get("agent_model"), str):
        normalized["agent_model"] = payload["agent_model"]
    return normalized, [], warnings


def require_role_resolutions(cases: list[dict], path: Path) -> dict:
    """Load resolutions when present. Missing or incomplete is not a hard stop.

    REGISTRY.md arbitration: try candidates in order and fall through on a
    decline. A resolutions file can still pin the first candidate.
    """
    payload = load_role_resolutions(path)
    if payload is None:
        return {"by_id": {}, "resolutions": []}
    return payload


def apply_role_resolution(item: dict, payload: dict) -> None:
    """Reduce a collision to one candidate when the agent picked a valid role."""
    roles = item.get("roles") or []
    if len(roles) <= 1:
        return
    case_id = role_case_id(item["label"], roles, item.get("role_context") or [])
    row = (payload.get("by_id") or {}).get(case_id) or {}
    chosen = row.get("chosen")
    if chosen in roles:
        item["roles"] = [chosen]
        item["role_context"] = list(item.get("role_context") or []) + [
            "role resolved by agent"
        ]


def walk_ast(node):
    if node is None:
        return
    yield node
    for arg in getattr(node, "args", ()):
        yield from walk_ast(arg)


def method_subbands(gold: Book, band: dict) -> list[dict]:
    """Selection already split mixed shapes by normalized formula pattern."""
    return [band]


def formula_profile(gold: Book, band: dict) -> dict:
    """One parsed representation shared by signature matching and prose.

    Parsing failure is reviewer-visible uncertainty. It never falls through to a
    guessed custom sentence.
    """
    cells = band.get("cell_keys", [])
    formulas = [gold.formula.get(c, "") for c in cells if gold.formula.get(c)]
    profile = {
        "complete": False,
        "band": band.get("band"),
        "cells": cells,
        "label": band.get("label") or (gold.row_label(cells[0]) if cells else ""),
        "formula": formulas[0] if formulas else "",
        "representative_cell": cells[0] if cells else "",
        "references": [],
        "reference_labels": [],
        "functions": [],
        "operators": [],
        "numbers": [],
        "pattern": band.get("pattern", ""),
        "error": None,
    }
    formula = profile["formula"]
    if not formula or not cells:
        profile["error"] = "missing_formula"
        return profile
    try:
        if str(REPO_ROOT) not in sys.path:
            sys.path.insert(0, str(REPO_ROOT))
        from xl_ast_graph import FormulaError, parse_formula  # type: ignore

        ast = parse_formula(formula)
    except Exception as exc:
        profile["error"] = f"parse_error:{exc}"
        return profile

    sheet = cells[0].split("!", 1)[0]
    references = refs_in(formula, sheet)
    labels = []
    for ref in references:
        label = gold.row_label(ref)
        if label and label not in labels:
            labels.append(label)
    functions, operators, numbers, text_literals, raw_references, literal_renderings = [], [], [], [], [], []
    for node in walk_ast(ast):
        if node.kind == "ref" and node.name not in raw_references:
            raw_references.append(node.name)
        if node.kind == "func" and node.name not in functions:
            functions.append(node.name)
        if node.kind in ("infix", "prefix", "postfix") and node.name not in operators:
            operators.append(node.name)
        if node.kind == "const" and node.name == "number":
            try:
                numbers.append(float(node.shape))
            except (TypeError, ValueError):
                pass
        if node.kind == "const" and node.name == "text":
            text_literals.append(str(node.shape))
        if node.kind == "const":
            rendered = describe_ast(node, gold, sheet, profile["label"])
            if rendered not in literal_renderings:
                literal_renderings.append(rendered)
    profile.update({
        "complete": True,
        "ast": ast,
        "references": references,
        "reference_labels": labels,
        "functions": functions,
        "operators": operators,
        "numbers": numbers,
        "text_literals": text_literals,
        "raw_references": raw_references,
        "literal_renderings": literal_renderings,
        "ast_nodes": sum(1 for _ in walk_ast(ast)),
    })
    return profile


def profile_has_label(profile: dict, *patterns: str) -> bool:
    text = " / ".join(profile.get("reference_labels", [])).lower()
    return any(re.search(pattern, text) for pattern in patterns)


def profile_has_all(profile: dict, *groups: tuple[str, ...]) -> bool:
    return all(profile_has_label(profile, *group) for group in groups)


def references_same_row(profile: dict) -> bool:
    cells = profile.get("cells") or []
    if not cells:
        return False
    sheet, coord = cells[0].split("!", 1)
    p = split_coord(coord)
    if not p:
        return False
    row = p[1]
    for ref in profile.get("references", []):
        rsheet, rcoord = ref.split("!", 1)
        rp = split_coord(rcoord)
        if rsheet == sheet and rp and rp[1] == row:
            return True
    return False


def ast_ref_cell(node, profile: dict) -> str | None:
    if node.kind != "ref":
        return None
    cells = profile.get("cells") or []
    sheet = cells[0].split("!", 1)[0] if cells else ""
    refs = refs_in("=" + node.name, sheet)
    return refs[0] if len(refs) == 1 else None


def exact_prior_period_growth(profile: dict) -> bool:
    """Exactly previous cell × (1 + one other reference), in either order."""
    ast = profile.get("ast")
    cells = profile.get("cells") or []
    if ast is None or not cells or ast.kind != "infix" or ast.name != "*" or len(ast.args) != 2:
        return False
    sheet, coord = cells[0].split("!", 1)
    p = split_coord(coord)
    if not p:
        return False
    current_col, current_row = col_to_num(p[0]), p[1]

    def is_previous(node) -> bool:
        ref = ast_ref_cell(node, profile)
        if not ref:
            return False
        rsheet, rcoord = ref.split("!", 1)
        rp = split_coord(rcoord)
        return bool(
            rsheet == sheet and rp
            and rp[1] == current_row
            and col_to_num(rp[0]) == current_col - 1
        )

    def is_one_plus_rate(node) -> bool:
        if node.kind != "infix" or node.name not in ("+", "-") or len(node.args) != 2:
            return False
        left, right = node.args
        const = left if left.kind == "const" else right if right.kind == "const" else None
        other = right if const is left else left
        try:
            one = const is not None and abs(float(const.shape) - 1.0) < 1e-12
        except (TypeError, ValueError):
            one = False
        return one and other.kind == "ref"

    left, right = ast.args
    return (is_previous(left) and is_one_plus_rate(right)) or (
        is_previous(right) and is_one_plus_rate(left)
    )


def root_is_function(profile: dict, name: str) -> bool:
    ast = profile.get("ast")
    return bool(ast is not None and ast.kind == "func" and ast.name.upper() == name.upper())


def catalog_signature(entry_id: str, profile: dict) -> str | None:
    """The catalogued variant matched by a complete profile, else None.

    These signatures are deliberately conservative. A positive match suppresses
    a custom hint, so each one requires both the formula shape and semantically
    appropriate labelled ingredients.
    """
    funcs = set(profile.get("functions", []))
    ops = set(profile.get("operators", []))
    formula = profile.get("formula", "").upper()

    if entry_id == "method_depreciation":
        if not profile_has_label(profile, r"\blife\b", r"useful"):
            return None
        if "AVERAGE" in funcs and profile_has_label(profile, r"opening", r"bop") and profile_has_label(profile, r"closing", r"eop"):
            return "average_balance"
        if any(abs(n - 0.5) < 1e-12 for n in profile.get("numbers", [])) and profile_has_label(profile, r"capex", r"capital"):
            return "midyear_capex"
        if profile_has_label(profile, r"capex", r"capital") and "/" in ops:
            return "bop_plus_capex_over_life"
        if profile_has_label(profile, r"asset", r"balance", r"basis") and "/" in ops:
            return "bop_over_life"
        return None

    if entry_id == "method_interest":
        if "IF" in funcs or "IFS" in funcs:
            return None
        if not profile_has_all(
            profile,
            (r"\brate\b", r"sofr", r"euribor", r"libor"),
            (r"balance", r"debt", r"cash", r"loan", r"principal", r"draw"),
        ):
            return None
        if "AVERAGE" in funcs:
            return "average_balance"
        if any(abs(n - 0.5) < 1e-12 for n in profile.get("numbers", [])):
            return "midyear_flow"
        if profile_has_label(profile, r"draw", r"repay") and ("+" in ops or "-" in ops):
            return "full_draw"
        if "*" in ops:
            return "opening_balance"
        return None

    if entry_id == "method_tax":
        if "IF" in funcs or "IFS" in funcs:
            return None
        if not profile_has_label(profile, r"\brate\b", r"tax %"):
            return None
        if profile_has_label(profile, r"\bnol\b", r"loss") and "MAX" in funcs:
            return "taxable_income_after_losses"
        # A clamp without a labelled loss balance is not any catalogued
        # variant: describing `MAX(EBT,0)*rate` as plain pretax_profit would
        # suppress the zero floor the agent cannot see (0256 CalcA!H56).
        if funcs & {"MAX", "MIN"}:
            return None
        if profile_has_label(profile, r"taxable income"):
            return "taxable_income"
        if profile_has_label(profile, r"\bebt\b", r"pre.?tax", r"profit before tax"):
            return "pretax_profit"
        return None

    if entry_id == "method_revenue":
        if "IF" in funcs or "IFS" in funcs:
            return None
        if exact_prior_period_growth(profile) and profile_has_label(profile, r"growth"):
            return "prior_period_growth"
        if profile_has_all(profile, (r"price", r"value"), (r"volume", r"units?")) and "*" in ops:
            return "price_times_volume"
        if root_is_function(profile, "SUM") and profile_has_label(profile, r"revenue", r"sales", r"turnover"):
            return "segment_sum"
        if profile_has_all(profile, (r"capacity",), (r"utili[sz]ation",), (r"price",)) and "*" in ops:
            return "capacity_utilisation"
        return None

    if entry_id == "method_operating_expense":
        if "IF" in funcs or "IFS" in funcs:
            return None
        if exact_prior_period_growth(profile) and profile_has_label(profile, r"growth", r"inflation"):
            return "prior_period_growth"
        if profile_has_all(profile, (r"revenue", r"sales"), (r"%", r"margin", r"rate")) and "*" in ops:
            return "percent_of_revenue"
        if profile_has_all(profile, (r"fixed",), (r"variable",)) and "+" in ops:
            return "fixed_plus_variable"
        if root_is_function(profile, "SUM"):
            return "component_sum"
        return None

    if entry_id == "method_working_capital":
        if "AVERAGE" in funcs and profile_has_label(profile, r"days?", r"dso", r"dio", r"dpo"):
            return "average_driver_days"
        if profile_has_label(profile, r"days?", r"dso", r"dio", r"dpo") and "/" in ops:
            return "days_of_driver"
        if profile_has_label(profile, r"%", r"rate") and "*" in ops:
            return "percent_of_driver"
        if profile_has_all(profile, (r"opening", r"bop"), (r"closing", r"eop")) and "-" in ops:
            return "balance_delta"
        return None

    if entry_id == "method_capex":
        if exact_prior_period_growth(profile) and profile_has_label(profile, r"growth"):
            return "prior_period_growth"
        if profile_has_all(profile, (r"revenue", r"sales"), (r"%", r"rate")) and "*" in ops:
            return "percent_of_revenue"
        if profile_has_all(profile, (r"maintenance",), (r"growth",)) and "+" in ops:
            return "maintenance_plus_growth"
        return None

    if entry_id == "method_debt_movement":
        if "MAX" in funcs and profile_has_label(profile, r"funding", r"shortfall"):
            return "required_funding"
        if profile_has_all(profile, (r"opening", r"bop", r"principal"), (r"amorti[sz]ation rate",)) and "*" in ops:
            return "fixed_amortisation"
        if profile_has_label(profile, r"maturity") and profile_has_label(profile, r"opening", r"principal"):
            return "maturity_repayment"
        if "MIN" in funcs and profile_has_all(profile, (r"cash",), (r"sweep",)):
            return "cash_sweep"
        return None

    if entry_id == "method_discounting":
        if "NPV" in funcs or "XNPV" in funcs:
            return "npv_plus_t0" if re.search(r"\)\s*[+\-]", formula) else "periodic_discount"
        if profile_has_all(profile, (r"growth",), (r"discount", r"wacc")) and "/" in ops:
            return "perpetuity_growth"
        if profile_has_label(profile, r"multiple") and "*" in ops:
            return "exit_multiple"
        if "^" in ops and profile_has_label(profile, r"discount", r"wacc"):
            return "midyear_discount" if any(abs(n - 0.5) < 1e-12 for n in profile.get("numbers", [])) else "periodic_discount"
        return None

    if entry_id == "method_returns":
        if "XIRR" in funcs:
            return "xirr_on_dated_series"
        if "IRR" in funcs:
            return "irr_on_series"
        if "/" in ops and profile_has_all(profile, (r"proceeds", r"return"), (r"invest", r"equity")):
            return "multiple_of_money"
        return None

    return None


def ast_is_ref_like(node) -> bool:
    if node.kind == "ref":
        return True
    if node.kind == "prefix" and node.name in ("+", "-") and len(node.args) == 1:
        return ast_is_ref_like(node.args[0])
    if node.kind == "infix" and node.name in ("*", "/") and len(node.args) == 2:
        left, right = node.args
        pairs = [(left, right)]
        if node.name == "*":
            pairs.append((right, left))
        for ref_node, const_node in pairs:
            if not (
                ast_is_ref_like(ref_node)
                and const_node.kind == "const"
                and const_node.name == "number"
            ):
                continue
            try:
                if abs(float(const_node.shape)) in BORING_METHOD_LITERALS:
                    return True
            except (TypeError, ValueError):
                continue
    return False


def structural_reason(profile: dict) -> str | None:
    """Structural/definitional plumbing is not custom finance logic."""
    ast = profile.get("ast")
    if ast is None:
        return None
    if ast_is_ref_like(ast):
        return "link_sign_or_unit_scale"
    funcs = set(profile.get("functions", []))
    ops = set(profile.get("operators", []))
    if funcs & {"INDEX", "MATCH", "VLOOKUP", "HLOOKUP", "XLOOKUP", "INDIRECT", "OFFSET"}:
        return "lookup_or_transcription"
    if profile.get("text_literals") and funcs & {"IF", "IFS", "OR", "AND", "ROUND"}:
        return "diagnostic_or_warning_logic"
    if funcs <= {"SUM"} and ops <= {"+", "-"}:
        return "component_aggregation"
    if not funcs and ops and ops <= {"+", "-"}:
        return "composition_or_rollforward"
    label = (profile.get("label") or "").lower()
    if not (funcs & {"IF", "IFS", "IFERROR", "CHOOSE", "SWITCH", "LOOKUP", "XLOOKUP"}) and any(
        token in label for token in ("/revenue", "% of ", " margin", " ratio")
    ):
        return "ratio_or_assumption_row"
    return None


def nonboring_numbers(profile: dict) -> list[float]:
    return [
        number for number in profile.get("numbers", [])
        if number not in BORING_METHOD_LITERALS
        and -number not in BORING_METHOD_LITERALS
    ]


def confident_custom_reason(entry_id: str, profile: dict) -> str | None:
    """Positive evidence that a no-match is genuinely out of catalogue.

    Merely failing a signature is uncertainty, not custom logic. Claiming the
    band requires a branch, a load-bearing literal, or a role-specific alternate
    driver/predicate that the catalogue explicitly excludes.
    """
    funcs = set(profile.get("functions", []))
    numbers = nonboring_numbers(profile)
    if funcs & {"IF", "IFS", "IFERROR", "CHOOSE", "SWITCH", "LOOKUP", "XLOOKUP"}:
        return "out_of_catalogue_branch"
    if numbers:
        return "embedded_threshold_or_parameter"

    if entry_id == "method_depreciation":
        if profile_has_all(
            profile,
            (r"revenue", r"sales"),
            (r"depreciat", r"capex", r"amorti[sz]"),
        ) and not profile_has_label(profile, r"\blife\b", r"useful"):
            return "alternate_revenue_driver"

    if entry_id == "method_interest":
        if profile_has_label(profile, r"revenue", r"sales") and not profile_has_label(
            profile, r"balance", r"debt", r"cash", r"loan", r"principal"
        ):
            return "alternate_revenue_driver"

    if entry_id == "method_tax":
        if "MAX" in funcs or "MIN" in funcs:
            return "uncatalogued_tax_clamp"

    if entry_id == "method_revenue":
        growth_rows = [
            label for label in profile.get("reference_labels", [])
            if re.search(r"growth|inflation|uplift", label, re.I)
        ]
        if len(growth_rows) > 1:
            return "multiple_growth_drivers"

    if entry_id == "method_operating_expense":
        if "SUM" in funcs and ("*" in set(profile.get("operators", [])) or "/" in set(profile.get("operators", []))):
            return "cumulative_base_escalation"
        if funcs & {"MAX", "MIN"}:
            return "uncatalogued_floor_or_cap"
        growth_rows = [
            label for label in profile.get("reference_labels", [])
            if re.search(r"growth|inflation|uplift|escalat", label, re.I)
        ]
        if len(growth_rows) > 1:
            return "multiple_growth_drivers"

    if entry_id == "method_working_capital":
        own = (profile.get("label") or "").lower()
        refs = " / ".join(profile.get("reference_labels", [])).lower()
        if "payable" in own and "revenue" in refs:
            return "nonstandard_payables_driver"
        if references_same_row(profile):
            return "recursive_working_capital_rule"

    if entry_id == "method_capex":
        if len(profile.get("reference_labels", [])) > 2:
            return "item_driver_build"

    if entry_id == "method_debt_movement":
        if profile.get("operators") and len(profile.get("references", [])) > 2:
            return "uncatalogued_debt_waterfall"

    if entry_id == "method_discounting":
        if "AVERAGE" in funcs or profile_has_label(profile, r"ebitda", r"multiple"):
            return "uncatalogued_valuation_basis"

    if entry_id == "method_returns":
        if "^" in set(profile.get("operators", [])):
            return "uncatalogued_annualisation"

    return None


def reference_lock_qualifier(raw: str) -> tuple[str, str]:
    """Describe Excel absolute-reference locks without exposing formula syntax."""
    match = REF_RE.search(raw)
    if not match:
        return "", ""
    coordinates = [match.group("a")]
    if match.group("b"):
        coordinates.append(match.group("b"))
    locks = []
    for coordinate in coordinates:
        parsed = re.fullmatch(r"(\$?)[A-Z]{1,3}(\$?)[0-9]{1,7}", coordinate)
        locks.append((bool(parsed and parsed.group(1)), bool(parsed and parsed.group(2))))
    if len(locks) == 1:
        fixed_column, fixed_row = locks[0]
        if fixed_column and fixed_row:
            return "fixed ", ""
        if fixed_column:
            return "", " with its column fixed when copied"
        if fixed_row:
            return "", " with its row fixed when copied"
        return "", ""
    if locks and all(fixed_column and fixed_row for fixed_column, fixed_row in locks):
        return "fixed ", ""
    details = []
    for coordinate, (fixed_column, fixed_row) in zip(coordinates, locks):
        if not fixed_column and not fixed_row:
            continue
        clean = coordinate.replace("$", "")
        if fixed_column and fixed_row:
            scope = "row and column"
        elif fixed_column:
            scope = "column"
        else:
            scope = "row"
        details.append(f"{clean}'s {scope}")
    return "", (" with " + " and ".join(details) + " fixed when copied") if details else ""


def describe_ref(gold: Book, raw: str, sheet: str, own_label: str) -> str:
    refs = refs_in("=" + raw, sheet)
    lock_prefix, lock_suffix = reference_lock_qualifier(raw)
    if len(refs) == 1:
        value = gold.value.get(refs[0])
        if (
            refs[0] not in gold.formula
            and isinstance(value, str)
            and value.strip()
        ):
            return f"{lock_prefix}cell {pretty(refs[0])}{lock_suffix} containing {q(value.strip())}"
    ref_labels = [gold.row_label(ref) for ref in refs]
    labels = []
    for label in ref_labels:
        if label and label not in labels:
            labels.append(label)
    if refs:
        if len(refs) == 1 and ":" not in raw:
            location = lock_prefix + "cell " + pretty(refs[0]) + lock_suffix
        else:
            first_sheet, first_coord = pretty(refs[0]).rsplit("!", 1)
            _, last_coord = pretty(refs[-1]).rsplit("!", 1)
            location = (
                f"{lock_prefix}range {first_sheet}!{first_coord}:{last_coord}"
                f"{lock_suffix}"
            )
        if labels and all(ref_labels):
            named = [q(label) for label in labels]
            label_text = named[0] if len(named) == 1 else ", ".join(named[:-1]) + " and " + named[-1]
            noun = "row" if len(named) == 1 else "rows"
            return f"{location} on the {noun} labelled {label_text}"
        return location
    return f'the named range {q(raw)}'


def describe_ast(node, gold: Book, sheet: str, own_label: str) -> str:
    """Render a parsed Excel AST as deterministic English, never formula text."""
    kind = node.kind
    if kind == "ref":
        return describe_ref(gold, node.name, sheet, own_label)
    if kind == "const":
        if node.name == "text":
            if node.shape == "":
                return "leave the cell blank"
            return q(str(node.shape))
        if node.name == "logical":
            return "true" if node.shape else "false"
        return str(node.value)
    if kind == "missing":
        return "a blank value"
    args = [describe_ast(arg, gold, sheet, own_label) for arg in node.args]
    if kind == "prefix":
        return f"take the negative of ({args[0]})" if node.name == "-" else args[0]
    if kind == "postfix":
        return f"{args[0]} percent" if node.name == "%" else args[0]
    if kind == "infix" and len(args) == 2:
        words = {
            "+": "add ({a}) and ({b})",
            "-": "subtract ({b}) from ({a})",
            "*": "multiply ({a}) by ({b})",
            "/": "divide ({a}) by ({b})",
            "^": "raise ({a}) to the power ({b})",
            "=": "({a}) equals ({b})",
            "<>": "({a}) does not equal ({b})",
            ">": "({a}) is greater than ({b})",
            "<": "({a}) is less than ({b})",
            ">=": "({a}) is at least ({b})",
            "<=": "({a}) is at most ({b})",
            "&": "join ({a}) with ({b})",
        }
        return words.get(node.name, "combine ({a}) with ({b})").format(a=args[0], b=args[1])
    if kind == "func":
        name = node.name.upper()
        if name == "IF" and len(args) >= 3:
            return f"when ({args[0]}) is true, use ({args[1]}); otherwise use ({args[2]})"
        if name == "IF" and len(args) == 2:
            return f"when ({args[0]}) is true, use ({args[1]}); otherwise use FALSE"
        if name == "IFERROR" and len(args) >= 2:
            return f"use ({args[0]}), or use ({args[1]}) if that calculation errors"
        if name == "MAX":
            return "take the greater of " + " and ".join(f"({arg})" for arg in args)
        if name == "MIN":
            return "take the lesser of " + " and ".join(f"({arg})" for arg in args)
        if name == "AVERAGE":
            return "take the mean of " + " and ".join(f"({arg})" for arg in args)
        if name == "SUM":
            return "add " + " and ".join(f"({arg})" for arg in args)
        if name == "CHOOSE" and len(args) >= 2:
            options = "; ".join(
                f"option {index}: ({value})" for index, value in enumerate(args[1:], 1)
            )
            return f"use ({args[0]}) as the option number, then choose {options}"
        if name == "MROUND" and len(args) >= 2:
            return f"round ({args[0]}) to the nearest multiple of ({args[1]})"
        if name == "SUMPRODUCT":
            return "sum the products of corresponding values in " + " and ".join(
                f"({arg})" for arg in args
            )
        if name == "SUMIFS" and len(args) >= 3:
            pairs = [
                f"({args[index]}) matches ({args[index + 1]})"
                for index in range(1, len(args) - 1, 2)
            ]
            return f"total the values in ({args[0]}) where " + " and ".join(pairs)
        if name == "SUMIF" and len(args) >= 2:
            summed = args[2] if len(args) >= 3 else args[0]
            return f"total the values in ({summed}) where ({args[0]}) matches ({args[1]})"
        if name == "AND":
            return " and ".join(args)
        if name == "OR":
            return " or ".join(args)
        if name in ("NPV", "XNPV", "IRR", "XIRR"):
            return f"the {name.lower()} of " + ", ".join(args)
        return f"{name.lower()} applied to " + ", ".join(args)
    if kind in ("union", "array"):
        return ", ".join(args)
    return " then ".join(args)


def custom_sentence_fields(profile: dict, gold: Book, lock_cells: list | None = None) -> dict:
    cells = profile.get("cells") or []
    sheet = cells[0].split("!", 1)[0] if cells else ""
    ast = profile.get("ast")
    steps = describe_ast(ast, gold, sheet, profile.get("label", "")) if ast is not None else ""
    if steps:
        # A copy pattern pinning some references while others advance must
        # say so, or the representative addresses misstate the band (0658).
        # Use the stated copied span when it is wider than the profile's
        # anchor cells, mirroring faithcheck's lock-cell selection (0248).
        span = lock_cells if lock_cells and len(lock_cells) > 1 else cells
        steps += copy_lock_note(gold, span)
    reference_renderings = [
        describe_ref(gold, raw, sheet, profile.get("label", ""))
        for raw in profile.get("raw_references", [])
    ]
    coverage_complete = all(rendered in steps for rendered in reference_renderings)
    coverage_complete = coverage_complete and all(
        rendered in steps for rendered in profile.get("literal_renderings", [])
    )
    band = profile.get("band") or compact_cells([pretty(c) for c in cells])
    representative = pretty(profile.get("representative_cell")) if profile.get("representative_cell") else band
    return {
        "label": q(profile.get("label", "")),
        "band": band,
        "representative": representative,
        "calculation_kind": (
            "calculation" if len(cells) == 1 else "copied-column calculation"
        ),
        "steps": steps,
        "_coverage_complete": coverage_complete,
    }


def expand_method_band(gold: Book, delivered: Book, band: dict) -> dict:
    """Include the full contiguous blank copied-formula run on the target row."""
    cells = list(band.get("cell_keys") or [])
    if not cells:
        return band
    first = split_coord(cells[0].split("!", 1)[1])
    if not first:
        return band
    sheet, row = band.get("sheet"), int(band.get("row") or first[1])
    pattern = band.get("pattern") or r1c1ish(
        gold.formula.get(cells[0], ""), row, col_to_num(first[0])
    )
    cols = [col_to_num(split_coord(c.split("!", 1)[1])[0]) for c in cells]
    lo, hi = min(cols), max(cols)

    def matches(col: int) -> bool:
        candidate = key(sheet, f"{num_to_col(col)}{row}")
        formula = gold.formula.get(candidate, "")
        return bool(
            formula
            and not delivered.has(candidate)
            and r1c1ish(formula, row, col) == pattern
        )

    while lo > 1 and matches(lo - 1):
        lo -= 1
    while matches(hi + 1):
        hi += 1
    if lo == min(cols) and hi == max(cols):
        return band
    return make_band(
        gold,
        sheet,
        row,
        pattern,
        [(col, key(sheet, f"{num_to_col(col)}{row}")) for col in range(lo, hi + 1)],
    )


def detect_custom_methods(
    gold: Book,
    delivered: Book,
    bands: list[dict],
    targets: set[str],
    resolutions_path: Path | None = None,
):
    """Run before convention detectors; only confident no-matches claim a band."""
    records, assessments, claimed = [], [], set()
    prepared = prepare_method_items(gold, bands)
    cases = collect_ambiguous_role_cases(prepared)
    if cases and resolutions_path is not None:
        payload = require_role_resolutions(cases, resolutions_path)
        for item in prepared:
            apply_role_resolution(item, payload)
    for item in prepared:
        band = expand_method_band(gold, delivered, item["band"])
        label = item["label"]
        if item.get("blank_label"):
            assessments.append({
                "band": band.get("band"),
                "cells": band.get("cells", []),
                "cell_keys": band.get("cell_keys", []),
                "label": "",
                "roles": [],
                "status": "unclassified",
                "reason": "blank_target_label",
            })
            continue
        roles = item["roles"]
        role_context = item["role_context"]
        base = {
            "band": band.get("band"),
            "cells": band.get("cells", []),
            "cell_keys": band.get("cell_keys", []),
            "label": label,
            "roles": roles,
            "role_context": role_context,
        }
        if not roles:
            assessments.append({
                **base, "status": "unclassified", "reason": "no_method_role"
            })
            continue
        # Registry order within Section 2. A resolutions file, when it chooses,
        # only pins the first candidate; a miss still falls through.
        ordered = [r for r in METHOD_ROLE_PATTERNS if r in roles]
        last_assessment = None
        claimed_here = False
        for entry_id in ordered:
            rec, assessment = try_custom_role(gold, band, label, base, entry_id)
            last_assessment = assessment
            if rec is not None:
                records.append(rec)
                claimed.update(rec.get("cell_keys", []))
                assessments.append(assessment)
                claimed_here = True
                break
        if not claimed_here and last_assessment is not None:
            if len(ordered) > 1:
                last_assessment = {
                    **base, "status": "unclassified", "reason": "ambiguous_role",
                    "tried": ordered,
                    "last": last_assessment,
                }
            assessments.append(last_assessment)
    return records, assessments, claimed


def try_custom_role(gold: Book, band: dict, label: str, base: dict, entry_id: str):
    """One method-entry attempt. A None record means fall through to the next."""
    profile = formula_profile(gold, band)
    if not profile.get("complete"):
        return None, {
            **base, "entry": entry_id, "status": "unclassified",
            "reason": profile.get("error", "profile_incomplete"),
        }
    variant = catalog_signature(entry_id, profile)
    if variant:
        return None, {
            **base, "entry": entry_id, "status": "standard",
            "variant": variant,
        }
    plumbing = structural_reason(profile)
    if plumbing:
        return None, {
            **base, "entry": entry_id, "status": "structural",
            "reason": plumbing,
        }
    custom_reason = confident_custom_reason(entry_id, profile)
    if not custom_reason:
        return None, {
            **base, "entry": entry_id, "status": "unclassified",
            "reason": "no_catalogue_match_but_no_positive_custom_signal",
        }
    fields = custom_sentence_fields(
        profile, gold, lock_cells=band.get("stated_cell_keys"))
    coverage_complete = bool(fields.pop("_coverage_complete", False))
    rec = record(
        entry_id,
        "out_of_catalogue",
        band.get("cell_keys", []),
        profile["formula"],
        registry().get(entry_id, {}).get("alternatives", []),
        f"Out-of-catalogue {entry_id} method on row {label!r}.",
        fields=fields,
        stated_cells=band.get("stated_cell_keys"),
    )
    rec.update({
        "source": "custom_method_detector",
        "custom_reason": custom_reason,
        "coverage_complete": coverage_complete,
        "method_profile": {
            k: v for k, v in profile.items()
            if k not in ("ast",)
        },
    })
    return rec, {
        **base, "entry": entry_id, "status": "out_of_catalogue",
        "reason": custom_reason,
    }


def calibrate_legacy_custom(task_dir: Path, assessments: list[dict], selected: set[str]) -> dict:
    """Compare deterministic outcomes with old custom_logic band labels.

    The old prose is never imported. It is a labelled coverage set only, and is
    deliberately not treated as a correctness oracle because 0523 proved it can
    be wrong.
    """
    path = task_dir / "tests" / "formula_hints.json"
    if not path.exists():
        return {"bands": 0, "outcomes": {}}
    hints = json.loads(path.read_text(encoding="utf-8")).get("hints", [])
    old = [
        band
        for hint in hints
        if "custom_logic" in hint.get("classes", [])
        for band in hint.get("bands", [])
    ]
    rows = []
    for band in old:
        cells = cells_from_band_ref(band)
        overlaps = [
            a for a in assessments
            if cells & set(a.get("cell_keys", []))
        ]
        statuses = sorted({a.get("status", "unknown") for a in overlaps})
        if not (cells & selected):
            outcome = "not_selected"
        elif "out_of_catalogue" in statuses:
            outcome = "out_of_catalogue"
        elif "standard" in statuses:
            outcome = "standard"
        elif "structural" in statuses:
            outcome = "structural"
        elif "unclassified" in statuses:
            outcome = "unclassified"
        else:
            outcome = "missed"
        rows.append({"band": band, "outcome": outcome, "statuses": statuses})
    return {
        "bands": len(old),
        "outcomes": dict(Counter(row["outcome"] for row in rows)),
        "details": rows,
    }


# --------------------------------------------------------------------------- detectors


HEADER_RE = re.compile(r"(statement|schedule|summary|assumptions|inputs)\s*$|:\s*$", re.I)


def selected_keys(selection: dict) -> set[str]:
    return {k for band in selection.get("bands", []) for k in band.get("cell_keys", [])}


def _fmt(cells, n=4):
    cells = list(cells)
    shown = ", ".join(pretty(c) for c in cells[:n])
    return shown + (f" +{len(cells) - n} more" if len(cells) > n else "")


def detect_discount_period(gold: Book, scope: set[str]) -> list[dict]:
    out, seen_rows = [], set()
    for k in sorted(scope):
        label = gold.row_label(k).lower()
        if "discount period" not in label and "discount factor period" not in label:
            continue
        sheet, coord = k.split("!", 1)
        p = split_coord(coord)
        if not p or (sheet, p[1]) in seen_rows:
            continue
        seen_rows.add((sheet, p[1]))
        row = [c for c in gold.row_cells(sheet, p[1]) if c in gold.formula or c in gold.value]
        nums = [gold.value.get(c) for c in row if isinstance(gold.value.get(c), (int, float))]
        formulas = [gold.formula[c] for c in row if c in gold.formula]
        value = None
        if any(isinstance(v, float) and abs(v - round(v)) > 0.4 for v in nums):
            value = "mid_year"
        elif nums and all(isinstance(v, (int, float)) and abs(v - round(v)) < 1e-9 for v in nums):
            value = "year_end"
        if value is None and any("/2" in f for f in formulas):
            value = "mid_year"
        if value:
            out.append(record(
                "discount_period", value, [c for c in row if c in gold.formula],
                formulas[0] if formulas else f"cached values {nums[:4]}",
                ["mid_year", "year_end", "day_weighted"],
                f"Row labelled {gold.row_label(k)!r}.",
                fields={"label": q(gold.row_label(k))},
            ))
    return out


def detect_inert_line(gold: Book, scope: set[str]) -> list[dict]:
    out, seen_rows = [], set()
    for k in sorted(scope):
        label = gold.row_label(k)
        if not label or len(label) < 4 or HEADER_RE.search(label) or label.isupper():
            continue
        if not re.search(
            r"waterfall|guard|flag|minimum cash|less:\s*minimum|closing balance|"
            r"cash closing|reserve|trap",
            label, re.I,
        ):
            continue
        sheet, coord = k.split("!", 1)
        p = split_coord(coord)
        if not p or (sheet, p[1]) in seen_rows:
            continue
        seen_rows.add((sheet, p[1]))
        row = [c for c in gold.row_cells(sheet, p[1]) if c in gold.formula]
        vals = [gold.value.get(c) for c in row]
        nums = [v for v in vals if isinstance(v, (int, float))]
        if not row or len(nums) < 2:
            continue
        if not all(abs(v) < 1e-9 for v in nums):
            continue
        # Claim only cells whose cached value is numeric zero. A guard row can
        # carry label-mirror formula cells caching text (0635's
        # Caclulation!B69 = B$18 -> "Post PPA flag"); the always_zero sentence
        # must be true of every claimed cell, matching the registry Detection
        # test ("every cached value is zero").
        zero_cells = [
            c for c in row if isinstance(gold.value.get(c), (int, float))
        ]
        out.append(record(
            "inert_line", "always_zero", zero_cells, gold.formula[zero_cells[0]],
            ["always_zero", "charged_once", "charged_every_period"],
            f"Row labelled {label!r} evaluates to zero in all periods.",
            fields={"label": q(label)},
        ))
    return out


def detect_terminal_value(gold: Book, scope: set[str]) -> list[dict]:
    out, seen_rows = [], set()
    for k in sorted(scope):
        label = gold.row_label(k).lower()
        if "terminal value" not in label and "exit value" not in label:
            continue
        sheet, coord = k.split("!", 1)
        p = split_coord(coord)
        if not p or (sheet, p[1]) in seen_rows:
            continue
        seen_rows.add((sheet, p[1]))
        row = [c for c in gold.row_cells(sheet, p[1]) if c in gold.formula]
        row_label = gold.row_label(k)
        if not row:
            out.append(record(
                "terminal_value", "absent", [], "no formula anywhere on the row",
                ["absent", "perpetuity_growth", "exit_multiple", "other"],
                f"Row labelled {row_label!r} is never populated.",
                row_ref=f"{quote_sheet(sheet)}!row {p[1]}",
                fields={"label": q(row_label)},
            ))
            continue
        formula = gold.formula[row[0]]
        if re.search(r"\(1\s*\+\s*[^)]+\)\s*/\s*\(", formula):
            value = "perpetuity_growth"
        elif re.search(r"[Mm]ultiple|Comps", formula):
            value = "exit_multiple"
        else:
            value = "other"
        out.append(record(
            "terminal_value", value, row, formula,
            ["absent", "perpetuity_growth", "exit_multiple", "other"],
            f"Row labelled {row_label!r}, populated in {len(row)} column(s).",
            fields={"label": q(row_label)},
        ))
    return out


def _row_once(gold: Book, k: str, seen_rows: set) -> tuple | None:
    """Resolve a scope cell to its sheet, row number and formula cells, once."""
    sheet, coord = k.split("!", 1)
    p = split_coord(coord)
    if not p or (sheet, p[1]) in seen_rows:
        return None
    seen_rows.add((sheet, p[1]))
    row = [c for c in gold.row_cells(sheet, p[1]) if c in gold.formula]
    return (sheet, p[1], col_to_num(p[0]), row) if row else None


def _ingredient_labels(gold: Book, formula: str, sheet: str, row: int, col: int) -> list[str]:
    return [
        gold.row_label(key(rs, "%s%d" % (num_to_col(rc), rr))).lower()
        for rs, rc, rr in _ingredients(formula, sheet, row, col)
    ]


DIST_CASH_TOKENS = (
    "cash available", "available cash", "cash left",
    "available for distribution", "available for dividend",
)


def _dist_cash_label(text: str) -> bool:
    return any(t in text for t in DIST_CASH_TOKENS)


def _parse_formula_ast(formula: str):
    try:
        if str(REPO_ROOT) not in sys.path:
            sys.path.insert(0, str(REPO_ROOT))
        from xl_ast_graph import parse_formula  # type: ignore

        return parse_formula(formula)
    except Exception:
        return None


def _max_floor_operand(node):
    """The X of MAX(X, 0) or MAX(0, X); None for any other shape."""
    if node is None or getattr(node, "kind", None) != "func":
        return None
    if node.name.upper() != "MAX" or len(node.args) != 2:
        return None

    def is_zero(n):
        try:
            return n.kind == "const" and n.name == "number" and abs(float(n.shape)) < 1e-12
        except (TypeError, ValueError):
            return False

    a, b = node.args
    if is_zero(b):
        return a
    if is_zero(a):
        return b
    return None


def _single_ref_cell(node, sheet: str) -> str | None:
    """The one cell a bare (possibly sign-prefixed) reference node points at."""
    if node is None:
        return None
    if getattr(node, "kind", None) == "prefix" and node.name == "+" and len(node.args) == 1:
        node = node.args[0]
    if getattr(node, "kind", None) != "ref":
        return None
    refs = refs_in("=" + node.name, sheet)
    return refs[0] if len(refs) == 1 else None


def _floor_note(gold: Book, formula: str, sheet: str) -> str:
    """A truthful floor clause when the row reads a MAX(...,0)-floored source.

    0256's `H29 = H28` and 0350's `= G157 * C66` both distribute a source that
    is already held at zero; saying nothing about the floor left the registry
    sentence free to claim negatives were possible. The clause names only the
    floor, never the source's full construction.
    """
    for ref in refs_in(formula, sheet):
        src = gold.formula.get(ref)
        if not src:
            continue
        if _max_floor_operand(_parse_formula_ast(src)) is None:
            continue
        label = gold.row_label(ref)
        if label and _dist_cash_label(label.lower()):
            return (
                "; the row labelled %s floors its own calculation at zero" % q(label)
            )
    return ""


def detect_distribution_policy(gold: Book, scope: set[str],
                               targets: set | None = None) -> list[dict]:
    """Which rule sizes a dividend or distribution row.

    A residual and a payout ratio are both ordinary practice and the delivered
    file keeps neither formula, so the choice is invisible. Worse, a surviving
    payout-ratio row on another sheet actively points at the wrong one.

    Only two shapes may still use a named-value sentence, because only there is
    the fixed wording provably complete: an exact `MAX(cash, 0)` residual and
    an exact two-reference payout product. Every other distribution shape - a
    `MIN` entitlement cap, a pass-through copy of a floored row, a share
    multiplier over floored cash, a lesser-of-two-rows clamp - is rendered
    mechanically from its parsed formula so no cap, floor or reference is
    dropped (0256, 0350, 0352, 0353, 0646).
    """
    values = [
        "residual_cash_floored", "residual_cash_unfloored", "payout_ratio",
        "capped_at_retained_earnings", "ownership_band_rate",
        "first_period_only", "formula_mechanics",
    ]
    targets = targets or set()
    out, seen_rows = [], set()
    for k in sorted(scope):
        label = gold.row_label(k)
        lowered = label.lower()
        if not any(t in lowered for t in DISTRIBUTION_LABELS):
            continue
        # "Available cash for dividends" contains the word but is the pool,
        # not the payment. 0638 rows 117-118 are that trap; row 119 is the
        # distribution.
        if re.search(r"available (cash|for)|cash available|retained earnings", lowered):
            continue
        resolved = _row_once(gold, k, seen_rows)
        if not resolved:
            continue
        sheet, rownum, col, row = resolved
        # Not the leftmost formula. 0638 row 119 holds `=BC` in D (a named
        # unit) and `=SUM(H119:AD119)` in F; the period cells `=H118` are
        # the residual. Classify on the strongest payment shape present.
        def dist_rank(cell: str) -> int:
            f = gold.formula[cell]
            u = f.upper()
            c = split_coord(cell.split("!", 1)[1])
            labs = _ingredient_labels(
                gold, f, sheet, rownum, col_to_num(c[0]) if c else col
            )
            cash = any(
                (
                    "cash available" in l
                    or "available cash" in l
                    or "cash left" in l
                    or "available for distribution" in l
                    or "available for dividend" in l
                )
                for l in labs
            )
            if cash and "MAX(" in u and re.search(r",\s*0\s*\)", u):
                return 5
            if "MIN(" in u and any("retained" in l for l in labs):
                return 4
            if any(("payout" in l or "dividend per" in l) for l in labs) and "*" in f:
                return 3
            if cash:
                return 2
            if re.match(r"^=\s*SUM\(", f.strip(), re.I) or not refs_in(f, sheet):
                return -1
            return 0

        cell = max(row, key=dist_rank)
        if dist_rank(cell) < 0:
            continue
        formula = gold.formula[cell]
        upper = formula.upper()
        labels = _ingredient_labels(
            gold, formula, sheet, rownum,
            col_to_num(split_coord(cell.split("!", 1)[1])[0]),
        )
        # Every value needs a positive signal. There is no default: a dividend
        # row whose shape says nothing is left uncovered, because asserting a
        # rule off the absence of a token is how this entry would become the
        # next `projection_rule` over-disclosure.
        cash_available = any(_dist_cash_label(l) for l in labels)
        payout_labelled = any(("payout" in l or "dividend per" in l) for l in labels)
        ast = _parse_formula_ast(formula)
        # The record claims the maximal contiguous same-mechanics span around
        # the classified cell, so a surviving unit stamp (`=BC` in 0638's D119)
        # never joins the band and the stated scope never truncates at an
        # arbitrary column.
        run = full_copied_scope(gold, [cell], targets)
        value, fields, mechanics_refs = None, {"label": q(label)}, []
        # Exact plain floored residual: MAX of a single cash-available
        # reference against zero. Anything else - MIN caps, pass-through
        # copies, share multipliers, lesser-of clamps - is not fully
        # described by any fixed sentence and must be rendered.
        floor_ref = _single_ref_cell(_max_floor_operand(ast), sheet)
        if floor_ref and _dist_cash_label(gold.row_label(floor_ref).lower()):
            value = "residual_cash_floored"
        if value is None and re.fullmatch(
            r'=IF\([^,]+="<20%",0\.7,IF\([^,]+=">80%",1(?:\.0)?,0\.8\)\)',
            re.sub(r"\s+|\$", "", formula.upper()),
        ):
            value = "ownership_band_rate"
        if value is None and (
            ast is not None
            and getattr(ast, "kind", None) == "infix"
            and ast.name == "*"
            and len(ast.args) == 2
        ):
            # Exact payout product: two bare references, one on a payout row.
            factor_cells = [_single_ref_cell(arg, sheet) for arg in ast.args]
            if all(factor_cells) and any(
                ("payout" in gold.row_label(fc).lower()
                 or "dividend per" in gold.row_label(fc).lower())
                for fc in factor_cells
            ):
                value = "payout_ratio"
        if value is None and ast is not None and (
            cash_available or payout_labelled or "MIN(" in upper
        ):
            steps = describe_ast(ast, gold, sheet, label)
            if _single_ref_cell(ast, sheet):
                ref_cell = _single_ref_cell(ast, sheet)
                own_p = split_coord(cell.split("!", 1)[1])
                ref_p = split_coord(ref_cell.split("!", 1)[1])
                same_col = bool(
                    own_p and ref_p and own_p[0] == ref_p[0]
                    and ref_cell.split("!", 1)[0] == sheet
                )
                lead = (
                    "copy the same-column value from"
                    if same_col
                    else "take the value from"
                )
                steps = "%s %s" % (lead, steps)
            steps += _floor_note(gold, formula, sheet)
            steps += copy_lock_note(gold, run)
            value = "formula_mechanics"
            fields = {
                "label": q(label),
                "band": compact_cells([pretty(c) for c in run]).strip("`"),
                "representative": pretty(cell),
                "steps": steps,
            }
            mechanics_refs = refs_in(formula, sheet)
        if value is None and (
            len(row) == 1
            and refs_in(formula, sheet)
            and not re.match(r"^=\s*SUM\(", formula.strip(), re.I)
            and not re.search(r"\b(IF|IFS|CHOOSE|LOOKUP|XLOOKUP)\s*\(", formula, re.I)
        ):
            value = "first_period_only"
        if value is None:
            continue
        # Attach the payment cells, not a surviving unit stamp. 0638 D119 is
        # `=BC` and remains as "EURm" in the delivered file; including it made
        # Ship when think the distribution row had survived.
        pay_cells = [c for c in row if dist_rank(c) >= 2] or [
            c for c in row if dist_rank(c) >= 0
        ]
        if value == "ownership_band_rate":
            refs = refs_in(formula, sheet)
            if refs:
                fields["ingredient"] = pretty(refs[0])
        rec = record(
            "distribution_policy", value, run or pay_cells, formula, values,
            f"Row labelled {label!r} sizes its distribution by {value}.",
            fields=fields,
        )
        if mechanics_refs:
            rec["mechanics_references"] = mechanics_refs
        out.append(rec)
    return out


def detect_liquidation_preference(gold: Book, scope: set[str]) -> list[dict]:
    """How exit proceeds divide between preferred and common holders.

    Ordered ahead of `stake_scaling`, which ships on `always` and would describe
    one of these rows as an ownership-share multiplication.
    """
    values = [
        "participating", "non_participating", "pro_rata_no_preference",
        "capped_participation",
    ]
    out, seen_rows = [], set()
    for k in sorted(scope):
        label = gold.row_label(k)
        lowered = label.lower()
        if not any(t in lowered for t in EXIT_SPLIT_LABELS):
            continue
        # Exit or cap-table block only. Elsewhere these words are commentary.
        sheet_l = k.split("!", 1)[0].lower()
        if "cap table" not in sheet_l and "exit" not in lowered and "waterfall" not in lowered:
            continue
        resolved = _row_once(gold, k, seen_rows)
        if not resolved:
            continue
        _sheet, _rownum, _col, row = resolved
        # Not the leftmost formula on the row. A waterfall row commonly holds a
        # plain pro-rata column beside the column that carries the preference -
        # 0668 row 43 is `=C42/$C$24` in C and the solved preference in D - so
        # classify on the most specific shape present, not on whichever comes
        # first.
        def rank(cell: str) -> int:
            f = gold.formula[cell]
            u = f.upper()
            if "MIN(" in u:
                return 3
            if "MAX(" in u or re.search(r"\bIF\s*\(", u):
                return 2
            if solves_for_divisor(f):
                return 1
            return 0
        cell = max(row, key=rank)
        rnk = rank(cell)
        # A product or a SUM labelled "Exit Equity Value" is not a split
        # convention. 0644's Flags row 420 is that shape. 0668 ships because
        # D43/D49 carry the solved preference (rank 1).
        if rnk == 0 and "cap table" not in k.split("!", 1)[0].lower():
            continue
        if rnk == 0 and "/" not in gold.formula[cell]:
            continue
        formula = gold.formula[cell]
        value = {
            3: "capped_participation",
            2: "non_participating",
            1: "participating",
            0: "pro_rata_no_preference",
        }[rnk]
        out.append(record(
            "liquidation_preference", value, row, formula, values,
            f"Row labelled {label!r} splits exit proceeds as {value}.",
            fields={"label": q(label)},
        ))
    return out


def solves_for_divisor(formula: str) -> bool:
    """A preference solved for carries its own pro-rata divisor in the numerator.

    0668's `=(C42-$C$21+$C$24*$C$21)/$C$24` is the observed shape: the share
    count divides the whole expression and also multiplies inside it, which is
    what solving for a participating preference leaves behind.
    """
    m = re.search(r"/\s*(\$?[A-Za-z]{1,3}\$?\d+)\s*\)?\s*$", formula.strip())
    if not m:
        return False
    ref = m.group(1).replace("$", "")
    head = formula[: m.start()].replace("$", "")
    return bool(re.search(r"\b%s\b" % re.escape(ref), head))


def detect_npv_timing(gold: Book, scope: set[str]) -> list[dict]:
    out = []
    for k in sorted(scope):
        formula = gold.formula.get(k)
        if not formula or "NPV(" not in formula.upper():
            continue
        tail = formula.upper().split("NPV(", 1)[1]
        extra = bool(re.search(r"\)\s*[+\-]\s*[A-Z$]", tail))
        out.append(record(
            "npv_timing",
            "t0_added_separately" if extra else "excel_default_one_period_out",
            [k],
            formula,
            ["excel_default_one_period_out", "t0_added_separately"],
            "Excel NPV discounts its first argument by one full period.",
            fields={"label": q(gold.row_label(k))},
        ))
    return out


def q(text: str) -> str:
    return '"%s"' % str(text).replace('"', "'")


def ingredient_phrase(gold: Book, evidence: str, cells: list[str], own_label: str = "") -> str:
    """Name the rows a formula reads, by their visible labels.

    Returns empty when the result would not be a clean sentence, which is the
    signal for the record not to ship at all. Three cases fail: no labelled
    ingredient to name, more than two of them, or one that merely repeats the
    target's own label on the same sheet. An off-sheet row with the same
    visible name is a different cell and is named (and the tab is included).
    In each failing case the sentence would be mush, and a mushy disclosure
    is worse than silence.
    """
    if not cells:
        return ""
    sheet, coord = cells[0].split("!", 1)
    p = split_coord(coord)
    if not p:
        return ""
    col, row = col_to_num(p[0]), p[1]
    labels = []
    for rsheet, rcol, rrow in _ingredients(evidence, sheet, row, col):
        lab = gold.row_label(key(rsheet, "%s%d" % (num_to_col(rcol), rrow)))
        if not lab:
            return ""
        # Same-string on the same sheet is the target naming itself. Same-string
        # on another sheet is a different row (0042: Summary I9 "Asset
        # Management Fee" is the rate; Operations 146 is the calculated line).
        if lab == own_label and rsheet == sheet:
            return ""
        if rsheet != sheet:
            piece = "the row labelled %s on the %s tab" % (q(lab), rsheet)
        else:
            piece = "the row labelled %s" % q(lab)
        if piece not in labels:
            labels.append(piece)
    if not labels or len(labels) > 2:
        return ""
    return labels[0] if len(labels) == 1 else " and ".join(labels)


def is_step_increment(body: str, prev: str) -> bool:
    """Prior cell of this row, plus or minus exactly one further term.

    `=+H59+$F$17` is the observed case. A SUM over members is not this shape.
    """
    text = re.sub(r"^\+", "", body.strip())
    prev_re = r"\$?%s" % re.escape(prev)
    other = r"(?:'[^']+'!)?\$?[A-Z]{1,3}\$?\d+"
    return bool(re.fullmatch(
        r"%s\s*[+\-]\s*%s" % (prev_re, other), text, re.I
    ))


def detect_projection_rule(gold: Book, delivered: Book, selected: set[str],
                           targets: set | None = None) -> list[dict]:
    rows = defaultdict(list)
    for k in selected:
        sheet, coord = k.split("!", 1)
        p = split_coord(coord)
        if p:
            rows[(sheet, p[1])].append((col_to_num(p[0]), k))
    out = []
    for (_sheet, _row), entries in sorted(rows.items()):
        if len(entries) < 3:
            continue
        entries.sort()
        classified = []
        for cnum, k in entries:
            formula = gold.formula.get(k, "")
            body = formula[1:] if formula.startswith("=") else formula
            prev = "%s%d" % (num_to_col(cnum - 1), _row)
            if re.fullmatch(r"\$?%s" % re.escape(prev), body.strip(), re.I):
                kind = "hold_level"
            elif re.search(r"\*\s*\(\s*1\s*[+\-]", body):
                kind = "hold_growth"
            elif is_step_increment(body, prev):
                kind = "step_increment"
            elif re.match(r"^AVERAGE\(", body.strip(), re.I):
                kind = "average_window"
            elif re.search(r"[A-Z]{1,3}\$?\d+\s*\*\s*", body):
                kind = "ratio_to_driver"
            else:
                kind = "other"
            # Two cells can share a rule kind and still be different formulas.
            # Splitting on kind alone fuses them into one band, and no single
            # sentence is true of both.
            classified.append((cnum, k, kind, r1c1ish(formula, _row, cnum)))
        runs: list[list] = []
        run: list[tuple] = []
        for item in classified:
            if run and (item[0] != run[-1][0] + 1
                        or item[2] != run[-1][2]
                        or item[3] != run[-1][3]):
                runs.append(run)
                run = []
            run.append(item)
        if run:
            runs.append(run)
        # A 1-cell seed plus one long run (0644 Info!H59 is implied share,
        # I59:M59 adds the increment) is not two bullets claiming the
        # horizon. Decline only when two shippable runs would compete.
        shippable = [r for r in runs if len(r) >= 3 and r[0][2] != "other"]
        for r in runs:
            append_projection_run(
                out, gold, r,
                covers_row=len(shippable) == 1 and r in shippable,
                targets=targets,
            )
    return out


def append_projection_run(out: list[dict], gold: Book, run: list[tuple],
                          covers_row: bool = True, targets: set | None = None):
    if len(run) < 3:
        return
    kind = run[0][2]
    if kind == "other":
        return
    label = gold.row_label(run[0][1])
    if not label:
        return
    cells = [item[1] for item in run]
    evidence = gold.formula.get(cells[min(1, len(cells) - 1)], "")
    sheet, coord = cells[0].split("!", 1)
    first = split_coord(coord)
    if kind == "hold_level" and first and run[0][0] >= 3:
        row = first[1]
        seed_coord = "%s%d" % (num_to_col(run[0][0] - 1), row)
        prior_coord = "%s%d" % (num_to_col(run[0][0] - 2), row)
        seed = key(sheet, seed_coord)
        seed_formula = gold.formula.get(seed, "")
        if re.fullmatch(
            r"=\s*MAX\(\s*\$?%s\s*,\s*0\s*\)" % re.escape(prior_coord),
            seed_formula,
            re.I,
        ):
            kind = "first_period_zero_floor_then_hold"
            cells.insert(0, seed)
            evidence = seed_formula
    # State the maximal golden same-mechanics span, not just the selected run.
    # A hold-flat run's seed (`AG31=AE31` ahead of `AH31=AG31`) and periods the
    # delivered file kept as cached text both belong to the copied scope.
    stated = full_copied_scope(gold, cells, targets)
    rec = record(
        "projection_rule", kind, cells, evidence,
        ["hold_level", "hold_growth", "step_increment", "average_window",
         "ratio_to_driver", "first_period_zero_floor_then_hold"],
        f"Forecast row labelled {label!r}; applies only to this contiguous {kind} run.",
        # hold_level names no ingredient, and its sentence has no slot for one.
        # Passing an empty field would make the renderer discard the record while
        # the disposition still claimed it shipped.
        fields=({"label": q(label)} if kind in (
                    "hold_level", "first_period_zero_floor_then_hold",
                ) else
                {"label": q(label),
                 "ingredient": ingredient_phrase(gold, evidence, cells, own_label=label)}),
        stated_cells=stated,
    )
    rec["covers_row"] = covers_row
    out.append(rec)


def _multiplies_ref(formula: str, ref: str) -> bool:
    """The reference actually participates in a multiplication.

    0666's `Calc_M!C386` is exactly `=Assumptions!C$162` - a direct read of an
    equity-investment amount with no operator at all - and the old detector
    still described it as an ownership-share multiplication.
    """
    sheet, coord = ref.split("!", 1)
    p = split_coord(coord)
    if not p:
        return False
    coord_re = r"\$?%s\$?%d(?![0-9])" % (p[0], p[1])
    sheet_re = r"(?:'%s'|%s)!" % (re.escape(sheet.replace("'", "''")), re.escape(sheet))
    ref_re = r"(?:%s)?%s" % (sheet_re, coord_re)
    return bool(
        re.search(r"%s\s*\*" % ref_re, formula)
        or re.search(r"\*\s*%s" % ref_re, formula)
    )


def detect_stake_scaling(gold: Book, scope: set[str]) -> list[dict]:
    pct_cells = set()
    for k, v in gold.value.items():
        label = gold.row_label(k)
        if not label or not re.search(r"equity investment|ownership|stake|% acquired", label, re.I):
            continue
        # An ownership share is a fraction. An "Initial Equity investment"
        # AMOUNT matches the label pattern too (0666), and describing a read
        # of that amount as share scaling states mechanics the formula does
        # not have.
        if isinstance(v, (int, float)) and not isinstance(v, bool) and 0 < v <= 1:
            pct_cells.add(k)
    byrow = defaultdict(list)
    for k in sorted(scope):
        formula = gold.formula.get(k)
        if not formula:
            continue
        try:
            if str(REPO_ROOT) not in sys.path:
                sys.path.insert(0, str(REPO_ROOT))
            from xl_ast_graph import parse_formula  # type: ignore

            ast = parse_formula(formula)
        except Exception:
            continue
        # This entry's sentence names both operands. Only claim a plain
        # two-reference multiplication; complements, sums, or added unscaled
        # terms require a different entry that can render their complete AST.
        if (
            ast.kind != "infix"
            or ast.name != "*"
            or len(ast.args) != 2
            or any(arg.kind != "ref" for arg in ast.args)
        ):
            continue
        references = refs_in(formula, k.split("!", 1)[0])
        hits = [ref for ref in references if ref in pct_cells]
        bases = [ref for ref in references if ref not in pct_cells]
        if len(hits) != 1 or len(bases) != 1:
            continue
        sheet, coord = k.split("!", 1)
        p = split_coord(coord)
        if p:
            byrow[(sheet, p[1])].append(
                (col_to_num(p[0]), k, formula, hits[0], bases[0])
            )
    out = []
    for (_sheet, _row), entries in sorted(byrow.items()):
        entries.sort()
        _, first, formula, share, base = entries[0]
        if not gold.row_label(first):
            continue
        # The fixed sentence has no slot for a sign flip or a subtraction.
        # 0523's `'NPV & IRR'!B11 = -Multiples!D26*B4` shipped as a bare
        # ownership-share claim and was blocked for omitting the source and
        # its leading negation; a partial claim is false by omission, so a
        # formula carrying any minus declines here. Silence is safer.
        body = re.sub(r"'(?:[^']|'')+'!", "S!", STRING_RE.sub('""', formula))
        if "-" in body:
            continue
        out.append(record(
            "stake_scaling", "applied", [k for _, k, _, _, _ in entries],
            formula, ["applied", "not_applied"],
            f"Row labelled {gold.row_label(first)!r} multiplies "
            f"{pretty(base)} by {pretty(share)}, the ownership share.",
            fields={
                "label": q(gold.row_label(first)),
                "base": "the row labelled %s" % q(gold.row_label(base))
                        if gold.row_label(base) else pretty(base),
                "ingredient": "the row labelled %s" % q(gold.row_label(share))
                              if gold.row_label(share) else pretty(share),
            },
        ))
    return out


def _row_key(ref: str) -> tuple[str, int] | None:
    sheet, coord = ref.split("!", 1)
    p = split_coord(coord)
    return (sheet, p[1]) if p else None


def is_whole_value_read(formula: str, gold: Book, sheet: str, row: int) -> str | None:
    """Single named source row, optionally sign-flipped or scaled by one factor."""
    refs = refs_in(formula, sheet)
    rows = []
    for ref in refs:
        rk = _row_key(ref)
        if rk and rk not in rows:
            rows.append(rk)
    if not rows:
        return None
    # Same-row references are a projection, not a source read.
    foreign = [(s, r) for s, r in rows if not (s == sheet and r == row)]
    if not foreign or len(foreign) > 2:
        return None
    source_sheet, source_row = foreign[0]
    sample = next(ref for ref in refs if _row_key(ref) == (source_sheet, source_row))
    if not gold.row_label(sample):
        return None
    if len(foreign) == 2:
        other = next(ref for ref in refs if _row_key(ref) == foreign[1])
        if not gold.row_label(other):
            return None
        if not re.search(r"\*", formula):
            return None
    return sample


def detect_source_selection(gold: Book, scope: set[str]) -> list[dict]:
    # Dedup by (sheet, row, source): a row whose columns each read a
    # *different* source cell (e.g. period columns linking to consecutive
    # blocks of another sheet) carries one material link per column, and the
    # faithcheck source_link rule demands a record for each. Same-source
    # repeats on a row still collapse to one record.
    out, seen_links = [], set()
    for k in sorted(scope):
        formula = gold.formula.get(k)
        if not formula:
            continue
        sheet, coord = k.split("!", 1)
        p = split_coord(coord)
        if not p:
            continue
        label = gold.row_label(k)
        if not label:
            continue
        source = is_whole_value_read(formula, gold, sheet, p[1])
        if not source:
            continue
        if (sheet, p[1], source) in seen_links:
            continue
        source_sheet = source.split("!", 1)[0]
        lowered = label.lower()
        purchase = any(t in lowered for t in (
            "initial investment", "purchase price", "entry", "consideration", "cv",
        ))
        rate_like = bool(re.search(
            r"\brate\b|hold start|blended|wacc|source|cost of (debt|equity|capital)",
            lowered,
        ))
        bare = bool(re.fullmatch(r"=\s*[+\-]?[A-Za-z0-9_ '$#!]+$", formula.strip()))
        # An unsigned bare single-cell read is exactly representable in the
        # entry's sentence, so the cross-sheet gate admits it even without a
        # purchase or rate label (0648's `'IRR-NPV Calculations'!DS6 =
        # Workings!D87`, the omitted terminal exit-value link).
        bare_unsigned = bool(BARE_LINK_RE.match(formula.strip()))
        if source_sheet != sheet:
            if not (purchase or rate_like or bare_unsigned):
                continue
        else:
            if gold.row_label(source) == label or not bare:
                continue
        seen_links.add((sheet, p[1], source))
        source_label = gold.row_label(source)
        rec = record(
            "source_selection", "source", [k], formula, ["source"],
            f"Row labelled {label!r} reads from {pretty(source)}.",
            fields={
                "label": q(label),
                # Name the exact cell as well as the row, so the link chain
                # is reproducible without the formula (0648's repair shape).
                "ingredient": describe_ref(gold, pretty(source), sheet, label),
            },
        )
        rec["source_cell"] = source
        rec["bare_unsigned_link"] = bare_unsigned
        out.append(rec)
    return out


def row_has_data(gold: Book, sheet: str, rownum: int) -> bool:
    for c in range(1, 61):
        k = key(sheet, "%s%d" % (num_to_col(c), rownum))
        if k in gold.formula or isinstance(gold.value.get(k), (int, float)):
            return True
    return False


def rows_inside_sums(gold: Book) -> dict:
    """Rows covered by some SUM range, keyed by sheet.

    An empty labelled row is worth disclosing only when a total spans it, since
    then its emptiness changes that total. Anywhere else it is a block header,
    and telling the agent to leave a header empty is noise.
    """
    out: dict = defaultdict(set)
    for k, formula in gold.formula.items():
        sheet = k.split("!", 1)[0]
        for m in re.finditer(r"SUM\(\s*(\$?[A-Z]{1,3}\$?\d+\s*:\s*\$?[A-Z]{1,3}\$?\d+)\s*\)",
                             formula, re.I):
            for cell in refs_in("=" + m.group(1), sheet):
                csheet, coord = cell.split("!", 1)
                p = split_coord(coord)
                if p:
                    out[csheet].add(p[1])
    return out


def detect_row_populated(gold: Book, delivered: Book, scope: set[str], targets: list[str]) -> list[dict]:
    out = []
    summed_rows = rows_inside_sums(gold)
    sheets = {t.split("!", 1)[0] for t in targets}
    scope_rows = defaultdict(set)
    for k in sorted(scope):
        sheet, coord = k.split("!", 1)
        p = split_coord(coord)
        if p:
            scope_rows[sheet].add(p[1])
    for sheet in sorted(sheets):
        rows = defaultdict(list)
        for k in list(gold.value) + list(gold.formula):
            s, coord = k.split("!", 1)
            if s != sheet:
                continue
            p = split_coord(coord)
            if p:
                rows[p[1]].append((col_to_num(p[0]), k))
        for rownum, entries in sorted(rows.items()):
            entries.sort()
            label, label_col, data = "", None, []
            for cnum, k in entries:
                v = gold.value.get(k)
                is_text = isinstance(v, str) and k not in gold.formula
                if is_text and not label:
                    label, label_col = v.strip(), cnum
                elif k in gold.formula or isinstance(v, (int, float)):
                    # Golden data anywhere on the row disproves "unused",
                    # including formulas left of the first literal text.
                    # 0605's Dashboard row 14 holds `D14 = Calc!B$140` before
                    # its "Operating metrics" heading in K14, and the old
                    # position gate declared the whole row empty.
                    data.append(k)
            if data or not label or len(label) < 4 or HEADER_RE.search(label) or label.isupper():
                continue
            # The row must look empty in the file the agent actually receives. Its
            # label survives the emptying pass by design, so only cells to the
            # right of the label count: a visible annotation or figure there means
            # this is not an unused row, and saying it is contradicts what the
            # agent can see.
            if any(delivered.has(k) for cnum, k in entries if cnum > (label_col or 0)):
                continue
            # An unused row only matters when a total actually spans it, which is
            # most of what separates a real empty member from a block header.
            if rownum not in summed_rows.get(sheet, ()):
                continue
            # A header also falls inside its block's total, so require data on
            # both sides. A member sits between populated rows; a header has its
            # block below it and nothing above.
            if not (row_has_data(gold, sheet, rownum - 1)
                    and row_has_data(gold, sheet, rownum + 1)):
                continue
            above = [r for r in (rownum - 1, rownum - 2) if r in scope_rows[sheet]]
            below = [r for r in (rownum + 1, rownum + 2) if r in scope_rows[sheet]]
            if above and below:
                out.append(record(
                    "row_populated", "unused", [], "no formula and no value anywhere on row",
                    ["unused", "populated", "populated_but_unread"],
                    f"Row labelled {label!r} is empty in the original but sits inside the block that feeds a graded answer.",
                    row_ref=f"{quote_sheet(sheet)}!row {rownum}",
                    fields={"label": q(label)},
                ))
    out.extend(detect_populated_but_unread(gold, delivered))
    return out


ASSUMPTION_LABEL_RE = re.compile(
    r"payout|ratio|assumption|yield|per annum|\bp\.a\.?\b|days\b|growth rate",
    re.I,
)


def detect_populated_but_unread(gold: Book, delivered: Book) -> list[dict]:
    """A surviving assumption row that nothing in the golden reads."""
    truncated: set[str] = set()
    referenced: set[str] = set()
    for k, formula in gold.formula.items():
        sheet = k.split("!", 1)[0]
        referenced.update(refs_in(formula, sheet, truncated=truncated))
    referenced_rows = set()
    for ref in referenced:
        rk = _row_key(ref)
        if rk:
            referenced_rows.add(rk)
    out, seen = [], set()
    for k in list(gold.value) + list(gold.formula):
        sheet, coord = k.split("!", 1)
        p = split_coord(coord)
        if not p or (sheet, p[1]) in seen:
            continue
        seen.add((sheet, p[1]))
        if sheet in truncated:
            continue
        if (sheet, p[1]) in referenced_rows:
            continue
        label = gold.row_label(k)
        if not label or len(label) < 4 or HEADER_RE.search(label) or label.isupper():
            continue
        if not ASSUMPTION_LABEL_RE.search(label):
            continue
        row_cells = gold.row_cells(sheet, p[1])
        surviving = [
            c for c in row_cells
            if c in delivered.formula
            or isinstance(delivered.value.get(c), (int, float))
        ]
        if not surviving:
            continue
        populated = any(
            c in gold.formula or isinstance(gold.value.get(c), (int, float))
            for c in row_cells
        )
        if not populated:
            continue
        out.append(record(
            "row_populated", "populated_but_unread",
            surviving,
            "populated; no inbound reference",
            ["unused", "populated", "populated_but_unread"],
            f"Row labelled {label!r} is populated but unread.",
            fields={"label": q(label)},
        ))
    return out


def detect_aggregate_scope(gold: Book, delivered: Book, targets: list[str]) -> list[dict]:
    """Name every member of equivalent vertical totals across target columns."""
    grouped = defaultdict(list)
    for k in targets:
        formula = gold.formula.get(k)
        if not formula:
            continue
        # A plain sum over a single span, nothing else. The old greedy pattern
        # also matched `=SUM(a-b/c*d)` and `=SUM(...)/-SUM(...)`, and described a
        # subtraction or a ratio as if it were an addition.
        m = re.match(r"^=\s*SUM\(\s*(\$?[A-Z]{1,3}\$?\d+\s*:\s*\$?[A-Z]{1,3}\$?\d+)\s*\)\s*$",
                     formula.strip(), re.I)
        if not m:
            continue
        sheet = k.split("!", 1)[0]
        spanned = refs_in("=" + m.group(1).strip(), sheet)
        # Vertical only: a same-row horizontal SUM names one row and renders as
        # "the total labelled X is the sum of the row labelled X".
        member_rows = {split_coord(c.split("!", 1)[1])[1] for c in spanned
                       if split_coord(c.split("!", 1)[1])}
        if len(member_rows) < 2:
            continue
        named = [gold.row_label(c) for c in spanned]
        if not all(named):
            continue
        if not all(row_has_data(gold, *c.split("!", 1)[0:1], split_coord(c.split("!", 1)[1])[1])
                   for c in spanned):
            continue
        labels = list(dict.fromkeys(named))
        members = [c for c in spanned if not delivered.has(c)]
        if not members or not labels:
            continue
        first = split_coord(spanned[0].split("!", 1)[1])
        last = split_coord(spanned[-1].split("!", 1)[1])
        target = split_coord(k.split("!", 1)[1])
        if not first or not last or not target:
            continue
        group_key = (sheet, first[1], last[1], tuple(labels), gold.row_label(k))
        grouped[group_key].append((col_to_num(target[0]), k, formula))

    out = []
    for (_sheet, first_row, last_row, labels, total_label), entries in sorted(grouped.items()):
        entries.sort()
        target_cells = [item[1] for item in entries]
        if not target_cells or not total_label:
            continue
        rec = record(
            "aggregate_scope", "member set", target_cells, entries[0][2], ["member set"],
            f"Total labelled {total_label!r} aggregates the complete named member set.",
            fields={
                "label": q(total_label),
                "members": ("the row labelled " + q(labels[0])) if len(labels) == 1
                           else "the rows labelled " + ", ".join(q(label) for label in labels),
            },
        )
        rec["accepted_target_reference"] = True
        rec["aggregate_member_rows"] = {
            "first": first_row,
            "last": last_row,
            "labels": list(labels),
        }
        out.append(rec)
    return out


# --------------------------------------------------------------------------- registry
#
# REGISTRY.md is the authority. Nothing may ship that it does not name, so the
# entry ids, alternatives and agent-facing sentences are read from it rather
# than duplicated here. Only the `Ship when` predicates live in Python, because
# they are conditions over a workbook rather than prose; `check_registry_drift`
# asserts the two stay in step.

REGISTRY_PATH = Path(__file__).resolve().parent.parent / "REGISTRY.md"
ENTRY_ID_RE = re.compile(r"^-\s+\*\*Id\.\*\*\s+`([a-z_]+)`", re.M)
FIELD_RE = re.compile(r"^-\s+\*\*([A-Z][A-Za-z ]*)\.\*\*\s*(.*)$")
SENTENCE_RE = re.compile(r"-\s+(?:`([^`]+)`|([A-Za-z][A-Za-z ]*?))\s+-\s+\"(.+?)\"", re.S)
_REGISTRY_CACHE: dict | None = None


def registry() -> dict:
    """Parse REGISTRY.md into {entry_id: {alternatives, ship_when, sentences}}."""
    global _REGISTRY_CACHE
    if _REGISTRY_CACHE is not None:
        return _REGISTRY_CACHE
    if not REGISTRY_PATH.exists():
        raise SystemExit(f"registry not found: {REGISTRY_PATH}")
    text = REGISTRY_PATH.read_text(encoding="utf-8")
    entries: dict = {}
    for block in text.split("\n## ")[1:]:
        ids = ENTRY_ID_RE.findall(block)
        if not ids:
            continue
        entry_id = ids[0]
        fields: dict = {}
        current = None
        for line in block.splitlines():
            m = FIELD_RE.match(line)
            if m:
                current = m.group(1).strip().lower()
                fields[current] = m.group(2).strip()
            elif current and line.startswith("  "):
                fields[current] += "\n" + line.strip()
        alts = [
            " ".join(a.split()).strip(" `")
            for a in fields.get("alternatives", "").split("|")
            if a.strip()
        ]
        sentences = {
            (m.group(1) or m.group(2) or "").strip(): " ".join(m.group(3).split())
            for m in SENTENCE_RE.finditer(fields.get("sentence", ""))
        }
        entries[entry_id] = {
            "alternatives": alts,
            "ship_when": fields.get("ship when", ""),
            "sentences": sentences,
            "question": fields.get("question", ""),
        }
    _REGISTRY_CACHE = entries
    return entries


def registry_always_ships(entry_id: str) -> bool:
    return registry().get(entry_id, {}).get("ship_when", "").lstrip().startswith("`always`")


def check_registry_drift(detector_entries: set[str]) -> list[str]:
    """Prove detector -> registry and registry -> detector coverage."""
    known = set(registry())
    faults = [f"detector emits `{e}` with no registry entry" for e in sorted(detector_entries - known)]
    expected = CONVENTION_ENTRY_IDS | METHOD_ENTRY_IDS
    faults.extend(
        f"registry entry `{entry_id}` has no detector mapping"
        for entry_id in sorted(known - expected)
    )
    faults.extend(
        f"detector mapping `{entry_id}` has no registry entry"
        for entry_id in sorted(expected - known)
    )
    for entry_id in sorted(detector_entries & known):
        if not registry_always_ships(entry_id) and entry_id not in SHIP_WHEN:
            faults.append(f"entry `{entry_id}` has a conditional Ship when but no predicate")
    for entry_id in sorted(METHOD_ENTRY_IDS):
        if entry_id not in SHIP_WHEN:
            faults.append(f"method entry `{entry_id}` has no shared Ship when predicate")
        entry = registry().get(entry_id, {})
        if "out_of_catalogue" not in entry.get("alternatives", []):
            faults.append(f"method entry `{entry_id}` has no out_of_catalogue alternative")
        if "out_of_catalogue" not in entry.get("sentences", {}):
            faults.append(f"method entry `{entry_id}` has no out_of_catalogue sentence")
    return faults


# ----------------------------------------------------------------- ship-when
#
# A predicate returns True to allow disclosure for this band. Entries whose
# registry `Ship when` is `always` need no predicate.


def _ingredients(evidence: str, sheet: str, row: int, col: int) -> list[tuple[str, int, int]]:
    """Referenced cells other than the immediately-prior column on the same row."""
    out = []
    for ref in refs_in(evidence, sheet):
        rsheet, coord = ref.split("!", 1)
        p = split_coord(coord)
        if not p:
            continue
        rcol, rrow = col_to_num(p[0]), p[1]
        if rsheet == sheet and rrow == row and abs(rcol - col) <= 1:
            continue
        out.append((rsheet, rcol, rrow))
    return out


def _available(ctx, rsheet: str, rcol: int, rrow: int, sheet: str, row: int) -> bool:
    """Visible in the delivered file, or structurally adjacent on the same sheet."""
    if ctx["delivered"].has(key(rsheet, "%s%d" % (num_to_col(rcol), rrow))):
        return True
    return rsheet == sheet and abs(rrow - row) <= 15


def ship_projection_rule(rec: dict, ctx: dict) -> bool:
    """Ship only when the rule's ingredient is out of the agent's reach.

    A driver, growth rate or averaged window that is visible or sitting a few
    rows away is ordinary spreadsheet reasoning. Disclosing it restates what the
    workbook already shows.
    """
    cells = rec.get("cell_keys") or []
    if not cells:
        return False
    if rec.get("value") not in (
        "hold_level", "first_period_zero_floor_then_hold",
    ):
        rec["declined_reason"] = (
            "only a literal hold-level convention is safe for agent-facing disclosure"
        )
        return False
    sheet, coord = cells[0].split("!", 1)
    p = split_coord(coord)
    if not p:
        return False
    col, row = col_to_num(p[0]), p[1]
    evidence = rec.get("evidence") or ""
    # A branch is not a projection rule. The shape classifier fires on the
    # multiplication inside an IF and mislabels genuinely custom logic.
    if re.search(r"\b(IF|IFS|CHOOSE|LOOKUP|XLOOKUP)\s*\(", evidence, re.I):
        return False
    if not evidence:
        return False
    if rec.get("value") == "average_window":
        # The sentence says the row IS an average. Anything wrapped around the
        # AVERAGE - a multiplier, an offset - makes that false, and flattening
        # the operand and the multiplier into one list reads as though both were
        # being averaged.
        if not re.fullmatch(r"=\s*AVERAGE\([^()]*\)\s*", evidence.strip(), re.I):
            return False
    # The sentences say "in each period" and "across the forecast". That is only
    # true when the run covers every cell the agent has to build on this row; a
    # row that switches rule partway gets two bullets, each claiming the whole
    # horizon, and at most one of them can be right.
    if rec.get("covers_row") is False:
        return False
    if rec.get("value") == "first_period_zero_floor_then_hold":
        return True
    # A record whose ingredient could not be named renders nothing. Deciding that
    # here keeps the disposition honest instead of marking it disclosed and
    # dropping it silently at render time.
    if rec.get("value") != "hold_level" and not (rec.get("fields") or {}).get("ingredient"):
        return False
    if rec.get("value") == "hold_level":
        # Holding flat needs no named ingredient, only a level to hold. Ship when
        # no numeric value earlier on the row is visible, so nothing tells the
        # agent what level to carry. The row's label is visible by design and is
        # not a level, so text cells do not count.
        return not any(
            isinstance(ctx["delivered"].value.get(key(sheet, "%s%d" % (num_to_col(c), row))),
                       (int, float))
            for c in range(1, col)
        )
    ingredients = _ingredients(evidence, sheet, row, col)
    if not ingredients:
        return False
    if rec.get("value") == "step_increment" and increment_label_is_rate_like(
        ctx, ingredients
    ):
        # Visibility of the increment is not visibility of add-vs-compound.
        return True
    return not all(_available(ctx, rs, rc, rr, sheet, row) for rs, rc, rr in ingredients)


def increment_label_is_rate_like(ctx: dict, ingredients: list) -> bool:
    gold = ctx["gold"]
    for rs, rc, rr in ingredients:
        label = gold.row_label(key(rs, "%s%d" % (num_to_col(rc), rr))).lower()
        if re.search(r"growth|p\.a\.|per annum|%", label):
            return True
    return False


def ship_source_selection(rec: dict, ctx: dict) -> bool:
    """Ship the wiring only in the exactly-representable case.

    0648 blocked because the model's sole terminal exit-value link
    (`'IRR-NPV Calculations'!DS6 = Workings!D87`) was omitted entirely. The
    old blanket decline ("until sign and period semantics are represented
    explicitly") is now met narrowly: a bare, unsigned, single-cell read from
    another sheet has no sign or scale to misstate, and the sentence names
    the exact source cell. Everything else stays reviewer-only:

    - a sign-flipped or scaled read has semantics the sentence cannot carry;
    - a same-sheet mirror is ordinary spreadsheet reasoning;
    - a source in the graded-output closure would name answer material;
    - a source that survives in the delivered file is one read away from a
      value, so naming it hands that value over.
    """
    cells = rec.get("cell_keys") or []
    src = rec.get("source_cell")
    if not rec.get("bare_unsigned_link") or not src or len(cells) != 1:
        rec["declined_reason"] = (
            "only a bare unsigned single-cell read renders sign and scale "
            "exactly; this shape stays reviewer-only"
        )
        return False
    if src.split("!", 1)[0] == cells[0].split("!", 1)[0]:
        rec["declined_reason"] = "same-sheet mirror is ordinary spreadsheet reasoning"
        return False
    if any(c in ctx["targets"] for c in cells) or src in ctx["targets"]:
        rec["declined_reason"] = "source link reaches the graded-output closure"
        return False
    if ctx["delivered"].has(src):
        rec["declined_reason"] = (
            "source survives in the delivered file; naming it hands the value over"
        )
        return False
    if not (rec.get("fields") or {}).get("ingredient"):
        rec["declined_reason"] = "source could not be named cleanly"
        return False
    return True


def ship_not_a_target(rec: dict, ctx: dict) -> bool:
    """Ship only when the band is not itself graded."""
    return not any(c in ctx["targets"] for c in rec.get("cell_keys") or [])


def ship_aggregate_scope(rec: dict, ctx: dict) -> bool:
    """Keep exact aggregate evidence reviewer-only when the total is graded."""
    if any(c in ctx["targets"] for c in rec.get("cell_keys") or []):
        rec["declined_reason"] = (
            "complete aggregate scope would disclose a graded target's calculation"
        )
        return False
    return bool(rec.get("aggregate_member_rows", {}).get("labels"))


def ship_row_populated(rec: dict, ctx: dict) -> bool:
    """Do not claim that values are hidden when they remain agent-visible."""
    if rec.get("value") == "populated_but_unread":
        rec["declined_reason"] = (
            "reviewer-only: cited values remain visible in the delivered workbook"
        )
        return False
    return True


def _band_survives(rec: dict, ctx: dict) -> bool:
    return any(ctx["delivered"].has(c) for c in rec.get("cell_keys") or [])


def ship_distribution_policy(rec: dict, ctx: dict) -> bool:
    """Ship only while the rule is still hidden.

    A payout rate that survives on a labelled row and is applied to one visible
    driver is ordinary reasoning, not a convention the agent cannot see.
    """
    if _band_survives(rec, ctx):
        rec["declined_reason"] = "distribution row survives in the delivered file"
        return False
    if rec.get("value") == "formula_mechanics":
        # Unlike the named-value sentences, the mechanics sentence states a
        # construction, so it follows the Section 2 graded-target policy.
        if any(c in ctx["targets"] for c in rec.get("cell_keys") or []):
            rec["declined_reason"] = "distribution mechanics band is itself graded"
            return False
        if set(rec.get("mechanics_references") or []) & ctx["targets"]:
            rec["declined_reason"] = "distribution mechanics would name a graded target"
            return False
    if rec.get("value") == "payout_ratio":
        ings = list(_ingredients(rec.get("evidence") or "", *_origin(rec)))
        if len(ings) == 1 and ctx["delivered"].has(
            key(ings[0][0], "%s%d" % (num_to_col(ings[0][1]), ings[0][2]))
        ):
            rec["declined_reason"] = "payout rate survives and applies to one visible driver"
            return False
    return True


def ship_liquidation_preference(rec: dict, ctx: dict) -> bool:
    """Ship only while the waterfall is blank and no preference multiple survives.

    Deliberately not gated on the band being graded. This is a Section 1
    convention: it names which of four defensible splits the author chose and
    states no construction over inputs the agent still holds.
    """
    if _band_survives(rec, ctx):
        rec["declined_reason"] = "waterfall row survives in the delivered file"
        return False
    for rs, rc, rr in _ingredients(rec.get("evidence") or "", *_origin(rec)):
        cell = key(rs, "%s%d" % (num_to_col(rc), rr))
        if "preference" in ctx["gold"].row_label(cell).lower() and ctx["delivered"].has(cell):
            rec["declined_reason"] = "a surviving labelled row states the preference multiple"
            return False
    return True


def ship_custom_method(rec: dict, ctx: dict) -> bool:
    """Shared gate for every out-of-catalogue method record.

    Thresholds, long ingredient lists, mixed source bands, constants, and full
    operator sequences are intentionally not decline reasons. They are rendered
    deterministically and then audited. Only direct answer disclosure or a
    sentence the parser could not build stops the record.
    """
    if any(c in ctx["targets"] for c in rec.get("cell_keys") or []):
        rec["declined_reason"] = "method band is itself graded"
        return False
    referenced = set((rec.get("method_profile") or {}).get("references", []))
    if referenced & ctx["targets"]:
        rec["declined_reason"] = "method sentence would name another graded target"
        return False
    fields = rec.get("fields") or {}
    if not fields.get("label") or not fields.get("steps"):
        rec["declined_reason"] = "method sentence fields are incomplete"
        return False
    if not rec.get("coverage_complete"):
        rec["declined_reason"] = "method sentence does not cover every reference and literal"
        return False
    return True


def _origin(rec: dict):
    cells = rec.get("cell_keys") or []
    if not cells:
        return "", 0, 0
    sheet, coord = cells[0].split("!", 1)
    p = split_coord(coord)
    return (sheet, p[1], col_to_num(p[0])) if p else (sheet, 0, 0)


SHIP_WHEN = {
    "projection_rule": ship_projection_rule,
    "source_selection": ship_source_selection,
    "npv_timing": ship_not_a_target,
    "aggregate_scope": ship_aggregate_scope,
    "row_populated": ship_row_populated,
    "distribution_policy": ship_distribution_policy,
    "liquidation_preference": ship_liquidation_preference,
    **{entry_id: ship_custom_method for entry_id in METHOD_ENTRY_IDS},
}


def apply_ship_when(records: list[dict], ctx: dict) -> list[dict]:
    """Set each record's disposition from its registry entry's Ship when."""
    for rec in records:
        entry_id = rec.get("entry")
        if not entry_id:
            rec["disposition"] = "unclassified"
            continue
        if not registry().get(entry_id, {}).get("sentences"):
            rec["disposition"] = "unclassified"
            rec["declined_reason"] = "entry has no sentence for any value"
            continue
        if rec.get("value") not in registry()[entry_id]["sentences"]:
            # A value the registry does not word, such as terminal_value `other`.
            rec["disposition"] = "unclassified"
            rec["declined_reason"] = f"no sentence for value {rec.get('value')!r}"
            continue
        predicate = SHIP_WHEN.get(entry_id)
        if predicate is not None and not predicate(rec, ctx):
            rec["disposition"] = "suppressed"
            rec.setdefault("declined_reason", "Ship when declined for this band")
            continue
        rec["disposition"] = "disclosed"
    return records


ENTRY_PRECEDENCE = (
    list(METHOD_ROLE_PATTERNS)
    + [
        "discount_period", "inert_line", "terminal_value", "row_populated",
        "npv_timing", "aggregate_scope", "projection_rule", "distribution_policy",
        "liquidation_preference", "stake_scaling", "source_selection",
    ]
)


def _entry_rank(rec: dict) -> int:
    entry = rec.get("entry") or rec.get("family") or ""
    try:
        return ENTRY_PRECEDENCE.index(entry)
    except ValueError:
        return len(ENTRY_PRECEDENCE)


def arbitrate_overlapping(records: list[dict]) -> list[dict]:
    """First candidate that ships wins the shared cells; the rest fall through.

    A withheld or declined record does not consume the band. Only a disclosed
    record claims its cells.
    """
    winners = [r for r in records if r.get("disposition") == "disclosed"]
    winners.sort(key=_entry_rank)
    claimed: set[str] = set()
    for rec in winners:
        cells = set(rec.get("cell_keys") or [])
        if cells & claimed:
            rec["disposition"] = "suppressed"
            rec["declined_reason"] = "earlier candidate already shipped"
            continue
        claimed |= cells
    return records


def resolve_unused_conflicts(records: list[dict]) -> list[dict]:
    """Never say a row is unused and also name it as part of a total.

    Both statements can be true of the golden - an empty row really can sit
    inside a summed range - but together they tell the agent to build and not
    build the same row. The membership sentence is the more useful of the two,
    so the unused claim yields.
    """
    named = set()
    for rec in records:
        if rec.get("disposition") == "disclosed" and rec.get("entry") == "aggregate_scope":
            # The members field is a rendered phrase, so strip its lead-in before
            # comparing. Splitting it raw leaves the first member glued to "the
            # rows labelled " and it never matches.
            phrase = re.sub(r"^the rows? labelled ", "",
                            str(rec.get("fields", {}).get("members", "")))
            named.update(m.strip() for m in phrase.split(", "))
    for rec in records:
        if rec.get("entry") != "row_populated" or rec.get("disposition") != "disclosed":
            continue
        if str(rec.get("fields", {}).get("label", "")) in named:
            rec["disposition"] = "suppressed"
            rec["declined_reason"] = "row is named as a member of a disclosed total"
    return records


def record(family, value, cells, evidence, alternatives, note, row_ref=None, fields=None,
           stated_cells=None):
    # stated_cells widens what the record *says* to the full copied-mechanics
    # span; cells remain what the record claims for arbitration, leak checks
    # and blankness verification.
    shown = stated_cells or cells
    return {
        "band": compact_cells([pretty(c) for c in shown]) if shown else row_ref,
        "cells": [pretty(c) for c in shown] if shown else ([row_ref] if row_ref else []),
        "cell_keys": cells,
        "label": note,
        "role": family,
        "family": family,
        "entry": family if family in registry() else None,
        "value": value,
        "alternatives": alternatives,
        # Set by apply_ship_when once the registry has been consulted.
        "disposition": "pending",
        "source": "convention_detector",
        "evidence": evidence,
        "note": note,
        "fields": fields or {},
        "leak_flag": False,
    }


def detect_defects(gold: Book, targets: list[str], scope: set[str]) -> list[dict]:
    out, seen_rows = [], set()
    for k in sorted(scope):
        formula = gold.formula.get(k)
        label = gold.row_label(k)
        if not (formula and label and re.search(r"\(\s*1\s*-\s*t\w*\s*\)", label, re.I)):
            continue
        m = re.search(r"\*\s*\(\s*1\s*\+\s*([^)]+)\)", formula)
        if not m:
            continue
        sheet, coord = k.split("!", 1)
        p = split_coord(coord)
        if not p or (sheet, p[1]) in seen_rows:
            continue
        rates = [gold.value.get(r) for r in refs_in("=" + m.group(1), sheet)]
        rates = [v for v in rates if isinstance(v, (int, float))]
        if rates and all(v < 0 for v in rates):
            continue
        seen_rows.add((sheet, p[1]))
        out.append({"kind": "label_sign_mismatch", "cell": pretty(k), "detail": f"Label {label!r} but formula multiplies by (1 + rate): {formula}"})
    for k in targets:
        formula = gold.formula.get(k)
        val = gold.value.get(k)
        if formula and isinstance(val, (int, float)) and abs(val) < 1e-12:
            empty = [r for r in refs_in(formula, k.split("!", 1)[0]) if not gold.has(r)]
            if empty:
                out.append({"kind": "empty_operand_zero_target", "cell": pretty(k), "detail": f"Target is zero because {_fmt(empty, 2)} is empty: {formula}"})
    return out


def overlap_claimed(record_cells: list[str], claimed: set[str]) -> bool:
    return any(c in claimed for c in record_cells)


def record_reach(rec: dict) -> set[str]:
    """Every cell a record claims, states, or names in its mechanics."""
    reach = set(rec.get("cell_keys") or [])
    for c in rec.get("cells") or []:
        if isinstance(c, str) and "!" in c and "!row " not in c:
            sheet, coord = c.rsplit("!", 1)
            if split_coord(coord):
                reach.add(key(sheet, coord))
    reach.update(rec.get("mechanics_references") or [])
    reach.update((rec.get("method_profile") or {}).get("references") or [])
    if rec.get("source_cell"):
        reach.add(rec["source_cell"])
    return reach


def suppress_closure_reach(records: list[dict], closure: set[str]) -> list[dict]:
    """Suppress any disclosed record that reaches the graded-output closure.

    The old gate suppressed only Section 2 methods whose own band was
    literally graded, which let through:

    - 0353: a zero-value claim over `Summary!B50:O50`, whose band contains
      requested `Summary!D50`, shipped from a Section 1 entry;
    - 0441: `Model!O24`'s method, where graded `Model!O1` is `=+O24`;
    - 0620/0672: mechanics for the same-row copied equivalent of a graded
      cell.

    `liquidation_preference` keeps its registry standing exception (0668: it
    names which of four defensible splits the author chose, not a
    construction), and `aggregate_scope` records marked
    `accepted_target_reference` name visible labels, never the target.
    """
    for rec in records:
        if rec.get("disposition") != "disclosed":
            continue
        if rec.get("entry") == "liquidation_preference" or rec.get("accepted_target_reference"):
            continue
        hit = record_reach(rec) & closure
        if hit:
            rec["disposition"] = "suppressed"
            rec["declined_reason"] = (
                "record reaches the graded-output closure at "
                + ", ".join(pretty(h) for h in sorted(hit)[:3])
            )
    return records


def import_legacy_method_records(task_dir: Path, delivered: Book, claimed_keys: set[str]) -> list[dict]:
    """Legacy hints are evidence, not shippable disclosure.

    Earlier versions imported non-redundant old hints as `method` records. The
    layer-3 verifier caught that this can re-ship known-bad prose (0523's
    "averaged forecast EBITDA" hint). Fresh method records must come from the
    model/context pass with reviewer evidence, not from stale instructions.
    """
    return []


def coalesce_same_band(records: list[dict]) -> list[dict]:
    """Enforce the single-record-per-band contract.

    A cell can legitimately encode two safe conventions, for example a purchase
    price can both read from another sheet and apply ownership scaling. The writer
    still needs one bullet for the band, so combine same-band convention records.
    """
    grouped = defaultdict(list)
    for rec in records:
        grouped[rec.get("band") or ""].append(rec)
    out = []
    for band, items in grouped.items():
        if len(items) == 1 or not band:
            out.extend(items)
            continue
        # Keep non-convention or review-only records separate if a collision ever
        # appears there; those need human attention rather than automatic merging.
        mergeable = [
            item for item in items
            if item.get("disposition") == "convention" and not item.get("review_only")
        ]
        others = [item for item in items if item not in mergeable]
        if len(mergeable) <= 1:
            out.extend(items)
            continue
        base = dict(mergeable[0])
        families = [item["family"] for item in mergeable]
        values = [f"{item['family']}={item['value']}" for item in mergeable]
        notes = [item.get("note", "") for item in mergeable if item.get("note")]
        alternatives = []
        evidence = []
        for item in mergeable:
            alternatives.extend(item.get("alternatives", []))
            if item.get("evidence"):
                evidence.append(item["evidence"])
        base.update({
            "role": "multiple_conventions",
            "family": "multiple_conventions",
            "value": "; ".join(values),
            "alternatives": list(dict.fromkeys(alternatives)),
            "note": " ".join(notes),
            "evidence": " | ".join(evidence),
            "merged_families": families,
        })
        out.append(base)
        out.extend(others)
    return out


def records_provenance(selection: dict) -> dict:
    """Carry immutable-generation bindings from selection through writing."""
    return {
        "golden": selection["golden"],
        "golden_sha256": selection["golden_sha256"],
        "pipeline_bindings": selection.get("pipeline_bindings"),
        "segmentation_mode": (
            (selection.get("segmentation_generation") or {}).get("mode")
            or "legacy"
        ),
        "segmentation_generation": selection.get("segmentation_generation"),
        "selection": selection.get("selection"),
        "ast_dir": selection.get("ast_dir"),
        "seg_root": selection.get("seg_root"),
    }


def detect_records(args) -> dict:
    task_dir = Path(args.task_dir).resolve()
    selection = read_stage(task_dir, "bands", Path(args.runs_root))
    gold = Book(Path(selection["golden"]))
    delivered = Book(Path(selection["delivered"]))
    scope = selected_keys(selection)
    targets = selection["target_keys"]
    target_set = set(targets)
    # Everything downstream that guards against describing a graded answer
    # guards against its aliases and copied equivalents too (0353, 0441,
    # 0620, 0672).
    closure = graded_closure(gold, target_set)

    # Custom methods get first claim. Standard, ambiguous and uncertain
    # assessments remain reviewer-visible but do not claim the band, so the
    # convention detectors still get their normal chance.
    resolutions_path = getattr(args, "role_resolutions", None)
    if resolutions_path:
        resolutions_path = Path(resolutions_path)
    else:
        resolutions_path = default_role_resolutions_path(task_dir, Path(args.runs_root))
    method_records, method_assessments, custom_claimed = detect_custom_methods(
        gold, delivered, selection.get("bands", []), target_set,
        resolutions_path=resolutions_path,
    )
    custom_calibration = calibrate_legacy_custom(task_dir, method_assessments, scope)
    convention_scope = scope - custom_claimed
    convention_records = []
    for fn in (
        detect_discount_period,
        detect_inert_line,
        detect_terminal_value,
        detect_npv_timing,
    ):
        convention_records.extend(fn(gold, convention_scope))
    convention_records.extend(
        detect_distribution_policy(gold, convention_scope, closure)
    )
    # Ahead of stake_scaling: order is precedence, and stake_scaling ships
    # on `always`, so it cannot be beaten by suppression.
    convention_records.extend(detect_liquidation_preference(gold, convention_scope))
    convention_records.extend(detect_stake_scaling(gold, convention_scope))
    convention_records.extend(detect_source_selection(gold, convention_scope))
    convention_records.extend(
        detect_projection_rule(gold, delivered, convention_scope, closure)
    )
    convention_records.extend(detect_row_populated(
        gold, delivered, convention_scope, targets
    ))
    convention_records.extend(detect_aggregate_scope(gold, delivered, targets))
    convention_records = [
        rec for rec in convention_records
        if not (set(rec.get("cell_keys", [])) & custom_claimed)
    ]
    records = method_records + convention_records

    unique = []
    seen = set()
    claimed_keys = set()
    for rec in records:
        sig = (rec["family"], tuple(rec["cells"]))
        if sig in seen:
            continue
        seen.add(sig)
        unique.append(rec)
        claimed_keys.update(rec.get("cell_keys", []))

    defects = detect_defects(gold, targets, scope)

    for rec in unique:
        rec["leak_flag"] = any(c in closure for c in rec.get("cell_keys", []))

    drift = check_registry_drift({r["family"] for r in unique})
    if drift:
        raise SystemExit("registry drift:\n  " + "\n  ".join(drift))

    ctx = {
        "gold": gold,
        "delivered": delivered,
        "targets": closure,
        "raw_targets": target_set,
    }
    unique = apply_ship_when(unique, ctx)
    unique = arbitrate_overlapping(unique)
    unique = resolve_unused_conflicts(unique)
    # Graded-output closure gate, replacing the old Section-2-only rule that
    # let 0353's zero-value claim over a graded band and 0441/0620/0672's
    # alias/copied-method leaks ship. liquidation_preference keeps its 0668
    # exception inside suppress_closure_reach.
    unique = suppress_closure_reach(unique, closure)

    claimed = {c for rec in unique for c in rec.get("cell_keys", [])}
    by_disposition = Counter(r["disposition"] for r in unique)
    payload = {
        "schema_version": "2.0",
        "task": task_dir.name,
        **records_provenance(selection),
        "records": sorted(unique, key=lambda r: (r["disposition"], r["family"], r["band"] or "")),
        "method_assessments": method_assessments,
        "custom_calibration": custom_calibration,
        "defects": defects,
        "summary": {
            "records": len(unique),
            "by_disposition": dict(by_disposition),
            "disclosed": by_disposition.get("disclosed", 0),
            "suppressed": by_disposition.get("suppressed", 0),
            "unclassified": by_disposition.get("unclassified", 0),
            "uncited_records": sum(1 for r in unique if not r.get("entry")),
            "selected_cells": len(scope),
            "claimed_selected_cells": len(scope & claimed),
            "unexplained_cells": len(scope - claimed),
            "graded_closure_cells": len(closure),
            "leak_flags": sum(1 for r in unique if r.get("leak_flag")),
            "method_assessments": dict(Counter(
                a.get("status", "unknown") for a in method_assessments
            )),
            "custom_claimed_cells": len(custom_claimed),
            "legacy_custom_calibration": custom_calibration.get("outcomes", {}),
        },
    }
    return payload


def cmd_roles(args):
    """Collect unique Filter 1 role collisions. No formulas or values."""
    task_dir = Path(args.task_dir).resolve()
    selection = read_stage(task_dir, "bands", Path(args.runs_root))
    gold = Book(Path(selection["golden"]))
    prepared = prepare_method_items(gold, selection.get("bands", []))
    cases = collect_ambiguous_role_cases(prepared)
    payload = {
        "schema_version": "1.0",
        "task": task_dir.name,
        "cases": cases,
        "case_count": len(cases),
        "band_count": sum(c["band_count"] for c in cases),
    }
    out = Path(args.out) if args.out else run_dir(task_dir, Path(args.runs_root)) / "ambiguous_roles.json"
    write_json(out, payload)
    print(
        f"{task_dir.name}: {payload['case_count']} ambiguous-role case(s) "
        f"covering {payload['band_count']} band(s) -> {out}"
    )


def cmd_detect(args):
    payload = detect_records(args)
    out = Path(args.out) if args.out else run_dir(Path(args.task_dir), Path(args.runs_root)) / "records.json"
    write_json(out, payload)
    s = payload["summary"]
    print(f"{payload['task']}: {s['records']} records "
          f"({s['disclosed']} disclosed, {s['suppressed']} suppressed, "
          f"{s['unclassified']} unclassified), {s['unexplained_cells']} cells no entry explains")


# --------------------------------------------------------------------------- probe/context/write/verify


def cmd_probe(args):
    task_dir = Path(args.task_dir).resolve()
    selection = read_stage(task_dir, "bands", Path(args.runs_root))
    started = time.time()
    payload = {
        "schema_version": "1.0",
        "task": task_dir.name,
        "status": "heuristic",
        "reason": "Per-alternative evaluator perturbation is not run in this pre-run implementation unless ast_out/seg_out artifacts are present.",
        "bands": [
            {
                "band": b["band"],
                "cells": b["cells"],
                "can_move_graded_answer": True,
                "measurement": "selected_by_closure",
                "divergence": {},
            }
            for b in selection.get("bands", [])
        ],
        "elapsed_s": round(time.time() - started, 3),
    }
    out = Path(args.out) if args.out else run_dir(task_dir, Path(args.runs_root)) / "probe.json"
    write_json(out, payload)
    print(f"{task_dir.name}: probe wrote {len(payload['bands'])} heuristic band measurements")


def role_hints(text: str) -> list[str]:
    patterns = {
        "depreciation_amortization": (r"\bdepreciat", r"\bamorti[sz]", r"\bd&a\b"),
        "interest_expense_income": (r"\binterest", r"\bfinance cost"),
        "tax": (r"\btax", r"\bcurrent tax", r"\bcash tax"),
        "revenue": (r"\brevenue", r"\bsales\b", r"\bturnover"),
        "operating_expense": (r"\bopex\b", r"\boperating expense", r"\bsg&a\b", r"\brent\b"),
        "working_capital": (r"\bworking capital\b", r"\breceivable", r"\binventor", r"\bpayable"),
        "capital_expenditure": (r"\bcapex\b", r"\bcapital expenditure"),
        "discounting_valuation": (r"\bdiscount", r"\bpresent value", r"\bterminal value", r"\bnpv\b"),
        "returns": (r"\birr\b", r"\bxirr\b", r"\bmoic\b"),
        "cash_flow_metric": (r"\bcash flow\b", r"\bfcf\b"),
        "profit_metric": (r"\bebitda\b", r"\bebit\b", r"\bnet profit\b"),
    }
    lowered = text.lower()
    return [role for role, pats in patterns.items() if any(re.search(p, lowered) for p in pats)]


def cmd_context(args):
    task_dir = Path(args.task_dir).resolve()
    selection = read_stage(task_dir, "bands", Path(args.runs_root))
    records_path = run_dir(task_dir, Path(args.runs_root)) / "records.json"
    claimed = set()
    if records_path.exists():
        records = json.loads(records_path.read_text(encoding="utf-8")).get("records", [])
        claimed = {c for r in records for c in r.get("cell_keys", [])}
    context = []
    for band in selection.get("bands", []):
        if any(c in claimed for c in band.get("cell_keys", [])):
            continue
        text = " ".join([band.get("label", "")] + band.get("formula_samples", []))
        context.append({
            "band": band["band"],
            "label": band.get("label", ""),
            "cells": band.get("cells", []),
            "formula_samples": band.get("formula_samples", []),
            "role_hints": role_hints(text),
            "instruction": "A model may propose role and plausible alternatives here; it must not decide disposition.",
        })
    payload = {"schema_version": "1.0", "task": task_dir.name, "residue": context}
    out = Path(args.out) if args.out else run_dir(task_dir, Path(args.runs_root)) / "context.json"
    write_json(out, payload)
    print(f"{task_dir.name}: {len(context)} residue bands for model context")


def compact_cells(cells: list[str]) -> str:
    if not cells:
        return ""
    if len(cells) == 1 or "!row " in cells[0] or " members" in cells[0]:
        return "`%s`" % cells[0]
    parsed = []
    for c in cells:
        sheet, coord = c.rsplit("!", 1)
        p = split_coord(coord)
        if not p:
            return ", ".join(f"`{x}`" for x in cells[:4])
        parsed.append((sheet, p[1], col_to_num(p[0]), coord))
    if len({p[0] for p in parsed}) == 1 and len({p[1] for p in parsed}) == 1:
        parsed.sort(key=lambda x: x[2])
        return f"`{parsed[0][0]}!{parsed[0][3]}:{parsed[-1][3]}`"
    # A copied-down family collapses the same way a row band does.
    if len({p[0] for p in parsed}) == 1 and len({p[2] for p in parsed}) == 1:
        parsed.sort(key=lambda x: x[1])
        return f"`{parsed[0][0]}!{parsed[0][3]}:{parsed[-1][3]}`"
    shown = ", ".join(f"`{c}`" for c in cells[:4])
    return shown + (f" and {len(cells) - 4} more" if len(cells) > 4 else "")


def stated_orientation(rec: dict) -> str:
    """How a record's stated cells lie: single, row, column, or mixed."""
    parsed = []
    for c in rec.get("cells") or []:
        if not isinstance(c, str) or "!" not in c or "!row " in c:
            continue
        sheet, coord = c.rsplit("!", 1)
        p = split_coord(coord)
        if p:
            parsed.append((sheet, p[1], col_to_num(p[0])))
    if not parsed:
        return "none"
    if len(parsed) == 1:
        return "single"
    if len({p[0] for p in parsed}) == 1 and len({p[1] for p in parsed}) == 1:
        return "row"
    if len({p[0] for p in parsed}) == 1 and len({p[2] for p in parsed}) == 1:
        return "column"
    return "mixed"


def agent_records(records: list[dict]) -> list[dict]:
    """Only cited, disclosed records reach the agent."""
    out = []
    for rec in records:
        if not rec.get("entry"):
            continue
        if rec.get("leak_flag") and rec.get("entry") in METHOD_ENTRY_IDS:
            continue
        if rec.get("disposition") != "disclosed":
            continue
        if not render_sentence(rec):
            continue
        out.append({k: v for k, v in rec.items()
                    if k not in ("evidence", "divergence", "declined_reason")})
    return out


def render_sentence(rec: dict) -> str:
    """Fill the registry entry's sentence for this record's value.

    Wording lives in REGISTRY.md so the phrasing sits beside the definition that
    licenses it. A record whose placeholders cannot be filled renders nothing
    rather than emitting a half-written sentence.
    """
    entry = registry().get(rec.get("entry") or "", {})
    template = entry.get("sentences", {}).get(rec.get("value") or "")
    if not template:
        return ""
    fields = rec.get("fields") or {}
    # An empty field means the detector could not name the thing cleanly. Render
    # nothing rather than a sentence with a hole in it. A quoted empty label
    # ('""') is the same hole wearing quotes: 0666 shipped a sentence claiming
    # its target sat on a row labelled "".
    if any(not str(v).strip() or str(v).strip() in ('""', "''") for v in fields.values()):
        return ""
    try:
        text = template.format(**fields)
    except (KeyError, IndexError):
        return ""
    if "{" in text:
        return ""
    # The registry method template says "copied-column calculation". For a
    # single-cell stated scope that claim is false (0248, 0462, 0518, 0537,
    # 0605), and for a copied-DOWN family the direction is wrong. Rewrite the
    # scope wording deterministically from the stated cells' geometry; the
    # registry keeps authority over everything else in the sentence.
    if "copied-column calculation" in text:
        orientation = stated_orientation(rec)
        if orientation == "single":
            text = text.replace("copied-column calculation", "single-cell calculation")
            representative = str(fields.get("representative") or "")
            if representative:
                text = text.replace(", shown for %s:" % representative, ":")
        elif orientation == "column":
            text = text.replace(
                "copied-column calculation", "calculation copied down the rows"
            )
    return text


def render_section(records: list[dict]) -> str:
    """One bullet per band, joining the sentences of every record on that band."""
    if not records:
        return ""
    lines = [
        SECTION_START,
        "",
        "These are choices the original model made that the delivered file no longer shows.",
        "They describe how the model works, not the figures you are asked to report.",
        "",
    ]
    by_band: dict = {}
    for rec in records:
        by_band.setdefault(rec.get("band") or "", []).append(rec)
    seen_text: set = set()
    for band in sorted(by_band):
        sentences = [s for s in (render_sentence(r) for r in by_band[band]) if s]
        if not sentences:
            continue
        body = " ".join(dict.fromkeys(sentences))
        cells = compact_cells(by_band[band][0].get("cells", []))
        # A colliding sentence is re-anchored, never dropped. 0618 and 0632 both
        # lost a disclosed record this way: same label, byte-identical body.
        if body in seen_text:
            sheet = ""
            first = (by_band[band][0].get("cells") or [""])[0]
            if isinstance(first, str) and "!" in first:
                sheet = first.split("!", 1)[0].strip("`'\"")
            candidates = []
            if sheet:
                candidates.append("On the %s sheet, %s%s" % (
                    sheet, body[0].lower(), body[1:],
                ))
            cells_plain = cells.strip("`")
            if cells_plain:
                candidates.append("At %s, %s%s" % (
                    cells_plain, body[0].lower(), body[1:],
                ))
            for cand in candidates:
                if cand not in seen_text:
                    body = cand
                    break
        seen_text.add(body)
        lines.append(f"- {cells}: {body}")
    return "\n".join(lines) + "\n"


def numeric_targets(task_dir: Path) -> list[float]:
    targets, _ = load_key(task_dir)
    return [
        float(v)
        for v in targets.values()
        if isinstance(v, (int, float)) and not isinstance(v, bool)
    ]


def internal_tokens() -> list[str]:
    """Entry ids and value tokens, none of which may appear in agent-facing text."""
    tokens = set(registry())
    for entry in registry().values():
        tokens.update(a for a in entry.get("alternatives", []) if "_" in a)
    return sorted(tokens)


def same_number(a: float, b: float) -> bool:
    return abs(a - b) <= 1e-12 * max(1.0, abs(a), abs(b))


def trusted_formula_literal_counts(records: list[dict], task_dir: Path) -> Counter:
    """Count independently sourced AST literals that may collide with targets.

    A formula literal is not evidence that a target value was disclosed merely
    because an unrelated graded result happens to have the same number. The
    exemption is deliberately provenance-bound: only a complete custom-method
    rendering on ungraded cells, with no graded references, receives a budget
    for the numeric AST literals that actually occur in its rendered sentence.
    Any additional occurrence of the same number remains a leak.
    """
    targets, _ = load_key(task_dir)
    target_cells = {
        parse_ref(ref, "")
        for ref in targets
        if "!" in ref
    }
    allowed: Counter = Counter()
    for rec in records:
        profile = rec.get("method_profile") or {}
        cells = set(profile.get("cells") or rec.get("cell_keys") or [])
        references = set(profile.get("references") or [])
        if (
            rec.get("source") != "custom_method_detector"
            or rec.get("disposition") != "disclosed"
            or rec.get("leak_flag")
            or not rec.get("coverage_complete")
            or not profile.get("complete")
            or target_cells & (cells | references)
        ):
            continue
        sentence = render_sentence(rec)
        rendered_numbers = []
        for raw in NUMBER_RE.findall(sentence):
            try:
                rendered_numbers.append(float(raw.replace(",", "")))
            except ValueError:
                continue
        used = set()
        for literal in profile.get("numbers") or []:
            if isinstance(literal, bool) or not isinstance(literal, (int, float)):
                continue
            for i, rendered in enumerate(rendered_numbers):
                if i not in used and same_number(rendered, float(literal)):
                    used.add(i)
                    allowed[rendered] += 1
                    break
    return allowed


def significant_digits(raw: str) -> int:
    """Digits a rendered numeric literal carries, leading zeros excluded."""
    mantissa = re.split(r"[eE]", raw)[0]
    digits = re.sub(r"[^0-9]", "", mantissa).lstrip("0")
    return len(digits)


def audit_text(section: str, task_dir: Path, records: list[dict] | None = None,
               closure: set[str] | None = None) -> list[str]:
    faults = []
    formula_lines = [line for line in section.splitlines() if FORMULA_RE.search(line)]
    if formula_lines:
        faults.append(f"{len(formula_lines)} formula-shaped line(s)")
    leaked = [t for t in internal_tokens() if re.search(r"\b%s\b" % re.escape(t), section)]
    if leaked:
        faults.append(f"internal taxonomy token(s) in agent text: {leaked[:5]}")
    targets = numeric_targets(task_dir)
    allowed = trusted_formula_literal_counts(records or [], task_dir)
    # Numeric-collision policy (0469). Two tiers:
    #
    # (a) Provenance. A literal rendered from a record that reaches the
    #     graded-output closure is answer material however it is spelled, so
    #     any target match refuses with no specificity floor. Records that
    #     reach the closure are already suppressed upstream; this is the
    #     fail-closed net for one that slips through.
    # (b) Pure value coincidence. Refuse only high-specificity matches: the
    #     literal must carry at least four significant digits. A small round
    #     integer such as the divisor 6 colliding with a graded 6.0 is a
    #     spurious collision (0469's refusal), not a leak - low-specificity
    #     constants are ordinary formula plumbing, and tier (a) has already
    #     proven the record they came from does not reach the closure.
    #     Zero and +/-1 targets stay exempt as control literals.
    closure_literals: set[float] = set()
    if records and closure:
        for rec in records:
            if record_reach(rec) & set(closure):
                for number in (rec.get("method_profile") or {}).get("numbers") or []:
                    try:
                        closure_literals.add(float(number))
                    except (TypeError, ValueError):
                        continue
    for raw in NUMBER_RE.findall(section):
        try:
            val = float(raw.replace(",", ""))
        except ValueError:
            continue
        matched_target = None
        for target in targets:
            if abs(float(target)) <= 1e-12 or (
                float(target).is_integer() and abs(target) <= 1
            ):
                continue
            if same_number(val, float(target)):
                matched_target = target
                break
        if matched_target is None:
            continue
        permitted = next(
            (number for number, count in allowed.items()
             if count and same_number(val, number)),
            None,
        )
        if permitted is not None:
            allowed[permitted] -= 1
            continue
        if any(same_number(val, lit) for lit in closure_literals):
            faults.append(
                f"numeric literal {raw} matches target {matched_target} and its "
                "provenance record reaches the graded-output closure"
            )
        elif significant_digits(raw) >= 4:
            faults.append(f"numeric literal {raw} matches target {matched_target}")
    return faults


def clone_task(source: Path, out: Path):
    if out.exists():
        shutil.rmtree(out)
    shutil.copytree(source, out, ignore=shutil.ignore_patterns(
        "*-traces", "__pycache__", "*.pyc", "gap_attrib.py", "verify_channels.py"
    ))


def write_disclosure(args) -> dict:
    source_task = Path(args.task_dir).resolve()
    records_payload = read_stage(source_task, "records", Path(args.runs_root))
    safe = agent_records(records_payload.get("records", []))
    section = render_section(safe)
    closure = None
    try:
        selection = read_stage(source_task, "bands", Path(args.runs_root))
        closure = graded_closure(
            Book(Path(selection["golden"])), set(selection["target_keys"])
        )
    except Exception:
        closure = None
    disclosed = [
        r for r in records_payload.get("records", [])
        if r.get("disposition") == "disclosed"
    ]
    faults = audit_text(
        section, source_task, records=disclosed or safe, closure=closure
    )
    if faults and not args.force:
        raise SystemExit("refusing to write disclosure: " + "; ".join(faults[:5]))
    dst = Path(args.out).resolve() if args.out else source_task
    if dst != source_task:
        clone_task(source_task, dst)
    instruction_path = dst / "instruction.md"
    text = instruction_path.read_text(encoding="utf-8")
    text = strip_agent_sections(text)
    if OUTPUT_HEADING not in text:
        raise SystemExit(f"{instruction_path} has no Output heading")
    text = text.replace(OUTPUT_HEADING, "\n" + section + "\n## Output\n", 1)
    if not args.dry_run:
        instruction_path.write_text(text, encoding="utf-8")
        reviewer = {
            **records_payload,
            "agent_records": safe,
            "audit": {"faults": faults, "written_records": len(safe)},
        }
        write_json(dst / "tests" / "disclosure.json", reviewer)
    return {"task": source_task.name, "out": str(dst), "written_records": len(safe), "faults": faults}


def strip_agent_sections(text: str) -> str:
    """Remove disclosure sections written by either old tool or this one."""
    for heading in STALE_AGENT_SECTIONS:
        while heading in text:
            head, rest = text.split(heading, 1)
            parts = rest.split("\n## ", 1)
            text = head + ("\n## " + parts[1] if len(parts) > 1 else "")
    return text


def cmd_write(args):
    result = write_disclosure(args)
    print(f"{result['task']}: wrote {result['written_records']} disclosure record(s) -> {result['out']}")


# --------------------------------------------------------------------------- faithcheck
#
# Runs after `write` and before `verify`. Every claim a written record makes
# is mechanically re-derived from the golden workbook and the staged
# artifacts; any divergence is a fault and the command exits nonzero. Pure
# read-and-check: it writes only runs/disclosure/<task>/faithcheck.json,
# which is never staged for the agent, and it never copies golden cached
# values anywhere.

FAITHCHECK_SCOPE_ENTRIES = METHOD_ENTRY_IDS | {"projection_rule", "distribution_policy"}


def _fault(faults: list, rec, kind: str, claim: str, expected: str = "", found: str = ""):
    faults.append({
        "record": rec.get("band") if isinstance(rec, dict) else rec,
        "entry": rec.get("entry") if isinstance(rec, dict) else None,
        "kind": kind,
        "claim": claim,
        "expected": expected,
        "found": found,
    })


def _unquote_label(text) -> str:
    text = str(text or "").strip()
    if len(text) >= 2 and text[0] == '"' and text[-1] == '"':
        return text[1:-1]
    return text


def _stated_keys(rec: dict) -> list[str]:
    out = []
    for c in rec.get("cells") or []:
        if isinstance(c, str) and "!" in c and "!row " not in c:
            sheet, coord = c.rsplit("!", 1)
            if split_coord(coord):
                out.append(key(sheet, coord))
    return out


def _claimed_rows(rec: dict) -> set[tuple]:
    rows = set()
    for c in rec.get("cell_keys") or []:
        rk = _row_key(c)
        if rk:
            rows.add(rk)
    for c in rec.get("cells") or []:
        m = re.match(r"^`?(.+)!row (\d+)`?$", str(c))
        if m:
            rows.add((unquote_sheet(m.group(1)), int(m.group(2))))
    return rows


def faithcheck_task(task_dir: Path, runs_root: Path = DEFAULT_RUNS_ROOT,
                    golden: str | None = None) -> dict:
    """Mechanically re-derive every claim in the written disclosure.

    Four claim families, mirroring the observed blocker classes:

    a. row labels (0598, 0605, 0622, 0660, 0661, 0669): every stated target
       and operand label is re-resolved with the leftward resolver;
    b. copied scope (0248, 0462, 0518, 0528, 0537, 0595, 0605, 0622): the
       stated band must be the maximal same-mechanics golden span, phrased
       for its true geometry;
    c. reference completeness (0523, 0648, 0658): every reference, sign, and
       literal of the representative golden formula must appear in the
       rendered mechanics, including pinned-reference lock behaviour and
       undisclosed bare cross-sheet source links;
    d. graded-output closure (0353, 0441, 0620, 0672): no disclosed record
       may reach a graded answer, its equality aliases, or its same-row
       copied equivalents.
    """
    disclosure_path = task_dir / "tests" / "disclosure.json"
    if not disclosure_path.exists():
        raise SystemExit(f"faithcheck needs {disclosure_path}; run write first")
    payload = json.loads(disclosure_path.read_text(encoding="utf-8"))
    gold = Book(find_golden(task_dir, golden))
    delivered = Book(find_environment(task_dir))
    targets_map, _ = load_key(task_dir)
    default_sheet = gold.sheets[0]
    targets = {parse_ref(t, default_sheet) for t in targets_map}
    closure = graded_closure(gold, targets)

    disclosed = [
        rec for rec in payload.get("records", [])
        if rec.get("disposition") == "disclosed"
        and rec.get("entry")
        and render_sentence(rec)
    ]

    faults: list[dict] = []

    truncated: set[str] = set()
    referenced_rows: set[tuple] = set()
    if any(rec.get("entry") == "row_populated" for rec in disclosed):
        for k, formula in gold.formula.items():
            for ref in refs_in(formula, k.split("!", 1)[0], truncated=truncated):
                rk = _row_key(ref)
                if rk:
                    referenced_rows.add(rk)

    for rec in disclosed:
        sentence = render_sentence(rec)
        entry = rec.get("entry")
        fields = rec.get("fields") or {}
        cells = list(rec.get("cell_keys") or [])
        stated = _stated_keys(rec)
        anchor_cells = stated or cells

        # ------------------------------------------------ (a) target label
        claimed_label = _unquote_label(fields.get("label"))
        if claimed_label and anchor_cells:
            expected_label = gold.row_label(anchor_cells[0])
            if expected_label != claimed_label:
                _fault(
                    faults, rec, "row_label",
                    "stated row label must equal the leftward-resolver label",
                    expected=expected_label, found=claimed_label,
                )

        # ------------------- (a)/(c) operand labels, references, literals
        steps = str(fields.get("steps") or "")
        if steps and cells:
            profile = formula_profile(gold, {
                "band": rec.get("band"),
                "cell_keys": cells,
                "label": gold.row_label(cells[0]),
                "pattern": "",
            })
            if not profile.get("complete"):
                _fault(
                    faults, rec, "reference",
                    "representative golden formula could not be re-parsed for the coverage proof",
                    expected="parseable formula", found=str(profile.get("error")),
                )
            else:
                sheet0 = cells[0].split("!", 1)[0]
                for raw in profile.get("raw_references") or []:
                    rendered = describe_ref(gold, raw, sheet0, profile.get("label", ""))
                    if rendered not in steps:
                        _fault(
                            faults, rec, "reference",
                            "every golden reference must appear in the rendered "
                            "mechanics with its re-derived label",
                            expected=rendered,
                            found="absent or differently labelled in the record's steps",
                        )
                for literal in profile.get("literal_renderings") or []:
                    if literal not in steps:
                        _fault(
                            faults, rec, "reference",
                            "every golden literal must appear in the rendered mechanics",
                            expected=literal, found="absent from the record's steps",
                        )
                lock_cells = anchor_cells if len(anchor_cells) > 1 else cells
                pinned, moving = _copy_lock_split(gold, lock_cells)
                if pinned and moving and "fixed" not in steps:
                    _fault(
                        faults, rec, "lock",
                        "a copy pattern pinning some references must state the lock behaviour",
                        expected="a clause naming the pinned reference(s): " + ", ".join(pinned),
                        found="no lock clause in the rendered mechanics",
                    )

        # ----------------------- (c) partial-claim entries: sign coverage
        if entry == "stake_scaling" and cells:
            formula = gold.formula.get(cells[0]) or ""
            body = re.sub(r"'(?:[^']|'')+'!", "S!", STRING_RE.sub('""', formula))
            if "-" in body:
                _fault(
                    faults, rec, "reference",
                    "ownership-share sentence omits a negation or subtraction "
                    "present in the golden formula",
                    expected="rendered mechanics carrying the sign",
                    found="fixed stake_scaling sentence with no sign slot",
                )

        # -------------------------------------------------- (b) scope
        if entry in FAITHCHECK_SCOPE_ENTRIES and anchor_cells:
            span_h = full_copied_scope(gold, [anchor_cells[0]], closure)
            span_v = full_copied_scope_vertical(gold, [anchor_cells[0]], closure)
            span = span_h if len(span_h) >= len(span_v) else span_v
            stated_multi = len(anchor_cells) > 1
            if "copied-column calculation" in sentence and not stated_multi and len(span) <= 1:
                _fault(
                    faults, rec, "scope",
                    "single-cell formula phrased as a copied-column calculation",
                    expected="single-cell phrasing",
                    found="copied-column claim over %s" % rec.get("band"),
                )
            if "single-cell calculation" in sentence and len(span) > 1:
                _fault(
                    faults, rec, "scope",
                    "copied family phrased as a single-cell calculation",
                    expected=compact_cells([pretty(c) for c in span]),
                    found=str(rec.get("band")),
                )
            if stated_multi:
                rows = {split_coord(c.split("!", 1)[1])[1] for c in anchor_cells}
                colsn = {col_to_num(split_coord(c.split("!", 1)[1])[0]) for c in anchor_cells}
                if len(rows) == 1:
                    expected_span = full_copied_scope(gold, anchor_cells, closure)
                elif len(colsn) == 1:
                    expected_span = full_copied_scope_vertical(gold, anchor_cells, closure)
                else:
                    expected_span = anchor_cells
                if set(expected_span) - set(anchor_cells):
                    _fault(
                        faults, rec, "scope",
                        "stated scope is narrower than the golden same-mechanics span",
                        expected=compact_cells([pretty(c) for c in expected_span]),
                        found=str(rec.get("band")),
                    )
                for a, b in zip(anchor_cells, anchor_cells[1:]):
                    if not copy_compatible(gold, a, b):
                        _fault(
                            faults, rec, "scope",
                            "stated band mixes different mechanics",
                            expected="one copy family per record",
                            found="%s against %s" % (pretty(a), pretty(b)),
                        )
                        break
            elif len(span) > 1:
                _fault(
                    faults, rec, "scope",
                    "stated scope is a fragment of a copied family",
                    expected=compact_cells([pretty(c) for c in span]),
                    found=str(rec.get("band")),
                )

        # ------------------------------ (b) row_populated truth re-checks
        if entry == "row_populated":
            for rk in sorted(_claimed_rows(rec)):
                sheet, rownum = rk
                if rec.get("value") == "populated_but_unread":
                    if sheet in truncated:
                        _fault(
                            faults, rec, "reference",
                            "unread claim is unverifiable: the reference scan "
                            "truncated an oversized range on this sheet",
                            expected="verifiable absence of inbound references",
                            found="oversized ranges skipped",
                        )
                    elif rk in referenced_rows:
                        _fault(
                            faults, rec, "reference",
                            "row claimed unread is read by a golden formula",
                            expected="no inbound reference anywhere in the golden",
                            found="a golden formula references the row",
                        )
                if rec.get("value") == "unused":
                    if any(
                        k in gold.formula or isinstance(gold.value.get(k), (int, float))
                        for k in gold.row_cells(sheet, rownum)
                    ):
                        _fault(
                            faults, rec, "scope",
                            "row claimed unused carries golden data",
                            expected="an empty golden row",
                            found="formulas or values on the row",
                        )

        # ------------------------------------------------ (d) closure
        if entry != "liquidation_preference" and not rec.get("accepted_target_reference"):
            hit = sorted(record_reach(rec) & closure)
            if hit:
                _fault(
                    faults, rec, "closure",
                    "record reaches the graded-output closure and must be "
                    "suppressed, not disclosed",
                    expected="suppressed",
                    found="disclosed; reaches " + ", ".join(pretty(h) for h in hit[:4]),
                )

    # -------------------- (c) undisclosed bare cross-sheet source links
    selection = None
    try:
        selection = read_stage(task_dir, "bands", runs_root)
    except (FileNotFoundError, OSError, json.JSONDecodeError):
        selection = None
    if selection is not None:
        covered = {c for r in disclosed for c in (r.get("cell_keys") or [])}
        for band in selection.get("bands", []):
            keys = band.get("cell_keys") or []
            if len(keys) != 1:
                continue
            k = keys[0]
            formula = (gold.formula.get(k) or "").strip()
            if not formula or not BARE_LINK_RE.match(formula):
                continue
            src = bare_alias_ref(gold, k)
            if not src or src.split("!", 1)[0] == k.split("!", 1)[0]:
                continue
            if k in closure or src in closure:
                continue
            if k in covered:
                continue
            if delivered.has(k) or delivered.has(src):
                continue
            if not gold.row_label(k) or not gold.row_label(src):
                continue
            _fault(
                faults, pretty(k), "source_link",
                "material bare cross-sheet source link is not disclosed",
                expected="a source_selection record naming %s and its source %s"
                         % (pretty(k), pretty(src)),
                found="no disclosed record covers this cell",
            )

    return {
        "schema_version": "1.0",
        "task": task_dir.name,
        "golden": str(find_golden(task_dir, golden)),
        "checked_records": len(disclosed),
        "graded_closure_cells": len(closure),
        "faults": faults,
        "passed": not faults,
    }


def cmd_faithcheck(args):
    task_dir = Path(args.task_dir).resolve()
    result = faithcheck_task(task_dir, Path(args.runs_root), getattr(args, "golden", None))
    out = Path(args.out) if args.out else run_dir(task_dir, Path(args.runs_root)) / "faithcheck.json"
    write_json(out, result)
    print(
        f"{result['task']}: faithcheck {'PASS' if result['passed'] else 'FAIL'} "
        f"({result['checked_records']} record(s), {len(result['faults'])} fault(s))"
    )
    for fault in result["faults"][:10]:
        print(f"  - [{fault['kind']}] {fault['record']}: {fault['claim']}")
    if result["faults"]:
        raise SystemExit(2)


def verify_task(task_dir: Path) -> dict:
    disclosure_path = task_dir / "tests" / "disclosure.json"
    faults = []
    if not disclosure_path.exists():
        return {"task": task_dir.name, "passed": False, "faults": ["missing tests/disclosure.json"]}
    payload = json.loads(disclosure_path.read_text(encoding="utf-8"))
    golden_path = Path(payload.get("golden", ""))
    expected_golden_hash = payload.get("golden_sha256")
    if (
        not golden_path.is_file()
        or not isinstance(expected_golden_hash, str)
        or hashlib.sha256(golden_path.read_bytes()).hexdigest()
        != expected_golden_hash
    ):
        faults.append("disclosure golden binding is missing or changed")
    if payload.get("segmentation_mode") != "strict":
        faults.append("production disclosure requires strict segmentation provenance")
    else:
        try:
            generation = payload.get("segmentation_generation")
            if not isinstance(generation, dict) or generation.get("mode") != "strict":
                raise ValueError("strict segmentation generation is missing")
            if payload.get("selection", {}).get("closure_source") != "ast":
                raise ValueError("strict disclosure did not use AST closure")
            generation_id = generation.get("generation_id")
            ast_root = Path(payload.get("ast_dir", ""))
            seg_root = Path(payload.get("seg_root", ""))
            if (
                not isinstance(generation_id, str)
                or not ast_root.is_dir()
                or not seg_root.is_dir()
            ):
                raise ValueError("strict disclosure provenance paths are missing")
            workbook = golden_path.stem
            pipeline_bindings = payload.get("pipeline_bindings")
            if pipeline_bindings is not None:
                if (
                    pipeline_bindings.get("segmentation_generation_id")
                    != generation_id
                    or pipeline_bindings.get("source_generation_id")
                    is None
                ):
                    raise ValueError("disclosure pipeline bindings changed")
                from xl_seg.publication import resolve_generation_by_id

                resolve_generation_by_id(
                    seg_root / workbook,
                    generation_id,
                    source_path=golden_path,
                    ast_dir=ast_root / workbook,
                    source_generation_dir=golden_path.parent.parent,
                    require_pass=True,
                    validate_live_evidence=True,
                )
            else:
                from xl_seg.publication import resolve_for_consumer

                resolve_for_consumer(
                    seg_root / workbook,
                    mode="strict",
                    source_path=golden_path,
                    ast_dir=ast_root / workbook,
                    require_pass=True,
                    expected_generation_id=generation_id,
                )
        except Exception as exc:
            faults.append(f"strict disclosure provenance failed: {exc}")
    delivered = Book(find_environment(task_dir))
    shipped = payload.get("agent_records", [])
    visible_cells = sorted({
        cell
        for record in shipped
        for cell in record.get("cell_keys", [])
    })
    nonblank = [pretty(c) for c in visible_cells if delivered.has(c)]
    if nonblank:
        faults.append(
            f"{len(nonblank)} agent-visible disclosure cells are non-blank "
            "in delivered workbook"
        )
    # One record per (band, entry). Two entries may legitimately describe one band;
    # the writer joins them into a single bullet.
    seen = Counter((r.get("band"), r.get("entry")) for r in payload.get("records", []))
    dupes = [b for b, count in seen.items() if b[0] and count > 1]
    if dupes:
        faults.append(f"duplicate record band/entry pairs: {dupes[:5]}")
    uncited = [r.get("band") for r in shipped if not r.get("entry")]
    if uncited:
        faults.append(f"{len(uncited)} shipped record(s) cite no registry entry")
    unknown = sorted({r.get("entry") for r in shipped if r.get("entry") not in registry()})
    if unknown:
        faults.append(f"shipped record(s) cite unknown registry entries: {unknown[:5]}")
    wrong_disposition = [r.get("band") for r in shipped if r.get("disposition") != "disclosed"]
    if wrong_disposition:
        faults.append(f"{len(wrong_disposition)} shipped record(s) are not disclosed")
    section = ""
    text = (task_dir / "instruction.md").read_text(encoding="utf-8")
    for heading in STALE_AGENT_SECTIONS[:-1]:
        if heading in text:
            faults.append(f"stale agent-facing section remains: {heading}")
    if SECTION_START in text:
        section = SECTION_START + text.split(SECTION_START, 1)[1].split("\n## Output", 1)[0]
    custom = [
        r for r in payload.get("records", [])
        if r.get("source") == "custom_method_detector"
        and r.get("disposition") == "disclosed"
        and not r.get("leak_flag")
    ]
    convention_cells = {
        c for r in payload.get("records", [])
        if r.get("source") == "convention_detector"
        for c in r.get("cell_keys", [])
    }
    overlap = sorted({
        c for r in custom for c in r.get("cell_keys", []) if c in convention_cells
    })
    if overlap:
        faults.append(f"custom and convention records overlap on {len(overlap)} cell(s)")
    missing_render = []
    for rec in custom:
        sentence = render_sentence(rec)
        cells_text = compact_cells(rec.get("cells", []))
        if not sentence or sentence not in section or cells_text not in section:
            missing_render.append(rec.get("band"))
        if not rec.get("coverage_complete"):
            faults.append(f"custom record lacks AST coverage proof: {rec.get('band')}")
    if missing_render:
        faults.append(f"{len(missing_render)} custom record(s) did not render: {missing_render[:5]}")

    gold_book = None
    try:
        gold_book = Book(find_golden(task_dir))
    except Exception:
        gold_book = None
    if custom:
        try:
            golden_path = Path(payload.get("golden", ""))
            expected_golden_hash = payload.get("golden_sha256")
            if (
                golden_path.is_file()
                and isinstance(expected_golden_hash, str)
                and hashlib.sha256(golden_path.read_bytes()).hexdigest()
                == expected_golden_hash
            ):
                gold = Book(golden_path)
            elif gold_book is not None:
                gold = gold_book
            else:
                raise ValueError("disclosure golden binding is missing or changed")
            structural = []
            for rec in custom:
                profile = formula_profile(gold, {
                    "band": rec.get("band"),
                    "cell_keys": rec.get("cell_keys", []),
                    "label": (rec.get("method_profile") or {}).get("label", ""),
                    "pattern": (rec.get("method_profile") or {}).get("pattern", ""),
                })
                if profile.get("complete") and structural_reason(profile):
                    structural.append(rec.get("band"))
            if structural:
                faults.append(
                    f"{len(structural)} custom record(s) are hard structural: {structural[:5]}"
                )
        except Exception as exc:
            faults.append(f"custom structural verification failed: {exc}")
    closure = None
    if gold_book is not None:
        try:
            targets_map, _ = load_key(task_dir)
            default_sheet = gold_book.sheets[0]
            closure = graded_closure(
                gold_book, {parse_ref(t, default_sheet) for t in targets_map}
            )
        except Exception:
            closure = None
    disclosed_records = [
        r for r in payload.get("records", [])
        if r.get("disposition") == "disclosed"
    ]
    faults.extend(audit_text(
        section, task_dir, records=disclosed_records or shipped, closure=closure
    ))
    return {"task": task_dir.name, "passed": not faults, "faults": faults}


def cmd_verify(args):
    result = verify_task(Path(args.task_dir).resolve())
    out = Path(args.out) if args.out else run_dir(Path(args.task_dir), Path(args.runs_root)) / "verify.json"
    write_json(out, result)
    print(f"{result['task']}: {'PASS' if result['passed'] else 'FAIL'}")
    for fault in result["faults"][:10]:
        print(f"  - {fault}")
    if result["faults"] and not args.no_fail:
        raise SystemExit(2)


# --------------------------------------------------------------------------- bench/migrate


def count_hint_cells(task_dir: Path):
    path = task_dir / "tests" / "formula_hints.json"
    if not path.exists():
        return 0, 0, 0, 0
    spec = json.loads(path.read_text(encoding="utf-8"))
    delivered = Book(find_environment(task_dir))
    hints = spec.get("hints", [])
    all_cells = []
    full_redundant = partial = 0
    for hint in hints:
        cells = sorted({c for band in hint.get("bands", []) for c in cells_from_band_ref(band)})
        filled = [c for c in cells if delivered.has(c)]
        all_cells.extend(cells)
        if cells and len(filled) == len(cells):
            full_redundant += 1
        elif filled:
            partial += 1
    filled_total = sum(1 for c in all_cells if delivered.has(c))
    return len(hints), len(set(all_cells)), filled_total, full_redundant + partial


def convention_counts(task_dir: Path):
    path = task_dir / "tests" / "conventions.json"
    if not path.exists():
        return 0, Counter()
    data = json.loads(path.read_text(encoding="utf-8"))
    return len(data.get("conventions", [])), Counter(r.get("family") for r in data.get("conventions", []))


def bench_tasks(tasks_root: Path, runs_root: Path = DEFAULT_RUNS_ROOT) -> dict:
    tasks = sorted(tasks_root.glob("*-outputs"))
    total = Counter()
    families = Counter()
    task_rows = []
    for task in tasks:
        hints, hint_cells, filled, redundant_hints = count_hint_cells(task)
        conventions, fams = convention_counts(task)
        families.update(fams)
        disclosure_path = task / "tests" / "disclosure.json"
        disclosed = 0
        if disclosure_path.exists():
            disclosed = len(json.loads(disclosure_path.read_text(encoding="utf-8")).get("agent_records", []))
        instruction = (task / "instruction.md").read_text(encoding="utf-8") if (task / "instruction.md").exists() else ""
        old_hint_section = int("## Custom formula hints" in instruction)
        old_manifest_section = int("## Modelling conventions in this workbook" in instruction)
        unified_section = int(SECTION_START in instruction)
        total.update({
            "tasks": 1,
            "hints": hints,
            "hint_cells": hint_cells,
            "filled_hint_cells": filled,
            "redundant_or_partial_hints": redundant_hints,
            "convention_records": conventions,
            "disclosure_records": disclosed,
            "agent_custom_hint_sections": old_hint_section,
            "agent_old_manifest_sections": old_manifest_section,
            "agent_unified_sections": unified_section,
        })
        task_rows.append({
            "task": task.name,
            "hints": hints,
            "hint_cells": hint_cells,
            "filled_hint_cells": filled,
            "redundant_or_partial_hints": redundant_hints,
            "convention_records": conventions,
            "disclosure_records": disclosed,
            "agent_custom_hint_section": bool(old_hint_section),
            "agent_old_manifest_section": bool(old_manifest_section),
            "agent_unified_section": bool(unified_section),
        })
    return {
        "schema_version": "1.0",
        "tasks_root": str(tasks_root),
        "totals": dict(total),
        "convention_families": dict(families),
        "tasks": task_rows,
    }


def cmd_bench(args):
    payload = bench_tasks(Path(args.tasks_root).resolve(), Path(args.runs_root))
    out = Path(args.out) if args.out else Path(args.runs_root) / "bench.json"
    write_json(out, payload)
    t = payload["totals"]
    print(
        f"{t.get('tasks', 0)} tasks; {t.get('hints', 0)} old hints; "
        f"{t.get('filled_hint_cells', 0)} filled hint cells; "
        f"{t.get('convention_records', 0)} old convention records"
    )


def ensure_pipeline(task: Path, args):
    class A:
        pass
    a = A()
    a.task_dir = str(task)
    a.golden = None
    a.ast_dir = args.ast_dir
    a.seg_root = args.seg_root
    a.segmentation_mode = args.segmentation_mode
    a.expected_generation_id = getattr(args, "expected_generation_id", None)
    a.runs_root = args.runs_root
    a.out = None
    a.role_resolutions = getattr(args, "role_resolutions", None)
    cmd_select(a)
    cmd_probe(a)
    cmd_roles(a)
    cmd_detect(a)
    cmd_context(a)


def cmd_migrate(args):
    tasks_root = Path(args.tasks_root).resolve()
    out_root = Path(args.out).resolve()
    out_root.mkdir(parents=True, exist_ok=True)
    rows = []
    for task in sorted(tasks_root.glob("*-outputs")):
        ensure_pipeline(task, args)
        class W:
            pass
        w = W()
        w.task_dir = str(task)
        w.runs_root = args.runs_root
        w.out = str(out_root / task.name)
        w.force = args.force
        w.dry_run = False
        result = write_disclosure(w)
        verify = verify_task(Path(result["out"]))
        rows.append({**result, "verify_passed": verify["passed"], "verify_faults": verify["faults"]})
        print(f"{task.name}: migrated, verify={'PASS' if verify['passed'] else 'FAIL'}")
    payload = {"schema_version": "1.0", "tasks_root": str(tasks_root), "out": str(out_root), "tasks": rows}
    write_json(Path(args.runs_root) / "migration-summary.json", payload)


def cmd_roles_validate(args):
    task_dir = Path(args.task_dir).resolve()
    runs_root = Path(args.runs_root)
    path = Path(args.file) if args.file else default_role_resolutions_path(
        task_dir, runs_root)
    if not path.exists():
        raise SystemExit(f"no arbitration file to validate at {path}")
    cases_path = run_dir(task_dir, runs_root) / "ambiguous_roles.json"
    cases = []
    if cases_path.exists():
        cases = json.loads(cases_path.read_text(encoding="utf-8")).get("cases") or []
    normalized, errors, warnings = validate_role_resolutions(path, cases)
    for warning in warnings:
        print(f"{task_dir.name}: WARNING {warning}")
    if errors:
        for error in errors:
            print(f"{task_dir.name}: ERROR {error}")
        raise SystemExit(
            f"{task_dir.name}: role_resolutions.json failed validation with "
            f"{len(errors)} error(s); re-launch the arbitration agent once "
            "with the exact errors above")
    unresolved = [
        case["case_id"] for case in cases
        if case["case_id"] not in {row["case_id"] for row in normalized["resolutions"]}
    ]
    if unresolved:
        print(f"{task_dir.name}: note: {len(unresolved)} case(s) left "
              f"unresolved (candidates fall through in registry order)")
    write_json(path, normalized)
    print(f"{task_dir.name}: roles-validate PASS "
          f"({len(normalized['resolutions'])} resolution(s)) -> {path}")


# --------------------------------------------------------------------------- cli


def main(argv=None):
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--runs-root", default=str(DEFAULT_RUNS_ROOT))
    sub = parser.add_subparsers(dest="command", required=True)

    b = sub.add_parser("bench")
    b.add_argument("--tasks-root", default=str(DEFAULT_TASKS_ROOT))
    b.add_argument("--out", default=None)

    for name in ("select", "probe", "roles", "roles-validate", "detect",
                 "context", "write", "faithcheck", "verify"):
        p = sub.add_parser(name)
        p.add_argument("--task-dir", required=True)
        p.add_argument("--out", default=None)
        if name in ("select", "faithcheck"):
            p.add_argument("--golden", default=None)
        if name == "select":
            p.add_argument("--ast-dir", default=None)
            p.add_argument("--seg-root", default=None)
            p.add_argument(
                "--segmentation-mode",
                choices=("strict", "shadow", "legacy"),
                default="strict",
            )
            p.add_argument("--expected-generation-id", default=None)
            p.add_argument("--release-root", default=None)
            p.add_argument("--release-id", default=None)
            p.add_argument("--source-generation-root", default="source_out")
            p.add_argument("--source-generation-id", default=None)
            p.add_argument("--segmentation-generation-id", default=None)
        if name == "roles-validate":
            p.add_argument("--file", default=None)
        if name == "detect":
            p.add_argument("--role-resolutions", default=None)
        if name == "write":
            p.add_argument("--force", action="store_true")
            p.add_argument("--dry-run", action="store_true")
        if name == "verify":
            p.add_argument("--no-fail", action="store_true")

    m = sub.add_parser("migrate")
    m.add_argument("--tasks-root", default=str(DEFAULT_TASKS_ROOT))
    m.add_argument("--out", required=True)
    m.add_argument("--ast-dir", default="ast_out")
    m.add_argument("--seg-root", default="seg_out")
    m.add_argument(
        "--segmentation-mode",
        choices=("strict", "shadow", "legacy"),
        default="strict",
    )
    m.add_argument("--expected-generation-id", default=None)
    m.add_argument("--force", action="store_true")
    m.add_argument("--role-resolutions", default=None)

    args = parser.parse_args(argv)
    {
        "bench": cmd_bench,
        "select": cmd_select,
        "probe": cmd_probe,
        "roles": cmd_roles,
        "roles-validate": cmd_roles_validate,
        "detect": cmd_detect,
        "context": cmd_context,
        "write": cmd_write,
        "faithcheck": cmd_faithcheck,
        "verify": cmd_verify,
        "migrate": cmd_migrate,
    }[args.command](args)


if __name__ == "__main__":
    main()

