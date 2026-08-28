#!/usr/bin/env python3
"""Focused regressions for the three systematic disclosure-writer defects.

1. Copied-column scope omissions: a record must state the maximal
   same-mechanics span in the golden workbook, not the selected subset
   (0233, 0518, 0522, 0528, 0533).
2. Distribution/waterfall mechanics misstatements: MIN caps, zero floors and
   share multipliers must be rendered, never dropped or contradicted
   (0256, 0350, 0352, 0353, 0646).
3. False row labels: resolve by walking left from the referenced cell,
   skipping formula numeric/error displays, blanks, units, scale stamps and
   scenario markers, accepting linked text labels (0233, 0350, 0353, 0518,
   0600, 0605, 0666).

Synthetic workbooks cover the mechanics that need exact shapes; the retained
golden workbooks under "FCP Workbooks/4-10 100" cover the observed cases.
Golden-dependent sections skip with a notice when the workbook is absent.
"""

from __future__ import annotations

import json
import tempfile
from pathlib import Path

import openpyxl

import disclose


GOLDEN_DIR = Path(__file__).resolve().parents[5] / "FCP Workbooks" / "4-10 100"
FAILED_BUNDLE_0233 = (
    Path(__file__).resolve().parents[5]
    / "runs" / "0233-variable-sources" / "failed-staged-bundle"
)

PASSED = []
SKIPPED = []


def check(name: str, condition: bool, detail: str = ""):
    if not condition:
        raise AssertionError(f"{name}: {detail}")
    PASSED.append(name)


def book_from(rows: dict) -> disclose.Book:
    """rows: {sheet: {cellref: value-or-formula-string}} -> Book."""
    wb = openpyxl.Workbook()
    default = wb.active
    first = True
    for sheet, cells in rows.items():
        ws = default if first else wb.create_sheet()
        ws.title = sheet
        first = False
        for ref, value in cells.items():
            ws[ref] = value
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as handle:
        path = Path(handle.name)
    wb.save(path)
    try:
        return disclose.Book(path)
    finally:
        path.unlink()


def golden(name: str) -> disclose.Book | None:
    path = GOLDEN_DIR / name
    if not path.exists():
        SKIPPED.append(name)
        return None
    return disclose.Book(path)


# ------------------------------------------------------------- copied scope


def test_scope_extends_over_unselected_copies():
    """0518 shape: E11 seeds the copy family but only F11:O11 were selected."""
    cells = {"Inputs": {"B11": "Revenue"}}
    for col_num in range(5, 16):  # E..O
        col = disclose.num_to_col(col_num)
        cells["Inputs"][f"{col}9"] = 10
        cells["Inputs"][f"{col}10"] = 2
        cells["Inputs"][f"{col}11"] = f"={col}9*{col}10*10^6"
    gold = book_from(cells)
    selected = [f"Inputs!{disclose.num_to_col(c)}11" for c in range(6, 16)]  # F..O
    bands = disclose.group_bands(gold, selected, set())
    check("scope/left-extension bands", len(bands) == 1, f"{len(bands)} bands")
    check(
        "scope/left-extension ref",
        bands[0]["band"] == "Inputs!E11:O11",
        bands[0]["band"],
    )
    check(
        "scope/selected-cells-preserved",
        bands[0]["cell_keys"] == selected,
        str(bands[0]["cell_keys"]),
    )


def test_scope_extension_stops_at_target():
    cells = {"S": {"B5": "Margin"}}
    for col_num in range(3, 10):  # C..I
        col = disclose.num_to_col(col_num)
        cells["S"][f"{col}4"] = 7
        cells["S"][f"{col}5"] = f"={col}4*2"
    gold = book_from(cells)
    selected = [f"S!{disclose.num_to_col(c)}5" for c in range(5, 10)]  # E..I
    bands = disclose.group_bands(gold, selected, targets={"S!C5"})
    check(
        "scope/target-guard",
        bands[0]["band"] == "S!D5:I5",
        bands[0]["band"],
    )


def test_scope_does_not_merge_different_mechanics():
    """0233 Model!224 shape: adjacent copy families pinning different cells."""
    cells = {"M": {"B2": "Repayment", "R1": 0.1, "S1": 0.2}}
    for col_num in range(3, 8):  # C..G pin $R$1
        col = disclose.num_to_col(col_num)
        cells["M"][f"{col}2"] = f"={col}3*$R$1"
        cells["M"][f"{col}3"] = 1
    for col_num in range(8, 13):  # H..L pin $S$1
        col = disclose.num_to_col(col_num)
        cells["M"][f"{col}2"] = f"={col}3*$S$1"
        cells["M"][f"{col}3"] = 1
    gold = book_from(cells)
    selected = [f"M!{disclose.num_to_col(c)}2" for c in range(3, 13)]
    bands = sorted(b["band"] for b in disclose.group_bands(gold, selected, set()))
    check(
        "scope/no-cross-mechanics-merge",
        bands == ["M!C2:G2", "M!H2:L2"],
        str(bands),
    )


def test_scope_pinned_anchor_counts_as_copy():
    """0533 shape: D19=D18*B19 and E19=E18*B19 are one copy family."""
    gold = book_from({"Multiples": {
        "B19": 5, "C19": "Non-Listed Market Comp",
        "D18": 100, "E18": 200,
        "D19": "=D18*B19", "E19": "=E18*B19",
    }})
    full = disclose.full_copied_scope(gold, ["Multiples!D19"])
    check(
        "scope/pinned-anchor",
        full == ["Multiples!D19", "Multiples!E19"],
        str(full),
    )


def test_scope_hold_flat_seed_included():
    """0233 DCF!31 shape: AG31=AE31 seeds the =prev hold-flat run."""
    cells = {"DCF": {"B31": "% net revenue", "E31": 0.4, "G31": "=E31"}}
    for col_num in range(8, 13):  # H..L each copy their predecessor
        col = disclose.num_to_col(col_num)
        prev = disclose.num_to_col(col_num - 1)
        cells["DCF"][f"{col}31"] = f"={prev}31"
    gold = book_from(cells)
    selected = {f"DCF!{disclose.num_to_col(c)}31" for c in range(8, 13)}  # H..L
    delivered = book_from({"DCF": {"B31": "% net revenue"}})
    records = disclose.detect_projection_rule(gold, delivered, selected, set())
    holds = [r for r in records if r["value"] == "hold_level"]
    check("scope/hold-flat-record", len(holds) == 1, f"{len(holds)} records")
    check(
        "scope/hold-flat-seed",
        holds[0]["band"] == "`DCF!G31:L31`",
        holds[0]["band"],
    )
    check(
        "scope/hold-flat-claims-selected-only",
        set(holds[0]["cell_keys"]) == selected,
        str(holds[0]["cell_keys"]),
    )


def test_scope_observed_0233_bands():
    gold = golden("0233.xlsx")
    if gold is None:
        return
    full = disclose.full_copied_scope(gold, ["Model!N223"])
    check(
        "scope/0233 Model!223 spans J:BT",
        full[0] == "Model!J223" and full[-1] == "Model!BT223",
        f"{full[0]}..{full[-1]}",
    )
    full = disclose.full_copied_scope(gold, ["Model!N224"])
    check(
        "scope/0233 Model!224 stops at R",
        full[0] == "Model!J224" and full[-1] == "Model!R224",
        f"{full[0]}..{full[-1]}",
    )
    full = disclose.full_copied_scope(gold, ["DCF!AE35"])
    check(
        "scope/0233 DCF!35 spans F:AE",
        full[0] == "DCF!F35" and full[-1] == "DCF!AE35",
        f"{full[0]}..{full[-1]}",
    )


def test_scope_observed_0533_band():
    gold = golden("0533.xlsx")
    if gold is None:
        return
    bands = disclose.group_bands(gold, ["Multiples!D19"], set())
    check(
        "scope/0533 Multiples!D19:E19",
        bands[0]["band"] == "Multiples!D19:E19",
        bands[0]["band"],
    )


# ------------------------------------------------- distribution mechanics


def test_distribution_plain_floored_residual_keeps_named_value():
    gold = book_from({"C": {
        "B10": "Cash available for dividends", "H10": "=H8-H9",
        "H8": 10, "H9": 4,
        "B12": "Dividends", "H12": "=MAX(H10,0)", "I12": "=MAX(I10,0)",
        "I10": "=I8-I9", "I8": 5, "I9": 9,
    }})
    records = disclose.detect_distribution_policy(gold, {"C!H12"}, set())
    check("mech/floored count", len(records) == 1, str(len(records)))
    check(
        "mech/floored named value",
        records[0]["value"] == "residual_cash_floored",
        records[0]["value"],
    )


def test_distribution_min_cap_renders_mechanics():
    """0352/0353 shape: =MIN(balance+accrual, cash*share) must keep the cap."""
    cells = {"W": {
        "B51": "Cash available for distribution", "B56": "Starting Balance",
        "B57": "Accrual", "B59": "Distribution", "E59": 0.8,
    }}
    for col_num in range(7, 12):  # G..K
        col = disclose.num_to_col(col_num)
        cells["W"][f"{col}51"] = "=MAX(%s50,0)" % col
        cells["W"][f"{col}50"] = 3
        cells["W"][f"{col}56"] = 1
        cells["W"][f"{col}57"] = 2
        cells["W"][f"{col}59"] = "=MIN(%s56+%s57,%s51*$E59)" % (col, col, col)
    gold = book_from(cells)
    records = disclose.detect_distribution_policy(gold, {"W!G59"}, set())
    check("mech/min-cap count", len(records) == 1, str(len(records)))
    rec = records[0]
    check("mech/min-cap value", rec["value"] == "formula_mechanics", rec["value"])
    check("mech/min-cap scope", rec["band"] == "`W!G59:K59`", rec["band"])
    sentence = disclose.render_sentence(rec)
    for needle in (
        "take the lesser of",
        "Starting Balance",
        "Accrual",
        "Cash available for distribution",
        "floors its own calculation at zero",
    ):
        check("mech/min-cap sentence has %r" % needle, needle in sentence, sentence)
    check(
        "mech/min-cap no negative claim",
        "negative" not in sentence,
        sentence,
    )


def test_distribution_pass_through_copy_states_floor():
    """0256 shape: =H28 copying a MAX(...,0) row is not an unfloored residual."""
    cells = {"CalcA": {
        "B28": "Cash available for dividends", "B29": "Dividends",
    }}
    for col_num in range(8, 18):  # H..Q
        col = disclose.num_to_col(col_num)
        cells["CalcA"][f"{col}27"] = 4
        cells["CalcA"][f"{col}28"] = f"=MAX({col}27-6,0)"
        cells["CalcA"][f"{col}29"] = f"={col}28"
    gold = book_from(cells)
    records = disclose.detect_distribution_policy(gold, {"CalcA!H29"}, set())
    check("mech/pass-through count", len(records) == 1, str(len(records)))
    rec = records[0]
    check(
        "mech/pass-through not unfloored",
        rec["value"] == "formula_mechanics",
        rec["value"],
    )
    check("mech/pass-through scope", rec["band"] == "`CalcA!H29:Q29`", rec["band"])
    sentence = disclose.render_sentence(rec)
    for needle in (
        "copy the same-column value from",
        "Cash available for dividends",
        "floors its own calculation at zero",
    ):
        check("mech/pass-through sentence has %r" % needle, needle in sentence, sentence)


def test_distribution_mechanics_declines_on_graded_band():
    gold = book_from({"C": {
        "B10": "Cash available for dividends", "H10": "=MAX(H8,0)", "H8": 5,
        "B12": "Dividends", "H12": "=MIN(H10,H8)",
    }})
    delivered = book_from({"C": {"B12": "Dividends"}})
    records = disclose.detect_distribution_policy(gold, {"C!H12"}, {"C!H12"})
    check("mech/graded-band record exists", len(records) == 1, str(len(records)))
    ctx = {"gold": gold, "delivered": delivered, "targets": {"C!H12"}}
    shipped = disclose.ship_distribution_policy(records[0], ctx)
    check("mech/graded-band declines", shipped is False, str(shipped))


def test_distribution_observed_goldens():
    expectations = {
        "0352.xlsx": (
            "'Equity Waterfall'!G59",
            "`'Equity Waterfall'!G59:CA59`",
            ("take the lesser of", "Starting Balance",
             "Cashflows Available for Distribution"),
        ),
        "0353.xlsx": (
            "'Monthly CFs'!G162",
            "`'Monthly CFs'!G162:DW162`",
            ("take the lesser of", "Beginning Balance",
             "floors its own calculation at zero"),
        ),
        "0646.xlsx": (
            "Calc!H203",
            "`Calc!H203:Q203`",
            ("take the greater of", "take the lesser of",
             "Cash available for dividends", "Available earnings for dividends"),
        ),
        "0350.xlsx": (
            "Cashflows!G143",
            "`Cashflows!G143:Q143`",
            ("LP Share", "floors its own calculation at zero"),
        ),
        "0256.xlsx": (
            "CalcA!H29",
            "`CalcA!H29:Q29`",
            ("copy the same-column value from",
             "Cash available for dividends",
             "floors its own calculation at zero"),
        ),
    }
    for book, (seed, band, needles) in sorted(expectations.items()):
        gold = golden(book)
        if gold is None:
            continue
        seed_key = disclose.parse_ref(seed, gold.sheets[0])
        records = disclose.detect_distribution_policy(gold, {seed_key}, set())
        check("mech/%s record count" % book, len(records) == 1, str(len(records)))
        rec = records[0]
        check("mech/%s value" % book, rec["value"] == "formula_mechanics", rec["value"])
        check("mech/%s scope" % book, rec["band"] == band, rec["band"])
        sentence = disclose.render_sentence(rec)
        for needle in needles:
            check("mech/%s sentence has %r" % (book, needle), needle in sentence, sentence)
        check("mech/%s no negative claim" % book, "negative" not in sentence, sentence)


def test_stake_scaling_requires_real_multiplication_by_share():
    """0666 shape: a direct read of an equity-investment amount is not scaling."""
    gold = book_from({
        "Assumptions": {"B162": "Initial Equity investment", "C162": 5000000},
        "Calc_M": {"B386": "Initial Equity investment",
                   "C386": "=Assumptions!$C$162"},
    })
    records = disclose.detect_stake_scaling(gold, {"Calc_M!C386"})
    check("stake/0666-shape emits nothing", records == [], str(records))
    gold_pos = book_from({"Deal": {
        "B7": "Equity Investment %", "C7": 0.6,
        "B9": "Entry consideration", "C8": 100, "C9": "=C8*C7",
    }})
    records = disclose.detect_stake_scaling(gold_pos, {"Deal!C9"})
    check("stake/real scaling still emits", len(records) == 1, str(records))
    check(
        "stake/real scaling value",
        records[0]["value"] == "applied",
        records[0]["value"],
    )


def test_render_sentence_rejects_quoted_empty_label():
    rec = {
        "entry": "stake_scaling",
        "value": "applied",
        "fields": {"label": '""', "ingredient": "the row labelled \"Share\""},
    }
    check(
        "render/empty-quoted-label",
        disclose.render_sentence(rec) == "",
        disclose.render_sentence(rec),
    )


# ----------------------------------------------------------- row labels


def test_row_label_rules():
    for text, expect in (
        ("US$", True), ("TWh", True), ("AED million", True), ("in '000", True),
        ("USD/Metric T", True), ("date", True), ("EURm", True),
        ("Facility Amount", False), ("% net revenue", False),
        ("First Availability Date", False),
    ):
        check(
            "label/is_unit_stamp(%r)=%s" % (text, expect),
            disclose.is_unit_stamp(text) is expect,
            str(disclose.is_unit_stamp(text)),
        )
    check("label/soft Year", disclose.is_soft_label("Year") is True, "")
    check("label/soft date not hard-lost", disclose.is_soft_label("dates") is True, "")


def test_row_label_soft_fallback_and_linked_labels():
    gold = book_from({"S": {
        # Year is skipped when a real name sits further left...
        "C6": "Uses & sources of funds", "I6": "Year", "J6": 2024,
        # ...but is the label when the row carries nothing else.
        "D9": "Year", "E9": "in '000", "J9": 2024,
    }})
    check(
        "label/skip Year for real name",
        gold.row_label("S!J6") == "Uses & sources of funds",
        gold.row_label("S!J6"),
    )
    check(
        "label/Year fallback",
        gold.row_label("S!J9") == "Year",
        gold.row_label("S!J9"),
    )


def test_row_label_observed_goldens():
    cases = {
        "0233.xlsx": {
            # The five pre-fix expectations from test_faithfulness_0233_0255.
            "Control!F58": "Facility Amount",
            "Control!F64": "Maturity",
            "DCF!AE31": "% net revenue",
            "DDM!H177": "Leverage (x)",
            "Model!AC86": "Personnel costs",
            # The reviewer's repairable-label corrections.
            "Control!I20": "ElecLink flows",
            "Model!H211": "First Availability Date",
            "Model!N292": "Year",
        },
        "0600.xlsx": {
            "InputT!L5": "Airfreight services price per ton",
            "InputT!L8": "Ocean freight and ocean services price per TEU",
        },
        "0518.xlsx": {
            "Inputs!F9": "Estimated market size - pre-cut fruits",
            "Inputs!F13": "Estimated market size - pre-cut vegetables",
            "Inputs!F17": "Estimated market size - fresh juices",
        },
        "0350.xlsx": {"Summary!J14": "Exit Year"},
        "0353.xlsx": {"Summary!J14": "Exit Month"},
        "0666.xlsx": {"Calc_M!C386": "Initial Equity investment"},
        "0646.xlsx": {"Calc!H203": "Dividends"},
    }
    for book, expectations in sorted(cases.items()):
        gold = golden(book)
        if gold is None:
            continue
        for cell, want in expectations.items():
            got = gold.row_label(cell)
            check("label/%s %s" % (book, cell), got == want, "%r != %r" % (got, want))


def test_row_populated_ignores_no_row_with_golden_formulas():
    """0605 shape: D14 holds a formula before the K14 heading; not unused."""
    cells = {"Dashboard": {
        "Z1": 1,
        "B13": "Revenue", "D13": 5,
        "D14": "=Z1", "K14": "Operating metrics",
        "B15": "Costs", "D15": 7,
        "B16": "Total", "D16": "=SUM(D13:D15)",
        # A genuinely empty labelled member row inside the same summed block.
        "B12": "Taxes paid",
    }}
    gold = book_from(cells)
    delivered = book_from({"Dashboard": {"B13": "Revenue", "K14": "Operating metrics"}})
    scope = {"Dashboard!D13", "Dashboard!D15", "Dashboard!D16"}
    records = disclose.detect_row_populated(
        gold, delivered, scope, ["Dashboard!D16"]
    )
    rows = [r.get("band") for r in records if r["value"] == "unused"]
    check(
        "row_populated/0605-shape row not unused",
        "`Dashboard!row 14`" not in rows,
        str(rows),
    )


def test_populated_but_unread_regression_0233():
    """Kept from test_faithfulness_0233_0255: blank removed rows stay silent."""
    gold = golden("0233.xlsx")
    if gold is None:
        return
    env = FAILED_BUNDLE_0233 / "environment"
    delivered_path = next(iter(sorted(env.glob("*.xlsx"))), None)
    if delivered_path is None:
        SKIPPED.append("0233 delivered")
        return
    delivered = disclose.Book(delivered_path)
    unread = disclose.detect_populated_but_unread(gold, delivered)
    forbidden = {"`DDM!F42:AE42`", "`DDM!F51:AE51`", "`Model!J13:BT13`"}
    hit = forbidden & {record.get("band") for record in unread}
    check("row_populated/0233 forbidden bands stay silent", not hit, str(hit))


# ---------------------------------------------- faithcheck & 08_27 fixes
#
# Coverage for the harbor-rerun blocker classes: wrong row labels (0598,
# 0599, 0605, 0622, 0660, 0661, 0669), singleton-vs-copied phrasing (0248,
# 0462, 0518, 0537), fragmented copied scope (0528), missing absolute-lock
# rendering (0658), omitted reference/negation (0523) and source link
# (0648), graded-output alias closure (0353, 0441, 0620, 0672), and the
# numeric-collision refusal policy (0469).


def book_path_from(rows: dict) -> Path:
    """Like book_from, but keeps the file on disk for path-based APIs."""
    wb = openpyxl.Workbook()
    default = wb.active
    first = True
    for sheet, cells in rows.items():
        ws = default if first else wb.create_sheet()
        ws.title = sheet
        first = False
        for ref, value in cells.items():
            ws[ref] = value
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as handle:
        path = Path(handle.name)
    wb.save(path)
    return path


def test_unit_stamp_new_classes():
    for text, expect in (
        ("000A$", True), ("'000 US$", True), ("US$m", True), ("US$mm", True),
        ("historical", True), ("Actuals", True), ("% of EBITDA", True),
        ("% of enterprise value", True), ("hours / week", True),
        ("hours", True), ("Empty", True),
        ("% net revenue", False), ("EBITDA", False),
        ("Historical Cost Basis", False), ("Equity invested", False),
    ):
        check(
            "stamp/is_unit_stamp(%r)=%s" % (text, expect),
            disclose.is_unit_stamp(text) is expect,
            str(disclose.is_unit_stamp(text)),
        )


def test_row_label_new_stamp_shapes():
    """0598's `000A$`, 0669's `historical`, 0660's `US$m` all keep walking."""
    gold = book_from({"Calc": {
        "B265": "Equity invested", "D265": "000A$", "J265": 1, "K265": "=J265*2",
        "B61": "Total Cost of Sales", "C61": "historical", "D61": "GBP", "I61": 5,
        "B46": "HoldCo debt repayment", "D46": "US$m", "J46": 7,
    }})
    for cell, want in (
        ("Calc!K265", "Equity invested"),
        ("Calc!I61", "Total Cost of Sales"),
        ("Calc!J46", "HoldCo debt repayment"),
    ):
        got = gold.row_label(cell)
        check("label/stamp-shape %s" % cell, got == want, "%r != %r" % (got, want))


def test_row_label_stamp_column_skips_basis_words():
    """0661 shape: a units/basis column stamps "EBITDA" beside the real name."""
    cells = {"Assumptions": {
        "C63": "Acquisition financing multiple", "F63": "EBITDA", "G63": 5,
        "C62": "Acquisition multiple", "F62": "EBITDA", "G62": 20,
        # Enough stamps to prove column F is a units/basis column.
        "F30": "hours", "F31": "#", "F32": "x", "F41": "k€",
        "F42": "% of EBITDA", "F55": "date", "F56": "date",
        "F64": "% of enterprise value", "F65": "years",
    }}
    gold = book_from(cells)
    check(
        "label/0661 basis column",
        gold.row_label("Assumptions!G63") == "Acquisition financing multiple",
        gold.row_label("Assumptions!G63"),
    )


def test_row_label_linked_vs_computed_cached_text():
    """A bare-link label is accepted; computed cached text is skipped."""
    gold = book_from({"S": {"A5": "Growth", "E5": 7, "B9": "Real Name", "E9": 3}})
    gold.formula["S!C5"] = "=A5"
    gold.value["S!C5"] = "Growth"
    gold.formula["S!C9"] = '=IF(1,"North","South")'
    gold.value["S!C9"] = "North"
    check("label/linked literal accepted", gold.row_label("S!E5") == "Growth",
          gold.row_label("S!E5"))
    check("label/computed text skipped", gold.row_label("S!E9") == "Real Name",
          gold.row_label("S!E9"))


def _method_rec(cells: list[str]) -> dict:
    return {
        "entry": "method_tax", "value": "out_of_catalogue",
        "cells": cells, "cell_keys": cells,
        "fields": {"label": '"Tax"', "band": "`X`",
                   "representative": "S!C5", "steps": "multiply (a) by (b)"},
    }


def test_render_scope_phrasing_matches_geometry():
    """0248/0462/0518/0537: a singleton must not read as copied-column."""
    single = disclose.render_sentence(_method_rec(["S!C5"]))
    check("phrase/singleton rewritten", "single-cell calculation" in single, single)
    check("phrase/singleton no copied-column", "copied-column" not in single, single)
    check("phrase/singleton drops shown-for", ", shown for" not in single, single)
    down = disclose.render_sentence(_method_rec(["S!C5", "S!C6"]))
    check("phrase/vertical copied-down", "calculation copied down the rows" in down, down)
    across = disclose.render_sentence(_method_rec(["S!C5", "S!D5"]))
    check("phrase/horizontal keeps copied-column", "copied-column calculation" in across, across)


def test_vertical_fragment_bands_merge():
    """0528 shape: a copied-down run selected cell by cell is one band."""
    gold = book_from({"Inputs": {
        "T144": "Alpha", "U137": 1, "U138": 2, "U139": 3,
        "U144": "=U137*5", "U145": "=U138*5", "U146": "=U139*5",
    }})
    bands = disclose.group_bands(
        gold, ["Inputs!U144", "Inputs!U145", "Inputs!U146"], set()
    )
    check("vscope/merged band count", len(bands) == 1, str(len(bands)))
    check("vscope/merged ref", bands[0]["band"] == "Inputs!U144:U146", bands[0]["band"])
    lone = disclose.group_bands(gold, ["Inputs!U145"], set())
    check("vscope/lone stated span", lone[0]["band"] == "Inputs!U144:U146", lone[0]["band"])
    check(
        "vscope/lone claims selected only",
        lone[0]["cell_keys"] == ["Inputs!U145"],
        str(lone[0]["cell_keys"]),
    )


def test_copy_lock_note():
    """0658 shape: a mixed pinned/advancing pattern names the pinned refs."""
    gold = book_from({"LBO": {
        "C9": 5, "K4": 1, "L4": 2, "K45": 1, "L45": 2,
        "K144": "=K4*$C$9*K45", "L144": "=L4*$C$9*L45",
    }})
    note = disclose.copy_lock_note(gold, ["LBO!K144", "LBO!L144"])
    check("lock/pinned named", "LBO!C9" in note and "fixed" in note, note)
    dup = book_from({"L": {
        "N83": 2, "O83": 3, "N84": "=$N$83*N83", "O84": "=$N$83*O83",
    }})
    note = disclose.copy_lock_note(dup, ["L!N84", "L!O84"])
    check("lock/duplicate operands disambiguated",
          "occurrence 1 of L!N83" in note, note)
    flat = book_from({"S": {"C4": 1, "D4": 2, "C5": "=C4*2", "D5": "=D4*2"}})
    check(
        "lock/all-advancing needs no note",
        disclose.copy_lock_note(flat, ["S!C5", "S!D5"]) == "",
        disclose.copy_lock_note(flat, ["S!C5", "S!D5"]),
    )


def test_stake_scaling_negation_declines():
    """0523 shape: `=-Multiples!D26*B4` cannot ship a bare share claim."""
    negated = book_from({"NPV": {
        "A11": "Intial Investment", "B11": "=-M26*B4",
        "A4": "Equity Investment %", "B4": 0.6, "M26": 100,
    }})
    check(
        "stake/negation declines",
        disclose.detect_stake_scaling(negated, {"NPV!B11"}) == [],
        "record emitted for negated formula",
    )
    clean = book_from({"NPV": {
        "A11": "Intial Investment", "B11": "=M26*B4",
        "A4": "Equity Investment %", "B4": 0.6, "M26": 100,
    }})
    check(
        "stake/plain product still emits",
        len(disclose.detect_stake_scaling(clean, {"NPV!B11"})) == 1,
        "no record for plain product",
    )


def test_source_selection_bare_cross_sheet_ships():
    """0648 shape: `'IRR'!D6 = Workings!D87` must ship with the exact cell."""
    gold = book_from({
        "IRR": {"C6": "Exit Value", "D6": "=Workings!D87"},
        "Workings": {"C87": "Equity Value", "D87": "=1+1"},
    })
    records = disclose.detect_source_selection(gold, {"IRR!D6"})
    check("srclink/detected", len(records) == 1, str(len(records)))
    rec = records[0]
    check("srclink/ingredient names cell",
          "Workings!D87" in rec["fields"]["ingredient"]
          and "Equity Value" in rec["fields"]["ingredient"],
          rec["fields"]["ingredient"])
    empty = book_from({"IRR": {"C6": "Exit Value"}})
    ctx = {"gold": gold, "delivered": empty, "targets": set()}
    check("srclink/ships when source hidden",
          disclose.ship_source_selection(dict(rec), ctx) is True, "declined")
    survived = book_from({"Workings": {"D87": 5}})
    ctx = {"gold": gold, "delivered": survived, "targets": set()}
    check("srclink/declines when source survives",
          disclose.ship_source_selection(dict(rec), ctx) is False, "shipped")
    ctx = {"gold": gold, "delivered": empty, "targets": {"Workings!D87"}}
    check("srclink/declines when source graded",
          disclose.ship_source_selection(dict(rec), ctx) is False, "shipped")
    flipped = book_from({
        "IRR": {"C6": "Exit Value", "D6": "=-Workings!D87"},
        "Workings": {"C87": "Equity Value", "D87": "=1+1"},
    })
    records = disclose.detect_source_selection(flipped, {"IRR!D6"})
    shipped = [
        r for r in records
        if disclose.ship_source_selection(r, {"gold": flipped, "delivered": empty, "targets": set()})
    ]
    check("srclink/sign-flip stays reviewer-only", not shipped, str(shipped))


def test_graded_closure_aliases_and_copies():
    gold = book_from({"M": {
        "O1": "=+O24", "O20": 2, "O22": 3, "O24": "=(O22/O20)^(1/5)-1",
        "C56": 1, "C57": 2, "C58": "=+C56/C57",
        "G56": 3, "G57": 4, "G58": "=+G56/G57",
        "C36": 1, "E36": 2, "C37": "=C36*2", "E37": "=E36*2",
    }})
    closure = disclose.graded_closure(gold, {"M!O1"})
    check("closure/0441 alias", "M!O24" in closure, str(sorted(closure)))
    closure = disclose.graded_closure(gold, {"M!C58"})
    check("closure/0672 same-row copy", "M!G58" in closure, str(sorted(closure)))
    closure = disclose.graded_closure(gold, {"M!E37"})
    check("closure/0620 same-row copy", "M!C37" in closure, str(sorted(closure)))


def test_closure_suppression_spares_standing_exception():
    records = [
        {"entry": "method_returns", "disposition": "disclosed",
         "cell_keys": ["M!O24"], "cells": ["M!O24"]},
        {"entry": "liquidation_preference", "disposition": "disclosed",
         "cell_keys": ["M!O1"], "cells": ["M!O1"]},
    ]
    disclose.suppress_closure_reach(records, {"M!O1", "M!O24"})
    check("closure/alias record suppressed",
          records[0]["disposition"] == "suppressed", records[0]["disposition"])
    check("closure/liquidation_preference spared",
          records[1]["disposition"] == "disclosed", records[1]["disposition"])


def _numeric_task(targets: dict) -> Path:
    task = Path(tempfile.mkdtemp(prefix="disclosure-task-"))
    (task / "tests").mkdir()
    (task / "tests" / "answer_key.json").write_text(
        json.dumps({"targets": targets}), encoding="utf-8"
    )
    return task


def test_numeric_collision_policy():
    """0469: 6 vs 6.0 passes; 4+ significant digits refuse; closure refuses."""
    task = _numeric_task({"S!C13": 6.0, "S!D13": 1234.5})
    head = "## Workbook disclosure\n- `S!A1`: "
    check(
        "numeric/small round integer passes",
        disclose.audit_text(head + "divide (x) by (6).\n", task) == [],
        str(disclose.audit_text(head + "divide (x) by (6).\n", task)),
    )
    faults = disclose.audit_text(head + "multiply (x) by (1234.5).\n", task)
    check("numeric/high-specificity refused", len(faults) == 1, str(faults))
    provenance = [{
        "cell_keys": ["S!C12"], "cells": ["S!C12"],
        "method_profile": {"references": [], "numbers": [6.0]},
    }]
    faults = disclose.audit_text(
        head + "divide (x) by (6).\n", task,
        records=provenance, closure={"S!C12"},
    )
    check("numeric/closure provenance refused", len(faults) == 1
          and "closure" in faults[0], str(faults))
    faults = disclose.audit_text(
        head + "divide (x) by (6).\n", task,
        records=provenance, closure={"S!Z99"},
    )
    check("numeric/non-closure provenance passes", faults == [], str(faults))


def test_faithcheck_synthetic():
    golden_path = book_path_from({
        "S": {
            "B5": "Revenue", "C4": 10, "D4": 10, "C5": "=C4*2", "D5": "=D4*2",
            "B9": "Costs", "C9": "=C4*3",
            "B12": "IRR", "C12": "=C4^2",
            "B13": "Answer", "C13": "=+C12",
            "T9": "Beta", "U7": 1, "U8": 2, "U9": "=U7*5", "U10": "=U8*5",
            "B20": "Exit Value", "C20": "=W!D8",
        },
        "W": {"C8": "Equity Value", "D8": "=1+1"},
    })
    gold = disclose.Book(golden_path)
    task = Path(tempfile.mkdtemp(prefix="disclosure-task-"))
    (task / "environment").mkdir()
    delivered_path = book_path_from({"S": {
        "B5": "Revenue", "B9": "Costs", "B12": "IRR", "B13": "Answer",
        "T9": "Beta", "B20": "Exit Value",
    }})
    (task / "environment" / "delivered.xlsx").write_bytes(delivered_path.read_bytes())
    (task / "tests").mkdir()
    (task / "tests" / "answer_key.json").write_text(
        json.dumps({"targets": {"S!C13": 100.0}}), encoding="utf-8"
    )

    def method_record(cells, fields, entry="method_revenue"):
        return {
            "band": disclose.compact_cells([disclose.pretty(c) for c in cells]).strip("`"),
            "cells": [disclose.pretty(c) for c in cells], "cell_keys": cells,
            "entry": entry, "family": entry, "value": "out_of_catalogue",
            "disposition": "disclosed", "fields": fields,
        }

    def derived_fields(cells, label):
        profile = disclose.formula_profile(gold, {
            "band": None, "cell_keys": cells, "label": label, "pattern": "",
        })
        fields = disclose.custom_sentence_fields(profile, gold)
        fields.pop("_coverage_complete", None)
        return fields

    records = [
        # Faithful: correct label, scope, coverage; must produce no fault.
        method_record(["S!C5", "S!D5"], derived_fields(["S!C5", "S!D5"], "Revenue")),
        # Wrong label and steps missing the reference and literal.
        method_record(
            ["S!C9"],
            {"label": '"Wrong"', "band": "`S!C9`",
             "representative": "S!C9", "steps": "do something"},
            entry="method_operating_expense",
        ),
        # Reaches the graded closure through the alias S!C13 = +C12.
        method_record(["S!C12"], derived_fields(["S!C12"], "IRR"),
                      entry="method_returns"),
        # Fragment of the vertical copy family U9:U10.
        method_record(
            ["S!U9"],
            {"label": '"Beta"', "band": "`S!U9`",
             "representative": "S!U9", "steps": "do something"},
        ),
    ]
    (task / "tests" / "disclosure.json").write_text(
        json.dumps({"records": records, "agent_records": []}, default=str),
        encoding="utf-8",
    )
    runs_root = Path(tempfile.mkdtemp(prefix="disclosure-runs-"))
    disclose.write_json(
        runs_root / task.name / "bands.json",
        {"bands": [{"cell_keys": ["S!C20"]}]},
    )
    result = disclose.faithcheck_task(task, runs_root, golden=str(golden_path))
    kinds = {fault["kind"] for fault in result["faults"]}
    check("faithcheck/records checked", result["checked_records"] == 4,
          str(result["checked_records"]))
    check("faithcheck/fails closed", result["passed"] is False, str(result))
    for kind in ("row_label", "reference", "closure", "scope", "source_link"):
        check("faithcheck/kind %s present" % kind, kind in kinds, str(kinds))
    good_band = records[0]["band"]
    check(
        "faithcheck/faithful record clean",
        not [f for f in result["faults"] if f["record"] == good_band],
        str([f for f in result["faults"] if f["record"] == good_band]),
    )
    wrong = [f for f in result["faults"]
             if f["kind"] == "row_label" and f["found"] == "Wrong"]
    check("faithcheck/label expected vs found", bool(wrong)
          and wrong[0]["expected"] == "Costs", str(wrong))


def test_faithcheck_observed_0441_0537():
    gold_0441 = golden("0441.xlsx")
    if gold_0441 is not None:
        closure = disclose.graded_closure(gold_0441, {"Model!O1"})
        check("closure/0441 golden alias", "Model!O24" in closure, str(len(closure)))
    gold_0537 = golden("0537.xlsx")
    if gold_0537 is not None:
        for cell in ("UFCF!L29", "UFCF!L33", "UFCF!L34"):
            span_h = disclose.full_copied_scope(gold_0537, [cell])
            span_v = disclose.full_copied_scope_vertical(gold_0537, [cell])
            check(
                "scope/0537 %s is a true singleton" % cell,
                span_h == [cell] and span_v == [cell],
                "%s / %s" % (span_h, span_v),
            )
        sentence = disclose.render_sentence(_method_rec(["UFCF!L29"]))
        check(
            "phrase/0537 singleton phrasing",
            "single-cell calculation" in sentence and "copied-column" not in sentence,
            sentence,
        )


def main() -> int:
    tests = [v for k, v in sorted(globals().items()) if k.startswith("test_")]
    for test in tests:
        test()
    print(f"{len(PASSED)} checks passed across {len(tests)} tests")
    if SKIPPED:
        print("skipped (missing data):", ", ".join(sorted(set(SKIPPED))))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
