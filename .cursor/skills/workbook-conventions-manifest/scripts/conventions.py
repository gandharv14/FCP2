#!/usr/bin/env python3
"""Build a conventions manifest for a GDPval rebuild task.

Reads the original ("golden") workbook, works backwards from the graded cells,
detects the modelling decisions the emptying pass destroyed, and emits them as
structured records.

Execute this; do not read it. Subcommands:

  closure     which deleted cells each graded answer depends on
  detect      the modelling decisions found in that closure
  emit        write conventions.json and the instruction section
  map         which cells the original actually used
  audit       completeness and answer-leakage checks
  defects     originals that are broken in ways no disclosure can fix

Requires openpyxl.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
from collections import defaultdict, deque

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl required: pip install openpyxl")


# --------------------------------------------------------------------- refs

# A cell or range reference, with an optional sheet qualifier.
REF_RE = re.compile(
    r"""(?:(?P<sheet>'(?:[^']|'')+'|[A-Za-z_\\][A-Za-z0-9_.]*)!)?
        (?P<a>\$?[A-Z]{1,3}\$?[0-9]{1,7})
        (?::(?P<b>\$?[A-Z]{1,3}\$?[0-9]{1,7}))?
        (?![A-Za-z0-9_(])""",
    re.X,
)
STRING_RE = re.compile(r'"(?:[^"]|"")*"')
COORD_RE = re.compile(r"^\$?([A-Z]{1,3})\$?([0-9]{1,7})$")

# Tokens that look like references but are function or error names.
NOT_A_REF = {"LOG10", "T", "N", "TRUE", "FALSE"}


def col_to_num(col: str) -> int:
    n = 0
    for ch in col:
        n = n * 26 + (ord(ch) - 64)
    return n


def num_to_col(n: int) -> str:
    s = ""
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def split_coord(coord: str):
    m = COORD_RE.match(coord)
    if not m:
        return None
    return m.group(1), int(m.group(2))


def unquote_sheet(name: str) -> str:
    if name.startswith("'") and name.endswith("'"):
        return name[1:-1].replace("''", "'")
    return name


def quote_sheet(name: str) -> str:
    return "'%s'" % name if re.search(r"[^A-Za-z0-9_.]", name) else name


def key(sheet: str, coord: str) -> str:
    return "%s!%s" % (sheet, coord.replace("$", ""))


def pretty(k: str) -> str:
    sheet, coord = k.split("!", 1)
    return "%s!%s" % (quote_sheet(sheet), coord)


def parse_ref(ref: str, default_sheet: str) -> str:
    """Normalise a user-supplied reference such as "'NPV & IRR'!B20"."""
    ref = ref.strip()
    if "!" in ref:
        sheet, coord = ref.rsplit("!", 1)
        return key(unquote_sheet(sheet), coord)
    return key(default_sheet, ref)


def expand(sheet: str, a: str, b: str | None, cap: int = 20000):
    """Expand a reference into individual cell keys."""
    pa = split_coord(a)
    if not pa:
        return []
    if not b:
        return [key(sheet, "%s%d" % pa)]
    pb = split_coord(b)
    if not pb:
        return [key(sheet, "%s%d" % pa)]
    c1, c2 = sorted([col_to_num(pa[0]), col_to_num(pb[0])])
    r1, r2 = sorted([pa[1], pb[1]])
    if (c2 - c1 + 1) * (r2 - r1 + 1) > cap:
        return []
    out = []
    for c in range(c1, c2 + 1):
        for r in range(r1, r2 + 1):
            out.append(key(sheet, "%s%d" % (num_to_col(c), r)))
    return out


def refs_in(formula: str, default_sheet: str) -> list[str]:
    """Every cell key a formula reads."""
    if not isinstance(formula, str) or not formula.startswith("="):
        return []
    body = STRING_RE.sub('""', formula)
    out = []
    for m in REF_RE.finditer(body):
        a = m.group("a")
        if a.upper() in NOT_A_REF:
            continue
        sheet = unquote_sheet(m.group("sheet")) if m.group("sheet") else default_sheet
        out.extend(expand(sheet, a, m.group("b")))
    return out


# ------------------------------------------------------------------ loading

class Book:
    """A workbook flattened to {sheet!coord: (formula, value)}."""

    def __init__(self, path: str):
        self.path = path
        self.formula = {}
        self.value = {}
        self.sheets = []
        wbf = openpyxl.load_workbook(path, data_only=False)
        wbv = openpyxl.load_workbook(path, data_only=True)
        for ws in wbf.worksheets:
            self.sheets.append(ws.title)
            vs = wbv[ws.title]
            for row in ws.iter_rows():
                for c in row:
                    if c.value is None:
                        continue
                    k = key(ws.title, c.coordinate)
                    if isinstance(c.value, str) and c.value.startswith("="):
                        self.formula[k] = c.value
                        self.value[k] = vs[c.coordinate].value
                    else:
                        self.value[k] = c.value
        self._labels = {}

    def has(self, k: str) -> bool:
        return k in self.value or k in self.formula

    def row_label(self, k: str) -> str:
        """The leftmost text on the cell's row, which is how these models label."""
        if k in self._labels:
            return self._labels[k]
        sheet, coord = k.split("!", 1)
        p = split_coord(coord)
        label = ""
        if p:
            for c in range(1, 7):
                v = self.value.get(key(sheet, "%s%d" % (num_to_col(c), p[1])))
                if isinstance(v, str) and v.strip() and not v.startswith("="):
                    label = v.strip()
                    break
        self._labels[k] = label
        return label

    def row_cells(self, sheet: str, rownum: int, lo: int = 1, hi: int = 60):
        return [key(sheet, "%s%d" % (num_to_col(c), rownum)) for c in range(lo, hi + 1)]


def task_paths(task_dir: str, golden: str | None):
    task_dir = os.path.abspath(task_dir)
    tid = os.path.basename(task_dir).split("-")[0]
    env = os.path.join(task_dir, "environment")
    inputs = None
    if os.path.isdir(env):
        for f in sorted(os.listdir(env)):
            if f.endswith(".xlsx"):
                inputs = os.path.join(env, f)
                break
    if golden is None:
        root = os.path.abspath(os.path.join(task_dir, "..", ".."))
        for base in ("FCP Workbooks",):
            d = os.path.join(root, base)
            if not os.path.isdir(d):
                continue
            for sub in sorted(os.listdir(d)):
                cand = os.path.join(d, sub, "%s.xlsx" % tid)
                if os.path.exists(cand):
                    golden = cand
                    break
            if golden:
                break
    return tid, inputs, golden


def load_targets(task_dir: str) -> dict:
    p = os.path.join(task_dir, "tests", "answer_key.json")
    key_json = json.load(open(p))
    return key_json.get("targets", {}), key_json.get("tolerance", {})


# ----------------------------------------------------------------- closure

def closure(gold: Book, targets: list[str], limit: int = 200000):
    """Every cell a graded answer transitively reads."""
    seen = set()
    order = []
    q = deque(targets)
    while q and len(seen) < limit:
        k = q.popleft()
        if k in seen:
            continue
        seen.add(k)
        order.append(k)
        f = gold.formula.get(k)
        if f:
            sheet = k.split("!", 1)[0]
            for r in refs_in(f, sheet):
                if r not in seen:
                    q.append(r)
    return order


def blanked(gold: Book, inp: Book, cells) -> list[str]:
    """Cells the original computes and the delivered file does not carry."""
    out = []
    for k in cells:
        if k in gold.formula and not inp.has(k):
            out.append(k)
    return out


# --------------------------------------------------------------- detectors
#
# Each detector returns records:
#   {family, value, cells, evidence, alternatives, note}
# A detector must only assert what the golden formula or structure shows.

MONEY_FN = re.compile(r"\b(NPV|IRR|XIRR|XNPV|MIRR)\s*\(", re.I)


def _fmt(cells, n=4):
    cells = list(cells)
    shown = ", ".join(pretty(c) for c in cells[:n])
    return shown + (" +%d more" % (len(cells) - n) if len(cells) > n else "")


def detect_discount_period(gold: Book, scope: set) -> list[dict]:
    """Mid-year against year-end, read off the discount-period row."""
    out = []
    seen_rows = set()
    for k in sorted(scope):
        lab = gold.row_label(k).lower()
        if "discount period" not in lab and "discount factor period" not in lab:
            continue
        sheet, coord = k.split("!", 1)
        p = split_coord(coord)
        if not p or (sheet, p[1]) in seen_rows:
            continue
        seen_rows.add((sheet, p[1]))
        row = [c for c in gold.row_cells(sheet, p[1]) if c in gold.formula or c in gold.value]
        vals = [gold.value.get(c) for c in row]
        nums = [v for v in vals if isinstance(v, (int, float))]
        fs = [gold.formula[c] for c in row if c in gold.formula]
        value = None
        if any(isinstance(v, float) and abs(v - round(v)) > 0.4 for v in nums):
            value = "mid_year"
        elif nums and all(isinstance(v, (int, float)) and abs(v - round(v)) < 1e-9 for v in nums):
            value = "year_end"
        if value is None and any("/2" in f for f in fs):
            value = "mid_year"
        if value:
            out.append({
                "family": "discount_period",
                "value": value,
                "cells": [pretty(c) for c in row if c in gold.formula],
                "evidence": fs[0] if fs else "cached values %s" % nums[:4],
                "alternatives": ["mid_year", "year_end", "day_weighted"],
                "note": "Row labelled %r." % gold.row_label(k),
            })
    return out


HEADER_RE = re.compile(r"(statement|schedule|summary|assumptions|inputs)\s*$|:\s*$", re.I)


def detect_row_populated(gold: Book, inp: Book, scope: set, targets: list[str]) -> list[dict]:
    """Labelled rows the original leaves empty, which look emptied in the delivered file.

    Restricted to sheets carrying a graded cell. Elsewhere an empty labelled row is
    almost always a section header or a sub-heading in an input list, and reporting
    those buries the ones that matter.
    """
    out = []
    sheets = {t.split("!", 1)[0] for t in targets}
    scope_rows = defaultdict(set)
    for k in sorted(scope):
        s, coord = k.split("!", 1)
        p = split_coord(coord)
        if p:
            scope_rows[s].add(p[1])
    for sheet in sorted(sheets):
        rows = defaultdict(list)
        for k in list(gold.value) + list(gold.formula):
            s, coord = k.split("!", 1)
            if s != sheet:
                continue
            p = split_coord(coord)
            if p:
                rows[p[1]].append((col_to_num(p[0]), k))
        for rownum, entries in sorted(rows.items()):
            entries.sort()
            label, label_col, data = "", None, []
            for cnum, k in entries:
                v = gold.value.get(k)
                is_text = isinstance(v, str) and k not in gold.formula
                if is_text and not label:
                    label, label_col = v.strip(), cnum
                elif label_col is not None and cnum > label_col:
                    # Anything computed or numeric counts as data, whatever column it sits in.
                    if k in gold.formula or isinstance(v, (int, float)):
                        data.append(k)
            if data or not label or len(label) < 4 or HEADER_RE.search(label):
                continue
            if label.isupper():
                continue
            # The row must sit inside the band that feeds a graded answer.
            above = [r for r in (rownum - 1, rownum - 2) if r in scope_rows[sheet]]
            below = [r for r in (rownum + 1, rownum + 2) if r in scope_rows[sheet]]
            if not (above and below):
                continue
            out.append({
                "family": "row_populated",
                "value": "unused",
                "cells": ["%s!row %d" % (quote_sheet(sheet), rownum)],
                "evidence": "no formula and no value anywhere on the row in the original",
                "alternatives": ["unused", "populated"],
                "note": "Row labelled %r is empty in the original but sits inside the block "
                        "that feeds a graded answer." % label,
            })
    return out


def detect_inert_row(gold: Book, scope: set) -> list[dict]:
    """Rows that are populated but evaluate to zero in every period."""
    out = []
    seen_rows = set()
    for k in sorted(scope):
        lab = gold.row_label(k).lower()
        if not any(t in lab for t in ("minimum cash", "less: minimum")):
            continue
        sheet, coord = k.split("!", 1)
        p = split_coord(coord)
        if not p or (sheet, p[1]) in seen_rows:
            continue
        seen_rows.add((sheet, p[1]))
        row = [c for c in gold.row_cells(sheet, p[1]) if c in gold.formula]
        if not row:
            continue
        vals = [gold.value.get(c) for c in row]
        nums = [v for v in vals if isinstance(v, (int, float))]
        if nums and all(abs(v) < 1e-9 for v in nums):
            out.append({
                "family": "inert_line",
                "value": "always_zero",
                "cells": [pretty(c) for c in row],
                "evidence": gold.formula[row[0]],
                "alternatives": ["always_zero", "charged_once", "charged_every_period"],
                "note": "Row labelled %r is populated but evaluates to zero in all %d periods."
                        % (gold.row_label(k), len(row)),
            })
    return out


def detect_terminal_value(gold: Book, scope: set) -> list[dict]:
    out = []
    seen_rows = set()
    for k in sorted(scope):
        lab = gold.row_label(k).lower()
        if "terminal value" not in lab and "exit value" not in lab:
            continue
        sheet, coord = k.split("!", 1)
        p = split_coord(coord)
        if not p or (sheet, p[1]) in seen_rows:
            continue
        seen_rows.add((sheet, p[1]))
        row = [c for c in gold.row_cells(sheet, p[1]) if c in gold.formula]
        label = gold.row_label(k)
        if not row:
            out.append({
                "family": "terminal_value",
                "value": "absent",
                "cells": ["%s!row %d" % (quote_sheet(sheet), p[1])],
                "evidence": "no formula anywhere on the row in the original",
                "alternatives": ["absent", "perpetuity_growth", "exit_multiple"],
                "note": "Row labelled %r is never populated." % label,
            })
            continue
        f = gold.formula[row[0]]
        if re.search(r"\(1\s*\+\s*[^)]+\)\s*/\s*\(", f):
            val = "perpetuity_growth"
        elif re.search(r"[Mm]ultiple|Comps", f):
            val = "exit_multiple"
        else:
            val = "other"
        out.append({
            "family": "terminal_value",
            "value": val,
            "cells": [pretty(c) for c in row],
            "evidence": f,
            "alternatives": ["absent", "perpetuity_growth", "exit_multiple"],
            "note": "Row labelled %r, populated in %d column(s)." % (label, len(row)),
        })
    return out


def detect_npv_timing(gold: Book, scope: set) -> list[dict]:
    out = []
    for k in sorted(scope):
        f = gold.formula.get(k)
        if not f or "NPV(" not in f.upper():
            continue
        tail = f.upper().split("NPV(", 1)[1]
        extra = bool(re.search(r"\)\s*[+\-]\s*[A-Z$]", tail))
        out.append({
            "family": "npv_timing",
            "value": "t0_added_separately" if extra else "excel_default_one_period_out",
            "cells": [pretty(k)],
            "evidence": f,
            "alternatives": ["excel_default_one_period_out", "t0_added_separately"],
            "note": "Excel NPV discounts its first argument by one full period.",
        })
    return out


def detect_aggregate_scope(gold: Book, targets: list[str]) -> list[dict]:
    """The exact span of a graded total."""
    out = []
    for k in targets:
        f = gold.formula.get(k)
        if not f:
            continue
        m = re.match(r"^=\s*SUM\((.+)\)\s*$", f.strip(), re.I)
        if not m:
            continue
        out.append({
            "family": "aggregate_scope",
            "value": m.group(1).strip(),
            "cells": [pretty(k)],
            "evidence": f,
            "alternatives": ["open-ended"],
            "note": "Graded total spans this range; row label %r." % gold.row_label(k),
        })
    return out


def detect_projection_rule(gold: Book, blanked_cells: list[str]) -> list[dict]:
    """How each forecast row is carried forward."""
    out = []
    rows = defaultdict(list)
    for k in blanked_cells:
        sheet, coord = k.split("!", 1)
        p = split_coord(coord)
        if p:
            rows[(sheet, p[1])].append((col_to_num(p[0]), k))
    for (sheet, rownum), entries in sorted(rows.items()):
        if len(entries) < 3:
            continue
        entries.sort()
        kinds = []
        for cnum, k in entries:
            f = gold.formula.get(k, "")
            body = f[1:] if f.startswith("=") else f
            prev = "%s%d" % (num_to_col(cnum - 1), rownum)
            if re.fullmatch(r"\$?%s" % re.escape(prev), body.strip(), re.I):
                kinds.append("hold_level")
            elif re.search(r"\*\s*\(\s*1\s*[+\-]", body):
                kinds.append("hold_growth")
            elif re.match(r"^AVERAGE\(", body.strip(), re.I):
                kinds.append("average_window")
            elif re.search(r"[A-Z]{1,3}\$?\d+\s*\*\s*", body):
                kinds.append("ratio_to_driver")
            else:
                kinds.append("other")
        top = sorted(set(kinds), key=lambda x: (-kinds.count(x), x))[0]
        if top == "other" or kinds.count(top) < max(2, len(kinds) // 2):
            continue
        label = gold.row_label(entries[0][1])
        if not label:
            continue
        out.append({
            "family": "projection_rule",
            "value": top,
            "cells": [pretty(k) for _, k in entries],
            "evidence": gold.formula.get(entries[min(1, len(entries) - 1)][1], ""),
            "alternatives": ["hold_level", "hold_growth", "average_window", "ratio_to_driver"],
            "note": "Forecast row labelled %r." % label,
        })
    return out


def detect_stake_scaling(gold: Book, scope: set) -> list[dict]:
    """Deal lines multiplied by an ownership-percentage cell."""
    out = []
    pct_cells = set()
    for k, v in gold.value.items():
        lab = gold.row_label(k)
        if lab and re.search(r"equity investment|ownership|stake|% acquired", lab, re.I):
            sheet, coord = k.split("!", 1)
            p = split_coord(coord)
            if p and isinstance(gold.value.get(k), (int, float)):
                pct_cells.add(k)
    if not pct_cells:
        return out
    # Group by row: the decision is per line item, not per period.
    byrow = defaultdict(list)
    for k in sorted(scope):
        f = gold.formula.get(k)
        if not f:
            continue
        hits = [r for r in refs_in(f, k.split("!", 1)[0]) if r in pct_cells]
        if not hits:
            continue
        sheet, coord = k.split("!", 1)
        p = split_coord(coord)
        if p:
            byrow[(sheet, p[1])].append((col_to_num(p[0]), k, f, hits))
    for (sheet, rownum), entries in sorted(byrow.items()):
        entries.sort()
        _, first, f, hits = entries[0]
        out.append({
            "family": "stake_scaling",
            "value": "applied",
            "cells": [pretty(k) for _, k, _, _ in entries],
            "evidence": f,
            "alternatives": ["applied", "not_applied"],
            "note": "Row labelled %r multiplies by %s, the ownership share."
                    % (gold.row_label(first), pretty(hits[0])),
        })
    return out


def detect_source_selection(gold: Book, scope: set) -> list[dict]:
    """Cells whose meaning is 'which of several valuations feeds this'."""
    out = []
    for k in sorted(scope):
        lab = gold.row_label(k).lower()
        if not any(t in lab for t in ("initial investment", "purchase price", "entry",
                                      "consideration", "cv")):
            continue
        f = gold.formula.get(k)
        if not f:
            continue
        sheet = k.split("!", 1)[0]
        ext = [r for r in refs_in(f, sheet) if r.split("!", 1)[0] != sheet]
        if not ext:
            continue
        out.append({
            "family": "source_selection",
            "value": _fmt(sorted(set(ext)), 3),
            "cells": [pretty(k)],
            "evidence": f,
            "alternatives": ["open-ended"],
            "note": "Row labelled %r reads from another sheet." % gold.row_label(k),
        })
    return out


DETECTORS_SCOPE = [
    detect_discount_period,
    detect_inert_row,
    detect_terminal_value,
    detect_npv_timing,
    detect_stake_scaling,
    detect_source_selection,
]


def detect_all(gold: Book, inp: Book, scope_cells, blanked_cells, targets):
    scope = set(scope_cells)
    records = []
    for fn in DETECTORS_SCOPE:
        try:
            records.extend(fn(gold, scope))
        except Exception as exc:  # a detector must never sink the run
            print("  ! %s failed: %s" % (fn.__name__, exc), file=sys.stderr)
    for fn, args in ((detect_row_populated, (gold, inp, scope, targets)),
                     (detect_aggregate_scope, (gold, targets)),
                     (detect_projection_rule, (gold, blanked_cells))):
        try:
            records.extend(fn(*args))
        except Exception as exc:
            print("  ! %s failed: %s" % (fn.__name__, exc), file=sys.stderr)
    # De-duplicate on family plus cell list.
    seen = set()
    unique = []
    for r in records:
        sig = (r["family"], tuple(r["cells"]))
        if sig not in seen:
            seen.add(sig)
            unique.append(r)
    # A row already described by a specific family does not also need the generic
    # "this row is unused" record.
    covered = set()
    for r in unique:
        if r["family"] == "row_populated":
            continue
        for c in r["cells"]:
            if "!row " in c:
                covered.add(c)
            else:
                sheet, coord = c.rsplit("!", 1)
                p = split_coord(coord)
                if p:
                    covered.add("%s!row %d" % (sheet, p[1]))
    kept = [r for r in unique
            if r["family"] != "row_populated" or r["cells"][0] not in covered]
    # Stable order so the emitted manifest is reproducible.
    kept.sort(key=lambda r: (r["family"], r["cells"][0] if r["cells"] else ""))
    return kept


# ------------------------------------------------------------------ defects

def detect_defects(gold: Book, targets: list[str], scope) -> list[dict]:
    """Originals that are broken in ways no disclosure can fix."""
    out = []
    seen_rows = set()
    for k in sorted(set(scope)):
        f = gold.formula.get(k)
        lab = gold.row_label(k)
        if not (f and lab and re.search(r"\(\s*1\s*-\s*t\w*\s*\)", lab, re.I)):
            continue
        m = re.search(r"\*\s*\(\s*1\s*\+\s*([^)]+)\)", f)
        if not m:
            continue
        sheet, coord = k.split("!", 1)
        p = split_coord(coord)
        if not p or (sheet, p[1]) in seen_rows:
            continue
        # A negative rate makes (1 + rate) the same as (1 - T). Only a positive
        # rate contradicts the label.
        rates = [gold.value.get(r) for r in refs_in("=" + m.group(1), sheet)]
        rates = [v for v in rates if isinstance(v, (int, float))]
        if rates and all(v < 0 for v in rates):
            continue
        seen_rows.add((sheet, p[1]))
        out.append({
            "kind": "label_sign_mismatch",
            "cell": pretty(k),
            "detail": "Label %r but formula multiplies by (1 + rate) and the rate is %s: %s"
                      % (lab, ("%.6g" % rates[0]) if rates else "not resolvable", f),
        })
    for k in targets:
        f = gold.formula.get(k)
        v = gold.value.get(k)
        if f and isinstance(v, (int, float)) and abs(v) < 1e-12:
            sheet = k.split("!", 1)[0]
            operands = refs_in(f, sheet)
            empty = [r for r in operands if not gold.has(r)]
            if empty:
                out.append({
                    "kind": "empty_operand_zero_target",
                    "cell": pretty(k),
                    "detail": "Graded target is 0 because %s is empty in the original: %s"
                              % (_fmt(empty, 2), f),
                })
    return out


# -------------------------------------------------------------------- audit

def audit(gold: Book, inp: Book, targets, tol, records, blanked_cells) -> dict:
    explained = set()
    for r in records:
        for c in r["cells"]:
            if "!row " in c:
                sheet, rownum = c.split("!row ")
                sheet = unquote_sheet(sheet)
                for k in blanked_cells:
                    s, coord = k.split("!", 1)
                    p = split_coord(coord)
                    if s == sheet and p and p[1] == int(rownum):
                        explained.add(k)
            else:
                sheet, coord = c.rsplit("!", 1)
                explained.add(key(unquote_sheet(sheet), coord))
    unexplained = [k for k in blanked_cells if k not in explained]

    # Answer leakage: a record whose cells include a graded target discloses the answer.
    tset = set(targets)
    leaks = []
    for r in records:
        for c in r["cells"]:
            if "!row " in c:
                continue
            sheet, coord = c.rsplit("!", 1)
            if key(unquote_sheet(sheet), coord) in tset:
                leaks.append({"family": r["family"], "cell": c})
    return {
        "blanked_in_closure": len(blanked_cells),
        "explained": len(blanked_cells) - len(unexplained),
        "unexplained": len(unexplained),
        "coverage": round(1 - len(unexplained) / len(blanked_cells), 4) if blanked_cells else 1.0,
        "unexplained_sample": [pretty(k) for k in unexplained[:25]],
        "answer_leaks": leaks,
    }


# ------------------------------------------------------------------ render

def compact_cells(cells: list[str]) -> str:
    """Collapse a run of cells on one row into a single range reference."""
    if not cells:
        return ""
    if len(cells) == 1 or "!row " in cells[0]:
        return "`%s`" % cells[0]
    parsed = []
    for c in cells:
        sheet, coord = c.rsplit("!", 1)
        p = split_coord(coord)
        if not p:
            return ", ".join("`%s`" % x for x in cells[:4])
        parsed.append((sheet, p[1], col_to_num(p[0]), coord))
    sheets = {p[0] for p in parsed}
    rows = {p[1] for p in parsed}
    if len(sheets) == 1 and len(rows) == 1:
        parsed.sort(key=lambda x: x[2])
        return "`%s!%s:%s`" % (parsed[0][0], parsed[0][3], parsed[-1][3])
    shown = ", ".join("`%s`" % c for c in cells[:4])
    return shown + (" and %d more" % (len(cells) - 4) if len(cells) > 4 else "")


def render_section(records: list[dict]) -> str:
    if not records:
        return ""
    by_family = defaultdict(list)
    for r in records:
        by_family[r["family"]].append(r)
    lines = [
        "## Modelling conventions in this workbook",
        "",
        "These state decisions the original model made that the emptied file no longer",
        "shows. They describe method, not the requested answers.",
        "",
    ]
    for fam in sorted(by_family):
        for r in by_family[fam]:
            lines.append("- **%s** (%s): `%s` — %s"
                         % (fam.replace("_", " "), compact_cells(r["cells"]),
                            r["value"], r["note"]))
    return "\n".join(lines) + "\n"


# ------------------------------------------------------------------ command

def prepare(args):
    tid, inputs, golden = task_paths(args.task_dir, args.golden)
    if not golden or not os.path.exists(golden):
        sys.exit("golden workbook not found; pass --golden")
    if not inputs or not os.path.exists(inputs):
        sys.exit("delivered workbook not found under %s/environment" % args.task_dir)
    targets_map, tol = load_targets(args.task_dir)
    gold = Book(golden)
    inp = Book(inputs)
    default_sheet = gold.sheets[0]
    targets = [parse_ref(t, default_sheet) for t in targets_map]
    print("task %s  golden=%s  targets=%d" % (tid, os.path.basename(golden), len(targets)))
    return tid, gold, inp, targets, targets_map, tol


def cmd_closure(args):
    tid, gold, inp, targets, targets_map, tol = prepare(args)
    allc = closure(gold, targets)
    bl = blanked(gold, inp, allc)
    print("closure %d cells, %d of them computed by the original and absent from the delivered file"
          % (len(allc), len(bl)))
    per = {}
    for t in targets:
        c = closure(gold, [t])
        b = blanked(gold, inp, c)
        per[t] = (len(c), len(b))
        print("  %-34s closure %5d  deleted %5d" % (pretty(t), len(c), len(b)))
    if args.list:
        for k in bl[:args.list]:
            print("    %-28s %s" % (pretty(k), (gold.formula.get(k) or "")[:70]))


def cmd_detect(args):
    tid, gold, inp, targets, targets_map, tol = prepare(args)
    allc = closure(gold, targets)
    bl = blanked(gold, inp, allc)
    records = detect_all(gold, inp, allc, bl, targets)
    print("%d convention record(s)" % len(records))
    for r in records:
        cells = r["cells"]
        shown = ", ".join(cells[:5]) + (" +%d more" % (len(cells) - 5) if len(cells) > 5 else "")
        print("  [%s] %s" % (r["family"], r["value"]))
        print("      cells    %s" % shown)
        print("      evidence %s" % str(r["evidence"])[:88])
        print("      %s" % r["note"])


def cmd_emit(args):
    tid, gold, inp, targets, targets_map, tol = prepare(args)
    allc = closure(gold, targets)
    bl = blanked(gold, inp, allc)
    records = detect_all(gold, inp, allc, bl, targets)
    rep = audit(gold, inp, targets, tol, records, bl)
    manifest = {
        "task": os.path.basename(os.path.abspath(args.task_dir)),
        "golden": os.path.basename(task_paths(args.task_dir, args.golden)[2]),
        "targets": list(targets_map),
        "conventions": records,
        "audit": rep,
    }
    out = args.out or os.path.join(args.task_dir, "tests", "conventions.json")
    if args.dry_run:
        print(json.dumps(manifest, indent=2, default=str)[:4000])
    else:
        with open(out, "w") as fh:
            json.dump(manifest, fh, indent=2, default=str)
            fh.write("\n")
        print("wrote %s (%d records, coverage %.1f%%)"
              % (out, len(records), 100 * rep["coverage"]))
    print()
    print(render_section(records))


STRUCTURAL_FAMILIES = {
    "terminal_value", "row_populated", "inert_line", "discount_period",
    "npv_timing", "stake_scaling", "source_selection",
}
SECTION_START = "## Modelling conventions in this workbook"
OUTPUT_HEADING = "\n## Output\n"
FORMULA_RE = re.compile(r"`=|\bIF\(|\bSUM\(|\bAVERAGE\(|\bNPV\(|\bIRR\(")


def agent_facing(records: list[dict], leaks: list[dict]) -> list[dict]:
    """Records safe to put in front of the model.

    Drops anything the audit flagged as naming a graded cell, and strips the
    evidence field so no golden formula can travel with the disclosure.
    """
    flagged = {l["cell"] for l in leaks}
    out = []
    for r in records:
        if any(c in flagged for c in r["cells"]):
            continue
        out.append({k: v for k, v in r.items() if k != "evidence"})
    return out


def cmd_apply(args):
    """Write the disclosure into instruction.md, where the model will actually read it."""
    tid, gold, inp, targets, targets_map, tol = prepare(args)
    allc = closure(gold, targets)
    bl = blanked(gold, inp, allc)
    records = detect_all(gold, inp, allc, bl, targets)
    rep = audit(gold, inp, targets, tol, records, bl)
    safe = records if args.force else agent_facing(records, rep["answer_leaks"])
    withheld_leak = len(records) - len(safe)
    if args.families == "structural":
        before = len(safe)
        safe = [r for r in safe if r["family"] in STRUCTURAL_FAMILIES]
        filtered = before - len(safe)
    else:
        filtered = 0
    section = render_section(safe)

    # A disclosure must never carry a formula.
    offending = [ln for ln in section.splitlines() if FORMULA_RE.search(ln)]
    if offending and not args.force:
        print("refusing to write: %d line(s) look like formulas" % len(offending))
        for ln in offending[:5]:
            print("   %s" % ln[:110])
        sys.exit(2)

    src = os.path.abspath(args.task_dir)
    dst = os.path.abspath(args.out) if args.out else src
    if dst != src:
        import shutil
        if os.path.exists(dst):
            shutil.rmtree(dst)
        # Clone the shipping bundle only. Run outputs and analysis scratch are not
        # task inputs and dominate the directory size.
        shutil.copytree(src, dst, ignore=shutil.ignore_patterns(
            "*-traces", "__pycache__", "gap_attrib.py", "*.pyc", "verify_channels.py"))

    path = os.path.join(dst, "instruction.md")
    text = open(path, encoding="utf-8").read()
    # Idempotent: drop any section this tool wrote before.
    if SECTION_START in text:
        head, rest = text.split(SECTION_START, 1)
        tail = rest.split("## ", 1)
        text = head + ("## " + tail[1] if len(tail) > 1 else "")
    if OUTPUT_HEADING not in text:
        sys.exit("instruction.md has no Output heading to insert before")
    text = text.replace(OUTPUT_HEADING, "\n" + section + "\n## Output\n", 1)
    if args.dry_run:
        print(section)
        print("(dry run; would write %s)" % path)
    else:
        open(path, "w", encoding="utf-8").write(text)
        print("wrote %s" % path)
    print("%d disclosed, %d withheld as answer-leak candidates, %d filtered out by --families %s"
          % (len(safe), withheld_leak, filtered, args.families))


def cmd_map(args):
    tid, gold, inp, targets, targets_map, tol = prepare(args)
    allc = set(closure(gold, targets))
    sheets = sorted({k.split("!", 1)[0] for k in allc})
    used = defaultdict(set)
    for k in list(gold.formula) + list(gold.value):
        s, coord = k.split("!", 1)
        p = split_coord(coord)
        if p and s in sheets:
            used[s].add(p[1])
    print("rows the original actually uses, on sheets a graded answer touches")
    for s in sheets:
        rows = sorted(used[s])
        if not rows:
            continue
        spans, start, prev = [], rows[0], rows[0]
        for r in rows[1:]:
            if r != prev + 1:
                spans.append((start, prev))
                start = r
            prev = r
        spans.append((start, prev))
        txt = ", ".join("%d" % a if a == b else "%d-%d" % (a, b) for a, b in spans[:14])
        print("  %-28s %s" % (quote_sheet(s), txt))


def cmd_audit(args):
    tid, gold, inp, targets, targets_map, tol = prepare(args)
    allc = closure(gold, targets)
    bl = blanked(gold, inp, allc)
    records = detect_all(gold, inp, allc, bl, targets)
    rep = audit(gold, inp, targets, tol, records, bl)
    print("deleted cells feeding a graded answer : %d" % rep["blanked_in_closure"])
    print("explained by a convention record      : %d" % rep["explained"])
    print("unexplained                           : %d  (coverage %.1f%%)"
          % (rep["unexplained"], 100 * rep["coverage"]))
    if rep["answer_leaks"]:
        print("ANSWER LEAK: a record names a graded cell directly")
        for l in rep["answer_leaks"]:
            print("   %s  %s" % (l["cell"], l["family"]))
    if rep["unexplained_sample"]:
        print("unexplained sample:")
        for k in rep["unexplained_sample"]:
            print("   %s" % k)


def cmd_defects(args):
    tid, gold, inp, targets, targets_map, tol = prepare(args)
    allc = closure(gold, targets)
    d = detect_defects(gold, targets, allc)
    if not d:
        print("no defect signature fired")
    for x in d:
        print("  [%s] %s" % (x["kind"], x["cell"]))
        print("      %s" % x["detail"])


def main(argv=None):
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("command", choices=["closure", "detect", "emit", "apply", "map",
                                        "audit", "defects"])
    ap.add_argument("--task-dir", required=True)
    ap.add_argument("--golden", default=None)
    ap.add_argument("--out", default=None)
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--list", type=int, default=0)
    ap.add_argument("--force", action="store_true",
                    help="disclose flagged records too (unsafe; review first)")
    ap.add_argument("--families", choices=["all", "structural"], default="all",
                    help="structural drops routine projection rules and totals")
    args = ap.parse_args(argv)
    {
        "closure": cmd_closure,
        "detect": cmd_detect,
        "emit": cmd_emit,
        "apply": cmd_apply,
        "map": cmd_map,
        "audit": cmd_audit,
        "defects": cmd_defects,
    }[args.command](args)


if __name__ == "__main__":
    main()
