#!/usr/bin/env python3
"""Generate a per-workbook atomic normalizer from a reviewed variable-source audit.

Applies the review rules established on the 0438 pilot:

* a draft row is included only when every referenced cell is a typed cell in the
  golden workbook and still present in the baseline inputs workbook;
* a value that also appears in an uncovered typed cell is a leak.  The leak is
  covered with ``extra_cells`` only when the duplicate sits on a data-validation
  or picklist sheet, which carries no modelling content.  Any other leak excludes
  the variable outright, because partial masking is forbidden;
* a multi-cell audit row is split into one atomic variable per cell.

The emitted script writes ``normalized.json``, ``exclusions.json`` and
``normalization_report.json`` atomically and is byte-stable across re-runs.
"""

from __future__ import annotations

import collections
import json
import os
import pprint
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.utils import column_index_from_string as ci, get_column_letter as gl

# Locate !A1 / !A1:B2, then take the sheet name immediately before the bang.
# Excel sheet names are at most 31 characters and cannot contain \ / ? * [ ] : !
CELL = r"\$?[A-Z]{1,3}\$?\d+(?::\$?[A-Z]{1,3}\$?\d+)?"
CELL_AT = re.compile(r"!(" + CELL + r")")
SHEET_BREAK = re.compile(r"`|<br\s*/?>|[\n;]", re.I)
PICKLIST = re.compile(r"data\s*validation|validation|dropdown|^lists?$|picklist"
                      r"|embedded\s*assumptions", re.I)

CAUSE_PRIORITY = (
    "unparsable_reference",
    "unknown_sheet",
    "formula_cell",
    "blank_in_baseline",
    "duplicate_value_leak",
    "forced_exclusion",
)

CAUSE_PROSE = {
    "unparsable_reference": (
        "No workbook cell reference could be parsed from the audit row."
    ),
    "unknown_sheet": (
        "Audit row references a sheet that does not exist in the golden workbook."
    ),
    "formula_cell": (
        "Audit row references a formula cell; it cannot be masked as a typed input."
    ),
    "blank_in_baseline": (
        "Audit row references a cell already blank in the baseline inputs; "
        "it cannot be masked as a typed input."
    ),
    "duplicate_value_leak": (
        "Value also occurs in typed cell(s) that carry unrelated modelling "
        "content, so the variable cannot be masked completely."
    ),
    "forced_exclusion": (
        "The delivered bundle exposes this value in a packager-generated "
        "cell that cannot be masked without removing task disclosure, so "
        "the variable is excluded (oracle feedback)."
    ),
}


def _source_root(wb):
    from pathlib import Path as _P
    return "4-10 100" if _P(f"4-10 100/{wb}.xlsx").exists() else "batch-src"


def extract_refs(text, sheets=None):
    """Return ``Sheet!A1`` / ``Sheet!A1:B2`` strings from audit prose."""
    refs = []
    seen = set()
    keys = {k.strip().upper(): k for k in (sheets or {})}
    for match in CELL_AT.finditer(text or ""):
        cell = match.group(1)
        prefix = text[: match.start()]
        quoted = re.search(r"'([^']+)'\s*$", prefix)
        if quoted:
            sheet = quoted.group(1).strip()
        else:
            chunk = SHEET_BREAK.split(prefix)[-1]
            sheet = chunk[-31:].strip(" \t\"=[")
        if not sheet or not cell:
            continue
        if keys:
            upper = sheet.strip().upper()
            if upper not in keys:
                resolved = None
                for length in range(len(upper), 0, -1):
                    cand = upper[-length:].strip()
                    if cand in keys:
                        resolved = keys[cand]
                        break
                if resolved is not None:
                    sheet = resolved
        ref = f"{sheet}!{cell}"
        if ref not in seen:
            seen.add(ref)
            refs.append(ref)
    return refs


def expand(ref, sheets):
    ref = ref.strip().replace("$", "").strip("'").strip("`")
    m = re.match(r"^(.+?)!([A-Z]{1,3})(\d+)(?::([A-Z]{1,3})(\d+))?$", ref, re.I)
    if not m:
        return None, []
    sh, c1, r1, c2, r2 = m.groups()
    real = sheets.get(sh.strip().upper())
    if real is None:
        return None, []
    c1, c2 = c1.upper(), (c2.upper() if c2 else None)
    if not c2:
        return real, [f"{c1}{r1}"]
    return real, [
        f"{gl(c)}{r}"
        for c in range(ci(c1), ci(c2) + 1)
        for r in range(int(r1), int(r2) + 1)
    ]


def primary_code(codes):
    present = [c for c in codes if c]
    for code in CAUSE_PRIORITY:
        if code in present:
            return code
    return present[0] if present else None


def slug(text, fallback="variable"):
    s = re.sub(r"[^a-z0-9]+", "-", str(text).lower()).strip("-")
    return s or fallback


def coerce(value):
    """Spec values must be JSON-serialisable; dates become ISO strings."""
    import datetime as _dt
    if isinstance(value, _dt.datetime):
        return value.date().isoformat(), True
    if isinstance(value, _dt.date):
        return value.isoformat(), True
    return value, False


def unit_for(value, name):
    low = name.lower()
    if isinstance(value, (int, float)) and -1.5 <= value <= 1.5 and value != int(value):
        if any(k in low for k in ("rate", "margin", "share", "premium", "growth",
                                  "percent", "%", "inflation", "discount", "tax")):
            return "ratio", True
    if any(k in low for k in ("days", "dso", "dpo")):
        return "days", False
    if any(k in low for k in ("year", "life", "term")):
        return "years", False
    if any(k in low for k in ("month",)):
        return "months", False
    return "value", False


def atomic_json(path, payload):
    path = Path(path)
    tmp = path.with_suffix(path.suffix + ".tmp")
    with tmp.open("w", encoding="utf-8") as handle:
        json.dump(payload, handle, indent=2, ensure_ascii=False, sort_keys=True)
        handle.write("\n")
        handle.flush()
        os.fsync(handle.fileno())
    os.replace(tmp, path)


def analyse(wb):
    run = Path(f"runs/{wb}-variable-sources")
    draft = json.loads((run / "draft.json").read_text(encoding="utf-8"))
    SR = _source_root(wb)
    G = openpyxl.load_workbook(f"{SR}/{wb}.xlsx", data_only=False)
    GV = openpyxl.load_workbook(f"{SR}/{wb}.xlsx", data_only=True)
    BI = openpyxl.load_workbook(f"inputs_out/{wb}-inputs.xlsx", data_only=False)
    sheets = {s.strip().upper(): s for s in G.sheetnames}

    def vkey(v):
        """Any value the agent could read back, not just numbers."""
        import datetime as _dt
        if v is True or v is False or v is None:
            return None
        if isinstance(v, (int, float)):
            return ("n", round(float(v), 10))
        if isinstance(v, (_dt.datetime, _dt.date)):
            return ("d", str(v)[:10])
        if isinstance(v, str) and v.strip():
            return ("s", v.strip())
        return None

    val_index = collections.defaultdict(list)
    for s in GV.sheetnames:
        for row in GV[s].iter_rows():
            for c in row:
                k = vkey(c.value)
                if k is not None:
                    val_index[k].append(f"{s}!{c.coordinate}")

    rows = []
    for row in draft["rows"]:
        cells = []
        codes = []
        refs = extract_refs(row["value_text"], sheets)
        if not refs:
            codes.append("unparsable_reference")
        for ref in refs:
            sh, cs = expand(ref, sheets)
            if sh is None:
                codes.append("unknown_sheet")
                continue
            for c in cs:
                gf = G[sh][c].value
                if isinstance(gf, str) and gf.startswith("="):
                    codes.append("formula_cell")
                    continue
                if BI[sh][c].value is None:
                    codes.append("blank_in_baseline")
                    continue
                cells.append((sh, c, GV[sh][c].value))
        unique_codes = []
        for code in codes:
            if code not in unique_codes:
                unique_codes.append(code)
        rows.append({
            "row": row,
            "cells": cells,
            "codes": unique_codes,
            "clean": bool(cells) and not unique_codes,
        })

    for r in rows:
        if not r["clean"]:
            code = primary_code(r["codes"])
            r["code"] = code
            r["reason"] = CAUSE_PROSE.get(code, CAUSE_PROSE["unparsable_reference"])
        r["extra"] = []

    # Dropping a row un-covers the values it was masking, which can create a new
    # leak elsewhere. Iterate until the surviving set is leak-free.
    while True:
        in_scope = {f"{sh}!{c}" for r in rows if r["clean"] for sh, c, _ in r["cells"]}
        covered = set(in_scope)
        for r in rows:
            if r["clean"]:
                covered |= set(r["extra"])
        newly_dropped = False
        for r in rows:
            if not r["clean"]:
                continue
            extra, leaks = list(r["extra"]), []
            for sh, c, val in r["cells"]:
                k = vkey(val)
                if k is None:
                    continue
                for other in val_index.get(k, []):
                    if other in covered or other in extra:
                        continue
                    osh, oc = other.split("!", 1)
                    # A formula cell can still ship its cached value, so the only
                    # question that matters is whether the delivered book holds it.
                    if BI[osh][oc].value is None:
                        continue
                    if PICKLIST.search(osh):
                        extra.append(other)
                    else:
                        leaks.append(other)
            r["extra"] = sorted(set(extra))
            if leaks:
                r["clean"] = False
                r["leaks"] = sorted(set(leaks))
                r["code"] = "duplicate_value_leak"
                r["codes"] = list(r.get("codes") or []) + ["duplicate_value_leak"]
                r["reason"] = (
                    "Value also occurs in typed cell(s) %s that carry unrelated "
                    "modelling content, so the variable cannot be masked completely."
                    % ", ".join(sorted(set(leaks))[:4])
                )
                newly_dropped = True
        if not newly_dropped:
            break
    return draft, rows


def emit(wb, draft, rows):
    # Cells the packager creates (e.g. the Embedded Assumptions disclosure sheet)
    # only exist in the delivered bundle, so the oracle is the first thing that can
    # see those duplicates. Feed its verdict back in here.
    forced = set()
    fp = Path(f"runs/{wb}-variable-sources/oracle_forced_exclusions.json")
    if fp.exists():
        raw = json.loads(fp.read_text(encoding="utf-8"))
        if isinstance(raw, list):
            forced = set(raw)
        elif isinstance(raw, dict):
            forced = set(raw.get("ids") or raw.get("exclusions") or [])

    included = [r for r in rows if r["clean"]]

    var_defs, inc_map, reasons = [], {}, {}
    used = set()
    for r in included:
        ids = []
        multi = len(r["cells"]) > 1
        base = slug(r["row"]["draft_id"])
        if base in forced or r["row"]["draft_id"] in forced:
            r["clean"] = False
            r["code"] = "forced_exclusion"
            r["codes"] = list(r.get("codes") or []) + ["forced_exclusion"]
            r["reason"] = CAUSE_PROSE["forced_exclusion"]
            continue
        for sh, c, raw_val in r["cells"]:
            val, is_date = coerce(raw_val)
            vid = f"{base}-{slug(sh)}-{c.lower()}" if multi else base
            n = 2
            while vid in used:
                vid = f"{base}-{slug(sh)}-{c.lower()}-{n}"
                n += 1
            used.add(vid)
            ids.append(vid)
            name = r["row"]["variable_text"].strip()
            unit, _ = ("date", False) if is_date else unit_for(val, name)
            display = val   # exact workbook value; never rescaled
            var_defs.append({
                "id": vid,
                "name": name if not multi else f"{name} ({sh}!{c})",
                "value": display,
                "raw": None,
                "unit": unit,
                "period": "Model horizon",
                "basis": "Externally sourced assumption",
                "question": (
                    "What %s should the model use, per the sourced research?"
                    % ((name[0].lower() + name[1:]) if name else "this assumption")
                ),
                "cells": [f"{sh}!{c}"],
                "extra_cells": r["extra"] if not multi else [],
            })
        inc_map[r["row"]["draft_id"]] = ids
        for e in r["extra"]:
            reasons[e] = (
                "Data-validation / picklist cell repeating the sourced value; "
                "masked so it cannot reveal the removed assumption."
            )
    excluded = [r for r in rows if not r["clean"]]
    exc_map = {}
    for r in excluded:
        code = r.get("code") or primary_code(r.get("codes") or []) or "unparsable_reference"
        exc_map[r["row"]["draft_id"]] = {
            "reason": r.get("reason") or CAUSE_PROSE[code],
            "code": code,
        }
    inc_map = {k: v for k, v in inc_map.items() if k not in exc_map}
    keep = {i for ids in inc_map.values() for i in ids}
    var_defs = [v for v in var_defs if v["id"] in keep]

    histogram = collections.Counter(item["code"] for item in exc_map.values())

    run = Path(f"runs/{wb}-variable-sources")
    first_path = run / "first_normalization.json"
    snapshot = {
        "schema_version": "1.0",
        "workbook": wb,
        "atomic_variables": len(var_defs),
        "draft_rows": len(draft["rows"]),
        "included_rows": len(inc_map),
        "excluded_rows": len(exc_map),
        "exclusion_reason_codes": dict(sorted(histogram.items())),
        "forced_exclusions": sorted(forced),
    }
    if not first_path.exists():
        atomic_json(first_path, snapshot)
    first = json.loads(first_path.read_text(encoding="utf-8"))

    script = run / f"normalize_{wb}.py"
    script.write_text(TEMPLATE.format(
        wb=wb,
        seed=int(wb),
        variables=pprint.pformat(var_defs, indent=4, width=96, sort_dicts=True),
        included=pprint.pformat(inc_map, indent=4, width=96, sort_dicts=True),
        excluded=pprint.pformat(exc_map, indent=4, width=96, sort_dicts=True),
        reasons=pprint.pformat(reasons, indent=4, width=96, sort_dicts=True),
        codes=pprint.pformat(dict(sorted(histogram.items())), indent=4,
                             width=96, sort_dicts=True),
        first_atomic=int(first.get("atomic_variables") or 0),
        forced=pprint.pformat(sorted(forced), indent=4, width=96),
    ), encoding="utf-8")
    print(f"{wb}: {len(draft['rows'])} draft rows -> {len(inc_map)} included, "
          f"{len(exc_map)} excluded, {len(var_defs)} atomic variables, "
          f"{len(reasons)} extra cells")
    return len(var_defs)


TEMPLATE = '''#!/usr/bin/env python3
"""Normalize the reviewed {wb} variable-source audit atomically (generated)."""

from __future__ import annotations

import json
import os
from pathlib import Path

RUN = Path(__file__).resolve().parent
DRAFT = RUN / "draft.json"
NORMALIZED = RUN / "normalized.json"
EXCLUSIONS = RUN / "exclusions.json"
REPORT = RUN / "normalization_report.json"
PROFILES = RUN / "source_profiles.json"

RESEARCH = {{
    "id": "sourced-market-research",
    "name": "Sourced market and contractual research",
    "url": "internal://research-service/sourced-assumptions",
    "role": "primary",
    "kind": "internal",
}}

VAR_DEFS = {variables}

INCLUDED = {included}

EXCLUDED = {excluded}

EXTRA_CELL_REASONS = {reasons}

EXCLUSION_REASON_CODES = {codes}

FIRST_ATOMIC_VARIABLES = {first_atomic}

FORCED_EXCLUSIONS = {forced}


def build_variables():
    out = []
    for d in VAR_DEFS:
        workbook = {{"cells": d["cells"]}}
        if d.get("raw") is not None:
            workbook["value"] = d["raw"]
        if d.get("extra_cells"):
            workbook["extra_cells"] = d["extra_cells"]
        out.append({{
            "id": d["id"],
            "name": d["name"],
            "value": d["value"],
            "unit": d["unit"],
            "entity": "Workbook {wb} subject company",
            "period": d["period"],
            "scenario": "Base case",
            "basis": d["basis"],
            "status": "Approved",
            "question": d["question"],
            "sources": [RESEARCH],
            "workbook": workbook,
        }})
    return out


def atomic_json(path, payload):
    tmp = path.with_suffix(path.suffix + ".tmp")
    with tmp.open("w", encoding="utf-8") as handle:
        json.dump(payload, handle, indent=2, ensure_ascii=False, sort_keys=True)
        handle.write("\\n")
        handle.flush()
        os.fsync(handle.fileno())
    os.replace(tmp, path)


def excluded_entry(draft_id):
    item = EXCLUDED[draft_id]
    if isinstance(item, dict):
        return item["reason"], item.get("code")
    return item, None


def main():
    draft = json.loads(DRAFT.read_text(encoding="utf-8"))
    draft_ids = [row["draft_id"] for row in draft["rows"]]
    dispositions = []
    for draft_id in draft_ids:
        if draft_id in INCLUDED:
            dispositions.append({{
                "draft_id": draft_id,
                "status": "included",
                "variable_ids": INCLUDED[draft_id],
            }})
        elif draft_id in EXCLUDED:
            reason, code = excluded_entry(draft_id)
            row = {{
                "draft_id": draft_id,
                "status": "excluded",
                "reason": reason,
            }}
            if code:
                row["code"] = code
            dispositions.append(row)
        else:
            raise SystemExit("unresolved draft row: %s" % draft_id)

    variables = build_variables()
    ids = {{v["id"] for v in variables}}
    referenced = {{
        vid for d in dispositions if d["status"] == "included"
        for vid in d["variable_ids"]
    }}
    if ids != referenced:
        raise SystemExit("variable/disposition mismatch")
    if len(dispositions) != len(set(d["draft_id"] for d in dispositions)):
        raise SystemExit("dispositions are not one-to-one")

    spec = {{
        "schema_version": "2.0.0",
        "environment_id": "fcp-{wb}",
        "seed": {seed},
        "variables": variables,
    }}
    if PROFILES.exists():
        spec["source_profiles"] = (
            json.loads(PROFILES.read_text(encoding="utf-8")).get("profiles") or [])

    atomic_json(NORMALIZED, spec)
    atomic_json(EXCLUSIONS, {{
        "schema_version": "1.0",
        "workbook": "{wb}",
        "exclusions": [
            {{
                "draft_id": k,
                **({{"reason": v, "code": None}} if not isinstance(v, dict) else v),
            }}
            for k, v in sorted(EXCLUDED.items())
        ],
    }})
    atomic_json(REPORT, {{
        "schema_version": "1.0",
        "workbook": "{wb}",
        "draft_rows": len(draft_ids),
        "included_rows": len(INCLUDED),
        "excluded_rows": len(EXCLUDED),
        "atomic_variables": len(variables),
        "first_atomic_variables": FIRST_ATOMIC_VARIABLES,
        "exclusion_reason_codes": EXCLUSION_REASON_CODES,
        "forced_exclusions": FORCED_EXCLUSIONS,
        "extra_cell_reasons": EXTRA_CELL_REASONS,
        "dispositions": dispositions,
    }})
    print("{wb}: %d draft rows -> %d included, %d excluded, %d atomic variables"
          % (len(draft_ids), len(INCLUDED), len(EXCLUDED), len(variables)))


if __name__ == "__main__":
    main()
'''


if __name__ == "__main__":
    wb = sys.argv[1]
    draft, rows = analyse(wb)
    emit(wb, draft, rows)
