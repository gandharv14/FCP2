"""Offline duplicate/leak detector and leakscan CLI tests.

These prove the detector extracted from the HTTP oracle
(``xl_mcp_oracle.scan_value_leaks``) finds planted duplicate representations
without a live server, respects an explicit allowlist, and that the
``xl_variable_mcp.py leakscan`` subcommand exits nonzero on any unapproved
leak and zero on a clean workbook.
"""

from __future__ import annotations

import hashlib
import io
import json
import tempfile
import unittest
from contextlib import redirect_stdout
from pathlib import Path

import xl_variable_mcp
from xl_variable_mcp import build_leak_patch, draft_chain
from xl_mcp_oracle import check_workbook, load_allowlist, scan_value_leaks


def _audit():
    return [
        {
            "variable_id": "asset-capacity",
            "refs": ["Model!A1"],
            "cell_value": 42,
            "mcp_value": 42,
            "unit": "MW",
        },
        {
            "variable_id": "ebitda-margin",
            "refs": ["Model!A2"],
            "cell_value": 0.125,
            "mcp_value": 12.5,
            "unit": "percent",
        },
    ]


MASK = ["Model!A1", "Model!A2"]


class ScanValueLeaksTests(unittest.TestCase):
    def setUp(self):
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl is not installed")
        self.openpyxl = openpyxl

    def _workbook(self, path, cells):
        book = self.openpyxl.Workbook()
        sheet = book.active
        sheet.title = "Model"
        for coordinate, value in cells.items():
            sheet[coordinate] = value
        book.save(path)
        return path

    def test_finds_planted_numeric_and_formatted_text_duplicates(self):
        with tempfile.TemporaryDirectory() as temp:
            workbook = self._workbook(Path(temp) / "golden.xlsx", {
                "A1": 42,          # masked cell still holds its raw value
                "A2": 0.125,       # masked percent cell, raw form
                "B1": 42,          # planted numeric duplicate
                "C1": "42 MW",     # planted formatted-text rendering
                "B2": "12.5%",     # planted percent rendering of 0.125
                "D1": "no leak here",
            })
            report = scan_value_leaks(
                workbook, MASK, _audit(), None, require_blank_mask=False
            )
            self.assertFalse(report["valid"])
            self.assertEqual(report["masked_cells_not_blank"], [])
            found = {
                (leak["cell"], leak["variable_id"])
                for leak in report["unapproved_value_leaks"]
            }
            self.assertEqual(found, {
                ("Model!B1", "asset-capacity"),
                ("Model!C1", "asset-capacity"),
                ("Model!B2", "ebitda-margin"),
            })

    def test_allowlist_is_respected_and_unused_entries_fail(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            workbook = self._workbook(root / "golden.xlsx", {
                "A1": 42,
                "A2": 0.125,
                "B1": 42,
            })
            allowlist = root / "allowlist.json"
            allowlist.write_text(json.dumps({
                "workbook_value_leaks": [{
                    "cell": "Model!B1",
                    "variable_id": "asset-capacity",
                    "reason": "required period-axis label",
                }]
            }), encoding="utf-8")
            report = scan_value_leaks(
                workbook, MASK, _audit(), allowlist, require_blank_mask=False
            )
            self.assertTrue(report["valid"], report)
            self.assertEqual(report["unapproved_value_leaks"], [])
            self.assertEqual(len(report["approved_value_leaks"]), 1)

            allowlist.write_text(json.dumps({
                "workbook_value_leaks": [
                    {
                        "cell": "Model!B1",
                        "variable_id": "asset-capacity",
                        "reason": "required period-axis label",
                    },
                    {
                        "cell": "Model!Z99",
                        "variable_id": "asset-capacity",
                        "reason": "stale entry",
                    },
                ]
            }), encoding="utf-8")
            report = scan_value_leaks(
                workbook, MASK, _audit(), allowlist, require_blank_mask=False
            )
            self.assertFalse(report["valid"])
            self.assertEqual(len(report["unused_allowlist_entries"]), 1)

    def test_blank_mask_enforcement_matches_the_masked_workbook_mode(self):
        with tempfile.TemporaryDirectory() as temp:
            workbook = self._workbook(Path(temp) / "masked.xlsx", {
                "A1": 42,
                "A2": None,
                "B1": "no leak here",
            })
            report = scan_value_leaks(workbook, MASK, _audit(), None)
            self.assertFalse(report["valid"])
            self.assertEqual(
                report["masked_cells_not_blank"][0]["cell"], "Model!A1"
            )

    def test_default_mode_is_byte_identical_to_the_oracle_composition(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            workbook = self._workbook(root / "shipped.xlsx", {
                "A1": None,
                "A2": None,
                "B1": "42 MW",
            })
            allowlist = root / "allowlist.json"
            allowlist.write_text(json.dumps({
                "workbook_value_leaks": [{
                    "cell": "Model!B1",
                    "variable_id": "asset-capacity",
                    "reason": "required label",
                }]
            }), encoding="utf-8")
            allowed, errors = load_allowlist(
                allowlist, {"asset-capacity", "ebitda-margin"}
            )
            expected = check_workbook(workbook, MASK, _audit(), allowed)
            expected["allowlist_errors"] = errors
            actual = scan_value_leaks(workbook, MASK, _audit(), allowlist)
            self.assertEqual(
                json.dumps(actual, sort_keys=True),
                json.dumps(expected, sort_keys=True),
            )


class RenderingEnrichmentTests(unittest.TestCase):
    """Every leak entry additionally reports its rendering kind (added key)."""

    def setUp(self):
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl is not installed")
        self.openpyxl = openpyxl

    def test_leak_entries_carry_the_closed_rendering_enum(self):
        with tempfile.TemporaryDirectory() as temp:
            book = self.openpyxl.Workbook()
            sheet = book.active
            sheet.title = "Model"
            sheet["A1"] = 42
            sheet["A2"] = 0.125
            sheet["B1"] = 42
            sheet["C1"] = "42 MW"
            sheet["B2"] = "12.5%"
            workbook = Path(temp) / "golden.xlsx"
            book.save(workbook)
            report = scan_value_leaks(
                workbook, MASK, _audit(), None, require_blank_mask=False
            )
            renderings = {
                (leak["cell"], leak["variable_id"]): leak["rendering"]
                for leak in report["unapproved_value_leaks"]
            }
            self.assertEqual(renderings, {
                ("Model!B1", "asset-capacity"): "raw_number",
                ("Model!C1", "asset-capacity"): "formatted_text",
                ("Model!B2", "ebitda-margin"): "percent_rendering",
            })


class LeakPatchTests(unittest.TestCase):
    def setUp(self):
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl is not installed")
        self.openpyxl = openpyxl

    def _spec(self, root):
        spec = root / "normalized.json"
        spec.write_text(json.dumps({
            "variables": [
                {
                    "id": "asset-capacity",
                    "name": "Asset capacity",
                    "value": 42,
                    "unit": "MW",
                    "workbook": {"cells": ["Model!A1"]},
                },
                {
                    "id": "ebitda-margin",
                    "name": "EBITDA margin",
                    "value": 12.5,
                    "unit": "percent",
                    "workbook": {"cells": ["Model!A2"]},
                },
            ]
        }), encoding="utf-8")
        return spec

    def _workbook(self, path, cells):
        book = self.openpyxl.Workbook()
        sheet = book.active
        sheet.title = "Model"
        for coordinate, value in cells.items():
            sheet[coordinate] = value
        book.save(path)
        return path

    def _run(self, argv):
        stdout = io.StringIO()
        with redirect_stdout(stdout):
            code = xl_variable_mcp.main(argv)
        return code, json.loads(stdout.getvalue())

    def test_actions_rendering_kinds_ordering_and_report_hash(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            spec = self._spec(root)
            workbook = self._workbook(root / "0999-inputs.xlsx", {
                "A1": 42,      # masked, raw
                "A2": 0.125,   # masked, raw
                "B1": 42,      # raw-number duplicate       -> mask
                "C1": "42 MW",                # formatted text -> extra_cell
                "B2": "12.5%",                # percent text   -> extra_cell
                "D1": "42 MW committed under the offtake contract",
            })
            report_path = root / "leakscan_report.json"
            patch_path = root / "leak_patch.json"
            code, _report = self._run([
                "leakscan", str(spec),
                "--inputs", str(workbook),
                "--report", str(report_path),
                "--emit-patch", str(patch_path),
            ])
            self.assertEqual(code, 1)
            patch = json.loads(patch_path.read_text(encoding="utf-8"))
            self.assertEqual(
                set(patch),
                {"workbook", "generated_at", "source_report_sha256",
                 "proposals"},
            )
            self.assertEqual(
                patch["source_report_sha256"],
                hashlib.sha256(report_path.read_bytes()).hexdigest(),
            )
            proposals = {
                (row["variable_id"], row["cell"]):
                    (row["action"], row["rendering"])
                for row in patch["proposals"]
            }
            self.assertEqual(proposals, {
                ("asset-capacity", "Model!B1"): ("mask", "raw_number"),
                ("asset-capacity", "Model!C1"):
                    ("extra_cell", "formatted_text"),
                ("asset-capacity", "Model!D1"):
                    ("exclude", "formatted_text"),
                ("ebitda-margin", "Model!B2"):
                    ("extra_cell", "percent_rendering"),
            })
            self.assertEqual(
                [(row["variable_id"], row["cell"])
                 for row in patch["proposals"]],
                sorted((row["variable_id"], row["cell"])
                       for row in patch["proposals"]),
            )
            by_key = {(row["variable_id"], row["cell"]): row
                      for row in patch["proposals"]}
            extra = by_key[("asset-capacity", "Model!C1")]["reason"]
            self.assertIn("formatted_text", extra)
            self.assertIn("Model!A1", extra)
            mixed = by_key[("asset-capacity", "Model!D1")]["reason"]
            self.assertIn("other required content", mixed)
            self.assertIn("draft chain", mixed)

    def test_allowlisted_leak_gets_no_proposal(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            spec = self._spec(root)
            workbook = self._workbook(root / "0999-inputs.xlsx", {
                "A1": 42,
                "A2": 0.125,
                "B1": 42,
            })
            allowlist = root / "allowlist.json"
            allowlist.write_text(json.dumps({
                "workbook_value_leaks": [{
                    "cell": "Model!B1",
                    "variable_id": "asset-capacity",
                    "reason": "required period-axis label",
                }]
            }), encoding="utf-8")
            patch_path = root / "leak_patch.json"
            code, report = self._run([
                "leakscan", str(spec),
                "--inputs", str(workbook),
                "--allowlist", str(allowlist),
                "--emit-patch", str(patch_path),
            ])
            self.assertEqual(code, 0)
            self.assertTrue(report["valid"], report)
            patch = json.loads(patch_path.read_text(encoding="utf-8"))
            self.assertEqual(patch["proposals"], [])

    def test_two_runs_write_byte_identical_patches(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            spec = self._spec(root)
            workbook = self._workbook(root / "0999-inputs.xlsx", {
                "A1": 42,
                "A2": 0.125,
                "B1": 42,
                "C1": "42 MW",
            })
            first = root / "patch-one.json"
            second = root / "patch-two.json"
            for path in (first, second):
                code, _report = self._run([
                    "leakscan", str(spec),
                    "--inputs", str(workbook),
                    "--emit-patch", str(path),
                ])
                self.assertEqual(code, 1)
            self.assertEqual(first.read_bytes(), second.read_bytes())

    def test_emit_patch_works_in_mcp_manifest_mode(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            spec = self._spec(root)
            mcp = root / "mcp"
            mcp.mkdir()
            (mcp / "mask_cells.json").write_text(
                json.dumps(["Model!A1"]), encoding="utf-8"
            )
            (mcp / "masked_inputs.json").write_text(json.dumps([{
                "variable_id": "asset-capacity",
                "refs": ["Model!A1"],
                "cell_value": 42,
                "mcp_value": 42,
                "unit": "MW",
            }]), encoding="utf-8")
            workbook = self._workbook(root / "0999-inputs.xlsx", {
                "A1": 42,
                "B1": 42,
            })
            patch_path = root / "leak_patch.json"
            code, _report = self._run([
                "leakscan", str(spec),
                "--mcp", str(mcp),
                "--inputs", str(workbook),
                "--emit-patch", str(patch_path),
            ])
            self.assertEqual(code, 1)
            patch = json.loads(patch_path.read_text(encoding="utf-8"))
            self.assertEqual(len(patch["proposals"]), 1)
            self.assertEqual(patch["proposals"][0]["action"], "mask")
            self.assertEqual(patch["proposals"][0]["cell"], "Model!B1")

    def test_shared_raw_cell_masks_one_variable_and_excludes_the_rest(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            workbook = self._workbook(root / "golden.xlsx", {
                "A1": 0.1,
                "A2": 0.1,
                "D2": 0.1,   # shared duplicate of both raw values
            })
            audit = [
                {
                    "variable_id": "cost-of-debt",
                    "refs": ["Model!A1"],
                    "cell_value": 0.1,
                    "mcp_value": 10,
                    "unit": "percent",
                },
                {
                    "variable_id": "market-return",
                    "refs": ["Model!A2"],
                    "cell_value": 0.1,
                    "mcp_value": 10,
                    "unit": "percent",
                },
            ]
            report = scan_value_leaks(
                workbook, ["Model!A1", "Model!A2"], audit, None,
                require_blank_mask=False,
            )
            patch = build_leak_patch(report, audit, workbook)
            actions = {
                row["variable_id"]: row["action"]
                for row in patch["proposals"]
            }
            self.assertEqual(actions, {
                "cost-of-debt": "mask",
                "market-return": "exclude",
            })
            excluded = next(row for row in patch["proposals"]
                            if row["action"] == "exclude")
            self.assertIn("cost-of-debt", excluded["reason"])
            self.assertIn("draft chain", excluded["reason"])

    def test_formula_cells_are_never_mask_proposals(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            workbook = self._workbook(root / "golden.xlsx", {
                "A1": "answer text",
                "B1": "answer text",   # typed duplicate of a string value
                "C1": '="answer" & " text"',
            })
            audit = [{
                "variable_id": "credit-rating",
                "refs": ["Model!A1"],
                "cell_value": "answer text",
                "mcp_value": "answer text",
                "unit": "text",
            }]
            report = scan_value_leaks(
                workbook, ["Model!A1"], audit, None,
                require_blank_mask=False,
            )
            patch = build_leak_patch(report, audit, workbook)
            by_cell = {row["cell"]: row for row in patch["proposals"]}
            self.assertEqual(by_cell["Model!B1"]["action"], "mask")
            if "Model!C1" in by_cell:
                self.assertEqual(by_cell["Model!C1"]["action"], "exclude")

    def test_draft_chain_recovers_the_normalizer_draft_id(self):
        self.assertEqual(
            draft_chain(
                "depreciation-and-amortization-projected-is-b19",
                ["Projected IS!B19"],
            ),
            "depreciation-and-amortization",
        )
        self.assertEqual(
            draft_chain("cost-of-debt", ["WACC!C4"]), "cost-of-debt"
        )


class LeakscanCliTests(unittest.TestCase):
    def setUp(self):
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl is not installed")
        self.openpyxl = openpyxl

    def _spec(self, root):
        spec = root / "normalized.json"
        spec.write_text(json.dumps({
            "variables": [{
                "id": "asset-capacity",
                "name": "Asset capacity",
                "value": 42,
                "unit": "MW",
                "workbook": {"cells": ["Model!A1"]},
            }]
        }), encoding="utf-8")
        return spec

    def _workbook(self, path, b1):
        book = self.openpyxl.Workbook()
        sheet = book.active
        sheet.title = "Model"
        sheet["A1"] = 42
        sheet["B1"] = b1
        book.save(path)
        return path

    def _run(self, argv):
        stdout = io.StringIO()
        with redirect_stdout(stdout):
            code = xl_variable_mcp.main(argv)
        return code, json.loads(stdout.getvalue())

    def test_exits_nonzero_on_a_leak_and_writes_the_report(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            spec = self._spec(root)
            workbook = self._workbook(root / "0999-inputs.xlsx", "42 MW")
            report_path = root / "leakscan_report.json"
            code, report = self._run([
                "leakscan", str(spec),
                "--inputs", str(workbook),
                "--report", str(report_path),
            ])
            self.assertEqual(code, 1)
            self.assertFalse(report["valid"])
            self.assertEqual(
                report["unapproved_value_leaks"][0]["cell"], "Model!B1"
            )
            written = json.loads(report_path.read_text(encoding="utf-8"))
            self.assertFalse(written["valid"])

    def test_exits_zero_on_a_clean_workbook(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            spec = self._spec(root)
            workbook = self._workbook(
                root / "0999-inputs.xlsx", "forty-two units of context"
            )
            code, report = self._run([
                "leakscan", str(spec), "--inputs", str(workbook),
            ])
            self.assertEqual(code, 0)
            self.assertTrue(report["valid"], report)
            self.assertEqual(report["unapproved_value_leaks"], [])

    def test_mcp_dir_manifests_and_require_blank_scan_a_masked_workbook(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            spec = self._spec(root)
            mcp = root / "mcp"
            mcp.mkdir()
            (mcp / "mask_cells.json").write_text(
                json.dumps(["Model!A1"]), encoding="utf-8"
            )
            (mcp / "masked_inputs.json").write_text(json.dumps([{
                "variable_id": "asset-capacity",
                "refs": ["Model!A1"],
                "cell_value": 42,
                "mcp_value": 42,
                "unit": "MW",
            }]), encoding="utf-8")
            book = self.openpyxl.Workbook()
            sheet = book.active
            sheet.title = "Model"
            sheet["B1"] = "no leak here"
            masked = root / "0999-inputs.xlsx"
            book.save(masked)
            code, report = self._run([
                "leakscan", str(spec),
                "--mcp", str(mcp),
                "--inputs", str(masked),
                "--require-blank",
            ])
            self.assertEqual(code, 0)
            self.assertTrue(report["valid"], report)


if __name__ == "__main__":
    unittest.main()
