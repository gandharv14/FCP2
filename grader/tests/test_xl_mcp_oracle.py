from __future__ import annotations

import asyncio
import json
import tempfile
import unittest
from pathlib import Path
from types import SimpleNamespace

from xl_mcp_oracle import (
    ValueTarget,
    build_parser,
    check_environment,
    check_mcp,
    check_profile_excerpts,
    check_rendered_profile_excerpts,
    check_workbook,
    load_allowlist,
    query_all,
    validate_task_rows,
    value_matches_target,
)


def _task():
    return {
        "task_id": "asset-capacity",
        "answer": {"value": 42, "unit": "MW", "tolerance": 0},
        "evidence": {
            "dataset_id": "dataset-1",
            "document_id": "doc-2",
            "record_id": "row-2",
            "release": "rel-2",
            "source_id": "source-1",
        },
        "required_dimensions": {
            "metric": "Asset capacity",
            "entity": "Asset",
            "period": "2026",
            "scenario": "All cases",
            "basis": "nameplate",
            "unit": "MW",
            "status": "final",
        },
    }


def _row(row_id, release, value, document_id, superseded_by=None, **changes):
    release_dates = {
        "rel-0": "2024-01-01",
        "rel-1": "2025-01-01",
        "rel-2": "2026-01-01",
        "rel-alt": "2025-06-01",
    }
    row = {
        "id": row_id,
        "dataset_id": "dataset-1",
        "document_id": document_id,
        "source_id": "source-1",
        "metric": "Asset capacity",
        "entity": "Asset",
        "period": "2026",
        "scenario": "All cases",
        "basis": "nameplate",
        "unit": "MW",
        "status": "final",
        "release": release,
        "published_at": release_dates[release],
        "superseded_by": superseded_by,
        "value": value,
    }
    row.update(changes)
    return row


class FakeClient:
    def __init__(self, _url):
        self.exact = [
            _row("row-0", "rel-0", 40, "doc-0", "rel-1"),
            _row("row-1", "rel-1", 41, "doc-1", "rel-2"),
            _row("row-2", "rel-2", 42, "doc-2"),
        ]
        self.broad = self.exact + [
            _row(
                "row-alt",
                "rel-alt",
                99,
                "doc-alt",
                period="2025",
            )
        ]
        self.documents = {
            "doc-0": {
                "id": "doc-0",
                "source_id": "source-1",
                "related_dataset_id": "dataset-1",
                "content": "release rel-0 was superseded by rel-1",
            },
            "doc-1": {
                "id": "doc-1",
                "source_id": "source-1",
                "related_dataset_id": "dataset-1",
                "content": "release rel-1 was superseded by rel-2",
            },
            "doc-2": {
                "id": "doc-2",
                "source_id": "source-1",
                "related_dataset_id": "dataset-1",
                "content": "release rel-2 supersedes rel-1",
            },
        }

    async def __aenter__(self):
        return self

    async def __aexit__(self, *_args):
        return False

    async def call_tool(self, name, arguments):
        if name == "list_sources":
            return SimpleNamespace(data=[{
                "id": "source-1",
                "name": "Source One",
                "origin_url": "https://example.test/source",
            }])
        if name == "list_datasets":
            return SimpleNamespace(data=[{"id": "dataset-1"}])
        if name == "fetch_document":
            return SimpleNamespace(data=self.documents[arguments["document_id"]])
        if name != "query_records":
            raise AssertionError(name)
        rows = self.exact if "period" in arguments else self.broad
        start = int(arguments["cursor"])
        limit = int(arguments["limit"])
        page = rows[start:start + limit]
        next_cursor = start + len(page)
        return SimpleNamespace(data={
            "rows": page,
            "returned": len(page),
            "total_matches": len(rows),
            "next_cursor": next_cursor if next_cursor < len(rows) else None,
        })


class PaginationAndMcpTests(unittest.TestCase):
    def test_cli_accepts_json_report_path(self):
        args = build_parser().parse_args([
            "--bundle", "task", "--mcp", "authoring",
            "--url", "http://127.0.0.1:8000/mcp",
            "--report", "oracle-report.json",
        ])
        self.assertEqual(args.report, Path("oracle-report.json"))

    def test_live_check_paginates_and_validates_chains_and_conflicts(self):
        target = ValueTarget("asset-capacity", 42, 42, "MW")
        report = asyncio.run(check_mcp(
            "http://example.test/mcp",
            [_task()],
            [target],
            [{
                "id": "source-1",
                "name": "Source One",
                "origin_url": "https://example.test/source",
            }],
            [{"id": "dataset-1"}],
            page_size=2,
            client_factory=FakeClient,
        ))
        self.assertTrue(report["valid"], report)
        self.assertEqual(report["exact_resolutions"], 1)
        self.assertEqual(report["provenance_chains"], 1)
        self.assertEqual(report["broad_queries_with_conflicts"], 1)
        self.assertEqual(report["pages_read"], 4)

    def test_query_all_rejects_repeated_cursor(self):
        class RepeatingClient:
            async def call_tool(self, _name, _arguments):
                return SimpleNamespace(data={
                    "rows": [{"id": "row"}],
                    "returned": 1,
                    "total_matches": 2,
                    "next_cursor": 0,
                })

        with self.assertRaisesRegex(RuntimeError, "cursor repeated"):
            asyncio.run(query_all(
                RepeatingClient(), {"dataset_id": "d", "metric": "m"},
            ))

    def test_exact_rows_require_a_provenance_chain(self):
        errors, _details = validate_task_rows(
            _task(), [_row("row-2", "rel-2", 42, "doc-2")]
        )
        self.assertIn("no stale provenance release was returned", errors)


class WorkbookLeakTests(unittest.TestCase):
    def setUp(self):
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl is not installed")
        self.openpyxl = openpyxl

    def _workbook(self, root):
        path = root / "model-inputs.xlsx"
        book = self.openpyxl.Workbook()
        sheet = book.active
        sheet.title = "Model"
        sheet["A1"] = None
        sheet["B1"] = "42 MW"
        book.save(path)
        return path

    def _audit(self):
        return [{
            "variable_id": "asset-capacity",
            "refs": ["Model!A1"],
            "cell_value": 42,
            "mcp_value": 42,
            "unit": "MW",
        }]

    def test_unapproved_representation_fails_and_explicit_allowlist_passes(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            workbook = self._workbook(root)
            failed = check_workbook(
                workbook, ["Model!A1"], self._audit(), {}
            )
            self.assertFalse(failed["valid"])
            self.assertEqual(
                failed["unapproved_value_leaks"][0]["cell"], "Model!B1"
            )

            allowlist_path = root / "allowlist.json"
            allowlist_path.write_text(json.dumps({
                "workbook_value_leaks": [{
                    "cell": "Model!B1",
                    "variable_id": "asset-capacity",
                    "reason": "required label",
                }]
            }), encoding="utf-8")
            allowed, errors = load_allowlist(
                allowlist_path, {"asset-capacity"}
            )
            self.assertEqual(errors, [])
            passed = check_workbook(
                workbook, ["Model!A1"], self._audit(), allowed
            )
            self.assertTrue(passed["valid"], passed)
            self.assertEqual(len(passed["approved_value_leaks"]), 1)

    def test_unused_allowlist_entry_fails(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            workbook = self._workbook(root)
            report = check_workbook(
                workbook,
                ["Model!A1"],
                self._audit(),
                {("Model!C1", "asset-capacity"): "expected label"},
            )
            self.assertFalse(report["valid"])
            self.assertEqual(len(report["unused_allowlist_entries"]), 1)

    def test_no_sheet_is_silently_excluded_from_leak_scanning(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            workbook = self._workbook(root)
            book = self.openpyxl.load_workbook(workbook)
            assumptions = book.create_sheet("Embedded Assumptions")
            assumptions["A1"] = 42
            book.save(workbook)
            report = check_workbook(
                workbook, ["Model!A1"], self._audit(), {}
            )
            leaked_cells = {
                item["cell"] for item in report["unapproved_value_leaks"]
            }
            self.assertIn("'Embedded Assumptions'!A1", leaked_cells)

    def test_common_numeric_display_forms_are_detected(self):
        target = ValueTarget("asset-capacity", 42, 42, "MW")
        self.assertTrue(value_matches_target("42 MW", target))
        self.assertTrue(value_matches_target("USD 42.00 million", target))
        self.assertFalse(value_matches_target("142 MW", target))


class IsolationAndProfileTests(unittest.TestCase):
    def _environment(self, root):
        environment = root / "environment"
        runtime = environment / "mcp-server" / "runtime"
        runtime.mkdir(parents=True)
        workbook = environment / "model-inputs.xlsx"
        workbook.touch()
        for relative in (
            "Dockerfile",
            "docker-compose.yaml",
            "mcp-server/Dockerfile",
            "mcp-server/server.py",
        ):
            (environment / relative).write_text("", encoding="utf-8")
        (runtime / "server.json").write_text("{}", encoding="utf-8")
        (runtime / "sources.json").write_text("[]", encoding="utf-8")
        (runtime / "datasets.json").write_text("[]", encoding="utf-8")
        (runtime / "documents.jsonl").write_text("", encoding="utf-8")
        (runtime / "records.jsonl").write_text("", encoding="utf-8")
        return workbook

    def test_unknown_environment_file_is_a_leak(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            workbook = self._environment(root)
            clean = check_environment(root, workbook)
            self.assertTrue(clean["valid"], clean)
            (root / "environment" / "normalized.json").write_text(
                "{}", encoding="utf-8"
            )
            leaked = check_environment(root, workbook)
            self.assertFalse(leaked["valid"])
            self.assertIn("normalized.json", leaked["unknown_files"])

    def test_dialogue_notes_allowed_after_apply(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            workbook = self._environment(root)
            (root / "environment" / "Dockerfile").write_text(
                "COPY additional-assumptions.md /app/additional-assumptions.md\n",
                encoding="utf-8",
            )
            (root / "environment" / "additional-assumptions.md").write_text(
                "notes\n", encoding="utf-8"
            )
            (root / "tests").mkdir()
            (root / "tests" / "dialogue-applied.json").write_text(
                "{}\n", encoding="utf-8"
            )
            report = check_environment(root, workbook)
            self.assertTrue(report["valid"], report)

    def test_stray_dialogue_notes_are_a_leak(self):
        with tempfile.TemporaryDirectory() as temp:
            root = Path(temp)
            workbook = self._environment(root)
            (root / "environment" / "additional-assumptions.md").write_text(
                "notes\n", encoding="utf-8"
            )
            report = check_environment(root, workbook)
            self.assertFalse(report["valid"])
            self.assertIn("additional-assumptions.md", report["unknown_files"])

    def test_golden_eval_and_source_profile_snapshots_are_not_shippable(self):
        extras = (
            "golden.xlsx",
            "mcp-server/eval/tasks.jsonl",
            "source_profiles.json",
        )
        for extra in extras:
            with self.subTest(extra=extra), tempfile.TemporaryDirectory() as temp:
                root = Path(temp)
                workbook = self._environment(root)
                leaked_path = root / "environment" / extra
                leaked_path.parent.mkdir(parents=True, exist_ok=True)
                leaked_path.write_text("{}", encoding="utf-8")
                report = check_environment(root, workbook)
                self.assertFalse(report["valid"])
                self.assertIn(extra, report["unknown_files"])

    def test_profile_excerpts_are_attributed_and_value_safe(self):
        target = ValueTarget("asset-capacity", 42, 42, "MW")
        source = {
            "id": "source-1",
            "name": "Source One",
            "origin_url": "https://example.test/source",
            "profile_excerpt": "Independent infrastructure research profile.",
        }
        self.assertTrue(check_profile_excerpts([source], [target])["valid"])

        source["profile_excerpt"] = {
            "text": "The profile reports 42 MW.",
            "source_id": "wrong-source",
        }
        report = check_profile_excerpts([source], [target])
        self.assertFalse(report["valid"])
        self.assertTrue(any("inconsistent" in item for item in report["violations"]))
        self.assertTrue(any("masked values" in item for item in report["violations"]))

    def test_rendered_document_excerpts_follow_profile_convention(self):
        target = ValueTarget("asset-capacity", 42, 42, "MW")
        sources = [{
            "id": "source-1",
            "name": "Source One",
            "origin_url": "https://example.test/source",
            "profile_id": "profile-1",
        }]
        datasets = [{
            "id": "dataset-1",
            "source_id": "source-1",
            "profile_id": "profile-1",
        }]
        documents = [{
            "id": "doc-1",
            "source_id": "source-1",
            "related_dataset_id": "dataset-1",
            "content": (
                "Synthetic record.\n\nShort attributed excerpt — Source One "
                "(release guide): “Uses a monthly release calendar.”"
            ),
        }]
        report = check_rendered_profile_excerpts(
            sources, datasets, documents, [target]
        )
        self.assertTrue(report["valid"], report)
        self.assertEqual(report["count"], 1)

        documents[0]["content"] = (
            "Short attributed excerpt — Source One "
            "(release guide): “The observed value is 42 MW.”"
        )
        report = check_rendered_profile_excerpts(
            sources, datasets, documents, [target]
        )
        self.assertFalse(report["valid"])
        self.assertIn("masked values", report["violations"][0])


if __name__ == "__main__":
    unittest.main()
