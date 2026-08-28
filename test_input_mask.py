import tempfile
import unittest
from pathlib import Path

import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

from xl_input_mask import (
    MASKED_VALUE_NOTE,
    _assumptions_sheet_xml,
    redact_assumptions,
    verify,
)
from xl_level_split import scrub_chart_caches


def entry(**overrides):
    base = {
        "sheet": "Model",
        "host": "Model!B2",
        "label": "Tax rate",
        "value": 0.25,
        "formula": "=Revenue*0.25",
    }
    base.update(overrides)
    return base


class RedactAssumptionsTests(unittest.TestCase):
    def test_masked_value_is_withheld_from_the_sheet(self):
        rows, redacted, stripped = redact_assumptions(
            [entry()], deny={}, deny_numbers={0.25}, deny_texts=set(),
            outputs=set())
        self.assertEqual(rows[0]["value"], MASKED_VALUE_NOTE)
        self.assertEqual(rows[0]["formula"], "")
        self.assertEqual(redacted, ["Model!B2"])
        self.assertEqual(stripped, [])

    def test_denied_host_cell_is_withheld(self):
        rows, redacted, _ = redact_assumptions(
            [entry()], deny={"Model": {(2, 2)}}, deny_numbers=set(),
            deny_texts=set(), outputs=set())
        self.assertEqual(rows[0]["value"], MASKED_VALUE_NOTE)
        self.assertEqual(rows[0]["formula"], "")
        self.assertEqual(redacted, ["Model!B2"])

    def test_output_host_loses_its_formula_but_keeps_the_constant(self):
        rows, redacted, stripped = redact_assumptions(
            [entry(host="Model!B2",
                   formula='=IFERROR(IRR(B18:F18,0.1),"NA")', value=0.1)],
            deny={}, deny_numbers=set(), deny_texts=set(),
            outputs={("Model", 2, 2)})
        self.assertEqual(rows[0]["value"], 0.1)
        self.assertEqual(rows[0]["formula"], "")
        self.assertEqual(redacted, [])
        self.assertEqual(stripped, ["Model!B2"])

    def test_formula_containing_a_masked_literal_is_withheld(self):
        rows, redacted, stripped = redact_assumptions(
            [entry(value=2100.0, formula="=102.37+2100")],
            deny={}, deny_numbers={102.37}, deny_texts=set(), outputs=set())
        self.assertEqual(rows[0]["value"], 2100.0)
        self.assertEqual(rows[0]["formula"], "")
        self.assertEqual(stripped, ["Model!B2"])

    def test_benign_entry_is_untouched(self):
        rows, redacted, stripped = redact_assumptions(
            [entry()], deny={}, deny_numbers={99.5}, deny_texts=set(),
            outputs={("Model", 9, 9)})
        self.assertEqual(rows[0]["value"], 0.25)
        self.assertEqual(rows[0]["formula"], "=Revenue*0.25")
        self.assertEqual(redacted, [])
        self.assertEqual(stripped, [])


class AssumptionsSheetXmlTests(unittest.TestCase):
    def test_formula_context_is_stored_as_text(self):
        xml = _assumptions_sheet_xml([entry()]).decode("utf-8")
        self.assertIn("'=Revenue*0.25", xml)

    def test_redacted_rows_render_the_note_and_no_formula(self):
        rows, _, _ = redact_assumptions(
            [entry()], deny={}, deny_numbers={0.25}, deny_texts=set(),
            outputs=set())
        xml = _assumptions_sheet_xml(rows).decode("utf-8")
        self.assertIn(MASKED_VALUE_NOTE, xml)
        self.assertNotIn("0.25", xml)
        self.assertNotIn("Revenue", xml)


class ChartCacheScrubTests(unittest.TestCase):
    CHART = (
        b'<c:chartSpace xmlns:c="http://schemas.openxmlformats.org/'
        b'drawingml/2006/chart"><c:ser>'
        b"<c:cat><c:strRef><c:f>FF!$B$9:$B$17</c:f>"
        b"<c:strCache><c:ptCount val=\"1\"/>"
        b"<c:pt idx=\"0\"><c:v>Revenue</c:v></c:pt></c:strCache>"
        b"</c:strRef></c:cat>"
        b"<c:val><c:numRef><c:f>FF!$J$9:$J$17</c:f>"
        b"<c:numCache><c:formatCode>General</c:formatCode>"
        b"<c:ptCount val=\"1\"/>"
        b"<c:pt idx=\"0\"><c:v>827.07855951256045</c:v></c:pt></c:numCache>"
        b"</c:numRef></c:val>"
        b"</c:ser></c:chartSpace>"
    )

    def test_cached_values_are_removed_and_references_survive(self):
        data, hits = scrub_chart_caches("xl/charts/chart1.xml", self.CHART)
        self.assertEqual(hits, 2)
        self.assertNotIn(b"827.07855951256045", data)
        self.assertNotIn(b"numCache", data)
        self.assertNotIn(b"strCache", data)
        self.assertIn(b"<c:f>FF!$J$9:$J$17</c:f>", data)

    def test_scrub_is_idempotent(self):
        data, _ = scrub_chart_caches("xl/charts/chart1.xml", self.CHART)
        again, hits = scrub_chart_caches("xl/charts/chart1.xml", data)
        self.assertEqual(hits, 0)
        self.assertEqual(again, data)

    def test_chartex_points_are_removed(self):
        blob = (
            b"<cx:numDim type=\"val\"><cx:f>Sheet1!$B$2:$B$5</cx:f>"
            b"<cx:lvl ptCount=\"2\">"
            b"<cx:pt idx=\"0\">1234.5</cx:pt><cx:pt idx=\"1\">6789.0</cx:pt>"
            b"</cx:lvl></cx:numDim>"
        )
        data, hits = scrub_chart_caches("xl/charts/chartEx1.xml", blob)
        self.assertEqual(hits, 2)
        self.assertNotIn(b"1234.5", data)
        self.assertIn(b"<cx:f>Sheet1!$B$2:$B$5</cx:f>", data)


class VerifyFormulaCensusTests(unittest.TestCase):
    """Workbook 0642 regression: typed text that merely looks like a formula
    (a literal "=" dashboard label) must not be reported as a survivor, while
    every genuinely stored formula -- string, shared, or array -- must be."""

    def _book(self, build):
        book = openpyxl.Workbook()
        build(book.active)
        path = Path(self._dir.name) / ("case%d.xlsx" % self._counter)
        self._counter += 1
        book.save(path)
        return path

    def setUp(self):
        self._dir = tempfile.TemporaryDirectory()
        self._counter = 0

    def tearDown(self):
        self._dir.cleanup()

    def _verify(self, path, derived=None):
        return verify(
            path, path, keep={}, frontier={},
            formula_coords=derived or {}, deny={},
        )

    def test_typed_equals_label_is_not_a_formula_fault(self):
        def build(ws):
            cell = ws.cell(row=12, column=13)
            cell.value = "="
            cell.data_type = "s"
        path = self._book(build)
        faults = self._verify(path)
        self.assertEqual(faults["formula"], [])

    def test_formula_like_text_on_a_derived_cell_is_refused(self):
        def build(ws):
            cell = ws.cell(row=2, column=1)
            cell.value = "=A1*2"
            cell.data_type = "s"
        path = self._book(build)
        faults = self._verify(path, derived={"Sheet": {(2, 1)}})
        self.assertTrue(any("formula-like derived text" in f
                            for f in faults["formula"]))

    def test_real_formula_is_caught_by_cell_and_census(self):
        def build(ws):
            ws["A1"] = "=1+1"
        path = self._book(build)
        faults = self._verify(path)
        self.assertTrue(any(f.startswith("Sheet!A1") for f in faults["formula"]))
        self.assertTrue(any("<f> element survived" in f
                            for f in faults["formula"]))

    def test_array_formula_is_caught(self):
        def build(ws):
            ws["A1"] = ArrayFormula("A1", "=SUM(1)")
        path = self._book(build)
        faults = self._verify(path)
        self.assertTrue(faults["formula"])


if __name__ == "__main__":
    unittest.main()
