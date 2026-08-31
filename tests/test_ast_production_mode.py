"""CSV-only Harbor production output leaves the graph and downstream results unchanged."""

from __future__ import annotations

import hashlib
import json
import re
import shutil
import sys
import zipfile
from argparse import Namespace
from pathlib import Path
from types import SimpleNamespace

import openpyxl
import pytest

import xl_ast_graph
import xl_input_mask
import xl_segment
from xl_seg import emit, publication


DEBUG_ARTIFACTS = ("graph.json", "graph.graphml", "graph.html")
REPO = Path(__file__).resolve().parents[1]
DISCLOSE = REPO / ".cursor/skills/task-disclosure/scripts"
WB = "prod01"


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    digest.update(path.read_bytes())
    return digest.hexdigest()


def _write_small_workbook(path: Path) -> None:
    book = openpyxl.Workbook()
    ws = book.active
    ws.title = "Summary"
    ws["A1"] = "Revenue"
    ws["B1"] = 100
    ws["A2"] = "Growth"
    ws["B2"] = 0.1
    ws["A3"] = "IRR"
    ws["B3"] = "=B1*(1+B2)"
    book.save(path)
    _inject_cached_value(path, "B3", "110")


def _inject_cached_value(path: Path, coord: str, value: str) -> None:
    with zipfile.ZipFile(path) as src:
        parts = {item.filename: src.read(item.filename) for item in src.infolist()}
    sheet = "xl/worksheets/sheet1.xml"
    xml = parts[sheet]
    needle = f'r="{coord}"'.encode()
    attr = xml.find(needle)
    assert attr != -1, f"{coord} missing from {sheet}"
    start = xml.rfind(b"<c", 0, attr)
    close = xml.find(b"</c>", attr)
    assert start != -1 and close != -1
    end = close + len(b"</c>")
    cell = xml[start:end]
    new_v = f"<v>{value}</v>".encode()
    if re.search(rb"<v\b", cell):
        cell = re.sub(rb"<v\b[^/]*/>|<v>[^<]*</v>", new_v, cell, count=1)
    else:
        cell = cell.replace(b"</c>", new_v + b"</c>", 1)
    parts[sheet] = xml[:start] + cell + xml[end:]
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as out:
        for name, data in parts.items():
            out.writestr(name, data)


def _blank_formula(src: Path, dest: Path, coord: str) -> None:
    shutil.copy2(src, dest)
    book = openpyxl.load_workbook(dest)
    ws = book.active
    ws[coord] = None
    book.save(dest)


def _build_ast(workbook: Path, out_dir: Path, *flags: str) -> None:
    xl_ast_graph.main([str(workbook), "-o", str(out_dir), "-q", *flags])


def _segment(wb: str, ast_root: Path, source: Path, seg_root: Path, env_file: Path):
    args = SimpleNamespace(
        ast_dir=str(ast_root),
        source=str(source),
        out=str(seg_root),
        threshold=6.0,
        top=40,
        sample=400,
        lineage_max=4000,
        recurate=False,
        llm=False,
        model="unused",
        env_file=str(env_file),
        no_verify=False,
    )
    return xl_segment.segment(wb, args)


def _select(
    task_dir: Path,
    golden: Path,
    ast_dir: Path,
    seg_root: Path | None = None,
    expected_generation_id: str | None = None,
) -> dict:
    sys.path.insert(0, str(DISCLOSE))
    try:
        import disclose
    finally:
        if str(DISCLOSE) in sys.path:
            sys.path.remove(str(DISCLOSE))
    args = Namespace(
        task_dir=str(task_dir),
        golden=str(golden),
        ast_dir=str(ast_dir),
        seg_root=str(seg_root) if seg_root is not None else None,
        segmentation_mode="strict",
        expected_generation_id=expected_generation_id,
        out=None,
        runs_root=str(task_dir / "runs"),
    )
    return disclose.select_payload(args)


@pytest.fixture
def small_model(tmp_path: Path) -> Path:
    source = tmp_path / "source"
    source.mkdir()
    workbook = source / f"{WB}.xlsx"
    _write_small_workbook(workbook)
    return workbook


def test_production_csv_matches_default_and_skips_debug_artifacts(small_model, tmp_path):
    source = small_model.parent
    default_root = tmp_path / "ast_default"
    production_root = tmp_path / "ast_production"
    _build_ast(small_model, default_root)
    _build_ast(small_model, production_root, "--production")

    default_dir = default_root / WB
    production_dir = production_root / WB
    assert _sha256(default_dir / "nodes.csv") == _sha256(production_dir / "nodes.csv")
    assert _sha256(default_dir / "edges.csv") == _sha256(production_dir / "edges.csv")
    assert sorted(path.name for path in production_dir.iterdir()) == [
        "edges.csv",
        "nodes.csv",
    ]
    for name in DEBUG_ARTIFACTS:
        assert not (production_dir / name).exists()
        assert (default_dir / name).is_file()

    default_seg = tmp_path / "seg_default"
    production_seg = tmp_path / "seg_production"
    env_file = tmp_path / "no.env"
    default_payload = _segment(WB, default_root, source, default_seg, env_file)
    production_payload = _segment(WB, production_root, source, production_seg, env_file)
    assert production_payload["verification"] == default_payload["verification"]
    assert production_payload["verification"].get("passed") is True
    assert (
        json.loads((default_seg / WB / "segments.json").read_text(encoding="utf-8"))
        ["verification"]
        == json.loads((production_seg / WB / "segments.json").read_text(encoding="utf-8"))
        ["verification"]
    )
    assert emit.read_curation(default_seg / WB / "curation.toml")
    assert emit.read_curation(production_seg / WB / "curation.toml")
    generation, manifest = publication.resolve_current_generation(
        production_seg / WB,
        source_path=small_model,
        ast_dir=production_root / WB,
        require_pass=True,
        validate_live_evidence=True,
    )
    assert generation.name == manifest["generation_id"]
    assert (generation / "generation-manifest.json").is_file()

    inputs_root = tmp_path / "inputs"
    assert xl_input_mask.main([
        WB,
        "--source", str(source),
        "--seg-dir", str(production_seg),
        "--ast-dir", str(production_root),
        "--segmentation-mode", "strict",
        "-o", str(inputs_root),
    ]) == 0
    assert (inputs_root / f"{WB}-inputs.xlsx").is_file()
    assert (inputs_root / f"{WB}-inputs.segmentation.json").is_file()

    task_dir = tmp_path / f"{WB}-outputs"
    env = task_dir / "environment"
    tests = task_dir / "tests"
    env.mkdir(parents=True)
    tests.mkdir()
    delivered = env / f"{WB}-inputs.xlsx"
    _blank_formula(small_model, delivered, "B3")
    (tests / "answer_key.json").write_text(
        json.dumps({"targets": {"Summary!B3": 110}, "tolerance": {}}, indent=2),
        encoding="utf-8",
    )
    default_select = _select(
        task_dir,
        small_model,
        default_root,
        seg_root=default_seg,
    )
    production_select = _select(
        task_dir,
        small_model,
        production_root,
        seg_root=production_seg,
    )
    shutil.copy2(
        generation / "generation-manifest.json",
        tests / "segmentation_generation_manifest.json",
    )
    delivered_sidecar = publication.write_inputs_sidecar(
        delivered,
        generation,
        manifest,
    )
    shutil.copy2(delivered_sidecar, tests / "inputs_generation.json")
    strict_select = _select(
        task_dir,
        small_model,
        production_root,
        production_seg,
        manifest["generation_id"],
    )
    for payload in (default_select, production_select):
        assert payload["selection"]["closure_source"] == "ast"
        assert payload["selection"]["ast_status"] == "ok"
    assert production_select["selection"] == default_select["selection"]
    assert [band["cell_keys"] for band in production_select["bands"]] == [
        band["cell_keys"] for band in default_select["bands"]
    ]
    assert (
        strict_select["segmentation_generation"]["generation_id"]
        == manifest["generation_id"]
    )
    pointer_path = production_seg / WB / "current.json"
    pointer = json.loads(pointer_path.read_text(encoding="utf-8"))
    pointer["generation_id"] = "b" * 64
    pointer["generation_path"] = f"generations/{'b' * 64}"
    pointer_path.write_text(json.dumps(pointer), encoding="utf-8")
    with pytest.raises(
        publication.GenerationValidationError,
        match="changed from the expected",
    ):
        _select(
            task_dir,
            small_model,
            production_root,
            production_seg,
            manifest["generation_id"],
        )


def test_production_rebuild_removes_stale_debug_artifacts(small_model, tmp_path):
    ast_root = tmp_path / "ast"
    _build_ast(small_model, ast_root)
    assert (ast_root / WB / "graph.json").is_file()

    _build_ast(small_model, ast_root, "--production")

    assert sorted(path.name for path in (ast_root / WB).iterdir()) == [
        "edges.csv",
        "nodes.csv",
    ]


def test_build_script_pins_generation_and_retains_ast():
    script = (REPO / "build_one.sh").read_text(encoding="utf-8")

    assert '--expected-generation-id "$GENERATION_ID"' in script
    assert "validate_inputs_sidecar" in script
    assert 'rm -rf "ast_out/$WB"' not in script


def test_disclosure_migration_propagates_generation_gate(monkeypatch, tmp_path):
    sys.path.insert(0, str(DISCLOSE))
    try:
        import disclose
    finally:
        sys.path.remove(str(DISCLOSE))
    captured = {}
    monkeypatch.setattr(
        disclose,
        "cmd_select",
        lambda args: captured.update({
            "seg_root": args.seg_root,
            "mode": args.segmentation_mode,
            "generation_id": args.expected_generation_id,
        }),
    )
    for name in ("cmd_probe", "cmd_roles", "cmd_detect", "cmd_context"):
        monkeypatch.setattr(disclose, name, lambda args: None)
    args = SimpleNamespace(
        ast_dir="ast_out",
        seg_root="seg_out",
        segmentation_mode="strict",
        expected_generation_id="a" * 64,
        runs_root="runs",
        role_resolutions=None,
    )

    disclose.ensure_pipeline(tmp_path / "task", args)

    assert captured == {
        "seg_root": "seg_out",
        "mode": "strict",
        "generation_id": "a" * 64,
    }


def test_harbor_skill_uses_only_full_validator_for_pass():
    skill = (REPO / ".cursor/skills/create-harbor-task/SKILL.md").read_text(
        encoding="utf-8"
    )

    assert "Only the full validator command above establishes PASS" in skill
    assert 'print("segmentation generation"' not in skill


def test_harbor_skill_binds_restricted_routes_to_approved_inventory():
    skill = (REPO / ".cursor/skills/create-harbor-task/SKILL.md").read_text(
        encoding="utf-8"
    )

    assert '[ "$SOURCE_ROUTE" = "restricted_pass" ]' in skill
    assert '[ "$SOURCE_ROUTE" = "restricted_recalc_pass" ]' in skill
    assert "--inventory-batch-id" in skill
    assert "--inventory-approval-registry" in skill
    assert "commit-approved batch inventory" in skill
    assert "inactive source and AST candidate" in skill
