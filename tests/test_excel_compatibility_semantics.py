from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.workbook.defined_name import DefinedName

from xl_ast_graph import AstGraph, write_csv
from xl_seg.evaluate import (
    CalculationMetadata,
    Evaluator,
    Unresolved,
    _excel_date_parts,
    _rate_value,
    to_serial,
)
from xl_seg.model import load
from xl_seg.project import build as build_cell_graph
from xl_segment import stabilize_runtime_proof


def _evaluate(path: Path):
    built = AstGraph(path, verbose=False).build()
    out = path.parent / f"{path.stem}-ast"
    out.mkdir()
    write_csv(out, built.nodes, built.edges)
    graph = load(out, path.stem)
    cell_graph = build_cell_graph(graph)
    inputs = {node.id for node in graph.cells() if node.kind == "input"}
    return built, graph, Evaluator(graph, cell_graph).run(inputs)


def _strict_evaluate(path: Path, outputs: set[str]):
    built = AstGraph(path, verbose=False).build()
    out = path.parent / f"{path.stem}-strict-ast"
    out.mkdir()
    write_csv(out, built.nodes, built.edges)
    graph = load(out, path.stem)
    cell_graph = build_cell_graph(graph)
    declared = {node.id for node in graph.cells() if node.kind in ("input", "label")}
    result, inputs, _, proof = stabilize_runtime_proof(
        graph,
        cell_graph,
        declared,
        outputs,
        calculation=CalculationMetadata(available=True, iterate=False),
    )
    return graph, result, inputs, proof


def test_rate_without_payments_resolves_root_near_negative_one() -> None:
    expected = -0.9999999

    actual = _rate_value(2, 0, -1, 1e-14, 0, 0.1)

    assert abs(actual - expected) < 1e-12


def test_rate_without_payments_matches_excel_zero_future_value_limit() -> None:
    actual = _rate_value(2, 0, -337.7, 0, 0, 0.1)

    assert abs(actual - -0.9999998807237069) < 1e-7


def test_subtotal_excludes_nested_subtotal_lineage_but_sum_does_not(tmp_path):
    path = tmp_path / "subtotals.xlsx"
    book = openpyxl.Workbook()
    ws = book.active
    ws.title = "Totals"
    ws["A1"], ws["A2"] = 10, 20
    ws["A3"] = "=SUBTOTAL(9,A1:A2)"
    ws["A4"] = "=SUBTOTAL(109,A1:A3)"
    ws["A5"] = "=SUM(A1:A3)"
    ws["A6"] = "=AGGREGATE(9,0,A1:A2)"
    ws["A7"] = "=SUBTOTAL(9,A1:A6)"
    book.save(path)

    _, _, result = _evaluate(path)

    assert result.values["Totals!A3"] == 30
    assert result.values["Totals!A4"] == 30
    assert result.values["Totals!A5"] == 60
    # A3/A4 are nested subtotals and A6 is AGGREGATE lineage. Ordinary A5
    # remains eligible, proving this is provenance filtering rather than dedupe.
    assert result.values["Totals!A7"] == 90


def test_unary_plus_preserves_lookup_text_and_coerces_numeric_text(tmp_path):
    path = tmp_path / "unary.xlsx"
    book = openpyxl.Workbook()
    ws = book.active
    ws.title = "Keys"
    ws["A1"] = "ALPHA"
    ws["A2"] = "=+A1"
    ws["A3"] = '=MATCH(+A1,A5:A6,0)'
    ws["A4"] = '=+"42"'
    ws["A5"], ws["A6"] = "ALPHA", "BETA"
    book.save(path)

    _, _, result = _evaluate(path)

    assert result.values["Keys!A2"] == "ALPHA"
    assert result.values["Keys!A3"] == 1
    assert result.values["Keys!A4"] == 42


def test_strict_proof_admits_only_active_text_dependencies(tmp_path):
    path = tmp_path / "strict-text.xlsx"
    book = openpyxl.Workbook()
    ws = book.active
    ws.title = "Text"
    ws["A1"], ws["A2"], ws["A3"] = "N", "Y", "unused"
    ws["B1"], ws["B2"] = 10, 20
    ws["C1"] = '=SUMIF(A1:A2,"Y",B1:B2)'
    ws["C2"] = '=MATCH("Y",A1:A2,0)'
    book.save(path)

    _, result, inputs, proof = _strict_evaluate(path, {"Text!C1", "Text!C2"})

    assert result.values["Text!C1"] == 20
    assert result.values["Text!C2"] == 2
    assert {"Text!A1", "Text!A2"} <= inputs
    assert "Text!A3" not in inputs
    assert proof["closure"]["stabilized"] is True


def test_bad_sumif_criteria_values_propagate_in_strict_proof(tmp_path):
    path = tmp_path / "bad-criteria.xlsx"
    book = openpyxl.Workbook()
    ws = book.active
    ws.title = "Text"
    ws["A1"] = "=_xlfn.NOT_A_REAL_FUNCTION()"
    ws["A2"] = "Y"
    ws["B1"], ws["B2"] = 10, 20
    ws["C1"] = '=SUMIF(A1:A2,"Y",B1:B2)'
    book.save(path)

    _, result, _, _ = _strict_evaluate(path, {"Text!C1"})

    assert isinstance(result.values["Text!C1"], Unresolved)


def test_numbervalue_and_rank_variants_cover_separators_ties_and_errors():
    evaluator = Evaluator(graph=None, cg=None)

    assert evaluator._fn_numbervalue(["1.234,56", ",", "."]) == 1234.56
    assert evaluator._fn_numbervalue(["12,5%", ",", "."]) == 0.125
    assert evaluator._fn_numbervalue(["1 234.5", ".", " "]) == 1234.5
    assert evaluator._fn_numbervalue(["1.2.3", ".", ","]).code == "#VALUE!"

    values = [100, 90, 90, 80, "ignored"]
    assert evaluator._fn_rank([90, values]) == 2
    assert evaluator._fn_rank_eq([90, values, 1]) == 2
    assert evaluator._fn_rank([80, values]) == 4
    assert evaluator._fn_rank_eq([80, values, 1]) == 1
    assert evaluator._fn_rank(["90", values]).code == "#VALUE!"
    assert evaluator._fn_rank([90, []]).code == "#N/A"


def test_xlfn_numbervalue_and_rank_dispatch_end_to_end(tmp_path):
    path = tmp_path / "future-functions.xlsx"
    book = openpyxl.Workbook()
    ws = book.active
    ws["A1"], ws["A2"], ws["A3"] = 3, 3, 1
    ws["B1"] = '=_xlfn.NUMBERVALUE("1.234,5",",",".")'
    ws["B2"] = "=_xlfn.RANK.EQ(3,A1:A3,0)"
    book.save(path)

    _, _, result = _evaluate(path)

    assert result.values["Sheet!B1"] == 1234.5
    assert result.values["Sheet!B2"] == 1


def test_excel_1900_serial_boundaries_and_modern_dates():
    evaluator = Evaluator(graph=None, cg=None)

    assert to_serial("1900-01-01T00:00:00") == 1
    assert [evaluator._fn_day([serial]) for serial in (1, 59, 60, 61)] == [
        1, 28, 29, 1
    ]
    assert evaluator._fn_month([60]) == 2
    assert evaluator._fn_year([60]) == 1900
    assert evaluator._fn_date([1900, 1, 1]) == 1
    assert evaluator._fn_date([1900, 2, 29]) == 60
    assert evaluator._fn_edate([60, 0]) == 60
    assert evaluator._fn_eomonth([1, 1]) == 60
    assert evaluator._fn_date([2024, 2, 29]) == 45351
    assert evaluator._fn_day([45351]) == 29

    before = _excel_date_parts.cache_info().hits
    _excel_date_parts(45351)
    _excel_date_parts(45351)
    assert _excel_date_parts.cache_info().hits >= before + 1


def test_sumif_alignment_and_active_dependencies_avoid_false_cycles(tmp_path):
    path = tmp_path / "aligned-sums.xlsx"
    book = openpyxl.Workbook()
    ws = book.active
    ws.title = "Align"
    ws["A1"], ws["A2"], ws["A3"] = "N", "Y", "Y"
    ws["C2"], ws["C3"], ws["C4"] = 10, 20, 30
    ws["D1"] = '=SUMIF(A1:A3,"Y",$C$2)'

    ws["G1"] = "=F1"
    ws["G2"] = 7
    ws["F1"] = '=SUMIF(A1:A2,"Y",G1:G2)'

    ws["K1"] = "=J1"
    ws["K2"] = 8
    ws["J1"] = '=SUMIFS(K1:K2,A1:A2,"Y")'

    # B starts earlier than A. Independent populated-cell clipping used to
    # shift these whole-column ranges into different coordinate frames.
    whole = book.create_sheet("Whole")
    whole["A5"], whole["A6"] = "Y", "N"
    whole["B2"], whole["B5"], whole["B6"] = 99, 40, 50
    whole["D1"] = '=SUMIF(A:A,"Y",B:B)'
    whole["D2"] = '=SUMIFS(B:B,A:A,"Y")'
    book.save(path)

    built, _, result = _evaluate(path)

    assert result.values["Align!D1"] == 50
    assert result.values["Align!F1"] == 7
    assert result.values["Align!J1"] == 8
    assert result.values["Whole!D1"] == 40
    assert result.values["Whole!D2"] == 40
    assert result.runtime_radj["Align!F1"] >= {"Align!A1", "Align!A2", "Align!G2"}
    assert "Align!G1" not in result.runtime_radj["Align!F1"]
    assert result.runtime_radj["Align!J1"] >= {"Align!A1", "Align!A2", "Align!K2"}
    assert "Align!K1" not in result.runtime_radj["Align!J1"]

    a_range = built.resolve("A:A", "Whole")
    b_range = built.resolve("B:B", "Whole")
    assert a_range == [("range", "Whole", "A1:A6", 6)]
    assert b_range == [("range", "Whole", "B1:B6", 6)]


def test_sumifs_rejects_unequal_range_shapes(tmp_path):
    path = tmp_path / "sumifs-shape.xlsx"
    book = openpyxl.Workbook()
    ws = book.active
    ws["A1"], ws["A2"] = 10, 20
    ws["B1"], ws["B2"] = 1, 1
    ws["C1"] = "=SUMIFS(A1:A2,B1:B1,1)"
    book.save(path)

    _, _, result = _evaluate(path)

    assert result.values["Sheet!C1"].code == "#VALUE!"


def test_ast_expands_global_and_sheet_scoped_names_that_look_like_cells(tmp_path):
    path = tmp_path / "names.xlsx"
    book = openpyxl.Workbook()
    data = book.active
    data.title = "Data"
    calc = book.create_sheet("Calc")
    data["A1"], data["A2"], data["A3"] = 11, 12, 13
    calc["A1"] = 17

    book.defined_names.add(DefinedName("Q1", attr_text="'Data'!$A$1"))
    book.defined_names.add(DefinedName("BC", attr_text="'Data'!$A$2"))
    data.defined_names.add(DefinedName("Rate", attr_text="$A$3"))
    calc.defined_names.add(DefinedName("Rate", attr_text="$A$1"))
    calc["B1"] = "=Q1+BC+Rate"
    calc["B2"] = "=Data!Rate"
    book.save(path)

    built, graph, result = _evaluate(path)

    assert built.resolve("Q1", "Calc") == [("cell", "Data", 1, 1)]
    assert built.resolve("BC", "Calc") == [("cell", "Data", 2, 1)]
    assert built.resolve("Rate", "Calc") == [("cell", "Calc", 1, 1)]
    assert built.resolve("Data!Rate", "Calc") == [("cell", "Data", 3, 1)]
    assert result.values["Calc!B1"] == 40
    assert result.values["Calc!B2"] == 13
    assert not [node for node in graph.nodes.values() if node.kind == "name"]
