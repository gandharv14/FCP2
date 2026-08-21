"""Regression tests for the four leak/blind-spot fixes:

1. no searchable marker distinguishes answer documents;
2. perturbed wrong values always clear the grader's acceptance band;
3. multi-cell array-formula members are formula cells, not inputs;
4. the pasted-answer detector covers frontier cells, rounded pastes,
   and text renderings.
"""

from __future__ import annotations

import zipfile
from pathlib import Path

import openpyxl
import pytest

from mcp_env.build import GRADER_TOLERANCE, make_document, perturb
from mcp_env.validate import within_grader_band
from xl_ast_graph import AstGraph
from xl_input_mask import pasted_answers


# -- 1. answer-document marker ------------------------------------------------

SOURCE = {"id": "src", "name": "Example Service", "url": "internal://example"}


def _row(**overrides):
    row = {
        "dataset_id": "ds", "source_id": "src", "document_id": "doc",
        "entity": "ACME", "metric": "Total revenue", "metric_aliases": [],
        "period": "FY2020", "scenario": "base", "basis": "nominal",
        "unit": "EUR", "status": "final", "release": "rel_1",
        "published_at": "2020-01-01", "superseded_by": None, "value": 42.5,
    }
    row.update(overrides)
    return row


def test_supported_and_distractor_documents_share_the_same_guidance():
    supported = make_document("d1", SOURCE, _row(), "data-release", "2020-01-01")
    distractor = make_document("d2", SOURCE, _row(value=40.0), "archive",
                               "2019-01-01")
    tail = lambda doc: doc["content"].rsplit("\n\n", 1)[-1]
    assert tail(supported) == tail(distractor)
    assert "authoritative" not in supported["content"]
    assert "supersed" not in supported["content"].casefold()


def test_stale_documents_still_carry_the_superseded_banner():
    stale = make_document("d3", SOURCE, _row(superseded_by="rel_2"),
                          "archive", "2018-01-01", superseded_by="rel_2")
    assert "SUPERSEDED" in stale["content"]
    assert "rel_2" in stale["content"]


# -- 2. perturbation floor ----------------------------------------------------

@pytest.mark.parametrize("value", [2e-5, 1e-6, 0.0, 0.016, 1.2, 827.078559,
                                   -4.2e-6, 12345678.9012345,
                                   15382792, 3, -20, 400000000])
@pytest.mark.parametrize("index", range(6))
def test_perturbed_values_clear_the_grader_band(value, index):
    candidate = perturb({"value": value}, index)
    band = 10 * max(GRADER_TOLERANCE, GRADER_TOLERANCE * abs(value))
    assert abs(candidate - value) > band
    assert type(candidate) is type(value)


def test_within_grader_band_matches_the_floor_semantics():
    assert within_grader_band(2.08e-5, 2e-5)          # inside 10x band
    assert not within_grader_band(0.017, 0.016)       # clearly outside
    assert not within_grader_band("n/a", 0.016)       # non-numeric never bands


# -- 3. array-formula members -------------------------------------------------

def _array_workbook(path: Path):
    """A workbook whose B1:B3 is a CSE array formula, written at XML level
    the way Excel serializes it: the formula only on the master cell."""
    book = openpyxl.Workbook()
    ws = book.active
    ws.title = "S1"
    for i, v in enumerate([10, 20, 30], start=1):
        ws.cell(row=i, column=1, value=v)
    for i, v in enumerate([20, 40, 60], start=1):
        ws.cell(row=i, column=2, value=v)
    book.save(path)
    with zipfile.ZipFile(path) as src:
        parts = {item.filename: src.read(item.filename) for item in src.infolist()}
    sheet = "xl/worksheets/sheet1.xml"
    xml = parts[sheet]
    replaced = xml.replace(
        b'<c r="B1" t="n"><v>20</v></c>',
        b'<c r="B1" t="n"><f t="array" ref="B1:B3">A1:A3*2</f><v>20</v></c>')
    assert replaced != xml, "master cell XML not found; openpyxl layout changed"
    xml = replaced
    parts[sheet] = xml
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as out:
        for name, data in parts.items():
            out.writestr(name, data)


def test_array_members_become_formula_cells_with_edges(tmp_path):
    path = tmp_path / "array.xlsx"
    _array_workbook(path)
    graph = AstGraph(path, verbose=False).build()
    for coord in ("B1", "B2", "B3"):
        node = graph.nodes["S1!%s" % coord]
        assert node["kind"] == "formula", coord
        assert node["array_formula"] is True, coord
    incoming = {edge["target"] for edge in graph.edges}
    assert "S1!B2" in incoming and "S1!B3" in incoming


# -- 4. pasted-answer blind spots ----------------------------------------------

def _paste_workbook():
    book = openpyxl.Workbook()
    ws = book.active
    ws.title = "M"
    ws["A1"] = 61.5989123456        # exact paste, typed, off-frontier
    ws["A2"] = 61.6                 # display-rounded paste
    ws["A3"] = "IRR came to 23.4%"  # text rendering of 0.234
    ws["A4"] = 61.5989123456        # exact paste inside the frontier
    ws["A5"] = 4.0                  # trivial small integer, ignored
    return book


def test_pasted_answer_classes():
    book = _paste_workbook()
    outputs = [61.5989123456, 0.234]
    frontier = {"M": {(4, 1)}}
    deny, report, suspects = pasted_answers(book, {}, frontier, outputs)

    assert (1, 1) in deny["M"]                       # exact -> blanked
    assert any("A1" in line for line in report)
    assert (2, 1) not in deny["M"]                   # rounded -> reported only
    assert any("A2" in line and "significant digits" in line
               for line in suspects)
    assert any("A3" in line and "renders output" in line for line in suspects)
    assert (4, 1) not in deny["M"]                   # frontier -> reported only
    assert any("A4" in line and "frontier" in line for line in suspects)
    assert not any("A5" in line for line in report + suspects)
