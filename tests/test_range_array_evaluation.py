from __future__ import annotations

import csv
from pathlib import Path

import openpyxl
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.formula import ArrayFormula

import xl_ast_graph
from xl_ast_graph import AstGraph, write_csv
from xl_seg.evaluate import (
    ExcelError,
    Evaluator,
    RangeValues,
    Unresolved,
    _binary,
    _intersection,
)
from xl_seg.model import load
from xl_seg.project import build as build_cell_graph


def _array_workbook(path: Path) -> None:
    book = openpyxl.Workbook()
    ws = book.active
    ws.title = "Arrays"
    ws["A1"], ws["B1"], ws["A2"], ws["B2"] = 1, 2, 3, 4
    ws["A4"], ws["B4"], ws["C4"] = 10, 20, 30
    ws["H1"], ws["H2"], ws["H3"] = 5, 6, 7

    ws["C1"] = ArrayFormula(ref="C1:D2", text="=A1:B2*10")
    ws["E1"] = ArrayFormula(ref="E1:F2", text="=TRANSPOSE(A1:B2)")
    ws["G1"] = "=SUMPRODUCT((A1:B2>2)*A1:B2)"
    ws["D4"] = "=INDEX(A4:C4,2)"
    ws["I1"] = "=INDEX(H1:H3,2)"
    ws["I2"] = "=SUM(INDEX(A1:B2,0,2))"
    ws["I3"] = "=COLUMNS(INDEX(A1:B2,2,0))"
    ws["J1"] = "=SUM(A1:B2 B1:B2)"
    book.save(path)


def _evaluate_workbook(path: Path):
    built = AstGraph(path, verbose=False).build()
    out = path.parent / "ast"
    out.mkdir()
    write_csv(out, built.nodes, built.edges)
    graph = load(out, path.stem)
    cell_graph = build_cell_graph(graph)
    inputs = {node.id for node in graph.cells() if node.kind == "input"}
    return built, graph, Evaluator(graph, cell_graph).run(inputs)


def test_range_binary_broadcast_comparison_and_error_propagation():
    values = RangeValues([1, 2, 3, 4], rows=2, cols=2)
    scaled = _binary("*", values, 10)
    assert isinstance(scaled, RangeValues)
    assert (scaled.rows, scaled.cols, list(scaled)) == (2, 2, [10, 20, 30, 40])

    mask = _binary(">", values, 2)
    assert (mask.rows, mask.cols, list(mask)) == (
        2, 2, [False, False, True, True]
    )
    ev = Evaluator(graph=None, cg=None)
    assert ev._fn_sumproduct([mask, values]) == 7

    errors = RangeValues(
        [1, ExcelError("#N/A"), Unresolved("missing"), 4], rows=2, cols=2
    )
    propagated = _binary("+", errors, 1)
    assert propagated[0] == 2
    assert isinstance(propagated[1], ExcelError)
    assert isinstance(propagated[2], Unresolved)

    mismatch = _binary("+", values, RangeValues([1, 2], rows=1, cols=2))
    assert isinstance(mismatch, ExcelError)
    assert mismatch.code == "#VALUE!"


def test_index_one_dimensional_and_reference_shapes():
    ev = Evaluator(graph=None, cg=None)
    horizontal = RangeValues([10, 20, 30], rows=1, cols=3)
    vertical = RangeValues([10, 20, 30], rows=3, cols=1)
    assert ev._fn_index([horizontal, 2]) == 20
    assert ev._fn_index([vertical, 2]) == 20

    table = RangeValues(
        [1, 2, 3, 4],
        rows=2,
        cols=2,
        refs=["S!A1", "S!B1", "S!A2", "S!B2"],
    )
    column = ev._fn_index([table, 0, 2])
    row = ev._fn_index([table, 2, 0])
    assert (column.rows, column.cols, list(column), column.refs) == (
        2, 1, [2, 4], ("S!B1", "S!B2")
    )
    assert (row.rows, row.cols, list(row), row.refs) == (
        1, 2, [3, 4], ("S!A2", "S!B2")
    )
    assert ev._fn_index([table, 2, 2]) == 4


def test_reference_intersection_uses_current_ast_operator_representation():
    left = RangeValues(
        [1, 2, 3, 4], 2, 2, refs=["S!A1", "S!B1", "S!A2", "S!B2"]
    )
    right = RangeValues(
        [20, 40], 2, 1, refs=["S!B1", "S!B2"]
    )
    result = _intersection(left, right)
    assert (result.rows, result.cols, list(result), result.refs) == (
        2, 1, [2, 4], ("S!B1", "S!B2")
    )
    assert _intersection(
        left, RangeValues([9], refs=["S!C3"])
    ).code == "#NULL!"


def test_excel_backed_spills_masks_and_index_end_to_end(tmp_path):
    workbook = tmp_path / "arrays.xlsx"
    _array_workbook(workbook)
    built, graph, result = _evaluate_workbook(workbook)

    assert [result.values[f"Arrays!{coord}"] for coord in ("C1", "D1", "C2", "D2")] \
        == [10, 20, 30, 40]
    assert [result.values[f"Arrays!{coord}"] for coord in ("E1", "F1", "E2", "F2")] \
        == [1, 3, 2, 4]
    assert result.values["Arrays!G1"] == 7
    assert result.values["Arrays!D4"] == 20
    assert result.values["Arrays!I1"] == 6
    assert result.values["Arrays!I2"] == 6
    assert result.values["Arrays!I3"] == 2
    assert result.values["Arrays!J1"] == 6

    expected_members = {
        "Arrays!C1": (0, 0), "Arrays!D1": (0, 1),
        "Arrays!C2": (1, 0), "Arrays!D2": (1, 1),
    }
    for node_id, position in expected_members.items():
        raw = built.nodes[node_id]
        loaded = graph.nodes[node_id]
        assert raw["array_anchor"] == "Arrays!C1"
        assert raw["array_ref"] == "C1:D2"
        assert (raw["array_row"], raw["array_col"]) == position
        assert (loaded.array_row, loaded.array_col) == position


def test_top_level_cse_range_has_one_root_per_projected_member(tmp_path):
    workbook = tmp_path / "cse-range.xlsx"
    book = openpyxl.Workbook()
    ws = book.active
    ws.title = "CSE"
    ws["A1"], ws["A2"], ws["A3"] = 10, 20, 30
    ws["B1"] = ArrayFormula(ref="B1:B3", text="=A1:A3")
    book.save(workbook)

    built, graph, result = _evaluate_workbook(workbook)

    assert not {
        error["code"] for error in graph.integrity_errors
    } & {"invalid_formula_root_count"}
    for row, expected in enumerate((10, 20, 30), 1):
        member = f"CSE!B{row}"
        incoming = graph.in_edges[member]
        assert len(incoming) == 1
        assert incoming[0].source == "CSE!A1:A3"
        assert result.values[member] == expected
    assert built.nodes["CSE!A1:A3"]["kind"] == "range"


def test_defined_name_to_explicit_blank_preserves_cell_reference(tmp_path):
    workbook = tmp_path / "blank-name.xlsx"
    book = openpyxl.Workbook()
    ws = book.active
    ws.title = "Blank"
    ws["B1"] = "=BLANK_CELL"
    book.defined_names.add(
        DefinedName("BLANK_CELL", attr_text="'Blank'!$A$1")
    )
    book.save(workbook)

    built, graph, result = _evaluate_workbook(workbook)

    assert built.resolve("BLANK_CELL", "Blank") == [
        ("cell", "Blank", 1, 1)
    ]
    assert graph.root_of("Blank!B1") == "Blank!A1"
    assert not (
        isinstance(result.values["Blank!B1"], ExcelError)
        and result.values["Blank!B1"].code == "#NAME?"
    )


def test_model_loader_accepts_legacy_node_columns(tmp_path):
    workbook = tmp_path / "arrays.xlsx"
    _array_workbook(workbook)
    built = AstGraph(workbook, verbose=False).build()
    out = tmp_path / "legacy"
    out.mkdir()
    write_csv(out, built.nodes, built.edges)

    nodes_path = out / "nodes.csv"
    with nodes_path.open(newline="", encoding="utf-8") as handle:
        rows = list(csv.DictReader(handle))
    legacy_fields = [
        field for field in xl_ast_graph.NODE_FIELDS
        if field not in {
            "array_anchor", "array_ref", "array_row", "array_col",
            "ast_schema_version", "parse_status", "parse_error",
            "value_type", "range_truncated",
        }
    ]
    with nodes_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle, fieldnames=legacy_fields, extrasaction="ignore"
        )
        writer.writeheader()
        writer.writerows(rows)

    graph = load(out, workbook.stem)
    member = graph.nodes["Arrays!D2"]
    assert member.array_formula is True
    assert member.array_anchor == ""
    assert member.array_ref == ""
    assert member.array_row is None
    assert member.array_col is None
    assert graph.capabilities["required_node_fields"] is False
    assert graph.capabilities["schema_version_v2"] is False


def test_formula_nodes_survive_parse_failure_with_bounded_diagnostics(
    tmp_path, monkeypatch
):
    path = tmp_path / "parse-error.xlsx"
    book = openpyxl.Workbook()
    book.active["A1"] = "=1+1"
    book.save(path)

    def fail_parse(_formula):
        raise xl_ast_graph.FormulaError("x" * 500)

    monkeypatch.setattr(xl_ast_graph, "parse_formula", fail_parse)
    built = AstGraph(path, verbose=False).build()
    node = built.nodes["Sheet!A1"]

    assert node["kind"] == "formula"
    assert node["parse_status"] == "error"
    assert len(node["parse_error"]) == xl_ast_graph.MAX_PARSE_ERROR_CHARS


def test_model_loader_reports_missing_endpoints_and_formula_roots(tmp_path):
    path = tmp_path / "integrity.xlsx"
    book = openpyxl.Workbook()
    ws = book.active
    ws["A1"] = 1
    ws["B1"] = "=A1"
    book.save(path)
    built = AstGraph(path, verbose=False).build()
    out = tmp_path / "integrity-ast"
    out.mkdir()
    write_csv(out, built.nodes, built.edges)

    edges_path = out / "edges.csv"
    with edges_path.open(newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        fields = reader.fieldnames
        rows = [
            row for row in reader
            if row["target"] != "Sheet!B1"
        ]
    missing = dict(rows[0]) if rows else {
        field: "" for field in fields
    }
    missing.update({
        "source": "Sheet!MISSING",
        "target": "Sheet!B1",
        "role": "identity",
        "arg_index": "0",
    })
    rows.append(missing)
    with edges_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)

    graph = load(out, path.stem)
    codes = {error["code"] for error in graph.integrity_errors}

    assert "missing_edge_endpoint" in codes
    assert "invalid_formula_root_count" in codes


def test_whole_columns_share_a_bounded_frame_with_unrelated_far_rows(tmp_path):
    path = tmp_path / "far-row.xlsx"
    book = openpyxl.Workbook()
    ws = book.active
    ws.title = "Far"
    ws["A5"], ws["A6"] = "Y", "N"
    ws["B5"], ws["B6"] = 40, 50
    ws["D1"] = '=SUMIF(A:A,"Y",B:B)'
    ws["Z1000000"] = 1
    book.save(path)

    built = AstGraph(path, max_range_expand=100, verbose=False).build()
    a_range = built.resolve("A:A", "Far")[0]
    b_range = built.resolve("B:B", "Far")[0]

    assert a_range[2] == "A1:A100"
    assert b_range[2] == "B1:B100"
    assert a_range[3] <= 100
    assert b_range[3] <= 100
    assert built.nodes["Far!A1:A100"]["range_truncated"] is False


def test_relevant_whole_column_truncation_fails_closed(tmp_path):
    path = tmp_path / "relevant-far-row.xlsx"
    book = openpyxl.Workbook()
    ws = book.active
    ws["A1"] = 1
    ws["A1000000"] = 2
    ws["B1"] = "=SUM(A:A)"
    book.save(path)

    _, graph, result = _evaluate_workbook(path)

    range_node = next(node for node in graph.nodes.values() if node.kind == "range")
    assert range_node.range_truncated is True
    assert isinstance(result.values["Sheet!B1"], Unresolved)
    assert result.values["Sheet!B1"].reason == "whole-column-range-exceeds-bound"


def test_array_fixture_production_csv_matches_default(tmp_path):
    workbook = tmp_path / "arrays.xlsx"
    _array_workbook(workbook)
    default_root = tmp_path / "default"
    production_root = tmp_path / "production"
    xl_ast_graph.main([str(workbook), "-o", str(default_root), "-q"])
    xl_ast_graph.main([
        str(workbook), "-o", str(production_root), "-q", "--production"
    ])

    default_dir = default_root / workbook.stem
    production_dir = production_root / workbook.stem
    assert (default_dir / "nodes.csv").read_bytes() == (
        production_dir / "nodes.csv"
    ).read_bytes()
    assert (default_dir / "edges.csv").read_bytes() == (
        production_dir / "edges.csv"
    ).read_bytes()
    assert {path.name for path in production_dir.iterdir()} == {
        "nodes.csv", "edges.csv"
    }
