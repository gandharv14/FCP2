from __future__ import annotations

import base64
import hashlib
import json
import os
import signal
import shutil
import subprocess
import sys
import time
import zipfile
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName

import xl_source_publication as publication
import xl_inventory_approval
from xl_inventory_approval import approval_claims, object_hash
from xl_source_health import inspect_workbook, sha256_file
from xl_source_recalc import (
    PF_NETWORK_RULES_SHA256,
    RESTRICTED_SOURCE_COHORT_MANIFEST,
    RecalculationError,
    MacOSExcelEngine,
    _validated_restricted_inventory_selection,
    create_identity_documents,
    create_restriction_documents,
    create_recalc_request,
    execute_recalc,
    prepare_source_generation,
    semantic_workbook_diff,
    validate_recalc_request,
    verify_signed_runner_receipt,
)
from xl_source_inventory import build_inventory_manifest


def _approve_inventory(
    tmp_path: Path,
    monkeypatch,
    inventory: dict,
    *,
    batch_id: str = "test-batch",
) -> Path:
    entry = approval_claims(inventory, batch_id=batch_id)
    core = {
        "schema_version": "source-inventory-approval-registry/v1",
        "approvals": [entry],
    }
    registry = {
        **core,
        "registry_sha256": object_hash(core),
    }
    path = tmp_path / "approved-source-inventories.json"
    path.write_text(json.dumps(registry), encoding="utf-8")
    monkeypatch.setattr(xl_inventory_approval, "DEFAULT_REGISTRY", path)
    return path


def _workbook(
    path: Path,
    *,
    input_value: str = "1",
    formula: str = "A1+1",
    cache: str | None = "2",
) -> None:
    value = "" if cache is None else f"<v>{cache}</v>"
    workbook = (
        '<workbook xmlns="http://schemas.openxmlformats.org/'
        'spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/'
        'officeDocument/2006/relationships"><sheets><sheet name="Sheet1" '
        'sheetId="1" r:id="rId1"/></sheets><calcPr calcMode="auto"/></workbook>'
    )
    relationships = (
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/'
        'relationships"><Relationship Id="rId1" Type="http://schemas.openxmlformats.org/'
        'officeDocument/2006/relationships/worksheet" '
        'Target="worksheets/sheet1.xml"/></Relationships>'
    )
    sheet = (
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/'
        'main"><sheetData><row r="1"><c r="A1">'
        f"<v>{input_value}</v></c><c r=\"B1\"><f>{formula}</f>{value}</c>"
        "</row></sheetData></worksheet>"
    )
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr(
            "[Content_Types].xml",
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/'
            'content-types"/>',
        )
        archive.writestr("xl/workbook.xml", workbook)
        archive.writestr("xl/_rels/workbook.xml.rels", relationships)
        archive.writestr("xl/worksheets/sheet1.xml", sheet)


def test_semantic_diff_separates_cache_changes_from_model_changes(tmp_path):
    before = tmp_path / "before.xlsx"
    cached = tmp_path / "cached.xlsx"
    changed_input = tmp_path / "changed-input.xlsx"
    _workbook(before, cache="2")
    _workbook(cached, cache="3")
    _workbook(changed_input, input_value="9", cache="10")

    cache_diff = semantic_workbook_diff(before, cached)
    input_diff = semantic_workbook_diff(before, changed_input)

    assert cache_diff["cache_only"] is True
    assert cache_diff["proven_stale_cache"] is True
    assert input_diff["equivalent_semantics"] is False
    assert input_diff["semantic_changes"] == ["inputs", "package_parts"]
    assert input_diff["proven_stale_cache"] is False


def test_semantic_diff_rejects_any_unapproved_package_part_change(tmp_path):
    before = tmp_path / "before.xlsx"
    after = tmp_path / "after.xlsx"
    _workbook(before)
    shutil.copyfile(before, after)
    with zipfile.ZipFile(after, "a") as archive:
        archive.writestr("docProps/custom.xml", "<properties><x>changed</x></properties>")

    difference = semantic_workbook_diff(before, after)

    assert difference["equivalent_semantics"] is False
    assert "package_parts" in difference["semantic_changes"]


def test_bound_request_and_injected_engine_preserve_original(tmp_path):
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "out" / "recalculated.xlsx"
    _workbook(source, cache="2")
    request = create_recalc_request(
        source,
        destination,
        request_id="request-0001",
        allowed_root=tmp_path / "out",
        engine_constraints={
            "allowed_engines": ["fake-engine"],
            "permitted_versions": ["test-v1"],
        },
    )
    original = source.read_bytes()

    class FakeEngine:
        name = "fake-engine"
        version = "test-v1"
        authoritative = True

        def capability(self):
            return True, "test"

        def execute(self, source_path, candidate, request_value):
            _workbook(candidate, cache="3")

    result = execute_recalc(
        request,
        FakeEngine(),
        allowed_root=tmp_path / "out",
    )

    assert source.read_bytes() == original
    assert destination.is_file()
    assert result["proven_stale_cache"] is True
    with pytest.raises(RecalculationError, match="destination_exists|replay"):
        execute_recalc(
            request,
            FakeEngine(),
            allowed_root=tmp_path / "out",
        )


def test_request_tamper_and_out_of_root_fail_closed(tmp_path):
    source = tmp_path / "source.xlsx"
    _workbook(source)
    request = create_recalc_request(
        source,
        tmp_path / "safe" / "out.xlsx",
        request_id="request-0002",
        allowed_root=tmp_path / "safe",
    )
    request["destination"] = str(tmp_path / "escape.xlsx")

    with pytest.raises(RecalculationError, match="request_sha256|destination_root"):
        validate_recalc_request(request, allowed_root=tmp_path / "safe")


def test_concurrent_recalculation_allows_one_publication(tmp_path):
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "out" / "recalculated.xlsx"
    _workbook(source, cache="2")
    request = create_recalc_request(
        source,
        destination,
        request_id="concurrent-request",
        allowed_root=tmp_path / "out",
        engine_constraints={
            "allowed_engines": ["slow-engine"],
            "permitted_versions": ["test-v1"],
        },
    )

    class SlowEngine:
        name = "slow-engine"
        version = "test-v1"
        authoritative = True

        def execute(self, source_path, candidate, request_value):
            time.sleep(0.05)
            _workbook(candidate, cache="3")

    def run_once():
        try:
            execute_recalc(
                request,
                SlowEngine(),
                allowed_root=tmp_path / "out",
            )
            return "published"
        except RecalculationError:
            return "rejected"

    with ThreadPoolExecutor(max_workers=2) as executor:
        outcomes = list(executor.map(lambda _: run_once(), range(2)))

    assert sorted(outcomes) == ["published", "rejected"]


def test_partial_engine_save_is_never_published(tmp_path):
    source = tmp_path / "source.xlsx"
    destination = tmp_path / "out" / "recalculated.xlsx"
    _workbook(source, cache="2")
    request = create_recalc_request(
        source,
        destination,
        request_id="partial-save",
        allowed_root=tmp_path / "out",
        engine_constraints={
            "allowed_engines": ["partial-engine"],
            "permitted_versions": ["test-v1"],
        },
    )

    class PartialEngine:
        name = "partial-engine"
        version = "test-v1"
        authoritative = True

        def execute(self, source_path, candidate, request_value):
            candidate.write_bytes(b"PK\x03\x04partial")

    with pytest.raises((RecalculationError, ValueError, zipfile.BadZipFile)):
        execute_recalc(
            request,
            PartialEngine(),
            allowed_root=tmp_path / "out",
        )

    assert not destination.exists()


def test_excel_timeout_terminates_only_owned_process_group(tmp_path, monkeypatch):
    destination = tmp_path / "candidate.xlsx"
    _workbook(destination)
    request = {
        "request_id": "timeout-request",
        "request_sha256": "a" * 64,
        "source_sha256": sha256_file(destination),
    }
    engine = MacOSExcelEngine(timeout_seconds=1)
    engine.sandbox_runner = tmp_path / "runner"
    engine.excel_app = tmp_path / "Excel.app"
    monkeypatch.setattr(engine, "capability", lambda: (True, "available"))

    class TimedOutProcess:
        pid = 4321
        returncode = None

        def communicate(self, timeout):
            raise subprocess.TimeoutExpired("runner", timeout)

        def wait(self, timeout):
            return 0

    killed = []
    monkeypatch.setattr(subprocess, "Popen", lambda *args, **kwargs: TimedOutProcess())
    monkeypatch.setattr(
        os,
        "killpg",
        lambda pid, signal_number: killed.append((pid, signal_number)),
    )

    with pytest.raises(RecalculationError, match="timed out"):
        engine.execute(destination, destination, request)

    assert killed == [(4321, signal.SIGTERM)]


def test_signed_runner_receipt_rejects_forgery(tmp_path):
    private_key = tmp_path / "private.pem"
    public_key = tmp_path / "public.pem"
    subprocess.run(
        [
            "/usr/bin/openssl",
            "genpkey",
            "-algorithm",
            "RSA",
            "-pkeyopt",
            "rsa_keygen_bits:2048",
            "-out",
            str(private_key),
        ],
        check=True,
        capture_output=True,
    )
    subprocess.run(
        [
            "/usr/bin/openssl",
            "pkey",
            "-in",
            str(private_key),
            "-pubout",
            "-out",
            str(public_key),
        ],
        check=True,
        capture_output=True,
    )
    public_key.chmod(0o600)
    payload = {
        "request_sha256": "a" * 64,
        "source_sha256": "b" * 64,
        "output_sha256": "c" * 64,
        "engine": "excel-macos",
        "engine_version": "16.99",
        "calculation_complete": True,
        "isolation_enforced": True,
        "network_isolation_mechanism": "macos-pf-anchor",
        "network_isolation_rules_sha256":
            PF_NETWORK_RULES_SHA256,
        "completed_at_ns": 123,
    }
    payload_path = tmp_path / "payload.json"
    signature_path = tmp_path / "signature.bin"
    payload_path.write_bytes(
        (
            json.dumps(payload, sort_keys=True, separators=(",", ":")) + "\n"
        ).encode("utf-8")
    )
    subprocess.run(
        [
            "/usr/bin/openssl",
            "dgst",
            "-sha256",
            "-sign",
            str(private_key),
            "-out",
            str(signature_path),
            str(payload_path),
        ],
        check=True,
        capture_output=True,
    )
    receipt = {
        "schema_version": "excel-runner-receipt/v1",
        "signed_payload": payload,
        "signature_base64": base64.b64encode(
            signature_path.read_bytes()
        ).decode("ascii"),
    }

    verified = verify_signed_runner_receipt(
        receipt,
        public_key,
        request_sha256="a" * 64,
        source_sha256="b" * 64,
        output_sha256="c" * 64,
        require_root_owner=False,
    )
    assert verified["engine_version"] == "16.99"
    with pytest.raises(RecalculationError, match="claims do not match"):
        verify_signed_runner_receipt(
            receipt,
            public_key,
            request_sha256="a" * 64,
            source_sha256="b" * 64,
            output_sha256="d" * 64,
            require_root_owner=False,
        )


def _ast(path: Path, marker: str) -> None:
    path.mkdir()
    (path / "nodes.csv").write_text(f"id\n{marker}\n", encoding="utf-8")
    (path / "edges.csv").write_text("source,target\n", encoding="utf-8")


def _publication_inputs(source: Path) -> tuple[dict, dict, dict]:
    request, result = create_identity_documents(source)
    return request, result, inspect_workbook(source)


def test_ast_provenance_tamper_and_interrupted_pointer_switch(
    tmp_path, monkeypatch
):
    monkeypatch.setattr(
        "xl_seg.publication.resolve_current_generation",
        lambda *args, **kwargs: ({}, {}),
    )
    source = tmp_path / "source.xlsx"
    _workbook(source)
    ast = tmp_path / "ast"
    _ast(ast, "first")
    request, result, health = _publication_inputs(source)
    root = tmp_path / "published"
    first, first_manifest = publication.publish_source_generation(
        source,
        ast,
        root,
        request=request,
        result=result,
        health=health,
        builder_args=["--production"],
    )
    publication.activate_source_generation(
        root,
        first_manifest["generation_id"],
        segmentation_dir=tmp_path / "seg",
    )
    resolved, _ = publication.resolve_current_source_generation(root)
    assert resolved == first

    first_ast = first / first_manifest["layout"]["ast_directory"]
    (first_ast / "nodes.csv").chmod(0o644)
    (first_ast / "nodes.csv").write_text("tampered\n", encoding="utf-8")
    with pytest.raises(publication.SourcePublicationError):
        publication.resolve_current_source_generation(root)

    # Restore the first immutable generation before testing rollback behavior.
    shutil.rmtree(root)
    first, first_manifest = publication.publish_source_generation(
        source,
        ast,
        root,
        request=request,
        result=result,
        health=health,
        builder_args=["--production"],
    )
    publication.activate_source_generation(
        root,
        first_manifest["generation_id"],
        segmentation_dir=tmp_path / "seg",
    )
    source2 = tmp_path / "source2.xlsx"
    _workbook(source2, formula="A1+2", cache="3")
    ast2 = tmp_path / "ast2"
    _ast(ast2, "second")
    request2, result2, health2 = _publication_inputs(source2)

    def interrupt(phase):
        if phase == "before_pointer_switch":
            raise RuntimeError("interrupted")

    with pytest.raises(RuntimeError, match="interrupted"):
        second, second_manifest = publication.publish_source_generation(
            source2,
            ast2,
            root,
            request=request2,
            result=result2,
            health=health2,
            builder_args=["--production"],
        )
        publication.activate_source_generation(
            root,
            second_manifest["generation_id"],
            segmentation_dir=tmp_path / "seg",
            fault=interrupt,
        )

    resolved, manifest = publication.resolve_current_source_generation(root)
    assert resolved == first
    assert manifest["generation_id"] == first_manifest["generation_id"]


def test_prepare_builds_inactive_bound_source_and_ast_generation(
    tmp_path, monkeypatch
):
    monkeypatch.setattr(
        "xl_seg.publication.resolve_current_generation",
        lambda *args, **kwargs: ({}, {}),
    )
    source = tmp_path / "raw-name.xlsx"
    book = Workbook()
    sheet = book.active
    sheet["A1"] = 1
    sheet["B1"] = 2
    book.save(source)
    health = inspect_workbook(source)
    root = tmp_path / "source-out" / "0001"

    generation, manifest = prepare_source_generation(
        source,
        "0001",
        root,
        health=health,
    )

    assert not (root / "current.json").exists()
    assert (generation / "source" / "0001.xlsx").is_file()
    assert (generation / "ast" / "0001" / "nodes.csv").is_file()
    assert (generation / "ast-provenance.json").is_file()
    with pytest.raises(
        publication.SourcePublicationError,
        match="strict segmentation proof",
    ):
        publication.activate_source_generation(
            root,
            manifest["generation_id"],
        )
    publication.activate_source_generation(
        root,
        manifest["generation_id"],
        segmentation_dir=tmp_path / "seg",
    )
    effective = publication.resolve_effective_source(root)
    assert effective["workbook_id"] == "0001"
    assert effective["source_sha256"] == sha256_file(source)
    assert effective["ast_root"] == str(generation / "ast")


def test_excel_adapter_requires_isolation_attestation():
    available, reason = MacOSExcelEngine().capability()
    assert available is False
    assert reason in {
        "platform_not_macos",
        "osascript_unavailable",
        "excel_unavailable",
        "isolation_attestation_required",
    }


def test_prepare_cli_emits_one_parseable_json_document(tmp_path):
    source = tmp_path / "source.xlsx"
    book = Workbook()
    book.active["A1"] = 1
    book.save(source)
    root = tmp_path / "published"

    completed = subprocess.run(
        [
            sys.executable,
            str(Path(__file__).resolve().parents[1] / "xl_source_recalc.py"),
            "prepare",
            str(source),
            "--workbook",
            "0002",
            "--publication-root",
            str(root),
        ],
        check=True,
        capture_output=True,
        text=True,
    )
    payload = json.loads(completed.stdout)

    assert payload["source_sha256"] == sha256_file(source)
    assert payload["generation_id"]
    assert Path(payload["source_path"]).name == "0002.xlsx"


def test_recalc_candidate_cannot_publish_with_identity_evidence(tmp_path):
    source = tmp_path / "source.xlsx"
    book = Workbook()
    book.active["A1"] = "=1+1"
    book.save(source)

    with pytest.raises(RecalculationError, match="identity evidence"):
        prepare_source_generation(
            source,
            "0003",
            tmp_path / "published",
        )


def test_restricted_prepare_rejects_unapproved_inventory(
    tmp_path, monkeypatch
):
    source_root = tmp_path / "sources"
    source_root.mkdir()
    source = source_root / "0004.xlsx"
    _workbook(source, formula="NOW()", cache="2")
    health = inspect_workbook(source)
    inventory = build_inventory_manifest(source_root, workbook_ids=["0004"])

    def fake_ast_main(args):
        output_root = Path(args[args.index("-o") + 1])
        output_root.mkdir()
        _ast(output_root / "0004", "restricted")

    monkeypatch.setattr("xl_ast_graph.main", fake_ast_main)

    with pytest.raises(RecalculationError, match="not approved"):
        prepare_source_generation(
            source,
            "0004",
            tmp_path / "published",
            health=health,
            inventory=inventory,
            inventory_batch_id="unapproved-batch",
        )


def test_approved_mixed_restricted_source_binds_signals_and_approval(
    tmp_path,
    monkeypatch,
):
    source_root = tmp_path / "sources"
    source_root.mkdir()
    source = source_root / "0004.xlsx"
    _workbook(source, formula="OFFSET(A1,0,0)", cache=None)
    health = inspect_workbook(source)
    inventory = build_inventory_manifest(source_root, workbook_ids=["0004"])
    registry = _approve_inventory(
        tmp_path,
        monkeypatch,
        inventory,
        batch_id="batch-mixed",
    )

    def fake_ast_main(args):
        output_root = Path(args[args.index("-o") + 1])
        output_root.mkdir()
        _ast(output_root / "0004", "mixed-restricted")

    monkeypatch.setattr("xl_ast_graph.main", fake_ast_main)

    generation, manifest = prepare_source_generation(
        source,
        "0004",
        tmp_path / "published",
        health=health,
        inventory=inventory,
        inventory_batch_id="batch-mixed",
        inventory_approval_registry=registry,
    )

    assert health["route"] == "restricted_recalc_pass"
    assert manifest["bindings"]["inventory_approval_sha256"]
    assert manifest["bindings"]["recalc_signals_sha256"] == (
        health["recalc_signals_sha256"]
    )
    assert generation.is_dir()
    assert not (tmp_path / "published" / "current.json").exists()


def test_restricted_evidence_rejects_inventory_health_route_drift(
    tmp_path,
    monkeypatch,
):
    source = tmp_path / "0004.xlsx"
    _workbook(source, formula="OFFSET(A1,0,0)", cache=None)
    health = inspect_workbook(source)
    selected = {
        "classification": "native_source",
        "health_report": health,
        "health_report_sha256": health["report_sha256"],
        "restriction_events_sha256": health[
            "restriction_events_sha256"
        ],
        "route": "restricted_pass",
    }
    monkeypatch.setattr(
        "xl_source_recalc._validated_restricted_inventory_selection",
        lambda *_args, **_kwargs: (
            selected,
            "1" * 64,
            "2" * 64,
            {"inventory_approval_sha256": "3" * 64},
        ),
    )

    with pytest.raises(
        RecalculationError,
        match="does not exactly match",
    ):
        create_restriction_documents(
            source,
            health,
            {"cohort": {"size": 1}},
            batch_id="batch-mixed",
        )


def test_restricted_publication_rejects_identity_and_unapproved_inventory(
    tmp_path,
):
    source_root = tmp_path / "sources"
    source_root.mkdir()
    source = source_root / "0005.xlsx"
    _workbook(source, formula='CELL("filename",A1)', cache="2")
    health = inspect_workbook(source)
    inventory = build_inventory_manifest(source_root, workbook_ids=["0005"])
    ast = tmp_path / "ast"
    _ast(ast, "restricted")
    identity_request, identity_result = create_identity_documents(source)

    with pytest.raises(
        publication.SourcePublicationError,
        match="approved inventory",
    ):
        publication.publish_source_generation(
            source,
            ast,
            tmp_path / "identity-published",
            request=identity_request,
            result=identity_result,
            health=health,
            inventory=inventory,
            inventory_batch_id="unapproved-batch",
        )

    with pytest.raises(RecalculationError, match="not approved"):
        create_restriction_documents(
            source,
            health,
            inventory,
            batch_id="unapproved-batch",
        )


def test_frozen_inventory_classification_controls_restricted_evidence():
    inventory = json.loads(RESTRICTED_SOURCE_COHORT_MANIFEST.read_text())
    records = {
        record["workbook_id"]: record for record in inventory["workbooks"]
    }
    unverified_id = inventory["classification_cohorts"][
        "conversion_unverified"
    ][0]
    allowed_ids = (
        inventory["classification_cohorts"]["conversion_equivalent"]
        + inventory["classification_cohorts"]["native_source"]
    )
    allowed_id = allowed_ids[0]

    with pytest.raises(RecalculationError, match="conversion_unverified"):
        _validated_restricted_inventory_selection(
            inventory,
            source_sha256=records[unverified_id]["sha256"],
            workbook_id=unverified_id,
            batch_id="batch-002-restricted-123-v2",
        )

    selected, cohort_hash, inventory_hash, approval = (
        _validated_restricted_inventory_selection(
            inventory,
            source_sha256=records[allowed_id]["sha256"],
            workbook_id=allowed_id,
            batch_id="batch-002-restricted-123-v2",
            expected_cohort_sha256=inventory["cohort"]["cohort_sha256"],
        )
    )
    assert selected["classification"] in {
        "native_source",
        "conversion_equivalent",
    }
    assert cohort_hash == inventory["cohort"]["cohort_sha256"]
    assert inventory_hash == inventory["inventory_sha256"]
    assert approval["inventory_approval_sha256"]


def test_generation_identity_ignores_receipt_request_id(tmp_path):
    source = tmp_path / "source.xlsx"
    _workbook(source)
    ast = tmp_path / "ast"
    _ast(ast, "same")
    request, result = create_identity_documents(source)
    health = inspect_workbook(source)
    root = tmp_path / "published"
    first, first_manifest = publication.publish_source_generation(
        source,
        ast,
        root,
        request=request,
        result=result,
        health=health,
    )
    request2 = {**request, "request_id": "identity-second"}
    request2.pop("request_sha256")
    request2["request_sha256"] = hashlib.sha256(
        (
            json.dumps(request2, sort_keys=True, separators=(",", ":")) + "\n"
        ).encode("utf-8")
    ).hexdigest()
    result2 = {
        **result,
        "request_id": request2["request_id"],
        "request_sha256": request2["request_sha256"],
    }
    result2.pop("result_sha256")
    result2["result_sha256"] = hashlib.sha256(
        (
            json.dumps(result2, sort_keys=True, separators=(",", ":")) + "\n"
        ).encode("utf-8")
    ).hexdigest()
    second, second_manifest = publication.publish_source_generation(
        source,
        ast,
        root,
        request=request2,
        result=result2,
        health=health,
    )

    assert second == first
    assert second_manifest["generation_id"] == first_manifest["generation_id"]
    assert (
        root
        / "receipts"
        / first_manifest["generation_id"]
        / "identity-second"
        / "result.json"
    ).is_file()


def test_live_builder_drift_invalidates_source_generation(tmp_path):
    source = tmp_path / "source.xlsx"
    _workbook(source)
    ast = tmp_path / "ast"
    _ast(ast, "same")
    builder = tmp_path / "builder.py"
    builder.write_text("VERSION = 1\n", encoding="utf-8")
    request, result = create_identity_documents(source)
    generation, _ = publication.publish_source_generation(
        source,
        ast,
        tmp_path / "published",
        request=request,
        result=result,
        health=inspect_workbook(source),
        builder_code_paths=[builder],
        builder_args=["--production"],
    )

    builder.write_text("VERSION = 2\n", encoding="utf-8")

    with pytest.raises(
        publication.SourcePublicationError,
        match="ast_provenance",
    ):
        publication.validate_source_generation(generation)


def test_publication_reobserves_source_health(tmp_path):
    source = tmp_path / "volatile-defined-name.xlsx"
    book = Workbook()
    book.active["A1"] = 1
    book.defined_names.add(DefinedName("Clock", attr_text="NOW()"))
    book.save(source)
    forged_health = inspect_workbook(source)
    assert forged_health["route"] == "unsupported"
    forged_health["route"] = "pass"
    forged_health["routing"] = "pass"
    forged_health["reason_codes"] = []
    forged_health.pop("report_sha256")
    forged_health["report_sha256"] = hashlib.sha256(
        (
            json.dumps(
                forged_health,
                sort_keys=True,
                separators=(",", ":"),
            )
            + "\n"
        ).encode("utf-8")
    ).hexdigest()
    ast = tmp_path / "ast"
    _ast(ast, "same")
    request, result = create_identity_documents(source)

    with pytest.raises(
        publication.SourcePublicationError,
        match="fresh source observation",
    ):
        publication.publish_source_generation(
            source,
            ast,
            tmp_path / "published",
            request=request,
            result=result,
            health=forged_health,
        )


@pytest.mark.real_excel
@pytest.mark.skipif(
    os.environ.get("FCP2_RUN_REAL_EXCEL") != "1",
    reason="set FCP2_RUN_REAL_EXCEL=1 with trusted runner paths",
)
def test_real_excel_recalculation_preserves_semantics(tmp_path):
    source = tmp_path / "real-excel-source.xlsx"
    book = Workbook()
    sheet = book.active
    sheet["A1"] = 2
    sheet["B1"] = "=A1*3"
    book.save(source)
    request = create_recalc_request(
        source,
        tmp_path / "real-excel-output.xlsx",
        allowed_root=tmp_path,
        trusted_runner_public_key=os.environ[
            "FCP2_EXCEL_RUNNER_PUBLIC_KEY"
        ],
        engine_constraints={
            "permitted_versions": [os.environ["FCP2_EXCEL_VERSION"]]
        },
    )
    engine = MacOSExcelEngine(
        isolation_attestation=os.environ["FCP2_EXCEL_ISOLATION_ATTESTATION"],
        sandbox_runner=os.environ["FCP2_EXCEL_SANDBOX_RUNNER"],
    )

    result = execute_recalc(
        request,
        engine,
        allowed_root=tmp_path,
        source_path=source,
    )

    assert result["authoritative"] is True
    assert result["semantic_diff"]["equivalent_semantics"] is True
    assert (tmp_path / "real-excel-output.xlsx").is_file()

