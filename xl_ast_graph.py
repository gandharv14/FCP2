#!/usr/bin/env python3
"""Turn an Excel workbook into an expression DAG.

Cells are variable nodes keyed ``Sheet!A1``. Every operator and function inside a
formula becomes its own intermediate node keyed ``<cell>#<i>:<op>``, so

    Calculations!T232 = SUM(T215:T231)*$C214

becomes ``T215..T231 -> op:SUM -> op:* -> T232`` with ``C214 -> op:*``. Formulas are
parsed into a real AST (shunting-yard over openpyxl's Tokenizer), and every edge
carries ``arg_index`` and ``role`` so that ``=A1-B1`` and ``=B1-A1`` are different
graphs. A formula that is nothing but a reference gets no operator node at all.

Outputs CSV, JSON, GraphML and a self-contained HTML viewer that draws one cell's
expression tree at a time.

Requires only openpyxl.
"""

from __future__ import annotations

import argparse
import bisect
import csv
import datetime as _dt
import json
import re
import sys
import time
import warnings
from collections import defaultdict
from pathlib import Path

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

try:
    import openpyxl
    from openpyxl.formula import Tokenizer
    from openpyxl.utils import get_column_letter, range_boundaries
    from openpyxl.worksheet.formula import ArrayFormula
except ImportError:  # pragma: no cover
    sys.exit("openpyxl is required:  python3 -m pip install openpyxl")


EXCEL_MAX_COL = 16384
EXCEL_MAX_ROW = 1048576

NODE_KINDS = ["formula", "input", "label", "op", "const", "range", "name", "external"]

ROLES = [
    "result", "identity", "constant", "lhs", "rhs", "operand", "summand", "arg",
    "condition", "then", "else", "criteria", "range", "element", "area", "index",
    "choice", "value", "fallback", "resolved",
]

# Arguments sitting directly inside one of these are `summand` edges.
AGGREGATION_FUNCS = {
    "SUM", "SUMIF", "SUMIFS", "SUMPRODUCT", "SUMSQ", "SUBTOTAL", "AGGREGATE",
    "AVERAGE", "AVERAGEA", "AVERAGEIF", "AVERAGEIFS", "MEDIAN", "COUNT",
    "COUNTA", "COUNTIF", "COUNTIFS", "MIN", "MINA", "MINIFS", "MAX", "MAXA",
    "MAXIFS", "PRODUCT", "NPV", "XNPV", "IRR", "XIRR", "MIRR", "STDEV",
    "STDEV.S", "STDEV.P", "STDEVA", "VAR", "VAR.S", "VAR.P", "MMULT",
}

# Strict shapes so a defined name is never mistaken for a range. openpyxl's
# range_boundaries("BC") happily returns column BC, which would be wrong.
_RE_CELL = re.compile(r"^\$?[A-Z]{1,3}\$?[0-9]{1,7}$", re.I)
_RE_RANGE = re.compile(r"^\$?[A-Z]{1,3}\$?[0-9]{1,7}:\$?[A-Z]{1,3}\$?[0-9]{1,7}$", re.I)
_RE_COLS = re.compile(r"^\$?[A-Z]{1,3}:\$?[A-Z]{1,3}$", re.I)
_RE_ROWS = re.compile(r"^\$?[0-9]{1,7}:\$?[0-9]{1,7}$")
_RE_EXTERNAL = re.compile(r"^\[([^\]]*)\]")


def is_range_like(body: str) -> bool:
    return bool(
        _RE_CELL.match(body)
        or _RE_RANGE.match(body)
        or _RE_COLS.match(body)
        or _RE_ROWS.match(body)
    )


def split_areas(ref: str):
    """Split a multi-area reference on top-level commas."""
    parts, depth, quoted, buf = [], 0, False, []
    for ch in ref:
        if ch == "'":
            quoted = not quoted
        if not quoted:
            if ch == "(":
                depth += 1
            elif ch == ")":
                depth -= 1
            elif ch == "," and depth == 0:
                parts.append("".join(buf))
                buf = []
                continue
        buf.append(ch)
    if buf:
        parts.append("".join(buf))
    return [p.strip() for p in parts if p.strip()]


def iter_cells(ws):
    """Yield only cells that actually exist.

    ws.iter_rows() walks the sheet's whole declared bounding box and instantiates a
    Cell for every empty coordinate inside it, which on sparse-but-large sheets means
    tens of millions of throwaway objects. The internal _cells dict is the populated
    set, so prefer it and keep iter_rows() as a fallback.
    """
    cells = getattr(ws, "_cells", None)
    if cells is not None:
        return list(cells.values())
    return [c for row in ws.iter_rows() for c in row]


def jsonable(value):
    if value is None or isinstance(value, (str, int, float, bool)):
        if isinstance(value, float) and (value != value or value in (float("inf"), float("-inf"))):
            return str(value)
        return value
    if isinstance(value, (_dt.datetime, _dt.date, _dt.time)):
        return value.isoformat()
    if isinstance(value, _dt.timedelta):
        return value.total_seconds()
    return str(value)


class SheetIndex:
    """Populated cells of one sheet, indexed for fast range intersection."""

    __slots__ = ("rows_by_col", "cols")

    def __init__(self):
        self.rows_by_col = defaultdict(list)
        self.cols = []

    def add(self, row, col):
        self.rows_by_col[col].append(row)

    def finalise(self):
        for rows in self.rows_by_col.values():
            rows.sort()
        self.cols = sorted(self.rows_by_col)

    def count(self, min_col, min_row, max_col, max_row):
        total = 0
        for col in self.cols:
            if col < min_col:
                continue
            if col > max_col:
                break
            rows = self.rows_by_col[col]
            total += bisect.bisect_right(rows, max_row) - bisect.bisect_left(rows, min_row)
        return total

    def cells(self, min_col, min_row, max_col, max_row):
        out = []
        for col in self.cols:
            if col < min_col:
                continue
            if col > max_col:
                break
            rows = self.rows_by_col[col]
            lo = bisect.bisect_left(rows, min_row)
            hi = bisect.bisect_right(rows, max_row)
            for i in range(lo, hi):
                out.append((rows[i], col))
        return out

    def clip(self, min_col, min_row, max_col, max_row):
        """Tight bounding box of the populated cells inside the given box.

        Whole-column references like ``A:I`` otherwise produce range nodes a
        million rows tall, which downstream consumers expand cell by cell.
        Returns ``None`` when nothing inside the box is populated.
        """
        lo_col = hi_col = None
        lo_row = hi_row = None
        for col in self.cols:
            if col < min_col:
                continue
            if col > max_col:
                break
            rows = self.rows_by_col[col]
            lo = bisect.bisect_left(rows, min_row)
            hi = bisect.bisect_right(rows, max_row)
            if lo >= hi:
                continue
            if lo_col is None:
                lo_col = col
            hi_col = col
            if lo_row is None or rows[lo] < lo_row:
                lo_row = rows[lo]
            if hi_row is None or rows[hi - 1] > hi_row:
                hi_row = rows[hi - 1]
        if lo_col is None:
            return None
        return lo_col, lo_row, hi_col, hi_row


# ---------------------------------------------------------------------------
# formula -> AST
# ---------------------------------------------------------------------------

# Operator precedence, lowest binds loosest. Excel puts unary minus above ^ so that
# -2^2 is 4, and the reference operators above everything.
PREC_INFIX = {
    "=": 1, "<": 1, ">": 1, "<=": 1, ">=": 1, "<>": 1,
    "&": 2,
    "+": 3, "-": 3,
    "*": 4, "/": 4,
    "^": 5,
    ":": 8, " ": 8, ",": 8,
}
PREC_POSTFIX = 6
PREC_PREFIX = 7
RIGHT_ASSOC = {"^"}

OP_KINDS = ("func", "infix", "prefix", "postfix", "array", "union")

# Argument roles for functions whose arguments are not interchangeable.
_FIXED_ROLES = {
    "IF": ("condition", "then", "else"),
    "IFERROR": ("value", "fallback"),
    "IFNA": ("value", "fallback"),
    "CHOOSE": ("index",),
    "SUMIF": ("range", "criteria", "summand"),
    "AVERAGEIF": ("range", "criteria", "summand"),
    "COUNTIF": ("range", "criteria"),
    "VLOOKUP": ("value", "range", "index", "criteria"),
    "HLOOKUP": ("value", "range", "index", "criteria"),
    "XLOOKUP": ("value", "range", "range", "fallback"),
    "INDEX": ("range", "index", "index"),
    "MATCH": ("value", "range", "criteria"),
    "OFFSET": ("range", "index", "index"),
    "ROUND": ("value", "index"),
    "ROUNDUP": ("value", "index"),
    "ROUNDDOWN": ("value", "index"),
}
_REST_ROLES = {"CHOOSE": "choice", "IF": "arg"}
# Functions taking (condition, result) pairs after an optional leading argument.
_ALTERNATING = {"IFS": (0, ("condition", "then")), "SWITCH": (1, ("criteria", "then"))}
_CRITERIA_PAIRS = {"SUMIFS", "AVERAGEIFS", "COUNTIFS", "MAXIFS", "MINIFS"}
_LOGICAL_FUNCS = {"AND", "OR", "NOT", "XOR"}


def role_for(kind, name, i, n):
    if kind == "infix":
        return "lhs" if i == 0 else "rhs"
    if kind in ("prefix", "postfix"):
        return "operand"
    if kind == "array":
        return "element"
    if kind == "union":
        return "area"
    fixed = _FIXED_ROLES.get(name)
    if fixed:
        if i < len(fixed):
            return fixed[i]
        return _REST_ROLES.get(name, "arg")
    alt = _ALTERNATING.get(name)
    if alt:
        offset, pair = alt
        if i < offset:
            return "value"
        return pair[(i - offset) % 2]
    if name in _CRITERIA_PAIRS:
        # COUNTIFS pairs from argument 0, the others start with the aggregated range.
        offset = 0 if name == "COUNTIFS" else 1
        if i < offset:
            return "summand"
        return ("range", "criteria")[(i - offset) % 2]
    if name in _LOGICAL_FUNCS:
        return "condition"
    if name in AGGREGATION_FUNCS:
        return "summand"
    return "arg"


class FormulaError(Exception):
    pass


class Ast:
    """One node of a parsed formula.

    `name` is the function name, the operator symbol, or the literal type. `value`
    holds a literal exactly as written; `shape` holds its parsed form for consts and
    the row count for array literals.
    """

    __slots__ = ("kind", "name", "args", "value", "shape")

    def __init__(self, kind, name="", args=(), value=None, shape=None):
        self.kind = kind
        self.name = name
        self.args = list(args)
        self.value = value
        self.shape = shape

    @property
    def is_op(self):
        return self.kind in OP_KINDS

    def tag(self):
        """The `<op>` part of the node id."""
        if self.kind == "func":
            return self.name
        if self.kind == "prefix":
            return "u" + self.name
        if self.kind == "array":
            return "{}"
        if self.kind == "union":
            return ","
        if self.kind == "infix" and self.name == " ":
            return "isect"
        return self.name

    def render(self, limit=200):
        text = self._render()
        return text if len(text) <= limit else text[:limit - 1] + "\u2026"

    def _render(self):
        if self.kind in ("ref", "name"):
            return self.name
        if self.kind == "const":
            return self.value if isinstance(self.value, str) else str(self.value)
        if self.kind == "missing":
            return ""
        if self.kind == "func":
            return "%s(%s)" % (self.name, ",".join(a._render() for a in self.args))
        if self.kind == "infix":
            joiner = self.name if self.name != " " else " "
            return "(%s)" % joiner.join(a._render() for a in self.args)
        if self.kind == "prefix":
            return "(%s%s)" % (self.name, self.args[0]._render())
        if self.kind == "postfix":
            return "(%s%s)" % (self.args[0]._render(), self.name)
        if self.kind == "union":
            return "(%s)" % ",".join(a._render() for a in self.args)
        if self.kind == "array":
            rows = self.shape or 1
            per = max(1, len(self.args) // max(1, rows))
            chunks = [",".join(a._render() for a in self.args[r * per:(r + 1) * per])
                      for r in range(rows)]
            return "{%s}" % ";".join(chunks)
        return self.name


def _const_from(token):
    """Build a const leaf from an OPERAND token, keeping the literal as written."""
    raw = token.value
    subtype = token.subtype
    if subtype == "NUMBER":
        try:
            parsed = float(raw)
            if parsed.is_integer() and "." not in raw and "e" not in raw.lower():
                parsed = int(parsed)
        except ValueError:
            parsed = raw
        return Ast("const", "number", value=raw, shape=parsed)
    if subtype == "TEXT":
        text = raw
        if len(text) >= 2 and text[0] == '"' and text[-1] == '"':
            text = text[1:-1].replace('""', '"')
        return Ast("const", "text", value=raw, shape=text)
    if subtype == "LOGICAL":
        return Ast("const", "logical", value=raw, shape=raw.upper() == "TRUE")
    if subtype == "ERROR":
        return Ast("const", "error", value=raw, shape=raw)
    return Ast("const", "text", value=raw, shape=raw)


_VALUE_END = ("OPERAND", "FUNC-CLOSE", "PAREN-CLOSE", "ARRAY-CLOSE", "OPERATOR-POSTFIX")
_VALUE_START = ("OPERAND", "FUNC-OPEN", "PAREN-OPEN", "ARRAY-OPEN", "OPERATOR-PREFIX")


def _slot(token):
    if token.type in ("FUNC", "PAREN", "ARRAY"):
        return "%s-%s" % (token.type, token.subtype)
    return token.type


def _normalise(tokens):
    """Drop cosmetic whitespace, but keep the whitespace that means intersection.

    Excel's intersection operator is a space between two operands. openpyxl reports it
    as a plain WHITE-SPACE token, indistinguishable at the token level from the padding
    in ``= V36 / SUM(...)``. It is only an operator where an infix operator could
    legally stand: after something that ends a value and before something that starts
    one.
    """
    kept = [t for t in tokens if t.type != "WHITE-SPACE"]
    if len(kept) == len(tokens):
        return [(t, False) for t in tokens]
    out = []
    prev = None
    for i, token in enumerate(tokens):
        if token.type != "WHITE-SPACE":
            out.append((token, False))
            prev = token
            continue
        if prev is None or _slot(prev) not in _VALUE_END:
            continue
        nxt = None
        for later in tokens[i + 1:]:
            if later.type != "WHITE-SPACE":
                nxt = later
                break
        if nxt is not None and _slot(nxt) in _VALUE_START:
            out.append((token, True))
            prev = token
    return out


class _Frame:
    """An open call, parenthesis or array literal on the operator stack."""

    __slots__ = ("kind", "name", "args", "start", "rows", "seen_sep")

    def __init__(self, kind, name, start):
        self.kind = kind
        self.name = name
        self.args = []
        self.start = start
        self.rows = []
        self.seen_sep = False


def parse_formula(text):
    """Shunting-yard over openpyxl's tokens -> Ast. Raises FormulaError."""
    try:
        tokens = Tokenizer(text).items
    except Exception as exc:  # openpyxl raises bare Exception on malformed input
        raise FormulaError(str(exc))

    out = []       # operand stack
    stack = []     # operators (dicts) and frames (_Frame)

    def floor():
        """Operands below the innermost open call belong to the enclosing expression."""
        for item in reversed(stack):
            if isinstance(item, _Frame):
                return item.start
        return 0

    def apply_operator(op):
        name, kind = op["op"], op["kind"]
        need = 2 if kind == "infix" else 1
        if len(out) - need < floor():
            raise FormulaError("operator %r is missing an operand" % name)
        if kind == "infix":
            right = out.pop()
            left = out.pop()
            out.append(Ast("infix", name, [left, right]))
        else:
            out.append(Ast(kind, name, [out.pop()]))

    def unwind(min_prec=None):
        while stack and isinstance(stack[-1], dict):
            top = stack[-1]
            if min_prec is not None:
                if top["prec"] < min_prec:
                    break
                if top["prec"] == min_prec and top["op"] in RIGHT_ASSOC:
                    break
            apply_operator(stack.pop())

    def collect_argument(frame):
        """Take the finished argument off the operand stack.

        Callers must have applied the argument's own operators first; unwinding here
        would also pop operators that belong outside the frame.
        """
        if len(out) > frame.start:
            value = out.pop()
            if len(out) > frame.start:
                raise FormulaError("unbalanced expression")
        elif frame.seen_sep or frame.kind == "func":
            value = Ast("missing")
        else:
            value = None
        frame.args.append(value)

    for token, is_intersect in _normalise(tokens):
        if is_intersect:
            unwind(PREC_INFIX[" "])
            stack.append({"op": " ", "kind": "infix", "prec": PREC_INFIX[" "]})
            continue

        ttype, subtype, value = token.type, token.subtype, token.value

        if ttype == "OPERAND":
            if subtype == "RANGE":
                out.append(Ast("ref", value))
            else:
                out.append(_const_from(token))
        elif ttype == "FUNC" and subtype == "OPEN":
            stack.append(_Frame("func", value.rstrip("( ").upper(), len(out)))
        elif ttype == "PAREN" and subtype == "OPEN":
            stack.append(_Frame("paren", "", len(out)))
        elif ttype == "ARRAY" and subtype == "OPEN":
            stack.append(_Frame("array", "", len(out)))
        elif ttype == "SEP":
            frame = next((f for f in reversed(stack) if isinstance(f, _Frame)), None)
            if frame is None:
                raise FormulaError("separator outside any call")
            unwind()
            collect_argument(frame)
            frame.seen_sep = True
            frame.start = len(out)
            if subtype == "ROW" and frame.kind == "array":
                frame.rows.append(len(frame.args))
        elif ttype in ("FUNC", "PAREN", "ARRAY") and subtype == "CLOSE":
            frame = None
            while stack:
                top = stack.pop()
                if isinstance(top, _Frame):
                    frame = top
                    break
                apply_operator(top)
            if frame is None:
                raise FormulaError("unbalanced closing bracket")
            collect_argument(frame)
            args = [a for a in frame.args if a is not None]
            if frame.kind == "func":
                if len(args) == 1 and args[0].kind == "missing" and not frame.seen_sep:
                    args = []
                out.append(Ast("func", frame.name, args))
            elif frame.kind == "array":
                frame.rows.append(len(frame.args))
                out.append(Ast("array", "{}", args, shape=max(1, len(frame.rows))))
            elif len(args) > 1:
                out.append(Ast("union", ",", args))
            elif args:
                out.append(args[0])  # plain grouping creates no node
            else:
                raise FormulaError("empty parentheses")
        elif ttype == "OPERATOR-PREFIX":
            stack.append({"op": value, "kind": "prefix", "prec": PREC_PREFIX})
        elif ttype == "OPERATOR-POSTFIX":
            if len(out) - 1 < floor():
                raise FormulaError("postfix %r is missing an operand" % value)
            out.append(Ast("postfix", value, [out.pop()]))
        elif ttype == "OPERATOR-INFIX":
            prec = PREC_INFIX.get(value)
            if prec is None:
                raise FormulaError("unknown operator %r" % value)
            unwind(prec)
            stack.append({"op": value, "kind": "infix", "prec": prec})
        elif ttype == "LITERAL":
            out.append(Ast("const", "text", value=value, shape=("text", value)))
        else:
            raise FormulaError("unexpected token %s/%s" % (ttype, subtype))

    unwind()
    if any(isinstance(f, _Frame) for f in stack):
        raise FormulaError("unclosed bracket")
    if len(out) != 1:
        raise FormulaError("expression left %d values on the stack" % len(out))
    return out[0]


# ---------------------------------------------------------------------------
# workbook -> expression DAG
# ---------------------------------------------------------------------------

class AstGraph:
    def __init__(self, path, max_range_expand=100, sheets=None, read_values=True,
                 consts=True, verbose=True):
        self.path = Path(path)
        self.max_range_expand = max_range_expand
        self.sheet_filter = set(sheets) if sheets else None
        self.read_values = read_values
        self.consts = consts
        self.verbose = verbose

        self.sheet_names = []
        self.index = {}
        self.values = {}
        self.formulas = {}
        self.text_cols_by_row = {}
        self.text_rows_by_col = {}
        self.names_global = {}
        self.names_local = {}

        self.nodes = {}
        self.edges = []
        self._edge_seen = set()
        self.cycles = []
        self.cyclic_nodes = set()
        self.warnings = []
        self.parsed = 0
        self.failed = 0
        self.elapsed = 0.0
        self.indirect_resolved = 0
        self.indirect_unresolved = 0
        self.adopted_static = 0

    def log(self, msg):
        if self.verbose:
            print("    " + msg, flush=True)

    # -- build -----------------------------------------------------------
    def build(self):
        started = time.time()
        self._load()
        self._collect_names()
        self._walk()
        self._adopt_static_cells()
        self._postprocess()
        self.elapsed = time.time() - started
        return self

    def _load(self):
        self.log("reading formulas...")
        wb = openpyxl.load_workbook(self.path, data_only=False)
        self.sheet_names = list(wb.sheetnames)

        for ws in wb.worksheets:
            idx = SheetIndex()
            formulas = {}
            for cell in iter_cells(ws):
                value = cell.value
                if value is None:
                    continue
                idx.add(cell.row, cell.column)
                if isinstance(value, ArrayFormula):
                    text = value.text
                    if isinstance(text, str) and text.startswith("="):
                        formulas[(cell.row, cell.column)] = (text, True)
                        # A multi-cell array (CSE or spill) formula lives only
                        # on its master cell; the member cells it fills carry
                        # nothing but a cached value in the file, and would
                        # otherwise be classified as hand-typed inputs -- a
                        # computed value handed to the rebuild as a given.
                        # Register every member as owning the same formula so
                        # it is a formula cell with real precedent edges.
                        ref = str(getattr(value, "ref", "") or "")
                        if ":" in ref:
                            min_col, min_row, max_col, max_row = \
                                range_boundaries(ref)
                            for member_row in range(min_row, max_row + 1):
                                for member_col in range(min_col, max_col + 1):
                                    spot = (member_row, member_col)
                                    if spot != (cell.row, cell.column):
                                        formulas.setdefault(spot, (text, True))
                elif cell.data_type == "f" and isinstance(value, str) and value.startswith("="):
                    formulas[(cell.row, cell.column)] = (value, False)
            idx.finalise()
            self.index[ws.title] = idx
            self.formulas[ws.title] = formulas

        self._wb = wb

        if not self.read_values:
            return

        self.log("reading cached values...")
        try:
            wbv = openpyxl.load_workbook(self.path, data_only=True)
        except Exception as exc:  # pragma: no cover
            self.warnings.append("could not read cached values: %s" % exc)
            return

        for ws in wbv.worksheets:
            vals = {}
            text_cols = defaultdict(list)
            text_rows = defaultdict(list)
            for cell in iter_cells(ws):
                value = cell.value
                if value is None:
                    continue
                vals[(cell.row, cell.column)] = value
                if isinstance(value, str) and 0 < len(value) <= 160:
                    text_cols[cell.row].append(cell.column)
                    text_rows[cell.column].append(cell.row)
            for cols in text_cols.values():
                cols.sort()
            for rows in text_rows.values():
                rows.sort()
            self.values[ws.title] = vals
            self.text_cols_by_row[ws.title] = text_cols
            self.text_rows_by_col[ws.title] = text_rows
        wbv.close()
        del wbv

    def _collect_names(self):
        wb = self._wb
        try:
            items = list(wb.defined_names.items())
        except AttributeError:  # pragma: no cover - very old openpyxl
            items = [(d.name, d) for d in wb.defined_names.definedName]
        for name, defn in items:
            value = getattr(defn, "value", None) or getattr(defn, "attr_text", None)
            if value:
                self.names_global[name.upper()] = value
        for ws in wb.worksheets:
            local = getattr(ws, "defined_names", None)
            if not local:
                continue
            try:
                local_items = list(local.items())
            except AttributeError:  # pragma: no cover
                continue
            for name, defn in local_items:
                value = getattr(defn, "value", None) or getattr(defn, "attr_text", None)
                if value:
                    self.names_local[(ws.title, name.upper())] = value

    # -- reference resolution -------------------------------------------
    def _split_ref(self, token):
        """-> (sheet_part or None, body, external_tag or None)"""
        if "!" in token:
            sheet_part, body = token.rsplit("!", 1)
            sheet_part = sheet_part.strip()
            if sheet_part.startswith("'") and sheet_part.endswith("'") and len(sheet_part) >= 2:
                sheet_part = sheet_part[1:-1].replace("''", "'")
            external = None
            match = _RE_EXTERNAL.match(sheet_part)
            if match:
                external = match.group(1)
                sheet_part = sheet_part[match.end():]
            return sheet_part, body.strip(), external
        return None, token.strip(), None

    def _sheet_span(self, first, last):
        try:
            i, j = self.sheet_names.index(first), self.sheet_names.index(last)
        except ValueError:
            return [s for s in (first, last) if s in self.index]
        if i > j:
            i, j = j, i
        return self.sheet_names[i:j + 1]

    def resolve(self, token, cur_sheet, depth=0):
        """Resolve a RANGE token to a list of targets.

        Defined names are looked up first, on purpose: range_boundaries("BC") returns
        column BC without complaint, so a name like BC would silently become a whole
        column instead of the single cell it points at.

        Each target is one of:
          ("cell", sheet, row, col)
          ("range", sheet, ref, n_cells)
          ("name", name)
          ("ext", token)
        """
        sheet_part, body, external = self._split_ref(token)

        if external is not None:
            return [("ext", token)]

        if sheet_part is None:
            key = body.upper()
            defn = self.names_local.get((cur_sheet, key))
            if defn is None:
                defn = self.names_global.get(key)
            if defn is not None and depth < 6:
                if "#REF" in defn:
                    return [("name", body)]
                out = []
                for area in split_areas(defn):
                    out.extend(self.resolve(area, cur_sheet, depth + 1))
                return out or [("name", body)]
            if not is_range_like(body):
                return [("name", body)]
            sheets = [cur_sheet]
        else:
            if not is_range_like(body):
                return [("name", token)]
            if ":" in sheet_part:
                first, last = sheet_part.split(":", 1)
                sheets = self._sheet_span(first.strip(), last.strip())
            else:
                sheets = [sheet_part]

        try:
            min_col, min_row, max_col, max_row = range_boundaries(body.replace("$", ""))
        except Exception:
            return [("name", token)]
        min_col = min_col or 1
        min_row = min_row or 1
        max_col = max_col or EXCEL_MAX_COL
        max_row = max_row or EXCEL_MAX_ROW

        out = []
        for sheet in sheets:
            idx = self.index.get(sheet)
            if idx is None:
                out.append(("name", token))
                continue
            n = idx.count(min_col, min_row, max_col, max_row)
            if n == 0:
                continue
            if n > self.max_range_expand:
                # Clamp to the populated bounding box so a whole-column A:I
                # does not become a range node a million rows tall.
                box = idx.clip(min_col, min_row, max_col, max_row)
                lo_col, lo_row, hi_col, hi_row = box if box else (
                    min_col, min_row, min(max_col, EXCEL_MAX_COL),
                    min(max_row, EXCEL_MAX_ROW))
                ref = "%s%d:%s%d" % (
                    get_column_letter(lo_col), lo_row,
                    get_column_letter(hi_col), hi_row,
                )
                out.append(("range", sheet, ref, n))
            else:
                for row, col in idx.cells(min_col, min_row, max_col, max_row):
                    out.append(("cell", sheet, row, col))
        return out

    # -- nodes -------------------------------------------------------------
    @staticmethod
    def cell_id(sheet, row, col):
        return "%s!%s%d" % (sheet, get_column_letter(col), row)

    def _blank_node(self, node_id, kind):
        node = {
            "id": node_id, "kind": kind, "sheet": "", "row": None, "col": None,
            "coordinate": "", "owner": "", "op": "", "op_kind": "", "arity": "",
            "expr": "", "label": "", "formula": "", "value": None,
            "array_formula": False,
        }
        self.nodes[node_id] = node
        return node

    def _cell_node(self, sheet, row, col):
        node_id = self.cell_id(sheet, row, col)
        node = self.nodes.get(node_id)
        if node is not None:
            return node
        entry = self.formulas.get(sheet, {}).get((row, col))
        value = self.values.get(sheet, {}).get((row, col)) if self.read_values else None
        if entry is not None:
            kind = "formula"
        elif isinstance(value, str):
            kind = "label"
        else:
            kind = "input"
        node = self._blank_node(node_id, kind)
        node.update({
            "sheet": sheet, "row": row, "col": col,
            "coordinate": "%s%d" % (get_column_letter(col), row),
            "formula": entry[0] if entry else "",
            "array_formula": bool(entry[1]) if entry else False,
            "value": jsonable(value),
            "label": self._label_for(sheet, row, col, value),
        })
        return node

    def _label_for(self, sheet, row, col, value):
        """Line-item name, any qualifier between it and the cell, and the column header.

        The row's leftmost text is preferred over the nearest text to the left, because
        in these models the leftmost cell holds the line item ("Revenues") while nearer
        cells hold qualifiers such as the currency ("AED").
        """
        if not self.read_values:
            return ""
        if isinstance(value, str) and value.strip():
            return value.strip()[:160]
        vals = self.values.get(sheet)
        if vals is None:
            return ""
        parts = []
        cols = self.text_cols_by_row.get(sheet, {}).get(row)
        if cols:
            i = bisect.bisect_left(cols, col) - 1
            if i >= 0:
                for j in (0, i):
                    text = vals.get((row, cols[j]))
                    if isinstance(text, str) and text.strip() and text.strip() not in parts:
                        parts.append(text.strip())
        rows = self.text_rows_by_col.get(sheet, {}).get(col)
        if rows:
            i = bisect.bisect_left(rows, row) - 1
            if i >= 0:
                text = vals.get((rows[i], col))
                if isinstance(text, str) and text.strip() and text.strip() not in parts:
                    parts.append(text.strip())
        return " / ".join(parts)[:160]

    def _add_edge(self, source, target, role, arg_index, op, ref, via_range, cell, sheet):
        key = (source, target, role, arg_index, ref)
        if key in self._edge_seen:
            return
        self._edge_seen.add(key)
        src_sheet = self.nodes[source].get("sheet") or ""
        self.edges.append({
            "source": source,
            "target": target,
            "role": role,
            "arg_index": arg_index,
            "op": op,
            "cell": cell,
            "ref": ref,
            "via_range": via_range,
            "cross_sheet": bool(src_sheet) and src_sheet != sheet,
        })

    # -- walk ---------------------------------------------------------------
    def _walk(self):
        self.log("parsing formulas into expression trees...")
        unknown = sorted(self.sheet_filter - set(self.sheet_names)) if self.sheet_filter else []
        if unknown:
            print("    no sheet named %s in this workbook; it has %s"
                  % (", ".join(repr(name) for name in unknown),
                     ", ".join(repr(name) for name in self.sheet_names)))
        unresolved = 0
        broken = 0
        for sheet in self.sheet_names:
            if self.sheet_filter and sheet not in self.sheet_filter:
                continue
            for (row, col), (formula, _) in self.formulas.get(sheet, {}).items():
                cell_id = self.cell_id(sheet, row, col)
                try:
                    ast = parse_formula(formula)
                except FormulaError as exc:
                    self.failed += 1
                    if len(self.warnings) < 25:
                        self.warnings.append("%s: could not parse %r (%s)"
                                             % (cell_id, formula[:60], exc))
                    continue
                self.parsed += 1
                self._cell_node(sheet, row, col)
                ctx = {"cell": cell_id, "sheet": sheet, "n": 0,
                       "unresolved": 0, "broken": 0}
                sources = self._emit(ast, ctx)
                if ast.is_op:
                    role = "result"
                elif ast.kind == "const":
                    role = "constant"
                else:
                    role = "identity"
                for src, ref, via in sources:
                    self._add_edge(src, cell_id, role, 0, ast.tag() if ast.is_op else "",
                                   ref, via, cell_id, sheet)
                unresolved += ctx["unresolved"]
                broken += ctx["broken"]

        if broken:
            self.warnings.append("%d formula reference(s) are broken #REF! links in the "
                                 "workbook itself" % broken)
        if unresolved:
            self.warnings.append("%d reference(s) could not be resolved to cells, e.g. table "
                                 "or external names (kept as `name` nodes)" % unresolved)
        if self.failed:
            self.warnings.append("%d formula(s) could not be parsed and were skipped"
                                 % self.failed)
        if self.indirect_resolved:
            self.log("resolved %d INDIRECT reference(s) from cached values"
                     % self.indirect_resolved)
        if self.indirect_unresolved:
            self.warnings.append("%d INDIRECT reference(s) could not be resolved from "
                                 "cached values; their true precedents are missing from "
                                 "the graph" % self.indirect_unresolved)

    def _adopt_static_cells(self):
        """Give every populated non-formula cell a node, referenced or not.

        The walk only creates nodes for cells some formula mentions, so a
        typed value nothing references never enters the graph and can never be
        classified. Downstream, the input mask must know about every supplied
        value -- an invisible cell is indistinguishable from a derived one.
        """
        if not self.read_values:
            return
        for sheet in self.sheet_names:
            if self.sheet_filter and sheet not in self.sheet_filter:
                continue
            formulas = self.formulas.get(sheet, {})
            for (row, col) in self.values.get(sheet, {}):
                if (row, col) in formulas:
                    continue
                if self.cell_id(sheet, row, col) not in self.nodes:
                    self._cell_node(sheet, row, col)
                    self.adopted_static += 1
        if self.adopted_static:
            self.log("adopted %d typed cell(s) no formula references"
                     % self.adopted_static)

    def _new_id(self, ctx, tag):
        node_id = "%s#%d:%s" % (ctx["cell"], ctx["n"], tag)
        ctx["n"] += 1
        return node_id

    # -- static resolution of dynamic references -----------------------------
    @staticmethod
    def _concat_text(value):
        """Render a cached value the way Excel's ``&`` would."""
        if isinstance(value, bool):
            return "TRUE" if value else "FALSE"
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value)

    def _static_value(self, node, sheet, depth=0):
        """Evaluate an AST subtree when cached values fully determine it.

        INDIRECT builds its reference out of text at run time, which is
        invisible to a static parse -- the single biggest source of missing
        edges. But the pieces of that text (constants, ``&`` concatenations
        and single-cell references) are all knowable from the cached values,
        so the reference string can usually be assembled here and turned into
        real edges. Returns ``None`` when the subtree is not decidable.
        """
        if node is None or depth > 8:
            return None
        if node.kind == "const":
            return node.shape
        if node.kind == "ref":
            items = self.resolve(node.name, sheet)
            if len(items) == 1 and items[0][0] == "cell":
                _, tgt_sheet, row, col = items[0]
                return self.values.get(tgt_sheet, {}).get((row, col))
            return None
        if ((node.kind == "infix" and node.name == "&")
                or (node.kind == "func" and node.name in ("CONCATENATE", "CONCAT"))):
            parts = [self._static_value(arg, sheet, depth + 1) for arg in node.args]
            if any(part is None for part in parts):
                return None
            return "".join(self._concat_text(part) for part in parts)
        return None

    def _emit(self, node, ctx):
        """Emit one AST node; returns [(source_id, ref_text, via_range)]."""
        if node.kind == "ref":
            return self._emit_ref(node, ctx)
        if node.kind == "const":
            if not self.consts:
                return []
            node_id = self._new_id(ctx, "const")
            const = self._blank_node(node_id, "const")
            const.update({
                "sheet": ctx["sheet"], "owner": ctx["cell"], "op": node.name,
                "op_kind": "const", "expr": node.render(),
                "value": jsonable(node.shape), "label": node.render(60),
            })
            return [(node_id, "", "")]
        if node.kind == "missing":
            return []
        if not node.is_op:  # defensive: unknown leaf
            return []

        # children first, so the innermost operator is #0
        child_sources = [self._emit(arg, ctx) for arg in node.args]
        node_id = self._new_id(ctx, node.tag())
        op = self._blank_node(node_id, "op")
        op.update({
            "sheet": ctx["sheet"], "owner": ctx["cell"], "op": node.tag(),
            "op_kind": node.kind, "arity": len(node.args), "expr": node.render(),
            "label": node.tag(),
        })
        total = len(node.args)
        for i, sources in enumerate(child_sources):
            role = role_for(node.kind, node.name, i, total)
            for src, ref, via in sources:
                self._add_edge(src, node_id, role, i, node.tag(), ref, via,
                               ctx["cell"], ctx["sheet"])

        # INDIRECT names its target as text assembled at run time, so the
        # ordinary walk records a dependency on the text's ingredients but not
        # on the cells the text points at. When cached values pin the string
        # down, resolve it here and wire the real precedents in.
        if node.kind == "func" and node.name == "INDIRECT" and node.args:
            resolved = self._static_value(node.args[0], ctx["sheet"])
            if isinstance(resolved, str) and resolved.strip():
                token = resolved.strip()
                for src, ref, via in self._emit_ref(Ast("ref", token), ctx):
                    self._add_edge(src, node_id, "resolved", total, node.tag(),
                                   ref, via or token, ctx["cell"], ctx["sheet"])
                self.indirect_resolved += 1
            else:
                self.indirect_unresolved += 1
                ctx["unresolved"] += 1
        return [(node_id, "", "")]

    def _emit_ref(self, node, ctx):
        token = node.name
        items = self.resolve(token, ctx["sheet"])
        multi = len(items) > 1 or (bool(items) and items[0][0] == "range")
        via = token if multi else ""
        out = []
        for item in items:
            if item[0] == "cell":
                _, sheet, row, col = item
                src = self._cell_node(sheet, row, col)
            elif item[0] == "range":
                _, sheet, ref, n = item
                node_id = "%s!%s" % (sheet, ref)
                src = self.nodes.get(node_id)
                if src is None:
                    bounds = range_boundaries(ref)
                    src = self._blank_node(node_id, "range")
                    src.update({"sheet": sheet, "row": bounds[1], "col": bounds[0],
                                "coordinate": ref, "value": n,
                                "label": "%s (%d cells)" % (ref, n)})
            elif item[0] == "name":
                node_id = "name:%s" % item[1]
                src = self.nodes.get(node_id) or self._blank_node(node_id, "name")
                src["label"] = item[1]
                if "#REF" in item[1].upper():
                    ctx["broken"] += 1
                else:
                    ctx["unresolved"] += 1
            else:
                node_id = "external:%s" % item[1]
                src = self.nodes.get(node_id) or self._blank_node(node_id, "external")
                src["label"] = item[1]
            out.append((src["id"], token, via))
        return out

    # -- post-processing ----------------------------------------------------
    def _postprocess(self):
        self.log("computing degrees and cycles...")
        for node in self.nodes.values():
            node["in_degree"] = 0
            node["out_degree"] = 0
        succ = defaultdict(list)
        for edge in self.edges:
            self.nodes[edge["target"]]["in_degree"] += 1
            self.nodes[edge["source"]]["out_degree"] += 1
            succ[edge["source"]].append(edge["target"])

        self.cycles = self._tarjan(succ)
        cyclic = {node_id for group in self.cycles for node_id in group}
        for node in self.nodes.values():
            node["in_cycle"] = node["id"] in cyclic
        for edge in self.edges:
            edge["in_cycle"] = edge["source"] in cyclic and edge["target"] in cyclic
        self.cyclic_nodes = cyclic

    def _tarjan(self, succ):
        """Iterative Tarjan; returns one node list per circular reference.

        Each list is a strongly connected component that a value can actually loop
        around: either several nodes feeding each other, or a single node feeding
        itself. Everything else is a component of one node with no self-edge.
        """
        index_of, low, on_stack, stack = {}, {}, set(), []
        counter = [0]
        cycles = []
        for root in self.nodes:
            if root in index_of:
                continue
            work = [(root, 0)]
            while work:
                node, pi = work[-1]
                if pi == 0:
                    index_of[node] = low[node] = counter[0]
                    counter[0] += 1
                    stack.append(node)
                    on_stack.add(node)
                recursed = False
                neighbours = succ.get(node, ())
                for i in range(pi, len(neighbours)):
                    nxt = neighbours[i]
                    if nxt not in index_of:
                        work[-1] = (node, i + 1)
                        work.append((nxt, 0))
                        recursed = True
                        break
                    if nxt in on_stack:
                        low[node] = min(low[node], index_of[nxt])
                if recursed:
                    continue
                if low[node] == index_of[node]:
                    component = []
                    while True:
                        w = stack.pop()
                        on_stack.discard(w)
                        component.append(w)
                        if w == node:
                            break
                    if len(component) > 1 or node in succ.get(node, ()):
                        cycles.append(component)
                work.pop()
                if work:
                    parent = work[-1][0]
                    low[parent] = min(low[parent], low[node])
        return cycles

    # -- views ---------------------------------------------------------------
    def match_node(self, node_id):
        if node_id in self.nodes:
            return node_id
        lowered = node_id.lower()
        for candidate in self.nodes:
            if candidate.lower() == lowered:
                return candidate
        raise KeyError(node_id)

    def _adjacency(self):
        preds, succs = defaultdict(list), defaultdict(list)
        for edge in self.edges:
            preds[edge["target"]].append(edge["source"])
            succs[edge["source"]].append(edge["target"])
        return preds, succs

    def _intermediate(self, node_id):
        node = self.nodes.get(node_id)
        return bool(node) and node["kind"] in ("op", "const")

    def _spread(self, start, depth, adjacency):
        """Walk `depth` cell hops; operator nodes are free and never cost a hop."""
        keep = {start}
        frontier = [start]
        expanded = set()
        for _ in range(max(0, depth)):
            next_cells = []
            for cell in frontier:
                if cell in expanded:
                    continue
                expanded.add(cell)
                stack = [cell]
                local = set()
                while stack:
                    current = stack.pop()
                    for other in adjacency.get(current, ()):
                        keep.add(other)
                        if self._intermediate(other):
                            if other not in local:
                                local.add(other)
                                stack.append(other)
                        elif other not in expanded:
                            next_cells.append(other)
            frontier = next_cells
            if not frontier:
                break
        return keep

    def focus(self, node_id, depth, direction="up"):
        node_id = self.match_node(node_id)
        preds, succs = self._adjacency()
        keep = set()
        if direction in ("up", "both"):
            keep |= self._spread(node_id, depth, preds)
        if direction in ("down", "both"):
            keep |= self._spread(node_id, depth, succs)
        keep.add(node_id)
        nodes = {k: v for k, v in self.nodes.items() if k in keep}
        edges = [e for e in self.edges if e["source"] in keep and e["target"] in keep]
        return node_id, nodes, edges

    def stats(self, nodes, edges):
        by_kind = defaultdict(int)
        for node in nodes.values():
            by_kind[node["kind"]] += 1
        by_role = defaultdict(int)
        by_op = defaultdict(int)
        cross = 0
        for edge in edges:
            by_role[edge["role"]] += 1
            if edge["cross_sheet"]:
                cross += 1
        for node in nodes.values():
            if node["kind"] == "op":
                by_op[node["op"]] += 1
        cycles = [g for g in self.cycles if any(k in nodes for k in g)]
        return {
            "workbook": self.path.name,
            "sheets": self.sheet_names,
            "nodes": len(nodes),
            "edges": len(edges),
            "formulas_parsed": self.parsed,
            "formulas_failed": self.failed,
            "nodes_by_kind": dict(by_kind),
            "edges_by_role": dict(sorted(by_role.items(), key=lambda kv: -kv[1])),
            "top_operators": dict(sorted(by_op.items(), key=lambda kv: -kv[1])[:15]),
            "cross_sheet_edges": cross,
            "cycles": len(cycles),
            "largest_cycle": max((len(g) for g in cycles), default=0),
            "cyclic_nodes": sum(1 for n in nodes.values() if n.get("in_cycle")),
            "max_range_expand": self.max_range_expand,
            "warnings": self.warnings,
        }


# ---------------------------------------------------------------------------
# writers
# ---------------------------------------------------------------------------

NODE_FIELDS = ["id", "kind", "sheet", "coordinate", "row", "col", "owner", "op",
               "op_kind", "arity", "expr", "label", "formula", "value",
               "array_formula", "in_degree", "out_degree", "in_cycle"]
EDGE_FIELDS = ["source", "target", "role", "arg_index", "op", "cell", "ref",
               "via_range", "cross_sheet", "in_cycle"]


def write_csv(out_dir, nodes, edges):
    with open(out_dir / "nodes.csv", "w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=NODE_FIELDS, extrasaction="ignore")
        writer.writeheader()
        for node in nodes.values():
            writer.writerow(node)
    with open(out_dir / "edges.csv", "w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=EDGE_FIELDS, extrasaction="ignore")
        writer.writeheader()
        for edge in edges:
            writer.writerow(edge)


def write_json(out_dir, nodes, edges, stats):
    payload = {"stats": stats, "nodes": list(nodes.values()), "edges": edges}
    with open(out_dir / "graph.json", "w", encoding="utf-8") as fh:
        json.dump(payload, fh, indent=1, default=str)


_GRAPHML_NODE_TYPES = {
    "kind": "string", "sheet": "string", "coordinate": "string", "row": "int",
    "col": "int", "owner": "string", "op": "string", "op_kind": "string",
    "arity": "int", "expr": "string", "label": "string", "formula": "string",
    "value": "string", "array_formula": "boolean", "in_degree": "int",
    "out_degree": "int", "in_cycle": "boolean",
}
_GRAPHML_EDGE_TYPES = {
    "role": "string", "arg_index": "int", "op": "string", "cell": "string",
    "ref": "string", "via_range": "string", "cross_sheet": "boolean",
    "in_cycle": "boolean",
}


def _xml_escape(value):
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    text = str(value)
    text = text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    text = text.replace('"', "&quot;")
    # strip control characters XML forbids
    return "".join(ch for ch in text if ch >= " " or ch in "\t\n\r")


def write_graphml(out_dir, nodes, edges):
    """Hand-rolled GraphML so the script needs no networkx."""
    with open(out_dir / "graph.graphml", "w", encoding="utf-8") as fh:
        fh.write('<?xml version="1.0" encoding="UTF-8"?>\n')
        fh.write('<graphml xmlns="http://graphml.graphdrawing.org/xmlns" '
                 'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" '
                 'xsi:schemaLocation="http://graphml.graphdrawing.org/xmlns '
                 'http://graphml.graphdrawing.org/xmlns/1.0/graphml.xsd">\n')
        nkeys, ekeys = {}, {}
        for i, (name, typ) in enumerate(_GRAPHML_NODE_TYPES.items()):
            key = "n%d" % i
            nkeys[name] = key
            fh.write('  <key id="%s" for="node" attr.name="%s" attr.type="%s"/>\n'
                     % (key, name, typ))
        for i, (name, typ) in enumerate(_GRAPHML_EDGE_TYPES.items()):
            key = "e%d" % i
            ekeys[name] = key
            fh.write('  <key id="%s" for="edge" attr.name="%s" attr.type="%s"/>\n'
                     % (key, name, typ))
        fh.write('  <graph id="G" edgedefault="directed">\n')
        for node in nodes.values():
            fh.write('    <node id="%s">\n' % _xml_escape(node["id"]))
            for name, key in nkeys.items():
                value = node.get(name)
                if value is None or value == "":
                    continue
                fh.write('      <data key="%s">%s</data>\n' % (key, _xml_escape(value)))
            fh.write("    </node>\n")
        for i, edge in enumerate(edges):
            fh.write('    <edge id="e%d" source="%s" target="%s">\n'
                     % (i, _xml_escape(edge["source"]), _xml_escape(edge["target"])))
            for name, key in ekeys.items():
                value = edge.get(name)
                if value is None or value == "":
                    continue
                fh.write('      <data key="%s">%s</data>\n' % (key, _xml_escape(value)))
            fh.write("    </edge>\n")
        fh.write("  </graph>\n</graphml>\n")


def pick_default_focus(nodes, edges):
    """A formula cell with a big tree makes the most useful first screen."""
    size = defaultdict(int)
    for edge in edges:
        if edge["cell"]:
            size[edge["cell"]] += 1
    for cell, _ in sorted(size.items(), key=lambda kv: (-kv[1], kv[0])):
        if cell in nodes:
            return cell
    for node_id, node in nodes.items():
        if node["kind"] == "formula":
            return node_id
    return next(iter(nodes), "")


def build_viewer_payload(graph, nodes, edges, stats, focus, depth, truncate, filtered):
    kidx = {k: i for i, k in enumerate(NODE_KINDS)}
    roles = sorted({e["role"] for e in edges}) or list(ROLES)
    ridx = {r: i for i, r in enumerate(roles)}
    present = {n.get("sheet") for n in nodes.values() if n.get("sheet")}
    sheets = [s for s in graph.sheet_names if s in present]
    sheets += sorted(present - set(sheets))
    order = {}
    out_nodes = []
    for i, (node_id, node) in enumerate(nodes.items()):
        order[node_id] = i
        value = node.get("value")
        out_nodes.append([
            node_id,
            kidx.get(node["kind"], 0),
            node.get("sheet") or "",
            node.get("label") or "",
            (node.get("formula") or "")[:truncate],
            "" if value is None else str(value)[:80],
            (node.get("expr") or "")[:truncate],
            node.get("op") or "",
            node.get("owner") or "",
            1 if node.get("in_cycle") else 0,
            node.get("row") if node.get("row") is not None else -1,
            node.get("col") if node.get("col") is not None else -1,
        ])
    out_edges = []
    for edge in edges:
        out_edges.append([
            order[edge["source"]], order[edge["target"]],
            ridx.get(edge["role"], 0), edge["arg_index"],
            edge.get("via_range") or edge.get("ref") or "",
        ])
    return {
        "workbook": graph.path.name,
        "kinds": NODE_KINDS,
        "roles": roles,
        "sheets": sheets,
        "stats": stats,
        "focus": order.get(focus, -1),
        "depth": depth,
        "filtered": bool(filtered),
        "nodes": out_nodes,
        "edges": out_edges,
    }


def write_html(out_dir, graph, nodes, edges, stats, focus, depth, truncate, filtered=False):
    payload = build_viewer_payload(graph, nodes, edges, stats, focus, depth, truncate,
                                   filtered)
    data = json.dumps(payload, separators=(",", ":"), default=str).replace("</", "<\\/")
    html = HTML_TEMPLATE.replace("__TITLE__", _xml_escape(graph.path.name))
    html = html.replace("__GRAPH_DATA__", data)
    with open(out_dir / "graph.html", "w", encoding="utf-8") as fh:
        fh.write(html)


HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>__TITLE__ &middot; expression trees</title>
<style>
  :root{
    --bg:#0e1117; --panel:#161b22; --panel2:#1c2230; --line:#2a3140;
    --fg:#e6edf3; --dim:#8b949e; --accent:#58a6ff;
  }
  *{box-sizing:border-box}
  html,body{margin:0;height:100%;background:var(--bg);color:var(--fg);
    font:13px/1.5 -apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,Helvetica,Arial,sans-serif}
  #app{display:flex;height:100%}
  #side{width:330px;min-width:330px;background:var(--panel);border-right:1px solid var(--line);
    display:flex;flex-direction:column;overflow:hidden}
  #side header{padding:14px 16px;border-bottom:1px solid var(--line)}
  #side h1{margin:0;font-size:14px;font-weight:600;letter-spacing:.2px}
  #side h1 span{display:block;font-weight:400;color:var(--dim);font-size:11px;margin-top:3px}
  .scroll{overflow-y:auto;flex:1}
  .sec{padding:12px 16px;border-bottom:1px solid var(--line)}
  .sec h2{margin:0 0 8px;font-size:11px;text-transform:uppercase;letter-spacing:.08em;
    color:var(--dim);font-weight:600}
  input[type=search]{width:100%;padding:7px 9px;background:var(--bg);color:var(--fg);
    border:1px solid var(--line);border-radius:6px;outline:none;font-size:12px}
  input[type=search]:focus{border-color:var(--accent)}
  label.row{display:flex;align-items:center;gap:7px;padding:2px 0;cursor:pointer;font-size:12px}
  label.row:hover{color:#fff}
  input[type=checkbox]{accent-color:var(--accent);margin:0}
  .btn{display:inline-block;padding:6px 10px;background:var(--panel2);border:1px solid var(--line);
    color:var(--fg);border-radius:6px;cursor:pointer;font-size:12px;user-select:none}
  .btn:hover{border-color:var(--accent);color:#fff}
  .btn.on{background:var(--accent);border-color:var(--accent);color:#04121f;font-weight:600}
  .btns{display:flex;gap:6px;flex-wrap:wrap}
  .kv{display:flex;justify-content:space-between;gap:8px;font-size:12px;padding:1px 0}
  .kv span:last-child{color:var(--dim);font-variant-numeric:tabular-nums}
  .hit{padding:4px 6px;border-radius:5px;cursor:pointer;font-size:12px;white-space:nowrap;
    overflow:hidden;text-overflow:ellipsis}
  .hit:hover{background:var(--panel2)}
  .hit em{color:var(--dim);font-style:normal}
  #results{margin-top:8px;max-height:210px;overflow-y:auto}
  #lineage{max-height:170px;overflow-y:auto}
  code{display:block;background:var(--bg);border:1px solid var(--line);border-radius:5px;
    padding:6px 8px;margin:5px 0;font-size:11px;word-break:break-all;color:#7ee787;
    font-family:ui-monospace,SFMono-Regular,Menlo,monospace}
  select{width:100%;padding:6px 8px;background:var(--bg);color:var(--fg);
    border:1px solid var(--line);border-radius:6px;outline:none;font-size:12px}
  #stage{flex:1;position:relative;overflow:hidden}
  svg,canvas{display:block;position:absolute;inset:0;width:100%;height:100%;
    cursor:grab;user-select:none}
  svg.drag,canvas.drag{cursor:grabbing}
  .nlabel{font-size:11.5px;fill:#e6edf3;font-family:ui-monospace,SFMono-Regular,Menlo,monospace}
  .nsub{font-size:10px;fill:#8b949e}
  .oplabel{font-size:12px;font-weight:600;fill:#0e1117;
    font-family:ui-monospace,SFMono-Regular,Menlo,monospace}
  .rlabel{font-size:9.5px;fill:#8b949e;letter-spacing:.03em}
  .hitbox{cursor:pointer}
  #tip{position:absolute;pointer-events:none;background:#0b0f16f2;border:1px solid var(--line);
    border-radius:6px;padding:7px 9px;font-size:12px;max-width:420px;display:none;z-index:5}
  #tip code{margin:4px 0 0}
  #banner{position:absolute;left:50%;top:14px;transform:translateX(-50%);background:#3d2b12ee;
    border:1px solid #7a5a26;color:#f0c674;padding:8px 13px;border-radius:7px;font-size:12px;
    display:none;z-index:6;max-width:70%}
  #hint{position:absolute;left:14px;bottom:12px;color:var(--dim);font-size:11px;z-index:3}
  .muted{color:var(--dim)}
  .swatch{width:9px;height:9px;border-radius:2px;flex:none;display:inline-block;margin-right:6px}
  #compList{max-height:250px;overflow-y:auto}
  .cmp{display:flex;align-items:center;gap:7px;padding:2px 4px;border-radius:5px;font-size:11.5px}
  .cmp:hover{background:var(--panel2)}
  .cmp input{flex:none}
  .cmp b{font-weight:600;flex:1;min-width:0;cursor:pointer;overflow:hidden;
    text-overflow:ellipsis;white-space:nowrap}
  .cmp b:hover{color:var(--accent)}
  .cmp i{font-style:normal;color:var(--dim);flex:none;font-variant-numeric:tabular-nums}
</style>
</head>
<body>
<div id="app">
  <aside id="side">
    <header><h1>__TITLE__<span id="sub"></span></h1></header>
    <div class="scroll">
      <div class="sec">
        <h2>View</h2>
        <div class="btns">
          <span class="btn on" id="bMap">Whole map</span>
          <span class="btn" id="bTree">One cell</span>
          <span class="btn" id="bFit">Fit</span>
        </div>
      </div>
      <div class="sec" id="secMap">
        <h2>Map</h2>
        <select id="sheetSel"></select>
        <div class="btns" style="margin-top:8px">
          <span class="btn on" id="bGrid">Spreadsheet</span>
          <span class="btn" id="bFlow">Dependency flow</span>
          <span class="btn" id="bComp" title="One panel per group of cells that shares no calculation with any other group">Independent graphs</span>
        </div>
        <div class="btns" style="margin-top:6px">
          <span class="btn on" id="bSrc" title="Nothing in view feeds these: inputs, constants and hardcodes">Initial</span>
          <span class="btn on" id="bSink" title="Nothing in view reads these: reported outputs and dead cells">Terminal</span>
        </div>
        <div style="margin-top:8px">
          <label class="row"><input type="checkbox" id="cOps"> Show operator nodes</label>
        </div>
        <div id="mapInfo" class="muted" style="margin-top:6px;font-size:11.5px"></div>
        <div id="mapHint" class="muted" style="margin-top:4px;font-size:11px">cells sit where they do in the spreadsheet, one lane per sheet</div>
      </div>
      <div class="sec" id="secComp">
        <h2>Independent graphs</h2>
        <div class="btns">
          <span class="btn" id="bCompAll">Show all</span>
          <span class="btn" id="bCompNone">Hide all</span>
          <span class="btn" id="bCompGiant">Largest only</span>
        </div>
        <div id="compInfo" class="muted" style="margin-top:7px;font-size:11.5px"></div>
        <div class="cmp muted" style="margin-top:4px">
          <span style="width:13px;flex:none"></span>
          <b style="font-weight:400">graph</b><i>cells &middot; edges</i>
        </div>
        <div id="compList"></div>
      </div>
      <div class="sec">
        <h2 id="focusTitle">Selected cell</h2>
        <div id="focusName" style="font-weight:600;word-break:break-all"></div>
        <div id="focusLabel" class="muted" style="font-size:11.5px"></div>
        <code id="focusFormula"></code>
        <div class="btns" style="margin-top:6px">
          <span class="btn" id="bOpen">Open its tree</span>
          <span class="btn" id="bBack">Back</span>
        </div>
      </div>
      <div class="sec">
        <h2>Find and highlight</h2>
        <input type="search" id="q" autocomplete="off"
               placeholder="Valuation!C103, Valuation!O102">
        <div class="btns" style="margin-top:6px">
          <span class="btn" id="bZoom">Zoom to matches</span>
          <span class="btn" id="bClear">Clear</span>
        </div>
        <div style="margin-top:6px">
          <label class="row"><input type="checkbox" id="cDim" checked> Fade the rest of the map</label>
        </div>
        <div id="qinfo" class="muted" style="margin-top:6px;font-size:11.5px"></div>
        <div id="qhint" class="muted" style="margin-top:4px;font-size:11px">comma-separated terms &middot; matches address, label, formula and value &middot; prefix with <b>value:</b> <b>formula:</b> <b>label:</b> or <b>cell:</b> to narrow</div>
        <div id="results"></div>
      </div>
      <div class="sec" id="secTree">
        <h2>Tree</h2>
        <label class="row" style="justify-content:space-between">
          <span>Depth (cell hops)</span><span id="dv" class="muted"></span>
        </label>
        <input type="range" id="depth" min="1" max="6" value="2" style="width:100%">
        <div class="btns" style="margin-top:8px">
          <span class="btn on" id="bUp">Precedents</span>
          <span class="btn" id="bDown">Dependents</span>
        </div>
        <div style="margin-top:8px">
          <label class="row"><input type="checkbox" id="cConst" checked> Show constants</label>
          <label class="row"><input type="checkbox" id="cRoles" checked> Show edge roles</label>
        </div>
      </div>
      <div class="sec"><h2 id="lineageTitle">Built from</h2><div id="lineage"></div></div>
      <div class="sec"><h2>Legend</h2><div id="legend"></div></div>
      <div class="sec"><h2>Workbook</h2><div id="stats"></div></div>
    </div>
  </aside>
  <main id="stage">
    <canvas id="map"></canvas>
    <svg id="cv" style="display:none"><g id="root"></g></svg>
    <div id="tip"></div>
    <div id="banner"></div>
    <div id="hint"></div>
  </main>
</div>
<script>const DATA = __GRAPH_DATA__;</script>
<script>
(function(){
"use strict";

var KINDS = DATA.kinds, ROLES = DATA.roles, N = DATA.nodes, E = DATA.edges;
var KIND_COLOR = {formula:"#58a6ff", input:"#f0c674", label:"#7d8590", op:"#e3924b",
                  const:"#9aa4b2", range:"#bc8cff", name:"#39d3c3", external:"#ff7b72"};
var ROLE_COLOR = {summand:"#3fb950", lhs:"#58a6ff", rhs:"#58a6ff", condition:"#ff7b72",
                  then:"#3fb950", else:"#e3924b", identity:"#3fb950", result:"#4f7dbb",
                  operand:"#bc8cff", criteria:"#ff7b72"};
var HIT_COLOR = "#ff6ac1";
// Roles that repeat inside one call, where the argument number carries the meaning.
var POSITIONAL = {arg:1, summand:1, element:1, area:1, choice:1, criteria:1, range:1};
var NS = "http://www.w3.org/2000/svg";
var NODE_H = 36, COL_W = 214, X_GAP = 78, Y_GAP = 9;
var CELL_W = 200, OP_W = 116, CONST_W = 116, MAX_NODES = 900;

var preds = [], succs = [];
for (var i = 0; i < N.length; i++){ preds.push([]); succs.push([]); }
E.forEach(function(e, i){ preds[e[1]].push(i); succs[e[0]].push(i); });

function kind(i){ return KINDS[N[i][1]]; }
function isInter(i){ var k = kind(i); return k === "op" || k === "const"; }
function owner(i){ return N[i][8] || N[i][0]; }
function role(ei){ return ROLES[E[ei][2]] || "arg"; }

var byId = {};
N.forEach(function(n, i){ byId[n[0]] = i; });
function byIdLookup(id){ return id in byId ? byId[id] : -1; }

var focus = DATA.focus >= 0 ? DATA.focus : 0;
if (isInter(focus) && byIdLookup(owner(focus)) >= 0) focus = byIdLookup(owner(focus));
var depth = DATA.depth || 2, dir = "up", showConst = true, showRoles = true;
var mode = "map", showOps = false, layout = "grid";
var showSources = true, showSinks = true;
var hiddenComps = {}, compGroupHidden = true;
var matches = null, matchList = [], dimOthers = true;
var history = [];
var view = {x: 60, y: 60, k: 1};
var tree = null, truncated = false;

var svg = document.getElementById("cv"), rootG = document.getElementById("root");
var cvs = document.getElementById("map"), ctx = cvs.getContext("2d");
var stage = document.getElementById("stage");
var tip = document.getElementById("tip"), banner = document.getElementById("banner");

// ---------- tree ----------
// Depth counts cell hops. Operator and constant nodes are part of the cell's own
// expression and never cost a hop, so depth 1 is exactly one formula.
function build(){
  var adj = dir === "down" ? succs : preds;
  var count = 0;
  truncated = false;
  function grow(i, budget, path){
    var node = {i: i, kids: [], cyc: false, x: 0, y: 0};
    if (count++ > MAX_NODES){ truncated = true; return node; }
    var inter = isInter(i);
    if (!inter && budget <= 0) return node;
    var next = inter ? budget : budget - 1;
    var list = adj[i].slice();
    list.sort(function(a, b){
      var d = E[a][3] - E[b][3];
      if (d) return d;
      var ia = dir === "down" ? E[a][1] : E[a][0];
      var ib = dir === "down" ? E[b][1] : E[b][0];
      return N[ia][0] < N[ib][0] ? -1 : 1;
    });
    for (var j = 0; j < list.length; j++){
      var ei = list[j], other = dir === "down" ? E[ei][1] : E[ei][0];
      if (!showConst && kind(other) === "const") continue;
      if (path[other]){
        node.kids.push({e: ei, child: {i: other, kids: [], cyc: true, x: 0, y: 0}});
        continue;
      }
      path[other] = 1;
      node.kids.push({e: ei, child: grow(other, next, path)});
      delete path[other];
    }
    return node;
  }
  var seed = {};
  seed[focus] = 1;
  tree = grow(focus, depth, seed);
  var cursor = 0;
  (function place(t, col){
    t.x = col * (COL_W + X_GAP);
    if (!t.kids.length){
      t.y = cursor;
      cursor += NODE_H + Y_GAP;
      return;
    }
    t.kids.forEach(function(k){ place(k.child, col + 1); });
    t.y = (t.kids[0].child.y + t.kids[t.kids.length - 1].child.y) / 2;
  })(tree, 0);
}

// ---------- drawing ----------
function el(tag, attrs){
  var node = document.createElementNS(NS, tag);
  for (var k in attrs) node.setAttribute(k, attrs[k]);
  return node;
}
function clip(text, n){
  text = text || "";
  return text.length > n ? text.slice(0, n - 1) + "\u2026" : text;
}
function widthOf(i){
  var k = kind(i);
  if (k === "op") return OP_W;
  if (k === "const") return CONST_W;
  return CELL_W;
}

function draw(){
  while (rootG.firstChild) rootG.removeChild(rootG.firstChild);
  var edgeLayer = el("g", {}), nodeLayer = el("g", {});
  rootG.appendChild(edgeLayer);
  rootG.appendChild(nodeLayer);

  (function walk(t){
    t.kids.forEach(function(k){
      var c = k.child;
      var x1 = c.x, y1 = c.y + NODE_H / 2;
      var x2 = t.x + widthOf(t.i), y2 = t.y + NODE_H / 2;
      var mx = (x1 + x2) / 2;
      var r = role(k.e), colour = ROLE_COLOR[r] || "#5a6472";
      // The arrow always follows the data. Tracing precedents that means right to
      // left, into the focused cell; tracing dependents it means the other way.
      var ax = dir === "down" ? x2 : x1, ay = dir === "down" ? y2 : y1;
      var bx = dir === "down" ? x1 : x2, by = dir === "down" ? y1 : y2;
      edgeLayer.appendChild(el("path", {
        d: "M" + ax + "," + ay + "C" + mx + "," + ay + " " + mx + "," + by + " " + bx + "," + by,
        fill: "none", stroke: colour, "stroke-width": 1.4, opacity: .75,
        "marker-end": "url(#arrow-" + (ROLE_COLOR[r] ? r : "plain") + ")"
      }));
      if (showRoles){
        // Sit the label next to the child, not at the midpoint: seventeen summands all
        // bend towards the same parent, and their midpoints pile up on each other.
        var label = el("text", {x: x1 - 7, y: y1 - 4, "text-anchor": "end",
                                class: "rlabel", fill: colour});
        var idx = E[k.e][3];
        label.textContent = idx > 0 && POSITIONAL[r] ? r + " " + idx : r;
        edgeLayer.appendChild(label);
      }
      walk(c);
    });
  })(tree);

  (function walk(t){
    nodeLayer.appendChild(shape(t));
    t.kids.forEach(function(k){ walk(k.child); });
  })(tree);

  markers();
  if (truncated) show(banner, "Tree is bigger than " + MAX_NODES +
                     " nodes and was cut short. Lower the depth.");
  else hide(banner);
}

function markers(){
  var defs = el("defs", {});
  var seen = {plain: "#5a6472"};
  for (var r in ROLE_COLOR) seen[r] = ROLE_COLOR[r];
  for (var name in seen){
    var m = el("marker", {id: "arrow-" + name, viewBox: "0 0 8 8", refX: 7, refY: 4,
                          markerWidth: 6, markerHeight: 6, orient: "auto"});
    m.appendChild(el("path", {d: "M0,1 L7,4 L0,7 z", fill: seen[name], opacity: .85}));
    defs.appendChild(m);
  }
  rootG.appendChild(defs);
}

function shape(t){
  var i = t.i, k = kind(i), rec = N[i], w = widthOf(i);
  var g = el("g", {transform: "translate(" + t.x + "," + t.y + ")", class: "hitbox"});
  var colour = KIND_COLOR[k] || "#7d8590";
  var isFocus = i === focus, isHit = !!(matches && matches[i]);
  if (isHit && k !== "op" && k !== "const"){
    g.appendChild(el("rect", {x: -5, y: -5, width: w + 10, height: NODE_H + 10, rx: 10,
                              fill: "none", stroke: HIT_COLOR, "stroke-width": 1,
                              opacity: .5}));
  }

  if (k === "op"){
    g.appendChild(el("polygon", {
      points: [w / 2 + ",0", w + "," + NODE_H / 2, w / 2 + "," + NODE_H, "0," + NODE_H / 2].join(" "),
      fill: colour, stroke: isFocus ? "#fff" : "#00000055", "stroke-width": isFocus ? 2 : 1
    }));
    var op = el("text", {x: w / 2, y: NODE_H / 2 + 4, "text-anchor": "middle", class: "oplabel"});
    op.textContent = clip(rec[7], 12);
    g.appendChild(op);
  } else if (k === "const"){
    g.appendChild(el("rect", {x: 0, y: 4, width: w, height: NODE_H - 8, rx: (NODE_H - 8) / 2,
                              fill: "#20262f", stroke: colour, "stroke-width": 1}));
    var cv = el("text", {x: w / 2, y: NODE_H / 2 + 4, "text-anchor": "middle", class: "nlabel",
                         fill: colour});
    cv.textContent = clip(rec[6] || rec[5], 14);
    g.appendChild(cv);
  } else {
    g.appendChild(el("rect", {x: 0, y: 0, width: w, height: NODE_H, rx: 6,
                              fill: t.cyc ? "#2b1d1d" : isHit ? "#2c1526" : "#1b222c",
                              stroke: isFocus ? "#fff" : isHit ? HIT_COLOR : colour,
                              "stroke-width": isFocus || isHit ? 2 : 1.2}));
    g.appendChild(el("rect", {x: 0, y: 0, width: 4, height: NODE_H, rx: 2, fill: colour}));
    var name = el("text", {x: 10, y: 15, class: "nlabel"});
    name.textContent = clip(rec[0], 26);
    g.appendChild(name);
    var sub = el("text", {x: 10, y: 28, class: "nsub"});
    sub.textContent = clip(rec[3] || rec[5] || kind(i), 30);
    g.appendChild(sub);
  }

  g.addEventListener("mouseenter", function(ev){ hover(ev, i, t.cyc); });
  g.addEventListener("mousemove", function(ev){ position(ev); });
  g.addEventListener("mouseleave", function(){ hide(tip); });
  g.addEventListener("click", function(ev){
    ev.stopPropagation();
    if (!isInter(i) && i !== focus) setFocus(i, true);
  });
  return g;
}

function esc(s){
  return String(s == null ? "" : s).replace(/&/g, "&amp;").replace(/</g, "&lt;")
                                   .replace(/>/g, "&gt;");
}
function hover(ev, i, cyc){
  var rec = N[i], k = kind(i), html = "<b>" + esc(rec[0]) + "</b> <span class=muted>" +
      esc(k) + "</span>";
  if (rec[3]) html += "<div class=muted>" + esc(clip(rec[3], 90)) + "</div>";
  if (rec[4]) html += "<code>" + esc(clip(rec[4], 240)) + "</code>";
  else if (rec[6]) html += "<code>" + esc(clip(rec[6], 240)) + "</code>";
  if (rec[5] !== "") html += "<div class=muted>value " + esc(rec[5]) + "</div>";
  if (cyc) html += "<div style='color:#ff7b72'>already on this path (circular)</div>";
  if (!isInter(i)) html += "<div class=muted>" +
    (mode === "map" ? "double-click to open its tree" : "click to focus") + "</div>";
  tip.innerHTML = html;
  show(tip);
  position(ev);
}
function position(ev){
  var box = stage.getBoundingClientRect();
  var x = ev.clientX - box.left + 14, y = ev.clientY - box.top + 14;
  if (x + tip.offsetWidth > box.width) x -= tip.offsetWidth + 26;
  if (y + tip.offsetHeight > box.height) y -= tip.offsetHeight + 26;
  tip.style.left = x + "px";
  tip.style.top = y + "px";
}
function show(node, text){ if (text != null) node.textContent = text; node.style.display = "block"; }
function hide(node){ node.style.display = "none"; }

// ---------- whole-workbook map ----------
// Every cell drawn where it actually sits in the spreadsheet, sheets side by side.
// Operator nodes fold into the cell that owns them unless asked for, so by default the
// map is a cell-level picture of the model and the trees are the drill-down.
var GX = 26, GY = 8, LANE_GAP = 150;
var LX = 170, LY = 10, LEVEL_ROWS = 400, SUBCOL_W = 13, LEVEL_GAP = 46;
var MAP = {vis: null, x: null, y: null, edges: [], lanes: [], hash: null,
           bbox: null, hidden: 0, shown: 0, comps: null, clipped: 0};

function asc(a, b){ return a - b; }
function ownerIdx(i){ return isInter(i) ? byIdLookup(owner(i)) : i; }

function visibleSet(want, sel){
  var n = N.length, vis = new Uint8Array(n), i, s;
  for (i = 0; i < n; i++){
    if (isInter(i)) continue;
    s = N[i][2];
    if (s){
      if (!want[s] || N[i][10] < 0) continue;
    } else if (sel) continue;  // defined names and links only show in the all-sheets view
    vis[i] = 1;
  }
  if (showOps){
    for (i = 0; i < n; i++){
      if (!isInter(i)) continue;
      var own = byIdLookup(owner(i));
      if (own >= 0 && vis[own]) vis[i] = 1;
    }
  }
  return vis;
}

function mapEdges(vis){
  var n = N.length, out = [], seen = new Set(), hidden = 0;
  for (var ei = 0; ei < E.length; ei++){
    var a = E[ei][0], b = E[ei][1];
    if (!showOps){ a = ownerIdx(a); b = ownerIdx(b); }
    if (a < 0 || b < 0 || a === b) continue;
    if (!vis[a] || !vis[b]){ hidden++; continue; }
    var key = a * n + b;
    if (seen.has(key)) continue;
    seen.add(key);
    out.push(a, b, E[ei][2]);
  }
  return {edges: out, hidden: hidden};
}

// Ends of the graph, judged against what is currently in view: with a sheet selected,
// a cell fed only from another sheet counts as initial, which is what makes the flow
// layout's level 0 and the initial set the same population.
function classify(vis, edges){
  var n = N.length, indeg = new Int32Array(n), outdeg = new Int32Array(n), i;
  for (var e = 0; e < edges.length; e += 3){ outdeg[edges[e]]++; indeg[edges[e + 1]]++; }
  var initial = new Uint8Array(n), terminal = new Uint8Array(n), starts = 0, ends = 0;
  for (i = 0; i < n; i++){
    if (!vis[i]) continue;
    if (!indeg[i]){ initial[i] = 1; starts++; }
    if (!outdeg[i]){ terminal[i] = 1; ends++; }
  }
  return {initial: initial, terminal: terminal, starts: starts, ends: ends};
}

// Cells where the spreadsheet puts them, sheets side by side.
function placeGrid(vis, sheets, sel){
  var n = N.length, X = new Float64Array(n), Y = new Float64Array(n), i, s;
  // collapse the used rows and columns to ordinals: a sheet using rows 1-40 and
  // 900-1000 would otherwise be mostly empty space
  var used = {};
  for (i = 0; i < n; i++){
    if (!vis[i] || isInter(i) || !N[i][2]) continue;
    var u = used[N[i][2]] || (used[N[i][2]] = {rows: {}, cols: {}});
    u.rows[N[i][10]] = 1;
    u.cols[N[i][11]] = 1;
  }
  var rank = {}, lanes = [], offset = {}, x = 0;
  sheets.forEach(function(sheet){
    if (!used[sheet]) return;
    var rr = {}, cc = {}, nr = 0, nc = 0;
    Object.keys(used[sheet].rows).map(Number).sort(asc).forEach(function(r){ rr[r] = nr++; });
    Object.keys(used[sheet].cols).map(Number).sort(asc).forEach(function(c){ cc[c] = nc++; });
    rank[sheet] = {r: rr, c: cc};
    offset[sheet] = x;
    lanes.push({name: sheet, x: x, w: Math.max(nc, 1) * GX, h: Math.max(nr, 1) * GY});
    x += Math.max(nc, 1) * GX + LANE_GAP;
  });
  for (i = 0; i < n; i++){
    if (!vis[i] || isInter(i)) continue;
    s = N[i][2];
    if (!s || !rank[s]) continue;
    X[i] = offset[s] + rank[s].c[N[i][11]] * GX;
    Y[i] = rank[s].r[N[i][10]] * GY;
  }
  var loose = 0;
  if (!sel){
    for (i = 0; i < n; i++){
      if (!vis[i] || isInter(i) || N[i][2]) continue;
      X[i] = x;
      Y[i] = (loose++) * GY * 2;
    }
    if (loose) lanes.push({name: "names & links", x: x, w: GX, h: loose * GY * 2});
  }
  if (showOps){
    var tail = {};
    for (i = 0; i < n; i++){
      if (!vis[i] || !isInter(i)) continue;
      var own = byIdLookup(owner(i));
      if (own < 0 || !vis[own]) continue;
      var k = tail[own] = (tail[own] || 0) + 1;
      X[i] = X[own] - k * 6;
      Y[i] = Y[own] - 3;
    }
  }
  return {x: X, y: Y, lanes: lanes};
}

// Cells by how deep they sit in the dependency chain: column 0 is everything nothing
// feeds, and each column to the right is one calculation step further downstream.
function placeLayers(vis, edges){
  var n = N.length, m = edges.length / 3, i, e, l;
  var X = new Float64Array(n), Y = new Float64Array(n);
  var outAt = new Int32Array(n + 1), inAt = new Int32Array(n + 1);
  for (e = 0; e < edges.length; e += 3){ outAt[edges[e] + 1]++; inAt[edges[e + 1] + 1]++; }
  for (i = 0; i < n; i++){ outAt[i + 1] += outAt[i]; inAt[i + 1] += inAt[i]; }
  var outAdj = new Int32Array(m), inAdj = new Int32Array(m);
  var oc = outAt.slice(0, n), ic = inAt.slice(0, n);
  for (e = 0; e < edges.length; e += 3){
    outAdj[oc[edges[e]]++] = edges[e + 1];
    inAdj[ic[edges[e + 1]]++] = edges[e];
  }

  // longest path from a source, so a cell always sits right of everything it reads
  var level = new Int32Array(n), pending = new Int32Array(n);
  for (i = 0; i < n; i++) pending[i] = inAt[i + 1] - inAt[i];
  var queue = [], head = 0;
  for (i = 0; i < n; i++) if (vis[i] && pending[i] === 0) queue.push(i);
  while (head < queue.length){
    var u = queue[head++];
    for (var j = outAt[u]; j < outAt[u + 1]; j++){
      var v = outAdj[j];
      if (level[v] < level[u] + 1) level[v] = level[u] + 1;
      if (--pending[v] === 0) queue.push(v);
    }
  }
  // cells inside a circular reference never reach in-degree zero, so a few bounded
  // relaxations put them past their acyclic precedents instead of in column 0
  for (var pass = 0; pass < 4; pass++){
    var moved = false;
    for (e = 0; e < edges.length; e += 3){
      var a = edges[e], b = edges[e + 1];
      if (pending[b] > 0 && level[b] < level[a] + 1){ level[b] = level[a] + 1; moved = true; }
    }
    if (!moved) break;
  }

  var deepest = 0;
  for (i = 0; i < n; i++) if (vis[i] && level[i] > deepest) deepest = level[i];
  var buckets = [];
  for (l = 0; l <= deepest; l++) buckets.push([]);
  for (i = 0; i < n; i++) if (vis[i]) buckets[level[i]].push(i);

  var sheetOrder = {};
  DATA.sheets.forEach(function(name, idx){ sheetOrder[name] = idx; });
  function where(i){
    var s = sheetOrder[N[i][2]];
    return s === undefined ? 999 : s;
  }
  buckets.forEach(function(b){
    b.sort(function(p, q){ return where(p) - where(q) || N[p][10] - N[q][10] || N[p][11] - N[q][11]; });
  });

  // A workbook's inputs all share level 0, so that one column can hold thousands of
  // cells and stretch the picture into a needle; wrap the tall ones into sub-columns.
  var tallest = 1;
  buckets.forEach(function(b){ if (b.length > tallest) tallest = b.length; });
  var rowsPer = Math.min(tallest, LEVEL_ROWS), bandX = [], bandW = [], span = 0;
  buckets.forEach(function(b, idx){
    var cols = Math.max(1, Math.ceil(b.length / rowsPer));
    bandX[idx] = span;
    bandW[idx] = Math.max(LX - 30, cols * SUBCOL_W);
    span += bandW[idx] + LEVEL_GAP;
  });
  function placeLevel(l){
    var b = buckets[l], rows = Math.min(b.length, rowsPer);
    var cols = Math.max(1, Math.ceil(b.length / rowsPer));
    var left = bandX[l] + (bandW[l] - (cols - 1) * SUBCOL_W) / 2;
    var top = (rowsPer - rows) / 2;
    for (var idx = 0; idx < b.length; idx++){
      X[b[idx]] = left + Math.floor(idx / rowsPer) * SUBCOL_W;
      Y[b[idx]] = (top + idx % rowsPer) * LY;
    }
  }
  for (l = 0; l <= deepest; l++) placeLevel(l);

  // barycentre sweeps: repeatedly order each column by the average height of the
  // neighbours already placed on one side, which unpicks most of the edge crossings
  function sweep(back){
    var at = back ? inAt : outAt, adj = back ? inAdj : outAdj;
    for (var step = 0; step <= deepest; step++){
      l = back ? step : deepest - step;
      var b = buckets[l];
      if (b.length < 2) continue;
      var key = new Float64Array(n);
      b.forEach(function(node){
        var sum = 0, count = 0;
        for (var j = at[node]; j < at[node + 1]; j++){
          var other = adj[j];
          if (back ? level[other] < l : level[other] > l){ sum += Y[other]; count++; }
        }
        key[node] = count ? sum / count : Y[node];
      });
      b.sort(function(p, q){ return key[p] - key[q] || Y[p] - Y[q]; });
      placeLevel(l);
    }
  }
  sweep(true);
  sweep(false);
  sweep(true);

  var lanes = [];
  buckets.forEach(function(b, idx){
    if (!b.length) return;
    var rows = Math.min(b.length, rowsPer);
    lanes.push({name: "L" + idx + " \u00b7 " + b.length, x: bandX[idx],
                w: bandW[idx], y: (rowsPer - rows) / 2 * LY - 6, h: rows * LY + 12});
  });
  return {x: X, y: Y, lanes: lanes, levels: deepest + 1};
}

// ---------- independent graphs ----------
// Weakly connected components of whatever is in view. Each one is a set of cells that
// shares no calculation with any other, so it can be placed and toggled on its own:
// a workbook's islands are usually pasted history, orphan header rows or label plumbing.
var CSUB = 13, CROW = 9, CCOL_GAP = 11, COMP_GAP_X = 34, COMP_GAP_Y = 30;
var COMP_ROWS = 150, COMP_WRAP = 2600, BAND_GAP = 20, LAYER_LABEL = 13;
var COMP_LIST_MAX = 200, COMP_DRAW_MAX = 4000;

function components(vis, edges){
  var n = N.length, parent = new Int32Array(n), i, e;
  for (i = 0; i < n; i++) parent[i] = i;
  function find(x){
    while (parent[x] !== x){ parent[x] = parent[parent[x]]; x = parent[x]; }
    return x;
  }
  for (e = 0; e < edges.length; e += 3){
    var ra = find(edges[e]), rb = find(edges[e + 1]);
    if (ra !== rb) parent[ra] = rb;
  }
  var groups = {}, order = [];
  for (i = 0; i < n; i++){
    if (!vis[i]) continue;
    var root = find(i), group = groups[root];
    if (!group){
      group = groups[root] = {nodes: [], cells: 0, links: 0, sheets: {}};
      order.push(group);
    }
    group.nodes.push(i);
    if (!isInter(i)) group.cells++;
    if (N[i][2]) group.sheets[N[i][2]] = 1;
  }
  for (e = 0; e < edges.length; e += 3) groups[find(edges[e])].links++;
  order.forEach(function(group){
    // lowest-numbered member: a stable name for the toggle while the filters hold
    group.key = N[group.nodes[0]][0];
    group.sheetList = Object.keys(group.sheets);
  });
  order.sort(function(a, b){
    return b.cells - a.cells || b.links - a.links || (a.key < b.key ? -1 : 1);
  });
  order.forEach(function(group, idx){ group.rank = idx; });
  return order;
}

// Each component gets its own panel holding its own layered drawing: L0 is everything
// nothing inside that component feeds, and each column to the right is one calculation
// step further downstream. Layer numbers restart per panel because a component is a whole
// dependency graph in itself, so its own sources are its layer 0.
function placeComponents(vis, edges, live){
  var n = N.length, X = new Float64Array(n), Y = new Float64Array(n), i, e, d, b;
  var m = edges.length / 3;
  var level = new Int32Array(n), pending = new Int32Array(n);
  var outAt = new Int32Array(n + 1), inAt = new Int32Array(n + 1);
  for (e = 0; e < edges.length; e += 3){
    outAt[edges[e] + 1]++;
    inAt[edges[e + 1] + 1]++;
    pending[edges[e + 1]]++;
  }
  for (i = 0; i < n; i++){ outAt[i + 1] += outAt[i]; inAt[i + 1] += inAt[i]; }
  var outAdj = new Int32Array(m), inAdj = new Int32Array(m);
  var oc = outAt.slice(0, n), ic = inAt.slice(0, n);
  for (e = 0; e < edges.length; e += 3){
    outAdj[oc[edges[e]]++] = edges[e + 1];
    inAdj[ic[edges[e + 1]]++] = edges[e];
  }

  // no edge crosses a component, so one global longest-path pass doubles as the local
  // layer number once each panel subtracts its own minimum
  var queue = [], head = 0;
  for (i = 0; i < n; i++) if (vis[i] && !pending[i]) queue.push(i);
  while (head < queue.length){
    var u = queue[head++];
    for (var j = outAt[u]; j < outAt[u + 1]; j++){
      var v = outAdj[j];
      if (level[v] < level[u] + 1) level[v] = level[u] + 1;
      if (--pending[v] === 0) queue.push(v);
    }
  }
  for (var pass = 0; pass < 4; pass++){
    var moved = false;
    for (e = 0; e < edges.length; e += 3){
      var a = edges[e], bb = edges[e + 1];
      if (pending[bb] > 0 && level[bb] < level[a] + 1){ level[bb] = level[a] + 1; moved = true; }
    }
    if (!moved) break;
  }

  var sheetOrder = {};
  DATA.sheets.forEach(function(name, idx){ sheetOrder[name] = idx; });
  function where(i){
    var s = sheetOrder[N[i][2]];
    return s === undefined ? 999 : s;
  }

  // height of a node inside its own layer, normalised to 0..1 so that layers of very
  // different sizes can still be compared when averaging neighbour positions
  var slot = new Float64Array(n), key = new Float64Array(n);
  var boxes = [];
  live.forEach(function(group){
    var lowest = Infinity;
    group.nodes.forEach(function(i){ if (level[i] < lowest) lowest = level[i]; });
    var layers = [];
    group.nodes.forEach(function(i){
      var at = level[i] - lowest;
      (layers[at] || (layers[at] = [])).push(i);
    });
    for (d = 0; d < layers.length; d++) if (!layers[d]) layers[d] = [];

    function reslot(list){
      for (var q = 0; q < list.length; q++) slot[list[q]] = (q + 0.5) / list.length;
    }
    layers.forEach(function(list){
      list.sort(function(p, q){
        return where(p) - where(q) || N[p][10] - N[q][10] || N[p][11] - N[q][11];
      });
      reslot(list);
    });
    // barycentre sweeps: repeatedly order each layer by the average height of the
    // neighbours already placed on one side, which unpicks most of the edge crossings
    function sweep(back){
      var at = back ? inAt : outAt, adj = back ? inAdj : outAdj;
      for (var step = 0; step < layers.length; step++){
        var list = layers[back ? step : layers.length - 1 - step];
        if (list.length < 2) continue;
        list.forEach(function(node){
          var sum = 0, count = 0;
          for (var q = at[node]; q < at[node + 1]; q++){
            var other = adj[q];
            if (back ? level[other] < level[node] : level[other] > level[node]){
              sum += slot[other];
              count++;
            }
          }
          key[node] = count ? sum / count : slot[node];
        });
        list.sort(function(p, q){ return key[p] - key[q] || slot[p] - slot[q]; });
        reslot(list);
      }
    }
    if (group.links){ sweep(true); sweep(false); sweep(true); }

    // A component's layer 0 holds all of its inputs and a long chain of periods can run to
    // a thousand layers, either of which would stretch the panel into a needle. Tall layers
    // split into sub-columns and the layers themselves wrap onto further rows, which the
    // layer numbers keep readable.
    var colX = [], colW = [], colRows = [], band = [], width = 0, x = 0, bands = 0;
    for (d = 0; d < layers.length; d++){
      var size = layers[d].length;
      var span = Math.max(1, Math.ceil(size / COMP_ROWS)) * CSUB;
      if (x && x + span > COMP_WRAP){ bands++; x = 0; }
      colX[d] = x;
      colW[d] = span;
      colRows[d] = Math.max(Math.min(size, COMP_ROWS), 1);
      band[d] = bands;
      x += span + CCOL_GAP;
      if (x - CCOL_GAP > width) width = x - CCOL_GAP;
    }
    var bandRows = [], bandTop = [], stack = 0;
    for (b = 0; b <= bands; b++) bandRows[b] = 1;
    for (d = 0; d < layers.length; d++)
      if (colRows[d] > bandRows[band[d]]) bandRows[band[d]] = colRows[d];
    for (b = 0; b <= bands; b++){
      bandTop[b] = stack + LAYER_LABEL;          // room above each row for its layer numbers
      stack = bandTop[b] + bandRows[b] * CROW + BAND_GAP;
    }
    boxes.push({group: group, layers: layers, colX: colX, colW: colW, colRows: colRows,
                band: band, bandRows: bandRows, bandTop: bandTop,
                w: Math.max(width, CSUB),
                h: Math.max(stack - BAND_GAP, CROW + LAYER_LABEL)});
  });

  // A workbook's giant component is hundreds of times the size of its islands, so plain
  // shelves would strand a screen of blank space beside it. This is a bottom-left skyline
  // pack instead: each panel drops into the lowest gap wide enough to hold it.
  var area = 0, widest = 1;
  boxes.forEach(function(box){
    area += (box.w + COMP_GAP_X) * (box.h + COMP_GAP_Y);
    if (box.w > widest) widest = box.w;
  });
  var target = Math.max(Math.sqrt(area * 1.9), widest);
  var skyline = [{x: 0, w: target, y: 0}];

  function lowestGap(need){
    var best = -1, bestY = Infinity, bestX = 0;
    for (var s = 0; s < skyline.length; s++){
      var span = 0, top = 0, t;
      for (t = s; t < skyline.length && span < need; t++){
        span += skyline[t].w;
        if (skyline[t].y > top) top = skyline[t].y;
      }
      if (span < need) break;              // not enough room from here to the right edge
      if (top < bestY){ bestY = top; bestX = skyline[s].x; best = s; }
    }
    if (best < 0){ bestX = 0; bestY = 0; skyline.forEach(function(seg){ bestY = Math.max(bestY, seg.y); }); }
    return {x: bestX, y: bestY};
  }
  function claim(x, y, w, h){
    var next = [], edge = x + w;
    skyline.forEach(function(seg){
      var left = Math.max(seg.x, x), right = Math.min(seg.x + seg.w, edge);
      if (right <= left){ next.push(seg); return; }
      if (seg.x < x) next.push({x: seg.x, w: x - seg.x, y: seg.y});
      if (seg.x + seg.w > edge) next.push({x: edge, w: seg.x + seg.w - edge, y: seg.y});
    });
    next.push({x: x, w: w, y: y + h});
    next.sort(function(a, b){ return a.x - b.x; });
    // merge neighbours at the same height so the profile stays short
    skyline = [];
    next.forEach(function(seg){
      var last = skyline[skyline.length - 1];
      if (last && last.y === seg.y && last.x + last.w === seg.x) last.w += seg.w;
      else skyline.push(seg);
    });
  }

  var panels = [], layerLanes = [], deepest = 0;
  boxes.forEach(function(box){
    var need = box.w + COMP_GAP_X, spot = lowestGap(need);
    box.x = spot.x; box.y = spot.y;
    claim(spot.x, spot.y, need, box.h + COMP_GAP_Y);
    if (box.layers.length > deepest) deepest = box.layers.length;
    box.layers.forEach(function(list, d){
      if (!list.length) return;
      var b = box.band[d], left = box.x + box.colX[d], top = box.y + box.bandTop[b];
      var lift = (box.bandRows[b] - box.colRows[d]) / 2 * CROW;
      list.forEach(function(i, idx){
        X[i] = left + Math.floor(idx / COMP_ROWS) * CSUB;
        Y[i] = top + lift + (idx % COMP_ROWS) * CROW;
      });
      layerLanes.push({name: "L" + d + " \u00b7 " + list.length, short: "L" + d, tier: 1,
                       x: left, y: top, w: box.colW[d], h: box.bandRows[b] * CROW,
                       pw: box.w});
    });
    panels.push({name: "#" + (box.group.rank + 1) + " \u00b7 " + box.group.cells +
                       " cells \u00b7 " + box.group.links + " edges \u00b7 " +
                       box.layers.length + (box.layers.length === 1 ? " layer" : " layers"),
                 short: "#" + (box.group.rank + 1),
                 x: box.x, y: box.y, w: box.w, h: box.h});
  });
  // drawn in reading order so the layer numbers can be thinned out where they would collide
  layerLanes.sort(function(a, b){ return a.y - b.y || a.x - b.x; });
  return {x: X, y: Y, lanes: panels.concat(layerLanes), levels: deepest};
}

function renderCompList(){
  var box = document.getElementById("compList");
  var info = document.getElementById("compInfo");
  box.innerHTML = "";
  var comps = MAP.comps;
  if (!comps){ info.textContent = ""; return; }
  if (!comps.length){ info.textContent = "nothing in view to group"; return; }

  var listed = Math.min(comps.length, COMP_LIST_MAX), shown = 0, tail = 0;
  comps.forEach(function(group){ if (isCompLive(group)) shown++; });
  info.textContent = comps.length + (comps.length === 1 ? " graph, " : " graphs, ") +
    shown + " shown" + (MAP.clipped ? " (" + MAP.clipped + " past the draw limit)" : "");

  function row(label, hint, cells, links, checked, toggle, jump){
    var line = document.createElement("div");
    line.className = "cmp";
    var tick = document.createElement("input");
    tick.type = "checkbox";
    tick.checked = checked;
    tick.onchange = function(){ toggle(this.checked); render(); };
    var name = document.createElement("b");
    name.textContent = label;
    if (hint) name.title = hint;
    if (jump) name.onclick = jump; else name.style.cursor = "default";
    var metric = document.createElement("i");
    metric.textContent = cells + " \u00b7 " + links;
    metric.title = cells + " cells, " + links + " edges";
    line.appendChild(tick);
    line.appendChild(name);
    line.appendChild(metric);
    box.appendChild(line);
  }

  for (var r = 0; r < listed; r++){
    (function(group){
      var where = group.sheetList.length === 1 ? group.sheetList[0]
                : group.sheetList.length ? group.sheetList.length + " sheets" : "names & links";
      row("#" + (group.rank + 1) + "  " + where,
          group.nodes.slice(0, 10).map(function(i){ return N[i][0]; }).join(", ") +
            (group.nodes.length > 10 ? ", \u2026" : ""),
          group.cells, group.links, !hiddenComps[group.key],
          function(on){ if (on) delete hiddenComps[group.key]; else hiddenComps[group.key] = 1; },
          function(){ goTo(group.nodes[0]); });
    })(comps[r]);
  }
  if (comps.length > listed){
    var cells = 0, links = 0;
    for (var j = listed; j < comps.length; j++){
      cells += comps[j].cells;
      links += comps[j].links;
      tail++;
    }
    row(tail + " smaller graphs", "everything below the " + listed + " largest, toggled together",
        cells, links, !compGroupHidden, function(on){ compGroupHidden = !on; }, null);
  }
}

function isCompLive(group){
  if (hiddenComps[group.key]) return false;
  if (group.rank >= COMP_LIST_MAX && compGroupHidden) return false;
  return true;
}

function layoutMap(){
  var n = N.length, i;
  var sel = document.getElementById("sheetSel").value;
  var sheets = sel ? [sel] : DATA.sheets.slice();
  var want = {};
  sheets.forEach(function(sheet){ want[sheet] = 1; });

  var vis = visibleSet(want, sel);
  var built = mapEdges(vis);
  var ends = classify(vis, built.edges);
  document.getElementById("bSrc").textContent = "Initial " + ends.starts;
  document.getElementById("bSink").textContent = "Terminal " + ends.ends;
  if (!showSources || !showSinks){
    for (i = 0; i < n; i++){
      if (!vis[i]) continue;
      if ((!showSources && ends.initial[i]) || (!showSinks && ends.terminal[i])) vis[i] = 0;
    }
    built = mapEdges(vis);
  }
  var placed, comps = null;
  MAP.clipped = 0;
  if (layout === "comp"){
    // grouped before any component is hidden, so the toggle list survives "Hide all"
    comps = components(vis, built.edges);
    var live = [];
    comps.forEach(function(group){
      if (!isCompLive(group)) return;
      if (live.length >= COMP_DRAW_MAX){ MAP.clipped++; return; }
      live.push(group);
    });
    if (live.length !== comps.length){
      vis.fill(0);
      live.forEach(function(group){
        group.nodes.forEach(function(i){ vis[i] = 1; });
      });
      built = mapEdges(vis);
    }
    placed = placeComponents(vis, built.edges, live);
  } else {
    placed = layout === "flow" ? placeLayers(vis, built.edges)
                               : placeGrid(vis, sheets, sel);
  }
  MAP.comps = comps;

  var shown = 0;
  for (i = 0; i < n; i++) if (vis[i]) shown++;
  MAP.vis = vis; MAP.x = placed.x; MAP.y = placed.y; MAP.lanes = placed.lanes;
  MAP.edges = built.edges; MAP.hidden = built.hidden; MAP.shown = shown;

  // same shape as SVGGraphicsElement.getBBox() so fit() can treat both views alike
  var minX = 0, maxX = 1, maxY = 1;
  placed.lanes.forEach(function(lane){
    minX = Math.min(minX, lane.x);
    maxX = Math.max(maxX, lane.x + lane.w);
    maxY = Math.max(maxY, (lane.y || 0) + lane.h);
  });
  MAP.bbox = {x: minX - 20, y: -30, width: maxX - minX + 40, height: maxY + 50};

  MAP.hash = {};
  for (i = 0; i < n; i++){
    if (!vis[i]) continue;
    var bucket = Math.floor(placed.x[i] / 24) + "," + Math.floor(placed.y[i] / 24);
    (MAP.hash[bucket] || (MAP.hash[bucket] = [])).push(i);
  }

  document.getElementById("mapInfo").textContent = shown
    ? shown + " nodes, " + (built.edges.length / 3) + " links" +
      (placed.levels ? comps ? ", deepest " + placed.levels +
                                 (placed.levels === 1 ? " layer" : " layers")
                             : ", " + placed.levels + " levels deep" : "") +
      (comps ? ", " + comps.length +
                 (comps.length === 1 ? " independent graph" : " independent graphs") : "") +
      (built.hidden ? comps ? " (" + built.hidden + " links inside switched-off graphs)"
                            : " (" + built.hidden + " links to hidden nodes)" : "")
    : "nothing left to draw: everything in view is switched off";
  if (layout === "comp") renderCompList();
}

function resizeCanvas(){
  var dpr = window.devicePixelRatio || 1;
  var w = stage.clientWidth, h = stage.clientHeight;
  cvs.width = Math.round(w * dpr);
  cvs.height = Math.round(h * dpr);
}

// Everything is drawn in screen space rather than through a canvas transform: a whole
// workbook fits at about k=0.1, where a cell-sized rectangle would be a third of a
// pixel and vanish. Marks keep a floor of a few pixels at any zoom.
function sx(i){ return MAP.x[i] * view.k + view.x; }
function sy(i){ return MAP.y[i] * view.k + view.y; }

function drawMap(){
  var dpr = window.devicePixelRatio || 1;
  var w = stage.clientWidth, h = stage.clientHeight;
  if (cvs.width !== Math.round(w * dpr)) resizeCanvas();
  ctx.setTransform(dpr, 0, 0, dpr, 0, 0);
  ctx.clearRect(0, 0, w, h);
  if (!MAP.vis) return;

  var k = view.k, edges = MAP.edges, vis = MAP.vis, i;

  ctx.font = "11px -apple-system,Segoe UI,Roboto,sans-serif";
  var labelEnd = -1e9, labelRow = -1e9;
  MAP.lanes.forEach(function(lane, idx){
    if (lane.tier){
      // a layer inside an independent graph: outline only, numbered where the number fits
      var lx = lane.x * k + view.x - 4, ly = lane.y * k + view.y - 3;
      var lww = lane.w * k + 8, lhh = lane.h * k + 6;
      if (lx + lww < 0 || lx > w || ly + lhh < 0 || ly > h) return;
      ctx.strokeStyle = "#242c3a";
      ctx.lineWidth = 1;
      ctx.strokeRect(lx + 0.5, ly + 0.5, lww, lhh);
      // hundreds of two-layer islands would drown the picture in L0s, so the numbers
      // only appear once a panel is wide enough on screen to be worth reading
      if (lane.pw * k < 90) return;
      var text = lww > lane.name.length * 6 ? lane.name : lane.short;
      var need = text.length * 6 + 8;
      if (Math.abs(ly - labelRow) > 4 || lx >= labelEnd){
        ctx.fillStyle = "#6e7681";
        ctx.fillText(text, lx + 1, ly - 3);
        labelEnd = lx + need;
        labelRow = ly;
      }
      return;
    }
    var x0 = lane.x * k + view.x - 8, y0 = (lane.y || 0) * k + view.y - 10;
    var lw = lane.w * k + 16, lh = lane.h * k + 18;
    if (x0 + lw < 0 || x0 > w) return;
    ctx.fillStyle = "#161d27";
    ctx.fillRect(x0, y0, lw, lh);
    ctx.strokeStyle = "#232b38";
    ctx.lineWidth = 1;
    ctx.strokeRect(x0 + 0.5, y0 + 0.5, lw, lh);
    var room = Math.floor(lw / 6);
    if (room < 4) return;
    // narrow bands sit too close for one row of labels, so drop every other one down
    var ty = Math.max(12, y0 - 5) + (lw < 90 && idx % 2 ? 13 : 0);
    ctx.fillStyle = "#8b949e";
    ctx.fillText(lane.short && room < lane.name.length ? lane.short : clip(lane.name, room),
                 x0 + 2, ty);
  });

  var byRole = {};
  for (i = 0; i < edges.length; i += 3){
    var r = ROLES[edges[i + 2]] || "arg";
    (byRole[r] || (byRole[r] = [])).push(i);
  }
  // a live search fades the model back so a couple of pink marks read at any zoom
  var fade = matches && dimOthers ? 0.16 : 1;
  ctx.lineWidth = k > 1 ? 1 : 0.7;
  ctx.globalAlpha = (k > 1 ? 0.5 : 0.32) * fade;
  for (var name in byRole){
    ctx.strokeStyle = ROLE_COLOR[name] || "#5a6472";
    ctx.beginPath();
    var list = byRole[name];
    for (var j = 0; j < list.length; j++){
      var e = list[j], a = edges[e], b = edges[e + 1];
      var ax = sx(a), ay = sy(a), bx = sx(b), by = sy(b);
      if ((ax < 0 && bx < 0) || (ax > w && bx > w)) continue;
      if ((ay < 0 && by < 0) || (ay > h && by > h)) continue;
      ctx.moveTo(ax, ay);
      ctx.lineTo(bx, by);
    }
    ctx.stroke();
  }
  ctx.globalAlpha = 1;

  var cw = Math.max(3, Math.min(GX * k - 1, 15));
  var ch = Math.max(3, Math.min(GY * k - 1, 10));
  ctx.globalAlpha = fade;
  for (i = 0; i < N.length; i++){
    if (!vis[i]) continue;
    var x = sx(i), y = sy(i);
    if (x < -20 || x > w + 20 || y < -20 || y > h + 20) continue;
    var kind_ = kind(i);
    ctx.fillStyle = KIND_COLOR[kind_] || "#7d8590";
    if (kind_ === "op" || kind_ === "const"){
      var s = Math.max(3, Math.min(ch, 7));
      ctx.fillRect(x - s / 2, y - s / 2, s, s);
    } else {
      ctx.fillRect(x - cw / 2, y - ch / 2, cw, ch);
    }
  }
  ctx.globalAlpha = 1;

  if (matches) drawHits(cw, ch, w, h);
  if (focus >= 0 && vis[focus]) ring(focus, "#ffffff", cw + 7);
  if (hoverIdx >= 0 && vis[hoverIdx] && hoverIdx !== focus) ring(hoverIdx, "#58a6ff", cw + 5);

  if (k > 1.4){
    ctx.fillStyle = "#c9d1d9";
    ctx.font = "9px ui-monospace,Menlo,monospace";
    for (i = 0; i < N.length; i++){
      if (!vis[i] || isInter(i)) continue;
      var lx = sx(i), ly = sy(i);
      if (lx < -60 || lx > w || ly < 0 || ly > h) continue;
      ctx.fillText(N[i][0].split("!").pop(), lx + cw / 2 + 3, ly + 3);
    }
  }
}

// Search hits sit on top of everything, and keep a floor of a few pixels so two cells
// out of five thousand are still findable with the whole workbook on screen.
function drawHits(cw, ch, w, h){
  var vis = MAP.vis, ring = Math.max(11, cw + 5), i, m;
  var label = matchList.length <= 40, placed = [];
  ctx.lineWidth = 1.8;
  ctx.font = "10px ui-monospace,Menlo,monospace";
  for (i = 0; i < matchList.length; i++){
    m = matchList[i];
    if (!vis[m]) continue;
    var x = sx(m), y = sy(m);
    if (x < -40 || x > w + 40 || y < -40 || y > h + 40) continue;
    ctx.fillStyle = HIT_COLOR;
    ctx.fillRect(x - cw / 2, y - ch / 2, cw, Math.max(ch, 4));
    ctx.strokeStyle = HIT_COLOR;
    ctx.strokeRect(x - ring / 2, y - ring / 2, ring, ring);
    if (!label) continue;
    // neighbouring hits would print their addresses on top of each other
    var clear = true;
    for (var p = 0; p < placed.length; p++){
      if (Math.abs(placed[p][1] - y) < 11 && Math.abs(placed[p][0] - x) < 70){ clear = false; break; }
    }
    if (!clear) continue;
    placed.push([x, y]);
    ctx.fillText(N[m][0].split("!").pop(), x + ring / 2 + 3, y + 3.5);
  }
}

function ring(i, colour, size){
  ctx.strokeStyle = colour;
  ctx.lineWidth = 1.6;
  ctx.strokeRect(sx(i) - size / 2, sy(i) - size / 2, size, size);
}

function pick(ev){
  var box = stage.getBoundingClientRect();
  var wx = (ev.clientX - box.left - view.x) / view.k;
  var wy = (ev.clientY - box.top - view.y) / view.k;
  var best = -1, bestD = (10 / view.k) * (10 / view.k);
  var cx = Math.floor(wx / 24), cy = Math.floor(wy / 24);
  for (var dx = -1; dx <= 1; dx++){
    for (var dy = -1; dy <= 1; dy++){
      var bucket = MAP.hash[(cx + dx) + "," + (cy + dy)];
      if (!bucket) continue;
      for (var j = 0; j < bucket.length; j++){
        var i = bucket[j];
        var d = (MAP.x[i] - wx) * (MAP.x[i] - wx) + (MAP.y[i] - wy) * (MAP.y[i] - wy);
        if (d < bestD){ bestD = d; best = i; }
      }
    }
  }
  return best;
}

// ---------- viewport ----------
function apply(){
  if (mode === "tree"){
    rootG.setAttribute("transform",
      "translate(" + view.x + "," + view.y + ") scale(" + view.k + ")");
  } else {
    drawMap();
  }
}
function fit(){
  var box, pad = 60;
  if (mode === "tree"){
    box = rootG.getBBox();
    pad = 80;
  } else {
    box = MAP.bbox;
  }
  if (!box || !box.width || !box.height) return;
  var w = stage.clientWidth, h = stage.clientHeight;
  var k = Math.min((w - pad) / box.width, (h - pad) / box.height, 1.4);
  view.k = Math.max(k, 0.02);
  view.x = (w - box.width * view.k) / 2 - box.x * view.k;
  view.y = (h - box.height * view.k) / 2 - box.y * view.k;
  apply();
}
stage.addEventListener("wheel", function(ev){
  ev.preventDefault();
  var box = stage.getBoundingClientRect();
  var mx = ev.clientX - box.left, my = ev.clientY - box.top;
  var k = view.k * Math.pow(1.0015, -ev.deltaY);
  k = Math.min(Math.max(k, 0.02), 8);
  view.x = mx - (mx - view.x) * (k / view.k);
  view.y = my - (my - view.y) * (k / view.k);
  view.k = k;
  apply();
}, {passive: false});
var drag = null, dragged = false;
stage.addEventListener("mousedown", function(ev){
  drag = {x: ev.clientX, y: ev.clientY, vx: view.x, vy: view.y};
  dragged = false;
  (mode === "tree" ? svg : cvs).classList.add("drag");
});
window.addEventListener("mousemove", function(ev){
  if (!drag) return;
  if (Math.abs(ev.clientX - drag.x) + Math.abs(ev.clientY - drag.y) > 3) dragged = true;
  view.x = drag.vx + (ev.clientX - drag.x);
  view.y = drag.vy + (ev.clientY - drag.y);
  apply();
});
window.addEventListener("mouseup", function(){
  drag = null;
  svg.classList.remove("drag");
  cvs.classList.remove("drag");
});

var hoverIdx = -1;
cvs.addEventListener("mousemove", function(ev){
  if (mode !== "map" || drag) return;
  var i = pick(ev);
  if (i !== hoverIdx){
    hoverIdx = i;
    drawMap();
  }
  if (i >= 0) hover(ev, i, false); else hide(tip);
});
cvs.addEventListener("mouseleave", function(){
  hide(tip);
  if (hoverIdx >= 0){ hoverIdx = -1; drawMap(); }
});
cvs.addEventListener("click", function(ev){
  if (dragged) return;
  var i = pick(ev);
  if (i < 0) return;
  select(isInter(i) ? (byIdLookup(owner(i)) >= 0 ? byIdLookup(owner(i)) : i) : i);
});
cvs.addEventListener("dblclick", function(ev){
  var i = pick(ev);
  if (i < 0) return;
  if (isInter(i) && byIdLookup(owner(i)) >= 0) i = byIdLookup(owner(i));
  setMode("tree");
  setFocus(i, true);
});

// ---------- sidebar ----------
function setFocus(i, push){
  if (push && i !== focus) history.push(focus);
  focus = i;
  render();
}
function select(i){
  // In the map a click only moves the selection; the tree is a double-click away.
  if (mode === "tree") return setFocus(i, true);
  if (i !== focus) history.push(focus);
  focus = i;
  panel();
  drawMap();
}
function centre(i){
  if (mode !== "map" || !MAP.vis || !MAP.vis[i]) return;
  view.k = Math.max(view.k, 1.1);
  view.x = stage.clientWidth / 2 - MAP.x[i] * view.k;
  view.y = stage.clientHeight / 2 - MAP.y[i] * view.k;
  drawMap();
}
function goTo(i){
  if (mode === "map"){ select(i); centre(i); }
  else setFocus(i, true);
}
function setMode(next){
  if (mode === next) return;
  mode = next;
  document.getElementById("bMap").classList.toggle("on", mode === "map");
  document.getElementById("bTree").classList.toggle("on", mode === "tree");
  document.getElementById("secMap").style.display = mode === "map" ? "" : "none";
  document.getElementById("secComp").style.display =
    mode === "map" && layout === "comp" ? "" : "none";
  document.getElementById("secTree").style.display = mode === "tree" ? "" : "none";
  svg.style.display = mode === "tree" ? "" : "none";
  cvs.style.display = mode === "map" ? "" : "none";
  document.getElementById("hint").textContent = mode === "map"
    ? "click a cell to select \u00b7 double-click to open its tree \u00b7 drag to pan \u00b7 scroll to zoom"
    : "click a cell to re-focus \u00b7 drag to pan \u00b7 scroll to zoom";
  render();
}
function render(){
  if (mode === "tree"){
    build();
    draw();
  } else {
    resizeCanvas();
    layoutMap();
  }
  fit();
  panel();
  paintResults();
}
function panel(){
  var rec = N[focus];
  document.getElementById("focusName").textContent = rec[0];
  document.getElementById("focusLabel").textContent = rec[3] || "";
  var code = document.getElementById("focusFormula");
  code.textContent = rec[4] || (rec[5] !== "" ? "value: " + rec[5] : "(no formula)");
  document.getElementById("focusTitle").textContent =
    mode === "map" ? "Selected cell" : "Focused cell";
  document.getElementById("dv").textContent = depth;
  lineage();
}
function lineage(){
  // Operator nodes belong to exactly one cell, so their owner is the cell to jump to.
  var box = document.getElementById("lineage");
  box.innerHTML = "";
  var adj = dir === "down" ? succs : preds, seen = {}, out = [];
  var stack = [focus];
  while (stack.length){
    var cur = stack.pop();
    adj[cur].forEach(function(ei){
      var other = dir === "down" ? E[ei][1] : E[ei][0];
      var cell = isInter(other) ? byIdLookup(owner(other)) : other;
      if (isInter(other) && dir === "up"){ if (!seen["i" + other]){ seen["i" + other] = 1; stack.push(other); } return; }
      if (cell < 0 || cell === focus || seen[cell]) return;
      seen[cell] = 1;
      out.push(cell);
    });
  }
  document.getElementById("lineageTitle").textContent =
    dir === "down" ? "Feeds into" : "Built from";
  if (!out.length){ box.innerHTML = "<div class=muted>none</div>"; return; }
  out.slice(0, 60).forEach(function(i){
    var row = document.createElement("div");
    row.className = "hit";
    row.innerHTML = esc(N[i][0]) + (N[i][3] ? " <em>" + esc(clip(N[i][3], 34)) + "</em>" : "");
    row.onclick = function(){ goTo(i); };
    box.appendChild(row);
  });
}

// ---------- find and highlight ----------
// A term that reads like a reference matches that cell exactly, so Valuation!C103 does
// not also drag in C1030; anything else is a substring of the address, label, formula
// or cached value. Terms are comma-separated, and a bare list of references may use
// spaces instead, which is how a cell address gets pasted out of Excel.
var searchable = [];
N.forEach(function(n, i){ if (!isInter(i)) searchable.push(i); });
var qbox = document.getElementById("q"), results = document.getElementById("results");
var FIELDS = {cell: 0, label: 3, formula: 4, value: 5};
var REF_RE = /^(?:'[^']+'!|[^\s!,]+!)?\$?[a-z]{1,3}\$?\d+$/i;

function looksLikeRef(term){ return REF_RE.test(term); }

function splitQuery(query){
  var parts = query.split(",").map(function(s){ return s.trim(); }).filter(Boolean);
  if (parts.length === 1 && /\s/.test(parts[0])){
    var words = parts[0].split(/\s+/);
    if (words.every(looksLikeRef)) return words;
  }
  return parts;
}

function sameCell(ref, i){
  var id = N[i][0].toLowerCase().replace(/\$/g, "");
  var want = ref.toLowerCase().replace(/[$']/g, "");
  return id === want || (want.indexOf("!") < 0 && id.split("!").pop() === want);
}

function hitsTerm(term, i, loose){
  var field = null, body = term, colon = term.indexOf(":");
  if (colon > 0){
    var key = term.slice(0, colon).toLowerCase();
    if (key in FIELDS){ field = FIELDS[key]; body = term.slice(colon + 1).trim(); }
  }
  if (!body) return false;
  var low = body.toLowerCase();
  if (field !== null) return String(N[i][field] || "").toLowerCase().indexOf(low) >= 0;
  if (!loose && looksLikeRef(body)) return sameCell(body, i);
  var rec = N[i];
  return rec[0].toLowerCase().indexOf(low) >= 0 ||
         String(rec[3] || "").toLowerCase().indexOf(low) >= 0 ||
         String(rec[4] || "").toLowerCase().indexOf(low) >= 0 ||
         String(rec[5] || "").toLowerCase().indexOf(low) >= 0;
}

function runSearch(){
  var query = qbox.value.trim();
  matchList = [];
  matches = null;
  if (query){
    var list = splitQuery(query), flags = new Uint8Array(N.length), j, i;
    list.forEach(function(term){
      var found = 0, k;
      for (k = 0; k < searchable.length; k++){
        if (hitsTerm(term, searchable[k], false)){ flags[searchable[k]] = 1; found++; }
      }
      // a partial address is still worth something when no cell carries it outright
      if (!found && looksLikeRef(term)){
        for (k = 0; k < searchable.length; k++){
          if (hitsTerm(term, searchable[k], true)) flags[searchable[k]] = 1;
        }
      }
    });
    for (j = 0; j < searchable.length; j++){
      i = searchable[j];
      if (flags[i]) matchList.push(i);
    }
    if (matchList.length) matches = flags;
  }
  paintResults();
  if (mode === "map") drawMap(); else draw();
}

function paintResults(){
  var info = document.getElementById("qinfo");
  results.innerHTML = "";
  if (!qbox.value.trim()){ info.textContent = ""; return; }
  if (!matchList.length){ info.textContent = "no match"; return; }
  var buried = 0;
  if (mode === "map" && MAP.vis)
    matchList.forEach(function(i){ if (!MAP.vis[i]) buried++; });
  info.textContent = matchList.length + (matchList.length === 1 ? " match" : " matches") +
    (buried ? " \u00b7 " + buried + " outside the current filters" : "");
  matchList.slice(0, 60).forEach(function(i){
    var row = document.createElement("div");
    row.className = "hit";
    row.innerHTML = esc(N[i][0]) +
      (N[i][3] ? " <em>" + esc(clip(N[i][3], 30)) + "</em>"
               : N[i][5] !== "" ? " <em>" + esc(clip(N[i][5], 24)) + "</em>" : "");
    row.onclick = function(){ goTo(i); };
    results.appendChild(row);
  });
  if (matchList.length > 60){
    var more = document.createElement("div");
    more.className = "muted";
    more.style.padding = "4px 6px";
    more.textContent = "and " + (matchList.length - 60) + " more, all highlighted";
    results.appendChild(more);
  }
}

function zoomToMatches(){
  if (!matchList.length) return;
  var w = stage.clientWidth, h = stage.clientHeight, pad = 150;
  var minX = Infinity, minY = Infinity, maxX = -Infinity, maxY = -Infinity, found = 0;
  if (mode === "map"){
    if (!MAP.vis) return;
    matchList.forEach(function(i){
      if (!MAP.vis[i]) return;
      found++;
      minX = Math.min(minX, MAP.x[i]); maxX = Math.max(maxX, MAP.x[i]);
      minY = Math.min(minY, MAP.y[i]); maxY = Math.max(maxY, MAP.y[i]);
    });
    if (!found){
      show(banner, "Every match is hidden by the sheet or node filters. Clear them to see it.");
      return;
    }
  } else {
    (function walk(t){
      if (matches[t.i]){
        found++;
        var cx = t.x + widthOf(t.i) / 2, cy = t.y + NODE_H / 2;
        minX = Math.min(minX, cx); maxX = Math.max(maxX, cx);
        minY = Math.min(minY, cy); maxY = Math.max(maxY, cy);
      }
      t.kids.forEach(function(k){ walk(k.child); });
    })(tree);
    // nothing matched inside this tree, so re-root it on the first match instead
    if (!found) return setFocus(matchList[0], true);
    minX -= COL_W / 2; maxX += COL_W / 2;
  }
  hide(banner);
  var bw = Math.max(maxX - minX, 60), bh = Math.max(maxY - minY, 60);
  view.k = Math.max(Math.min((w - pad) / bw, (h - pad) / bh, 2.4), 0.02);
  view.x = w / 2 - (minX + maxX) / 2 * view.k;
  view.y = h / 2 - (minY + maxY) / 2 * view.k;
  apply();
}

qbox.addEventListener("input", runSearch);
qbox.addEventListener("keydown", function(ev){ if (ev.key === "Enter") zoomToMatches(); });
document.getElementById("bZoom").addEventListener("click", zoomToMatches);
document.getElementById("bClear").addEventListener("click", function(){
  qbox.value = "";
  runSearch();
});
document.getElementById("cDim").addEventListener("change", function(){
  dimOthers = this.checked;
  if (mode === "map") drawMap();
});

document.getElementById("depth").value = depth;
document.getElementById("depth").addEventListener("input", function(){
  depth = +this.value;
  render();
});
document.getElementById("cConst").addEventListener("change", function(){
  showConst = this.checked;
  render();
});
document.getElementById("cRoles").addEventListener("change", function(){
  showRoles = this.checked;
  render();
});
document.getElementById("bUp").addEventListener("click", function(){
  dir = "up";
  this.classList.add("on");
  document.getElementById("bDown").classList.remove("on");
  render();
});
document.getElementById("bDown").addEventListener("click", function(){
  dir = "down";
  this.classList.add("on");
  document.getElementById("bUp").classList.remove("on");
  render();
});
document.getElementById("bBack").addEventListener("click", function(){
  if (!history.length) return;
  var previous = history.pop();
  if (mode === "map"){ focus = previous; panel(); centre(previous); }
  else setFocus(previous, false);
});
document.getElementById("bOpen").addEventListener("click", function(){
  setMode("tree");
  render();
});
document.getElementById("bMap").addEventListener("click", function(){ setMode("map"); });
document.getElementById("bTree").addEventListener("click", function(){ setMode("tree"); });
document.getElementById("cOps").addEventListener("change", function(){
  showOps = this.checked;
  render();
});
function setLayout(next){
  if (layout === next) return;
  layout = next;
  document.getElementById("bGrid").classList.toggle("on", layout === "grid");
  document.getElementById("bFlow").classList.toggle("on", layout === "flow");
  document.getElementById("bComp").classList.toggle("on", layout === "comp");
  document.getElementById("secComp").style.display = layout === "comp" ? "" : "none";
  document.getElementById("mapHint").textContent = layout === "flow"
    ? "columns are dependency depth: inputs on the left, final outputs on the right"
    : layout === "comp"
    ? "one panel per group of cells with no calculation in common, each laid out in its own layers from L0"
    : "cells sit where they do in the spreadsheet, one lane per sheet";
  render();
}
document.getElementById("bGrid").addEventListener("click", function(){ setLayout("grid"); });
document.getElementById("bFlow").addEventListener("click", function(){ setLayout("flow"); });
document.getElementById("bComp").addEventListener("click", function(){ setLayout("comp"); });
document.getElementById("bCompAll").addEventListener("click", function(){
  hiddenComps = {};
  compGroupHidden = false;
  render();
});
document.getElementById("bCompNone").addEventListener("click", function(){
  if (!MAP.comps) return;
  MAP.comps.forEach(function(group){ hiddenComps[group.key] = 1; });
  compGroupHidden = true;
  render();
});
document.getElementById("bCompGiant").addEventListener("click", function(){
  if (!MAP.comps) return;
  MAP.comps.forEach(function(group){
    if (group.rank) hiddenComps[group.key] = 1; else delete hiddenComps[group.key];
  });
  compGroupHidden = true;
  render();
});
document.getElementById("bSrc").addEventListener("click", function(){
  showSources = !showSources;
  this.classList.toggle("on", showSources);
  render();
});
document.getElementById("bSink").addEventListener("click", function(){
  showSinks = !showSinks;
  this.classList.toggle("on", showSinks);
  render();
});
document.getElementById("bFit").addEventListener("click", fit);
window.addEventListener("resize", function(){
  if (mode === "map") resizeCanvas();
  fit();
});

(function sheetPicker(){
  var sel = document.getElementById("sheetSel");
  var all = document.createElement("option");
  all.value = "";
  all.textContent = "All sheets (" + DATA.sheets.length + ")";
  sel.appendChild(all);
  DATA.sheets.forEach(function(name){
    var option = document.createElement("option");
    option.value = name;
    option.textContent = name;
    sel.appendChild(option);
  });
  sel.addEventListener("change", render);
})();

(function legend(){
  var box = document.getElementById("legend");
  KINDS.forEach(function(k){
    var row = document.createElement("div");
    row.style.fontSize = "12px";
    row.innerHTML = "<i class=swatch style='background:" + KIND_COLOR[k] + "'></i>" + k;
    box.appendChild(row);
  });
  var hit = document.createElement("div");
  hit.style.fontSize = "12px";
  hit.innerHTML = "<i class=swatch style='background:" + HIT_COLOR + "'></i>search match";
  box.appendChild(hit);
})();

(function stats(){
  var s = DATA.stats, box = document.getElementById("stats");
  var rows = [["formulas", s.formulas_parsed], ["nodes", s.nodes], ["edges", s.edges],
              ["cross-sheet edges", s.cross_sheet_edges], ["sheets", s.sheets.length]];
  if (s.formulas_failed) rows.push(["unparsed formulas", s.formulas_failed]);
  rows.push(["circular references", s.cycles]);
  if (s.cycles) rows.push(["nodes in a cycle", s.cyclic_nodes]);
  rows.forEach(function(r){
    var row = document.createElement("div");
    row.className = "kv";
    row.innerHTML = "<span>" + r[0] + "</span><span>" + r[1] + "</span>";
    box.appendChild(row);
  });
  document.getElementById("sub").textContent =
    s.nodes + " nodes \u00b7 " + s.edges + " edges \u00b7 " + s.formulas_parsed + " formulas";
})();

document.getElementById("secTree").style.display = "none";
document.getElementById("secComp").style.display = "none";
document.getElementById("hint").textContent =
  "click a cell to select \u00b7 double-click to open its tree \u00b7 drag to pan \u00b7 scroll to zoom";
render();
if (DATA.filtered){
  show(banner, "This file was exported with --focus, so it holds only " + N.length +
       " nodes. Re-run without --focus for the whole workbook.");
}
})();
</script>
</body>
</html>
"""


# ---------------------------------------------------------------------------
# driver
# ---------------------------------------------------------------------------

def human_size(num_bytes):
    n = float(num_bytes)
    for unit in ("B", "KB", "MB", "GB"):
        if n < 1024 or unit == "GB":
            return "%d%s" % (n, unit) if unit == "B" else "%.1f%s" % (n, unit)
        n /= 1024.0


def process(path, args):
    print("%s" % path.name, flush=True)
    if path.suffix.lower() == ".xls":
        print("    skipped: legacy .xls holds no readable formulas for openpyxl.\n"
              "             convert first, e.g.  soffice --headless --convert-to xlsx %s"
              % path.name)
        return None

    graph = AstGraph(
        path,
        max_range_expand=args.max_range_expand,
        sheets=args.sheet,
        read_values=not args.no_values,
        consts=not args.no_consts,
        verbose=not args.quiet,
    )
    try:
        graph.build()
    except Exception as exc:
        print("    failed: %s: %s" % (type(exc).__name__, exc))
        return None

    nodes, edges = graph.nodes, graph.edges
    focus = ""
    filtered = False
    if args.focus:
        try:
            focus, nodes, edges = graph.focus(args.focus, args.depth, args.direction)
        except KeyError:
            print("    focus node %r not found in this workbook; keeping the whole graph"
                  % args.focus)
            focus = ""
        else:
            filtered = True
            print("    focus %s depth %d (%s) -> %d nodes, %d edges"
                  % (focus, args.depth, args.direction, len(nodes), len(edges)))
    if not focus:
        focus = pick_default_focus(nodes, edges)

    stats = graph.stats(nodes, edges)
    stats["elapsed_sec"] = round(graph.elapsed, 2)
    stats["focus"] = focus

    out_dir = Path(args.out) / path.stem
    out_dir.mkdir(parents=True, exist_ok=True)

    if not args.no_csv:
        write_csv(out_dir, nodes, edges)
    if not args.no_json:
        write_json(out_dir, nodes, edges, stats)
    if not args.no_graphml:
        write_graphml(out_dir, nodes, edges)
    if not args.no_html:
        if len(nodes) > args.html_max_nodes:
            print("    skipping graph.html: %d nodes is past --html-max-nodes (%d). "
                  "Narrow it with --sheet or --focus."
                  % (len(nodes), args.html_max_nodes))
        else:
            write_html(out_dir, graph, nodes, edges, stats, focus, args.depth,
                       args.html_text_limit, filtered)
            if filtered:
                print("    note: graph.html holds only the focused subgraph; run without "
                      "--focus for the whole-workbook map")

    print("    %d nodes, %d edges from %d formulas in %.1fs"
          % (stats["nodes"], stats["edges"], stats["formulas_parsed"], graph.elapsed))
    kinds = stats["nodes_by_kind"]
    print("    %s" % ", ".join("%s %d" % (k, kinds[k]) for k in NODE_KINDS if k in kinds))
    if stats["cycles"]:
        print("    %d circular reference%s over %d node(s); largest loops through %d"
              % (stats["cycles"], "" if stats["cycles"] == 1 else "s",
                 stats["cyclic_nodes"], stats["largest_cycle"]))
        present = [g for g in graph.cycles if any(k in nodes for k in g)]
        for group in sorted(present, key=len, reverse=True)[:3]:
            members = sorted(group)
            shown = ", ".join(members[:4])
            print("      %d nodes: %s%s"
                  % (len(group), shown, ", ..." if len(members) > 4 else ""))
    else:
        print("    no circular references")
    for warning in stats["warnings"]:
        print("    note: %s" % warning)
    written = sum(p.stat().st_size for p in out_dir.iterdir() if p.is_file())
    stats["output_bytes"] = written
    if written > 100 * 1024 * 1024:
        print("    warning: wrote %s. Lower --max-range-expand, add --focus, or drop "
              "formats with --no-graphml/--no-json to shrink this." % human_size(written))
    print("    -> %s (%s)" % (out_dir, human_size(written)))
    stats["output"] = str(out_dir)
    return stats


def main(argv=None):
    parser = argparse.ArgumentParser(
        description="Build an expression DAG (cells + operator nodes) from an Excel workbook.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""examples:
  python3 xl_ast_graph.py "4-10 100/0248.xlsx"
  python3 xl_ast_graph.py "4-10 100/0248.xlsx" --focus "Calculations!T232" --depth 2
  python3 xl_ast_graph.py "4-10 100" --glob '*.xlsx' --no-graphml
""")
    parser.add_argument("path", help="workbook, or a directory to process in batch")
    parser.add_argument("--glob", default="*.xls[xm]",
                        help="pattern used when path is a directory (default: %(default)s)")
    parser.add_argument("-o", "--out", default="ast_out", help="output directory")
    parser.add_argument("--sheet", action="append",
                        help="only parse formulas on this sheet (repeatable); precedents "
                             "on other sheets are still included")
    parser.add_argument("--focus", help="keep only this cell's expression tree, "
                                        "e.g. 'Calculations!T232'")
    parser.add_argument("--depth", type=int, default=2,
                        help="cell hops to expand around --focus; operator nodes are free, "
                             "so depth 1 is one formula (default: %(default)s)")
    parser.add_argument("--direction", choices=("up", "down", "both"), default="up",
                        help="follow precedents, dependents, or both (default: %(default)s)")
    parser.add_argument("--max-range-expand", type=int, default=100,
                        help="ranges resolving to more populated cells than this collapse "
                             "into a single range node (default: %(default)s)")
    parser.add_argument("--no-consts", action="store_true",
                        help="drop literal numbers and strings instead of giving them nodes")
    parser.add_argument("--no-values", action="store_true",
                        help="skip the cached-value pass (faster, but no labels)")
    parser.add_argument("--no-html", action="store_true")
    parser.add_argument("--no-json", action="store_true")
    parser.add_argument("--no-graphml", action="store_true")
    parser.add_argument("--no-csv", action="store_true")
    parser.add_argument("--html-max-nodes", type=int, default=150000,
                        help="skip the viewer above this many nodes; the payload is "
                             "inlined, so a bigger graph makes a file no browser enjoys "
                             "(default: %(default)s)")
    parser.add_argument("--html-text-limit", type=int, default=300,
                        help="truncate formula and expression text in the viewer at this "
                             "many characters (default: %(default)s)")
    parser.add_argument("-q", "--quiet", action="store_true")
    args = parser.parse_args(argv)

    root = Path(args.path)
    if not root.exists():
        sys.exit("no such path: %s" % root)

    if root.is_dir():
        targets = sorted(p for p in root.glob(args.glob) if not p.name.startswith("~$"))
        if not targets:
            sys.exit("no files matching %r in %s" % (args.glob, root))
    else:
        targets = [root]

    print("%d workbook(s) -> %s\n" % (len(targets), Path(args.out).resolve()))
    results = []
    for path in targets:
        stats = process(path, args)
        if stats:
            results.append(stats)
        print()

    if len(targets) > 1 and results:
        out = Path(args.out)
        out.mkdir(parents=True, exist_ok=True)
        fields = ["workbook", "nodes", "edges", "formulas_parsed", "formulas_failed",
                  "cross_sheet_edges", "cycles", "cyclic_nodes", "largest_cycle",
                  "elapsed_sec", "output_bytes", "output"]
        with open(out / "summary.csv", "w", newline="", encoding="utf-8") as fh:
            writer = csv.DictWriter(fh, fieldnames=fields, extrasaction="ignore")
            writer.writeheader()
            for stats in results:
                writer.writerow(stats)
        print("summary -> %s" % (out / "summary.csv"))

    if not results:
        sys.exit(1)


if __name__ == "__main__":
    main()
