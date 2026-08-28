#!/usr/bin/env python3
"""Write an inputs-only copy of a segmented workbook.

``xl_segment.py`` decides which cells a model has to be given and which ones it
is supposed to work out. This takes that verdict and produces the workbook you
would hand to someone asked to rebuild the model from scratch: every input value
is present, and every derived cell is empty.

A surviving formula is replaced by its cached result, so the output holds no
formulas at all. Keeping one would be worse than useless -- its precedents have
just been blanked, so it would quietly recalculate to zero. Blanked cells keep
their style, and the rewrite runs over the sheet XML rather than through
openpyxl, so column widths, number formats, conditional formatting and charts
come through untouched.

The golden rule: **a hand-typed cell is never blanked.** A typed value can never
be recomputed by whoever rebuilds the model, so removing one is only safe if the
dependency graph proves nothing needs it -- and the graph can have holes
(dynamic references like INDIRECT, values pasted between sheets). Only formula
cells are blanked; every static cell survives. The one exception is a typed cell
whose value exactly duplicates a chosen output's value: that is a pasted answer,
and it is redacted and reported.

``--keep`` still controls how much *derived* presentation content survives:

  inputs   Frozen formula values only for the input frontier.
  labels   Same; text cells are typed and always survive anyway.
  headers  (default) Also freezes computed period headers -- axis formula cells
           that read as a year or carry a date format, like ``=K1+1`` dragged
           across a year row. Nothing computes off those values, but the sheet
           is unreadable without them.

Every run is checked against the original before it is accepted: no formula may
remain, no formula-derived number outside the input frontier may survive, and
every typed cell and every input must still hold the value it started with. The
frontier-sufficiency proof from stage 9 is reported when it exists; when it was
skipped, keeping every typed cell is what makes the mask safe regardless.
"""

from __future__ import annotations

import argparse
import csv
import datetime
import json
import os
import re
import sys
import tempfile
import warnings
import zipfile
from collections import Counter, defaultdict
from pathlib import Path

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.cell import coordinate_to_tuple
    from openpyxl.worksheet.formula import ArrayFormula
except ImportError:  # pragma: no cover
    sys.exit("openpyxl is required:  python3 -m pip install openpyxl")

# Zip parts that can carry <f> formula elements.
FORMULA_PART_PREFIXES = ("xl/worksheets/", "xl/macrosheets/", "xl/dialogsheets/")

try:
    from xl_level_split import (
        CALC_CHAIN,
        CALC_OVERRIDE_RE,
        CALC_REL_RE,
        CELL_RE,
        FORMULA_RE,
        SHEET_TAG_RE,
        VALUE_RE,
        attr,
        build_cell,
        is_chart_part,
        scrub_chart_caches,
        sheet_parts,
    )
except ImportError:  # pragma: no cover
    sys.exit("xl_input_mask.py expects xl_level_split.py to sit next to it")

try:
    from xl_ast_graph import human_size
except ImportError:  # pragma: no cover
    def human_size(n):
        return "%.1f KB" % (n / 1024.0)

from xl_seg.proof import load_contract
from xl_seg.publication import (
    GenerationValidationError,
    resolve_for_consumer,
    write_inputs_sidecar,
)


# In sheet XML a static cell is a number unless it says otherwise; these are the
# three ways it can say it holds text.
TEXT_TYPES = {b"s", b"str", b"inlineStr"}

# Number-format fragments that mean a numeric cell is displayed as a date, so a
# bare serial like 45658 is really a period header.
DATE_TOKENS = ("yy", "mmm", "dd")

KEEP_MODES = ("inputs", "labels", "headers")


# ---------------------------------------------------------------------------
# which cells are allowed to live
# ---------------------------------------------------------------------------

def period_headers(book, sheet, cells):
    """Axis cells that read as a period label rather than a computed number.

    ``axis`` is assigned from a cell's shape, not its role, so the class holds
    real year headers alongside numbers that merely look date-ish. Only the ones
    that still look like a period after a second check are let through.
    """
    ws = book[sheet]
    out = set()
    for row, col in cells:
        cell = ws.cell(row=row, column=col)
        value = cell.value
        if isinstance(value, (datetime.datetime, datetime.date)):
            out.add((row, col))
            continue
        fmt = str(cell.number_format).lower()
        if any(token in fmt for token in DATE_TOKENS):
            out.add((row, col))
            continue
        if (isinstance(value, (int, float)) and not isinstance(value, bool)
                and float(value).is_integer() and 1990 <= value <= 2100):
            out.add((row, col))
    return out


def keep_cells(bands_path, mode, book=None):
    """``({sheet: {(row, col)}}, {(sheet, row, col): bucket})``.

    The second value is the period headers that were let in despite not being
    inputs, tagged with the bucket they came from, so a run can report what it
    admitted rather than leaving it to be discovered.
    """
    keep = defaultdict(set)
    axis = defaultdict(set)
    origin = {}
    with open(bands_path, newline="", encoding="utf-8") as fh:
        bands = list(csv.DictReader(fh))
    outputs = {
        (band["sheet"], int(band["row"]), col)
        for band in bands if band["bucket"] == "output"
        for col in range(int(band["col_lo"]), int(band["col_hi"]) + 1)
    }
    for band in bands:
        sheet = band["sheet"]
        row = int(band["row"])
        span = [(row, c) for c in
                range(int(band["col_lo"]), int(band["col_hi"]) + 1)]
        # A synthetic literal-source band marks a formula cell whose hardcoded
        # constant is an input. Keep and freeze its host unless that host is
        # itself a curated output; output cells must remain blank, and their
        # constants are still carried by the Embedded Assumptions sheet.
        if band["kind"] == "literal":
            if band["bucket"] == "input":
                keep[sheet].update(
                    cell for cell in span
                    if (sheet,) + cell not in outputs
                )
            continue
        if band["bucket"] == "input":
            keep[sheet].update(span)
        elif (mode != "inputs" and band["kind"] != "formula"
              and band["vtype"] in ("text", "unit")):
            # Typed text/unit cells survive on their own (mask_sheet never
            # blanks typed cells); admitting a *formula* band here would
            # freeze its cached display -- e.g. IFERROR fallback strings like
            # "-" -- which is derived content and pre-discloses the branch a
            # blanked calculation takes (0261 Soft Inc survivors).
            keep[sheet].update(span)
        elif mode == "headers" and band["vtype"] == "axis":
            axis[sheet].update(span)
            for cell in span:
                origin[(sheet,) + cell] = band["bucket"]

    admitted = {}
    for sheet, cells in axis.items():
        allowed = period_headers(book, sheet, cells)
        keep[sheet].update(allowed)
        for cell in allowed:
            admitted[(sheet,) + cell] = origin.get((sheet,) + cell, "?")
    return keep, admitted


def input_cells(bands_path):
    """``{sheet: {(row, col)}}`` for the input frontier alone."""
    frontier = defaultdict(set)
    with open(bands_path, newline="", encoding="utf-8") as fh:
        bands = list(csv.DictReader(fh))
    outputs = {
        (band["sheet"], int(band["row"]), col)
        for band in bands if band["bucket"] == "output"
        for col in range(int(band["col_lo"]), int(band["col_hi"]) + 1)
    }
    for band in bands:
        if band["bucket"] != "input":
            continue
        row = int(band["row"])
        frontier[band["sheet"]].update(
            (row, col)
            for col in range(int(band["col_lo"]), int(band["col_hi"]) + 1)
            if (band["sheet"], row, col) not in outputs
        )
    return frontier


def stabilized_proof_inputs(seg_dir):
    """Effective strict-proof cells, or ``None`` for legacy segmentation."""
    proof = load_contract(seg_dir)
    if proof is None:
        return None
    refs = proof.get("effective_inputs")
    if not isinstance(refs, list):
        return None
    cells = defaultdict(set)
    for ref in refs:
        sheet, sep, coordinate = str(ref).rpartition("!")
        if not sep:
            continue
        try:
            row, col = coordinate_to_tuple(coordinate.replace("$", ""))
        except ValueError:
            continue
        cells[sheet.strip("'")].add((row, col))
    return cells


def output_values(bands_path, book):
    """Original values of the chosen output cells, for the pasted-answer audit."""
    values = []
    with open(bands_path, newline="", encoding="utf-8") as fh:
        for band in csv.DictReader(fh):
            if band["bucket"] != "output" or band["sheet"] not in book.sheetnames:
                continue
            ws = book[band["sheet"]]
            row = int(band["row"])
            for col in range(int(band["col_lo"]), int(band["col_hi"]) + 1):
                value = ws.cell(row=row, column=col).value
                if isinstance(value, (int, float)) and not isinstance(value, bool):
                    values.append(float(value))
    return values


def output_cells(bands_path):
    """``{(sheet, row, col)}`` for every curated output cell."""
    cells = set()
    with open(bands_path, newline="", encoding="utf-8") as fh:
        for band in csv.DictReader(fh):
            if band["bucket"] != "output":
                continue
            row = int(band["row"])
            for col in range(int(band["col_lo"]), int(band["col_hi"]) + 1):
                cells.add((band["sheet"], row, col))
    return cells


def formula_coordinates(src_path):
    """``{sheet: {(row, col)}}`` of every cell that carries a formula."""
    book = openpyxl.load_workbook(src_path, data_only=False)
    coords = defaultdict(set)
    for ws in book.worksheets:
        cells = getattr(ws, "_cells", None)
        iterable = cells.values() if cells is not None else (
            c for row in ws.iter_rows() for c in row)
        for cell in iterable:
            if cell.data_type == "f" and cell.value is not None:
                coords[ws.title].add((cell.row, cell.column))
    book.close()
    return coords


TEXT_NUMBER_RE = re.compile(r"\d+(?:\.\d+)?(?:[eE][+-]?\d+)?")


def _sig4(value):
    """The value rounded to 4 significant digits, for display-paste matching."""
    return float("%.4g" % value)


def pasted_answers(book, formula_coords, frontier, outputs):
    """Typed cells whose value duplicates a chosen output's value.

    Keeping every typed cell is what makes the mask safe, but a value-paste of
    an answer would hand the rebuild its target. Three classes are caught:

    - a full-precision float match on a typed cell outside the input frontier
      is no coincidence: it is blanked and reported, as before;
    - a typed cell *inside* the frontier that equals an output, and a typed
      cell agreeing with an output only to 4 significant digits (a paste
      rounded for display), are reported as suspects for a human to arbitrate
      -- blanking a frontier cell would break the proven-sufficient input set,
      and a rounded collision can be legitimate;
    - a text cell whose formatted rendering contains an output value (either
      at face value or as a percentage, "23.4%" for 0.234) is likewise
      reported as a suspect.

    Trivial round integers (whole numbers under 10,000) are never matched:
    they collide by accident, not by paste, and would drown the report.
    """
    deny = defaultdict(set)
    report = []
    suspects = []
    interesting = [o for o in outputs
                   if not (float(o).is_integer() and abs(o) < 10000)]
    if not interesting:
        return deny, report, suspects

    def exact(v):
        return next((t for t in interesting
                     if abs(v - t) <= 1e-12 * max(1.0, abs(t))), None)

    def rounded(v):
        return next((t for t in interesting if _sig4(v) == _sig4(t)), None)

    for ws in book.worksheets:
        formulas = formula_coords.get(ws.title, set())
        safe = frontier.get(ws.title, set())
        cells = getattr(ws, "_cells", None)
        iterable = cells.values() if cells is not None else (
            c for row in ws.iter_rows() for c in row)
        for cell in iterable:
            value = cell.value
            spot = (cell.row, cell.column)
            if spot in formulas:
                continue
            where = "%s!%s%d" % (ws.title, get_column_letter(cell.column),
                                 cell.row)

            if isinstance(value, str):
                for token in TEXT_NUMBER_RE.findall(value.replace(",", "")):
                    try:
                        x = float(token)
                    except ValueError:  # pragma: no cover
                        continue
                    if x.is_integer() and abs(x) < 10000:
                        continue
                    hit = rounded(x) or rounded(x / 100.0)
                    if hit is not None:
                        suspects.append(
                            "%s=%r renders output value %r" % (where, value, hit))
                        break
                continue

            if not isinstance(value, (int, float)) or isinstance(value, bool):
                continue
            v = float(value)
            if float(v).is_integer() and abs(v) < 10000:
                continue
            if spot in safe:
                hit = exact(v)
                if hit is not None:
                    suspects.append(
                        "%s=%r (input-frontier cell equals an output; "
                        "review curation)" % (where, value))
                continue
            if exact(v) is not None:
                deny[ws.title].add(spot)
                report.append("%s=%r" % (where, value))
                continue
            hit = rounded(v)
            if hit is not None:
                suspects.append(
                    "%s=%r matches output %r at 4 significant digits"
                    % (where, value, hit))
    return deny, report, suspects


def load_mask_cells(path):
    """``{sheet: {(row, col)}}`` from a JSON list of refs, ranges allowed.

    These are externally-sourced variables served through a mock MCP research
    service; blanking them is the one sanctioned exception to the golden rule
    that a typed cell is never removed, because the value stays retrievable --
    just through the research tools instead of the sheet.
    """
    from openpyxl.utils import range_boundaries
    refs = json.loads(Path(path).read_text(encoding="utf-8"))
    cells = defaultdict(set)
    for ref in refs:
        sheet, _, coords = ref.rpartition("!")
        sheet = sheet.strip("'")
        min_col, min_row, max_col, max_row = range_boundaries(coords)
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cells[sheet].add((row, col))
    return cells


def frontier_proof(seg_dir):
    """What stage 9 established for this workbook: PASS, FAIL or SKIPPED."""
    path = Path(seg_dir) / "segments.json"
    if not path.exists():
        return "MISSING"
    verdict = json.loads(path.read_text(encoding="utf-8")).get("verification", {})
    if verdict.get("skipped"):
        return "SKIPPED"
    return "PASS" if (
        verdict.get("status") == "pass"
        and verdict.get("disposition") == "pass"
        and verdict.get("blocking_reasons") == []
    ) else "FAIL"


# ---------------------------------------------------------------------------
# embedded assumptions
# ---------------------------------------------------------------------------

ASSUMPTIONS_SHEET = "Embedded Assumptions"
ASSUMPTIONS_PART = "xl/worksheets/sheetEmbeddedAssumptions.xml"
WORKSHEET_CT = (b'<Override PartName="/' + ASSUMPTIONS_PART.encode("ascii")
                + b'" ContentType="application/vnd.openxmlformats-officedocument.'
                  b'spreadsheetml.worksheet+xml"/>')
WORKSHEET_REL_TYPE = (b"http://schemas.openxmlformats.org/officeDocument/2006/"
                      b"relationships/worksheet")
SHEET_ID_RE = re.compile(rb'sheetId="(\d+)"')
REL_ID_RE = re.compile(rb'Id="([^"]+)"')


def embedded_assumptions(seg_dir):
    """Promoted literals the segmentation classified as inputs.

    These constants live inside derived formulas, so they own no cell the mask
    could keep -- without this list the inputs-only workbook silently loses
    them and the model cannot be rebuilt.
    """
    path = Path(seg_dir) / "segments.json"
    if not path.exists():
        return []
    data = json.loads(path.read_text(encoding="utf-8"))
    rows = [e for e in data.get("embedded_literals", [])
            if e.get("bucket") == "input" and "host" in e]
    rows.sort(key=lambda e: (e.get("sheet", ""), e.get("host", "")))
    return rows


MASKED_VALUE_NOTE = "(masked: retrieve via the research data service)"

FORMULA_NUMBER_RE = re.compile(r"\d+(?:\.\d+)?(?:[eE][+-]?\d+)?")


def _value_matches(value, deny_numbers, deny_texts):
    if isinstance(value, bool) or value is None:
        return False
    if isinstance(value, (int, float)):
        v = float(value)
        return any(abs(abs(v) - abs(t)) <= 1e-9 * max(1.0, abs(t))
                   for t in deny_numbers)
    text = str(value).strip()
    return bool(text) and text in deny_texts


def _formula_mentions(formula, deny_numbers):
    for token in FORMULA_NUMBER_RE.findall(str(formula)):
        try:
            v = float(token)
        except ValueError:  # pragma: no cover
            continue
        if any(abs(v - abs(t)) <= 1e-9 * max(1.0, abs(t)) for t in deny_numbers):
            return True
    return False


def _entry_host(entry):
    """``(sheet, row, col)`` of an assumption entry's host cell, or None."""
    host = entry.get("host", "")
    sheet, _, coords = host.rpartition("!")
    if not sheet or not coords:
        return None
    try:
        row, col = coordinate_to_tuple(coords)
    except ValueError:
        return None
    return (sheet, row, col)


def redact_assumptions(rows, deny, deny_numbers, deny_texts, outputs):
    """Withhold masked and answer-revealing content from the assumption rows.

    The sheet documents constants hardcoded inside formulas, which creates two
    side channels the grid mask cannot see. Mirroring the mask itself:

    - an entry whose host cell is denied, or whose constant equals a denied
      cell's value, loses both value and formula -- the value stays
      retrievable, through the research tools instead of the sheet;
    - an entry whose formula merely contains a denied value as a literal
      loses the formula text;
    - an entry hosted in a curated output cell loses the formula text, which
      is the answer cell's exact derivation. Its constant is still an input
      and stays.
    """
    redacted, stripped = [], []
    out = []
    for entry in rows:
        entry = dict(entry)
        where = _entry_host(entry)
        host_denied = (where is not None
                       and where[1:] in deny.get(where[0], set()))
        formula = str(entry.get("formula", ""))
        if host_denied or _value_matches(entry.get("value"), deny_numbers, deny_texts):
            entry["value"] = MASKED_VALUE_NOTE
            entry["formula"] = ""
            redacted.append(entry.get("host", ""))
        elif formula and (
            (where is not None and where in outputs)
            or _formula_mentions(formula, deny_numbers)
        ):
            entry["formula"] = ""
            stripped.append(entry.get("host", ""))
        out.append(entry)
    return out, redacted, stripped


def _xml_escape(text):
    return (str(text).replace("&", "&amp;").replace("<", "&lt;")
            .replace(">", "&gt;").replace('"', "&quot;"))


def _assumptions_sheet_xml(rows):
    def cell(ref, value):
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            return '<c r="%s"><v>%.17g</v></c>' % (ref, value)
        return ('<c r="%s" t="inlineStr"><is><t xml:space="preserve">%s</t></is></c>'
                % (ref, _xml_escape(value)))

    table = [["Sheet", "Cells", "Line item", "Hardcoded value", "Formula"]]
    for entry in rows:
        host = entry.get("host", "")
        coords = host.split("!", 1)[1] if "!" in host else host
        formula = entry.get("formula", "")
        if str(formula).startswith("="):
            formula = "'" + str(formula)
        table.append([
            entry.get("sheet", ""), coords, entry.get("label", ""),
            entry.get("value", ""), formula,
        ])

    body = []
    for r, values in enumerate(table, start=1):
        cells = "".join(
            cell("%s%d" % (get_column_letter(c), r), v)
            for c, v in enumerate(values, start=1)
            if v != ""
        )
        body.append('<row r="%d">%s</row>' % (r, cells))
    xml = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
           '<worksheet xmlns="http://schemas.openxmlformats.org/'
           'spreadsheetml/2006/main"><sheetData>%s</sheetData></worksheet>'
           % "".join(body))
    return xml.encode("utf-8")


def inject_assumptions(out_path, rows, existing_sheets):
    """Append an assumptions sheet at the XML level, sparing charts and styles.

    ``rows`` must already have been through :func:`redact_assumptions`."""
    name = ASSUMPTIONS_SHEET
    while name in existing_sheets:
        name += " (pipeline)"

    with zipfile.ZipFile(out_path) as src:
        items = [(item, src.read(item.filename)) for item in src.infolist()]

    parts = {item.filename: data for item, data in items}
    workbook = parts["xl/workbook.xml"]
    rels = parts["xl/_rels/workbook.xml.rels"]
    types = parts["[Content_Types].xml"]

    next_sheet_id = max((int(m.group(1)) for m in SHEET_ID_RE.finditer(workbook)),
                        default=0) + 1
    taken = {m.group(1) for m in REL_ID_RE.finditer(rels)}
    n = 1
    while ("rIdEmbA%d" % n).encode("ascii") in taken:
        n += 1
    rel_id = ("rIdEmbA%d" % n).encode("ascii")

    first_sheet = SHEET_TAG_RE.search(workbook)
    prefix = b"r"
    if first_sheet:
        found = re.search(rb'\s([A-Za-z0-9]+):id="', first_sheet.group(0))
        if found:
            prefix = found.group(1)

    sheet_tag = (b'<sheet name="' + _xml_escape(name).encode("utf-8")
                 + b'" sheetId="' + str(next_sheet_id).encode("ascii")
                 + b'" ' + prefix + b':id="' + rel_id + b'"/>')
    rel_tag = (b'<Relationship Id="' + rel_id + b'" Type="' + WORKSHEET_REL_TYPE
               + b'" Target="worksheets/sheetEmbeddedAssumptions.xml"/>')

    parts["xl/workbook.xml"] = workbook.replace(b"</sheets>", sheet_tag + b"</sheets>", 1)
    parts["xl/_rels/workbook.xml.rels"] = rels.replace(
        b"</Relationships>", rel_tag + b"</Relationships>", 1)
    parts["[Content_Types].xml"] = types.replace(b"</Types>", WORKSHEET_CT + b"</Types>", 1)

    fd, tmp = tempfile.mkstemp(dir=str(Path(out_path).parent), suffix=".xlsx")
    os.close(fd)
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as out:
        for item, _ in items:
            info = zipfile.ZipInfo(item.filename, date_time=item.date_time)
            info.compress_type = item.compress_type
            info.external_attr = item.external_attr
            out.writestr(info, parts[item.filename])
        out.writestr(ASSUMPTIONS_PART, _assumptions_sheet_xml(rows))
    os.replace(tmp, out_path)
    return name


# ---------------------------------------------------------------------------
# rewriting
# ---------------------------------------------------------------------------

def mask_sheet(xml, keep, deny, tally):
    """Blank every formula cell outside ``keep``; typed cells always survive.

    A typed value can never be recomputed downstream, so blanking one is only
    safe under a completeness guarantee the graph cannot always give. The only
    typed cells removed are the ``deny`` set -- pasted duplicates of output
    values, which would hand the rebuild its answer.
    """
    def fix(match):
        cell = match.group(0)
        head = cell[:cell.index(b">")] if b">" in cell else cell
        ref = attr(head, b"r")
        if ref is None:
            return cell
        style = attr(head, b"s")
        kind = attr(head, b"t")
        row, col = coordinate_to_tuple(ref.decode("ascii"))
        keeper = (row, col) in keep
        blanked = build_cell(ref, style, None, b"")

        if FORMULA_RE.search(cell) is None:
            if b"<v" not in cell and b"<is" not in cell:
                return cell  # carries nothing but formatting
            if (row, col) in deny:
                tally["redacted"] += 1
                return blanked
            tally["kept"] += 1
            return cell

        if not keeper:
            tally["blanked"] += 1
            return blanked

        value = VALUE_RE.search(cell)
        if value is None:
            # a formula Excel never calculated has no result to keep
            tally["uncached"] += 1
            return blanked
        text = value.group(1) or b""
        if kind == b"str":
            # a formula's string result; with the formula gone it has to inline
            tally["frozen"] += 1
            return build_cell(ref, style, b"inlineStr",
                              b'<is><t xml:space="preserve">' + text + b"</t></is>")
        if not text.strip():
            tally["blanked"] += 1
            return blanked
        tally["frozen"] += 1
        return build_cell(ref, style, kind, b"<v>" + text + b"</v>")

    return CELL_RE.sub(fix, xml)


def write_masked(src_path, out_path, keep, deny, tally):
    with zipfile.ZipFile(src_path) as src:
        part_keep = {part: (keep.get(name, set()), deny.get(name, set()))
                     for name, part in sheet_parts(src)}
        with zipfile.ZipFile(out_path, "w", zipfile.ZIP_DEFLATED) as out:
            for item in src.infolist():
                if item.filename == CALC_CHAIN:
                    continue  # nothing left to calculate
                data = src.read(item.filename)
                if item.filename in part_keep:
                    sheet_keep, sheet_deny = part_keep[item.filename]
                    data = mask_sheet(data, sheet_keep, sheet_deny, tally)
                elif is_chart_part(item.filename):
                    # charts cache every plotted value, including cells the
                    # mask just blanked; Excel rebuilds the cache from the
                    # grid on open
                    data, hits = scrub_chart_caches(item.filename, data)
                    tally["chart_caches"] += hits
                elif item.filename == "[Content_Types].xml":
                    data = CALC_OVERRIDE_RE.sub(b"", data)
                elif item.filename == "xl/_rels/workbook.xml.rels":
                    data = CALC_REL_RE.sub(b"", data)
                info = zipfile.ZipInfo(item.filename, date_time=item.date_time)
                info.compress_type = item.compress_type
                info.external_attr = item.external_attr
                out.writestr(info, data)


# ---------------------------------------------------------------------------
# checking
# ---------------------------------------------------------------------------

def verify(out_path, src_path, keep, frontier, formula_coords, deny,
           assumptions_sheet=None, deny_numbers=(), deny_texts=(),
           forbidden_formulas=()):
    """Confirm the mask leaked nothing and cost nothing.

    Ways this can go wrong, all worth catching: a formula survives and gives
    the answer away, a formula-derived number survives outside the keep set, a
    redacted pasted answer survives, a chart part still caches blanked values,
    the assumptions sheet prints a masked value or an output cell's formula,
    or a typed cell / input is lost or changed so the workbook can no longer
    be rebuilt.
    """
    warnings.simplefilter("ignore")
    masked = openpyxl.load_workbook(out_path)
    original = openpyxl.load_workbook(src_path, data_only=True)
    faults = {"formula": [], "leaked": [], "lost": [], "changed": []}

    # No chart part may still carry cached series values: re-scrubbing the
    # written file must be a no-op. And a strict zero-formula census: no sheet
    # part of the written file may retain any <f> element at all -- ordinary,
    # shared, array, or dataTable -- regardless of how openpyxl would present
    # the cell.
    with zipfile.ZipFile(out_path) as zf:
        for name in zf.namelist():
            if is_chart_part(name):
                _, hits = scrub_chart_caches(name, zf.read(name))
                if hits:
                    faults["leaked"].append("%s (chart value cache)" % name)
            elif name.startswith(FORMULA_PART_PREFIXES) and name.endswith(".xml"):
                match = FORMULA_RE.search(zf.read(name))
                if match:
                    faults["formula"].append(
                        "%s (<f> element survived: %s)"
                        % (name, match.group(0)[:80].decode("utf-8", "replace")))

    forbidden_formulas = [f for f in forbidden_formulas if str(f).strip()]

    for ws in masked.worksheets:
        allowed = keep.get(ws.title, set())
        derived = formula_coords.get(ws.title, set())
        denied = deny.get(ws.title, set())
        on_assumptions = assumptions_sheet is not None and ws.title == assumptions_sheet
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if value is None:
                    continue
                spot = (cell.row, cell.column)
                where = "%s!%s%d" % (ws.title, get_column_letter(cell.column),
                                     cell.row)
                if isinstance(value, ArrayFormula):
                    faults["formula"].append("%s (array formula)" % where)
                    continue
                if isinstance(value, str):
                    if cell.data_type == "f":
                        faults["formula"].append(where)
                    elif value.startswith("=") and spot in derived:
                        # A frozen formula result that still reads like a
                        # formula is indistinguishable from a leak; refuse it.
                        # Typed text that merely looks like a formula (e.g. a
                        # literal "=" label on a dashboard) is original golden
                        # content the mask must preserve, not a survivor.
                        faults["formula"].append(
                            "%s (formula-like derived text)" % where)
                    elif any(f in value for f in forbidden_formulas):
                        faults["formula"].append("%s (output cell formula text)"
                                                 % where)
                    elif on_assumptions and _value_matches(
                            value, deny_numbers, deny_texts):
                        faults["leaked"].append("%s=%r (masked value on "
                                                "assumptions sheet)" % (where, value))
                    continue
                if on_assumptions:
                    if _value_matches(value, deny_numbers, deny_texts):
                        faults["leaked"].append("%s=%r (masked value on "
                                                "assumptions sheet)" % (where, value))
                    continue
                if spot in denied:
                    faults["leaked"].append("%s=%r (pasted answer)" % (where, value))
                elif spot in derived and spot not in allowed:
                    faults["leaked"].append("%s=%r" % (where, value))

    # Every typed cell of the original must survive unchanged, and so must the
    # frozen values of the input frontier.
    for ws in original.worksheets:
        if ws.title not in masked.sheetnames:
            continue
        derived = formula_coords.get(ws.title, set())
        denied = deny.get(ws.title, set())
        must_hold = frontier.get(ws.title, set())
        now = masked[ws.title]
        cells = getattr(ws, "_cells", None)
        iterable = cells.values() if cells is not None else (
            c for row in ws.iter_rows() for c in row)
        for cell in iterable:
            before = cell.value
            if before is None:
                continue
            spot = (cell.row, cell.column)
            if spot in denied:
                continue
            typed = spot not in derived
            if not typed and spot not in must_hold:
                continue
            after = now.cell(row=cell.row, column=cell.column).value
            where = "%s!%s%d" % (ws.title, get_column_letter(cell.column),
                                 cell.row)
            if after is None:
                faults["lost"].append(where)
            elif isinstance(before, float) and isinstance(after, float):
                if abs(before - after) > 1e-9 * max(1.0, abs(before)):
                    faults["changed"].append(where)
            elif isinstance(before, (datetime.datetime, datetime.date)):
                continue  # date serials round-trip through XML formatting
            elif before != after:
                faults["changed"].append(where)

    return faults


# ---------------------------------------------------------------------------
# driver
# ---------------------------------------------------------------------------

def process(wb_id, args):
    pipeline_context = None
    pinned = bool(
        getattr(args, "release_root", None)
        or getattr(args, "source_generation_id", None)
        or getattr(args, "segmentation_generation_id", None)
    )
    if pinned:
        try:
            from xl_release_publication import (
                ReleasePublicationError,
                resolve_build_context,
            )

            pipeline_context = resolve_build_context(
                wb_id,
                release_root=(
                    Path(args.release_root) / wb_id
                    if getattr(args, "release_root", None)
                    else None
                ),
                release_id=getattr(args, "release_id", None),
                source_root=Path(args.source_generation_root) / wb_id,
                source_generation_id=getattr(args, "source_generation_id", None),
                segmentation_root=Path(args.seg_dir) / wb_id,
                segmentation_generation_id=getattr(
                    args, "segmentation_generation_id", None
                ),
            )
            expected = getattr(args, "expected_generation_id", None)
            if (
                expected is not None
                and pipeline_context["bindings"]["segmentation_generation_id"]
                != expected
            ):
                raise ReleasePublicationError(
                    "pinned segmentation generation changed"
                )
        except (OSError, ReleasePublicationError) as exc:
            print("  %s: pinned release/candidate gate failed: %s" % (wb_id, exc))
            return False
        source = Path(pipeline_context["source_path"])
        seg_dir = Path(pipeline_context["segmentation_dir"])
        generation_manifest = pipeline_context["segmentation_manifest"]
        ast_root = Path(pipeline_context["ast_root"])
    else:
        source = None
        for suffix in (".xlsx", ".xlsm"):
            candidate = Path(args.source) / (wb_id + suffix)
            if candidate.exists():
                source = candidate
                break
        if source is None:
            print("  %s: no workbook in %s" % (wb_id, args.source))
            return False
        ast_root = Path(getattr(args, "ast_dir", "ast_out"))
    segmentation_mode = getattr(args, "segmentation_mode", "legacy")
    if not pinned:
        try:
            seg_dir, generation_manifest = resolve_for_consumer(
                Path(args.seg_dir) / wb_id,
                mode=segmentation_mode,
                source_path=source,
                ast_dir=ast_root / wb_id,
                require_pass=True,
                expected_generation_id=getattr(
                    args, "expected_generation_id", None
                ),
            )
        except GenerationValidationError as exc:
            print("  %s: segmentation gate failed: %s" % (wb_id, exc))
            return False
    bands = seg_dir / "bands.csv"
    if not bands.exists():
        print("  %s: no bands.csv under %s" % (wb_id, bands.parent))
        return False

    warnings.simplefilter("ignore")
    book = openpyxl.load_workbook(source, data_only=True)
    keep, admitted = keep_cells(bands, args.keep, book)
    frontier = input_cells(bands)
    proof_frontier = stabilized_proof_inputs(seg_dir)
    if proof_frontier is not None:
        # ``keep`` only controls which formula caches are frozen; typed cells
        # survive independently. Replace declared formula-frontier admissions
        # with the strict proof frontier so formula literals remain blank and
        # must be rebuilt through their AST logic.
        for sheet, spots in frontier.items():
            keep[sheet].difference_update(spots)
        for sheet, spots in proof_frontier.items():
            keep[sheet].update(spots)
        frontier = proof_frontier
    formula_coords = formula_coordinates(source)

    proof = frontier_proof(seg_dir)
    outputs = output_values(bands, book)
    deny, denied_report, paste_suspects = pasted_answers(
        book, formula_coords, frontier, outputs)

    # Externally-sourced variables: blank them everywhere. Removing them from
    # ``keep``/``frontier`` blanks frozen formula hosts too and tells verify()
    # the blanks are intentional.
    masked_external = 0
    if getattr(args, "mask_cells", ""):
        for sheet, spots in load_mask_cells(args.mask_cells).items():
            masked_external += len(spots)
            deny[sheet].update(spots)
            keep[sheet].difference_update(spots)
            frontier[sheet].difference_update(spots)

    out_dir = Path(args.out)
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / ("%s-inputs%s" % (wb_id, source.suffix))

    # Original values of every denied cell: the assumptions sheet must not
    # print a number the grid mask just removed.
    deny_numbers, deny_texts = set(), set()
    for sheet, spots in deny.items():
        if sheet not in book.sheetnames:
            continue
        ws = book[sheet]
        for row, col in spots:
            value = ws.cell(row=row, column=col).value
            if value is None or isinstance(value, bool):
                continue
            if isinstance(value, (int, float)):
                deny_numbers.add(float(value))
            elif isinstance(value, str) and value.strip():
                deny_texts.add(value.strip())

    tally = Counter()
    write_masked(source, out_path, keep, deny, tally)

    outputs_at = output_cells(bands)
    assumptions = embedded_assumptions(seg_dir)
    # The exact derivations of answer cells must not appear anywhere as text.
    forbidden_formulas = sorted({
        str(e.get("formula", "")) for e in assumptions
        if _entry_host(e) in outputs_at and str(e.get("formula", "")).strip()
    })
    assumptions, redacted_rows, stripped_rows = redact_assumptions(
        assumptions, deny, deny_numbers, deny_texts, outputs_at
    )
    assumptions_sheet = None
    if assumptions:
        with zipfile.ZipFile(source) as zf:
            existing = {name for name, _ in sheet_parts(zf)}
        assumptions_sheet = inject_assumptions(out_path, assumptions, existing)

    faults = verify(out_path, source, keep, frontier, formula_coords, deny,
                    assumptions_sheet, deny_numbers, deny_texts,
                    forbidden_formulas)
    broken = sum(len(v) for v in faults.values())
    print("  %s  kept %d  frozen %d  blanked %d  redacted %d  ->  %s (%s)"
          % (wb_id, tally["kept"], tally["frozen"], tally["blanked"],
             tally["redacted"], out_path.name, human_size(out_path.stat().st_size)))
    print("       inputs %d cells over %d sheets; every typed cell preserved"
          % (sum(len(c) for c in frontier.values()), len(frontier)))
    print("       frontier sufficiency proof (stage 9): %s" % proof)
    if proof != "PASS":
        print("       note: without a PASS, only the typed-cell guarantee "
              "protects rebuildability; nothing typed was blanked")
    if denied_report:
        print("       redacted %d typed duplicate(s) of output values: %s"
              % (len(denied_report), ", ".join(denied_report[:6])))
    if paste_suspects:
        print("       WARNING: %d possible pasted answer(s) kept -- review "
              "before shipping:" % len(paste_suspects))
        for line in paste_suspects[:10]:
            print("         %s" % line)
        if len(paste_suspects) > 10:
            print("         ... and %d more" % (len(paste_suspects) - 10))
    if masked_external:
        print("       masked %d externally-sourced cell(s) listed in %s"
              % (masked_external, args.mask_cells))
    if tally["chart_caches"]:
        print("       %d cached chart series scrubbed (values re-read from "
              "the grid on open)" % tally["chart_caches"])
    if assumptions_sheet:
        print("       %d hardcoded formula constants -> sheet %r"
              % (len(assumptions), assumptions_sheet))
    if redacted_rows:
        print("       %d assumption row(s) redacted (masked values): %s"
              % (len(redacted_rows), ", ".join(redacted_rows[:6])))
    if stripped_rows:
        print("       %d assumption formula(s) withheld (output cells or "
              "masked literals): %s"
              % (len(stripped_rows), ", ".join(stripped_rows[:6])))
    if admitted:
        derived = Counter(b for b in admitted.values()
                          if b in ("middle", "output"))
        note = "" if not derived else "; %d from %s" % (
            sum(derived.values()), "/".join(sorted(derived)))
        print("       period headers %d kept though not inputs%s"
              % (len(admitted), note))
    if not broken:
        if generation_manifest is not None:
            write_inputs_sidecar(
                out_path,
                seg_dir,
                generation_manifest,
                pipeline_bindings=(
                    pipeline_context["bindings"]
                    if pipeline_context is not None
                    else None
                ),
            )
        print("       verified: no formulas, no derived numbers, typed cells intact")
        return True
    for name, hits in faults.items():
        if hits:
            print("       FAIL %s (%d): %s" % (name, len(hits),
                                               ", ".join(hits[:6])))
    return False


def main(argv=None):
    parser = argparse.ArgumentParser(
        description=__doc__.split("\n")[0],
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="python3 xl_input_mask.py 0248 0262 0449 0450")
    parser.add_argument("workbooks", nargs="+",
                        help="workbook ids under --seg-dir")
    parser.add_argument("--seg-dir", default="seg_out",
                        help="where xl_segment.py wrote bands.csv")
    parser.add_argument(
        "--ast-dir",
        default="ast_out",
        help="AST root used to validate strict segmentation fingerprints",
    )
    parser.add_argument(
        "--segmentation-mode",
        choices=("strict", "shadow", "legacy"),
        default="strict",
        help="strict production gate (default), versioned shadow, or explicit "
             "unversioned legacy downgrade",
    )
    parser.add_argument(
        "--expected-generation-id",
        default=None,
        help="fail if current.json no longer names this generation",
    )
    parser.add_argument(
        "--release-root",
        default=None,
        help="root containing <workbook>/current-release.json",
    )
    parser.add_argument("--release-id", default=None)
    parser.add_argument(
        "--source-generation-root",
        default="source_out",
        help="root containing <workbook>/generations/<source-generation-id>",
    )
    parser.add_argument("--source-generation-id", default=None)
    parser.add_argument("--segmentation-generation-id", default=None)
    parser.add_argument("--source", default="4-10 100",
                        help="folder holding <wb>.xlsx or <wb>.xlsm")
    parser.add_argument("-o", "--out", default="inputs_out")
    parser.add_argument("--keep", choices=KEEP_MODES, default="headers",
                        help="how much non-input content survives")
    parser.add_argument("--mask-cells", default="",
                        help="JSON list of Sheet!Cell refs (ranges allowed) to "
                             "additionally blank; used for variables served "
                             "through a mock MCP research service")
    args = parser.parse_args(argv)
    if args.release_id and not args.release_root:
        parser.error("--release-id requires --release-root")
    if args.release_root and (
        args.source_generation_id or args.segmentation_generation_id
    ):
        parser.error(
            "--release-root is mutually exclusive with explicit candidate IDs"
        )
    if bool(args.source_generation_id) != bool(args.segmentation_generation_id):
        parser.error(
            "explicit staging requires both --source-generation-id and "
            "--segmentation-generation-id"
        )

    print("masking to inputs only (--keep %s)" % args.keep)
    ok = [process(wb, args) for wb in args.workbooks]
    print("%d/%d verified" % (sum(ok), len(ok)))
    return 0 if all(ok) else 1


if __name__ == "__main__":
    sys.exit(main())
