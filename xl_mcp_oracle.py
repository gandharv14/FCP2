#!/usr/bin/env python3
"""Validate a shipped workbook/MCP bundle against its live MCP endpoint.

The authoring MCP directory supplies evaluation-only manifests; it must remain
outside the shipped task bundle.  ``fastmcp`` and ``openpyxl`` are imported only
inside the checks that need them, so this module remains cheap to import in
tests and other pipeline commands.

Example:

    uv run --python 3.12 --with fastmcp --with openpyxl \
      python xl_mcp_oracle.py \
      --bundle ../tasks_outputs/0233-outputs \
      --mcp ../runs/0233-variable-sources/mcp \
      --url http://127.0.0.1:18233/mcp

Legitimate duplicate workbook values require an explicit JSON allowlist:

    {
      "workbook_value_leaks": [
        {
          "cell": "Model!A1",
          "variable_id": "some-variable",
          "reason": "required period-axis label"
        }
      ]
    }

Unknown, duplicate, or unused allowlist entries are failures.  There are no
implicit sheet, value, or path exclusions.
"""

from __future__ import annotations

import argparse
import asyncio
import datetime as dt
import decimal
import json
import math
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable, Iterable

from plain_eligibility import DIALOGUE_NOTES_NAME, dialogue_notes_expected


DIMENSIONS = ("metric", "entity", "period", "scenario", "basis", "unit", "status")
FORBIDDEN_RUNTIME_KEYS = {
    "answer",
    "cell_value",
    "correct_answer",
    "gold",
    "gold_evidence",
    "is_truth",
    "mcp_value",
    "normalized",
    "required_dimensions",
    "resolution_rule",
    "source_snapshot",
    "supported",
    "target_value",
    "task_id",
    "workbook",
}
PROFILE_KEYS = {"profile_excerpt", "profile_excerpts"}
RENDERED_EXCERPT_RE = re.compile(
    r"Short attributed excerpt\s*[—-]\s*"
    r"(?P<attribution>[^\n(]+?)\s*\((?P<locator>[^)\n]+)\):\s*"
    r"[“\"](?P<text>.*?)[”\"]",
)
NUMBER_RE = re.compile(
    r"(?<![A-Za-z0-9.])[-+]?(?:\d{1,3}(?:,\d{3})+|\d+)"
    r"(?:\.\d+)?(?:[eE][-+]?\d+)?%?(?![A-Za-z0-9.])"
)
SIMPLE_NUMBER_RE = re.compile(
    r"^\s*(?:[A-Z]{3}\s*|[$€£]\s*)?"
    r"(?P<open>\()?\s*(?P<number>[-+]?(?:\d{1,3}(?:[,\s]\d{3})+|\d+)"
    r"(?:\.\d+)?(?:[eE][-+]?\d+)?)\s*(?P<percent>%?)\s*(?(open)\))"
    r"(?:\s*[A-Za-z][A-Za-z0-9 ./()_-]*)?\s*$"
)
DATE_FORMATS = (
    "%Y-%m-%d",
    "%Y/%m/%d",
    "%m/%d/%Y",
    "%d/%m/%Y",
    "%d-%b-%Y",
    "%d %b %Y",
    "%b %d, %Y",
)


class OracleInputError(ValueError):
    """The command inputs or manifests are malformed."""


@dataclass(frozen=True)
class ValueTarget:
    variable_id: str
    raw: Any
    mcp: Any
    unit: str


def load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def read_jsonl(path: Path) -> list[dict[str, Any]]:
    return [
        json.loads(line)
        for line in path.read_text(encoding="utf-8").splitlines()
        if line.strip()
    ]


def json_key(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def normalized_text(value: Any) -> str:
    return " ".join(str(value).strip().casefold().split())


def normalized_ref(ref: str) -> str:
    sheet, separator, coordinate = str(ref).rpartition("!")
    if not separator or not sheet or not coordinate:
        raise OracleInputError("invalid workbook reference: %r" % ref)
    return "%s!%s" % (sheet.strip().strip("'"), coordinate.upper())


def display_ref(ref: str) -> str:
    sheet, _, coordinate = normalized_ref(ref).rpartition("!")
    if any(char in sheet for char in " -&()"):
        sheet = "'%s'" % sheet.replace("'", "''")
    return "%s!%s" % (sheet, coordinate)


def expand_refs(refs: Iterable[str]) -> list[str]:
    """Expand qualified ranges while keeping openpyxl an optional import."""
    from openpyxl.utils import get_column_letter, range_boundaries

    expanded: list[str] = []
    for original in refs:
        ref = normalized_ref(original)
        sheet, _, coordinates = ref.rpartition("!")
        try:
            min_col, min_row, max_col, max_row = range_boundaries(coordinates)
        except ValueError as exc:
            raise OracleInputError("invalid workbook reference: %r" % original) from exc
        for row in range(min_row, max_row + 1):
            for column in range(min_col, max_col + 1):
                expanded.append(
                    "%s!%s%d" % (sheet, get_column_letter(column), row)
                )
    return expanded


def recursive_keys(value: Any) -> set[str]:
    found: set[str] = set()
    if isinstance(value, dict):
        for key, child in value.items():
            found.add(str(key).casefold())
            found.update(recursive_keys(child))
    elif isinstance(value, list):
        for child in value:
            found.update(recursive_keys(child))
    return found


def _as_date(value: Any) -> dt.date | None:
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    if not isinstance(value, str):
        return None
    text = value.strip()
    for date_format in DATE_FORMATS:
        try:
            return dt.datetime.strptime(text, date_format).date()
        except ValueError:
            pass
    return None


def _as_decimal(value: Any) -> decimal.Decimal | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float, decimal.Decimal)):
        if isinstance(value, float) and not math.isfinite(value):
            return None
        return decimal.Decimal(str(value))
    return None


def _numeric_string(value: str) -> tuple[decimal.Decimal, bool] | None:
    match = SIMPLE_NUMBER_RE.fullmatch(value)
    if not match:
        return None
    token = match.group("number").replace(",", "").replace(" ", "")
    try:
        number = decimal.Decimal(token)
    except decimal.InvalidOperation:
        return None
    if match.group("open"):
        number = -abs(number)
    return number, bool(match.group("percent"))


def _decimals_equal(left: decimal.Decimal, right: decimal.Decimal) -> bool:
    tolerance = decimal.Decimal("1e-9") * max(
        decimal.Decimal(1), abs(left), abs(right)
    )
    return abs(left - right) <= tolerance


def value_matches_target(value: Any, target: ValueTarget) -> bool:
    """Match typed and common displayed representations of a masked value."""
    candidates = (target.raw, target.mcp)
    boolean_candidates = [
        candidate for candidate in candidates if isinstance(candidate, bool)
    ]
    if boolean_candidates:
        if isinstance(value, bool) and value in boolean_candidates:
            return True
        if isinstance(value, str) and normalized_text(value) in {
            str(candidate).casefold() for candidate in boolean_candidates
        }:
            return True

    value_date = _as_date(value)
    if value_date is not None:
        if any(_as_date(candidate) == value_date for candidate in candidates):
            return True

    numeric_candidates = [
        number for number in (_as_decimal(candidate) for candidate in candidates)
        if number is not None
    ]
    value_number = _as_decimal(value)
    if value_number is not None and any(
        _decimals_equal(value_number, candidate) for candidate in numeric_candidates
    ):
        return True

    if isinstance(value, str):
        parsed = _numeric_string(value)
        if parsed is not None:
            number, is_percent = parsed
            represented = [number]
            if is_percent:
                represented.append(number / decimal.Decimal(100))
            if any(
                _decimals_equal(form, candidate)
                for form in represented
                for candidate in numeric_candidates
            ):
                return True

    string_candidates = [
        normalized_text(candidate)
        for candidate in candidates
        if isinstance(candidate, str) and _as_date(candidate) is None
    ]
    return isinstance(value, str) and normalized_text(value) in string_candidates


RENDERING_KINDS = (
    "raw_number",
    "formatted_text",
    "percent_rendering",
    "date_rendering",
    "other",
)


def classify_rendering(value: Any, target: ValueTarget) -> str:
    """How a duplicated masked value appears in a leaking cell (closed enum).

    One of ``RENDERING_KINDS``: "raw_number" for a typed numeric cell equal to
    the raw value, "percent_rendering" for percent-scaled numerics or "12.5%"
    style text, "date_rendering" for typed dates or date strings,
    "formatted_text" for the value embedded in a text rendering ("42 MW",
    concatenated labels, exact string duplicates), and "other" for anything
    else (booleans, unclassifiable representations).
    """
    if isinstance(value, bool):
        return "other"
    if isinstance(value, (dt.date, dt.datetime)):
        return "date_rendering"
    number = _as_decimal(value)
    if number is not None:
        raw = _as_decimal(target.raw)
        if raw is not None and _decimals_equal(number, raw):
            return "raw_number"
        if target.unit.strip().casefold() == "percent":
            return "percent_rendering"
        return "other"
    if isinstance(value, str):
        if _as_date(value) is not None:
            return "date_rendering"
        parsed = _numeric_string(value)
        if parsed is not None and parsed[1]:
            return "percent_rendering"
        return "formatted_text"
    return "other"


def text_leaks_target(text: str, target: ValueTarget) -> bool:
    """Whether free text contains a masked value representation."""
    for candidate in (target.raw, target.mcp):
        if isinstance(candidate, bool) and re.search(
            r"(?<!\w)%s(?!\w)" % str(candidate),
            text,
            re.IGNORECASE,
        ):
            return True
        date = _as_date(candidate)
        if date is not None and date.isoformat() in text:
            return True
        if isinstance(candidate, str) and date is None:
            phrase = candidate.strip()
            if phrase and re.search(
                r"(?<!\w)%s(?!\w)" % re.escape(phrase), text, re.IGNORECASE
            ):
                return True

    numeric_candidates = [
        number
        for number in (_as_decimal(target.raw), _as_decimal(target.mcp))
        if number is not None
    ]
    for match in NUMBER_RE.finditer(text):
        parsed = _numeric_string(match.group())
        if parsed is None:
            continue
        number, is_percent = parsed
        represented = [number]
        if is_percent:
            represented.append(number / decimal.Decimal(100))
        if any(
            _decimals_equal(form, candidate)
            for form in represented
            for candidate in numeric_candidates
        ):
            return True
    return False


def value_targets(audit: list[dict[str, Any]]) -> list[ValueTarget]:
    targets: list[ValueTarget] = []
    seen: set[str] = set()
    for index, entry in enumerate(audit):
        variable_id = entry.get("variable_id")
        if not isinstance(variable_id, str) or not variable_id:
            raise OracleInputError("masked_inputs[%d] has no variable_id" % index)
        if variable_id in seen:
            raise OracleInputError("duplicate masked variable_id: %s" % variable_id)
        seen.add(variable_id)
        if "cell_value" not in entry or "mcp_value" not in entry:
            raise OracleInputError(
                "%s lacks cell_value or mcp_value" % variable_id
            )
        targets.append(
            ValueTarget(
                variable_id=variable_id,
                raw=entry["cell_value"],
                mcp=entry["mcp_value"],
                unit=str(entry.get("unit") or ""),
            )
        )
    return targets


def load_allowlist(
    path: Path | None, variable_ids: set[str]
) -> tuple[dict[tuple[str, str], str], list[str]]:
    if path is None:
        return {}, []
    data = load_json(path)
    if not isinstance(data, dict):
        raise OracleInputError("allowlist must be a JSON object")
    unknown_keys = sorted(set(data) - {"workbook_value_leaks"})
    if unknown_keys:
        raise OracleInputError(
            "unknown allowlist keys: %s" % ", ".join(unknown_keys)
        )
    rows = data.get("workbook_value_leaks", [])
    if not isinstance(rows, list):
        raise OracleInputError("workbook_value_leaks must be a JSON list")

    allowed: dict[tuple[str, str], str] = {}
    errors: list[str] = []
    for index, row in enumerate(rows):
        if not isinstance(row, dict):
            errors.append("allowlist entry %d is not an object" % index)
            continue
        unknown = sorted(set(row) - {"cell", "variable_id", "reason"})
        if unknown:
            errors.append(
                "allowlist entry %d has unknown keys: %s"
                % (index, ", ".join(unknown))
            )
            continue
        variable_id = row.get("variable_id")
        reason = row.get("reason")
        try:
            cell = normalized_ref(row.get("cell", ""))
        except OracleInputError as exc:
            errors.append(str(exc))
            continue
        if variable_id not in variable_ids:
            errors.append(
                "allowlist entry %d has unknown variable_id %r"
                % (index, variable_id)
            )
            continue
        if not isinstance(reason, str) or not reason.strip():
            errors.append("allowlist entry %d requires a reason" % index)
            continue
        key = (cell, variable_id)
        if key in allowed:
            errors.append(
                "duplicate allowlist entry for %s / %s"
                % (display_ref(cell), variable_id)
            )
            continue
        allowed[key] = reason.strip()
    return allowed, errors


def resolve_workbook(bundle: Path, requested: Path | None = None) -> Path:
    environment = bundle / "environment"
    if requested is not None:
        candidate = requested if requested.is_absolute() else environment / requested
        candidate = candidate.resolve()
        try:
            candidate.relative_to(environment.resolve())
        except ValueError as exc:
            raise OracleInputError("--workbook must be inside bundle/environment") from exc
        if candidate.parent != environment.resolve():
            raise OracleInputError("the shipped workbook must be at environment root")
        if not candidate.is_file():
            raise OracleInputError("workbook does not exist: %s" % candidate)
        return candidate

    candidates = sorted(
        path for path in environment.iterdir()
        if path.is_file() and path.suffix.casefold() in {".xlsx", ".xlsm"}
    )
    if len(candidates) != 1:
        raise OracleInputError(
            "expected exactly one workbook in %s; found %d (use --workbook)"
            % (environment, len(candidates))
        )
    return candidates[0]


def check_environment(bundle: Path, workbook: Path) -> dict[str, Any]:
    """Enforce the intentionally small shipped-environment file surface."""
    environment = bundle / "environment"
    compose_candidates = [
        name for name in ("docker-compose.yaml", "docker-compose.yml")
        if (environment / name).is_file()
    ]
    expected = {
        workbook.name,
        "Dockerfile",
        "mcp-server/Dockerfile",
        "mcp-server/server.py",
        "mcp-server/runtime/server.json",
        "mcp-server/runtime/sources.json",
        "mcp-server/runtime/datasets.json",
        "mcp-server/runtime/documents.jsonl",
        "mcp-server/runtime/records.jsonl",
    }
    if len(compose_candidates) == 1:
        expected.add(compose_candidates[0])
    if dialogue_notes_expected(bundle):
        expected.add(DIALOGUE_NOTES_NAME)

    actual: set[str] = set()
    symlinks: list[str] = []
    for path in environment.rglob("*"):
        relative = path.relative_to(environment).as_posix()
        if path.is_symlink():
            symlinks.append(relative)
        if path.is_file() or path.is_symlink():
            actual.add(relative)

    missing = sorted(expected - actual)
    unknown = sorted(actual - expected)
    if len(compose_candidates) != 1:
        missing.append("exactly one docker-compose.yaml/docker-compose.yml")

    runtime_root = environment / "mcp-server" / "runtime"
    runtime_key_leaks: list[str] = []
    runtime_objects: list[Any] = []
    for name in ("server.json", "sources.json", "datasets.json"):
        path = runtime_root / name
        if path.is_file():
            runtime_objects.append(load_json(path))
    for name in ("documents.jsonl", "records.jsonl"):
        path = runtime_root / name
        if path.is_file():
            runtime_objects.extend(read_jsonl(path))
    for key in sorted(
        set().union(*(recursive_keys(value) for value in runtime_objects))
        & FORBIDDEN_RUNTIME_KEYS
    ):
        runtime_key_leaks.append(key)

    workbook_files = sorted(
        relative for relative in actual
        if Path(relative).suffix.casefold() in {".xlsx", ".xlsm"}
    )
    return {
        "valid": not (missing or unknown or symlinks or runtime_key_leaks),
        "missing_files": missing,
        "unknown_files": unknown,
        "symlinks": sorted(symlinks),
        "workbooks": workbook_files,
        "forbidden_runtime_keys": runtime_key_leaks,
    }


def check_workbook(
    workbook: Path,
    mask_refs: list[str],
    audit: list[dict[str, Any]],
    allowlist: dict[tuple[str, str], str],
) -> dict[str, Any]:
    import openpyxl

    book = openpyxl.load_workbook(workbook, data_only=False, read_only=True)
    try:
        expanded_mask = set(expand_refs(mask_refs))
        targets = value_targets(audit)
        audit_refs: set[str] = set()
        manifest_errors: list[str] = []
        for entry in audit:
            refs = entry.get("refs")
            if not isinstance(refs, list) or not refs:
                manifest_errors.append(
                    "%s has no refs" % entry.get("variable_id", "<unknown>")
                )
                continue
            audit_refs.update(expand_refs(refs))
        if audit_refs != expanded_mask:
            for ref in sorted(expanded_mask - audit_refs):
                manifest_errors.append("mask ref absent from audit: %s" % display_ref(ref))
            for ref in sorted(audit_refs - expanded_mask):
                manifest_errors.append("audit ref absent from mask: %s" % display_ref(ref))

        not_blank: list[dict[str, Any]] = []
        for ref in sorted(expanded_mask):
            sheet, _, coordinate = ref.rpartition("!")
            if sheet not in book.sheetnames:
                not_blank.append({"cell": display_ref(ref), "value": "<unknown sheet>"})
                continue
            value = book[sheet][coordinate].value
            if value is not None:
                not_blank.append({"cell": display_ref(ref), "value": repr(value)})

        found: dict[tuple[str, str], dict[str, Any]] = {}
        for sheet in book.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    ref = "%s!%s" % (sheet.title, cell.coordinate)
                    if ref in expanded_mask:
                        continue
                    for target in targets:
                        if value_matches_target(cell.value, target):
                            key = (ref, target.variable_id)
                            found[key] = {
                                "cell": display_ref(ref),
                                "variable_id": target.variable_id,
                                "value": repr(cell.value),
                                "rendering": classify_rendering(
                                    cell.value, target
                                ),
                            }

        found_keys = set(found)
        approved_keys = found_keys & set(allowlist)
        unapproved = [
            found[key] for key in sorted(found_keys - approved_keys)
        ]
        approved = [
            {**found[key], "reason": allowlist[key]}
            for key in sorted(approved_keys)
        ]
        unused = [
            {
                "cell": display_ref(cell),
                "variable_id": variable_id,
                "reason": reason,
            }
            for (cell, variable_id), reason in sorted(allowlist.items())
            if (cell, variable_id) not in found_keys
        ]
        return {
            "valid": not (manifest_errors or not_blank or unapproved or unused),
            "mask_cells": len(expanded_mask),
            "masked_cells_not_blank": not_blank,
            "manifest_errors": manifest_errors,
            "unapproved_value_leaks": unapproved,
            "approved_value_leaks": approved,
            "unused_allowlist_entries": unused,
        }
    finally:
        book.close()


def scan_value_leaks(
    workbook: Path,
    mask_refs: list[str],
    audit: list[dict[str, Any]],
    allowlist_path: Path | None = None,
    *,
    require_blank_mask: bool = True,
) -> dict[str, Any]:
    """Offline masked-value duplicate/leak scan of one workbook.

    This is the exact detector the HTTP oracle applies to the delivered
    workbook: the masked-cell blanking check, the duplicate representation
    scan (typed values plus text/date/percent renderings), and explicit
    allowlist handling.  It needs no live MCP server and no docker, so it can
    run at maskability-review time against the golden or baseline inputs
    workbook.  Pass ``require_blank_mask=False`` for such pre-masking scans:
    masked cells are then expected to still hold their raw values and only
    duplicate representations outside the mask (and allowlist hygiene) decide
    validity.
    """
    targets = value_targets(audit)
    allowlist, allowlist_errors = load_allowlist(
        allowlist_path, {target.variable_id for target in targets}
    )
    report = check_workbook(workbook, mask_refs, audit, allowlist)
    if not require_blank_mask:
        report["masked_cells_not_blank"] = []
        report["valid"] = not (
            report["manifest_errors"]
            or report["unapproved_value_leaks"]
            or report["unused_allowlist_entries"]
        )
    report["allowlist_errors"] = allowlist_errors
    if allowlist_errors:
        report["valid"] = False
    return report


def profile_excerpt_items(source: dict[str, Any]) -> list[tuple[str, Any]]:
    items: list[tuple[str, Any]] = []
    for key in PROFILE_KEYS:
        if key not in source:
            continue
        value = source[key]
        if key == "profile_excerpts" and isinstance(value, list):
            items.extend((key, item) for item in value)
        else:
            items.append((key, value))
    return items


def check_profile_excerpts(
    sources: list[dict[str, Any]], targets: list[ValueTarget]
) -> dict[str, Any]:
    violations: list[str] = []
    count = 0
    for source in sources:
        source_id = str(source.get("id") or "")
        source_name = str(source.get("name") or "")
        source_url = str(source.get("origin_url") or "")
        for key, item in profile_excerpt_items(source):
            count += 1
            if isinstance(item, str):
                text = item
                explicit_attribution: list[str] = []
            elif isinstance(item, dict):
                text = next(
                    (
                        str(item[name])
                        for name in ("excerpt", "text", "content")
                        if isinstance(item.get(name), str)
                    ),
                    "",
                )
                explicit_attribution = [
                    str(item[name])
                    for name in (
                        "attribution",
                        "origin_url",
                        "source_id",
                        "source_name",
                        "url",
                    )
                    if item.get(name)
                ]
            else:
                violations.append(
                    "%s.%s is neither text nor an excerpt object"
                    % (source_id or "<unknown>", key)
                )
                continue

            if not text.strip():
                violations.append("%s.%s is empty" % (source_id or "<unknown>", key))
                continue
            if not (source_id and source_name and source_url):
                violations.append(
                    "%s.%s lacks parent source id/name/origin_url attribution"
                    % (source_id or "<unknown>", key)
                )
            if explicit_attribution:
                normalized = {normalized_text(value) for value in explicit_attribution}
                expected = {
                    normalized_text(value)
                    for value in (source_id, source_name, source_url)
                    if value
                }
                if not (normalized & expected):
                    violations.append(
                        "%s.%s has attribution inconsistent with its source"
                        % (source_id or "<unknown>", key)
                    )
            leaking = sorted(
                target.variable_id
                for target in targets
                if text_leaks_target(text, target)
            )
            if leaking:
                violations.append(
                    "%s.%s contains masked values for: %s"
                    % (source_id or "<unknown>", key, ", ".join(leaking))
                )
    return {"valid": not violations, "count": count, "violations": violations}


def check_rendered_profile_excerpts(
    sources: list[dict[str, Any]],
    datasets: list[dict[str, Any]],
    documents: list[dict[str, Any]],
    targets: list[ValueTarget],
) -> dict[str, Any]:
    """Validate profile excerpts rendered into synthetic MCP documents."""
    source_by_id = {row.get("id"): row for row in sources}
    dataset_by_id = {row.get("id"): row for row in datasets}
    violations: list[str] = []
    count = 0
    for document in documents:
        document_id = str(document.get("id") or "<unknown>")
        content = str(document.get("content") or "")
        matches = list(RENDERED_EXCERPT_RE.finditer(content))
        if "Short attributed excerpt" in content and not matches:
            violations.append(
                "%s has a malformed rendered profile excerpt" % document_id
            )
            continue
        for match in matches:
            count += 1
            source = source_by_id.get(document.get("source_id")) or {}
            dataset = dataset_by_id.get(document.get("related_dataset_id")) or {}
            source_profile = source.get("profile_id")
            dataset_profile = dataset.get("profile_id")
            if not source_profile:
                violations.append(
                    "%s renders an excerpt without profiled source metadata"
                    % document_id
                )
            elif (
                document.get("source_id") == dataset.get("source_id")
                and dataset_profile
                and source_profile != dataset_profile
            ):
                violations.append(
                    "%s renders an excerpt with inconsistent profile attribution"
                    % document_id
                )
            if not str(source.get("origin_url") or "").startswith(
                ("http://", "https://")
            ):
                violations.append(
                    "%s excerpt source lacks a public origin URL" % document_id
                )
            attribution = match.group("attribution").strip()
            locator = match.group("locator").strip()
            text = match.group("text").strip()
            if not attribution or not locator:
                violations.append(
                    "%s excerpt lacks attribution or locator" % document_id
                )
            if not text or len(text) > 240:
                violations.append(
                    "%s excerpt is empty or exceeds 240 characters" % document_id
                )
            leaking = sorted(
                target.variable_id
                for target in targets
                if text_leaks_target(text, target)
            )
            if leaking:
                violations.append(
                    "%s excerpt contains masked values for: %s"
                    % (document_id, ", ".join(leaking))
                )
    return {"valid": not violations, "count": count, "violations": violations}


def merge_profile_reports(*reports: dict[str, Any]) -> dict[str, Any]:
    violations = [
        violation
        for report in reports
        for violation in report.get("violations", [])
    ]
    return {
        "valid": not violations,
        "count": sum(int(report.get("count", 0)) for report in reports),
        "violations": violations,
    }


def _tool_data(result: Any) -> Any:
    if getattr(result, "is_error", False):
        raise RuntimeError("MCP tool returned an error: %s" % result)
    return result.data if hasattr(result, "data") else result


async def query_all(
    client: Any,
    arguments: dict[str, Any],
    *,
    page_size: int = 2,
    max_pages: int = 1000,
) -> tuple[list[dict[str, Any]], int]:
    rows: list[dict[str, Any]] = []
    cursor: Any = 0
    seen_cursors: set[str] = set()
    expected_total: int | None = None
    pages = 0
    while True:
        cursor_key = json_key(cursor)
        if cursor_key in seen_cursors:
            raise RuntimeError("query_records pagination cursor repeated: %r" % cursor)
        seen_cursors.add(cursor_key)
        if pages >= max_pages:
            raise RuntimeError("query_records exceeded %d pages" % max_pages)
        payload = _tool_data(
            await client.call_tool(
                "query_records",
                {**arguments, "limit": page_size, "cursor": cursor},
            )
        )
        pages += 1
        if not isinstance(payload, dict) or not isinstance(payload.get("rows"), list):
            raise RuntimeError("query_records returned a malformed page")
        page = payload["rows"]
        if payload.get("returned") is not None and payload["returned"] != len(page):
            raise RuntimeError("query_records returned count does not match rows")
        total = payload.get("total_matches")
        if total is not None:
            if not isinstance(total, int) or total < 0:
                raise RuntimeError("query_records returned invalid total_matches")
            if expected_total is None:
                expected_total = total
            elif total != expected_total:
                raise RuntimeError("query_records total_matches changed between pages")
        rows.extend(page)
        next_cursor = payload.get("next_cursor")
        if next_cursor is None:
            break
        cursor = next_cursor

    ids = [str(row.get("id")) for row in rows]
    if len(ids) != len(set(ids)):
        raise RuntimeError("query_records repeated rows across pages")
    if expected_total is not None and len(rows) != expected_total:
        raise RuntimeError(
            "query_records pagination returned %d of %d rows"
            % (len(rows), expected_total)
        )
    return rows, pages


def _dimensions_equal(row: dict[str, Any], wanted: dict[str, Any]) -> bool:
    return all(
        normalized_text(row.get(field)) == normalized_text(wanted.get(field))
        for field in DIMENSIONS
    )


def validate_task_rows(
    task: dict[str, Any],
    rows: list[dict[str, Any]],
) -> tuple[list[str], dict[str, Any]]:
    task_id = str(task.get("task_id") or "<unknown>")
    evidence = task["evidence"]
    wanted = task["required_dimensions"]
    errors: list[str] = []
    if any(not _dimensions_equal(row, wanted) for row in rows):
        errors.append("fully-filtered query returned non-exact dimensions")

    current = [row for row in rows if not row.get("superseded_by")]
    if len(current) != 1:
        errors.append("expected one unsuperseded row; got %d" % len(current))
        return errors, {"task_id": task_id, "rows": len(rows), "stale": 0}
    supported = current[0]
    evidence_checks = {
        "id": "record_id",
        "dataset_id": "dataset_id",
        "document_id": "document_id",
        "release": "release",
        "source_id": "source_id",
    }
    for row_key, evidence_key in evidence_checks.items():
        if supported.get(row_key) != evidence.get(evidence_key):
            errors.append(
                "unsuperseded %s does not match evidence.%s"
                % (row_key, evidence_key)
            )
    if json_key(supported.get("value")) != json_key(task["answer"]["value"]):
        errors.append("unsuperseded value does not exactly match the answer")

    by_release: dict[str, dict[str, Any]] = {}
    for row in rows:
        release = row.get("release")
        if not isinstance(release, str) or not release:
            errors.append("row %s has no release" % row.get("id"))
            continue
        if release in by_release:
            errors.append("duplicate release %s" % release)
        by_release[release] = row

    stale = [row for row in rows if row.get("superseded_by")]
    if not stale:
        errors.append("no stale provenance release was returned")
    predecessor_counts = {release: 0 for release in by_release}
    for row in stale:
        successor = row.get("superseded_by")
        if successor in predecessor_counts:
            predecessor_counts[successor] += 1
    if stale and any(count > 1 for count in predecessor_counts.values()):
        errors.append("provenance releases branch instead of forming one chain")

    for row in rows:
        walk = row
        seen: set[str] = set()
        while walk.get("superseded_by"):
            release = walk.get("release")
            if release in seen:
                errors.append("provenance cycle from release %s" % row.get("release"))
                break
            seen.add(release)
            successor = by_release.get(walk["superseded_by"])
            if successor is None:
                errors.append(
                    "release %s points to unknown successor %s"
                    % (release, walk["superseded_by"])
                )
                break
            try:
                published = dt.date.fromisoformat(str(walk.get("published_at")))
                successor_published = dt.date.fromisoformat(
                    str(successor.get("published_at"))
                )
            except ValueError:
                errors.append(
                    "release %s has invalid provenance publication dates" % release
                )
                break
            if successor_published <= published:
                errors.append(
                    "release %s does not precede successor %s"
                    % (release, successor.get("release"))
                )
                break
            walk = successor
        else:
            if walk.get("id") != evidence.get("record_id"):
                errors.append(
                    "chain from %s does not terminate at evidence"
                    % row.get("release")
                )

    current_value = json_key(supported.get("value"))
    if any(json_key(row.get("value")) == current_value for row in stale):
        errors.append("a stale release duplicates the current value")
    # Even a non-identical stale value is a leak when it sits inside the
    # grader's acceptance band: picking it would still be graded correct.
    if isinstance(supported.get("value"), (int, float)) and not isinstance(
            supported.get("value"), bool):
        expected = float(supported["value"])
        band = 10 * max(1e-6, 1e-6 * abs(expected))
        for row in stale:
            value = row.get("value")
            if isinstance(value, (int, float)) and not isinstance(value, bool) \
                    and abs(float(value) - expected) <= band:
                errors.append(
                    "stale release %s is within grader tolerance of the "
                    "current value" % row.get("release"))
    return list(dict.fromkeys(errors)), {
        "task_id": task_id,
        "rows": len(rows),
        "stale": len(stale),
        "supported": supported,
    }


async def validate_documents(
    client: Any,
    rows: list[dict[str, Any]],
    evidence_release: str,
    sources: list[dict[str, Any]],
    datasets: list[dict[str, Any]],
    targets: list[ValueTarget],
) -> tuple[list[str], dict[str, Any]]:
    errors: list[str] = []
    fetched: list[dict[str, Any]] = []
    by_successor: dict[str, list[str]] = {}
    for row in rows:
        if row.get("superseded_by"):
            by_successor.setdefault(row["superseded_by"], []).append(row["release"])
    for row in rows:
        document_id = row.get("document_id")
        try:
            document = _tool_data(
                await client.call_tool(
                    "fetch_document", {"document_id": document_id}
                )
            )
        except Exception as exc:
            errors.append("cannot fetch document %s: %s" % (document_id, exc))
            continue
        fetched.append(document)
        for key, expected in (
            ("id", document_id),
            ("source_id", row.get("source_id")),
            ("related_dataset_id", row.get("dataset_id")),
        ):
            if document.get(key) != expected:
                errors.append("%s has wrong %s" % (document_id, key))
        content = str(document.get("content") or "")
        if row.get("release") not in content:
            errors.append("%s does not attribute release %s" % (
                document_id, row.get("release")
            ))
        successor = row.get("superseded_by")
        if successor and (
            successor not in content or "supersed" not in content.casefold()
        ):
            errors.append("%s does not attribute successor %s" % (
                document_id, successor
            ))
        # The current release's document must carry no marker that
        # distinguishes it from distractor documents: no supersession
        # narrative, no predecessor release ids, no 'authoritative figure'
        # phrase. Any of those would make the answer documents searchable in
        # one query; the chain is conveyed by the records' superseded_by
        # fields and the SUPERSEDED banner on stale documents alone.
        predecessors = by_successor.get(row.get("release"), [])
        if not successor:
            lowered = content.casefold()
            if "authoritative figure" in lowered:
                errors.append(
                    "%s carries the 'authoritative figure' answer marker"
                    % document_id)
            if "supersed" in lowered:
                errors.append(
                    "unsuperseded document %s carries supersession text"
                    % document_id)
            for predecessor in predecessors:
                if predecessor in content:
                    errors.append(
                        "%s names its predecessor release %s"
                        % (document_id, predecessor))
    return errors, check_rendered_profile_excerpts(
        sources, datasets, fetched, targets
    )


async def wait_until_ready(
    url: str,
    client_factory: Callable[[str], Any],
    *,
    timeout: float = 60.0,
    interval: float = 2.0,
) -> dict[str, Any]:
    """Poll the sidecar until it answers a trivial tool call, or time out.

    A container that is still importing (or crash-looping) otherwise surfaces
    as one opaque ``mcp_connection`` failure mid-run (workbook 0256). Probing
    first separates "never became ready" from "broke mid-run": only the former
    justifies the single pre-semantic container restart the skill permits.
    """
    import time

    deadline = time.monotonic() + timeout
    attempts = 0
    last_error = ""
    while True:
        attempts += 1
        try:
            async with client_factory(url) as client:
                await client.call_tool("list_sources", {"limit": 1})
            return {"ready": True, "attempts": attempts, "last_error": ""}
        except Exception as exc:  # noqa: BLE001 - any failure means not ready yet
            last_error = str(exc)
        if time.monotonic() >= deadline:
            return {"ready": False, "attempts": attempts, "last_error": last_error}
        await asyncio.sleep(interval)


async def check_mcp(
    url: str,
    tasks: list[dict[str, Any]],
    targets: list[ValueTarget],
    expected_sources: list[dict[str, Any]],
    expected_datasets: list[dict[str, Any]],
    *,
    page_size: int = 2,
    max_pages: int = 1000,
    client_factory: Callable[[str], Any] | None = None,
    readiness_timeout: float = 60.0,
) -> dict[str, Any]:
    if client_factory is None:
        try:
            from fastmcp import Client
        except ImportError as exc:
            raise OracleInputError(
                "fastmcp is required for live MCP checks; install fastmcp>=2"
            ) from exc
        client_factory = Client

    failures: list[dict[str, Any]] = []
    resolution_passed = 0
    broad_passed = 0
    chains_passed = 0
    pages_read = 0
    profile_report = {"valid": True, "count": 0, "violations": []}
    broad_cache: dict[tuple[str, str, str], tuple[list[dict[str, Any]], int]] = {}

    readiness = await wait_until_ready(
        url, client_factory, timeout=readiness_timeout
    )
    if not readiness["ready"]:
        return {
            "valid": False,
            "url": url,
            "tasks": len(tasks),
            "exact_resolutions": 0,
            "provenance_chains": 0,
            "broad_queries_with_conflicts": 0,
            "pages_read": 0,
            "profile_excerpts": profile_report,
            "readiness": readiness,
            "failures": [{
                "check": "mcp_not_ready",
                "detail": "server did not become ready within %.0fs after %d "
                          "attempt(s); last error: %s. Capture container logs "
                          "and restart once before any semantic check."
                          % (readiness_timeout, readiness["attempts"],
                             readiness["last_error"]),
            }],
        }

    try:
        async with client_factory(url) as client:
            live_sources = _tool_data(
                await client.call_tool("list_sources", {"limit": 100})
            )
            live_datasets = _tool_data(
                await client.call_tool("list_datasets", {"limit": 100})
            )
            if not isinstance(live_sources, list) or not isinstance(live_datasets, list):
                raise RuntimeError("source/dataset discovery returned malformed data")
            expected_source_ids = {row["id"] for row in expected_sources}
            live_source_ids = {row.get("id") for row in live_sources}
            if live_source_ids != expected_source_ids:
                failures.append({
                    "check": "live_bundle_identity",
                    "detail": "live source IDs do not match shipped runtime",
                })
            expected_dataset_ids = {row["id"] for row in expected_datasets}
            live_dataset_ids = {row.get("id") for row in live_datasets}
            if live_dataset_ids != expected_dataset_ids:
                failures.append({
                    "check": "live_bundle_identity",
                    "detail": "live dataset IDs do not match shipped runtime",
                })
            source_profile_report = check_profile_excerpts(live_sources, targets)
            profile_report = merge_profile_reports(source_profile_report)
            for violation in source_profile_report["violations"]:
                failures.append({"check": "profile_excerpt", "detail": violation})

            for task in tasks:
                task_id = str(task.get("task_id") or "<unknown>")
                wanted = task.get("required_dimensions") or {}
                evidence = task.get("evidence") or {}
                missing = [
                    field for field in DIMENSIONS
                    if field not in wanted
                ]
                if missing:
                    failures.append({
                        "check": "task_manifest",
                        "task_id": task_id,
                        "detail": "missing dimensions: %s" % ", ".join(missing),
                    })
                    continue
                try:
                    exact_rows, exact_pages = await query_all(
                        client,
                        {
                            "dataset_id": evidence["dataset_id"],
                            **{field: str(wanted[field]) for field in DIMENSIONS},
                        },
                        page_size=page_size,
                        max_pages=max_pages,
                    )
                    pages_read += exact_pages
                    row_errors, details = validate_task_rows(task, exact_rows)
                    document_errors, document_profiles = await validate_documents(
                        client,
                        exact_rows,
                        str(evidence.get("release") or ""),
                        live_sources,
                        live_datasets,
                        targets,
                    )
                    row_errors.extend(document_errors)
                    profile_report = merge_profile_reports(
                        profile_report, document_profiles
                    )
                    for detail in document_profiles["violations"]:
                        failures.append({
                            "check": "profile_excerpt",
                            "task_id": task_id,
                            "detail": detail,
                        })
                    if row_errors:
                        for detail in row_errors:
                            failures.append({
                                "check": "exact_resolution",
                                "task_id": task_id,
                                "detail": detail,
                            })
                    else:
                        resolution_passed += 1
                        if details["stale"]:
                            chains_passed += 1
                except Exception as exc:
                    failures.append({
                        "check": "exact_resolution",
                        "task_id": task_id,
                        "detail": str(exc),
                    })
                    continue

                try:
                    broad_key = (
                        str(evidence["dataset_id"]),
                        str(wanted["metric"]),
                        str(wanted["entity"]),
                    )
                    if broad_key not in broad_cache:
                        broad_cache[broad_key] = await query_all(
                            client,
                            {
                                "dataset_id": evidence["dataset_id"],
                                "metric": str(wanted["metric"]),
                                "entity": str(wanted["entity"]),
                            },
                            page_size=page_size,
                            max_pages=max_pages,
                        )
                        pages_read += broad_cache[broad_key][1]
                    broad_rows, _broad_pages = broad_cache[broad_key]
                    current_values = {
                        json_key(row.get("value"))
                        for row in broad_rows
                        if not row.get("superseded_by")
                    }
                    if len(current_values) < 2:
                        failures.append({
                            "check": "broad_query_conflict",
                            "task_id": task_id,
                            "detail": (
                                "broad query has fewer than two conflicting "
                                "unsuperseded values"
                            ),
                        })
                    else:
                        broad_passed += 1
                except Exception as exc:
                    failures.append({
                        "check": "broad_query_conflict",
                        "task_id": task_id,
                        "detail": str(exc),
                    })
    except Exception as exc:
        failures.append({"check": "mcp_connection", "detail": str(exc)})

    return {
        "valid": not failures,
        "url": url,
        "tasks": len(tasks),
        "readiness": readiness,
        "exact_resolutions": resolution_passed,
        "provenance_chains": chains_passed,
        "broad_queries_with_conflicts": broad_passed,
        "pages_read": pages_read,
        "profile_excerpts": profile_report,
        "failures": failures,
    }


def validate_manifests(
    tasks: list[dict[str, Any]],
    audit: list[dict[str, Any]],
) -> list[str]:
    errors: list[str] = []
    task_by_id: dict[str, dict[str, Any]] = {}
    for task in tasks:
        task_id = task.get("task_id")
        if not isinstance(task_id, str) or not task_id:
            errors.append("task without task_id")
        elif task_id in task_by_id:
            errors.append("duplicate task_id: %s" % task_id)
        else:
            task_by_id[task_id] = task
    audit_by_id = {entry.get("variable_id"): entry for entry in audit}
    if set(task_by_id) != set(audit_by_id):
        for task_id in sorted(set(task_by_id) - set(audit_by_id)):
            errors.append("task missing from masked_inputs: %s" % task_id)
        for variable_id in sorted(set(audit_by_id) - set(task_by_id)):
            errors.append("masked variable missing from tasks: %s" % variable_id)
    for variable_id in sorted(set(task_by_id) & set(audit_by_id)):
        task = task_by_id[variable_id]
        entry = audit_by_id[variable_id]
        if task.get("evidence") != entry.get("evidence"):
            errors.append("%s has inconsistent evidence manifests" % variable_id)
        if json_key((task.get("answer") or {}).get("value")) != json_key(
            entry.get("mcp_value")
        ):
            errors.append("%s has inconsistent MCP values" % variable_id)
    return errors


async def run_oracle(
    bundle: Path,
    mcp: Path,
    url: str,
    *,
    workbook: Path | None = None,
    allowlist_path: Path | None = None,
    page_size: int = 2,
    max_pages: int = 1000,
    client_factory: Callable[[str], Any] | None = None,
    readiness_timeout: float = 60.0,
) -> dict[str, Any]:
    bundle = bundle.resolve()
    mcp = mcp.resolve()
    if not bundle.is_dir():
        raise OracleInputError("bundle directory does not exist: %s" % bundle)
    required_manifests = {
        "tasks": mcp / "eval" / "tasks.jsonl",
        "mask": mcp / "mask_cells.json",
        "audit": mcp / "masked_inputs.json",
    }
    missing = [str(path) for path in required_manifests.values() if not path.is_file()]
    if missing:
        raise OracleInputError("missing authoring manifests: %s" % ", ".join(missing))

    tasks = read_jsonl(required_manifests["tasks"])
    mask_refs = load_json(required_manifests["mask"])
    audit = load_json(required_manifests["audit"])
    if not isinstance(mask_refs, list) or not all(
        isinstance(ref, str) for ref in mask_refs
    ):
        raise OracleInputError("mask_cells.json must be a JSON list of references")
    if not isinstance(audit, list):
        raise OracleInputError("masked_inputs.json must be a JSON list")
    if not tasks:
        raise OracleInputError("eval/tasks.jsonl contains no tasks")

    targets = value_targets(audit)
    shipped_workbook = resolve_workbook(bundle, workbook)
    isolation = check_environment(bundle, shipped_workbook)
    workbook_report = scan_value_leaks(
        shipped_workbook, mask_refs, audit, allowlist_path
    )
    manifest_errors = validate_manifests(tasks, audit)

    runtime = bundle / "environment" / "mcp-server" / "runtime"
    sources_path = runtime / "sources.json"
    datasets_path = runtime / "datasets.json"
    documents_path = runtime / "documents.jsonl"
    expected_sources = load_json(sources_path) if sources_path.is_file() else []
    expected_datasets = load_json(datasets_path) if datasets_path.is_file() else []
    expected_documents = (
        read_jsonl(documents_path) if documents_path.is_file() else []
    )
    static_profiles = merge_profile_reports(
        check_profile_excerpts(expected_sources, targets),
        check_rendered_profile_excerpts(
            expected_sources, expected_datasets, expected_documents, targets
        ),
    )
    mcp_report = await check_mcp(
        url,
        tasks,
        targets,
        expected_sources,
        expected_datasets,
        page_size=page_size,
        max_pages=max_pages,
        client_factory=client_factory,
        readiness_timeout=readiness_timeout,
    )

    valid = (
        isolation["valid"]
        and workbook_report["valid"]
        and not manifest_errors
        and static_profiles["valid"]
        and mcp_report["valid"]
    )
    return {
        "valid": valid,
        "bundle": str(bundle),
        "mcp_authoring_bundle": str(mcp),
        "workbook": str(shipped_workbook),
        "tasks": len(tasks),
        "manifest_errors": manifest_errors,
        "environment": isolation,
        "workbook_checks": workbook_report,
        "static_profile_excerpts": static_profiles,
        "mcp_checks": mcp_report,
    }


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=__doc__.splitlines()[0],
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--bundle", type=Path, required=True,
                        help="shipped task bundle containing environment/")
    parser.add_argument("--mcp", type=Path, required=True,
                        help="unshipped authoring MCP directory with eval manifests")
    parser.add_argument("--url", required=True,
                        help="live streamable-HTTP MCP URL")
    parser.add_argument("--workbook", type=Path,
                        help="workbook path relative to environment/; inferred by default")
    parser.add_argument("--allowlist", type=Path,
                        help="explicit JSON allowlist for legitimate workbook duplicates")
    parser.add_argument("--report", type=Path,
                        help="also write the complete JSON report to this path")
    parser.add_argument("--page-size", type=int, default=2,
                        help="query_records page size (default: 2)")
    parser.add_argument("--max-pages", type=int, default=1000,
                        help="hard pagination limit per query (default: 1000)")
    return parser


def main(argv: list[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    if args.page_size < 1:
        parser.error("--page-size must be positive")
    if args.max_pages < 1:
        parser.error("--max-pages must be positive")
    try:
        report = asyncio.run(
            run_oracle(
                args.bundle,
                args.mcp,
                args.url,
                workbook=args.workbook,
                allowlist_path=args.allowlist,
                page_size=args.page_size,
                max_pages=args.max_pages,
            )
        )
    except (OracleInputError, OSError, json.JSONDecodeError) as exc:
        report = {"valid": False, "input_error": str(exc)}
        if args.report:
            args.report.parent.mkdir(parents=True, exist_ok=True)
            args.report.write_text(
                json.dumps(report, indent=2, ensure_ascii=False) + "\n",
                encoding="utf-8",
            )
        print(json.dumps(report, indent=2))
        print("ORACLE: ERROR")
        return 2
    if args.report:
        args.report.parent.mkdir(parents=True, exist_ok=True)
        args.report.write_text(
            json.dumps(report, indent=2, ensure_ascii=False) + "\n",
            encoding="utf-8",
        )
    print(json.dumps(report, indent=2, ensure_ascii=False))
    print("ORACLE:", "PASS" if report["valid"] else "FAIL")
    return 0 if report["valid"] else 1


if __name__ == "__main__":
    sys.exit(main())
