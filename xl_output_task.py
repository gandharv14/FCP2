#!/usr/bin/env python3
"""Build "rebuild the model" tasks from the segmentation.

Unlike ``xl_task_build.py``, which blanks cells by dependency level, this takes
the inputs-only workbook from ``xl_input_mask.py`` and asks for the outputs the
segmentation curated in ``seg_out/<id>/curation.toml`` - the headline figures
the model exists to produce. Stage 9 already proved the input frontier is
sufficient to recompute them, so the ask is answerable from the artifact alone.

    python3 xl_output_task.py 0248 0262 0449 0450 -o tasks_outputs

The framing paragraph goes through the same naturalizer as the level-split
tasks; the target list is appended verbatim so the deliverable stays exact.
Before packaging, a separate GPT 5.6 Sol stage writes a variable/source audit
and deterministic inventory under ``runs/<id>-variable-sources/``.
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import re
import shutil
import sys
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter, range_boundaries

try:
    import tomli
except ImportError:  # pragma: no cover
    import tomllib as tomli

from xl_task_build import (Instance, PROD_ENDPOINT, naturalize, read_env_key,
                           toml_table)
from xl_harbor_prep import DOCKERFILE
from mcp_env.server_assets import COMPOSE_YAML
from xl_variable_source_audit import (
    DEFAULT_MODEL as DEFAULT_AUDIT_MODEL,
    generate_audit,
)
from xl_formula_hint_tasks import (
    load_formula_artifacts,
    render_section as render_custom_formula_section,
)

PIPELINE_VERSION = "1.0.0"
TIMEOUT_BASE_SEC = 2400.0
TIMEOUT_PER_TARGET_SEC = 20.0
TIMEOUT_MAX_SEC = 10800.0

TEST_SH = """\
#!/bin/bash
set -uo pipefail
mkdir -p /logs/verifier
exec python3 /tests/run_grader.py \\
    --workspace /app \\
    --answer-key /tests/answer_key.json \\
    --output-dir /logs/verifier \\
    --mode continuous
"""

GRADER_ROOT = Path(__file__).resolve().parent / "grader"

FAMILY_PHRASE = {
    "three_statement": "an integrated three-statement and valuation model",
    "scenario_driven": "a scenario-driven investment model",
    "dcf_valuation": "a discounted cash flow valuation model",
    "lbo_returns": "a leveraged buyout returns model",
    "real_estate_pf": "a real estate project finance model",
}

SEMANTIC_HINTS = {
    "three_statement": [
        "Treat the projected statements as one connected model. Before relying "
        "on a valuation result, make sure profit, cash movement and the balance "
        "sheet tell a consistent story across the forecast.",
        "The valuation sheet presents several methods side by side. Keep the "
        "dividend, equity-cash-flow and enterprise-value approaches distinct "
        "rather than forcing them to share the same cash-flow definition.",
        "Pay close attention to whether a line is an enterprise value or an "
        "equity value, and to where debt and cash belong in that bridge.",
        "Terminal and exit figures should use a mature forecast period. Check "
        "that the period, units and sign convention agree with the surrounding "
        "projection before carrying the figure into a return calculation.",
        "For IRR, think in terms of the investor's signed cash-flow sequence: "
        "investment first, distributions during the hold, and exit proceeds at "
        "the end. Accounting profit is not itself that sequence.",
        "Use the dashboard outputs as reasonableness checks. If a headline "
        "valuation conflicts with the operating and working-capital trends, "
        "revisit the upstream interpretation rather than adjusting the output.",
    ],
    "scenario_driven": [
        "Resolve which scenario is active before calculating anything else, "
        "then use that same case consistently in the operating, financing and "
        "investment sections.",
        "Separate the acquisition-date funding decision from the later exit "
        "calculation. Sources and uses should reconcile before they feed any "
        "multiple or return measure.",
        "Follow cash generation through working capital, investment and debt "
        "service. A profitable forecast can still require fresh equity if its "
        "cash timing or financing obligations demand it.",
        "Keep acquisition and exit assumptions separate, especially ownership "
        "percentage and entry/exit multiples. Mixing their periods can make a "
        "plausible-looking equity value materially wrong.",
        "For equity returns, preserve the investor perspective and sign "
        "convention across acquisition funding, interim dividends or injections, "
        "and exit proceeds.",
        "Use the dashboard ratios as cross-checks against the underlying sources, "
        "uses and operating measures; do not infer one ratio from another when "
        "their numerators represent different funding components.",
    ],
    "dcf_valuation": [
        "Keep the trading-comps summaries, cost-of-capital build and DCF as "
        "separate stages. A statistic from the comps sheet is an input or check, "
        "not automatically the DCF conclusion.",
        "The workbook shows both exit-multiple and perpetuity-growth terminal "
        "methods. Do not blend the assumptions or terminal values between them.",
        "Match each forecast cash flow and terminal value to its intended "
        "discount period. Timing inconsistency is often larger than any rounding "
        "difference.",
        "Before interpreting present enterprise value, check that the cash-flow "
        "measure is the one appropriate to enterprise value and that the "
        "discount rate is on the same basis.",
        "Treat CAGR, mean and median outputs as summaries of their labelled "
        "source series. Confirm the historical/forecast endpoints and units "
        "rather than extrapolating from a nearby displayed statistic.",
        "Use the relationship between discounted forecast cash flows, discounted "
        "terminal value and total present enterprise value as a reconciliation "
        "check, without assuming either terminal method must produce the same "
        "answer.",
    ],
    "lbo_returns": [
        "Settle the active operating case first. The transaction and debt model "
        "should use one internally consistent EBITDA and cash-flow forecast.",
        "Make sources and uses reconcile before building returns. Entry equity "
        "is the residual funding requirement after debt and other sources, not "
        "a free-standing assumption.",
        "Work through each debt instrument separately, respecting cash interest, "
        "PIK interest, amortization, fees and minimum cash. Circular interest "
        "effects should be resolved consistently rather than ignored.",
        "At exit, distinguish enterprise value from equity proceeds: operating "
        "performance and the exit multiple establish enterprise value, while "
        "the debt and cash position determine what remains for equity holders.",
        "Apply management and co-investor economics only after establishing the "
        "total equity proceeds. Check that allocations reconcile back to the "
        "same total.",
        "IRR is sensitive to dates and signed cash-flow timing; MOIC is not. If "
        "the two imply incompatible economics, recheck the cash-flow sequence "
        "instead of tuning either output.",
        "Use leverage, interest-coverage and fixed-charge metrics as diagnostics "
        "for the debt schedule. A returns result built on an inconsistent credit "
        "profile is unlikely to be reliable.",
    ],
}


def quote_sheet(sheet):
    return "'%s'" % sheet if any(c in sheet for c in " -&()") else sheet


def display_name(label, sheet):
    """Turn a segmentation label into something an analyst would write.

    Labels arrive as ``"Terminal value / AED / Projection"`` - the line item
    followed by whatever context the band sat in, sometimes including the sheet
    name itself. Keep the line item, fold the rest into a parenthetical, and
    drop the sheet because the table already has a column for it.
    """
    noise = {"-", "na", "n/a"}
    parts = [p.strip() for p in str(label).split(" / ") if p.strip()]
    parts = [p for p in parts
             if p.lower() != sheet.lower() and p.lower() not in noise]
    if not parts:
        return str(label)
    head, context = parts[0], parts[1:]
    if not context:
        return head
    if len(context) == 1 and context[0].startswith("("):
        return "%s %s" % (head, context[0])
    return "%s (%s)" % (head, ", ".join(context))


def band_cells(band):
    """['Sheet!C3', ...] for a band that may be a single cell or a run."""
    sheet, _, ref = band.rpartition("!")
    min_col, min_row, max_col, max_row = range_boundaries(ref)
    return [(sheet, "%s%d" % (get_column_letter(col), row))
            for row in range(min_row, max_row + 1)
            for col in range(min_col, max_col + 1)]


def curated_outputs(seg_dir):
    data = tomli.loads((seg_dir / "curation.toml").read_text(encoding="utf-8"))
    return [entry for entry in data.get("output", []) if entry.get("include")]


def load_lineage(seg_dir):
    data = json.loads(
        (seg_dir / "lineage.json").read_text(encoding="utf-8")
    )
    return {trace["output"]: trace for trace in data.get("traces", [])}


def formula_functions(formula):
    """Excel functions used by a formula, without constants or values."""
    return list(dict.fromkeys(re.findall(
        r"\b([A-Z][A-Z0-9.]*)\s*\(", str(formula).upper()
    )))


def longest_band_path(trace):
    """One deterministic maximum-depth route from an input to the output."""
    steps = trace.get("band_steps") or []
    by_node = {step["node"]: step for step in steps}
    current = by_node.get(trace["output"])
    if current is None:
        return []
    path = [current]
    seen = {current["node"]}
    while current.get("inputs"):
        predecessors = [
            by_node[node] for node in current["inputs"]
            if node in by_node and node not in seen
        ]
        if not predecessors:
            break
        current = max(
            predecessors,
            key=lambda step: (step.get("depth", 0), step["node"]),
        )
        seen.add(current["node"])
        path.append(current)
    return list(reversed(path))


def lineage_hints(outputs, lineage):
    """Compact, answer-free hints grounded in each output's lineage trace."""
    hints = []
    for output in outputs:
        trace = lineage.get(output["band"])
        if not trace:
            raise SystemExit(
                "no lineage trace for curated output %s" % output["band"]
            )
        path = longest_band_path(trace)
        # The end of the longest path is the useful part: it tells the solver
        # which calculated rows must exist immediately before the output.
        route_steps = (
            [path[0]] + path[-4:] if len(path) > 5 else path
        )
        route = " → ".join(
            "`%s` (%s)" % (
                step["node"],
                display_name(step.get("label") or step["node"], step["sheet"]),
            )
            for step in route_steps
        )
        target = path[-1] if path else {}
        direct_nodes = target.get("inputs") or []
        direct = ["`%s`" % node for node in direct_nodes[:5]]
        if len(direct_nodes) > 5:
            direct.append(
                "%d other direct bands" % (len(direct_nodes) - 5)
            )
        functions = formula_functions(target.get("formula", ""))
        suffix = []
        if direct:
            suffix.append("direct predecessors: %s" % ", ".join(direct))
        if functions:
            suffix.append("terminal operation: `%s`" % "/".join(functions))
        hints.append({
            "name": output["name"],
            "band": output["band"],
            "route": route,
            "direct_inputs": direct_nodes,
            "functions": functions,
            "text": (
                "**%s**: %s%s"
                % (
                    output["name"],
                    route,
                    "; " + "; ".join(suffix) if suffix else "",
                )
            ),
        })
    return hints


def hint_section(hints):
    return "\n".join("- " + hint["text"] for hint in hints)


def semantic_hints(family):
    texts = SEMANTIC_HINTS.get(family)
    if not texts:
        raise SystemExit("no semantic hint template for family %r" % family)
    return [
        {"name": "Analyst hint %d" % index, "text": text}
        for index, text in enumerate(texts, 1)
    ]


def verify_mcp_mask(mcp_dir, artifact):
    """Every cell the MCP bundle serves must be blank in the shipped workbook.

    Packaging is the last line of defense: if the bundle was rebuilt (new
    mask_cells.json) without re-running xl_input_mask.py, "research-only"
    variables would still be sitting in the sheets, silently defeating the
    research component.
    """
    from xl_input_mask import load_mask_cells
    mask_path = mcp_dir / "mask_cells.json"
    if not mask_path.is_file():
        raise SystemExit("--mcp bundle %s has no mask_cells.json" % mcp_dir)
    book = openpyxl.load_workbook(artifact, data_only=True)
    populated = []
    for sheet, spots in load_mask_cells(mask_path).items():
        if sheet not in book.sheetnames:
            raise SystemExit(
                "%s names unknown sheet %r" % (mask_path, sheet))
        grid = book[sheet]
        for row, col in sorted(spots):
            if grid.cell(row=row, column=col).value is not None:
                populated.append(
                    "%s!%s%d" % (sheet, get_column_letter(col), row))
    if populated:
        raise SystemExit(
            "%s was not masked with %s: %d MCP-served cell(s) are still "
            "populated, e.g. %s -- re-run xl_input_mask.py --mask-cells "
            "against this bundle"
            % (artifact.name, mask_path, len(populated),
               ", ".join(populated[:8])))


def collect(workbook, source_dir, seg_root, inputs_root):
    """(outputs, targets) - curated figures and their golden cell values."""
    seg_dir = Path(seg_root) / workbook
    outputs = curated_outputs(seg_dir)
    if not outputs:
        raise SystemExit("%s: curation.toml includes no outputs" % workbook)

    gold = openpyxl.load_workbook(Path(source_dir) / ("%s.xlsx" % workbook),
                                  data_only=True)
    masked = openpyxl.load_workbook(
        Path(inputs_root) / ("%s-inputs.xlsx" % workbook), data_only=True)

    resolved, targets = [], {}
    for entry in outputs:
        refs = []
        for sheet, coord in band_cells(entry["band"]):
            if sheet not in gold.sheetnames:
                raise SystemExit("%s: unknown sheet %r" % (workbook, sheet))
            value = gold[sheet][coord].value
            if not isinstance(value, (int, float)) or isinstance(value, bool):
                continue
            # the artifact must actually be missing the figure we are asking for
            if masked[sheet][coord].value is not None:
                raise SystemExit("%s: %s!%s is not blank in the inputs workbook"
                                 % (workbook, sheet, coord))
            ref = "%s!%s" % (quote_sheet(sheet), coord)
            targets[ref] = value
            refs.append(ref)
        if refs:
            raw = entry.get("name") or entry["label"]
            resolved.append({"name": display_name(raw, entry["sheet"]),
                             "raw_label": raw,
                             "sheet": entry["sheet"], "band": entry["band"],
                             "score": entry.get("score"), "refs": refs})
    return resolved, targets


def scenario_for(workbook, family, outputs, mcp=False):
    phrase = FAMILY_PHRASE.get(family, "a financial model")
    headline = ", ".join(o["name"] for o in outputs[:3])
    if mcp:
        return (
            "A colleague has sent you %s with every calculated figure stripped "
            "out, and this time the externally-sourced input assumptions have "
            "been removed from the sheets as well: market rates, tax rates, "
            "macro assumptions, contractual terms and opening balances have to "
            "be researched through the data service described below before "
            "anything can be computed. Rebuild the calculations from the "
            "remaining inputs plus your research, and report the %d headline "
            "figures the model exists to produce, such as %s. The full list of "
            "figures and the cells they belong in is given below."
            % (phrase, len(outputs), headline))
    return (
        "A colleague has sent you %s with every input assumption in place and "
        "every calculated figure stripped out. Nothing has been computed yet: "
        "the drivers, historicals and assumptions are all there, but the cells "
        "the model works out for itself are empty. Rebuild the calculations "
        "from those inputs and report the %d headline figures the model exists "
        "to produce, such as %s. The full list of figures and the cells they "
        "belong in is given below." % (phrase, len(outputs), headline))


def target_table(outputs):
    lines = ["| # | figure | cells |", "| --- | --- | --- |"]
    for i, out in enumerate(outputs, 1):
        refs = ", ".join("`%s`" % r for r in out["refs"])
        lines.append("| %d | %s | %s |" % (i, out["name"], refs))
    return "\n".join(lines)


def answer_example(targets):
    """A shape-only example built from this task's own refs."""
    refs = list(targets)[:2]
    body = ", ".join('"%s": <number>' % r for r in refs)
    return "{%s}" % body


MCP_URL = "http://mcp-server:8000/mcp"

INPUT_SECTION_PLAIN = """\
The workbook `%s` is in your working directory. Every input the model needs is
present - assumptions, drivers, historicals and labels - and every cell the
model is meant to work out is blank. The artifact has been checked to contain
no formulas or derived numbers and to preserve the identified input cells.
You may install Python packages (for example `openpyxl`) to read it.
"""

INPUT_SECTION_MCP = """\
The workbook `%s` is in your working directory. Every cell the model is meant
to work out is blank, and a set of externally-sourced input assumptions has
additionally been removed from the sheets. Those removed inputs are only
retrievable through the research data service described in the next section;
every other input - drivers, historicals and labels - is present. The artifact
has been checked to contain no formulas or derived numbers. You may install
Python packages (for example `openpyxl`) to read it.
"""

RESEARCH_SECTION = """\
## Research data service

The removed input assumptions (rates, tax and macro assumptions, contractual
terms, opening balances, dates and similar externally-sourced values) are
served by a mock research MCP server at `%s` (streamable HTTP transport). It
exposes five tools:

- `list_sources` - the organizations and data platforms available
- `list_datasets` - structured datasets and their filter dimensions
- `search_documents` / `fetch_document` - keyword search over source
  documents, then fetch one by id
- `query_records` - filter dataset records by entity, metric, period,
  scenario, basis, unit and status

Records vary across those dimensions and broad queries return conflicting
candidate values from adjacent periods, other scenario cases and earlier
releases. `query_records` requires at least two filter dimensions and returns
at most 5 rows per page; follow the returned `next_cursor` to page through
larger result sets. Records exist in multiple releases: a row whose
`superseded_by` field names a later release has been replaced and must not be
used - only the unsuperseded release of a record is authoritative. Filter
every dimension that matters to your question. Convert each retrieved value
into the units and scale of the sheet it belongs on (rates are reported in
percent while sheets may store decimal fractions; monetary amounts state
their unit on the record).

Any MCP client works, for example from Python:

```
pip install fastmcp
```

```python
import asyncio
from fastmcp import Client

async def main():
    async with Client("%s") as client:
        print(await client.call_tool("list_sources", {}))

asyncio.run(main())
```
"""


def build_instruction(scenario, artifact, outputs, targets, hints=None,
                      hint_style="", mcp=False):
    n_cells = len(targets)
    hints_text = ""
    if hints:
        if hint_style == "lineage":
            heading = "Hints from the dependency paths"
            introduction = """\
These are route markers, not answers. Each bullet follows a longest path from
the supplied inputs through the model's intermediate calculations to that
output. Build the referenced rows in order; the direct predecessors are the
last cells or bands to reconcile before calculating the requested output."""
        else:
            heading = "Analyst hints"
            introduction = """\
Use these as modelling checks, not as a recipe. They flag the conceptual
distinctions that matter in this workbook but deliberately provide no formulas,
dependency routes, intermediate values or answers."""
        hints_text = """\

## %s

%s

%s
""" % (heading, introduction, hint_section(hints))
    input_section = (INPUT_SECTION_MCP if mcp else INPUT_SECTION_PLAIN) % artifact
    research = ("\n" + RESEARCH_SECTION % (MCP_URL, MCP_URL)) if mcp else ""
    return """\
%s

## Input

%s%s
## What to compute

%s

%s

## Output

Write a JSON object to `/app/answers.json` mapping each cell reference above to
the value the model computes for it, keyed exactly as written in the table:

```
%s
```

Report %d values in total, one per cell listed. Give each number exactly as the
model computes it, in the same units and scale as the sheet it sits on - do not
round, rescale or convert percentages.
""" % (scenario.strip(), input_section, research, target_table(outputs),
       hints_text, answer_example(targets), n_cells)


def add_custom_formula_hints(instruction, spec):
    section = render_custom_formula_section(spec)
    if not section:
        return instruction
    marker = "\n## Output\n"
    if marker not in instruction:
        raise ValueError("generated instruction has no Output heading")
    return instruction.replace(marker, "\n" + section + "## Output\n", 1)


TIMEOUT_PER_MCP_VARIABLE_SEC = 15.0


def agent_timeout(n_cells, n_mcp_variables=0):
    return min(TIMEOUT_MAX_SEC,
               TIMEOUT_BASE_SEC + TIMEOUT_PER_TARGET_SEC * n_cells
               + TIMEOUT_PER_MCP_VARIABLE_SEC * n_mcp_variables)


def mcp_variable_count(mcp_dir):
    tasks = (mcp_dir / "eval" / "tasks.jsonl").read_text(encoding="utf-8")
    return sum(1 for line in tasks.splitlines() if line.strip())


def load_plain_meta(workbook, audit_root="runs"):
    """In-bundle marker fields for a no-MCP task, plus the exclusions path."""
    run = Path(audit_root) / ("%s-variable-sources" % workbook)
    report_path = run / "normalization_report.json"
    if not report_path.is_file():
        return None, None
    report = json.loads(report_path.read_text(encoding="utf-8"))
    elig_path = run / "plain_eligibility.json"
    elig = {}
    if elig_path.is_file():
        elig = json.loads(elig_path.read_text(encoding="utf-8"))
    codes = report.get("exclusion_reason_codes") or {}
    reason = elig.get("plain_reason") or elig.get("reason") or (
        "all %s audit rows excluded" % report.get("excluded_rows", 0)
    )
    meta = {
        "research_service": False,
        "n_mcp_variables": 0,
        "draft_rows": report.get("draft_rows"),
        "excluded_rows": report.get("excluded_rows"),
        "exclusion_reason_codes": codes,
        "plain_reason": reason,
    }
    exclusions = run / "exclusions.json"
    return meta, exclusions if exclusions.is_file() else None


def emit(out_dir, workbook, family, artifact, instruction, targets, outputs,
         nat_meta, hints=None, hint_style="", mcp_dir=None, audit_meta=None,
         formula_report=None, formula_hints=None, audit_root="runs"):
    if out_dir.exists():
        shutil.rmtree(out_dir)
    (out_dir / "environment").mkdir(parents=True)
    (out_dir / "tests").mkdir()

    shutil.copy2(artifact, out_dir / "environment" / artifact.name)
    (out_dir / "environment" / "Dockerfile").write_text(
        DOCKERFILE % (artifact.name, artifact.name), encoding="utf-8")

    n_mcp_variables = 0
    if mcp_dir is not None:
        # Sidecar image: server + runtime data only. eval/ and the normalized
        # spec must never enter the task environment.
        sidecar = out_dir / "environment" / "mcp-server"
        sidecar.mkdir()
        shutil.copy2(mcp_dir / "server.py", sidecar / "server.py")
        shutil.copy2(mcp_dir / "Dockerfile", sidecar / "Dockerfile")
        shutil.copytree(mcp_dir / "runtime", sidecar / "runtime")
        (out_dir / "environment" / "docker-compose.yaml").write_text(
            COMPOSE_YAML, encoding="utf-8")
        n_mcp_variables = mcp_variable_count(mcp_dir)

    # Harbor's loader requires instruction.md
    (out_dir / "instruction.md").write_text(instruction, encoding="utf-8")

    hinted = bool(hints)
    if hint_style == "semantic":
        template = "outputs_semantic_hints"
    elif hinted:
        template = "outputs_hinted"
    else:
        template = "outputs"
    metadata = {
        "workbook": workbook,
        "source_file": "%s.xlsx" % workbook,
        "artifact": artifact.name,
        "template": template,
        "financebench_question_type": "metrics-generated",
        "financebench_reasoning": "numerical reasoning",
        "answer_kind": "cell_value",
        "reward_type": "continuous_scoring_function",
        "n_outputs": len(outputs),
        "n_answer_cells": len(targets),
        "answer_cells": list(targets),
        "output_names": [o["name"] for o in outputs],
        "taxonomy_primary": family,
        "curation_source": "seg_out/%s/curation.toml" % workbook,
        "hint_source": (
            "finance-domain template grounded in workbook family"
            if hint_style == "semantic"
            else "seg_out/%s/lineage.json" % workbook if hinted else ""
        ),
        "hint_policy": (
            "conceptual checks; no formulas, routes, or values"
            if hint_style == "semantic"
            else "longest dependency route + direct predecessors"
            if hinted else ""
        ),
        "pipeline_version": PIPELINE_VERSION,
        "created_at": dt.datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
    }
    if audit_meta is not None:
        metadata.update({
            "variable_source_audit": audit_meta["markdown"],
            "variable_source_audit_model": audit_meta["model"],
            "variable_source_audit_inventory_sha256":
                audit_meta["inventory_sha256"],
            "variable_source_audit_rows": audit_meta["inventory_rows"],
        })
    if mcp_dir is not None:
        metadata.update({
            "mcp_environment": str(mcp_dir),
            "mcp_url": MCP_URL,
            "n_mcp_variables": n_mcp_variables,
            "mcp_masked_cells": json.loads(
                (mcp_dir / "mask_cells.json").read_text(encoding="utf-8")
            ).__len__(),
        })
    else:
        plain_meta, exclusions_path = load_plain_meta(workbook, audit_root)
        if plain_meta:
            metadata.update(plain_meta)
        if exclusions_path is not None:
            shutil.copy2(exclusions_path,
                         out_dir / "tests" / "normalization_exclusions.json")
    if formula_report is not None:
        metadata.update({
            "custom_formula_gate_model":
                formula_report["generator"]["model"],
            "custom_formula_gate_verdict": formula_report["verdict"],
            "custom_formula_hint_groups":
                len((formula_hints or {}).get("hints") or []),
        })

    sections = [
        'schema_version = "1.4"',
        "",
        toml_table("task", {
            "name": "fcp/%s-%s" % (workbook, template.replace("_", "-")),
            "version": "1.0.0",
            "description": "Rebuild %s from its inputs%s and report the %d "
                           "curated outputs"
                           % (
                               family,
                               (
                                   " using analyst hints"
                                   if hint_style == "semantic"
                                   else " using lineage hints" if hinted else ""
                               ),
                               len(outputs),
                           ),
            "keywords": [
                family, template, "spreadsheet", "lineage_hints"
            ] if hinted else [family, template, "spreadsheet"],
        }),
        "",
        toml_table("metadata", metadata),
        "",
        toml_table("metadata.naturalizer", nat_meta),
        "",
        toml_table("agent", {"timeout_sec": agent_timeout(len(targets),
                                                          n_mcp_variables)}),
        "",
        toml_table("verifier", {"timeout_sec": 300.0}),
        "",
        toml_table("environment", {"cpus": 2, "memory_mb": 4096}),
        "",
    ]
    if mcp_dir is not None:
        sections += [
            "[[environment.mcp_servers]]",
            'name = "research"',
            'transport = "streamable-http"',
            'url = "%s"' % MCP_URL,
            "",
        ]
    (out_dir / "task.toml").write_text("\n".join(sections), encoding="utf-8")

    test_path = out_dir / "tests" / "test.sh"
    test_path.write_text(TEST_SH, encoding="utf-8")
    test_path.chmod(0o755)
    runner_path = out_dir / "tests" / "run_grader.py"
    shutil.copy2(GRADER_ROOT / "run_grader.py", runner_path)
    runner_path.chmod(0o755)
    shutil.copytree(
        GRADER_ROOT / "finance_grader",
        out_dir / "tests" / "finance_grader",
        ignore=shutil.ignore_patterns("__pycache__", "*.pyc"),
    )
    # Group each output's cells under its band so the grader can weight per
    # curated figure rather than per cell; a 7-period fill row counts once.
    groups = {o["band"]: o["refs"] for o in outputs}
    with open(out_dir / "tests" / "answer_key.json", "w", encoding="utf-8") as fh:
        json.dump({"kind": "cell_value",
                   "tolerance": {"numeric_abs": 1e-6, "numeric_rel": 1e-6},
                   "targets": targets,
                   "groups": groups}, fh, indent=1)
    with open(out_dir / "tests" / "outputs.json", "w", encoding="utf-8") as fh:
        json.dump(outputs, fh, indent=1)
    if hints:
        with open(out_dir / "tests" / "hints.json", "w", encoding="utf-8") as fh:
            json.dump(hints, fh, indent=1)
    if formula_hints is not None:
        with open(out_dir / "tests" / "formula_hints.json",
                  "w", encoding="utf-8") as fh:
            json.dump(formula_hints, fh, indent=2)
    if mcp_dir is not None:
        # Audit map of masked inputs and their MCP evidence. tests/ is mounted
        # only at verification time, so like answer_key.json it never reaches
        # the agent.
        shutil.copy2(mcp_dir / "masked_inputs.json",
                     out_dir / "tests" / "masked_inputs.json")


def main(argv=None):
    parser = argparse.ArgumentParser(
        description="Build rebuild-the-model tasks from curated outputs")
    parser.add_argument("workbooks", nargs="+")
    parser.add_argument("--source", default="4-10 100")
    parser.add_argument("--seg-root", default="seg_out")
    parser.add_argument("--inputs-root", default="inputs_out")
    parser.add_argument("--taxonomy", default="taxonomy_out/workbooks.json")
    parser.add_argument("--env-file", default=".env")
    parser.add_argument("--project-id", default="cms6m4urm006n07z8ecxi1oi2")
    parser.add_argument("--model", default="openai/gpt-5.6-luna")
    parser.add_argument("--no-naturalize", action="store_true")
    parser.add_argument(
        "--variable-source-audit-root", default="runs",
        help="root for <workbook>-variable-sources audit artifacts",
    )
    parser.add_argument(
        "--variable-source-audit-inputs-root", default="inputs_out",
        help="unredacted inputs-only workbooks audited before optional MCP masking",
    )
    parser.add_argument(
        "--variable-source-audit-model", default=DEFAULT_AUDIT_MODEL,
        help="LiteLLM model used only for variable/source Markdown",
    )
    parser.add_argument(
        "--no-variable-source-audit", action="store_true",
        help="skip the GPT-generated variable/source audit stage",
    )
    parser.add_argument(
        "--refresh-variable-source-audit", action="store_true",
        help="regenerate audit Markdown even when its inventory hash matches",
    )
    parser.add_argument(
        "--hints", action="store_true",
        help="append answer-free routes derived from lineage.json",
    )
    parser.add_argument(
        "--semantic-hints", action="store_true",
        help="append finance-domain checks without formulas or routes",
    )
    parser.add_argument(
        "--mcp", default="",
        help="MCP bundle from xl_variable_mcp.py; adds the research sidecar, "
             "docker-compose.yaml and the task.toml mcp_servers entry. The "
             "inputs workbook must have been masked with the bundle's "
             "mask_cells.json.",
    )
    parser.add_argument(
        "--custom-formula-context",
        default="",
        help="extracted formula context paired with the Terra report and hints",
    )
    parser.add_argument(
        "--custom-formula-report",
        default="",
        help="validated GPT-5.6 Terra custom-formula report for this workbook",
    )
    parser.add_argument(
        "--custom-formula-hints",
        default="",
        help="validated method-only hint spec paired with --custom-formula-report",
    )
    parser.add_argument("-o", "--out", default="tasks_outputs")
    args = parser.parse_args(argv)
    if args.hints and args.semantic_hints:
        parser.error("--hints and --semantic-hints are mutually exclusive")
    formula_paths = (
        args.custom_formula_context,
        args.custom_formula_report,
        args.custom_formula_hints,
    )
    if any(formula_paths) and not all(formula_paths):
        parser.error(
            "--custom-formula-context, --custom-formula-report, and "
            "--custom-formula-hints are required together"
        )
    if args.custom_formula_report and len(args.workbooks) != 1:
        parser.error("custom-formula artifacts require exactly one workbook")

    taxonomy = json.loads(Path(args.taxonomy).read_text(encoding="utf-8"))
    families = {Path(k).stem: v.get("primary", "") for k, v in taxonomy.items()}
    config = {
        "enabled": not args.no_naturalize,
        "base_url": PROD_ENDPOINT,
        "api_key": read_env_key(args.env_file),
        "model": args.model,
        "project_id": args.project_id,
    }

    mcp_dir = Path(args.mcp) if args.mcp else None
    if mcp_dir is not None and not (mcp_dir / "runtime").is_dir():
        parser.error("--mcp %s has no runtime/ directory" % mcp_dir)
    if mcp_dir is not None and len(args.workbooks) != 1:
        parser.error("an --mcp bundle is specific to one workbook; pass "
                     "exactly one workbook id per invocation")

    out_root = Path(args.out)
    out_root.mkdir(parents=True, exist_ok=True)
    for workbook in args.workbooks:
        family = families.get(workbook, "")
        outputs, targets = collect(workbook, args.source, args.seg_root,
                                   args.inputs_root)
        formula_report = None
        formula_hints = None
        if args.custom_formula_report:
            formula_report, formula_hints = load_formula_artifacts(
                "%s-outputs" % workbook,
                Path(args.custom_formula_report),
                Path(args.custom_formula_hints),
                [
                    float(value)
                    for value in targets.values()
                    if isinstance(value, (int, float))
                    and not isinstance(value, bool)
                ],
                context_path=Path(args.custom_formula_context),
            )
        artifact = Path(args.inputs_root) / ("%s-inputs.xlsx" % workbook)
        if mcp_dir is not None:
            verify_mcp_mask(mcp_dir, artifact)
        audit_meta = None
        if not args.no_variable_source_audit:
            audit_artifact = (
                Path(args.variable_source_audit_inputs_root)
                / ("%s-inputs.xlsx" % workbook)
            )
            audit_meta = generate_audit(
                workbook,
                audit_artifact,
                Path(args.seg_root) / workbook,
                Path(args.variable_source_audit_root)
                / ("%s-variable-sources" % workbook),
                config["api_key"],
                model=args.variable_source_audit_model,
                endpoint=PROD_ENDPOINT,
                project_id=args.project_id,
                refresh=args.refresh_variable_source_audit,
                log=lambda message: print("   ", message),
            )
        hints = None
        hint_style = ""
        if args.hints:
            seg_dir = Path(args.seg_root) / workbook
            hints = lineage_hints(outputs, load_lineage(seg_dir))
            hint_style = "lineage"
        elif args.semantic_hints:
            hints = semantic_hints(family)
            hint_style = "semantic"
        scenario = scenario_for(workbook, family, outputs,
                                mcp=mcp_dir is not None)
        instance = Instance(
            template_id="outputs", financebench="metrics-generated",
            slots={}, scenario=scenario, output_format="", answer_key={},
            answer_cells=[], facts_required=[], facts_refs=[],
            forbidden=[repr(v) for v in targets.values()])
        text, nat_meta = naturalize(instance, config,
                                    lambda m: print("   ", m))
        instruction = build_instruction(
            text, artifact.name, outputs, targets, hints=hints,
            hint_style=hint_style, mcp=mcp_dir is not None,
        )
        if formula_hints is not None:
            instruction = add_custom_formula_hints(
                instruction, formula_hints
            )
        if args.semantic_hints:
            suffix = "outputs_semantic_hints"
        elif args.hints:
            suffix = "outputs_hinted"
        else:
            suffix = "outputs"
        out_dir = out_root / ("%s-%s" % (workbook, suffix))
        emit(out_dir, workbook, family, artifact, instruction, targets,
             outputs, nat_meta, hints=hints, hint_style=hint_style,
             mcp_dir=mcp_dir, audit_meta=audit_meta,
             formula_report=formula_report, formula_hints=formula_hints,
             audit_root=args.variable_source_audit_root)
        n_vars = mcp_variable_count(mcp_dir) if mcp_dir is not None else 0
        print("%s  %-16s %2d outputs, %3d cells%s, timeout %.0fs -> %s"
              % (workbook, family, len(outputs), len(targets),
                 ", %d mcp variables" % n_vars if n_vars else "",
                 agent_timeout(len(targets), n_vars), out_dir))
    print("\n%d task(s) -> %s" % (len(args.workbooks), out_root.resolve()))


if __name__ == "__main__":
    main()
