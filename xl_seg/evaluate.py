"""Stage 8: recompute the workbook from the input frontier alone.

This is what makes the segmentation falsifiable. Input cells are seeded with their
cached values and nothing else is; every other cell is rebuilt from its parsed AST.
If the recomputed outputs match the workbook, the input set is provably sufficient.

Only 24 distinct functions and 14 operators appear across these four workbooks, so
the surface is small. Anything outside it yields ``Unresolved``, which propagates
and is reported rather than silently coerced to zero.
"""

from __future__ import annotations

import math
import os
import re
import stat
import zipfile
from calendar import monthrange
from collections import defaultdict
from dataclasses import dataclass, field as dataclass_field
from datetime import datetime, timedelta
from functools import lru_cache
from pathlib import Path
from xml.etree import ElementTree

from .condense import strongly_connected, topo_order
from .model import Graph, a1, range_members, split_ref
from .project import CellGraph


def _split(cid: str):
    parsed = split_ref(cid)
    return (parsed[0], parsed[1], parsed[2]) if parsed else (cid, 0, 0)

EPOCH = datetime(1899, 12, 30)
EXCEL_EPOCH = datetime(1899, 12, 31)
ISO_RE = re.compile(r"^\d{4}-\d{2}-\d{2}[T ]")
CRITERIA_RE = re.compile(r"^\s*(>=|<=|<>|>|<|=)?\s*(.*)$")
A1_RANGE_RE = re.compile(
    r"(?:(?:'(?P<quoted>(?:[^']|'')+)'|(?P<bare>[A-Za-z0-9_ .&-]+))!)?"
    r"\$?(?P<col0>[A-Za-z]{1,3})\$?(?P<row0>\d+)"
    r"\s*:\s*\$?(?P<col1>[A-Za-z]{1,3})\$?(?P<row1>\d+)"
)
A1_REFERENCE_RE = re.compile(
    r"^\s*(?:(?:'(?P<quoted>(?:[^']|'')+)'|(?P<bare>[A-Za-z0-9_ .&-]+))!)?"
    r"\$?(?P<col0>[A-Za-z]{1,3})\$?(?P<row0>\d+)"
    r"(?:\s*:\s*\$?(?P<col1>[A-Za-z]{1,3})\$?(?P<row1>\d+))?\s*$"
)
OFFSET_BASE_RE = re.compile(
    r"^OFFSET\(\s*(?:(?:'(?P<quoted>(?:[^']|'')+)'|(?P<bare>[A-Za-z0-9_ .&-]+))!)?"
    r"\$?(?P<col>[A-Za-z]{1,3})\$?(?P<row>\d+)",
    re.IGNORECASE,
)
MAX_ITERATIONS = 5000
MAX_SWEEPS = 12
MAX_ACTIVE_PASSES = 12
EXCEL_DEFAULT_ITERATIONS = 100
EXCEL_DEFAULT_CHANGE = 0.001
MAX_WORKBOOK_ITERATIONS = 10_000
MAX_ITERATION_DELTA = 1.0
CONVERGENCE = 1e-12


class Unresolved:
    """A value the evaluator refuses to guess at; poisons anything downstream."""

    __slots__ = ("reason",)

    def __init__(self, reason: str):
        self.reason = reason

    def __repr__(self):
        return f"Unresolved({self.reason})"


class ExcelError:
    __slots__ = ("code",)

    def __init__(self, code: str):
        self.code = code

    def __repr__(self):
        return self.code


class RangeValues(list):
    """Flat range values with enough shape metadata for lookup/array functions."""

    def __init__(self, values, rows=1, cols=None, refs=None):
        super().__init__(values)
        self.rows = max(int(rows or 1), 1)
        self.cols = max(int(cols or len(values) or 1), 1)
        self.refs = tuple(refs) if refs is not None and len(refs) == len(self) else None


def is_bad(value) -> bool:
    return isinstance(value, (Unresolved, ExcelError))


@lru_cache(maxsize=8192)
def _datetime_to_serial(stamp: datetime) -> float:
    """Convert a real datetime to Excel's 1900-system serial."""
    delta = stamp - EXCEL_EPOCH
    serial = delta.days + delta.seconds / 86400.0 + delta.microseconds / 86400000000.0
    if stamp >= datetime(1900, 3, 1):
        serial += 1.0
    return serial


@lru_cache(maxsize=8192)
def _excel_date_parts(serial: int) -> tuple[int, int, int]:
    """Return Excel Y/M/D fields, including its fictitious serial-60 date."""
    serial = int(serial)
    if serial == 60:
        return 1900, 2, 29
    offset = serial if serial < 60 else serial - 1
    stamp = EXCEL_EPOCH + timedelta(days=offset)
    return stamp.year, stamp.month, stamp.day


def _excel_month_days(year: int, month: int) -> int:
    return 29 if (year, month) == (1900, 2) else monthrange(year, month)[1]


def _excel_date_serial(year: int, month: int, day: int) -> float:
    """Excel DATE normalization expressed in the same serial coordinate system."""
    absolute_month = year * 12 + month - 1
    normalized_year, month0 = divmod(absolute_month, 12)
    try:
        first = datetime(normalized_year, month0 + 1, 1)
    except (OverflowError, ValueError):
        raise ValueError("date outside supported range")
    return _datetime_to_serial(first) + day - 1


@lru_cache(maxsize=8192)
def to_serial(text: str):
    try:
        stamp = datetime.fromisoformat(text.replace("T", " "))
    except ValueError:
        return None
    return _datetime_to_serial(stamp)


def literal(raw: str):
    if raw is None or raw == "":
        return None
    if not isinstance(raw, str):
        return raw
    if ISO_RE.match(raw):
        serial = to_serial(raw)
        if serial is not None:
            return serial
    try:
        return float(raw)
    except ValueError:
        pass
    if raw in ("TRUE", "FALSE"):
        return raw == "TRUE"
    return raw


def num(value, default=0.0):
    if value is None:
        return default
    if isinstance(value, bool):
        return 1.0 if value else 0.0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        if ISO_RE.match(value):
            serial = to_serial(value)
            if serial is not None:
                return serial
        try:
            return float(value)
        except ValueError:
            return default
    return default


def flatten(args):
    for arg in args:
        if isinstance(arg, list):
            yield from flatten(arg)
        else:
            yield arg


def numbers(args):
    return [num(v) for v in flatten(args) if isinstance(v, (int, float)) and not isinstance(v, bool)]


def truthy(value) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    if isinstance(value, str):
        return value.upper() == "TRUE"
    return False


def _matches(value, criteria) -> bool:
    if isinstance(criteria, (int, float)) and not isinstance(criteria, bool):
        return isinstance(value, (int, float)) and float(value) == float(criteria)
    text = str(criteria)
    op, operand = CRITERIA_RE.match(text).groups()
    try:
        target = float(operand)
        left = num(value, math.nan)
    except ValueError:
        target = operand.strip().lower()
        left = str(value).strip().lower() if value is not None else ""
    if op in (None, "", "="):
        return left == target
    if isinstance(left, float) and isinstance(target, float):
        if math.isnan(left):
            return False
        return {
            ">": left > target, "<": left < target,
            ">=": left >= target, "<=": left <= target, "<>": left != target,
        }[op]
    return left != target if op == "<>" else False


def _eomonth(start, months):
    year, month, _ = _excel_date_parts(int(num(start)))
    absolute = year * 12 + month - 1 + int(num(months))
    target_year, month0 = divmod(absolute, 12)
    target_month = month0 + 1
    return _excel_date_serial(
        target_year, target_month, _excel_month_days(target_year, target_month)
    )


def _edate(start, months):
    base_year, base_month, base_day = _excel_date_parts(int(num(start)))
    absolute = base_year * 12 + base_month - 1 + int(num(months))
    year, month0 = divmod(absolute, 12)
    month = month0 + 1
    day = min(base_day, _excel_month_days(year, month))
    return _excel_date_serial(year, month, day)


def _solve_rate(exponents, values, guess=0.1, scan=True):
    """Excel-compatible rate solve.

    A cash flow series can cross zero more than once, so bisecting a wide bracket
    is liable to land on a root Excel never reports. Excel runs Newton from the
    guess and returns the root nearest it; bisection is only the fallback.

    ``scan=False`` skips the grid scan when the caller has proven the root is
    unique (a single sign change in the series); RATE-heavy workbooks call this
    hundreds of times and the unconditional scan would dominate the run.
    """
    def npv(rate):
        return sum(v * (1.0 + rate) ** -t for v, t in zip(values, exponents))

    def slope(rate):
        return sum(-t * v * (1.0 + rate) ** (-t - 1.0) for v, t in zip(values, exponents))

    rate = guess
    newton_root = None
    for _ in range(100):
        try:
            value, derivative = npv(rate), slope(rate)
        except (OverflowError, ZeroDivisionError, ValueError):
            break
        if abs(derivative) < 1e-14:
            break
        step = value / derivative
        nxt = rate - step
        if nxt <= -1.0:
            nxt = (rate - 1.0) / 2.0
        if abs(nxt - rate) < 1e-12:
            newton_root = nxt
            break
        rate = nxt
    else:
        newton_root = rate

    if not scan and newton_root is not None and newton_root > -1.0:
        try:
            if abs(npv(newton_root)) < 1e-7:
                return newton_root
        except (OverflowError, ZeroDivisionError, ValueError):
            pass

    # Multiple sign changes yield multiple mathematically valid IRRs. Excel
    # selects the root nearest the supplied guess; an unconstrained Newton step
    # can jump over that root and settle on a remote one. Find every bracketed
    # root on a mixed linear/log grid and use the closest.
    points = [-0.9999 + index * 1.9999 / 1000 for index in range(1001)]
    points += [10.0 ** (index / 100.0) - 1.0 for index in range(31, 601)]
    roots = []
    previous_rate, previous_value = None, None
    for candidate in points:
        try:
            value = npv(candidate)
        except (OverflowError, ZeroDivisionError, ValueError):
            previous_rate, previous_value = None, None
            continue
        if abs(value) < 1e-10:
            roots.append(candidate)
        elif previous_value is not None and value * previous_value < 0:
            low, high = previous_rate, candidate
            f_low = previous_value
            for _ in range(200):
                mid = (low + high) / 2.0
                f_mid = npv(mid)
                if abs(f_mid) < 1e-12 or high - low < 1e-14:
                    break
                if f_low * f_mid < 0:
                    high = mid
                else:
                    low, f_low = mid, f_mid
            roots.append((low + high) / 2.0)
        previous_rate, previous_value = candidate, value
    if roots:
        return min(roots, key=lambda root: abs(root - guess))
    return newton_root if newton_root is not None else ExcelError("#NUM!")


def _irr(values, guess=0.1):
    if not values:
        return ExcelError("#NUM!")
    return _solve_rate(list(range(len(values))), values, guess)


def _xirr(values, dates, guess=0.1):
    if len(values) != len(dates) or not values:
        return ExcelError("#NUM!")
    start = dates[0]
    return _solve_rate([(d - start) / 365.0 for d in dates], values, guess)


def _div(a, b):
    return ExcelError("#DIV/0!") if b == 0 else a / b


def _fv_value(rate, nper, pmt, pv, ptype):
    """Balance after ``nper`` periods (Excel FV sign convention)."""
    if rate == 0.0:
        return -(pv + pmt * nper)
    growth = (1.0 + rate) ** nper
    return -(pv * growth + pmt * (1.0 + rate * ptype) * (growth - 1.0) / rate)


def _pmt_value(rate, nper, pv, fv, ptype):
    if nper == 0:
        return ExcelError("#DIV/0!")
    if rate == 0.0:
        return -(pv + fv) / nper
    growth = (1.0 + rate) ** nper
    denominator = (growth - 1.0) * (1.0 + rate * ptype)
    if denominator == 0:
        return ExcelError("#DIV/0!")
    return -(pv * growth + fv) * rate / denominator


def _ipmt_value(rate, per, nper, pv, fv, ptype):
    """Interest share of period ``per``: rate on the balance carried into it."""
    payment = _pmt_value(rate, nper, pv, fv, ptype)
    if isinstance(payment, ExcelError):
        return payment
    if per < 1 or per > nper:
        return ExcelError("#NUM!")
    if ptype and per == 1:
        return 0.0
    balance = _fv_value(rate, per - 1, payment, pv, ptype)
    interest = balance * rate
    return interest / (1.0 + rate) if ptype else interest


_RATE_CACHE: dict = {}


def _rate_value(nper, pmt, pv, fv, ptype, guess):
    """RATE as a root of the same NPV equation IRR solves.

    The flow series is ``pv`` now, ``pmt`` each period (shifted one period
    earlier when payments are due at the start), and ``fv`` at the end.
    """
    if nper <= 0:
        return ExcelError("#NUM!")
    key = (nper, pmt, pv, fv, ptype, guess)
    if key in _RATE_CACHE:
        return _RATE_CACHE[key]
    periods = int(nper)
    if periods == nper:
        shift = 1.0 if ptype else 0.0
        exponents = [0.0] + [t - shift for t in range(1, periods + 1)] + [float(periods)]
        values = [pv] + [pmt] * periods + [fv]
        nonzero = [v for v in values if v]
        changes = sum(1 for a, b in zip(nonzero, nonzero[1:]) if a * b < 0)
        result = _solve_rate(exponents, values, guess, scan=changes > 1)
    else:
        # Fractional periods have no discrete flow series; Newton on the
        # closed-form annuity equation, which is what Excel evaluates.
        def balance(rate):
            if rate == 0.0:
                return pv + pmt * nper + fv
            growth = (1.0 + rate) ** nper
            return pv * growth + pmt * (1.0 + rate * ptype) * (growth - 1.0) / rate + fv

        rate = guess
        result = ExcelError("#NUM!")
        for _ in range(100):
            try:
                value = balance(rate)
                step = max(abs(rate), 1e-5) * 1e-6
                derivative = (balance(rate + step) - balance(rate - step)) / (2.0 * step)
            except (OverflowError, ZeroDivisionError, ValueError):
                break
            if abs(derivative) < 1e-14:
                break
            nxt = rate - value / derivative
            if nxt <= -1.0:
                nxt = (rate - 1.0) / 2.0
            if abs(nxt - rate) < 1e-12:
                result = nxt
                break
            rate = nxt
    if len(_RATE_CACHE) > 4096:
        _RATE_CACHE.clear()
    _RATE_CACHE[key] = result
    return result


def _norm_cdf(z):
    return 0.5 * (1.0 + math.erf(z / math.sqrt(2.0)))


def _norm_pdf(z):
    return math.exp(-0.5 * z * z) / math.sqrt(2.0 * math.pi)


def _norm_s_inv(p):
    """Acklam's rational approximation, polished with one Halley step."""
    if not 0.0 < p < 1.0:
        return ExcelError("#NUM!")
    a = (-3.969683028665376e+01, 2.209460984245205e+02, -2.759285104469687e+02,
         1.383577518672690e+02, -3.066479806614716e+01, 2.506628277459239e+00)
    b = (-5.447609879822406e+01, 1.615858368580409e+02, -1.556989798598866e+02,
         6.680131188771972e+01, -1.328068155288572e+01)
    c = (-7.784894002430293e-03, -3.223964580411365e-01, -2.400758277161838e+00,
         -2.549732539343734e+00, 4.374664141464968e+00, 2.938163982698783e+00)
    d = (7.784695709041462e-03, 3.224671290700398e-01, 2.445134137142996e+00,
         3.754408661907416e+00)
    p_low, p_high = 0.02425, 1.0 - 0.02425
    if p < p_low:
        q = math.sqrt(-2.0 * math.log(p))
        z = (((((c[0] * q + c[1]) * q + c[2]) * q + c[3]) * q + c[4]) * q + c[5]) / \
            ((((d[0] * q + d[1]) * q + d[2]) * q + d[3]) * q + 1.0)
    elif p <= p_high:
        q = p - 0.5
        r = q * q
        z = (((((a[0] * r + a[1]) * r + a[2]) * r + a[3]) * r + a[4]) * r + a[5]) * q / \
            (((((b[0] * r + b[1]) * r + b[2]) * r + b[3]) * r + b[4]) * r + 1.0)
    else:
        q = math.sqrt(-2.0 * math.log(1.0 - p))
        z = -(((((c[0] * q + c[1]) * q + c[2]) * q + c[3]) * q + c[4]) * q + c[5]) / \
            ((((d[0] * q + d[1]) * q + d[2]) * q + d[3]) * q + 1.0)
    error = _norm_cdf(z) - p
    density = _norm_pdf(z)
    if density > 0:
        u = error / density
        z -= u / (1.0 + z * u / 2.0)
    return z


def _day_count_30_360(start: datetime, end: datetime, european: bool):
    d1, d2 = start.day, end.day
    if european:
        d1, d2 = min(d1, 30), min(d2, 30)
    else:
        start_feb_end = start.month == 2 and d1 == monthrange(start.year, 2)[1]
        end_feb_end = end.month == 2 and d2 == monthrange(end.year, 2)[1]
        if start_feb_end and end_feb_end:
            d2 = 30
        if start_feb_end:
            d1 = 30
        if d2 == 31 and d1 >= 30:
            d2 = 30
        if d1 == 31:
            d1 = 30
    return (end.year - start.year) * 360 + (end.month - start.month) * 30 + (d2 - d1)


def snap(result: float, *operands: float) -> float:
    """Mimic Excel's cancellation rule for sums.

    Excel zeroes a sum whose magnitude is negligible against its operands, so
    ``=SUM(a,-a)`` is exactly 0 and any ``IF(x>0)`` downstream takes the false
    branch. Plain IEEE arithmetic leaves a ~1e-15 residue and flips the branch.
    """
    scale = max((abs(x) for x in operands), default=0.0)
    return 0.0 if scale and abs(result) < scale * 1e-13 else result


BINARY = {
    "+": lambda a, b: snap(num(a) + num(b), num(a), num(b)),
    "-": lambda a, b: snap(num(a) - num(b), num(a), num(b)),
    "*": lambda a, b: num(a) * num(b),
    "/": lambda a, b: _div(num(a), num(b)),
    "^": lambda a, b: _power(num(a), num(b)),
    "&": lambda a, b: f"{_text(a)}{_text(b)}",
    "=": lambda a, b: _equal(a, b),
    "<>": lambda a, b: not _equal(a, b),
    ">": lambda a, b: num(a) > num(b),
    "<": lambda a, b: num(a) < num(b),
    ">=": lambda a, b: num(a) >= num(b),
    "<=": lambda a, b: num(a) <= num(b),
}


def _unary_plus(value):
    """Excel's reference-style unary plus keeps nonnumeric text as text."""
    if isinstance(value, str):
        try:
            return float(value)
        except ValueError:
            return value
    return num(value)


UNARY = {
    "u-": lambda a: -num(a),
    "u+": _unary_plus,
    "%": lambda a: num(a) / 100.0,
}


def _power(a, b):
    try:
        result = a ** b
    except (ValueError, OverflowError, ZeroDivisionError):
        return ExcelError("#NUM!")
    return ExcelError("#NUM!") if isinstance(result, complex) else result


def _text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _equal(a, b):
    if isinstance(a, str) or isinstance(b, str):
        return _text(a).strip().lower() == _text(b).strip().lower()
    return num(a) == num(b)


def _binary_value(op, left, right):
    """Apply one scalar operator without swallowing spreadsheet errors."""
    for error_type in (Unresolved, ExcelError):
        if isinstance(left, error_type):
            return left
        if isinstance(right, error_type):
            return right
    return BINARY[op](left, right)


def _binary(op, left, right):
    """Excel array arithmetic: scalar broadcast or exact-shape elementwise."""
    left_range = isinstance(left, RangeValues)
    right_range = isinstance(right, RangeValues)
    if not left_range and not right_range:
        return _binary_value(op, left, right)
    if left_range and right_range:
        if (
            (left.rows, left.cols) != (right.rows, right.cols)
            or len(left) != len(right)
        ):
            return ExcelError("#VALUE!")
        values = [
            _binary_value(op, left_value, right_value)
            for left_value, right_value in zip(left, right)
        ]
        return RangeValues(values, left.rows, left.cols)
    source = left if left_range else right
    scalar = right if left_range else left
    values = [
        _binary_value(op, value, scalar) if left_range
        else _binary_value(op, scalar, value)
        for value in source
    ]
    return RangeValues(values, source.rows, source.cols)


def _intersection(left, right):
    """Return the overlapping cells of two reference-valued ranges."""
    for error_type in (Unresolved, ExcelError):
        if isinstance(left, error_type):
            return left
        if isinstance(right, error_type):
            return right
    if not isinstance(left, RangeValues) or not isinstance(right, RangeValues):
        return ExcelError("#NULL!")
    if left.refs is None or right.refs is None:
        return ExcelError("#NULL!")
    right_refs = set(right.refs)
    selected = [
        (value, ref) for value, ref in zip(left, left.refs) if ref in right_refs
    ]
    if not selected:
        return ExcelError("#NULL!")
    coordinates = [split_ref(ref) for _, ref in selected]
    valid = [coordinate for coordinate in coordinates if coordinate is not None]
    rows = len({(sheet, row) for sheet, row, _ in valid}) or 1
    cols = len({(sheet, col) for sheet, _, col in valid}) or len(selected)
    return RangeValues(
        [value for value, _ in selected],
        rows,
        cols,
        refs=[ref for _, ref in selected],
    )


@dataclass
class EvalResult:
    values: dict
    unresolved: dict
    iterated: list
    coverage: dict
    # target cell -> source cells actually read on the final evaluation pass,
    # including dynamic references (OFFSET) invisible to the static graph.
    runtime_radj: dict = dataclass_field(default_factory=dict)
    # Address-only origins used to resolve dynamic references.  These are
    # deliberately separate from value precedents so OFFSET(A1, ...) does not
    # make A1 a proof input merely because its coordinate anchors the address.
    runtime_address_radj: dict = dataclass_field(default_factory=dict)
    # Dynamic/selector targets chosen on the final pass, sorted for stable
    # diagnostics and proof-closure signatures.
    resolved_targets: dict = dataclass_field(default_factory=dict)
    # The same targets keyed by the exact AST operation that resolved them.
    # Formula-level targets above are retained for dependency diagnostics, but
    # cannot prove two dynamic calls in one formula independently.
    resolved_operation_targets: dict = dataclass_field(default_factory=dict)
    # cell id served by the workbook cached-value oracle -> the formula cells
    # that consumed it. Every entry is a value the evaluator could not derive
    # and copied from the golden workbook instead.
    oracle_reads: dict = dataclass_field(default_factory=dict)
    oracle_accesses: dict = dataclass_field(default_factory=dict)
    # Every grid read attempted by a formula, including reads which produced no
    # value.  This is diagnostic provenance, not a new frontier-closure policy.
    read_attempts: dict = dataclass_field(default_factory=dict)
    missing_reads: dict = dataclass_field(default_factory=dict)
    seed_provenance: dict = dataclass_field(default_factory=dict)
    seeded_cells: set = dataclass_field(default_factory=set)
    # Complete formula/input scope authorized for this run. ``values`` also
    # contains cached initialization state outside a strict proof scope, so it
    # cannot by itself establish that a cell was executed.
    evaluated_cells: set = dataclass_field(default_factory=set)
    cycle_membership: dict = dataclass_field(default_factory=dict)
    # Verification-only state.  These are intentionally not emitted directly:
    # diagnostics carry bounded member samples while this complete mapping lets
    # the post-proof verifier reason about every member without truncation.
    cycle_groups: dict = dataclass_field(default_factory=dict)
    endpoint_equation_checker: object = None
    strict_proof: bool = False


@dataclass(frozen=True)
class CalculationMetadata:
    """Calculation settings obtained without opening the expected-value cache."""

    available: bool
    iterate: bool | None = None
    iterate_count: int | None = None
    iterate_delta: float | None = None
    source: str = ""
    reason: str = ""
    iterate_origin: str = "unknown"
    iterate_count_origin: str = "unknown"
    iterate_delta_origin: str = "unknown"
    raw_calc_pr: dict = dataclass_field(default_factory=dict)
    calc_mode: str | None = None
    full_calc_on_load: bool | None = None
    force_full_calc: bool | None = None
    calc_mode_origin: str = "unknown"
    full_calc_on_load_origin: str = "unknown"
    force_full_calc_origin: str = "unknown"

    @property
    def calcMode(self):
        """OOXML spelling retained for metadata consumers."""
        return self.calc_mode

    @property
    def fullCalcOnLoad(self):
        """OOXML spelling retained for metadata consumers."""
        return self.full_calc_on_load

    @property
    def forceFullCalc(self):
        """OOXML spelling retained for metadata consumers."""
        return self.force_full_calc


class Evaluator:
    def __init__(
        self,
        graph: Graph,
        cg: CellGraph,
        oracle=None,
        *,
        strict_proof: bool = False,
        calculation=None,
        proof_outputs=None,
        run_probes: bool = True,
        proof_scope=None,
    ):
        self.graph = graph
        self.cg = cg
        self.oracle = oracle
        self.strict_proof = strict_proof
        self.calculation = calculation if calculation is not None else oracle
        self.proof_outputs = set(proof_outputs) if proof_outputs is not None else None
        self.run_probes = run_probes
        self.proof_scope = set(proof_scope) if proof_scope is not None else None
        self.values: dict = {}
        self.unresolved: dict = {}
        self.unknown_ops: dict = defaultdict(int)
        self._arg_cache: dict = {}
        self._subtotal_provenance: dict[str, bool] = {}
        self.oracle_reads: dict = {}
        self.oracle_accesses: dict = {}
        self.read_attempts: dict = {}
        self.missing_reads: dict = {}
        self.runtime_address_radj: dict = {}
        self._resolved_targets: dict = {}
        self._resolved_operation_targets: dict = {}
        self._selective_sources: dict = {}
        self._current_cell = None
        self._proof_endpoint_values: dict = {}
        self._proof_active_radj: dict = {}
        self._proof_resolved_targets: dict = {}
        self._proof_resolved_operation_targets: dict = {}

    # -- AST walking ------------------------------------------------------
    def _arg_specs(self, node):
        """Cached positional argument specifications for an op node.

        Slots are indexed by ``arg_index`` rather than packed, because the parser
        emits no edge for an operand that points at an empty cell. Packing would
        silently slide ``=J3+1`` into a one-argument addition.
        """
        cached = self._arg_cache.get(node.id)
        if cached is None:
            grouped: dict = {}
            for edge in self.graph.in_edges.get(node.id, ()):
                grouped.setdefault(edge.arg_index, []).append(edge)
            declared = int(node.arity) if str(node.arity).isdigit() else 0
            size = max([declared] + [k + 1 for k in grouped]) if (grouped or declared) else 0
            cached = [self._spec(grouped.get(i)) for i in range(size)]
            self._arg_cache[node.id] = cached
        return cached

    def _arg_value(self, spec):
        kind, payload = spec
        if kind == "none":
            return None
        if kind == "scalar":
            return self._eval_node(payload)
        if kind == "span":
            values = [self._read(cid) for cid in payload]
            coordinates = [_split(cid) for cid in payload]
            rows = len({row for _, row, _ in coordinates})
            cols = len({col for _, _, col in coordinates})
            return RangeValues(values, rows, cols, refs=payload)
        return [self._eval_node(src) for src in payload]

    def _arg(self, node, index):
        specs = self._arg_specs(node)
        return self._arg_value(specs[index]) if index < len(specs) else None

    def _args(self, node):
        return [self._arg_value(spec) for spec in self._arg_specs(node)]

    def _reference_arg(self, node, index, *, read=True):
        """Read an argument while retaining any grid coordinates it represents."""
        specs = self._arg_specs(node)
        if index >= len(specs):
            return RangeValues([], 1, 1)
        spec = specs[index]
        kind, payload = spec
        if not read:
            if kind == "span":
                coordinates = [split_ref(ref) for ref in payload]
                rows = len({item[1] for item in coordinates if item is not None}) or 1
                cols = len({item[2] for item in coordinates if item is not None}) or 1
                return RangeValues([None] * len(payload), rows, cols, refs=payload)
            if kind == "scalar":
                source = self.graph.nodes.get(payload)
                if source is not None and source.kind == "range":
                    refs = range_members(source.sheet, source.coordinate or "")
                    return RangeValues(
                        [None] * len(refs),
                        len({split_ref(ref)[1] for ref in refs}) or 1,
                        len({split_ref(ref)[2] for ref in refs}) or 1,
                        refs=refs,
                    )
                if source is not None and source.is_cell:
                    return RangeValues([None], 1, 1, refs=[source.id])
            if kind == "many":
                refs = [
                    source_id for source_id in payload
                    if self.graph.nodes.get(source_id) is not None
                    and self.graph.nodes[source_id].is_cell
                ]
                if len(refs) == len(payload):
                    coordinates = [split_ref(ref) for ref in refs]
                    rows = len({item[1] for item in coordinates if item is not None}) or 1
                    cols = len({item[2] for item in coordinates if item is not None}) or 1
                    return RangeValues([None] * len(refs), rows, cols, refs=refs)
        value = self._arg_value(spec)
        if isinstance(value, RangeValues):
            return value
        if kind == "scalar":
            source = self.graph.nodes.get(payload)
            if source is not None and source.is_cell:
                return RangeValues([value], 1, 1, refs=[source.id])
        if kind == "many":
            refs = []
            for source_id in payload:
                source = self.graph.nodes.get(source_id)
                if source is None or not source.is_cell:
                    refs = []
                    break
                refs.append(source.id)
            if refs:
                coordinates = [split_ref(ref) for ref in refs]
                rows = len({item[1] for item in coordinates if item is not None}) or 1
                cols = len({item[2] for item in coordinates if item is not None}) or 1
                return RangeValues(list(flatten([value])), rows, cols, refs=refs)
        values = list(flatten([value]))
        return RangeValues(values, 1, max(len(values), 1))

    @staticmethod
    def _reference_anchor(source):
        parsed = [item for ref in (source.refs or ()) if (item := split_ref(ref))]
        if not parsed:
            return None
        sheet = parsed[0][0]
        same_sheet = [item for item in parsed if item[0] == sheet]
        return sheet, min(item[1] for item in same_sheet), min(item[2] for item in same_sheet)

    def _aligned_reference(self, source, rows, cols, *, read=True):
        """Resize a reference from its absolute top-left cell, as Excel does."""
        rows, cols = max(int(rows), 1), max(int(cols), 1)
        anchor = self._reference_anchor(source)
        if anchor is None:
            values = list(source)
            size = rows * cols
            return RangeValues(
                values[:size] + [None] * max(size - len(values), 0), rows, cols
            )
        sheet, top, left = anchor
        cached = dict(zip(source.refs or (), source))
        refs = [
            a1(sheet, top + row, left + col)
            for row in range(rows)
            for col in range(cols)
        ]
        values = [
            cached[ref] if ref in cached else (self._read(ref) if read else None)
            for ref in refs
        ]
        return RangeValues(values, rows, cols, refs=refs)

    @staticmethod
    def _bad_in(values):
        return next((value for value in flatten(values) if is_bad(value)), None)

    def _conditional_layout(self, node):
        """Return condition values and aligned summand coordinates."""
        is_sumifs = node.op == "SUMIFS"
        target_index = 0 if is_sumifs else 2
        first_range = 1 if is_sumifs else 0
        criterion_pairs = []

        if is_sumifs:
            target = self._reference_arg(node, 0, read=False)
            rows, cols = target.rows, target.cols
        else:
            first = self._reference_arg(node, 0)
            rows, cols = first.rows, first.cols
            target = self._reference_arg(node, target_index, read=False) \
                if target_index < len(self._arg_specs(node)) else first

        for index in range(first_range, len(self._arg_specs(node)) - 1, 2):
            raw_source = self._reference_arg(node, index)
            if is_sumifs and (
                raw_source.rows != rows
                or raw_source.cols != cols
                or len(raw_source) != rows * cols
            ):
                return target, rows, cols, criterion_pairs, [], ExcelError("#VALUE!")
            source = self._aligned_reference(raw_source, rows, cols)
            criterion_pairs.append((source, self._arg(node, index + 1), index + 1))

        if not is_sumifs:
            criterion_pairs = [
                (self._aligned_reference(first, rows, cols), self._arg(node, 1), 1)
            ]
        bad = self._bad_in([
            item
            for source, criterion, _ in criterion_pairs
            for item in (source, criterion)
        ])
        if bad is not None:
            return target, rows, cols, criterion_pairs, [], bad
        matches = [
            all(_matches(source[offset], criterion) for source, criterion, _ in criterion_pairs)
            for offset in range(rows * cols)
        ]
        return target, rows, cols, criterion_pairs, matches, None

    def _conditional_sum(self, node):
        target, rows, cols, _, matches, bad = self._conditional_layout(node)
        if bad is not None:
            return bad
        anchor = self._reference_anchor(target)
        cached = dict(zip(target.refs or (), target))
        values = list(target)
        selected = []
        for offset, matched in enumerate(matches):
            if not matched:
                continue
            row, col = divmod(offset, cols)
            if anchor is not None:
                sheet, top, left = anchor
                ref = a1(sheet, top + row, left + col)
                selected.append(
                    self._read(ref) if self.graph is not None else cached.get(ref)
                )
            elif offset < len(values):
                selected.append(values[offset])
        bad = self._bad_in(selected)
        if bad is not None:
            return bad
        nums = [num(value) for value in selected]
        return snap(sum(nums), *nums)

    def _conditional_sources(self, node):
        target, rows, cols, criterion_pairs, matches, _ = self._conditional_layout(node)
        sources = set()
        for source, _, criterion_index in criterion_pairs:
            sources.update(source.refs or ())
            spec = self._arg_specs(node)[criterion_index]
            sources.update(self._active_spec_sources(spec))

        anchor = self._reference_anchor(target)
        if anchor is not None:
            sheet, top, left = anchor
            for offset, matched in enumerate(matches):
                if matched:
                    row, col = divmod(offset, cols)
                    sources.add(a1(sheet, top + row, left + col))
        elif target.refs:
            sources.update(
                ref for ref, matched in zip(target.refs, matches) if matched
            )
        return sources

    def _active_spec_sources(self, spec, seen=None):
        kind, payload = spec
        if kind == "span":
            return set(payload)
        if kind == "scalar":
            return self._active_node_sources(payload, seen)
        if kind == "many":
            sources = set()
            for source in payload:
                sources.update(self._active_node_sources(source, seen))
            return sources
        return set()

    def _has_subtotal_provenance(self, node_id, seen=None):
        """Whether a cell's evaluated lineage contains SUBTOTAL or AGGREGATE."""
        cached = self._subtotal_provenance.get(node_id)
        if cached is not None:
            return cached
        seen = set() if seen is None else seen
        if node_id in seen:
            return False
        seen.add(node_id)
        node = self.graph.nodes.get(node_id)
        if node is None:
            return False
        if node.is_cell:
            root = self.graph.root_of(node.id)
            result = root is not None and self._has_subtotal_provenance(root, seen)
        elif node.kind == "op":
            bare = node.op.upper()
            for prefix in ("_XLFN.", "_XLWS."):
                if bare.startswith(prefix):
                    bare = bare[len(prefix):]
            result = bare in {"SUBTOTAL", "AGGREGATE"} or any(
                self.graph.nodes.get(edge.source) is not None
                and not self.graph.nodes[edge.source].is_cell
                and self._has_subtotal_provenance(edge.source, seen)
                for edge in self.graph.in_edges.get(node.id, ())
            )
        else:
            result = False
        self._subtotal_provenance[node_id] = result
        return result

    def _subtotal(self, node):
        code_value = self._arg(node, 0)
        if is_bad(code_value):
            return code_value
        code = int(num(code_value)) % 100
        kept = []
        for index in range(1, len(self._arg_specs(node))):
            source = self._reference_arg(node, index)
            if source.refs:
                kept.extend(
                    value for value, ref in zip(source, source.refs)
                    if not self._has_subtotal_provenance(ref)
                )
            else:
                kept.extend(source)
        bad = self._bad_in(kept)
        if bad is not None:
            return bad
        if code == 1:
            return self._fn_average(kept)
        if code == 2:
            return self._fn_count(kept)
        if code == 3:
            return float(sum(value is not None and value != "" for value in kept))
        if code == 4:
            return self._fn_max(kept)
        if code == 5:
            return self._fn_min(kept)
        if code == 6:
            return self._fn_product(kept)
        if code == 9:
            return self._fn_sum(kept)
        return Unresolved(f"unsupported:SUBTOTAL({code})")

    def _spec(self, group):
        """Decide once how an argument slot should be read.

        Range arguments are re-derived from their A1 span rather than from the
        edges, because the parser emits no edge for an empty cell. Reading them
        edge-wise silently shortens the list, which slides every later element
        onto the wrong position in paired-range functions like XIRR.
        """
        if not group:
            return ("none", None)
        span = group[0].via_range
        if span:
            sheet = self.graph.nodes[group[0].source].sheet
            members = range_members(sheet, span.split("!")[-1].replace("$", ""))
            if members:
                return ("span", members)
        if len(group) == 1:
            return ("scalar", group[0].source)
        return ("many", [e.source for e in group])

    def _read(self, cid: str):
        self.read_attempts.setdefault(cid, set()).add(self._current_cell)
        if cid in self.values:
            return self.values[cid]
        if cid in self.graph.nodes:
            return Unresolved("range-uncomputed")
        if self.strict_proof:
            # Proof mode may record the absence, but it cannot ask the golden
            # workbook for a saved answer.  Comparison cache access happens
            # only after ``run`` has returned.
            self.missing_reads.setdefault(cid, set()).add(self._current_cell)
            return None
        if self.oracle is not None:
            self.oracle_accesses.setdefault(cid, set()).add(self._current_cell)
            raw = self.oracle(*_split(cid))
            value = literal(raw) if isinstance(raw, str) else raw
            if value is not None:
                # A populated cell the graph never recorded: its value comes
                # straight from the golden workbook's cache, not from any
                # rebuild. Record who consumed it so verification can refuse
                # to count outputs fed this way.
                self.oracle_reads.setdefault(cid, set()).add(self._current_cell)
            return value
        return None

    def _eval_node(self, node_id):
        node = self.graph.nodes.get(node_id)
        if node is None:
            return Unresolved("missing-node")
        if node.kind == "const":
            if node.op == "text" and node.value == "":
                return ""
            return literal(node.value if node.value != "" else node.expr)
        if node.kind == "range":
            return self._expand_range(node)
        if node.kind == "name":
            return ExcelError("#NAME?")
        if node.kind == "external":
            return Unresolved("external-reference")
        if node.is_cell:
            if node_id in self.values:
                return self.values[node_id]
            return Unresolved("uncomputed-precedent")
        return self._apply(node)

    def _expand_range(self, node):
        """A range node the parser kept whole; read its cells one by one."""
        if getattr(node, "range_truncated", False):
            return Unresolved("whole-column-range-exceeds-bound")
        members = range_members(node.sheet, node.coordinate or "")
        if not members:
            return Unresolved("opaque-range")
        coordinates = [_split(cid) for cid in members]
        rows = len({row for _, row, _ in coordinates})
        cols = len({col for _, _, col in coordinates})
        return RangeValues(
            [self._read(cid) for cid in members], rows, cols, refs=members
        )

    def _apply(self, node):
        op = node.op
        if op == "OFFSET":
            return self._offset(node)
        if op == "SUBTOTAL":
            return self._subtotal(node)
        if op in ("SUMIF", "SUMIFS"):
            return self._conditional_sum(node)
        if op == "CELL":
            return self._cell_info(node)
        if op == "INDIRECT":
            return self._indirect(node)
        if op == "ROW":
            return self._row(node)
        if op == "COUNTA":
            return self._counta(node)
        if op == "IFERROR":
            if not self._arg_specs(node):
                return Unresolved("iferror-arity")
            value = self._arg(node, 0)
            if isinstance(value, ExcelError):
                fallback = self._arg(node, 1)
                return 0.0 if fallback is None else fallback
            return value
        if op == "IF":
            cond = self._arg(node, 0)
            if is_bad(cond):
                return cond
            chosen = 1 if truthy(cond) else 2
            value = self._arg(node, chosen)
            if value is not None:
                return value
            # Excel: an omitted value_if_true yields 0, an omitted
            # value_if_false yields FALSE.
            return False if chosen == 2 else 0.0
        if op == "CHOOSE":
            index = self._arg(node, 0)
            if index is None:
                return Unresolved("choose-arity")
            if is_bad(index):
                return index
            selected = int(num(index))
            if not 1 <= selected < len(self._arg_specs(node)):
                return ExcelError("#VALUE!")
            return self._arg(node, selected)
        if op in ("IFS", "_XLFN.IFS"):
            specs = self._arg_specs(node)
            for index in range(0, len(specs) - 1, 2):
                cond = self._arg(node, index)
                if is_bad(cond):
                    return cond
                if truthy(cond):
                    return self._arg(node, index + 1)
            return ExcelError("#N/A")
        args = self._args(node)
        if node.op_kind == "infix":
            if op in BINARY:
                return _binary(op, args[0], args[1]) if len(args) > 1 \
                    else Unresolved("arity")
            if op in ("isect", " "):
                return _intersection(args[0], args[1]) if len(args) > 1 \
                    else Unresolved("arity")
        selective = (
            "VLOOKUP", "HLOOKUP", "XLOOKUP", "_XLFN.XLOOKUP",
            "LOOKUP", "MATCH", "INDEX",
        )
        if op in selective:
            # Lookups tolerate bad values inside the searched range -- only the
            # matched row matters -- but a bad key/index/mode must poison the
            # result. Letting it fall through to num() would silently search
            # for 0 and fabricate a match Excel never produces.
            scalar_slots = {
                "VLOOKUP": (0, 2, 3), "HLOOKUP": (0, 2, 3),
                "XLOOKUP": (0,), "_XLFN.XLOOKUP": (0,),
                "LOOKUP": (0,),
                "MATCH": (0, 2), "INDEX": (1, 2),
            }[op]
            for slot in scalar_slots:
                if slot < len(args) and is_bad(args[slot]):
                    return args[slot]
        bad = next((a for a in flatten(args) if isinstance(a, Unresolved)), None)
        if bad is not None and op not in selective:
            return bad
        err = next((a for a in flatten(args) if isinstance(a, ExcelError)), None)
        if err is not None and op not in (
            "COUNTIF", "COUNTIFS", "SUMIF", "SUMIFS", "AVERAGEIFS", "AVERAGEIF",
            "MAXIFS", "_XLFN.MAXIFS", "MINIFS", "_XLFN.MINIFS",
            "ISERR", "ISERROR", "ISNA", "ISNUMBER", "IFNA", "_XLFN.IFNA",
            *selective
        ):
            return err
        if op == "{}":
            values = list(flatten(args))
            rows = (node.expr or "").count(";") + 1
            if not values or len(values) % rows:
                return Unresolved("array-constant-shape")
            return RangeValues(values, rows, len(values) // rows)
        if node.op_kind in ("prefix", "postfix") and op in UNARY:
            if not args:
                return Unresolved("arity")
            if isinstance(args[0], RangeValues):
                return RangeValues(
                    [UNARY[op](value) for value in args[0]],
                    args[0].rows,
                    args[0].cols,
                    refs=args[0].refs,
                )
            return UNARY[op](args[0])
        # Excel stores post-2007 functions with a future-function prefix
        # (``_xlfn.NORM.DIST``), which the tokenizer keeps verbatim. Try the
        # bare name first, but keep the raw lookup as fallback: handlers like
        # ``_fn__xlfn_vstack`` exist only under the prefixed spelling.
        bare = op
        for prefix in ("_XLFN.", "_XLWS."):
            if bare.startswith(prefix):
                bare = bare[len(prefix):]
        handler = getattr(self, f"_fn_{bare.replace('.', '_').lower()}", None)
        if handler is None and bare != op:
            handler = getattr(self, f"_fn_{op.replace('.', '_').lower()}", None)
        if handler is None:
            self.unknown_ops[op] += 1
            return Unresolved(f"unsupported:{op}")
        value = handler(args)
        if op in selective:
            self._selective_sources[node.id] = self._selective_dependencies(
                node, args, value
            )
        return value

    # -- functions --------------------------------------------------------
    def _fn_sum(self, args):
        vals = numbers(args)
        return snap(sum(vals), *vals)

    def _fn_average(self, args):
        vals = numbers(args)
        return sum(vals) / len(vals) if vals else ExcelError("#DIV/0!")

    def _fn_count(self, args):
        return float(len(numbers(args)))

    def _fn_min(self, args):
        vals = numbers(args)
        return min(vals) if vals else 0.0

    def _fn_max(self, args):
        vals = numbers(args)
        return max(vals) if vals else 0.0

    def _fn_median(self, args):
        vals = sorted(numbers(args))
        if not vals:
            return ExcelError("#NUM!")
        mid = len(vals) // 2
        return vals[mid] if len(vals) % 2 else (vals[mid - 1] + vals[mid]) / 2

    def _fn_round(self, args):
        digits = int(num(args[1])) if len(args) > 1 else 0
        factor = 10.0 ** digits
        return math.floor(abs(num(args[0])) * factor + 0.5) / factor * (1 if num(args[0]) >= 0 else -1)

    def _fn_roundup(self, args):
        value = num(args[0])
        digits = int(num(args[1])) if len(args) > 1 else 0
        factor = 10.0 ** digits
        return math.ceil(abs(value) * factor - 1e-12) / factor * (
            1 if value >= 0 else -1
        )

    def _fn_rounddown(self, args):
        value = num(args[0])
        digits = int(num(args[1])) if len(args) > 1 else 0
        factor = 10.0 ** digits
        return math.floor(abs(value) * factor + 1e-12) / factor * (
            1 if value >= 0 else -1
        )

    def _fn_mround(self, args):
        value = num(args[0])
        multiple = num(args[1]) if len(args) > 1 else 0.0
        if multiple == 0:
            return 0.0
        if value * multiple < 0:
            return ExcelError("#NUM!")
        rounded = math.floor(abs(value / multiple) + 0.5 + 1e-12)
        return math.copysign(rounded * abs(multiple), value)

    def _fn_abs(self, args):
        return abs(num(args[0]))

    def _fn_not(self, args):
        return not truthy(args[0])

    def _fn_and(self, args):
        return all(truthy(value) for value in flatten(args) if value is not None)

    def _fn_or(self, args):
        return any(truthy(value) for value in flatten(args) if value is not None)

    def _fn_year(self, args):
        return float(_excel_date_parts(int(num(args[0])))[0])

    def _fn_month(self, args):
        return float(_excel_date_parts(int(num(args[0])))[1])

    def _fn_day(self, args):
        return float(_excel_date_parts(int(num(args[0])))[2])

    def _fn_date(self, args):
        year = int(num(args[0]))
        month = int(num(args[1]))
        day = int(num(args[2]))
        if 0 <= year < 1900:
            year += 1900
        try:
            return float(_excel_date_serial(year, month, day))
        except (OverflowError, ValueError):
            return ExcelError("#NUM!")

    def _fn_eomonth(self, args):
        return float(_eomonth(args[0], args[1]))

    def _fn_edate(self, args):
        return float(_edate(args[0], args[1]))

    def _fn_numbervalue(self, args):
        if not args:
            return ExcelError("#VALUE!")
        text = _text(args[0]).strip().replace("\u00a0", " ")
        decimal = _text(args[1]) if len(args) > 1 and args[1] is not None else "."
        group = _text(args[2]) if len(args) > 2 and args[2] is not None else ","
        if (
            len(decimal) > 1 or len(group) > 1
            or (decimal and group and decimal == group)
        ):
            return ExcelError("#VALUE!")
        percent = 0
        while text.endswith("%"):
            percent += 1
            text = text[:-1].rstrip()
        if "%" in text:
            return ExcelError("#VALUE!")
        text = text.replace(" ", "")
        if decimal:
            if text.count(decimal) > 1:
                return ExcelError("#VALUE!")
            integer, separator, fraction = text.partition(decimal)
        else:
            integer, separator, fraction = text, "", ""
        if group and group in fraction:
            return ExcelError("#VALUE!")
        if group:
            integer = integer.replace(group, "")
        normalized = integer + (("." + fraction) if separator else "")
        try:
            return float(normalized) / (100.0 ** percent)
        except ValueError:
            return ExcelError("#VALUE!")

    def _fn_rank(self, args):
        if len(args) < 2:
            return ExcelError("#N/A")
        number = args[0]
        if not isinstance(number, (int, float)) or isinstance(number, bool):
            return ExcelError("#VALUE!")
        values = numbers([args[1]])
        if not values:
            return ExcelError("#N/A")
        ascending = len(args) > 2 and num(args[2]) != 0
        if ascending:
            return float(1 + sum(value < float(number) for value in values))
        return float(1 + sum(value > float(number) for value in values))

    _fn_rank_eq = _fn_rank

    def _fn_days(self, args):
        return num(args[0]) - num(args[1])

    _fn__xlfn_days = _fn_days

    def _fn_today(self, args):
        return float(math.floor(_datetime_to_serial(datetime.today())))

    def _fn_choose(self, args):
        idx = int(num(args[0]))
        return args[idx] if 1 <= idx < len(args) else ExcelError("#VALUE!")

    def _fn_index(self, args):
        source = args[0]
        pool = list(flatten([source]))
        if not isinstance(source, RangeValues):
            row = int(num(args[1])) if len(args) > 1 else 1
            return pool[row - 1] if 1 <= row <= len(pool) else ExcelError("#REF!")
        selector = int(num(args[1])) if len(args) > 1 else 1
        has_col = len(args) > 2 and args[2] is not None
        if not has_col and source.rows == 1:
            row, col = 1, selector
        else:
            row = selector
            col = int(num(args[2])) if has_col else 1
        if row == 0 and col == 0:
            return source
        if row == 0:
            if not 1 <= col <= source.cols:
                return ExcelError("#REF!")
            indices = list(range(col - 1, len(pool), source.cols))
            refs = [source.refs[index] for index in indices] \
                if source.refs is not None else None
            return RangeValues(
                [pool[index] for index in indices], len(indices), 1, refs=refs
            )
        if col == 0:
            if not 1 <= row <= source.rows:
                return ExcelError("#REF!")
            start = (row - 1) * source.cols
            indices = list(range(start, min(start + source.cols, len(pool))))
            refs = [source.refs[index] for index in indices] \
                if source.refs is not None else None
            return RangeValues(
                [pool[index] for index in indices], 1, source.cols, refs=refs
            )
        index = (row - 1) * source.cols + col - 1
        if not (1 <= row <= source.rows and 1 <= col <= source.cols and index < len(pool)):
            return ExcelError("#REF!")
        return pool[index]

    def _fn_countif(self, args):
        pool = list(flatten([args[0]]))
        return float(sum(1 for v in pool if _matches(v, args[1])))

    def _fn_sumif(self, args):
        pool = list(flatten([args[0]]))
        target = list(flatten([args[2]])) if len(args) > 2 else pool
        total = 0.0
        for i, value in enumerate(pool):
            if _matches(value, args[1]) and i < len(target):
                total += num(target[i])
        return total

    @staticmethod
    def _criteria_matches(args, start):
        ranges = []
        for index in range(start, len(args) - 1, 2):
            pool = list(flatten([args[index]]))
            ranges.append((pool, args[index + 1]))
        width = min((len(pool) for pool, _ in ranges), default=0)
        return [
            all(_matches(pool[index], criterion) for pool, criterion in ranges)
            for index in range(width)
        ]

    def _fn_sumifs(self, args):
        target = list(flatten([args[0]])) if args else []
        matches = self._criteria_matches(args, 1)
        vals = [
            num(target[index]) for index, matched in enumerate(matches)
            if matched and index < len(target)
        ]
        return snap(sum(vals), *vals)

    def _fn_averageifs(self, args):
        target = list(flatten([args[0]])) if args else []
        matches = self._criteria_matches(args, 1)
        vals = [
            num(target[index]) for index, matched in enumerate(matches)
            if matched and index < len(target)
            and isinstance(target[index], (int, float))
            and not isinstance(target[index], bool)
        ]
        return sum(vals) / len(vals) if vals else ExcelError("#DIV/0!")

    def _fn_countifs(self, args):
        return float(sum(self._criteria_matches(args, 0)))

    def _fn_npv(self, args):
        rate = num(args[0]) if args else 0.0
        vals = numbers(args[1:])
        if rate == -1.0:
            return ExcelError("#DIV/0!")
        try:
            return sum(value / (1.0 + rate) ** period
                       for period, value in enumerate(vals, 1))
        except (OverflowError, ZeroDivisionError, ValueError):
            return ExcelError("#NUM!")

    def _fn_rri(self, args):
        periods = num(args[0])
        present = num(args[1])
        future = num(args[2])
        if periods == 0 or present == 0 or future / present < 0:
            return ExcelError("#NUM!")
        return _power(future / present, 1.0 / periods) - 1.0

    _fn__xlfn_rri = _fn_rri

    def _fn_sumproduct(self, args):
        arrays = [list(flatten([arg])) for arg in args]
        if not arrays:
            return 0.0
        width = min(len(array) for array in arrays)
        vals = [
            math.prod(num(array[index]) for array in arrays)
            for index in range(width)
        ]
        return snap(sum(vals), *vals)

    def _fn_forecast(self, args):
        x = num(args[0])
        known_y = numbers([args[1]]) if len(args) > 1 else []
        known_x = numbers([args[2]]) if len(args) > 2 else []
        if not known_y or len(known_y) != len(known_x):
            return ExcelError("#N/A")
        mean_x = sum(known_x) / len(known_x)
        mean_y = sum(known_y) / len(known_y)
        denominator = sum((value - mean_x) ** 2 for value in known_x)
        if denominator == 0:
            return ExcelError("#DIV/0!")
        slope = sum(
            (x_value - mean_x) * (y_value - mean_y)
            for x_value, y_value in zip(known_x, known_y)
        ) / denominator
        return mean_y + slope * (x - mean_x)

    _fn_forecast_linear = _fn_forecast
    _fn__xlfn_forecast_linear = _fn_forecast

    def _fn_len(self, args):
        return float(len(_text(args[0])))

    def _fn_upper(self, args):
        return _text(args[0]).upper()

    def _fn_proper(self, args):
        return _text(args[0]).title()

    def _fn_text(self, args):
        value = args[0]
        pattern = _text(args[1]).strip() if len(args) > 1 else ""
        lower = pattern.lower()
        if any(token in lower for token in ("yy", "mm", "dd")):
            stamp = EPOCH + timedelta(days=num(value))
            replacements = (
                ("yyyy", f"{stamp.year:04d}"), ("yy", f"{stamp.year % 100:02d}"),
                ("mmmm", stamp.strftime("%B")), ("mmm", stamp.strftime("%b")),
                ("mm", f"{stamp.month:02d}"), ("dd", f"{stamp.day:02d}"),
            )
            text = pattern
            for token, rendered in replacements:
                text = re.sub(token, rendered, text, flags=re.IGNORECASE)
            return text
        percent = "%" in pattern
        scaled = num(value) * (100.0 if percent else 1.0)
        decimals = len(pattern.split(".", 1)[1].split("%", 1)[0]) if "." in pattern else 0
        grouping = "," in pattern.split(".", 1)[0]
        rendered = f"{scaled:,.{decimals}f}" if grouping else f"{scaled:.{decimals}f}"
        return rendered + ("%" if percent else "")

    def _fn_right(self, args):
        count = int(num(args[1])) if len(args) > 1 else 1
        return _text(args[0])[-count:] if count else ""

    def _fn_mid(self, args):
        start = int(num(args[1]))
        count = int(num(args[2]))
        return _text(args[0])[start - 1: start - 1 + count]

    def _fn_find(self, args):
        pos = _text(args[1]).find(_text(args[0]))
        return float(pos + 1) if pos >= 0 else ExcelError("#VALUE!")

    def _fn_xlookup(self, args):
        keys = list(flatten([args[1]]))
        pool = list(flatten([args[2]])) if len(args) > 2 else keys
        for i, key in enumerate(keys):
            if _equal(key, args[0]) and i < len(pool):
                return pool[i]
        return args[3] if len(args) > 3 and args[3] is not None else ExcelError("#N/A")

    _fn__xlfn_xlookup = _fn_xlookup

    def _fn_match(self, args):
        lookup = args[0]
        pool = list(flatten([args[1]])) if len(args) > 1 else []
        mode = int(num(args[2], 1.0)) if len(args) > 2 and args[2] is not None else 1
        if mode == 0:
            for index, value in enumerate(pool):
                if _equal(value, lookup):
                    return float(index + 1)
            return ExcelError("#N/A")
        candidates = [
            (index, value) for index, value in enumerate(pool)
            if isinstance(value, (int, float))
            and ((mode > 0 and value <= num(lookup))
                 or (mode < 0 and value >= num(lookup)))
        ]
        if not candidates:
            return ExcelError("#N/A")
        selected = max(candidates, key=lambda pair: pair[1]) if mode > 0 else \
            min(candidates, key=lambda pair: pair[1])
        return float(selected[0] + 1)

    def _fn_vlookup(self, args):
        lookup = args[0]
        table = args[1] if len(args) > 1 else []
        values = list(flatten([table]))
        column = int(num(args[2])) if len(args) > 2 else 1
        exact = len(args) > 3 and not truthy(args[3])
        cols = table.cols if isinstance(table, RangeValues) else max(column, 2)
        if column < 1 or column > cols:
            return ExcelError("#REF!")
        rows = [values[index:index + cols] for index in range(0, len(values), cols)]
        selected = None
        for row in rows:
            if len(row) < column:
                continue
            if _equal(row[0], lookup):
                return row[column - 1]
            if not exact and isinstance(row[0], (int, float)) and row[0] <= num(lookup):
                selected = row
        return selected[column - 1] if selected is not None else ExcelError("#N/A")

    def _fn_hlookup(self, args):
        lookup = args[0]
        table = args[1] if len(args) > 1 else []
        values = list(flatten([table]))
        row = int(num(args[2])) if len(args) > 2 else 1
        exact = len(args) > 3 and not truthy(args[3])
        cols = table.cols if isinstance(table, RangeValues) else len(values)
        rows = table.rows if isinstance(table, RangeValues) else 1
        if row < 1 or row > rows or cols < 1:
            return ExcelError("#REF!")
        header = values[:cols]
        selected = None
        for index, value in enumerate(header):
            if _equal(value, lookup):
                selected = index
                break
            if not exact and isinstance(value, (int, float)) and value <= num(lookup):
                selected = index
        if selected is None:
            return ExcelError("#N/A")
        index = (row - 1) * cols + selected
        return values[index] if index < len(values) else ExcelError("#REF!")

    def _fn_transpose(self, args):
        source = args[0] if args else []
        values = list(flatten([source]))
        if not isinstance(source, RangeValues):
            return RangeValues(values, len(values), 1)
        matrix = [
            values[index:index + source.cols]
            for index in range(0, len(values), source.cols)
        ]
        transposed = [
            matrix[row][col]
            for col in range(source.cols)
            for row in range(min(source.rows, len(matrix)))
        ]
        return RangeValues(transposed, source.cols, source.rows)

    def _fn__xlfn_vstack(self, args):
        values = []
        cols = 1
        rows = 0
        for arg in args:
            chunk = list(flatten([arg]))
            values.extend(chunk)
            if isinstance(arg, RangeValues):
                cols = max(cols, arg.cols)
                rows += arg.rows
            else:
                rows += max(len(chunk), 1)
        return RangeValues(values, rows, cols)

    def _fn_irr(self, args):
        return _irr(numbers([args[0]]))

    def _fn_xirr(self, args):
        # Blanks are kept as zeros on both sides rather than dropped. A blank date
        # is serial 0, which discounts its amount over a century -- negligible on
        # its own, but so is everything else once the other flows are discounted
        # that far, and Excel's answer depends on it.
        flows = list(flatten([args[0]]))
        dates = list(flatten([args[1]]))
        if not flows or not dates:
            return ExcelError("#NUM!")
        pairs = [(num(v, 0.0), num(d, 0.0)) for v, d in zip(flows, dates)]
        return _xirr([p[0] for p in pairs], [p[1] for p in pairs])

    def _fn_xnpv(self, args):
        rate = num(args[0])
        flows = list(flatten([args[1]])) if len(args) > 1 else []
        dates = list(flatten([args[2]])) if len(args) > 2 else []
        if not flows or len(flows) != len(dates) or rate <= -1.0:
            return ExcelError("#NUM!")
        start = num(dates[0], 0.0)
        try:
            return sum(
                num(value, 0.0) * (1.0 + rate) ** (-(num(date, 0.0) - start) / 365.0)
                for value, date in zip(flows, dates)
            )
        except (OverflowError, ZeroDivisionError, ValueError):
            return ExcelError("#NUM!")

    def _fn_mirr(self, args):
        flows = [
            num(v) for v in flatten([args[0]])
            if isinstance(v, (int, float)) and not isinstance(v, bool)
        ]
        finance_rate = num(args[1]) if len(args) > 1 else 0.0
        reinvest_rate = num(args[2]) if len(args) > 2 else 0.0
        n = len(flows)
        npv_positive = sum(
            v / (1.0 + reinvest_rate) ** i for i, v in enumerate(flows) if v > 0
        )
        npv_negative = sum(
            v / (1.0 + finance_rate) ** i for i, v in enumerate(flows) if v < 0
        )
        if n < 2 or npv_positive == 0 or npv_negative == 0:
            return ExcelError("#DIV/0!")
        try:
            ratio = npv_positive / -npv_negative
            return ratio ** (1.0 / (n - 1)) * (1.0 + reinvest_rate) - 1.0
        except (OverflowError, ZeroDivisionError, ValueError):
            return ExcelError("#NUM!")

    def _fn_pmt(self, args):
        fv = num(args[3]) if len(args) > 3 else 0.0
        ptype = 1 if len(args) > 4 and truthy(args[4]) else 0
        return _pmt_value(num(args[0]), num(args[1]), num(args[2]), fv, ptype)

    def _fn_ipmt(self, args):
        fv = num(args[4]) if len(args) > 4 else 0.0
        ptype = 1 if len(args) > 5 and truthy(args[5]) else 0
        return _ipmt_value(
            num(args[0]), int(num(args[1])), num(args[2]), num(args[3]), fv, ptype)

    def _fn_ppmt(self, args):
        fv = num(args[4]) if len(args) > 4 else 0.0
        ptype = 1 if len(args) > 5 and truthy(args[5]) else 0
        rate, per = num(args[0]), int(num(args[1]))
        nper, pv = num(args[2]), num(args[3])
        payment = _pmt_value(rate, nper, pv, fv, ptype)
        interest = _ipmt_value(rate, per, nper, pv, fv, ptype)
        if isinstance(payment, ExcelError):
            return payment
        if isinstance(interest, ExcelError):
            return interest
        return payment - interest

    def _fn_pv(self, args):
        rate, nper, pmt = num(args[0]), num(args[1]), num(args[2])
        fv = num(args[3]) if len(args) > 3 else 0.0
        ptype = 1 if len(args) > 4 and truthy(args[4]) else 0
        if rate == 0.0:
            return -(fv + pmt * nper)
        try:
            growth = (1.0 + rate) ** nper
        except (OverflowError, ZeroDivisionError, ValueError):
            return ExcelError("#NUM!")
        return -(fv + pmt * (1.0 + rate * ptype) * (growth - 1.0) / rate) / growth

    def _fn_fv(self, args):
        rate, nper, pmt = num(args[0]), num(args[1]), num(args[2])
        pv = num(args[3]) if len(args) > 3 else 0.0
        ptype = 1 if len(args) > 4 and truthy(args[4]) else 0
        try:
            return _fv_value(rate, nper, pmt, pv, ptype)
        except (OverflowError, ZeroDivisionError, ValueError):
            return ExcelError("#NUM!")

    def _fn_nper(self, args):
        rate, pmt, pv = num(args[0]), num(args[1]), num(args[2])
        fv = num(args[3]) if len(args) > 3 else 0.0
        ptype = 1 if len(args) > 4 and truthy(args[4]) else 0
        if rate == 0.0:
            return ExcelError("#DIV/0!") if pmt == 0 else -(pv + fv) / pmt
        adjusted = pmt * (1.0 + rate * ptype) / rate
        denominator = adjusted + pv
        if denominator == 0:
            return ExcelError("#DIV/0!")
        ratio = (adjusted - fv) / denominator
        if ratio <= 0 or rate <= -1.0:
            return ExcelError("#NUM!")
        return math.log(ratio) / math.log(1.0 + rate)

    def _fn_rate(self, args):
        nper, pmt, pv = num(args[0]), num(args[1]), num(args[2])
        fv = num(args[3]) if len(args) > 3 else 0.0
        ptype = 1 if len(args) > 4 and truthy(args[4]) else 0
        guess = num(args[5], 0.1) if len(args) > 5 else 0.1
        return _rate_value(nper, pmt, pv, fv, ptype, guess)

    def _fn_cumipmt(self, args):
        rate, nper, pv = num(args[0]), num(args[1]), num(args[2])
        start, end = int(num(args[3])), int(num(args[4]))
        ptype = 1 if len(args) > 5 and truthy(args[5]) else 0
        if rate <= 0 or nper <= 0 or pv <= 0 or start < 1 or end < start or end > nper:
            return ExcelError("#NUM!")
        total = 0.0
        for per in range(start, end + 1):
            interest = _ipmt_value(rate, per, nper, pv, 0.0, ptype)
            if isinstance(interest, ExcelError):
                return interest
            total += interest
        return total

    def _fn_cumprinc(self, args):
        rate, nper, pv = num(args[0]), num(args[1]), num(args[2])
        start, end = int(num(args[3])), int(num(args[4]))
        ptype = 1 if len(args) > 5 and truthy(args[5]) else 0
        if rate <= 0 or nper <= 0 or pv <= 0 or start < 1 or end < start or end > nper:
            return ExcelError("#NUM!")
        payment = _pmt_value(rate, nper, pv, 0.0, ptype)
        if isinstance(payment, ExcelError):
            return payment
        total = 0.0
        for per in range(start, end + 1):
            interest = _ipmt_value(rate, per, nper, pv, 0.0, ptype)
            if isinstance(interest, ExcelError):
                return interest
            total += payment - interest
        return total

    def _fn_mod(self, args):
        divisor = num(args[1]) if len(args) > 1 else 0.0
        if divisor == 0:
            return ExcelError("#DIV/0!")
        return num(args[0]) % divisor

    def _fn_networkdays(self, args):
        start, end = int(num(args[0])), int(num(args[1]))
        holidays = set()
        if len(args) > 2 and args[2] is not None:
            for value in flatten([args[2]]):
                serial = num(value, math.nan)
                if not math.isnan(serial):
                    holidays.add(int(serial))
        sign = 1
        if start > end:
            start, end, sign = end, start, -1
        count = sum(
            1 for serial in range(start, end + 1)
            if (EPOCH + timedelta(days=serial)).weekday() < 5
            and serial not in holidays
        )
        return float(sign * count)

    def _fn_yearfrac(self, args):
        start_serial = num(args[0])
        end_serial = num(args[1]) if len(args) > 1 else 0.0
        basis = int(num(args[2])) if len(args) > 2 and args[2] is not None else 0
        if start_serial > end_serial:
            start_serial, end_serial = end_serial, start_serial
        start = EPOCH + timedelta(days=int(start_serial))
        end = EPOCH + timedelta(days=int(end_serial))
        actual_days = int(end_serial) - int(start_serial)
        if basis == 0:
            return _day_count_30_360(start, end, european=False) / 360.0
        if basis == 4:
            return _day_count_30_360(start, end, european=True) / 360.0
        if basis == 2:
            return actual_days / 360.0
        if basis == 3:
            return actual_days / 365.0
        if basis == 1:
            def leap(year):
                return year % 4 == 0 and (year % 100 != 0 or year % 400 == 0)
            if start.year == end.year:
                denominator = 366.0 if leap(start.year) else 365.0
            elif end <= start.replace(year=start.year + 1, day=min(start.day, 28)) \
                    or (end - start).days <= 366:
                feb29_hit = any(
                    leap(year)
                    and start <= datetime(year, 2, 29) <= end
                    for year in (start.year, end.year)
                )
                denominator = 366.0 if feb29_hit else 365.0
            else:
                years = range(start.year, end.year + 1)
                denominator = sum(366.0 if leap(y) else 365.0 for y in years) / len(years)
            return actual_days / denominator
        return ExcelError("#NUM!")

    def _fn_norm_dist(self, args):
        x, mean, sd = num(args[0]), num(args[1]), num(args[2])
        if sd <= 0:
            return ExcelError("#NUM!")
        cumulative = truthy(args[3]) if len(args) > 3 else True
        z = (x - mean) / sd
        return _norm_cdf(z) if cumulative else _norm_pdf(z) / sd

    _fn_normdist = _fn_norm_dist

    def _fn_norm_s_dist(self, args):
        # Legacy NORMSDIST takes no flag and is always cumulative; NORM.S.DIST
        # passes one, so a missing second argument defaults to cumulative.
        z = num(args[0])
        cumulative = truthy(args[1]) if len(args) > 1 and args[1] is not None else True
        return _norm_cdf(z) if cumulative else _norm_pdf(z)

    _fn_normsdist = _fn_norm_s_dist

    def _fn_norm_inv(self, args):
        p, mean, sd = num(args[0]), num(args[1]), num(args[2])
        if sd <= 0:
            return ExcelError("#NUM!")
        z = _norm_s_inv(p)
        return z if isinstance(z, ExcelError) else mean + sd * z

    _fn_norminv = _fn_norm_inv

    def _fn_norm_s_inv(self, args):
        return _norm_s_inv(num(args[0]))

    _fn_normsinv = _fn_norm_s_inv

    def _fn_product(self, args):
        vals = numbers(args)
        return math.prod(vals) if vals else 0.0

    def _fn_power(self, args):
        return _power(num(args[0]), num(args[1]) if len(args) > 1 else 0.0)

    def _fn_exp(self, args):
        try:
            return math.exp(num(args[0]))
        except OverflowError:
            return ExcelError("#NUM!")

    def _fn_ln(self, args):
        value = num(args[0])
        return math.log(value) if value > 0 else ExcelError("#NUM!")

    def _fn_concatenate(self, args):
        return "".join(_text(value) for value in flatten(args))

    # -- census-driven long tail -------------------------------------------
    def _fn_averageif(self, args):
        pool = list(flatten([args[0]]))
        target = list(flatten([args[2]])) if len(args) > 2 and args[2] is not None else pool
        vals = [
            num(target[i]) for i, value in enumerate(pool)
            if _matches(value, args[1]) and i < len(target)
            and isinstance(target[i], (int, float)) and not isinstance(target[i], bool)
        ]
        return sum(vals) / len(vals) if vals else ExcelError("#DIV/0!")

    def _fn_maxifs(self, args):
        target = list(flatten([args[0]])) if args else []
        matches = self._criteria_matches(args, 1)
        vals = [
            num(target[i]) for i, matched in enumerate(matches)
            if matched and i < len(target)
            and isinstance(target[i], (int, float)) and not isinstance(target[i], bool)
        ]
        return max(vals) if vals else 0.0

    def _fn_minifs(self, args):
        target = list(flatten([args[0]])) if args else []
        matches = self._criteria_matches(args, 1)
        vals = [
            num(target[i]) for i, matched in enumerate(matches)
            if matched and i < len(target)
            and isinstance(target[i], (int, float)) and not isinstance(target[i], bool)
        ]
        return min(vals) if vals else 0.0

    def _fn_days360(self, args):
        start = EPOCH + timedelta(days=int(num(args[0])))
        end = EPOCH + timedelta(days=int(num(args[1] if len(args) > 1 else None)))
        european = len(args) > 2 and truthy(args[2])
        return float(_day_count_30_360(start, end, european))

    def _fn_lookup(self, args):
        lookup = args[0]
        table = args[1] if len(args) > 1 else []
        if isinstance(table, RangeValues) and table.rows > 1 and table.cols > 1:
            # Array form: search the longer edge, return the opposite edge.
            values = list(table)
            if table.cols >= table.rows:
                keys = values[: table.cols]
                results = values[(table.rows - 1) * table.cols:]
            else:
                keys = values[:: table.cols]
                results = values[table.cols - 1:: table.cols]
        else:
            keys = list(flatten([table]))
            results = list(flatten([args[2]])) if len(args) > 2 and args[2] is not None else keys
        selected = None
        for index, key in enumerate(keys):
            if _equal(key, lookup):
                selected = index
                break
            if isinstance(key, (int, float)) and not isinstance(key, bool) \
                    and key <= num(lookup):
                selected = index
        if selected is None or selected >= len(results):
            return ExcelError("#N/A")
        return results[selected]

    def _fn_subtotal(self, args):
        # Hidden-row variants (1xx) computed over everything: the graph has no
        # visibility data, and models that hide rows fail verification honestly.
        code = int(num(args[0])) % 100
        rest = args[1:]
        if code == 1:
            return self._fn_average(rest)
        if code == 2:
            return self._fn_count(rest)
        if code == 3:
            values = list(flatten(rest))
            return float(sum(v is not None and v != "" for v in values))
        if code == 4:
            return self._fn_max(rest)
        if code == 5:
            return self._fn_min(rest)
        if code == 6:
            return self._fn_product(rest)
        if code == 9:
            return self._fn_sum(rest)
        return Unresolved(f"unsupported:SUBTOTAL({code})")

    def _fn_int(self, args):
        return float(math.floor(num(args[0])))

    def _fn_n(self, args):
        value = self._scalar(args[0] if args else None)
        if isinstance(value, bool):
            return 1.0 if value else 0.0
        if isinstance(value, (int, float)):
            return float(value)
        return 0.0

    def _fn_now(self, args):
        return _datetime_to_serial(datetime.now())

    def _fn_workday(self, args):
        serial = int(num(args[0]))
        remaining = int(num(args[1] if len(args) > 1 else None))
        holidays = set()
        if len(args) > 2 and args[2] is not None:
            for value in flatten([args[2]]):
                day = num(value, math.nan)
                if not math.isnan(day):
                    holidays.add(int(day))
        step = 1 if remaining >= 0 else -1
        remaining = abs(remaining)
        while remaining:
            serial += step
            if (EPOCH + timedelta(days=serial)).weekday() < 5 and serial not in holidays:
                remaining -= 1
        return float(serial)

    def _fn_weekday(self, args):
        monday0 = (EPOCH + timedelta(days=int(num(args[0])))).weekday()
        mode = int(num(args[1], 1.0)) if len(args) > 1 and args[1] is not None else 1
        if mode == 1:
            return float((monday0 + 1) % 7 + 1)
        if mode == 2:
            return float(monday0 + 1)
        if mode == 3:
            return float(monday0)
        return ExcelError("#NUM!")

    def _fn_datedif(self, args):
        start = EPOCH + timedelta(days=int(num(args[0])))
        end = EPOCH + timedelta(days=int(num(args[1] if len(args) > 1 else None)))
        unit = _text(args[2] if len(args) > 2 else "").strip().upper()
        if end < start:
            return ExcelError("#NUM!")
        months = (end.year - start.year) * 12 + end.month - start.month
        if end.day < start.day:
            months -= 1
        if unit == "D":
            return float((end - start).days)
        if unit == "Y":
            return float(months // 12)
        if unit == "M":
            return float(months)
        if unit == "YM":
            return float(months % 12)
        if unit == "MD":
            anchor = _edate((start - EPOCH).days, months)
            return float((end - EPOCH).days - anchor)
        if unit == "YD":
            anchor = _edate((start - EPOCH).days, (months // 12) * 12)
            return float((end - EPOCH).days - anchor)
        return ExcelError("#NUM!")

    def _fn_rows(self, args):
        source = args[0] if args else None
        return float(source.rows) if isinstance(source, RangeValues) else 1.0

    def _fn_columns(self, args):
        source = args[0] if args else None
        return float(source.cols) if isinstance(source, RangeValues) else 1.0

    def _fn_countblank(self, args):
        values = list(flatten(args))
        return float(sum(v is None or v == "" for v in values))

    def _fn_mina(self, args):
        vals = [num(v) for v in flatten(args) if v is not None]
        return min(vals) if vals else 0.0

    def _fn_maxa(self, args):
        vals = [num(v) for v in flatten(args) if v is not None]
        return max(vals) if vals else 0.0

    def _fn_small(self, args):
        vals = sorted(numbers([args[0]]))
        k = int(num(args[1])) if len(args) > 1 else 1
        return vals[k - 1] if 1 <= k <= len(vals) else ExcelError("#NUM!")

    def _fn_large(self, args):
        vals = sorted(numbers([args[0]]), reverse=True)
        k = int(num(args[1])) if len(args) > 1 else 1
        return vals[k - 1] if 1 <= k <= len(vals) else ExcelError("#NUM!")

    def _fn_quartile(self, args):
        vals = sorted(numbers([args[0]]))
        quart = int(num(args[1])) if len(args) > 1 else 0
        if not vals or not 0 <= quart <= 4:
            return ExcelError("#NUM!")
        position = quart / 4.0 * (len(vals) - 1)
        low = int(math.floor(position))
        frac = position - low
        if low + 1 < len(vals):
            return vals[low] + frac * (vals[low + 1] - vals[low])
        return vals[low]

    _fn_quartile_inc = _fn_quartile

    def _fn_gcd(self, args):
        vals = [int(num(v)) for v in flatten(args)
                if isinstance(v, (int, float)) and not isinstance(v, bool)]
        if not vals or any(v < 0 for v in vals):
            return ExcelError("#NUM!")
        out = 0
        for v in vals:
            out = math.gcd(out, v)
        return float(out)

    def _fn_slope(self, args):
        known_y = numbers([args[0]]) if args else []
        known_x = numbers([args[1]]) if len(args) > 1 else []
        if not known_y or len(known_y) != len(known_x):
            return ExcelError("#N/A")
        mean_x = sum(known_x) / len(known_x)
        mean_y = sum(known_y) / len(known_y)
        denominator = sum((x - mean_x) ** 2 for x in known_x)
        if denominator == 0:
            return ExcelError("#DIV/0!")
        return sum((x - mean_x) * (y - mean_y)
                   for x, y in zip(known_x, known_y)) / denominator

    def _fn_rsq(self, args):
        known_y = numbers([args[0]]) if args else []
        known_x = numbers([args[1]]) if len(args) > 1 else []
        if not known_y or len(known_y) != len(known_x):
            return ExcelError("#N/A")
        mean_x = sum(known_x) / len(known_x)
        mean_y = sum(known_y) / len(known_y)
        var_x = sum((x - mean_x) ** 2 for x in known_x)
        var_y = sum((y - mean_y) ** 2 for y in known_y)
        if var_x == 0 or var_y == 0:
            return ExcelError("#DIV/0!")
        cov = sum((x - mean_x) * (y - mean_y) for x, y in zip(known_x, known_y))
        return cov * cov / (var_x * var_y)

    def _fn_avedev(self, args):
        vals = numbers(args)
        if not vals:
            return ExcelError("#NUM!")
        mean = sum(vals) / len(vals)
        return sum(abs(v - mean) for v in vals) / len(vals)

    def _fn_trim(self, args):
        return re.sub(r" +", " ", _text(args[0])).strip(" ")

    def _fn_left(self, args):
        count = int(num(args[1])) if len(args) > 1 and args[1] is not None else 1
        return _text(args[0])[:count] if count >= 0 else ExcelError("#VALUE!")

    def _fn_lower(self, args):
        return _text(args[0]).lower()

    def _fn_substitute(self, args):
        text = _text(args[0])
        old = _text(args[1] if len(args) > 1 else "")
        new = _text(args[2] if len(args) > 2 else "")
        if not old:
            return text
        if len(args) > 3 and args[3] is not None:
            instance = int(num(args[3]))
            if instance < 1:
                return ExcelError("#VALUE!")
            parts = text.split(old)
            if instance >= len(parts):
                return text
            return old.join(parts[:instance]) + new + old.join(parts[instance:])
        return text.replace(old, new)

    def _fn_search(self, args):
        start = int(num(args[2], 1.0)) if len(args) > 2 and args[2] is not None else 1
        haystack = _text(args[1] if len(args) > 1 else "").lower()
        position = haystack.find(_text(args[0]).lower(), max(start - 1, 0))
        return float(position + 1) if position >= 0 else ExcelError("#VALUE!")

    def _fn_fixed(self, args):
        decimals = int(num(args[1], 2.0)) if len(args) > 1 and args[1] is not None else 2
        no_commas = len(args) > 2 and truthy(args[2])
        value = round(num(args[0]), decimals)
        rendered = f"{value:.{max(decimals, 0)}f}" if no_commas else \
            f"{value:,.{max(decimals, 0)}f}"
        return rendered

    def _fn_textjoin(self, args):
        delimiter = _text(args[0] if args else "")
        ignore_empty = truthy(args[1]) if len(args) > 1 else True
        parts = [_text(v) for v in flatten(args[2:])]
        if ignore_empty:
            parts = [p for p in parts if p != ""]
        return delimiter.join(parts)

    def _fn_concat(self, args):
        return self._fn_concatenate(args)

    def _fn_ifna(self, args):
        value = args[0] if args else None
        if isinstance(value, ExcelError) and value.code == "#N/A":
            fallback = args[1] if len(args) > 1 else None
            return 0.0 if fallback is None else fallback
        return value

    @staticmethod
    def _scalar(value):
        """Implicit intersection for predicates handed a range."""
        if isinstance(value, RangeValues):
            return value[0] if value else None
        return value

    def _fn_isnumber(self, args):
        value = self._scalar(args[0] if args else None)
        return isinstance(value, (int, float)) and not isinstance(value, bool)

    def _fn_isblank(self, args):
        return self._scalar(args[0] if args else None) is None

    def _fn_iserr(self, args):
        value = self._scalar(args[0] if args else None)
        return isinstance(value, ExcelError) and value.code != "#N/A"

    def _fn_iserror(self, args):
        return isinstance(self._scalar(args[0] if args else None), ExcelError)

    def _fn_isna(self, args):
        value = self._scalar(args[0] if args else None)
        return isinstance(value, ExcelError) and value.code == "#N/A"

    def _cell_info(self, node):
        """Only ``CELL("filename", ...)`` appears here, feeding sheet-title cells."""
        args = self._args(node)
        kind = _text(args[0]).strip().lower() if args else ""
        if kind != "filename":
            return Unresolved("unsupported:CELL")
        target = node.sheet
        for edge in self.graph.in_edges.get(node.id, ()):
            ref = self.graph.nodes.get(edge.source)
            if ref is not None and ref.is_cell:
                target = ref.sheet
                break
        return f"[{self.graph.wb}.xlsx]{target}"

    def _record_resolved_targets(self, node, targets):
        if node.kind != "op" or not node.owner:
            return
        owner = self._current_cell or node.owner
        self._resolved_targets.setdefault(owner, set()).update(targets)
        self._resolved_operation_targets.setdefault(node.id, set()).update(targets)

    def _indirect_targets(self, node):
        """Resolve an A1-style INDIRECT string without consulting saved values."""
        reference = self._arg(node, 0)
        if is_bad(reference):
            return reference
        if not isinstance(reference, str):
            return ExcelError("#REF!")
        declared = int(node.arity) if str(node.arity).isdigit() else 1
        if declared > 1:
            a1_style = self._arg(node, 1)
            if a1_style is not None and not truthy(a1_style):
                return Unresolved("unsupported:INDIRECT-R1C1")
        match = A1_REFERENCE_RE.match(reference)
        if not match:
            # Legacy/static ASTs may already carry a resolved named reference
            # in an extra argument slot. Retain that path without reading its
            # value merely to discover its coordinates.
            attached = self._reference_arg(node, declared, read=False)
            if attached.refs:
                return list(attached.refs)
            return ExcelError("#REF!")
        sheet = match.group("quoted") or match.group("bare") or node.sheet
        sheet = sheet.replace("''", "'").strip()
        start = f"{match.group('col0')}{match.group('row0')}"
        end_col, end_row = match.group("col1"), match.group("row1")
        if end_col is None:
            return [f"{sheet}!{start}"]
        return range_members(sheet, f"{start}:{end_col}{end_row}")

    def _indirect(self, node):
        """Resolve and read the reference text computed by INDIRECT."""
        targets = self._indirect_targets(node)
        if is_bad(targets):
            return targets
        self._record_resolved_targets(node, targets)
        values = [self._read(target) for target in targets]
        if len(targets) == 1:
            return values[0]
        coordinates = [_split(target) for target in targets]
        rows = len({row for _, row, _ in coordinates}) or 1
        cols = len({col for _, _, col in coordinates}) or 1
        return RangeValues(values, rows, cols, refs=targets)

    def _row(self, node):
        """Excel ROW, using the referenced range or the formula owner's row."""
        incoming = sorted(self.graph.in_edges.get(node.id, ()), key=lambda edge: edge.arg_index)
        if incoming:
            source = self.graph.nodes.get(incoming[0].source)
            if source is not None and source.row is not None:
                return float(source.row)
        owner = self.graph.nodes.get(node.owner)
        return float(owner.row) if owner is not None and owner.row is not None else ExcelError("#REF!")

    def _counta(self, node):
        """Count nonblank values, recovering a range omitted by older AST graphs."""
        args = self._args(node)
        values = list(flatten(args))
        if not values or all(value is None for value in values):
            match = A1_RANGE_RE.search(node.expr or "")
            if match:
                sheet = (match.group("quoted") or match.group("bare") or node.sheet)
                sheet = sheet.replace("''", "'").strip()
                span = (
                    f"{match.group('col0')}{match.group('row0')}:"
                    f"{match.group('col1')}{match.group('row1')}"
                )
                values = [self._read(cid) for cid in range_members(sheet, span)]
        bad = next((value for value in values if isinstance(value, Unresolved)), None)
        if bad is not None:
            return bad
        return float(sum(value is not None and value != "" for value in values))

    def _offset(self, node):
        """Resolve a dynamic reference, including optional range dimensions."""
        targets = self._offset_targets(node)
        if is_bad(targets):
            return targets
        self._record_resolved_targets(node, targets)
        values = [self._read(target) for target in targets]
        groups = {
            edge.arg_index for edge in self.graph.in_edges.get(node.id, ())
        }
        height = int(num(self._arg(node, 3), 1.0)) if 3 in groups else 1
        width = int(num(self._arg(node, 4), 1.0)) if 4 in groups else 1
        if height == 1 and width == 1:
            return values[0]
        return RangeValues(values, height, width, refs=targets)

    def _offset_targets(self, node):
        """Resolve the cell ids addressed by OFFSET without reading their values."""
        groups: dict = {}
        for edge in self.graph.in_edges.get(node.id, ()):
            groups.setdefault(edge.arg_index, []).append(edge)
        if 0 in groups:
            base_id = groups[0][0].source
            base = self.graph.nodes.get(base_id)
            if base is None or base.row is None:
                return Unresolved("offset-base-not-cell")
            base_sheet, base_row, base_col = base.sheet, base.row, base.col
        else:
            # Older ASTs omitted an OFFSET base when that anchor cell was blank.
            # The coordinate is still a reference origin, not a value dependency.
            match = OFFSET_BASE_RE.match(node.expr or "")
            if not match:
                return Unresolved("offset-no-base")
            base_sheet = match.group("quoted") or match.group("bare") or node.sheet
            base_sheet = base_sheet.replace("''", "'").strip()
            parsed = _split(f"{base_sheet}!{match.group('col')}{match.group('row')}")
            if parsed is None:
                return Unresolved("offset-base-not-cell")
            base_sheet, base_row, base_col = parsed
            base_id = a1(base_sheet, base_row, base_col)
        if self._current_cell is not None:
            self.runtime_address_radj.setdefault(
                self._current_cell, set()
            ).add(base_id)
        rows = self._arg(node, 1) if 1 in groups else 0
        cols = self._arg(node, 2) if 2 in groups else 0
        if is_bad(rows) or is_bad(cols):
            return rows if is_bad(rows) else cols
        target_row = base_row + int(num(rows))
        target_col = base_col + int(num(cols))
        height = self._arg(node, 3) if 3 in groups else 1
        width = self._arg(node, 4) if 4 in groups else 1
        if is_bad(height) or is_bad(width):
            return height if is_bad(height) else width
        height, width = int(num(height, 1.0)), int(num(width, 1.0))
        if height < 1 or width < 1 or target_row < 1 or target_col < 1:
            return ExcelError("#REF!")
        return [
            a1(base_sheet, row, col)
            for row in range(target_row, target_row + height)
            for col in range(target_col, target_col + width)
        ]

    @staticmethod
    def _refs(value):
        return list(value.refs or ()) if isinstance(value, RangeValues) else []

    def _selective_dependencies(self, node, args, result):
        """Value precedents used by a selector, excluding unselected result cells."""
        op = node.op
        specs = self._arg_specs(node)
        sources = set()

        def add_slot(index):
            if index < len(specs):
                sources.update(self._active_spec_sources(specs[index]))

        selected_refs = []
        if op == "INDEX":
            add_slot(1)
            add_slot(2)
            source = args[0] if args else RangeValues([])
            refs = self._refs(source)
            if refs and isinstance(source, RangeValues):
                row = int(num(args[1])) if len(args) > 1 else 1
                has_col = len(args) > 2 and args[2] is not None
                if not has_col and source.rows == 1:
                    row, col = 1, row
                else:
                    col = int(num(args[2])) if has_col else 1
                if row == 0 and col == 0:
                    selected_refs = refs
                elif row == 0 and 1 <= col <= source.cols:
                    selected_refs = refs[col - 1::source.cols]
                elif col == 0 and 1 <= row <= source.rows:
                    start = (row - 1) * source.cols
                    selected_refs = refs[start:start + source.cols]
                elif 1 <= row <= source.rows and 1 <= col <= source.cols:
                    index = (row - 1) * source.cols + col - 1
                    if index < len(refs):
                        selected_refs = [refs[index]]
            elif specs:
                add_slot(0)
        elif op == "MATCH":
            add_slot(0)
            add_slot(2)
            if len(args) > 1:
                refs = self._refs(args[1])
                sources.update(refs)
                if not refs:
                    add_slot(1)
        elif op in ("XLOOKUP", "_XLFN.XLOOKUP"):
            add_slot(0)
            keys = list(flatten([args[1]])) if len(args) > 1 else []
            key_refs = self._refs(args[1]) if len(args) > 1 else []
            sources.update(key_refs)
            if not key_refs:
                add_slot(1)
            selected = next(
                (index for index, key in enumerate(keys) if _equal(key, args[0])),
                None,
            )
            if selected is not None and len(args) > 2:
                result_refs = self._refs(args[2])
                if selected < len(result_refs):
                    selected_refs = [result_refs[selected]]
                elif not result_refs:
                    add_slot(2)
            elif selected is None:
                add_slot(3)
            add_slot(4)
            add_slot(5)
        elif op == "VLOOKUP":
            add_slot(0)
            add_slot(2)
            add_slot(3)
            table = args[1] if len(args) > 1 else RangeValues([])
            refs = self._refs(table)
            values = list(flatten([table]))
            cols = table.cols if isinstance(table, RangeValues) else 1
            column = int(num(args[2])) if len(args) > 2 else 1
            exact = len(args) > 3 and not truthy(args[3])
            selected = None
            for offset in range(0, len(values), max(cols, 1)):
                if offset < len(refs):
                    sources.add(refs[offset])
                key = values[offset]
                if _equal(key, args[0]):
                    selected = offset
                    break
                if not exact and isinstance(key, (int, float)) and key <= num(args[0]):
                    selected = offset
            if selected is not None:
                target = selected + column - 1
                if target < len(refs):
                    selected_refs = [refs[target]]
            if not refs:
                add_slot(1)
        elif op == "HLOOKUP":
            add_slot(0)
            add_slot(2)
            add_slot(3)
            table = args[1] if len(args) > 1 else RangeValues([])
            refs = self._refs(table)
            values = list(flatten([table]))
            cols = table.cols if isinstance(table, RangeValues) else len(values)
            row = int(num(args[2])) if len(args) > 2 else 1
            exact = len(args) > 3 and not truthy(args[3])
            selected = None
            for index, key in enumerate(values[:cols]):
                if index < len(refs):
                    sources.add(refs[index])
                if _equal(key, args[0]):
                    selected = index
                    break
                if not exact and isinstance(key, (int, float)) and key <= num(args[0]):
                    selected = index
            target = (row - 1) * cols + selected if selected is not None else -1
            if 0 <= target < len(refs):
                selected_refs = [refs[target]]
            if not refs:
                add_slot(1)
        elif op == "LOOKUP":
            add_slot(0)
            table = args[1] if len(args) > 1 else RangeValues([])
            refs = self._refs(table)
            sources.update(refs)
            if not refs:
                add_slot(1)
            if len(args) > 2 and args[2] is not None:
                result_refs = self._refs(args[2])
                if isinstance(result, RangeValues):
                    selected_refs = list(result.refs or ())
                elif result_refs:
                    values = list(flatten([table]))
                    selected = None
                    for index, key in enumerate(values):
                        if _equal(key, args[0]):
                            selected = index
                            break
                        if isinstance(key, (int, float)) and key <= num(args[0]):
                            selected = index
                    if selected is not None and selected < len(result_refs):
                        selected_refs = [result_refs[selected]]
                else:
                    add_slot(2)

        sources.update(selected_refs)
        self._record_resolved_targets(node, selected_refs)
        return sources

    def _active_node_sources(self, node_id, seen=None):
        """Cell precedents actually evaluated under the current selector values."""
        seen = set() if seen is None else seen
        if node_id in seen:
            return set()
        seen.add(node_id)
        node = self.graph.nodes.get(node_id)
        if node is None:
            return set()
        if node.is_cell:
            return {node.id}
        if node.kind == "range":
            return set(range_members(node.sheet, node.coordinate or ""))
        if node.kind != "op":
            return set()

        if node.op in ("SUMIF", "SUMIFS"):
            return self._conditional_sources(node)
        if node.op in (
            "VLOOKUP", "HLOOKUP", "XLOOKUP", "_XLFN.XLOOKUP",
            "LOOKUP", "MATCH", "INDEX",
        ):
            selected = self._selective_sources.get(node.id)
            if selected is not None:
                return set(selected)

        specs = self._arg_specs(node)
        indices = list(range(len(specs)))
        if node.op == "OFFSET":
            # The first argument supplies an address origin, not a value. The
            # resolved target below is the actual value dependency.
            indices = list(range(1, len(specs)))
        elif node.op == "IF":
            cond = self._arg(node, 0)
            # Discovery must over-approximate an unresolved selector. Keeping
            # only the condition creates a self-sustaining under-closed cone:
            # neither branch is admitted, so the condition can never become
            # computable. Once the selector has a value this remains exactly
            # branch-specific.
            indices = (
                list(range(len(specs)))
                if is_bad(cond) or cond is None
                else [0, 1 if truthy(cond) else 2]
            )
        elif node.op == "CHOOSE":
            selected = self._arg(node, 0)
            if is_bad(selected) or selected is None:
                indices = list(range(len(specs)))
            else:
                index = int(num(selected))
                indices = [0] + ([index] if 1 <= index < len(specs) else [])
        elif node.op in ("IFS", "_XLFN.IFS"):
            indices = []
            for index in range(0, len(specs) - 1, 2):
                indices.append(index)
                cond = self._arg(node, index)
                if is_bad(cond) or cond is None:
                    # Every remaining condition/value pair can still become
                    # active after proof closure supplies the missing selector.
                    indices.extend(range(index + 1, len(specs)))
                    break
                if truthy(cond):
                    indices.append(index + 1)
                    break
        elif node.op == "IFERROR":
            indices = [0]
            if isinstance(self._arg(node, 0), ExcelError) and len(specs) > 1:
                indices.append(1)

        sources = set()
        for index in indices:
            if index >= len(specs):
                continue
            kind, payload = specs[index]
            if kind == "span":
                sources.update(payload)
            elif kind == "scalar":
                sources.update(self._active_node_sources(payload, seen))
            elif kind == "many":
                for source in payload:
                    sources.update(self._active_node_sources(source, seen))
        if node.op == "OFFSET":
            targets = self._offset_targets(node)
            if isinstance(targets, list):
                sources.update(targets)
                self._record_resolved_targets(node, targets)
        elif node.op == "INDIRECT":
            targets = self._indirect_targets(node)
            if isinstance(targets, list):
                sources.update(targets)
                self._record_resolved_targets(node, targets)
        return sources

    def _active_graph(self, cells):
        """Build source-to-consumer edges after pruning inactive formula branches."""
        adj: dict[str, set] = {}
        radj: dict[str, set] = {}
        for target in cells:
            root = self.graph.root_of(target)
            if root is None:
                continue
            radj[target] = set()
            self._current_cell = target
            for source in self._active_node_sources(root):
                adj.setdefault(source, set()).add(target)
                radj.setdefault(target, set()).add(source)
        return adj, radj

    @staticmethod
    def _cell_key(cell):
        sheet, row, col = _split(cell)
        return str(sheet), int(row), int(col), str(cell)

    def _component_plan(self, cells, active_adj):
        """Return deterministic SCC and topological metadata for an active graph."""
        ordered_cells = sorted(cells, key=self._cell_key)
        ordered_adj = {
            source: tuple(sorted(targets, key=self._cell_key))
            for source, targets in active_adj.items()
        }
        groups = strongly_connected(ordered_cells, ordered_adj)
        groups = [tuple(sorted(group, key=self._cell_key)) for group in groups]
        groups.sort(key=lambda group: self._cell_key(group[0]))
        comp_of = {
            cell: index for index, group in enumerate(groups) for cell in group
        }
        comp_adj = {index: set() for index in range(len(groups))}
        indegree = {index: 0 for index in range(len(groups))}
        for source, targets in active_adj.items():
            source_comp = comp_of.get(source)
            for target in targets:
                target_comp = comp_of.get(target)
                if (
                    source_comp is None
                    or target_comp is None
                    or source_comp == target_comp
                    or target_comp in comp_adj[source_comp]
                ):
                    continue
                comp_adj[source_comp].add(target_comp)
                indegree[target_comp] += 1
        ready = sorted(
            (index for index, count in indegree.items() if count == 0),
            key=lambda index: self._cell_key(groups[index][0]),
        )
        topo = []
        while ready:
            index = ready.pop(0)
            topo.append(index)
            for target in sorted(
                comp_adj[index],
                key=lambda item: self._cell_key(groups[item][0]),
            ):
                indegree[target] -= 1
                if indegree[target] == 0:
                    ready.append(target)
                    ready.sort(key=lambda item: self._cell_key(groups[item][0]))
        cyclic = {
            index
            for index, group in enumerate(groups)
            if len(group) > 1
            or any(cell in active_adj.get(cell, ()) for cell in group)
        }
        return groups, topo, cyclic

    def _evaluate_acyclic(self, groups, topo, cyclic, evaluable, seeded):
        for index in topo:
            if index in cyclic:
                continue
            for cell in groups[index]:
                if cell in evaluable and cell not in seeded:
                    self.values[cell] = self._eval_cell(cell)

    @staticmethod
    def _graph_signature(active_adj):
        return tuple(sorted(
            (source, target)
            for source, targets in active_adj.items()
            for target in targets
        ))

    def _target_signature(self):
        return tuple(
            (target, tuple(sorted(sources)))
            for target, sources in sorted(
                self._resolved_operation_targets.items()
            )
        )

    def _equation_residual(self, members, state=None):
        """Evaluate every equation against one simultaneous endpoint state."""
        saved = {cell: self.values.get(cell) for cell in members}
        if state is not None:
            self.values.update(state)
        changes = []
        errors = []
        unresolved = []
        try:
            for cell in members:
                before = self.values.get(cell)
                after = self._eval_cell(cell)
                if isinstance(after, ExcelError):
                    errors.append({"cell": cell, "value": repr(after)})
                if isinstance(after, Unresolved):
                    unresolved.append({
                        "cell": cell,
                        "value": repr(after),
                        "reason": after.reason,
                    })
                changes.append(self._value_change(before, after))
        finally:
            self.values.update(saved)
        if any(math.isinf(change) for change in changes):
            residual = None
        else:
            residual = max(changes, default=0.0)
        return residual, errors, unresolved

    @staticmethod
    def _bounded_history(history, limit=20):
        if len(history) <= limit:
            return list(history)
        half = limit // 2
        return list(history[:half]) + list(history[-half:])

    def _endpoint_equation_check(self, members, state):
        """Check a cached/recomputed endpoint against the stabilized equations."""
        members = tuple(members)
        if not members or set(members) - set(self._proof_endpoint_values):
            return {"safe": False, "reason": "missing_proof_endpoint_state"}
        if set(members) - set(state):
            return {"safe": False, "reason": "incomplete_cycle_endpoint"}
        if not all(
            isinstance(state[cell], (int, float))
            and not isinstance(state[cell], bool)
            and math.isfinite(float(state[cell]))
            for cell in members
        ):
            return {"safe": False, "reason": "non_numeric_cycle_endpoint"}

        checker = Evaluator(
            self.graph,
            self.cg,
            strict_proof=True,
            calculation=self.calculation,
            proof_outputs=self.proof_outputs,
            proof_scope=self.proof_scope,
        )
        checker.values = dict(self._proof_endpoint_values)
        checker.values.update(state)
        computed = {}
        for cell in members:
            computed[cell] = checker._eval_cell(cell)
        active_adj, active_radj = checker._active_graph(self.cg.info)
        del active_adj

        expected_equations = {
            cell: tuple(sorted(self._proof_active_radj.get(cell, ())))
            for cell in members
        }
        actual_equations = {
            cell: tuple(sorted(active_radj.get(cell, ())))
            for cell in members
        }
        if actual_equations != expected_equations:
            return {
                "safe": False,
                "reason": "cycle_endpoint_equations_changed",
                "expected_equations": expected_equations,
                "actual_equations": actual_equations,
            }
        expected_targets = {
            cell: tuple(sorted(self._proof_resolved_targets.get(cell, ())))
            for cell in members
        }
        actual_targets = {
            cell: tuple(sorted(checker._resolved_targets.get(cell, ())))
            for cell in members
        }
        if actual_targets != expected_targets:
            return {
                "safe": False,
                "reason": "cycle_endpoint_targets_changed",
                "expected_targets": expected_targets,
                "actual_targets": actual_targets,
            }
        expected_operation_targets = {
            operation: tuple(sorted(targets))
            for operation, targets in sorted(
                self._proof_resolved_operation_targets.items()
            )
            if self.graph.nodes.get(operation) is not None
            and self.graph.nodes[operation].owner in members
        }
        actual_operation_targets = {
            operation: tuple(sorted(targets))
            for operation, targets in sorted(
                checker._resolved_operation_targets.items()
            )
            if self.graph.nodes.get(operation) is not None
            and self.graph.nodes[operation].owner in members
        }
        if actual_operation_targets != expected_operation_targets:
            return {
                "safe": False,
                "reason": "cycle_endpoint_operation_targets_changed",
                "expected_operation_targets": expected_operation_targets,
                "actual_operation_targets": actual_operation_targets,
            }

        changes = [
            self._value_change(state[cell], computed[cell]) for cell in members
        ]
        if any(math.isinf(change) for change in changes):
            return {
                "safe": False,
                "reason": "cycle_endpoint_error_or_unresolved",
            }
        return {
            "safe": True,
            "reason": None,
            "residual": max(changes, default=0.0),
            "equations": actual_equations,
        }

    def _affine_cycle_certificate(self, members):
        """Certify a bounded affine SCC ``x = A*x + b``.

        The accepted expression subset is deliberately narrow: scalar
        arithmetic, SUM/AVERAGE, and selectors whose conditions are constants
        with respect to the SCC.  The induced infinity norm ``||A||∞ < 1`` is
        a sound contraction bound, not an empirical convergence guess.
        """
        members = tuple(members)
        if not members:
            return {"certified": False, "reason": "empty_cycle"}
        position = {cell: index for index, cell in enumerate(members)}
        dimension = len(members)
        if not all(
            isinstance(self.values.get(cell), (int, float))
            and not isinstance(self.values.get(cell), bool)
            and math.isfinite(float(self.values[cell]))
            for cell in members
        ):
            return {
                "certified": False,
                "reason": "non_numeric_affine_endpoint",
            }
        branch_guards = []
        unsupported_nodes = []

        def reject(node_id, reason):
            if len(unsupported_nodes) < 20:
                node = self.graph.nodes.get(node_id)
                unsupported_nodes.append({
                    "node": node_id,
                    "kind": getattr(node, "kind", None),
                    "op": getattr(node, "op", None),
                    "reason": reason,
                })
            return None

        def constant(value):
            zeros = [0.0] * dimension
            return (zeros, float(value), list(zeros))

        def combine(left, right, sign=1.0):
            return (
                [
                    left[0][index] + sign * right[0][index]
                    for index in range(dimension)
                ],
                left[1] + sign * right[1],
                [
                    left[2][index] + right[2][index]
                    for index in range(dimension)
                ],
            )

        def scale(value, factor):
            return (
                [coefficient * factor for coefficient in value[0]],
                value[1] * factor,
                [bound * abs(factor) for bound in value[2]],
            )

        def is_constant(value):
            return all(abs(coefficient) <= CONVERGENCE for coefficient in value[0])

        def derive_spec(spec, seen, *, sum_blank_as_zero=False):
            kind, payload = spec
            if kind == "none":
                return []
            if kind == "span":
                sources = payload
            elif kind == "many":
                sources = payload
            elif kind == "scalar":
                source = self.graph.nodes.get(payload)
                if source is not None and source.kind == "range":
                    sources = range_members(
                        source.sheet, source.coordinate or ""
                    )
                else:
                    sources = (payload,)
            else:
                return None
            values = [
                (
                    constant(0.0)
                    if sum_blank_as_zero
                    and (
                        source not in self.graph.nodes
                        or self.values.get(source) in (None, "")
                    )
                    else derive(source, seen)
                )
                for source in sources
            ]
            return None if any(value is None for value in values) else values

        def derive_condition(spec, seen):
            kind, payload = spec
            node = self.graph.nodes.get(payload) if kind == "scalar" else None
            if node is not None and node.kind == "op" and node.op in {
                "=", "<>", "<", "<=", ">", ">=",
            }:
                comparison_specs = self._arg_specs(node)
                if len(comparison_specs) != 2:
                    return None
                sides = [
                    derive_spec(item, seen | {node.id})
                    for item in comparison_specs
                ]
                if any(side is None or len(side) != 1 for side in sides):
                    return None
                left, right = sides[0][0], sides[1][0]
                left_value = sum(
                    left[0][index] * float(self.values[members[index]])
                    for index in range(dimension)
                ) + left[1]
                right_value = sum(
                    right[0][index] * float(self.values[members[index]])
                    for index in range(dimension)
                ) + right[1]
                gap = abs(left_value - right_value)
                sensitivity = sum(
                    abs(left[0][index] - right[0][index])
                    for index in range(dimension)
                )
                if gap <= CONVERGENCE and sensitivity > CONVERGENCE:
                    return None
                if sensitivity > CONVERGENCE:
                    branch_guards.append((left, right, gap))
                return truthy(self._eval_node(node.id))
            values = derive_spec(spec, seen)
            if values is None or len(values) != 1 or not is_constant(values[0]):
                return None
            return truthy(values[0][1])

        def derive(node_id, seen):
            if node_id in seen:
                return reject(node_id, "recursive_ast_path")
            node = self.graph.nodes.get(node_id)
            if node is None:
                return reject(node_id, "missing_node")
            if node.is_cell:
                if node.id in position:
                    coefficients = [0.0] * dimension
                    coefficients[position[node.id]] = 1.0
                    return coefficients, 0.0, [abs(value) for value in coefficients]
                value = self.values.get(node.id)
                if (
                    isinstance(value, (int, float))
                    and math.isfinite(float(value))
                ):
                    return constant(value)
                return reject(node_id, "non_numeric_external_cell")
            if node.kind == "const":
                value = literal(node.value if node.value != "" else node.expr)
                if (
                    isinstance(value, (int, float))
                    and math.isfinite(float(value))
                ):
                    return constant(value)
                return reject(node_id, "non_numeric_constant")
            if node.kind != "op":
                return reject(node_id, "unsupported_node_kind")
            specs = self._arg_specs(node)
            if node.op == "IF" and len(specs) >= 2:
                condition = derive_condition(specs[0], seen | {node_id})
                if condition is None:
                    return reject(node_id, "non_constant_if_selector")
                branch = 1 if condition else 2
                if branch >= len(specs):
                    return constant(0.0)
                selected = derive_spec(specs[branch], seen | {node_id})
                return selected[0] if selected is not None and len(selected) == 1 else None
            if node.op == "CHOOSE" and specs:
                selectors = derive_spec(specs[0], seen | {node_id})
                if (
                    selectors is None
                    or len(selectors) != 1
                    or not is_constant(selectors[0])
                ):
                    return reject(node_id, "non_constant_choose_selector")
                branch = int(num(selectors[0][1]))
                if not 1 <= branch < len(specs):
                    return None
                selected = derive_spec(specs[branch], seen | {node_id})
                return selected[0] if selected is not None and len(selected) == 1 else None
            if node.op in ("IFS", "_XLFN.IFS"):
                for index in range(0, len(specs) - 1, 2):
                    condition = derive_condition(
                        specs[index], seen | {node_id}
                    )
                    if condition is None:
                        return reject(node_id, "non_constant_ifs_selector")
                    if condition:
                        selected = derive_spec(
                            specs[index + 1], seen | {node_id}
                        )
                        return (
                            selected[0]
                            if selected is not None and len(selected) == 1
                            else None
                        )
                return None
            grouped_args = [
                derive_spec(
                    spec,
                    seen | {node_id},
                    # Blank cells have no AST node. Excel SUM treats blank
                    # references as zero, so their absence is affine rather
                    # than an unsupported dependency.
                    sum_blank_as_zero=node.op == "SUM",
                )
                for spec in specs
            ]
            if any(arg is None for arg in grouped_args):
                return reject(node_id, "unsupported_argument")
            args = [value for group in grouped_args for value in group]
            if node.op in ("u+", "u-") and len(args) == 1:
                return scale(args[0], -1.0 if node.op == "u-" else 1.0)
            if node.op in ("SUM", "AVERAGE"):
                total = constant(0.0)
                for arg in args:
                    total = combine(total, arg)
                if node.op == "AVERAGE" and args:
                    total = scale(total, 1.0 / len(args))
                return total
            if node.op in ("MIN", "MAX") and args:
                endpoint = [
                    sum(
                        value[0][index] * float(self.values[members[index]])
                        for index in range(dimension)
                    ) + value[1]
                    for value in args
                ]
                selected_index = (
                    min(range(len(args)), key=endpoint.__getitem__)
                    if node.op == "MIN"
                    else max(range(len(args)), key=endpoint.__getitem__)
                )
                selected = args[selected_index]
                tied = []
                for index, other in enumerate(args):
                    if index == selected_index:
                        continue
                    gap = abs(endpoint[selected_index] - endpoint[index])
                    sensitivity = sum(
                        abs(selected[0][member_index] - other[0][member_index])
                        for member_index in range(dimension)
                    )
                    constant_gap = abs(selected[1] - other[1])
                    if (
                        gap <= CONVERGENCE
                        and (
                            sensitivity > CONVERGENCE
                            or constant_gap > CONVERGENCE
                        )
                    ):
                        tied.append(other)
                        continue
                    if sensitivity > CONVERGENCE or constant_gap > CONVERGENCE:
                        branch_guards.append((selected, other, gap))
                if tied:
                    selected = (
                        selected[0],
                        selected[1],
                        [
                            max(value[2][member_index] for value in [selected, *tied])
                            for member_index in range(dimension)
                        ],
                    )
                return selected
            if any(len(group) != 1 for group in grouped_args):
                return reject(node_id, "non_scalar_arithmetic_argument")
            if args and all(is_constant(arg) for arg in args):
                actual = self._eval_node(node_id)
                if (
                    isinstance(actual, (int, float))
                    and math.isfinite(float(actual))
                ):
                    return constant(actual)
            if len(args) != 2:
                return reject(node_id, "unsupported_operation_arity")
            left, right = args
            if node.op == "+":
                return combine(left, right)
            if node.op == "-":
                return combine(left, right, -1.0)
            if node.op == "*":
                left_constant = is_constant(left)
                right_constant = is_constant(right)
                if not left_constant and not right_constant:
                    return reject(node_id, "bilinear_product")
                if left_constant:
                    return scale(right, left[1])
                return scale(left, right[1])
            if (
                node.op == "/"
                and is_constant(right)
                and abs(right[1]) > CONVERGENCE
            ):
                return scale(left, 1.0 / right[1])
            return reject(node_id, "unsupported_operation")

        matrix = []
        derivative_bounds = []
        constants = []
        for variable in members:
            root = self.graph.root_of(variable)
            if root is None:
                return {"certified": False, "reason": "missing_formula_root"}
            affine = derive(root, set())
            if affine is None:
                return {
                    "certified": False,
                    "reason": "unsupported_affine_structure",
                    "unsupported_nodes": unsupported_nodes,
                }
            matrix.append(affine[0])
            derivative_bounds.append(affine[2])
            constants.append(affine[1])
        if not all(
            math.isfinite(value)
            for row in matrix for value in row
        ) or not all(math.isfinite(value) for value in constants):
            return {"certified": False, "reason": "non_finite_affine_terms"}

        def matrix_multiply(left, right):
            return [
                [
                    sum(left[row][k] * right[k][col] for k in range(dimension))
                    for col in range(dimension)
                ]
                for row in range(dimension)
            ]

        def matrix_norm(value):
            return max(
                sum(abs(coefficient) for coefficient in row)
                for row in value
            )

        identity = [
            [1.0 if row == col else 0.0 for col in range(dimension)]
            for row in range(dimension)
        ]
        power = identity
        prefix_norm_sum = 0.0
        contraction_bound = None
        contraction_power = None
        inverse_norm_bound = None
        for exponent in range(1, min(2 * dimension, 100) + 1):
            prefix_norm_sum += matrix_norm(power)
            power = matrix_multiply(power, derivative_bounds)
            candidate = matrix_norm(power)
            if candidate < 1.0:
                contraction_bound = candidate
                contraction_power = exponent
                inverse_norm_bound = prefix_norm_sum / (1.0 - candidate)
                break
        contractive = contraction_power is not None
        if not contractive:
            return {
                "certified": False,
                "reason": "affine_system_not_bounded_contractive",
                "kind": "affine_system",
                "dimension": dimension,
                "contraction_bound": matrix_norm(derivative_bounds),
                "contractive": False,
            }

        # Banach's theorem gives a unique fixed point and the useful endpoint
        # error bound ||x-x*||∞ <= residual / (1-||A||∞).
        residual = self._equation_residual(members)[0]
        endpoint_error_bound = (
            None
            if residual is None
            else residual * inverse_norm_bound
        )
        if endpoint_error_bound is None:
            return {
                "certified": False,
                "reason": "missing_affine_endpoint_residual",
            }
        for selected, other, gap in branch_guards:
            guard_sensitivity = sum(
                abs(selected[0][index] - other[0][index])
                for index in range(dimension)
            )
            if gap <= guard_sensitivity * endpoint_error_bound:
                return {
                    "certified": False,
                    "reason": "unstable_piecewise_affine_branch",
                    "kind": "piecewise_affine_system",
                    "dimension": dimension,
                    "branch_gap": gap,
                    "branch_error_bound": (
                        guard_sensitivity * endpoint_error_bound
                    ),
                }
        if dimension == 1:
            coefficient = matrix[0][0]
            denominator = 1.0 - coefficient
            return {
                "certified": True,
                "kind": "scalar_affine",
                "coefficient": coefficient,
                "constant": constants[0],
                "fixed_point": constants[0] / denominator,
                "contractive": True,
                "contraction_bound": contraction_bound,
                "contraction_power": contraction_power,
                "endpoint_error_bound": endpoint_error_bound,
            }
        return {
            "certified": True,
            "kind": "affine_system",
            "dimension": dimension,
            "contractive": True,
            "contraction_bound": contraction_bound,
            "contraction_power": contraction_power,
            "inverse_norm_bound": inverse_norm_bound,
            "endpoint_error_bound": endpoint_error_bound,
        }

    # -- driver -----------------------------------------------------------
    def run(self, inputs: set) -> EvalResult:
        """Seed primitive graph values, rebuild formulas, and report the gaps."""
        cells = [
            cell for cell in self.cg.info
            if self.proof_scope is None or cell in self.proof_scope
        ]
        seed_categories = {
            "requested_frontier": set(),
            "primitive_input": set(),
            "primitive_text": set(),
            "label": set(),
            "formula_literal": set(),
            "formula_cache_rejected": set(),
            "requested_ineligible_rejected": set(),
            "implicit_seed_rejected": set(),
        }
        for cid in cells:
            info = self.cg.info[cid]
            requested = cid in inputs
            formula_cache_seed = info.node.kind == "formula" and (
                requested or info.is_literal
            )
            if self.strict_proof:
                if formula_cache_seed:
                    # Formula ``node.value`` is a saved workbook answer. Even
                    # when old frontier logic asks for it, strict proof must
                    # evaluate its AST (or fail) instead of seeding that answer.
                    seed_categories["formula_cache_rejected"].add(cid)
                if requested and info.node.kind in ("input", "label"):
                    self.values[cid] = literal(info.node.value)
                    seed_categories["requested_frontier"].add(cid)
                    seed_categories[
                        "primitive_input" if info.node.kind == "input"
                        else "primitive_text"
                    ].add(cid)
                elif requested:
                    seed_categories["requested_ineligible_rejected"].add(cid)
                elif info.node.kind in ("input", "label") or info.is_literal:
                    seed_categories["implicit_seed_rejected"].add(cid)
                continue
            if requested or info.node.kind in ("input", "label") or info.is_literal:
                self.values[cid] = literal(info.node.value)
                if requested:
                    seed_categories["requested_frontier"].add(cid)
                if info.node.kind == "input":
                    seed_categories["primitive_input"].add(cid)
                if info.node.kind == "label":
                    seed_categories["label"].add(cid)
                if info.is_literal:
                    seed_categories["formula_literal"].add(cid)
        seeded = set(self.values)
        # Excel initializes every iterative formula at zero. Doing this globally
        # for formulas also makes selector evaluation deterministic before the
        # first active dependency graph is built. Strict proof must not apply
        # that initialization to primitive cells: an unauthorized workbook
        # value is unknown, not zero.
        evaluable = {
            cid for cid in cells if self.cg.info[cid].node.kind == "formula"
        }
        for cid in cells:
            if cid not in seeded:
                self.values[cid] = (
                    0.0 if cid in evaluable
                    else Unresolved("unauthorized-primitive")
                    if self.strict_proof
                    else 0.0
                )
        indeterminate: set[str] = set()

        settings = self.calculation
        iterate_setting = getattr(settings, "iterate", None)
        # Preserve the evaluator's historical unknown-metadata behavior while
        # retaining enough provenance for verification to fail closed.
        iterate = True if iterate_setting is None else bool(iterate_setting)
        raw_limit = getattr(settings, "iterate_count", None)
        iteration_limit = (
            int(raw_limit)
            if isinstance(raw_limit, (int, float)) and int(raw_limit) > 0
            else EXCEL_DEFAULT_ITERATIONS
        )
        raw_delta = getattr(settings, "iterate_delta", None)
        iteration_delta = (
            float(raw_delta)
            if isinstance(raw_delta, (int, float)) and float(raw_delta) >= 0
            else EXCEL_DEFAULT_CHANGE
        )

        # First stabilize selectors, dynamic targets, and the active topology.
        # Cyclic components are deliberately not advanced in this phase: their
        # workbook iteration budget starts only after topology discovery.
        previous_graph = None
        previous_targets = None
        topology_stable = False
        targets_stable = False
        topology_history = []
        graph_passes = 0
        active_adj, active_radj = {}, {}
        groups, topo, cyclic = [], [], set()
        for graph_pass in range(1, MAX_ACTIVE_PASSES + 1):
            graph_passes = graph_pass
            self._resolved_targets = {}
            self._resolved_operation_targets = {}
            self.runtime_address_radj = {}
            active_adj, active_radj = self._active_graph(cells)
            groups, topo, cyclic = self._component_plan(cells, active_adj)
            self._evaluate_acyclic(groups, topo, cyclic, evaluable, seeded)

            # Rebuild after the acyclic sweep so branch and address choices made
            # by newly computed selectors are what stabilization compares.
            self._resolved_targets = {}
            self._resolved_operation_targets = {}
            self.runtime_address_radj = {}
            active_adj, active_radj = self._active_graph(cells)
            graph_signature = self._graph_signature(active_adj)
            target_signature = self._target_signature()
            graph_same = graph_signature == previous_graph
            targets_same = target_signature == previous_targets
            topology_history.append({
                "pass": graph_pass,
                "edge_count": len(graph_signature),
                "resolved_target_count": sum(
                    len(sources) for _, sources in target_signature
                ),
                "topology_unchanged": graph_same,
                "targets_unchanged": targets_same,
            })
            topology_stable = graph_same
            targets_stable = targets_same
            if topology_stable and targets_stable:
                break
            previous_graph = graph_signature
            previous_targets = target_signature

        stabilized_graph_signature = self._graph_signature(active_adj)
        stabilized_target_signature = self._target_signature()
        groups, topo, cyclic = self._component_plan(cells, active_adj)
        self._evaluate_acyclic(groups, topo, cyclic, evaluable, seeded)

        if self.proof_outputs is None:
            relevant_cone = set(cells)
        else:
            relevant_cone = set(self.proof_outputs)
            stack = list(self.proof_outputs)
            while stack:
                target = stack.pop()
                for source in active_radj.get(target, ()):
                    if source not in relevant_cone:
                        relevant_cone.add(source)
                        stack.append(source)

        cycle_specs = []
        for index in sorted(cyclic, key=lambda item: self._cell_key(groups[item][0])):
            complete_members = tuple(groups[index])
            pending = tuple(
                cell for cell in complete_members
                if cell in evaluable and cell not in seeded
            )
            if not pending:
                continue
            cycle_specs.append({
                "members": complete_members,
                "pending": pending,
                "output_relevant": bool(set(complete_members) & relevant_cone),
                "history": [],
            })

        relevant_specs = [
            spec for spec in cycle_specs if spec["output_relevant"]
        ]
        workbook_iterations = 0
        if iterate and relevant_specs:
            for workbook_iteration in range(1, iteration_limit + 1):
                workbook_iterations = workbook_iteration
                changes = {spec["members"]: 0.0 for spec in relevant_specs}
                for spec in relevant_specs:
                    for cell in spec["pending"]:
                        before = self.values[cell]
                        after = self._eval_cell(cell)
                        self.values[cell] = after
                        changes[spec["members"]] = max(
                            changes[spec["members"]],
                            self._value_change(before, after),
                        )
                # Recompute deterministic acyclic dependants between workbook
                # iterations without granting any SCC an extra advance.
                self._evaluate_acyclic(groups, topo, cyclic, evaluable, seeded)
                for spec in relevant_specs:
                    spec["history"].append(changes[spec["members"]])
                if all(
                    not math.isinf(changes[spec["members"]])
                    and changes[spec["members"]] <= iteration_delta
                    for spec in relevant_specs
                ):
                    break

        # The final graph must be the same graph whose cycles were iterated.
        self._resolved_targets = {}
        self._resolved_operation_targets = {}
        self.runtime_address_radj = {}
        final_active_adj, final_active_radj = self._active_graph(cells)
        final_graph_signature = self._graph_signature(final_active_adj)
        final_target_signature = self._target_signature()
        pre_iteration_topology_stable = topology_stable
        pre_iteration_targets_stable = targets_stable

        # Iteration may legitimately move an IF/MAX/lookup selector away from
        # its zero-initialized branch. Certify the graph at the converged values
        # instead of requiring it to equal that pre-iteration graph. Continue
        # coordinated value/topology sweeps until both values and active
        # dependencies stop changing, or the workbook's own iteration budget
        # is exhausted. Oscillating selectors therefore remain fail-closed.
        joint_topology_stable = (
            final_graph_signature == stabilized_graph_signature
        )
        joint_targets_stable = (
            final_target_signature == stabilized_target_signature
        )
        joint_values_stable = all(
            spec["history"]
            and not math.isinf(spec["history"][-1])
            and spec["history"][-1] <= iteration_delta
            for spec in relevant_specs
        )
        while (
            iterate
            and relevant_specs
            and not (
                joint_topology_stable
                and joint_targets_stable
                and joint_values_stable
            )
            and workbook_iterations < iteration_limit
        ):
            reference_graph = final_graph_signature
            reference_targets = final_target_signature
            groups, topo, cyclic = self._component_plan(
                cells, final_active_adj
            )
            changes = {spec["members"]: 0.0 for spec in relevant_specs}
            for spec in relevant_specs:
                for cell in spec["pending"]:
                    before = self.values[cell]
                    after = self._eval_cell(cell)
                    self.values[cell] = after
                    changes[spec["members"]] = max(
                        changes[spec["members"]],
                        self._value_change(before, after),
                    )
            self._evaluate_acyclic(
                groups, topo, cyclic, evaluable, seeded
            )
            workbook_iterations += 1
            for spec in relevant_specs:
                spec["history"].append(changes[spec["members"]])

            self._resolved_targets = {}
            self._resolved_operation_targets = {}
            self.runtime_address_radj = {}
            final_active_adj, final_active_radj = self._active_graph(cells)
            final_graph_signature = self._graph_signature(final_active_adj)
            final_target_signature = self._target_signature()
            joint_topology_stable = (
                final_graph_signature == reference_graph
            )
            joint_targets_stable = (
                final_target_signature == reference_targets
            )
            joint_values_stable = all(
                not math.isinf(changes[spec["members"]])
                and changes[spec["members"]] <= iteration_delta
                for spec in relevant_specs
            )

        canonical_resolved_targets = {
            target: set(sources) for target, sources in self._resolved_targets.items()
        }
        canonical_resolved_operation_targets = {
            operation: set(targets)
            for operation, targets in self._resolved_operation_targets.items()
        }
        canonical_runtime_address_radj = {
            target: set(sources)
            for target, sources in self.runtime_address_radj.items()
        }
        topology_stable = (
            pre_iteration_topology_stable and joint_topology_stable
        )
        targets_stable = (
            pre_iteration_targets_stable and joint_targets_stable
        )

        cycle_diagnostics = []
        diagnostics_by_members = {}
        for spec in cycle_specs:
            members = spec["members"]
            pending = spec["pending"]
            output_relevant = spec["output_relevant"]
            residual, residual_errors, residual_unresolved = (
                self._equation_residual(pending)
                if output_relevant and iterate
                else (None, [], [])
            )
            current_errors = [
                {"cell": cell, "value": repr(self.values[cell]), "phase": "state"}
                for cell in pending if isinstance(self.values[cell], ExcelError)
            ]
            current_unresolved = [
                {
                    "cell": cell,
                    "value": repr(self.values[cell]),
                    "reason": self.values[cell].reason,
                    "phase": "state",
                }
                for cell in pending if isinstance(self.values[cell], Unresolved)
            ]
            equation_errors = [
                {**record, "phase": "equation"} for record in residual_errors
            ]
            equation_unresolved = [
                {**record, "phase": "equation"} for record in residual_unresolved
            ]
            error_records = current_errors + equation_errors
            unresolved_records = current_unresolved + equation_unresolved
            error_cells = sorted({record["cell"] for record in error_records})
            unresolved_cells = sorted({
                record["cell"] for record in unresolved_records
            })
            final_change = (
                spec["history"][-1] if spec["history"] else None
            )
            converged = (
                bool(iterate)
                and output_relevant
                and final_change is not None
                and not math.isinf(final_change)
                and final_change <= iteration_delta
                and residual is not None
                and residual <= iteration_delta
                and not error_records
                and not unresolved_records
            )
            certification = self._affine_cycle_certificate(pending)
            certified = bool(
                certification.get("certified")
                and certification.get("contractive")
            )
            diagnostic = {
                "size": len(members),
                "member_count": len(members),
                "members": list(members[:20]),
                "members_sample": list(members[:20]),
                "members_sample_count": min(len(members), 20),
                "seed": members[0],
                "graph_pass": graph_passes,
                "iterations": workbook_iterations if output_relevant and iterate else 0,
                "workbook_iterations": (
                    workbook_iterations if output_relevant and iterate else 0
                ),
                "iteration_enabled": bool(iterate),
                "iteration_limit": iteration_limit,
                "iteration_delta": iteration_delta,
                "evaluated": bool(output_relevant and iterate),
                "output_relevant": output_relevant,
                "converged": converged,
                "certified": certified,
                "certification": certification,
                "max_change": (
                    None if final_change is None or math.isinf(final_change)
                    else final_change
                ),
                "max_change_history": self._bounded_history([
                    None if math.isinf(change) else change
                    for change in spec["history"]
                ]),
                "max_change_history_complete_count": len(spec["history"]),
                "residual": residual,
                "errors": len(error_records),
                "error_members": error_cells[:20],
                "error_member_count": len(error_cells),
                "error_values": error_records[:20],
                "error_value_count": len(error_records),
                "unresolved": len(unresolved_records),
                "unresolved_members": unresolved_cells[:20],
                "unresolved_member_count": len(unresolved_cells),
                "unresolved_values": unresolved_records[:20],
                "unresolved_value_count": len(unresolved_records),
                "budget_exhausted": bool(
                    output_relevant
                    and iterate
                    and workbook_iterations >= iteration_limit
                    and not converged
                ),
                "topology_stable": topology_stable,
                "targets_stable": targets_stable,
                "uniqueness": (
                    "not_disproven" if converged and topology_stable and targets_stable
                    else "unknown"
                ),
                # Compatibility key: agreement never becomes ``True``.
                "unique": None,
                "reason": (
                    "detached-cycle"
                    if not output_relevant
                    else "iteration-disabled"
                    if not iterate
                    else "active-topology-not-stable"
                    if not topology_stable
                    else "runtime-targets-not-stable"
                    if not targets_stable
                    else "iteration-budget-exhausted"
                    if workbook_iterations >= iteration_limit and not converged
                    else "non-converged"
                    if not converged
                    else None
                ),
            }
            cycle_diagnostics.append(diagnostic)
            diagnostics_by_members[members] = diagnostic

        # A multi-start probe can exhibit a second fixed point, but agreement is
        # only "not_disproven".  All active SCCs share the probe's workbook
        # iteration counter and deterministic advance order as well.
        def strict_fixed_point_tolerance(spec, values):
            scale = max(
                [
                    abs(float(values[cell]))
                    for cell in spec["pending"]
                    if isinstance(values[cell], (int, float))
                    and not isinstance(values[cell], bool)
                    and math.isfinite(float(values[cell]))
                ]
                + [1.0]
            )
            return CONVERGENCE * scale

        probe_specs = [
            spec for spec in relevant_specs
            if diagnostics_by_members[spec["members"]]["converged"]
            and topology_stable
            and targets_stable
            and diagnostics_by_members[spec["members"]]["residual"]
            <= strict_fixed_point_tolerance(spec, self.values)
        ]
        if probe_specs and self.run_probes:
            canonical_values = dict(self.values)
            for spec in probe_specs:
                canonical = {
                    cell: canonical_values[cell] for cell in spec["pending"]
                }
                seed_value = 1.0
                if all(
                    isinstance(value, (int, float))
                    and not isinstance(value, bool)
                    and abs(float(value) - seed_value) <= iteration_delta
                    for value in canonical.values()
                ):
                    seed_value = -1.0
                for cell in spec["pending"]:
                    self.values[cell] = seed_value

            probe_changes = {spec["members"]: math.inf for spec in probe_specs}
            probe_iterations = 0
            for probe_iteration in range(1, iteration_limit + 1):
                probe_iterations = probe_iteration
                for spec in probe_specs:
                    change = 0.0
                    for cell in spec["pending"]:
                        before = self.values[cell]
                        after = self._eval_cell(cell)
                        self.values[cell] = after
                        change = max(change, self._value_change(before, after))
                    probe_changes[spec["members"]] = change
                self._evaluate_acyclic(groups, topo, cyclic, evaluable, seeded)
                if all(
                    not math.isinf(probe_changes[spec["members"]])
                    and probe_changes[spec["members"]]
                    <= strict_fixed_point_tolerance(spec, self.values)
                    for spec in probe_specs
                ):
                    break

            alternate_values = dict(self.values)
            self._resolved_targets = {}
            self._resolved_operation_targets = {}
            probe_adj, _ = self._active_graph(cells)
            probe_graph_signature = self._graph_signature(probe_adj)
            probe_target_signature = self._target_signature()
            probe_global_safe = (
                probe_graph_signature == final_graph_signature
                and probe_target_signature == final_target_signature
            )
            probe_residuals = {}
            for spec in probe_specs:
                probe_residuals[spec["members"]] = self._equation_residual(
                    spec["pending"]
                )
            probe_global_safe = probe_global_safe and all(
                residual is not None
                and residual <= strict_fixed_point_tolerance(spec, alternate_values)
                and not errors
                and not unresolved
                for spec in probe_specs
                for residual, errors, unresolved in [
                    probe_residuals[spec["members"]]
                ]
            )

            for spec in probe_specs:
                members = spec["members"]
                diagnostic = diagnostics_by_members[members]
                diagnostic["probe_iterations"] = probe_iterations
                diagnostic["probe_converged"] = probe_global_safe
                if not probe_global_safe:
                    diagnostic["uniqueness"] = "unknown"
                    continue
                distinct = any(
                    self._value_change(
                        canonical_values[cell], alternate_values[cell]
                    ) > iteration_delta
                    for cell in spec["pending"]
                )
                if not distinct:
                    continue
                diagnostic.update({
                    "unique": False,
                    "uniqueness": "demonstrated_non_unique",
                    "reason": "non-unique-fixed-point",
                    "canonical_fixed_point": {
                        cell: canonical_values[cell] for cell in spec["pending"][:20]
                    },
                    "alternate_fixed_point": {
                        cell: alternate_values[cell] for cell in spec["pending"][:20]
                    },
                    "alternate_fixed_point_member_count": len(spec["pending"]),
                })
            self.values = canonical_values
            self._resolved_targets = canonical_resolved_targets
            self._resolved_operation_targets = (
                canonical_resolved_operation_targets
            )
            self.runtime_address_radj = canonical_runtime_address_radj

        # Preserve the canonical proof endpoint before compatibility unresolved
        # markers are applied to disabled or demonstrated-non-unique cycles.
        self._proof_endpoint_values = dict(self.values)
        self._proof_active_radj = {
            target: set(sources) for target, sources in final_active_radj.items()
        }
        self._proof_resolved_targets = {
            target: set(sources)
            for target, sources in canonical_resolved_targets.items()
        }
        self._proof_resolved_operation_targets = {
            operation: set(targets)
            for operation, targets in canonical_resolved_operation_targets.items()
        }
        for spec in relevant_specs:
            diagnostic = diagnostics_by_members[spec["members"]]
            if not iterate:
                reason = "circular-reference"
            elif diagnostic["uniqueness"] == "demonstrated_non_unique":
                reason = "non-unique-circular-reference"
            else:
                continue
            for cell in spec["pending"]:
                self.values[cell] = Unresolved(reason)
                indeterminate.add(cell)

        cycle_diagnostics.sort(key=lambda item: self._cell_key(item["seed"]))
        cycle_groups = {
            index: tuple(spec["pending"])
            for index, spec in enumerate(cycle_specs, 1)
        }
        cycle_membership = {
            cell: index
            for index, members_in_cycle in cycle_groups.items()
            for cell in members_in_cycle
        }
        iterated = cycle_diagnostics

        for cid, value in self.values.items():
            if isinstance(value, Unresolved):
                self.unresolved[cid] = value.reason

        computed = sum(1 for c in cells if not isinstance(self.values.get(c), Unresolved))
        errors = sum(isinstance(self.values.get(c), ExcelError) for c in cells)
        coverage = {
            "cells": len(cells),
            "computed": computed,
            "unresolved": len(self.unresolved),
            "errors": errors,
            "unknown_ops": dict(self.unknown_ops),
            "active_graph_passes": graph_passes,
            "iteration_enabled": iterate,
            "iteration_setting": iterate_setting,
            "iteration_limit": iteration_limit,
            "iteration_delta": iteration_delta,
            "workbook_iterations": workbook_iterations,
            "active_cycles": len(cycle_diagnostics),
            "active_topology_stable": topology_stable,
            "runtime_targets_stable": targets_stable,
            "active_topology_history": topology_history,
            "calculation_metadata": {
                "available": getattr(settings, "available", False),
                "iterate": iterate_setting,
                "iterate_origin": getattr(settings, "iterate_origin", "unknown"),
                "iterate_count": raw_limit,
                "iterate_count_origin": getattr(
                    settings, "iterate_count_origin", "unknown"
                ),
                "iterate_delta": raw_delta,
                "iterate_delta_origin": getattr(
                    settings, "iterate_delta_origin", "unknown"
                ),
                "raw_calc_pr": getattr(settings, "raw_calc_pr", {}),
                "reason": getattr(settings, "reason", ""),
            },
            "strict_proof": self.strict_proof,
            "expected_cache_reads": sum(
                len(v) for v in self.oracle_accesses.values()
            ),
            "runtime_read_cells": len(self.read_attempts),
            "missing_read_cells": len(self.missing_reads),
            "stable_self_cycles": sum(
                diagnostic["size"] == 1 and diagnostic["converged"]
                and diagnostic.get("uniqueness") != "demonstrated_non_unique"
                for diagnostic in cycle_diagnostics
            ),
        }
        # The final pass's active graph carries the dependencies actually read,
        # including runtime-resolved dynamic references (OFFSET targets) that
        # the static edge list cannot see. Verification walks the output cone
        # over these edges as well as the static ones.
        runtime_radj = {
            target: set(sources) for target, sources in final_active_radj.items()
        }
        runtime_address_radj = {
            target: set(sources)
            for target, sources in self.runtime_address_radj.items()
        }
        return EvalResult(
            self.values, self.unresolved, iterated, coverage,
            runtime_radj=runtime_radj,
            runtime_address_radj=runtime_address_radj,
            resolved_targets={
                target: tuple(sorted(sources))
                for target, sources in sorted(self._resolved_targets.items())
            },
            resolved_operation_targets={
                operation: tuple(sorted(targets))
                for operation, targets in sorted(
                    self._resolved_operation_targets.items()
                )
            },
            oracle_reads=self.oracle_reads,
            oracle_accesses=self.oracle_accesses,
            read_attempts=self.read_attempts,
            missing_reads=self.missing_reads,
            seed_provenance={
                "effective": {
                    "count": len(seeded),
                    "sample": sorted(seeded)[:20],
                },
                "categories": {
                    name: {
                        "count": len(ids),
                        "sample": sorted(ids)[:20],
                    }
                    for name, ids in sorted(seed_categories.items())
                },
            },
            seeded_cells=seeded,
            evaluated_cells=set(cells),
            cycle_membership=cycle_membership,
            cycle_groups=cycle_groups,
            endpoint_equation_checker=self._endpoint_equation_check,
            strict_proof=self.strict_proof,
        )

    @staticmethod
    def _value_change(before, after):
        if isinstance(before, (int, float)) and isinstance(after, (int, float)):
            if math.isfinite(float(before)) and math.isfinite(float(after)):
                return abs(float(after) - float(before))
        return 0.0 if type(before) is type(after) and repr(before) == repr(after) else math.inf

    def _stabilize(self, order, seeded):
        """Bounded fixed-point sweeps for dependencies hidden by dynamic references."""
        for _ in range(MAX_SWEEPS):
            changed = 0
            # The static order is optimal for known edges. The reverse pass
            # quickly carries values across dynamic edges whose direction was
            # unavailable to the graph builder.
            for sweep in (order, reversed(order)):
                for cell in sweep:
                    if cell in seeded:
                        continue
                    before = self.values.get(cell)
                    after = self._eval_cell(cell)
                    self.values[cell] = after
                    if isinstance(before, (int, float)) and isinstance(after, (int, float)):
                        scale = max(abs(float(before)), abs(float(after)), 1.0)
                        changed += abs(float(after) - float(before)) > CONVERGENCE * scale
                    else:
                        changed += type(before) is not type(after) or repr(before) != repr(after)
            if not changed:
                break

    def _eval_cell(self, cid: str):
        self._current_cell = cid
        info = self.cg.info[cid]
        root = self.graph.root_of(cid)
        if root is not None:
            value = self._eval_node(root)
            if isinstance(value, RangeValues):
                node = info.node
                if getattr(node, "array_ref", ""):
                    member_row = getattr(node, "array_row", None)
                    member_col = getattr(node, "array_col", None)
                    member_row = 0 if member_row is None else member_row
                    member_col = 0 if member_col is None else member_col
                    if (
                        member_row < 0 or member_col < 0
                        or member_row >= value.rows or member_col >= value.cols
                    ):
                        return ExcelError("#N/A")
                    index = member_row * value.cols + member_col
                    return value[index] if index < len(value) else ExcelError("#N/A")
                # Legacy graphs did not persist array-member coordinates. Keep
                # their historical anchor behavior while new graphs project
                # every CSE/spill member through the metadata above.
                return value[0] if value else None
            return value
        if info.empty_ref:
            return 0.0
        return Unresolved("no-ast-root")


def _xml_bool(value):
    if value is None:
        return None
    normalized = str(value).strip().lower()
    if normalized in ("1", "true", "yes", "on"):
        return True
    if normalized in ("0", "false", "no", "off"):
        return False
    return None


def workbook_calculation_metadata(path) -> CalculationMetadata:
    """Read only OOXML calculation settings, never worksheet cached values."""
    path = Path(path)
    try:
        metadata = path.lstat()
    except OSError:
        return CalculationMetadata(
            available=False, source=str(path), reason="file_not_found"
        )
    if path.is_symlink() or not stat.S_ISREG(metadata.st_mode):
        return CalculationMetadata(
            available=False, source=str(path), reason="source_not_regular_file"
        )
    try:
        flags = os.O_RDONLY | getattr(os, "O_NOFOLLOW", 0)
        descriptor = os.open(path, flags)
        try:
            opened = os.fstat(descriptor)
            if not stat.S_ISREG(opened.st_mode):
                raise OSError("source_not_regular_file")
            with os.fdopen(descriptor, "rb") as handle:
                descriptor = -1
                with zipfile.ZipFile(handle) as archive:
                    root = ElementTree.fromstring(archive.read("xl/workbook.xml"))
        finally:
            if descriptor >= 0:
                os.close(descriptor)
    except (OSError, KeyError, zipfile.BadZipFile, ElementTree.ParseError) as exc:
        return CalculationMetadata(
            available=False,
            source=str(path),
            reason=f"{type(exc).__name__}: {exc}",
        )
    calc = root.find("{*}calcPr")
    if calc is None:
        return CalculationMetadata(
            available=True,
            iterate=False,
            source=str(path),
            reason="calcPr_absent",
            iterate_origin="default",
            iterate_count_origin="default",
            iterate_delta_origin="default",
            calc_mode_origin="default",
            full_calc_on_load_origin="default",
            force_full_calc_origin="default",
            raw_calc_pr={},
        )
    count_raw = calc.attrib.get("iterateCount")
    delta_raw = calc.attrib.get("iterateDelta")
    try:
        count = int(count_raw) if count_raw else None
        count_valid = (
            count_raw is None
            or 0 < count <= MAX_WORKBOOK_ITERATIONS
        )
        if not count_valid:
            count = None
    except (TypeError, ValueError):
        count = None
        count_valid = False
    try:
        delta = float(delta_raw) if delta_raw else None
        delta_valid = (
            delta_raw is None
            or math.isfinite(delta)
            and 0 <= delta <= MAX_ITERATION_DELTA
        )
        if not delta_valid:
            delta = None
    except (TypeError, ValueError):
        delta = None
        delta_valid = False
    raw = {str(key): str(value) for key, value in sorted(calc.attrib.items())}
    iterate_raw = calc.attrib.get("iterate")
    parsed_iterate = _xml_bool(iterate_raw)
    calc_mode_raw = calc.attrib.get("calcMode")
    parsed_calc_mode = (
        calc_mode_raw
        if calc_mode_raw in {"auto", "autoNoTable", "manual"}
        else None
    )
    full_calc_raw = calc.attrib.get("fullCalcOnLoad")
    parsed_full_calc = _xml_bool(full_calc_raw)
    force_full_raw = calc.attrib.get("forceFullCalc")
    parsed_force_full = _xml_bool(force_full_raw)
    return CalculationMetadata(
        available=True,
        iterate=parsed_iterate if iterate_raw is not None else False,
        iterate_count=count,
        iterate_delta=delta,
        calc_mode=parsed_calc_mode,
        full_calc_on_load=parsed_full_calc,
        force_full_calc=parsed_force_full,
        source=str(path),
        iterate_origin=(
            "unknown"
            if iterate_raw is not None and parsed_iterate is None
            else "explicit"
            if iterate_raw is not None
            else "default"
        ),
        iterate_count_origin=(
            "unknown"
            if not count_valid
            else "explicit"
            if count_raw is not None
            else "default"
        ),
        iterate_delta_origin=(
            "unknown"
            if not delta_valid
            else "explicit"
            if delta_raw is not None
            else "default"
        ),
        calc_mode_origin=(
            "unknown"
            if calc_mode_raw is not None and parsed_calc_mode is None
            else "explicit"
            if calc_mode_raw is not None
            else "default"
        ),
        full_calc_on_load_origin=(
            "unknown"
            if full_calc_raw is not None and parsed_full_calc is None
            else "explicit"
            if full_calc_raw is not None
            else "default"
        ),
        force_full_calc_origin=(
            "unknown"
            if force_full_raw is not None and parsed_force_full is None
            else "explicit"
            if force_full_raw is not None
            else "default"
        ),
        raw_calc_pr=raw,
    )


class ExpectedValueCache:
    """Comparison-only workbook cache with explicit read provenance."""

    def __init__(self, path, book):
        self.path = str(path)
        self.book = book
        self.cache: dict = {}
        self.reads: dict = {}

    def __call__(self, sheet, row, col):
        cid = a1(sheet, row, col)
        self.reads[cid] = self.reads.get(cid, 0) + 1
        if sheet not in self.cache:
            if sheet not in self.book.sheetnames:
                self.cache[sheet] = {}
            else:
                grid = {}
                for line in self.book[sheet].iter_rows():
                    for cell in line:
                        if cell.value is not None:
                            grid[(cell.row, cell.column)] = cell.value
                self.cache[sheet] = grid
        value = self.cache[sheet].get((row, col))
        if isinstance(value, datetime):
            return _datetime_to_serial(value)
        return value

    def close(self):
        self.book.close()


def workbook_expected_cache(path):
    """Open the saved-value cache for post-proof comparison only."""
    try:
        import openpyxl
    except ImportError:
        return None
    try:
        book = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception:
        return None
    return ExpectedValueCache(path, book)


def workbook_oracle(path):
    """Backward-compatible combined cache; strict proof never uses this API."""
    cache = workbook_expected_cache(path)
    if cache is None:
        return None
    calculation = workbook_calculation_metadata(path)
    cache.iterate = calculation.iterate
    cache.iterate_count = calculation.iterate_count
    cache.iterate_delta = calculation.iterate_delta
    return cache


def compare(expected: str, actual, tolerance=1e-6):
    """Relative comparison against the workbook's cached value."""
    if isinstance(actual, Unresolved):
        return "unresolved", None
    want = literal(expected)
    if want is None:
        # The workbook cached no value, so there is no ground truth to compare
        # against. Calling a recomputed 0/blank a "match" here would verify
        # nothing (and Python's False == 0.0 made even FALSE pass); report the
        # cell as unverifiable instead of inflating the match count.
        return "unverifiable", None
    if isinstance(want, str) or isinstance(actual, str):
        return ("match" if _text(want).strip() == _text(actual).strip() else "mismatch"), None
    got, ref = num(actual, math.nan), num(want, math.nan)
    if math.isnan(got) or math.isnan(ref):
        return "unresolved", None
    diff = abs(got - ref)
    scale = max(abs(ref), 1.0)
    return ("match" if diff / scale <= tolerance else "mismatch"), diff
