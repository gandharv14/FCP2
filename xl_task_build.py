#!/usr/bin/env python3
"""Build Harbor-format evaluation tasks from level-split workbooks.

For every sampled (workbook, cutoff, template) triple this emits one Harbor
task directory:

    tasks_out/{id}-L{xx}-{template}/
        task.toml            config + full pipeline metadata
        instruction.md       naturalized (or skeleton) prompt
        environment/Lxx.xlsx the artifact the agent works on
        tests/test.sh        stub verifier (graders land later)
        tests/answer_key.json  expected values, verifier-side only
        tests/facts.json     facts the instruction asserts (audit / grading)

Pipeline stages, all deterministic except stage 3:
  1. spec compiler   targets = formula cells above the cutoff, minus cells in
                     cycles, unparsed formulas, volatile ancestors and error
                     values; label resolution names cells "EBITDA @ FY2027"
  2. assembler       template eligibility + seeded sampling from
                     task_templates.yaml, rendering a prompt skeleton + facts
  3. naturalizer     GPT 5.6 Luna via the Labelbox LiteLLM proxy rephrases the
                     scenario losslessly (--no-naturalize skips it)
  4. fact verifier   every reference/label/number in the rewrite must trace
                     back to the facts; bounded retries, skeleton fallback
  5. harbor emitter  writes the bundle above

Requires: openpyxl, pyyaml, and xl_ast_graph.py / xl_level_split.py next to it.
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import os
import random
import re
import shutil
import sys
import time
import urllib.error
import urllib.request
from collections import defaultdict
from pathlib import Path

try:
    import yaml
except ImportError:  # pragma: no cover
    sys.exit("pyyaml is required:  python3 -m pip install pyyaml")

try:
    from xl_ast_graph import AstGraph, jsonable
    from xl_level_split import UNKNOWN_LEVEL, levels_by_sheet
except ImportError:  # pragma: no cover
    sys.exit("xl_task_build.py expects xl_ast_graph.py and xl_level_split.py "
             "next to it")

from openpyxl.utils import get_column_letter

PIPELINE_VERSION = "1.0.0"
DEFAULT_PROJECT_ID = "cms6m4urm006n07z8ecxi1oi2"
PROD_ENDPOINT = "https://litellm.labelbox.com"
DEV_ENDPOINT = "https://litellm.lb-dev.xyz"

VOLATILE_RE = re.compile(r"\b(NOW|TODAY|RAND|RANDBETWEEN|RANDARRAY|INDIRECT)\s*\(",
                         re.I)
SUM_RUN_RE = re.compile(r"^=\s*SUM\(\s*\$?([A-Z]{1,3})\$?(\d+)\s*:\s*"
                        r"\$?([A-Z]{1,3})\$?(\d+)\s*\)\s*$", re.I)
RULE_RE = re.compile(r"^\s*(>=|<=|>|<)\s*([\d.]+)\s*$")
SNAPSHOT_RE = re.compile(r"^L(\d+)\.(xlsx|xlsm)$", re.I)

# templates needing headless recalculation for their answer keys are deferred
# to the grader iteration
DEFERRED_TEMPLATES = {"whatif_input_shift", "whatif_scenario"}

MAX_FRONTIER_LISTED = 40
MAX_SLOT_CANDIDATES = 400
NATURALIZE_RETRIES = 2

# words that disqualify a label from matching a plain line-item pattern
# ("Revenue growth %" must not match "revenue")
NEGATIVE_LABEL_WORDS = ("%", "margin", "growth", "ratio", "per ", "yoy",
                        "y-o-y", "change", "non current", "non-current",
                        "noncurrent")

HURDLES = {
    "irr":  {"display": "15%",  "op": ">=", "threshold": 0.15, "percent_aware": True},
    "moic": {"display": "2.0x", "op": ">=", "threshold": 2.0,  "percent_aware": False},
    "npv":  {"display": "0",    "op": ">",  "threshold": 0.0,  "percent_aware": False},
    "dscr": {"display": "1.2x", "op": ">=", "threshold": 1.2,  "percent_aware": False},
}

TRAJECTORY_LABEL_RE = re.compile(r"margin|ratio|rate|%", re.I)

# a text header only counts as a period if it looks like one ("FY2027", "Q3",
# "2026A", "Jan-27", "Year 5"); bare dashes and stray notes are skipped
PERIOD_TEXT_RE = re.compile(
    r"\d|fy|q[1-4]|jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec"
    r"|year|month|quarter|actual|budget|forecast|hist|proj|ltm|ttm", re.I)


def plausible_row_label(text):
    return bool(re.search(r"[A-Za-z0-9]", text)) and len(text) >= 2


def qualifier_like(text):
    """Currency codes, units and similar short qualifiers ("AED", "US$m")."""
    return len(text) <= 5 and (text.isupper() or any(s in text for s in "$€£%"))


def is_number(value):
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def ref_str(sheet, row, col):
    """Excel-style qualified reference, quoting the sheet when needed."""
    name = sheet if re.fullmatch(r"[A-Za-z0-9_]+", sheet) else "'%s'" % sheet
    return "%s!%s%d" % (name, get_column_letter(col), row)


def fmt_number(value):
    if isinstance(value, float) and value == int(value) and abs(value) < 1e15:
        return str(int(value))
    return repr(value) if isinstance(value, float) else str(value)


def compare(op, value, threshold):
    return {"": False, ">": value > threshold, ">=": value >= threshold,
            "<": value < threshold, "<=": value <= threshold}[op]


# ---------------------------------------------------------------------------
# stage 1: spec compiler
# ---------------------------------------------------------------------------

class Spec:
    """Everything stage 2 needs, computed once per (workbook, cutoff)."""

    def __init__(self, graph, cutoff):
        self.graph = graph
        self.cutoff = cutoff
        self.placed, self.deepest, _, _ = levels_by_sheet(graph)
        self.volatile = self._volatile_cells()
        self.targets = []          # gradable (sheet, row, col, level, value)
        self.excluded = defaultdict(int)
        self._collect_targets()
        self.target_set = {(s, r, c) for s, r, c, _, _ in self.targets}
        self._label_cache = {}
        self._series_cache = {}
        self._labels_cache = {}

    # -- targets ----------------------------------------------------------

    def _volatile_cells(self):
        """Cells whose own formula or any ancestor's formula is volatile."""
        own = set()
        for sheet, cells in self.graph.formulas.items():
            for coord, (formula, _) in cells.items():
                if VOLATILE_RE.search(formula):
                    own.add((sheet,) + coord)
        if not own:
            return set()

        succ = defaultdict(set)
        for edge in self.graph.edges:
            src = self.graph.nodes.get(edge["source"])
            dst = self.graph.nodes.get(edge["target"])
            if not src or not dst:
                continue
            s_owner = src["owner"] or edge["source"]
            d_owner = dst["owner"] or edge["target"]
            s = self.graph.nodes.get(s_owner, src)
            d = self.graph.nodes.get(d_owner, dst)
            if s["row"] is None or d["row"] is None:
                continue
            succ[(s["sheet"], s["row"], s["col"])].add((d["sheet"], d["row"], d["col"]))

        tainted = set(own)
        frontier = list(own)
        while frontier:
            cell = frontier.pop()
            for other in succ.get(cell, ()):
                if other not in tainted:
                    tainted.add(other)
                    frontier.append(other)
        return tainted

    def _collect_targets(self):
        for sheet, cells in self.graph.formulas.items():
            values = self.graph.values.get(sheet, {})
            for coord in cells:
                row, col = coord
                level = self.placed.get(sheet, {}).get(coord, UNKNOWN_LEVEL)
                if level == UNKNOWN_LEVEL:
                    self.excluded["unparsed"] += 1
                    continue
                if level <= self.cutoff:
                    continue  # still visible in Lx
                node = self.graph.nodes.get(AstGraph.cell_id(sheet, row, col))
                value = values.get(coord)
                if node is not None and node.get("in_cycle"):
                    self.excluded["cycle"] += 1
                elif (sheet, row, col) in self.volatile:
                    self.excluded["volatile"] += 1
                elif isinstance(value, str) and value.startswith("#"):
                    self.excluded["error_value"] += 1
                elif value is None:
                    self.excluded["no_cached_value"] += 1
                else:
                    self.targets.append((sheet, row, col, level, value))
        self.targets.sort(key=lambda t: (t[0], t[1], t[2]))

    @property
    def frontier(self):
        if not self.targets:
            return []
        lowest = min(t[3] for t in self.targets)
        return [t for t in self.targets if t[3] == lowest]

    def is_target(self, sheet, row, col):
        return (sheet, row, col) in self.target_set

    # -- visibility & labels ----------------------------------------------

    def visible(self, sheet, row, col):
        """True when the cell still shows a value in the Lx snapshot."""
        entry = self.graph.formulas.get(sheet, {}).get((row, col))
        if entry is None:
            return True  # static text / typed input survives every snapshot
        level = self.placed.get(sheet, {}).get((row, col), UNKNOWN_LEVEL)
        if level == UNKNOWN_LEVEL or level > self.cutoff:
            return False
        return self.graph.values.get(sheet, {}).get((row, col)) is not None

    def row_label(self, sheet, row, col):
        """Line-item text left of the cell: nearest non-qualifier wins,
        because section banners sit at the far left while the item name sits
        next to the numbers; falls back to the leftmost plausible text."""
        key = ("r", sheet, row, col)
        if key in self._label_cache:
            return self._label_cache[key]
        values = self.graph.values.get(sheet, {})
        texts = []
        for c in self.graph.text_cols_by_row.get(sheet, {}).get(row, []):
            if c >= col:
                break
            text = values.get((row, c))
            if (isinstance(text, str) and plausible_row_label(text.strip())
                    and self.visible(sheet, row, c)):
                texts.append(text.strip())
        label = ""
        for text in reversed(texts):
            if not qualifier_like(text):
                label = text
                break
        if not label and texts:
            label = texts[0]
        self._label_cache[key] = label
        return label

    def period_label(self, sheet, row, col):
        """Nearest visible period-like header above the cell, same column."""
        key = ("p", sheet, row, col)
        if key in self._label_cache:
            return self._label_cache[key]
        label = ""
        values = self.graph.values.get(sheet, {})
        for r in range(row - 1, max(0, row - 200) - 1, -1):
            value = values.get((r, col))
            if value is None or not self.visible(sheet, r, col):
                continue
            if isinstance(value, str):
                text = value.strip()
                if (text and len(text) <= 40 and PERIOD_TEXT_RE.search(text)):
                    label = text
                    break
                continue
            if isinstance(value, dt.datetime):
                label = value.date().isoformat()
                break
            if isinstance(value, dt.date):
                label = value.isoformat()
                break
            if (is_number(value) and 1900 <= value <= 2100
                    and float(value).is_integer()):
                label = str(int(value))
                break
        self._label_cache[key] = label
        return label

    def sheet_row_labels(self, sheet):
        """{row: full-row label} using the same preference as row_label."""
        if sheet in self._labels_cache:
            return self._labels_cache[sheet]
        labels = {}
        for row in self.graph.text_cols_by_row.get(sheet, {}):
            label = self.row_label(sheet, row, 10 ** 4)
            if label:
                labels[row] = label
        self._labels_cache[sheet] = labels
        return labels

    def rows_with_label(self, sheet, label):
        """All rows on the sheet whose leftmost text equals the label."""
        return {row for row, lbl in self.sheet_row_labels(sheet).items()
                if lbl == label}

    def period_series(self, sheet, row):
        """Columns in the row that hold numbers under a period header."""
        key = (sheet, row)
        if key in self._series_cache:
            return self._series_cache[key]
        values = self.graph.values.get(sheet, {})
        series = sorted(c for (r, c), value in values.items()
                        if r == row and is_number(value)
                        and self.period_label(sheet, row, c))
        self._series_cache[key] = series
        return series


# ---------------------------------------------------------------------------
# stage 2: template instantiation
# ---------------------------------------------------------------------------

class Instance:
    """One concrete task: prompt pieces, facts and (verifier-side) answer."""

    def __init__(self, template_id, financebench, slots, scenario,
                 output_format, answer_key, answer_cells, facts_required,
                 facts_refs, forbidden):
        self.template_id = template_id
        self.financebench = financebench
        self.slots = slots
        self.scenario = scenario
        self.output_format = output_format
        self.answer_key = answer_key
        self.answer_cells = answer_cells        # [(sheet,row,col,level)]
        self.facts_required = facts_required    # strings that must survive
        self.facts_refs = facts_refs            # cell refs allowed to appear
        self.forbidden = forbidden               # answer strings that must not


def label_ok(label, patterns):
    low = label.lower()
    if any(word in low for word in NEGATIVE_LABEL_WORDS):
        return False
    return any(p in low for p in patterns)


def named_candidates(spec):
    """Gradable numeric targets with unambiguous row + period labels."""
    out = []
    for sheet, row, col, level, value in spec.targets[:MAX_SLOT_CANDIDATES]:
        if not is_number(value):
            continue
        if 1900 <= value <= 2100 and float(value).is_integer():
            continue  # almost certainly a computed year header, not a metric
        metric = spec.row_label(sheet, row, col)
        period = spec.period_label(sheet, row, col)
        if not metric or not period or len(metric) > 80:
            continue
        if len(spec.rows_with_label(sheet, metric)) != 1:
            continue
        out.append((sheet, row, col, level, value, metric, period))
    return out


def build_recon_full(spec, tpl, rng):
    if not spec.targets:
        return None
    key = {"kind": "cell_value",
           "tolerance": {"numeric_abs": 1e-6, "numeric_rel": 1e-6},
           "targets": {ref_str(s, r, c): jsonable(v)
                       for s, r, c, _, v in spec.targets}}
    return Instance(
        tpl["id"], tpl["financebench"], {},
        scenario=tpl["prompt"].strip(),
        output_format=('Write a JSON object to `/app/answers.json` mapping '
                       'every reconstructed cell reference (e.g. "Sheet!A1") '
                       'to its value.'),
        answer_key=key,
        answer_cells=[(s, r, c, l) for s, r, c, l, _ in spec.targets],
        facts_required=[], facts_refs=[], forbidden=[])


def build_recon_frontier(spec, tpl, rng):
    frontier = spec.frontier
    if not frontier:
        return None
    if len(frontier) > MAX_FRONTIER_LISTED:
        frontier = sorted(rng.sample(frontier, MAX_FRONTIER_LISTED),
                          key=lambda t: (t[0], t[1], t[2]))
    refs = [ref_str(s, r, c) for s, r, c, _, _ in frontier]
    listed = ", ".join(refs)
    return Instance(
        tpl["id"], tpl["financebench"],
        {"frontier_refs": listed},
        scenario=tpl["prompt"].strip().format(frontier_refs=listed),
        output_format=('Write a JSON object to `/app/answers.json` mapping '
                       'each listed cell reference to its value.'),
        answer_key={"kind": "cell_value",
                    "tolerance": {"numeric_abs": 1e-6, "numeric_rel": 1e-6},
                    "targets": {ref_str(s, r, c): jsonable(v)
                                for s, r, c, _, v in frontier}},
        answer_cells=[(s, r, c, l) for s, r, c, l, _ in frontier],
        facts_required=refs, facts_refs=refs, forbidden=[])


def build_extract(spec, tpl, rng):
    candidates = named_candidates(spec)
    if not candidates:
        return None
    sheet, row, col, level, value, metric, period = rng.choice(candidates)
    slots = {"metric": metric, "period": period, "sheet": sheet}
    return Instance(
        tpl["id"], tpl["financebench"], slots,
        scenario=tpl["prompt"].strip().format(**slots),
        output_format='Write `{"answer": <number>}` to `/app/answers.json`.',
        answer_key={"kind": "cell_value", "tolerance": {"numeric_rel": 1e-4},
                    "targets": {ref_str(sheet, row, col): jsonable(value)}},
        answer_cells=[(sheet, row, col, level)],
        facts_required=[metric, period, sheet],
        facts_refs=[], forbidden=[fmt_number(value)])


def _ratio_pairs(spec, library):
    """(entry_name, entry, sheet, period_col, num cell, den cell) combos."""
    combos = []
    for name, entry in library.items():
        nums = [p.lower() for p in entry["num"]]
        dens = [p.lower() for p in entry["den"]]
        for sheet in spec.graph.values:
            labels = spec.sheet_row_labels(sheet)
            num_rows = [r for r, lbl in labels.items() if label_ok(lbl, nums)
                        and len(spec.rows_with_label(sheet, lbl)) == 1]
            den_rows = [r for r, lbl in labels.items() if label_ok(lbl, dens)
                        and len(spec.rows_with_label(sheet, lbl)) == 1]
            for nr in num_rows[:4]:
                for dr in den_rows[:4]:
                    if nr == dr:
                        continue
                    shared = set(spec.period_series(sheet, nr)) & \
                        set(spec.period_series(sheet, dr))
                    for col in sorted(shared):
                        values = spec.graph.values.get(sheet, {})
                        nv, dv = values.get((nr, col)), values.get((dr, col))
                        if not (is_number(nv) and is_number(dv)) or dv == 0:
                            continue
                        blanked = [spec.is_target(sheet, nr, col),
                                   spec.is_target(sheet, dr, col)]
                        if not any(blanked):
                            continue
                        combos.append((name, entry, sheet, col, nr, nv, dr, dv))
    return combos


def build_compute_ratio(spec, tpl, rng):
    combos = _ratio_pairs(spec, tpl["ratio_library"])
    if not combos:
        return None
    name, entry, sheet, col, nr, nv, dr, dv = rng.choice(combos)
    num_label = spec.sheet_row_labels(sheet).get(nr, "")
    den_label = spec.sheet_row_labels(sheet).get(dr, "")
    period = spec.period_label(sheet, nr, col) or spec.period_label(sheet, dr, col)
    if not (num_label and den_label and period):
        return None
    ratio_name = name.replace("_", " ")
    definition = "%s / %s" % (num_label, den_label)
    slots = {"ratio_name": ratio_name, "definition": definition,
             "period": period}
    expected = round(nv / dv, 2)
    answer_cells = [(sheet, r, col, spec.placed[sheet].get((r, col), 0))
                    for r in (nr, dr) if spec.is_target(sheet, r, col)]
    return Instance(
        tpl["id"], tpl["financebench"], slots,
        scenario=tpl["prompt"].strip().format(**slots),
        output_format=('Write `{"answer": <number rounded to two decimal '
                       'places>}` to `/app/answers.json`.'),
        answer_key={"kind": "derived", "tolerance": {"numeric_abs": 0.005},
                    "answer": expected,
                    "detail": {"numerator": ref_str(sheet, nr, col),
                               "denominator": ref_str(sheet, dr, col),
                               "numerator_value": jsonable(nv),
                               "denominator_value": jsonable(dv)}},
        answer_cells=answer_cells,
        facts_required=[ratio_name, definition, period],
        facts_refs=[], forbidden=[fmt_number(expected)])


def build_compute_growth(spec, tpl, rng):
    candidates = []
    for sheet, row, col, level, value, metric, period_b in named_candidates(spec):
        series = spec.period_series(sheet, row)
        if col not in series:
            continue
        i = series.index(col)
        if i == 0:
            continue
        col_a = series[i - 1]
        value_a = spec.graph.values.get(sheet, {}).get((row, col_a))
        if not is_number(value_a) or value_a == 0:
            continue
        if not spec.visible(sheet, row, col_a):
            continue
        period_a = spec.period_label(sheet, row, col_a)
        if not period_a or period_a == period_b:
            continue
        candidates.append((sheet, row, col, col_a, level, value, value_a,
                           metric, period_a, period_b))
    if not candidates:
        return None
    (sheet, row, col, col_a, level, value, value_a,
     metric, period_a, period_b) = rng.choice(candidates)
    slots = {"metric": metric, "period_a": period_a, "period_b": period_b}
    expected = round((value - value_a) / abs(value_a) * 100.0, 1)
    return Instance(
        tpl["id"], tpl["financebench"], slots,
        scenario=tpl["prompt"].strip().format(**slots),
        output_format=('Write `{"answer": <growth in percent, one decimal '
                       'place>}` to `/app/answers.json`.'),
        answer_key={"kind": "derived", "tolerance": {"numeric_abs": 0.05},
                    "answer": expected, "units": "percent",
                    "detail": {"later": ref_str(sheet, row, col),
                               "earlier": ref_str(sheet, row, col_a),
                               "later_value": jsonable(value),
                               "earlier_value": jsonable(value_a)}},
        answer_cells=[(sheet, row, col, level)],
        facts_required=[metric, period_a, period_b],
        facts_refs=[], forbidden=[fmt_number(expected)])


def build_judge_threshold(spec, tpl, rng, ratio_library):
    combos = []
    for judgment, entry in tpl["judgment_library"].items():
        ratio_entry = ratio_library.get(entry["ratio"])
        if not ratio_entry:
            continue
        for combo in _ratio_pairs(spec, {entry["ratio"]: ratio_entry}):
            combos.append((judgment, entry, combo))
    if not combos:
        return None
    judgment, entry, combo = rng.choice(combos)
    _, _, sheet, col, nr, nv, dr, dv = combo
    num_label = spec.sheet_row_labels(sheet).get(nr, "")
    den_label = spec.sheet_row_labels(sheet).get(dr, "")
    period = spec.period_label(sheet, nr, col) or spec.period_label(sheet, dr, col)
    if not (num_label and den_label and period):
        return None
    match = RULE_RE.match(entry["rule"])
    op, threshold = match.group(1), float(match.group(2))
    ratio_name = entry["ratio"].replace("_", " ")
    frame = ("is the business %s as of %s, judged by %s (%s / %s) with the "
             "threshold %s %s?"
             % (entry["verdict_true"], period, ratio_name, num_label,
                den_label, op, fmt_number(threshold)))
    frame = frame[0].upper() + frame[1:]
    value = nv / dv
    verdict = "yes" if compare(op, value, threshold) else "no"
    slots = {"question_frame": frame}
    answer_cells = [(sheet, r, col, spec.placed[sheet].get((r, col), 0))
                    for r in (nr, dr) if spec.is_target(sheet, r, col)]
    return Instance(
        tpl["id"], tpl["financebench"], slots,
        scenario=tpl["prompt"].strip().format(**slots),
        output_format=('Write `{"verdict": "yes"|"no", "value": <the metric '
                       'you computed>}` to `/app/answers.json`.'),
        answer_key={"kind": "boolean+value",
                    "tolerance": {"numeric_abs": 0.005},
                    "verdict": verdict, "value": round(value, 4),
                    "detail": {"judgment": judgment, "rule": entry["rule"],
                               "numerator": ref_str(sheet, nr, col),
                               "denominator": ref_str(sheet, dr, col)}},
        answer_cells=answer_cells,
        facts_required=[frame],
        facts_refs=[], forbidden=[fmt_number(round(value, 4))])


def build_judge_trajectory(spec, tpl, rng):
    candidates = []
    for sheet, row, col, level, value, metric, _period in named_candidates(spec):
        if not TRAJECTORY_LABEL_RE.search(metric):
            continue
        series = spec.period_series(sheet, row)
        if len(series) < 3 or series[-1] != col:
            continue
        first_col = series[0]
        if not spec.visible(sheet, row, first_col):
            continue
        first = spec.graph.values.get(sheet, {}).get((row, first_col))
        if not is_number(first):
            continue
        p_first = spec.period_label(sheet, row, first_col)
        p_last = spec.period_label(sheet, row, col)
        if not p_first or not p_last or p_first == p_last:
            continue
        candidates.append((sheet, row, col, level, value, metric,
                           first, p_first, p_last))
    if not candidates:
        return None
    (sheet, row, col, level, value, metric,
     first, p_first, p_last) = rng.choice(candidates)
    periods = "%s to %s" % (p_first, p_last)
    slots = {"metric": metric, "periods": periods}
    verdict = "improving" if value > first else "deteriorating"
    return Instance(
        tpl["id"], tpl["financebench"], slots,
        scenario=tpl["prompt"].strip().format(**slots),
        output_format=('Write `{"verdict": "improving"|"deteriorating", '
                       '"first": <number>, "last": <number>}` to '
                       '`/app/answers.json`.'),
        answer_key={"kind": "boolean+value",
                    "tolerance": {"numeric_abs": 0.005},
                    "verdict": verdict,
                    "first": jsonable(first), "last": jsonable(value),
                    "detail": {"first_ref": None,
                               "last_ref": ref_str(sheet, row, col)}},
        answer_cells=[(sheet, row, col, level)],
        facts_required=[metric, p_first, p_last],
        facts_refs=[], forbidden=[fmt_number(value)])


def build_judge_hurdle(spec, tpl, rng):
    candidates = []
    for sheet, row, col, level, value in spec.targets:
        if not is_number(value):
            continue
        label = spec.row_label(sheet, row, col)
        if not label:
            continue
        for key, hurdle in HURDLES.items():
            if re.search(r"\b%s\b" % key, label, re.I):
                candidates.append((sheet, row, col, level, value, label, key))
                break
    if not candidates:
        return None
    sheet, row, col, level, value, label, key = rng.choice(candidates)
    hurdle = HURDLES[key]
    threshold = hurdle["threshold"]
    metric_value = value
    interpretation = "as-is"
    if hurdle["percent_aware"] and abs(value) > 1.5:
        metric_value = value / 100.0       # sheet stores percent points
        interpretation = "percent_points"
    verdict = "yes" if compare(hurdle["op"], metric_value, threshold) else "no"
    slots = {"metric": key.upper(), "hurdle": hurdle["display"]}
    return Instance(
        tpl["id"], tpl["financebench"], slots,
        scenario=tpl["prompt"].strip().format(**slots),
        output_format=('Write `{"verdict": "yes"|"no", "value": <the computed '
                       '%s>}` to `/app/answers.json`.' % key.upper()),
        answer_key={"kind": "boolean+value",
                    "tolerance": {"numeric_rel": 0.005},
                    "verdict": verdict, "value": jsonable(value),
                    "detail": {"cell": ref_str(sheet, row, col),
                               "row_label": label, "hurdle": hurdle["display"],
                               "value_interpretation": interpretation}},
        answer_cells=[(sheet, row, col, level)],
        facts_required=[key.upper(), hurdle["display"]],
        facts_refs=[], forbidden=[fmt_number(value)])


def build_trace_driver(spec, tpl, rng):
    candidates = []
    for sheet, row, col, level, value in spec.targets:
        entry = spec.graph.formulas.get(sheet, {}).get((row, col))
        if not entry:
            continue
        match = SUM_RUN_RE.match(entry[0])
        if not match:
            continue
        col_l, row_1, col_r, row_2 = match.groups()
        if col_l.upper() != col_r.upper():
            continue
        if get_column_letter(col) != col_l.upper():
            continue
        rows = range(min(int(row_1), int(row_2)), max(int(row_1), int(row_2)) + 1)
        values = spec.graph.values.get(sheet, {})
        parts = []
        for r in rows:
            part_value = values.get((r, col))
            part_label = spec.row_label(sheet, r, col)
            if is_number(part_value) and part_label:
                parts.append((r, part_label, part_value))
        if len(parts) < 3:
            continue
        metric = spec.row_label(sheet, row, col)
        period_b = spec.period_label(sheet, row, col)
        if not metric or not period_b:
            continue
        series = spec.period_series(sheet, row)
        if col not in series or series.index(col) == 0:
            continue
        col_a = series[series.index(col) - 1]
        period_a = spec.period_label(sheet, row, col_a)
        if not period_a or period_a == period_b:
            continue
        earlier = []
        for r, part_label, part_value in parts:
            v_a = values.get((r, col_a))
            if not is_number(v_a):
                break
            earlier.append((r, part_label, part_value, v_a))
        if len(earlier) != len(parts):
            continue
        candidates.append((sheet, row, col, level, metric,
                           period_a, period_b, earlier))
    if not candidates:
        return None
    sheet, row, col, level, metric, period_a, period_b, parts = rng.choice(candidates)
    deltas = [(label, value_b - value_a) for _, label, value_b, value_a in parts]
    top_label, top_delta = max(deltas, key=lambda d: abs(d[1]))
    slots = {"metric": metric, "period_a": period_a, "period_b": period_b}
    return Instance(
        tpl["id"], tpl["financebench"], slots,
        scenario=tpl["prompt"].strip().format(**slots),
        output_format=('Write `{"line_item": "<name>", "change": <number>}` '
                       'to `/app/answers.json`.'),
        answer_key={"kind": "derived", "tolerance": {"numeric_rel": 0.005},
                    "line_item": top_label, "change": jsonable(top_delta),
                    "detail": {"aggregate": ref_str(sheet, row, col),
                               "components": [
                                   {"label": lbl, "delta": jsonable(d)}
                                   for lbl, d in deltas]}},
        answer_cells=[(sheet, row, col, level)],
        facts_required=[metric, period_a, period_b],
        facts_refs=[], forbidden=[top_label, fmt_number(top_delta)])


BUILDERS = {
    "recon_full": build_recon_full,
    "recon_frontier": build_recon_frontier,
    "extract_named_metric": build_extract,
    "compute_ratio": build_compute_ratio,
    "compute_growth": build_compute_growth,
    "judge_threshold": build_judge_threshold,
    "judge_trajectory": build_judge_trajectory,
    "judge_hurdle": build_judge_hurdle,
    "trace_driver": build_trace_driver,
}


# ---------------------------------------------------------------------------
# stage 3 + 4: naturalizer + fact verifier
# ---------------------------------------------------------------------------

NATURALIZER_SYSTEM = """\
You rewrite spreadsheet-model task briefs into natural, analyst-style
scenarios. Hard rules:
1. Preserve every cell reference, sheet name, line-item label, period label,
   number, threshold and definition EXACTLY as written; they must appear
   verbatim in your rewrite.
2. Do not introduce any new cell references, sheet names, labels or numbers.
3. Never reveal, estimate or hint at any answer value.
4. Keep the deliverable identical to the brief. Stay under 180 words.
Output only the rewritten brief, no preamble."""


def read_env_key(env_file):
    for source in (env_file,):
        path = Path(source)
        if not path.is_file():
            continue
        for line in path.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, _, value = line.partition("=")
            if key.strip().lower() == "lbx_api_key":
                return value.strip().strip('"').strip("'")
    return os.environ.get("LBX_API_KEY") or os.environ.get("lbx_api_key")


def call_naturalizer(base_url, api_key, model, project_id, messages):
    # gpt-5.x chat models reject explicit temperature; default is the only value
    body = json.dumps({"model": model, "messages": messages}).encode("utf-8")
    request = urllib.request.Request(
        base_url.rstrip("/") + "/chat/completions",
        data=body, method="POST",
        headers={
            "Authorization": "Bearer " + api_key,
            "Content-Type": "application/json",
            "x-labelbox-context": json.dumps({"project_id": project_id}),
        })
    with urllib.request.urlopen(request, timeout=180) as response:
        payload = json.loads(response.read().decode("utf-8"))
    return payload["choices"][0]["message"]["content"].strip()


BARE_REF_RE = re.compile(r"\b\$?[A-Z]{1,2}\$?\d{1,5}\b")
QUALIFIED_REF_RE = re.compile(r"(?:'[^']+'|[A-Za-z0-9_]+)!\$?[A-Z]{1,3}\$?\d{1,7}")
NUMBER_RE = re.compile(r"\d+(?:\.\d+)?")


def verify_facts(text, instance):
    """Violation strings, empty when the rewrite is faithful."""
    violations = []
    # everything the skeleton itself states is an allowed fact by construction
    allowed_text = " ".join(instance.facts_required + instance.facts_refs
                            + list(instance.slots.values())
                            + [instance.scenario])

    for fact in instance.facts_required:
        if fact and fact not in text:
            violations.append("missing required fact: %r" % fact)

    allowed_refs = {ref.replace("$", "") for ref in instance.facts_refs}
    for match in QUALIFIED_REF_RE.finditer(text):
        ref = match.group(0).replace("$", "")
        if ref not in allowed_refs and match.group(0) not in allowed_text:
            violations.append("new cell reference: %r" % match.group(0))
    for match in BARE_REF_RE.finditer(text):
        token = match.group(0)
        if token in allowed_text:
            continue
        if token.replace("$", "") in {r.split("!")[-1] for r in allowed_refs}:
            continue
        violations.append("possible new cell reference: %r" % token)

    allowed_numbers = set(NUMBER_RE.findall(allowed_text))
    allowed_numbers.update({"180", "1", "2", "3"})  # counting words are fine
    for token in NUMBER_RE.findall(text):
        if token not in allowed_numbers:
            violations.append("new number: %r" % token)

    for secret in instance.forbidden:
        if secret and len(secret) >= 2 and secret in text \
                and secret not in allowed_text:
            violations.append("leaks answer value: %r" % secret)
    return violations


def naturalize(instance, config, log):
    """(text, metadata) - naturalized scenario or skeleton fallback."""
    meta = {"model": config["model"], "endpoint": config["base_url"],
            "attempts": 0, "naturalized": False, "fallback_reason": ""}
    if not config["enabled"]:
        meta["fallback_reason"] = "naturalizer disabled"
        return instance.scenario, meta
    if not config["api_key"]:
        meta["fallback_reason"] = "no lbx_api_key found"
        log("    naturalizer skipped: no lbx_api_key in .env or environment")
        return instance.scenario, meta

    messages = [{"role": "system", "content": NATURALIZER_SYSTEM},
                {"role": "user", "content": instance.scenario}]
    for attempt in range(1 + NATURALIZE_RETRIES):
        meta["attempts"] = attempt + 1
        try:
            text = call_naturalizer(config["base_url"], config["api_key"],
                                    config["model"], config["project_id"],
                                    messages)
        except (urllib.error.URLError, urllib.error.HTTPError, OSError,
                KeyError, json.JSONDecodeError) as exc:
            meta["fallback_reason"] = "llm call failed: %s" % exc
            log("    naturalizer call failed (%s), using skeleton" % exc)
            return instance.scenario, meta
        violations = verify_facts(text, instance)
        if not violations:
            meta["naturalized"] = True
            return text, meta
        log("    naturalizer attempt %d rejected: %s"
            % (attempt + 1, "; ".join(violations[:4])))
        messages.append({"role": "assistant", "content": text})
        messages.append({"role": "user", "content":
                         "Your rewrite violated these rules, fix them and "
                         "output only the corrected brief:\n- "
                         + "\n- ".join(violations)})
    meta["fallback_reason"] = "fact verification failed after retries"
    return instance.scenario, meta


# ---------------------------------------------------------------------------
# stage 5: harbor emitter
# ---------------------------------------------------------------------------

def toml_value(value):
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (int, float)):
        return fmt_number(value) if isinstance(value, float) else str(value)
    if isinstance(value, dict):
        items = []
        for key, inner in value.items():
            safe = key if re.fullmatch(r"[A-Za-z0-9_-]+", str(key)) else '"%s"' % key
            items.append("%s = %s" % (safe, toml_value(inner)))
        return "{%s}" % ", ".join(items)
    if isinstance(value, (list, tuple)):
        return "[" + ", ".join(toml_value(v) for v in value) + "]"
    text = str(value).replace("\\", "\\\\").replace('"', '\\"')
    text = text.replace("\n", "\\n")
    return '"%s"' % text


def toml_table(name, mapping):
    lines = ["[%s]" % name]
    for key, value in mapping.items():
        if value is None:
            continue
        safe = key if re.fullmatch(r"[A-Za-z0-9_-]+", key) else '"%s"' % key
        lines.append("%s = %s" % (safe, toml_value(value)))
    return "\n".join(lines)


STUB_TEST = """\
#!/bin/bash
# Grader not implemented yet: this pipeline iteration only builds tasks.
# The answer key for the future grader sits next to this script at
# /tests/answer_key.json; the agent's answers land at /app/answers.json.
mkdir -p /logs/verifier
echo 0 > /logs/verifier/reward.txt
"""


def emit_bundle(out_dir, task_name, instance, instruction, snapshot,
                metadata, facts):
    if out_dir.exists():
        shutil.rmtree(out_dir)
    (out_dir / "environment").mkdir(parents=True)
    (out_dir / "tests").mkdir()

    shutil.copy2(snapshot, out_dir / "environment" / snapshot.name)
    (out_dir / "instruction.md").write_text(instruction, encoding="utf-8")

    sections = [
        'schema_version = "1.4"',
        "",
        toml_table("task", {
            "name": task_name,
            "version": "1.0.0",
            "description": metadata["description"],
            "keywords": metadata["keywords"],
        }),
        "",
        toml_table("metadata", {k: v for k, v in metadata.items()
                                if k not in ("description", "keywords",
                                             "slots", "naturalizer")}),
        "",
        toml_table("metadata.slots", instance.slots),
        "",
        toml_table("metadata.naturalizer", metadata["naturalizer"]),
        "",
        # 1.5x the original 1,800-second allowance. Agentic spreadsheet
        # reconstruction now includes paginated MCP research before modelling.
        toml_table("agent", {"timeout_sec": 2700.0}),
        "",
        toml_table("verifier", {"timeout_sec": 300.0}),
        "",
        toml_table("environment", {
            "docker_image": "python:3.11-slim",
            "cpus": 1,
            "memory_mb": 2048,
        }),
        "",
    ]
    (out_dir / "task.toml").write_text("\n".join(sections), encoding="utf-8")

    test_path = out_dir / "tests" / "test.sh"
    test_path.write_text(STUB_TEST, encoding="utf-8")
    test_path.chmod(0o755)
    with open(out_dir / "tests" / "answer_key.json", "w", encoding="utf-8") as fh:
        json.dump(instance.answer_key, fh, indent=1)
    with open(out_dir / "tests" / "facts.json", "w", encoding="utf-8") as fh:
        json.dump(facts, fh, indent=1)


def build_instruction(scenario, artifact_name, output_format):
    return """\
%s

## Input

The workbook `%s` is in your working directory. All static inputs, labels and
shallower calculated values are present; deeper calculated cells are blank.
You may install Python packages (for example `openpyxl`) to read it.

## Output

%s
""" % (scenario.strip(), artifact_name, output_format)


# ---------------------------------------------------------------------------
# driver
# ---------------------------------------------------------------------------

def load_templates(path):
    with open(path, encoding="utf-8") as fh:
        spec = yaml.safe_load(fh)
    return {tpl["id"]: tpl for tpl in spec["templates"]}


def snapshots_for(levels_dir, stem):
    """{cutoff: snapshot path} discovered from level_out/{id}/L*.xlsx."""
    folder = Path(levels_dir) / stem
    found = {}
    if folder.is_dir():
        for path in folder.iterdir():
            match = SNAPSHOT_RE.match(path.name)
            if match:
                found[int(match.group(1))] = path
    return found


def pick_cutoffs(available, deepest, count, rng):
    candidates = sorted(c for c in available if c < deepest)
    if not candidates:
        return []
    if len(candidates) <= count:
        return candidates
    picks = {candidates[int(round(q * (len(candidates) - 1)))]
             for q in (0.25, 0.5, 0.75)[:count]}
    while len(picks) < count:
        extra = rng.choice(candidates)
        picks.add(extra)
    return sorted(picks)


def slugify(text):
    return re.sub(r"[^a-z0-9]+", "-", text.lower()).strip("-")


def process_workbook(path, args, templates, taxonomy, nat_config, made):
    stem = path.stem
    print(path.name, flush=True)
    snapshots = snapshots_for(args.levels_dir, stem)
    if not snapshots:
        print("    skipped: no snapshots under %s/%s (run xl_level_split.py "
              "first)" % (args.levels_dir, stem))
        return 0

    graph = AstGraph(path, max_range_expand=args.max_range_expand,
                     read_values=True, verbose=False)
    try:
        graph.build()
    except Exception as exc:
        print("    failed: %s: %s" % (type(exc).__name__, exc))
        return 0

    deepest = max(snapshots)
    if args.cutoffs:
        cutoffs = [c for c in args.cutoffs if c in snapshots and c < deepest]
        missing = [c for c in args.cutoffs if c not in snapshots]
        if missing:
            print("    no snapshot for cutoff(s) %s" % missing)
    else:
        rng = random.Random("%s:%s:cutoffs" % (args.seed, stem))
        cutoffs = pick_cutoffs(snapshots, deepest, args.per_workbook, rng)
    if not cutoffs:
        print("    skipped: no usable cutoffs")
        return 0

    tax = taxonomy.get(path.name, {})
    wanted = args.templates or [t for t in BUILDERS if t in templates]
    count = 0
    log = (lambda msg: None) if args.quiet else (lambda msg: print(msg, flush=True))

    for cutoff in cutoffs:
        spec = Spec(graph, cutoff)
        if not spec.targets:
            log("    L%02d: no gradable targets" % cutoff)
            continue
        for template_id in wanted:
            if args.max_tasks and made[0] >= args.max_tasks:
                return count
            tpl = templates.get(template_id)
            if tpl is None or template_id in DEFERRED_TEMPLATES:
                continue
            builder = BUILDERS.get(template_id)
            if builder is None:
                continue
            rng = random.Random("%s:%s:%d:%s"
                                % (args.seed, stem, cutoff, template_id))
            try:
                if template_id == "judge_threshold":
                    ratio_library = templates["compute_ratio"]["ratio_library"]
                    instance = builder(spec, tpl, rng, ratio_library)
                else:
                    instance = builder(spec, tpl, rng)
            except Exception as exc:
                log("    L%02d %s: builder failed (%s: %s)"
                    % (cutoff, template_id, type(exc).__name__, exc))
                continue
            if instance is None:
                continue

            scenario, nat_meta = naturalize(instance, nat_config, log)
            snapshot = snapshots[cutoff]
            instruction = build_instruction(scenario, snapshot.name,
                                            instance.output_format)

            answer_levels = [l for _, _, _, l in instance.answer_cells]
            depth = (max(answer_levels) - cutoff) if answer_levels else 0
            fb = instance.financebench or {}
            task_name = "%s/%s-l%02d-%s" % (args.org, stem, cutoff,
                                            slugify(template_id))
            metadata = {
                "description": ("Reconstruct blanked cells of a %s model "
                                "(cutoff L%02d, template %s)"
                                % (tax.get("primary", "financial"), cutoff,
                                   template_id)),
                "keywords": sorted(set(["spreadsheet", template_id]
                                       + tax.get("tags", []))),
                "workbook": stem,
                "source_file": path.name,
                "cutoff": cutoff,
                "deepest_level": deepest,
                "depth": depth,
                "template": template_id,
                "financebench_question_type": fb.get("question_type", ""),
                "financebench_reasoning": fb.get("reasoning", ""),
                "answer_kind": instance.answer_key.get("kind", ""),
                "n_targets_gradable": len(spec.targets),
                "n_answer_cells": len(instance.answer_cells),
                "answer_cells": [ref_str(s, r, c)
                                 for s, r, c, _ in instance.answer_cells[:60]],
                "excluded_cycle": spec.excluded.get("cycle", 0),
                "excluded_volatile": spec.excluded.get("volatile", 0),
                "excluded_unparsed": spec.excluded.get("unparsed", 0),
                "excluded_error_value": spec.excluded.get("error_value", 0),
                "excluded_no_cached_value": spec.excluded.get("no_cached_value", 0),
                "taxonomy_primary": tax.get("primary", ""),
                "taxonomy_tags": tax.get("tags", []),
                "seed": args.seed,
                "pipeline_version": PIPELINE_VERSION,
                "created_at": dt.datetime.now(dt.timezone.utc)
                                .strftime("%Y-%m-%dT%H:%M:%SZ"),
                "naturalizer": nat_meta,
                "slots": instance.slots,
            }
            facts = {
                "required": instance.facts_required,
                "refs": instance.facts_refs,
                "forbidden": instance.forbidden,
                "skeleton": instance.scenario,
            }
            out_dir = Path(args.out) / ("%s-L%02d-%s"
                                        % (stem, cutoff, template_id))
            emit_bundle(out_dir, task_name, instance, instruction,
                        snapshot, metadata, facts)
            made[0] += 1
            count += 1
            log("    L%02d %-22s -> %s%s"
                % (cutoff, template_id, out_dir.name,
                   "" if nat_meta["naturalized"] else "  [skeleton]"))
    return count


def main(argv=None):
    parser = argparse.ArgumentParser(
        description="Emit Harbor task bundles from level-split workbooks")
    parser.add_argument("path", help="workbook, or a directory of workbooks")
    parser.add_argument("--glob", default="*.xls[xm]")
    parser.add_argument("-o", "--out", default="tasks_out")
    parser.add_argument("--levels-dir", default="level_out")
    parser.add_argument("--taxonomy", default="taxonomy_out/workbooks.json")
    parser.add_argument("--templates-file", default="task_templates.yaml")
    parser.add_argument("--templates", default="",
                        help="comma-separated template ids (default: all "
                             "implemented, minus recalc-based ones)")
    parser.add_argument("--cutoff", dest="cutoffs", action="append", type=int,
                        help="explicit cutoff level; repeatable")
    parser.add_argument("--per-workbook", type=int, default=3,
                        help="cutoffs sampled per workbook when --cutoff is "
                             "not given (default: %(default)s)")
    parser.add_argument("--max-tasks", type=int, default=0,
                        help="stop after this many tasks overall (default: no cap)")
    parser.add_argument("--seed", default="42")
    parser.add_argument("--org", default="fcp", help="harbor task org prefix")
    parser.add_argument("--max-range-expand", type=int, default=100)
    parser.add_argument("--no-naturalize", action="store_true",
                        help="skip the LLM pass; instruction uses the skeleton")
    parser.add_argument("--model", default="openai/gpt-5.6-luna",
                        help="model id in the proxy catalog (default: %(default)s)")
    parser.add_argument("--dev", action="store_true",
                        help="use the dev LiteLLM endpoint")
    parser.add_argument("--project-id", default=DEFAULT_PROJECT_ID)
    parser.add_argument("--env-file", default=".env")
    parser.add_argument("-q", "--quiet", action="store_true")
    args = parser.parse_args(argv)

    args.templates = [t.strip() for t in args.templates.split(",") if t.strip()]
    for template_id in args.templates:
        if template_id in DEFERRED_TEMPLATES:
            sys.exit("%s needs headless recalculation for its answer key and "
                     "is deferred to the grader iteration" % template_id)
        if template_id not in BUILDERS:
            sys.exit("unknown template %r (have: %s)"
                     % (template_id, ", ".join(sorted(BUILDERS))))

    templates = load_templates(args.templates_file)
    taxonomy = {}
    if Path(args.taxonomy).is_file():
        with open(args.taxonomy, encoding="utf-8") as fh:
            taxonomy = json.load(fh)
    else:
        print("note: %s not found, taxonomy tags will be empty" % args.taxonomy)

    nat_config = {
        "enabled": not args.no_naturalize,
        "base_url": DEV_ENDPOINT if args.dev else PROD_ENDPOINT,
        "model": args.model,
        "project_id": args.project_id,
        "api_key": read_env_key(args.env_file) if not args.no_naturalize else "",
    }

    root = Path(args.path)
    if root.is_dir():
        targets = sorted(p for p in root.glob(args.glob)
                         if not p.name.startswith("~$"))
    else:
        targets = [root]
    if not targets:
        sys.exit("nothing to do: no workbook matched %s" % args.path)

    started = time.time()
    made = [0]
    for target in targets:
        process_workbook(target, args, templates, taxonomy, nat_config, made)
        if args.max_tasks and made[0] >= args.max_tasks:
            break
        print()
    print("%d task(s) -> %s in %.1fs"
          % (made[0], Path(args.out).resolve(), time.time() - started))


if __name__ == "__main__":
    main()
