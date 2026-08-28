#!/usr/bin/env python3
"""Turn an audited variable-sources table into a mock MCP research environment.

Three subcommands, run in order:

    # 1. Preserve every Markdown row as a draft review queue (audit artifact)
    python3 xl_variable_mcp.py import \\
        runs/0233-variable-sources/0233-inputs-variable-sources.md \\
        runs/0233-variable-sources/draft.json

    # 2. Build the environment from a hand-normalized spec. Every variable's
    #    raw cell value is verified against the golden workbook before anything
    #    is emitted -- the MCP must serve exactly what the masked cells held.
    python3 xl_variable_mcp.py build \\
        runs/0233-variable-sources/normalized.json \\
        runs/0233-variable-sources/mcp \\
        --workbook 0233 --source "4-10 100"

    # 3. Exercise the generated server with an in-memory MCP client
    python3 xl_variable_mcp.py smoke runs/0233-variable-sources/mcp

``build`` writes, under the output directory:

    runtime/            the served data (sources, documents, datasets, records)
    eval/tasks.jsonl    prompts, typed answers, evidence ids -- NEVER shipped
    server.py           standalone FastMCP sidecar (streamable-http :8000/mcp)
    Dockerfile          sidecar image
    mask_cells.json     every workbook cell the masker must additionally blank
    masked_inputs.json  audit map: variable -> refs, raw value, MCP evidence

The normalized spec follows the scenario schema of the build-variable-source-mcp
skill, extended with per-variable ``workbook`` metadata::

    "workbook": {"cells": ["Model!F44", "Control!I62:J62"], "value": 25}

``workbook.value`` is the raw value the cells hold. When omitted it is derived
from ``value``: percent units divide by 100, ISO date strings compare against
date cells, everything else compares as-is.
"""

from __future__ import annotations

import argparse
import datetime
import hashlib
import importlib.util
import json
import re
import sys
import warnings
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from mcp_env.build import (
    access_status,
    build,
    load,
    profile_catalog,
    validate_spec,
    write_json,
)
from mcp_env.import_table import import_table
from mcp_env.server_assets import SERVER_PY, SIDECAR_DOCKERFILE
from mcp_env.validate import validate


# ---------------------------------------------------------------------------
# golden-workbook verification
# ---------------------------------------------------------------------------

def split_ref(ref):
    sheet, _, coords = ref.rpartition("!")
    return sheet.strip("'"), coords


def expand_cells(refs):
    """['Model!F44', 'Control!I62:J62'] -> ['Model!F44', 'Control!I62', ...]."""
    from openpyxl.utils import get_column_letter, range_boundaries
    out = []
    for ref in refs:
        sheet, coords = split_ref(ref)
        min_col, min_row, max_col, max_row = range_boundaries(coords)
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                out.append("%s!%s%d" % (sheet, get_column_letter(col), row))
    return out


def raw_value(variable):
    workbook = variable.get("workbook") or {}
    if "value" in workbook:
        return workbook["value"]
    value = variable["value"]
    if str(variable.get("unit", "")).casefold() == "percent" and \
            isinstance(value, (int, float)) and not isinstance(value, bool):
        return value / 100.0
    return value


def value_matches(cell_value, expected):
    if cell_value is None:
        return False
    if isinstance(expected, str):
        try:
            wanted = datetime.date.fromisoformat(expected)
        except ValueError:
            return str(cell_value).strip().casefold() == expected.strip().casefold()
        if isinstance(cell_value, datetime.datetime):
            return cell_value.date() == wanted
        if isinstance(cell_value, datetime.date):
            return cell_value == wanted
        return False
    if isinstance(cell_value, bool) or not isinstance(cell_value, (int, float)):
        return False
    return abs(float(cell_value) - float(expected)) <= \
        1e-9 * max(1.0, abs(float(expected)))


def verify_against_golden(spec, golden_path):
    """Every masked cell must hold exactly the value the MCP will serve."""
    import openpyxl
    warnings.simplefilter("ignore")
    book = openpyxl.load_workbook(golden_path, data_only=True)
    faults = []
    for variable in spec["variables"]:
        workbook = variable.get("workbook") or {}
        refs = workbook.get("cells") or []
        if not refs:
            continue
        expected = raw_value(variable)
        for ref in expand_cells(refs):
            sheet, coords = split_ref(ref)
            if sheet not in book.sheetnames:
                faults.append("%s: unknown sheet in %s" % (variable["id"], ref))
                continue
            actual = book[sheet][coords].value
            if not value_matches(actual, expected):
                faults.append("%s: %s holds %r, spec says %r"
                              % (variable["id"], ref, actual, expected))
    book.close()
    return faults


# ---------------------------------------------------------------------------
# leak-patch proposals (leakscan --emit-patch)
# ---------------------------------------------------------------------------

def _slug(text):
    """gen_normalizer's variable-id slug, kept in sync for draft recovery."""
    out = re.sub(r"[^a-z0-9]+", "-", str(text).lower()).strip("-")
    return out or "variable"


def draft_chain(variable_id, refs):
    """Best-effort draft_id behind a normalized per-cell variable id.

    gen_normalizer derives per-cell variable ids as
    ``{draft-slug}-{sheet-slug}-{coord}[-{n}]``; single-cell draft rows keep
    the draft slug unchanged.  Stripping a matching cell suffix recovers the
    draft id so the per-workbook normalize script can update dispositions.
    """
    for ref in refs:
        sheet, coords = split_ref(ref)
        suffix = "-%s-%s" % (_slug(sheet), coords.lower())
        match = re.fullmatch(
            "(?P<base>.+)%s(?:-\\d+)?" % re.escape(suffix), variable_id)
        if match:
            return match.group("base")
    return variable_id


def _trailing_label(text):
    """Text following a simple numeric rendering, or None when the string is
    not number-shaped at all (e.g. an exact string duplicate)."""
    from xl_mcp_oracle import SIMPLE_NUMBER_RE
    match = SIMPLE_NUMBER_RE.fullmatch(text)
    if match is None:
        return None
    end = max(match.end("number"), match.end("percent"))
    return text[end:].strip(" \t)")


def _leak_cell_state(book, cell_ref):
    from xl_mcp_oracle import normalized_ref
    sheet, _, coords = normalized_ref(cell_ref).rpartition("!")
    cell = book[sheet][coords]
    is_formula = cell.data_type == "f" or (
        isinstance(cell.value, str) and cell.value.startswith("="))
    return cell.value, is_formula


def build_leak_patch(report, audit, workbook_path):
    """Deterministic per-leak fix proposals for the normalize script.

    Exactly one proposal per unapproved leak (approved/allowlisted leaks get
    none):

    * ``mask``       the leaking cell is typed (non-formula) and holds the
                     variable's raw value, so golden verification would accept
                     it: append the cell to that variable's ``workbook.cells``.
                     When several variables share the raw value in one cell,
                     only the first variable_id gets ``mask``.
    * ``extra_cell`` the value appears as a non-comparable rendering
                     (formatted text, percent, or date) with no other needed
                     content: blank it via ``workbook.extra_cells``.
    * ``exclude``    the value cannot be blanked safely (formula cell, text
                     mixing the value with other required content, or the cell
                     is already claimed by another variable): exclude the
                     variable, surfacing the draft_id chain so the normalizer
                     can update its dispositions.

    The patch is a PROPOSAL only; nothing is applied to normalized.json.
    ``generated_at`` is derived from the scanned workbook's mtime and
    ``source_report_sha256`` hashes the report exactly as ``--report`` writes
    it, so identical inputs yield byte-identical patches.
    """
    import openpyxl
    warnings.simplefilter("ignore")

    entries = {entry["variable_id"]: entry for entry in audit}
    leaks = report.get("unapproved_value_leaks") or []

    states = {}
    if leaks:
        book = openpyxl.load_workbook(
            workbook_path, data_only=False, read_only=True)
        try:
            for leak in leaks:
                if leak["cell"] not in states:
                    states[leak["cell"]] = _leak_cell_state(book, leak["cell"])
        finally:
            book.close()

    maskable = {}
    for leak in leaks:
        value, is_formula = states[leak["cell"]]
        entry = entries[leak["variable_id"]]
        if not is_formula and value_matches(value, entry["cell_value"]):
            maskable.setdefault(leak["cell"], []).append(leak["variable_id"])

    proposals = []
    for leak in sorted(leaks, key=lambda row: (row["variable_id"], row["cell"])):
        variable_id = leak["variable_id"]
        cell = leak["cell"]
        rendering = leak.get("rendering", "other")
        refs = sorted(entries[variable_id].get("refs") or [])
        source_cell = refs[0] if refs else "<unknown>"
        chain = draft_chain(variable_id, refs)
        value, is_formula = states[cell]

        if is_formula:
            action = "exclude"
            reason = (
                "%s is a formula cell that reproduces the value and cannot "
                "be blanked as a typed input; draft chain: %s -> %s"
                % (cell, chain, variable_id))
        elif variable_id in maskable.get(cell, []):
            winner = min(maskable[cell])
            if variable_id == winner:
                action = "mask"
                reason = (
                    "typed cell %s holds the raw value of %s; append it to "
                    "workbook.cells" % (cell, variable_id))
            else:
                action = "exclude"
                reason = (
                    "cell %s holds the shared raw value but can only serve "
                    "one variable (mask proposed for %s); draft chain: "
                    "%s -> %s" % (cell, winner, chain, variable_id))
        else:
            label = _trailing_label(value) if isinstance(value, str) else ""
            if label is not None and len(label.split()) > 2:
                action = "exclude"
                reason = (
                    "text cell %s mixes a %s rendering of the value with "
                    "other required content (%r); draft chain: %s -> %s"
                    % (cell, rendering, label, chain, variable_id))
            else:
                action = "extra_cell"
                reason = (
                    "%s rendering of %s (source cell %s) at %s; blank it "
                    "via workbook.extra_cells"
                    % (rendering, variable_id, source_cell, cell))

        proposals.append({
            "variable_id": variable_id,
            "cell": cell,
            "rendering": rendering,
            "action": action,
            "reason": reason,
        })

    report_bytes = (
        json.dumps(report, indent=2, ensure_ascii=False, sort_keys=True) + "\n"
    ).encode("utf-8")
    generated_at = datetime.datetime.fromtimestamp(
        Path(workbook_path).stat().st_mtime, datetime.timezone.utc
    ).replace(microsecond=0).isoformat().replace("+00:00", "Z")
    return {
        "workbook": str(workbook_path),
        "generated_at": generated_at,
        "source_report_sha256": hashlib.sha256(report_bytes).hexdigest(),
        "proposals": proposals,
    }


# ---------------------------------------------------------------------------
# subcommands
# ---------------------------------------------------------------------------

def cmd_import(args):
    result = import_table(Path(args.input))
    output = Path(args.output)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(result, indent=2, ensure_ascii=False) + "\n",
                      encoding="utf-8")
    print(json.dumps({"rows": result["row_count"], "output": str(output)}, indent=2))
    return 0


def cmd_validate_spec(args):
    """Validate normalized variables and reviewed profiles without building."""
    spec = load(Path(args.spec))
    validate_spec(spec)
    profiles = profile_catalog(spec)
    public = sum(access_status(profile) == "public"
                 for profile in profiles.values())
    skipped = len(profiles) - public
    print(json.dumps({
        "valid": True,
        "variables": len(spec["variables"]),
        "source_profiles": len(profiles),
        "approved_public_profiles": public,
        "skipped_profiles": skipped,
    }, indent=2))
    return 0


def cmd_build(args):
    spec = load(Path(args.spec))
    validate_spec(spec)

    golden = None
    if args.workbook:
        for suffix in (".xlsx", ".xlsm"):
            candidate = Path(args.source) / (args.workbook + suffix)
            if candidate.exists():
                golden = candidate
                break
        if golden is None:
            sys.exit("no golden workbook %s under %s" % (args.workbook, args.source))
        faults = verify_against_golden(spec, golden)
        if faults:
            for fault in faults:
                print("  GOLDEN MISMATCH  %s" % fault)
            sys.exit("%d golden-value mismatch(es); fix the spec" % len(faults))

    output = Path(args.output)
    summary = build(spec, output)
    (output / "server.py").write_text(SERVER_PY, encoding="utf-8")
    (output / "Dockerfile").write_text(SIDECAR_DOCKERFILE, encoding="utf-8")

    mask_refs = sorted({
        ref
        for variable in spec["variables"]
        for ref in expand_cells(
            ((variable.get("workbook") or {}).get("cells") or [])
            + ((variable.get("workbook") or {}).get("extra_cells") or []))
    })
    write_json(output / "mask_cells.json", mask_refs)

    tasks = {row["task_id"]: row for row in (
        json.loads(line)
        for line in (output / "eval" / "tasks.jsonl").read_text(
            encoding="utf-8").splitlines() if line.strip())}
    audit = []
    for variable in spec["variables"]:
        workbook = variable.get("workbook") or {}
        refs = workbook.get("cells") or []
        if not refs:
            continue
        raw = raw_value(variable)
        if isinstance(raw, (datetime.date, datetime.datetime)):
            raw = raw.isoformat()
        audit.append({
            "variable_id": variable["id"],
            "name": variable["name"],
            "refs": expand_cells(refs + (workbook.get("extra_cells") or [])),
            "cell_value": raw,
            "mcp_value": variable["value"],
            "unit": variable["unit"],
            "question": variable["question"],
            "evidence": tasks[variable["id"]]["evidence"],
        })
    write_json(output / "masked_inputs.json", audit)

    report = validate(output)
    summary.update({
        "validation": report,
        "mask_cells": len(mask_refs),
        "golden_checked_against": str(golden) if golden else None,
    })
    print(json.dumps(summary, indent=2))
    return 0


def cmd_leakscan(args):
    """Run the HTTP oracle's workbook duplicate/leak detector offline."""
    from xl_mcp_oracle import scan_value_leaks

    if args.mcp:
        mcp = Path(args.mcp)
        mask_refs = load(mcp / "mask_cells.json")
        audit = load(mcp / "masked_inputs.json")
    else:
        spec = load(Path(args.spec))
        mask_refs = sorted({
            ref
            for variable in spec["variables"]
            for ref in expand_cells(
                ((variable.get("workbook") or {}).get("cells") or [])
                + ((variable.get("workbook") or {}).get("extra_cells") or []))
        })
        audit = []
        for variable in spec["variables"]:
            workbook = variable.get("workbook") or {}
            refs = workbook.get("cells") or []
            if not refs:
                continue
            raw = raw_value(variable)
            if isinstance(raw, (datetime.date, datetime.datetime)):
                raw = raw.isoformat()
            audit.append({
                "variable_id": variable["id"],
                "refs": expand_cells(refs + (workbook.get("extra_cells") or [])),
                "cell_value": raw,
                "mcp_value": variable["value"],
                "unit": variable["unit"],
            })
    if not mask_refs or not audit:
        sys.exit("nothing to scan: the spec/mcp dir yields no masked cells")

    if args.inputs:
        target = Path(args.inputs)
    elif args.workbook:
        target = None
        for suffix in (".xlsx", ".xlsm"):
            candidate = Path(args.source) / (args.workbook + suffix)
            if candidate.exists():
                target = candidate
                break
        if target is None:
            sys.exit("no golden workbook %s under %s"
                     % (args.workbook, args.source))
    else:
        sys.exit("pass --inputs or --workbook to choose the workbook to scan")
    if not target.is_file():
        sys.exit("workbook to scan does not exist: %s" % target)

    report = scan_value_leaks(
        target, mask_refs, audit,
        Path(args.allowlist) if args.allowlist else None,
        require_blank_mask=args.require_blank)
    report["workbook_scanned"] = str(target)
    if args.report:
        write_json(Path(args.report), report)
    if args.emit_patch:
        write_json(Path(args.emit_patch),
                   build_leak_patch(report, audit, target))
    print(json.dumps(report, indent=2, ensure_ascii=False))
    return 0 if report["valid"] else 1


def cmd_smoke(args):
    import asyncio

    bundle = Path(args.bundle)
    spec = importlib.util.spec_from_file_location("generated_mcp_server",
                                                  bundle / "server.py")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    server = module.build_server(bundle / "runtime")

    async def run():
        from fastmcp import Client

        async def query_all(client, args, max_pages=40):
            rows, cursor = [], 0
            for _ in range(max_pages):
                result = await client.call_tool(
                    "query_records", {**args, "limit": 5, "cursor": cursor})
                rows.extend(result.data["rows"])
                cursor = result.data.get("next_cursor")
                if cursor is None:
                    break
            return rows

        async with Client(server) as client:
            tools = sorted(tool.name for tool in await client.list_tools())
            datasets = await client.call_tool("list_datasets", {"limit": 100})
            checks = {"tools": tools, "datasets": len(datasets.data), "resolved": 0,
                      "broad_conflicts": 0, "chains_seen": 0, "failures": []}
            tasks = [json.loads(line) for line in
                     (bundle / "eval" / "tasks.jsonl").read_text(
                         encoding="utf-8").splitlines() if line.strip()]
            for task in tasks:
                wanted = task["required_dimensions"]
                evidence = task["evidence"]
                metric = wanted["metric"]
                rows = await query_all(client, {
                    "dataset_id": evidence["dataset_id"], "metric": metric,
                    "entity": wanted["entity"], "period": str(wanted["period"]),
                    "scenario": wanted["scenario"], "basis": wanted["basis"],
                    "unit": wanted["unit"], "status": wanted["status"],
                })
                current = [row for row in rows if not row.get("superseded_by")]
                if len(rows) > 1:
                    checks["chains_seen"] += 1
                if len(current) == 1 and \
                        current[0]["id"] == evidence["record_id"] and \
                        current[0]["value"] == task["answer"]["value"]:
                    checks["resolved"] += 1
                else:
                    checks["failures"].append(task["task_id"])
                broad = await query_all(client, {
                    "dataset_id": evidence["dataset_id"], "metric": metric,
                    "entity": wanted["entity"],
                })
                values = {json.dumps(row["value"], sort_keys=True)
                          for row in broad}
                if len(values) > 1:
                    checks["broad_conflicts"] += 1
            return checks

    checks = asyncio.run(run())
    print(json.dumps(checks, indent=2))
    return 1 if checks["failures"] else 0


def main(argv=None):
    parser = argparse.ArgumentParser(
        description=__doc__.split("\n")[0],
        formatter_class=argparse.RawDescriptionHelpFormatter)
    sub = parser.add_subparsers(dest="command", required=True)

    p_import = sub.add_parser("import", help="Markdown table -> draft review queue")
    p_import.add_argument("input")
    p_import.add_argument("output")
    p_import.set_defaults(func=cmd_import)

    p_build = sub.add_parser("build", help="normalized spec -> MCP bundle")
    p_build.add_argument("spec")
    p_build.add_argument("output")
    p_build.add_argument("--workbook", default="",
                         help="golden workbook id to verify raw cell values against")
    p_build.add_argument("--source", default="4-10 100",
                         help="folder holding the golden <wb>.xlsx")
    p_build.set_defaults(func=cmd_build)

    p_validate = sub.add_parser(
        "validate-spec",
        help="validate normalized variables and reviewed source profiles")
    p_validate.add_argument("spec")
    p_validate.set_defaults(func=cmd_validate_spec)

    p_leak = sub.add_parser(
        "leakscan",
        help="offline oracle duplicate/leak scan of a workbook (no server)")
    p_leak.add_argument("spec", help="normalized spec used to derive the mask "
                                     "and masked-value targets")
    p_leak.add_argument("--workbook", default="",
                        help="golden workbook id to scan when --inputs is absent")
    p_leak.add_argument("--source", default="4-10 100",
                        help="folder holding the golden <wb>.xlsx")
    p_leak.add_argument("--inputs", default="",
                        help="inputs workbook to scan instead of the golden")
    p_leak.add_argument("--mcp", default="",
                        help="built MCP dir; use its mask_cells.json and "
                             "masked_inputs.json instead of the spec")
    p_leak.add_argument("--allowlist", default="",
                        help="explicit JSON allowlist for legitimate duplicates")
    p_leak.add_argument("--report", default="",
                        help="also write the JSON report to this path")
    p_leak.add_argument("--emit-patch", default="",
                        help="write a deterministic leak_patch.json proposing "
                             "one mask/extra_cell/exclude action per "
                             "unapproved leak (a proposal for the normalize "
                             "script; never auto-applied)")
    p_leak.add_argument("--require-blank", action="store_true",
                        help="also require every masked cell to be blank "
                             "(when scanning an already-masked workbook)")
    p_leak.set_defaults(func=cmd_leakscan)

    p_smoke = sub.add_parser("smoke", help="exercise the server with an MCP client")
    p_smoke.add_argument("bundle")
    p_smoke.set_defaults(func=cmd_smoke)

    args = parser.parse_args(argv)
    return args.func(args)


if __name__ == "__main__":
    sys.exit(main())
