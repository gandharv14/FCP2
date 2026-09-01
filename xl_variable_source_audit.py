#!/usr/bin/env python3
"""Generate auditable variable-to-source Markdown from an inputs workbook.

The deterministic half of this stage extracts compact input-band summaries from
``seg_out/<workbook>/segments.json`` and the corresponding inputs-only workbook.
GPT 5.6 Sol then selects externally sourced candidates and proposes plausible
source classes through Labelbox's LiteLLM endpoint.
"""

from __future__ import annotations

import argparse
import datetime as dt
import decimal
import hashlib
import json
import math
import os
import re
import sys
import urllib.error
import urllib.request
from pathlib import Path

import openpyxl
from openpyxl.utils import range_boundaries

from xl_task_build import DEFAULT_PROJECT_ID, PROD_ENDPOINT, read_env_key


GENERATOR_VERSION = "1.0.0"
PROMPT_VERSION = "variable-source-audit-v1"
DEFAULT_MODEL = "openai/gpt-5.6-sol"
MAX_CHUNK_CHARS = 100_000

SYSTEM_PROMPT = """\
You are GPT 5.6 Sol producing a review artifact for a financial-spreadsheet
pipeline. The user message contains workbook metadata and deterministic NDJSON
inventory rows extracted from an inputs-only workbook.

Identify inputs that are plausibly externally sourced: market and macro data,
tax or regulatory rates, contractual and financing terms, asset facts,
historical accounting balances, transaction terms, and operating data that
would normally come from filings, contracts, advisers, or authoritative data
providers. Exclude labels, blank cells, calculations, and purely internal model
choices unless a contract, filing, or external study is a plausible source.

Return exactly one Markdown table with these columns:
| Variable | Workbook cells and values | Plausible external source(s) |

Rules:
1. Use only cell ranges and values present in the supplied inventory. Never
   alter, infer, or invent a workbook value.
2. Group related period series or repeated assumptions when that improves
   auditability, but preserve the supplied ranges and compact value summaries.
3. Sources are candidates, not proof. Say when exact-value provenance is
   unverified or likely private/internal.
4. Prefer authoritative organizations, filings, contracts, and official data
   portals. Add a URL only when confident it is the canonical public page;
   otherwise name the source without fabricating a link.
5. Do not include a preamble, headings, commentary, or fenced code block.
If the batch has no externally sourced candidates, return the table header and
one row stating that no candidates were identified."""


def _json_value(value):
    if isinstance(value, dt.datetime):
        return value.isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        return str(value)
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    return str(value)


def _value_summary(values):
    values = [_json_value(value) for value in values if value is not None]
    if not values:
        return {"count": 0}
    if all(value == values[0] for value in values):
        return {"count": len(values), "value": values[0]}
    if len(values) <= 16:
        return {"count": len(values), "values": values}
    numeric = [
        float(value) for value in values
        if isinstance(value, (int, float)) and not isinstance(value, bool)
    ]
    summary = {
        "count": len(values),
        "first": values[:3],
        "last": values[-3:],
        "distinct_count": len({json.dumps(value, sort_keys=True)
                               for value in values}),
    }
    if len(numeric) == len(values):
        summary.update({"minimum": min(numeric), "maximum": max(numeric)})
    return summary


def _cell_value(book, ref):
    sheet, _, coord = ref.rpartition("!")
    sheet = sheet.strip("'")
    if not sheet or sheet not in book.sheetnames:
        return None
    return book[sheet][coord].value


def build_inventory(workbook, artifact, seg_dir, *, segmentation=None):
    """Return a compact, deterministic inventory of relevant input bands."""
    segments = json.loads(
        (Path(seg_dir) / "segments.json").read_text(encoding="utf-8")
    )
    book = openpyxl.load_workbook(artifact, data_only=True, read_only=True)
    try:
        rows = []
        seen = set()
        candidates = list(segments.get("inputs") or [])
        candidates.extend(segments.get("embedded_literals") or [])
        for entry in candidates:
            band = str(entry.get("band") or "")
            if not band or band in seen or entry.get("kind") == "label":
                continue
            seen.add(band)
            cells = entry.get("cells") or []
            values = [_cell_value(book, ref) for ref in cells]
            summary = _value_summary(values)
            # Empty formula-host bands carry no auditable supplied value. A
            # promoted literal has its hardcoded value in the label/band.
            if not values and entry.get("kind") != "literal":
                continue
            rows.append({
                "band": band,
                "sheet": entry.get("sheet") or band.partition("!")[0],
                "label": entry.get("label") or "",
                "kind": entry.get("kind") or "",
                "value_type": entry.get("vtype") or "",
                "value_summary": summary,
            })
        rows.sort(key=lambda row: (
            str(row["label"]).casefold(), str(row["sheet"]).casefold(),
            str(row["band"]),
        ))
        props = book.properties
        metadata = {
            "workbook": workbook,
            "artifact": Path(artifact).name,
            "sheet_names": list(book.sheetnames),
            "document_properties": {
                key: _json_value(getattr(props, key, None))
                for key in ("title", "subject", "creator", "description",
                            "keywords", "category", "company")
                if getattr(props, key, None)
            },
        }
        if segmentation is not None:
            metadata["segmentation"] = segmentation
        return {"metadata": metadata, "variables": rows}
    finally:
        book.close()


def _split_chunks(inventory, max_chars=MAX_CHUNK_CHARS):
    header = json.dumps(inventory["metadata"], ensure_ascii=False,
                        separators=(",", ":"))
    chunks, current, size = [], [], len(header)
    for row in inventory["variables"]:
        encoded = json.dumps(row, ensure_ascii=False, separators=(",", ":"))
        if current and size + len(encoded) + 1 > max_chars:
            chunks.append(current)
            current, size = [], len(header)
        current.append(encoded)
        size += len(encoded) + 1
    if current or not chunks:
        chunks.append(current)
    return [(header, rows) for rows in chunks]


def _call_model(endpoint, api_key, model, project_id, messages):
    body = json.dumps({"model": model, "messages": messages}).encode("utf-8")
    request = urllib.request.Request(
        endpoint.rstrip("/") + "/chat/completions",
        data=body,
        method="POST",
        headers={
            "Authorization": "Bearer " + api_key,
            "Content-Type": "application/json",
            "x-labelbox-context": json.dumps({"project_id": project_id}),
        },
    )
    with urllib.request.urlopen(request, timeout=600) as response:
        payload = json.loads(response.read().decode("utf-8"))
    choice = payload["choices"][0]
    return choice["message"]["content"].strip(), choice.get("finish_reason", "")


def _clean_table(text):
    text = text.strip()
    if text.startswith("```") and text.endswith("```"):
        lines = text.splitlines()
        text = "\n".join(lines[1:-1]).strip()
    match = re.search(
        r"(?im)^\|\s*Variable\s*\|\s*Workbook cells and values\s*\|"
        r"\s*Plausible external source\(s\)\s*\|",
        text,
    )
    if not match:
        raise ValueError("model response did not contain the required table")
    return text[match.start():].strip()


VALUE_NUMBER_RE = re.compile(
    r"(?<![A-Za-z])[-+]?(?:\d+\.\d+|\d+)(?:[eE][-+]?\d+)?"
)


def _normalized_number(token):
    try:
        return decimal.Decimal(token).normalize()
    except decimal.InvalidOperation:
        return None


def validate_table(table, encoded_rows, workbook_metadata):
    """Return hallucination violations for refs and values in a GPT table."""
    rows = [json.loads(row) for row in encoded_rows]
    allowed_cells = set()
    for row in rows:
        band = str(row["band"]).split("#lit=", 1)[0]
        sheet, _, coords = band.rpartition("!")
        sheet = sheet.strip("'")
        try:
            min_col, min_row, max_col, max_row = range_boundaries(coords)
        except ValueError:
            continue
        for row_num in range(min_row, max_row + 1):
            for col_num in range(min_col, max_col + 1):
                allowed_cells.add((sheet, row_num, col_num))

    violations = []
    sheets = sorted(workbook_metadata.get("sheet_names") or [],
                    key=len, reverse=True)
    if sheets:
        alternatives = "|".join(
            "(?:%s|'%s')" % (re.escape(sheet), re.escape(sheet))
            for sheet in sheets
        )
        ref_re = re.compile(
            r"(?<![A-Za-z0-9_])(?P<sheet>%s)!"
            r"(?P<coords>[A-Z]{1,3}\d+(?::[A-Z]{1,3}\d+)?)"
            % alternatives
        )
        for match in ref_re.finditer(table):
            sheet = match.group("sheet").strip("'")
            try:
                min_col, min_row, max_col, max_row = range_boundaries(
                    match.group("coords")
                )
            except ValueError:
                violations.append("invalid reference: %s" % match.group(0))
                continue
            unknown = [
                (row_num, col_num)
                for row_num in range(min_row, max_row + 1)
                for col_num in range(min_col, max_col + 1)
                if (sheet, row_num, col_num) not in allowed_cells
            ]
            if unknown:
                violations.append("reference not in inventory: %s"
                                  % match.group(0))

    allowed_numbers = {
        value for value in (
            _normalized_number(token)
            for token in VALUE_NUMBER_RE.findall("\n".join(encoded_rows))
        )
        if value is not None
    }
    for line in table.splitlines():
        if not line.lstrip().startswith("|"):
            continue
        columns = line.strip().strip("|").split("|")
        if len(columns) < 3 or "Workbook cells and values" in columns[1]:
            continue
        for token in VALUE_NUMBER_RE.findall(columns[1]):
            value = _normalized_number(token)
            if value is not None and value not in allowed_numbers:
                violations.append(
                    "value not in inventory: %s (row %s)"
                    % (token, columns[0].strip())
                )
    return list(dict.fromkeys(violations))


def _canonical_hash(value):
    encoded = json.dumps(value, ensure_ascii=False, sort_keys=True,
                         separators=(",", ":")).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def generate_audit(
        workbook, artifact, seg_dir, output_dir, api_key, *,
        model=DEFAULT_MODEL, endpoint=PROD_ENDPOINT,
        project_id=DEFAULT_PROJECT_ID, refresh=False, log=print,
        segmentation=None):
    """Generate (or reuse) the Markdown audit and return its metadata."""
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    stem = "%s-inputs-variable-sources" % workbook
    markdown_path = output_dir / (stem + ".md")
    inventory_path = output_dir / (stem + ".inventory.json")
    metadata_path = output_dir / (stem + ".metadata.json")

    inventory = build_inventory(
        workbook, artifact, seg_dir, segmentation=segmentation
    )
    inventory_hash = _canonical_hash(inventory)
    inventory_path.write_text(
        json.dumps(inventory, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )

    if not refresh and markdown_path.is_file() and metadata_path.is_file():
        previous = json.loads(metadata_path.read_text(encoding="utf-8"))
        if (
            previous.get("status") == "complete"
            and previous.get("inventory_sha256") == inventory_hash
            and previous.get("model") == model
            and previous.get("prompt_version") == PROMPT_VERSION
            and (previous.get("validation") or {}).get("status") == "passed"
        ):
            result = dict(previous)
            result["cache_hit"] = True
            log("variable-source audit reused: %s" % markdown_path)
            return result

    if not api_key:
        raise RuntimeError(
            "no lbx_api_key found; use --no-variable-source-audit to opt out"
        )

    chunks = _split_chunks(inventory)
    generated_at = dt.datetime.now(dt.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    tables, finishes = [], []
    try:
        for index, (header, rows) in enumerate(chunks, 1):
            user = (
                "Workbook metadata:\n%s\n\nInventory batch %d of %d (NDJSON):\n%s"
                % (header, index, len(chunks), "\n".join(rows))
            )
            log("variable-source audit %s: GPT batch %d/%d"
                % (workbook, index, len(chunks)))
            messages = [{"role": "system", "content": SYSTEM_PROMPT},
                        {"role": "user", "content": user}]
            for attempt in range(2):
                text, finish = _call_model(
                    endpoint, api_key, model, project_id, messages
                )
                finishes.append(finish)
                table = _clean_table(text)
                violations = validate_table(
                    table, rows, inventory["metadata"]
                )
                if not violations:
                    tables.append(table)
                    break
                if attempt:
                    raise ValueError("; ".join(violations[:8]))
                log("variable-source audit %s: batch %d rejected; retrying"
                    % (workbook, index))
                messages.extend([
                    {"role": "assistant", "content": text},
                    {"role": "user", "content":
                     "Correct the table. These items were not present in the "
                     "inventory:\n- " + "\n- ".join(violations[:20])},
                ])
    except (urllib.error.URLError, urllib.error.HTTPError, OSError, KeyError,
            json.JSONDecodeError, ValueError) as exc:
        failure = {
            "schema_version": 1,
            "generator_version": GENERATOR_VERSION,
            "prompt_version": PROMPT_VERSION,
            "status": "failed",
            "workbook": workbook,
            "model": model,
            "endpoint": endpoint,
            "inventory_sha256": inventory_hash,
            "error": str(exc),
            "generated_at": generated_at,
        }
        metadata_path.write_text(
            json.dumps(failure, indent=2) + "\n", encoding="utf-8"
        )
        raise RuntimeError("variable-source audit failed: %s" % exc) from exc

    warning = (
        "This model-generated audit maps workbook-supplied values to plausible "
        "sources. It does not prove that any source supplied the exact value; "
        "review source claims before using the table as evidence."
    )
    sections = [
        "# `%s` — variable values and plausible external sources"
        % Path(artifact).name,
        "",
        "Generated by `%s` through the Labelbox LiteLLM proxy on %s."
        % (model, generated_at),
        "",
        "> **Audit warning:** %s" % warning,
        "",
    ]
    for index, table in enumerate(tables, 1):
        if len(tables) > 1:
            sections.extend(["## Inventory batch %d of %d"
                             % (index, len(tables)), ""])
        sections.extend([table, ""])
    markdown_path.write_text("\n".join(sections).rstrip() + "\n",
                             encoding="utf-8")

    metadata = {
        "schema_version": 1,
        "generator_version": GENERATOR_VERSION,
        "prompt_version": PROMPT_VERSION,
        "status": "complete",
        "workbook": workbook,
        "artifact": str(artifact),
        "model": model,
        "endpoint": endpoint,
        "project_id": project_id,
        "generated_at": generated_at,
        "inventory_sha256": inventory_hash,
        "inventory_rows": len(inventory["variables"]),
        "api_calls": len(finishes),
        "finish_reasons": finishes,
        "validation": {
            "status": "passed",
            "checks": ["qualified_cell_refs", "workbook_value_numbers"],
        },
        "markdown": str(markdown_path),
        "inventory": str(inventory_path),
        "cache_hit": False,
    }
    metadata_path.write_text(
        json.dumps(metadata, indent=2) + "\n", encoding="utf-8"
    )
    return metadata


def resolve_segmentation_directory(
        seg_root, workbook, *, segmentation_generation_id=None,
        source_generation_root=None, source_generation_id=None):
    """Resolve a pinned inactive generation, or the explicit legacy directory."""
    seg_dir = Path(seg_root) / workbook
    pinned = (
        segmentation_generation_id,
        source_generation_root,
        source_generation_id,
    )
    if not any(pinned):
        return seg_dir, None
    if not all(pinned):
        raise RuntimeError(
            "pinned audit requires segmentation generation ID, source "
            "generation root, and source generation ID"
        )

    from xl_seg.publication import resolve_generation_by_id
    from xl_source_publication import resolve_source_generation_by_id

    source_dir, source_manifest = resolve_source_generation_by_id(
        Path(source_generation_root) / workbook,
        source_generation_id,
    )
    generation_dir, generation_manifest = resolve_generation_by_id(
        seg_dir,
        segmentation_generation_id,
        source_generation_dir=source_dir,
        require_pass=True,
    )
    return generation_dir, {
        "generation_id": generation_manifest["generation_id"],
        "inventory_approval_sha256": (
            source_manifest.get("bindings") or {}
        ).get("inventory_approval_sha256"),
        "manifest_sha256": hashlib.sha256(
            (generation_dir / "generation-manifest.json").read_bytes()
        ).hexdigest(),
        "recalc_signals_sha256": (
            source_manifest.get("bindings") or {}
        ).get("recalc_signals_sha256"),
        "source_generation_id": source_manifest["generation_id"],
    }


def main(argv=None):
    parser = argparse.ArgumentParser(
        description="Generate variable/source audit Markdown with GPT 5.6 Sol"
    )
    parser.add_argument("workbooks", nargs="+")
    parser.add_argument("--inputs-root", default="inputs_out")
    parser.add_argument("--seg-root", default="seg_out")
    parser.add_argument("--segmentation-generation-id")
    parser.add_argument("--source-generation-root")
    parser.add_argument("--source-generation-id")
    parser.add_argument("--audit-root", default="runs")
    parser.add_argument("--env-file", default=".env")
    parser.add_argument("--model", default=DEFAULT_MODEL)
    parser.add_argument("--endpoint", default=PROD_ENDPOINT)
    parser.add_argument("--project-id", default=DEFAULT_PROJECT_ID)
    parser.add_argument("--refresh", action="store_true")
    parser.add_argument(
        "--inventory-only", action="store_true",
        help="write deterministic inventories without calling the model",
    )
    args = parser.parse_args(argv)

    if args.segmentation_generation_id and len(args.workbooks) != 1:
        parser.error("a pinned segmentation generation requires one workbook")

    api_key = "" if args.inventory_only else read_env_key(args.env_file)
    results = []
    for workbook in args.workbooks:
        artifact = Path(args.inputs_root) / ("%s-inputs.xlsx" % workbook)
        output_dir = Path(args.audit_root) / ("%s-variable-sources" % workbook)
        try:
            seg_dir, segmentation = resolve_segmentation_directory(
                args.seg_root,
                workbook,
                segmentation_generation_id=args.segmentation_generation_id,
                source_generation_root=args.source_generation_root,
                source_generation_id=args.source_generation_id,
            )
        except (RuntimeError, ValueError, OSError) as exc:
            parser.error(str(exc))
        if args.inventory_only:
            output_dir.mkdir(parents=True, exist_ok=True)
            inventory = build_inventory(
                workbook,
                artifact,
                seg_dir,
                segmentation=segmentation,
            )
            path = output_dir / ("%s-inputs-variable-sources.inventory.json"
                                 % workbook)
            path.write_text(
                json.dumps(inventory, indent=2, ensure_ascii=False) + "\n",
                encoding="utf-8",
            )
            results.append({"workbook": workbook, "inventory": str(path),
                            "rows": len(inventory["variables"])})
        else:
            results.append(generate_audit(
                workbook, artifact, seg_dir, output_dir,
                api_key, model=args.model, endpoint=args.endpoint,
                project_id=args.project_id, refresh=args.refresh,
                segmentation=segmentation,
            ))
    print(json.dumps(results, indent=2))
    return 0


if __name__ == "__main__":
    sys.exit(main())
